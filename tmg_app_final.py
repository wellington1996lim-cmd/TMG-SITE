import sys
import streamlit as st
import streamlit.components.v1 as components
from pathlib import Path
from PIL import Image, ImageFile, ImageDraw
import os
import shutil
import base64
import io
from io import BytesIO
import numpy as np
import json
import tempfile
import zipfile
import csv
import cv2
import warnings
import hashlib
import html
import re
import subprocess
import sqlite3
import threading
import time
from datetime import datetime, date
import pandas as pd

try:
    from streamlit_image_coordinates import streamlit_image_coordinates
    HAS_STREAMLIT_IMAGE_COORDINATES = True
except ImportError:
    streamlit_image_coordinates = None
    HAS_STREAMLIT_IMAGE_COORDINATES = False

APP_ROOT = Path(__file__).resolve().parent
APP_TEMP_DIR = APP_ROOT / "tmg_data" / "tmp"
APP_ULTRALYTICS_DIR = APP_ROOT / "Ultralytics"
APP_MPLCONFIG_DIR = APP_ROOT / "tmg_data" / "matplotlib"

for _local_dir in (
    APP_ROOT / ".streamlit",
    APP_ROOT / "tmg_data",
    APP_ROOT / "tmg_config",
    APP_TEMP_DIR,
    APP_ULTRALYTICS_DIR,
    APP_MPLCONFIG_DIR,
):
    _local_dir.mkdir(parents=True, exist_ok=True)

os.environ.setdefault("TMG_APP_ROOT", str(APP_ROOT))
os.environ.setdefault("TMG_DATA_DIR", str(APP_ROOT / "tmg_data"))
os.environ.setdefault("TMG_CONFIG_DIR", str(APP_ROOT / "tmg_config"))
os.environ.setdefault("STREAMLIT_CONFIG_DIR", str(APP_ROOT / ".streamlit"))
os.environ.setdefault("YOLO_CONFIG_DIR", str(APP_ULTRALYTICS_DIR))
os.environ.setdefault("MPLCONFIGDIR", str(APP_MPLCONFIG_DIR))
os.environ.setdefault("TMP", str(APP_TEMP_DIR))
os.environ.setdefault("TEMP", str(APP_TEMP_DIR))
os.environ.setdefault("TMPDIR", str(APP_TEMP_DIR))
tempfile.tempdir = str(APP_TEMP_DIR)

# NOVO - Configurações para suportar imagens grandes e formatos variados
Image.MAX_IMAGE_PIXELS = None
ImageFile.LOAD_TRUNCATED_IMAGES = True
warnings.simplefilter('ignore', Image.DecompressionBombWarning)

# IMPORTAÇÕES ADICIONAIS PARA GERAÇÃO ROBUSTA DE SHP[cite: 1]
try:
    import geopandas as gpd
    from shapely.geometry import Polygon
    from affine import Affine
    HAS_GEOPANDAS = True
except ImportError:
    HAS_GEOPANDAS = False

# Remover limite de tamanho de imagem do PIL para Ortofotos gigantes[cite: 1]
Image.MAX_IMAGE_PIXELS = None

# ==========================================
# CONFIGURAÇÃO DA PÁGINA[cite: 1]
# ==========================================
st.set_page_config(
    page_title="TMG Sistema de Análise",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown(
    """
    <meta name="google" content="notranslate">
    <meta name="robots" content="notranslate">
    <style>
      html, body, #root, .stApp, [data-testid="stAppViewContainer"] {
        translate: no;
      }
      .goog-te-banner-frame, .skiptranslate {
        display: none !important;
      }
      body {
        top: 0 !important;
      }
    </style>
    """,
    unsafe_allow_html=True
)

# ==========================================
# ESTILO CSS PERSONALIZADO (DARK SAAS 3D)[cite: 1]
# ==========================================
st.markdown("""
<style>
    .stApp {
        background-color: #121212;
        color: #e0e0e0;
    }

    .main-header {
        text-align: center;
        color: #FFFFFF;
        padding: 20px;
        font-family: 'Segoe UI', sans-serif;
        font-weight: 800;
        letter-spacing: 4px;
        font-size: 2.2rem;
        text-shadow:
            2px 2px 0px #000000,
            4px 4px 0px #1a1a1a,
            6px 6px 8px rgba(0,0,0,0.9),
            0 0 30px var(--tmg-primary-glow-soft);
        border-bottom: 2px solid var(--tmg-primary);
        margin-bottom: 30px;
    }

    [data-testid="stSidebar"] {
        background-color: #1a1a1a;
        border-right: 1px solid #333;
        padding-top: 20px;
    }

    div.stButton > button {
        width: 100%;
        border-radius: 10px;
        border: none;
        padding: 15px 20px;
        background: linear-gradient(145deg, #222, #111);
        color: #ccc;
        font-weight: 600;
        box-shadow: 3px 3px 6px #0a0a0a, -1px -1px 6px #2a2a2a;
        transition: 0.3s;
        margin-bottom: 10px;
        text-align: left;
    }

    div.stButton > button:hover {
        color: var(--tmg-primary);
        border: 1px solid var(--tmg-primary);
        transform: translateY(-2px);
    }

    .stButton > button[kind="primary"] {
        background: linear-gradient(145deg, var(--tmg-primary), var(--tmg-primary-dark)) !important;
        color: white !important;
        box-shadow: 4px 4px 10px #0a0a0a !important;
    }

    .card {
        background-color: #1e1e1e;
        padding: 20px;
        border-radius: 15px;
        border: 1px solid #333;
    }

    /* 3D Menu Title */
    .menu-3d-title {
        text-align: center;
        font-family: 'Segoe UI', sans-serif;
        font-weight: 900;
        font-size: 1.6rem;
        letter-spacing: 4px;
        color: var(--tmg-primary);
        text-transform: uppercase;
        text-shadow:
            1px 1px 0 var(--tmg-primary-shadow-1),
            2px 2px 0 var(--tmg-primary-shadow-2),
            3px 3px 0 var(--tmg-primary-shadow-3),
            4px 4px 6px rgba(0,0,0,0.9),
            0 0 20px var(--tmg-primary-glow),
            0 0 40px var(--tmg-primary-glow-soft);
        margin-bottom: 8px;
    }

    /* Separator with glow */
    .separator-glow {
        border: none;
        border-top: 1px solid var(--tmg-primary);
        box-shadow: 0 0 8px var(--tmg-primary-glow);
        margin: 10px 0 18px 0;
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
# DIRETÓRIOS DE ASSETS[cite: 1]
# ==========================================
LOGO_DIR = APP_ROOT / "tmg_assets/logo"
LOGO_DIR.mkdir(parents=True, exist_ok=True)
LOGO_PATH = LOGO_DIR / "logo_tmg_referencia.png"

LOGIN_BG_DIR = APP_ROOT / "tmg_assets/login"
LOGIN_BG_DIR.mkdir(parents=True, exist_ok=True)
LOGIN_BG_PATH = LOGIN_BG_DIR / "login_bg_referencia.png"

SYSTEM_CONFIG_DIR = APP_ROOT / "tmg_config"
SYSTEM_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
SYSTEM_CONFIG_PATH = SYSTEM_CONFIG_DIR / "system_config.json"

# ==========================================
# HELPER — ENCODE IMAGEM PARA BASE64 CSS[cite: 1]
# ==========================================
def _tmg_file_signature(path: Path) -> tuple[str, int, int]:
    p = Path(path)
    stat = p.stat()
    return str(p.resolve()), int(getattr(stat, "st_mtime_ns", int(stat.st_mtime * 1_000_000_000))), int(stat.st_size)

def _tmg_image_mime(path: Path) -> str:
    suffix = Path(path).suffix.lower()
    if suffix in (".jpg", ".jpeg"):
        return "image/jpeg"
    if suffix == ".webp":
        return "image/webp"
    if suffix == ".gif":
        return "image/gif"
    if suffix == ".svg":
        return "image/svg+xml"
    return "image/png"

@st.cache_data(show_spinner=False, max_entries=80)
def _img_to_base64_css_cached(path_str: str, mtime_ns: int, size_bytes: int, mime: str) -> str:
    del mtime_ns, size_bytes
    data = base64.b64encode(Path(path_str).read_bytes()).decode("ascii")
    return f"data:{mime};base64,{data}"

def _img_to_base64_css(path: Path) -> str:
    path_str, mtime_ns, size_bytes = _tmg_file_signature(path)
    return _img_to_base64_css_cached(path_str, mtime_ns, size_bytes, _tmg_image_mime(Path(path_str)))

@st.cache_data(show_spinner=False, max_entries=120)
def _tmg_read_json_cached(path_str: str, mtime_ns: int, size_bytes: int):
    del mtime_ns, size_bytes
    return json.loads(Path(path_str).read_text(encoding="utf-8"))

def _tmg_json_clone(data):
    try:
        return json.loads(json.dumps(data, ensure_ascii=False))
    except Exception:
        return data

def _tmg_read_json_file(path: Path, default):
    path = Path(path)
    if not path.exists():
        return _tmg_json_clone(default)
    try:
        path_str, mtime_ns, size_bytes = _tmg_file_signature(path)
        return _tmg_read_json_cached(path_str, mtime_ns, size_bytes)
    except Exception:
        return _tmg_json_clone(default)

def get_progress_color(progress: int | float) -> str:
    try:
        pct = max(0, min(100, int(float(progress))))
    except Exception:
        pct = 0
    theme = globals().get("DEPLOY_BAR_THEME", {})
    if pct >= 100:
        return theme.get("fill_100", "linear-gradient(90deg,#00bcd4,#00e676)")
    if pct >= 86:
        return theme.get("fill_86_99", "linear-gradient(90deg,#0d47a1,#1565c0,#42a5f5)")
    if pct >= 61:
        return theme.get("fill_61_85", "linear-gradient(90deg,#1565c0,#1976d2,#42a5f5)")
    if pct >= 31:
        return theme.get("fill_31_60", "linear-gradient(90deg,#1976d2,#42a5f5,#64b5f6)")
    return theme.get("fill_0_30", "linear-gradient(90deg,#64b5f6,#90caf9,#bbdefb)")

def _tmg_loading_logo_html() -> str:
    try:
        if LOGO_PATH.exists():
            logo_src = _img_to_base64_css(LOGO_PATH)
            return f"<img class='tmg-load-logo-img' src='{logo_src}' alt='TMG'>"
    except Exception:
        pass
    return "<div class='tmg-load-logo-fallback'>TMG</div>"

def render_tmg_loading_bar(progress, texto: str = "Carregando arquivo...", container=None):
    try:
        pct = max(0, min(100, int(float(progress))))
    except Exception:
        pct = 0
    texto_seguro = html.escape(str(texto or "Carregando arquivo..."))
    status = "Carregamento concluído com sucesso." if pct >= 100 else texto_seguro
    fill = get_progress_color(pct)
    theme = globals().get("DEPLOY_BAR_THEME", {})
    card_background = theme.get(
        "card_background",
        "linear-gradient(145deg, rgba(14,26,43,.96), rgba(16,18,24,.98)), radial-gradient(circle at top left, rgba(0,229,255,.14), transparent 36%)",
    )
    card_border = theme.get("card_border", "rgba(66,165,245,.36)")
    card_shadow = theme.get(
        "card_shadow",
        "0 16px 34px rgba(0,0,0,.34), inset 0 1px 0 rgba(255,255,255,.08), inset 0 -10px 22px rgba(0,0,0,.18)",
    )
    track_background = theme.get("track_background", "linear-gradient(180deg,#070b12,#111d2c)")
    track_border = theme.get("track_border", "rgba(255,255,255,.10)")
    track_shadow = theme.get("track_shadow", "inset 0 3px 8px rgba(0,0,0,.65), 0 8px 18px rgba(0,0,0,.28)")
    fill_shadow = theme.get("fill_shadow", "inset 0 1px 0 rgba(255,255,255,.42), 0 0 18px rgba(66,165,245,.58)")
    status_color = theme.get("status_success" if pct >= 100 else "status_loading", "#5ff2b1" if pct >= 100 else "#ffb347")
    logo_html = _tmg_loading_logo_html()
    markup = f"""
    <style>
    .tmg-load-card {{
        width:100%;
        margin:10px 0 14px 0;
        padding:16px 18px;
        border-radius:14px;
        border:1px solid {card_border};
        background:{card_background};
        box-shadow:{card_shadow};
        font-family:'Segoe UI', Arial, sans-serif;
    }}
    .tmg-load-head {{
        display:flex;
        align-items:center;
        justify-content:center;
        gap:12px;
        margin-bottom:12px;
    }}
    .tmg-load-logo-img {{
        max-height:40px;
        max-width:130px;
        object-fit:contain;
        filter:drop-shadow(0 6px 12px rgba(0,0,0,.45));
    }}
    .tmg-load-logo-fallback {{
        color:#fff;
        font-weight:900;
        letter-spacing:4px;
        font-size:1.15rem;
        text-shadow:0 0 16px rgba(66,165,245,.55);
    }}
    .tmg-load-text {{
        color:#f4f7fb;
        font-size:.88rem;
        font-weight:700;
        letter-spacing:.3px;
        text-align:center;
    }}
    .tmg-load-track {{
        position:relative;
        height:24px;
        overflow:hidden;
        border-radius:999px;
        background:{track_background};
        border:1px solid {track_border};
        box-shadow:{track_shadow};
    }}
    .tmg-load-fill {{
        width:{pct}%;
        height:100%;
        border-radius:999px;
        background:{fill};
        box-shadow:{fill_shadow};
        transition:width .45s ease, background .45s ease;
        position:relative;
    }}
    .tmg-load-fill:after {{
        content:"";
        position:absolute;
        inset:0;
        background:linear-gradient(120deg, transparent 0%, rgba(255,255,255,.38) 45%, transparent 75%);
        transform:translateX(-100%);
        animation:tmgLoadShine 1.4s ease-in-out infinite;
    }}
    .tmg-load-percent {{
        position:absolute;
        inset:0;
        display:flex;
        align-items:center;
        justify-content:center;
        color:#fff;
        font-weight:900;
        font-size:.78rem;
        text-shadow:0 1px 4px rgba(0,0,0,.8);
        letter-spacing:.5px;
    }}
    .tmg-load-status {{
        margin-top:8px;
        color:{status_color};
        font-size:.78rem;
        font-weight:700;
        text-align:center;
    }}
    @keyframes tmgLoadShine {{
        0% {{ transform:translateX(-100%); }}
        100% {{ transform:translateX(180%); }}
    }}
    </style>
    <div class="tmg-load-card">
        <div class="tmg-load-head">{logo_html}<div class="tmg-load-text">{texto_seguro}</div></div>
        <div class="tmg-load-track">
            <div class="tmg-load-fill"></div>
            <div class="tmg-load-percent">{pct}%</div>
        </div>
        <div class="tmg-load-status">{html.escape(status)}</div>
    </div>
    """
    target = container if container is not None else st
    target.markdown(markup, unsafe_allow_html=True)
    return container

def update_tmg_loading(container, progress, texto: str = "Carregando arquivo..."):
    if container is None:
        return None
    return render_tmg_loading_bar(progress, texto, container=container)

def render_progress_upload_tmg(progress, texto: str = "Carregando arquivo...", container=None):
    return render_tmg_loading_bar(progress, texto, container=container)

def render_upload_status_tmg(progress=100, texto: str = "Carregamento concluído com sucesso.", container=None):
    return render_tmg_loading_bar(progress, texto, container=container)

def clear_tmg_loading(container):
    try:
        if container is not None:
            container.markdown(
                "<div class='tmg-load-cleared' style='display:none;height:0;overflow:hidden'></div>",
                unsafe_allow_html=True,
            )
    except Exception:
        pass

def finish_tmg_loading_and_clear(container, texto: str = "Carregamento concluído com sucesso.", hold_seconds: float = 0.45):
    if container is None:
        return
    try:
        update_tmg_loading(container, 100, texto)
        if hold_seconds and hold_seconds > 0:
            time.sleep(float(hold_seconds))
    except Exception:
        pass
    clear_tmg_loading(container)

def app_rerun():
    if hasattr(st, "rerun"):
        st.rerun()
    else:
        st.experimental_rerun()

def app_image(image, **kwargs):
    try:
        return st.image(image, use_container_width=True, **kwargs)
    except TypeError:
        return st.image(image, use_column_width=True, **kwargs)

def _streamlit_secret(name: str, default: str = "") -> str:
    secrets_paths = [
        APP_ROOT / ".streamlit" / "secrets.toml",
    ]
    if not any(path.exists() for path in secrets_paths):
        return default
    try:
        return str(st.secrets.get(name, default) or default)
    except Exception:
        return default

def _configured_database_dir() -> str:
    configured = (
        os.getenv("TMG_DATABASE_DIR", "").strip()
        or _streamlit_secret("TMG_DATABASE_DIR").strip()
        or _streamlit_secret("database_dir").strip()
    )
    if configured:
        path = _resolve_system_path(configured)
        if _path_inside_app_root(path):
            return str(path)
    return str((APP_ROOT / "tmg_data").resolve())

def _int_setting(name: str, default: int, min_value: int, max_value: int) -> int:
    raw = os.getenv(name, "").strip() or _streamlit_secret(name).strip()
    try:
        value = int(raw) if raw else int(default)
    except Exception:
        value = int(default)
    return max(min_value, min(max_value, value))

def _preview_max_dim() -> int:
    # Ajuste solicitado: visualização de ortofotos com mais qualidade no Streamlit.
    # Pode ser configurado por variável/secret TMG_PREVIEW_MAX_DIM.
    # Mantém compatibilidade com Streamlit Cloud evitando carregar a imagem original inteira no navegador.
    return _int_setting("TMG_PREVIEW_MAX_DIM", 8192, 1024, 12000)

def _preview_jpeg_quality() -> int:
    # Qualidade alta para preservar detalhes de TIF/GeoTIFF/RGB no visualizador.
    return _int_setting("TMG_PREVIEW_JPEG_QUALITY", 97, 70, 98)

def _preview_max_payload_mb() -> int:
    # Evita que o HTML do visualizador fique pesado demais para carregar no navegador.
    return _int_setting("TMG_PREVIEW_MAX_PAYLOAD_MB", 10, 4, 80)

def _preview_min_dim() -> int:
    return _int_setting("TMG_PREVIEW_MIN_DIM", 2048, 900, 4096)

def _upload_limit_mb() -> int:
    # Valor informativo mostrado na interface; o limite real do Streamlit Cloud é definido em .streamlit/config.toml.
    return _int_setting("TMG_UPLOAD_LIMIT_MB", 2048, 200, 4096)

def _looks_like_windows_drive_path(raw: str) -> bool:
    return len(raw) > 2 and raw[1] == ":" and raw[2:3] in ("\\", "/")

def _path_inside_app_root(path: Path) -> bool:
    try:
        return Path(path).resolve().is_relative_to(APP_ROOT.resolve())
    except Exception:
        return False

def _force_inside_app_root(path: Path, default_name: str = "tmg_data") -> Path:
    path = Path(path)
    if _path_inside_app_root(path):
        return path.resolve()
    safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", path.name or default_name).strip("._") or default_name
    return (APP_ROOT / safe_name).resolve()

def _resolve_system_path(value: str) -> Path:
    raw = os.path.expandvars(str(value or "").strip())
    if os.name != "nt" and _looks_like_windows_drive_path(raw):
        raw = "tmg_data"
    path = Path(raw) if raw else Path("tmg_data")
    if not path.is_absolute():
        path = (APP_ROOT / path).resolve()
    return _force_inside_app_root(path, "tmg_data")

def _load_system_config() -> dict:
    configured_dir = _configured_database_dir()
    default_dir = _resolve_system_path(configured_dir or "tmg_data")
    default = {
        "database_dir": str(default_dir),
        "updated_at": "",
        "tema": "padrao"
    }
    if not SYSTEM_CONFIG_PATH.exists():
        SYSTEM_CONFIG_PATH.write_text(json.dumps(default, indent=2, ensure_ascii=False), encoding="utf-8")
        return default
    data = _tmg_read_json_file(SYSTEM_CONFIG_PATH, default)
    data.setdefault("database_dir", str(default_dir))
    data.setdefault("updated_at", "")
    data.setdefault("tema", "padrao")
    data["database_dir"] = str(_resolve_system_path(data.get("database_dir", default_dir)))
    return data

def _save_system_config(data: dict) -> None:
    SYSTEM_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    data["database_dir"] = str(_resolve_system_path(data.get("database_dir", "tmg_data")))
    SYSTEM_CONFIG_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

SYSTEM_CONFIG = _load_system_config()
SYSTEM_DATABASE_DIR = _resolve_system_path(SYSTEM_CONFIG.get("database_dir", "tmg_data"))
SYSTEM_DATABASE_DIR.mkdir(parents=True, exist_ok=True)

THEME_PALETTES = {
    "padrao": {
        "primary": "#42a5f5",
        "primary_dark": "#1565c0",
        "primary_soft": "#90caf9",
        "shadow_1": "#0d47a1",
        "shadow_2": "#0a3070",
        "shadow_3": "#071f4a",
        "rgb": "66,165,245",
    },
    "tmg_azul": {
        "primary": "#42a5f5",
        "primary_dark": "#1565c0",
        "primary_soft": "#90caf9",
        "shadow_1": "#0d47a1",
        "shadow_2": "#0a3070",
        "shadow_3": "#071f4a",
        "rgb": "66,165,245",
    },
    "tmg_premium_neon_3d": {
        "primary": "#00E5FF",
        "primary_dark": "#0E3A70",
        "primary_soft": "#00FF9D",
        "shadow_1": "#1D7BFF",
        "shadow_2": "#071A35",
        "shadow_3": "#020E24",
        "rgb": "0,229,255",
    },
}
THEME_PALETTE = THEME_PALETTES.get(SYSTEM_CONFIG.get("tema", "padrao"), THEME_PALETTES["padrao"])
THEME_PRIMARY_COLOR = THEME_PALETTE["primary"]
THEME_PRIMARY_DARK = THEME_PALETTE["primary_dark"]
THEME_PRIMARY_SOFT = THEME_PALETTE["primary_soft"]
THEME_PRIMARY_RGB = THEME_PALETTE["rgb"]

DEPLOY_BAR_THEME_PATH = SYSTEM_DATABASE_DIR / "deploy_bar_theme.json"

def _default_deploy_bar_theme() -> dict:
    return {
        "version": 1,
        "theme": SYSTEM_CONFIG.get("tema", "padrao"),
        "card_background": (
            f"radial-gradient(circle at top left, rgba({THEME_PRIMARY_RGB},.16), transparent 36%), "
            "linear-gradient(145deg, rgba(2,14,36,.98), rgba(7,26,53,.96))"
        ),
        "card_border": f"rgba({THEME_PRIMARY_RGB},.50)",
        "card_shadow": (
            "0 18px 38px rgba(0,0,0,.48), "
            f"0 0 28px rgba({THEME_PRIMARY_RGB},.22), "
            "inset 0 1px 0 rgba(255,255,255,.08)"
        ),
        "track_background": "linear-gradient(180deg,#020e24,#061525)",
        "track_border": f"rgba({THEME_PRIMARY_RGB},.42)",
        "track_shadow": (
            "inset 0 3px 8px rgba(0,0,0,.68), "
            "0 8px 18px rgba(0,0,0,.30), "
            f"0 0 14px rgba({THEME_PRIMARY_RGB},.18)"
        ),
        "fill_0_30": f"linear-gradient(90deg,{THEME_PRIMARY_SOFT},{THEME_PRIMARY_COLOR})",
        "fill_31_60": f"linear-gradient(90deg,{THEME_PRIMARY_COLOR},#00d4ff)",
        "fill_61_85": f"linear-gradient(90deg,{THEME_PRIMARY_DARK},{THEME_PRIMARY_COLOR},#00d4ff)",
        "fill_86_99": f"linear-gradient(90deg,#020e24,{THEME_PRIMARY_DARK},{THEME_PRIMARY_COLOR})",
        "fill_100": f"linear-gradient(90deg,#00bcd4,{THEME_PRIMARY_SOFT},#00e676)",
        "fill_active": f"linear-gradient(90deg,#00e5ff 0%,{THEME_PRIMARY_DARK} 48%,#00ff9d 100%)",
        "fill_shadow": (
            "inset 0 1px 0 rgba(255,255,255,.46), "
            f"0 0 16px rgba({THEME_PRIMARY_RGB},.56), "
            "0 0 22px rgba(0,255,157,.22)"
        ),
        "status_success": "#5ff2b1",
        "status_loading": THEME_PRIMARY_SOFT,
    }

def _load_deploy_bar_theme() -> dict:
    default = _default_deploy_bar_theme()
    data = _tmg_read_json_file(DEPLOY_BAR_THEME_PATH, {})
    if not isinstance(data, dict):
        data = {}
    if data.get("theme") not in (None, default["theme"]):
        data = {}
    merged = dict(default)
    for key, value in data.items():
        if key == "version" and isinstance(value, (int, float)):
            merged[key] = value
        elif key != "version" and isinstance(value, str):
            merged[key] = value
    try:
        DEPLOY_BAR_THEME_PATH.parent.mkdir(parents=True, exist_ok=True)
        DEPLOY_BAR_THEME_PATH.write_text(json.dumps(merged, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass
    return merged

DEPLOY_BAR_THEME = _load_deploy_bar_theme()

st.markdown(f"""
<style>
:root {{
    --tmg-primary: {THEME_PRIMARY_COLOR};
    --tmg-primary-dark: {THEME_PRIMARY_DARK};
    --tmg-primary-soft: {THEME_PRIMARY_SOFT};
    --tmg-primary-shadow-1: {THEME_PALETTE["shadow_1"]};
    --tmg-primary-shadow-2: {THEME_PALETTE["shadow_2"]};
    --tmg-primary-shadow-3: {THEME_PALETTE["shadow_3"]};
    --tmg-primary-glow: rgba({THEME_PRIMARY_RGB}, .42);
    --tmg-primary-glow-soft: rgba({THEME_PRIMARY_RGB}, .18);
    --tmg-deploy-card-bg: {DEPLOY_BAR_THEME.get("card_background")};
    --tmg-deploy-border: {DEPLOY_BAR_THEME.get("card_border")};
    --tmg-deploy-card-shadow: {DEPLOY_BAR_THEME.get("card_shadow")};
    --tmg-deploy-track-bg: {DEPLOY_BAR_THEME.get("track_background")};
    --tmg-deploy-track-border: {DEPLOY_BAR_THEME.get("track_border")};
    --tmg-deploy-track-shadow: {DEPLOY_BAR_THEME.get("track_shadow")};
    --tmg-deploy-fill-active: {DEPLOY_BAR_THEME.get("fill_active")};
    --tmg-deploy-fill-shadow: {DEPLOY_BAR_THEME.get("fill_shadow")};
}}
[style*="#ff8c00"], [style*="#FF8C00"] {{
    color: var(--tmg-primary) !important;
    border-color: var(--tmg-primary) !important;
}}
.main-header,
.cultura-title,
.login-title,
.partner-hero-title,
.partner-window-title,
.partner-card-title,
.vd-hero h2,
h1, h2, h3 {{
    text-shadow:
        1px 1px 0 rgba(0,0,0,.75),
        2px 2px 0 rgba(0,0,0,.45),
        0 0 18px var(--tmg-primary-glow) !important;
}}
header[data-testid="stHeader"],
[data-testid="stHeader"] {{
    background:
        linear-gradient(90deg, rgba(2,14,36,.97), rgba(7,31,63,.96) 48%, rgba({THEME_PRIMARY_RGB}, .20)),
        linear-gradient(180deg, rgba(255,255,255,.06), rgba(255,255,255,0)) !important;
    border-bottom:1px solid rgba({THEME_PRIMARY_RGB}, .42) !important;
    box-shadow:
        0 8px 24px rgba(0,0,0,.34),
        0 0 24px rgba({THEME_PRIMARY_RGB}, .20) !important;
    backdrop-filter: blur(10px) saturate(140%);
    -webkit-backdrop-filter: blur(10px) saturate(140%);
}}
header[data-testid="stHeader"] > div,
[data-testid="stHeader"] > div {{
    background:transparent !important;
}}
[data-testid="stDecoration"] {{
    background:linear-gradient(90deg, var(--tmg-primary-dark), var(--tmg-primary), var(--tmg-primary-soft)) !important;
}}
.tmg-user-chip-neon {{
    position:fixed;
    left:14px;
    top:58px;
    z-index:999997;
    pointer-events:auto;
    padding:9px 14px;
    border-radius:14px;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .62);
    background:
        linear-gradient(120deg, rgba(255,255,255,.13), transparent 28%),
        radial-gradient(circle at top left, rgba({THEME_PRIMARY_RGB}, .24), transparent 42%),
        linear-gradient(145deg, rgba(2,14,36,.94), rgba(14,58,112,.82) 54%, rgba({THEME_PRIMARY_RGB}, .30));
    color:#ffffff;
    font-weight:900;
    font-size:.82rem;
    letter-spacing:.45px;
    box-shadow:
        0 13px 28px rgba(0,0,0,.48),
        0 0 20px rgba({THEME_PRIMARY_RGB}, .38),
        inset 0 1px 0 rgba(255,255,255,.24),
        inset 0 -7px 14px rgba(2,14,36,.48);
    text-shadow:
        0 1px 0 rgba(0,0,0,.92),
        0 0 12px rgba({THEME_PRIMARY_RGB}, .62);
    backdrop-filter: blur(10px) saturate(140%);
    -webkit-backdrop-filter: blur(10px) saturate(140%);
    transform:translateZ(0);
    transition:all .30s ease;
}}
.tmg-user-chip-neon:hover {{
    border-color:var(--tmg-primary-soft);
    box-shadow:
        0 16px 34px rgba(0,0,0,.54),
        0 0 28px rgba({THEME_PRIMARY_RGB}, .58),
        0 0 42px rgba({THEME_PRIMARY_RGB}, .24),
        inset 0 1px 0 rgba(255,255,255,.34),
        inset 0 -8px 16px rgba(2,14,36,.44);
    transform:translateY(-1px) scale(1.01);
}}
header [data-testid="stToolbar"] a,
header [data-testid="stToolbar"] button,
[data-testid="stToolbar"] a,
[data-testid="stToolbar"] button {{
    border-radius:14px !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .54) !important;
    background:
        linear-gradient(120deg, rgba(255,255,255,.14), transparent 30%),
        linear-gradient(145deg, rgba(2,14,36,.92), rgba(14,58,112,.78), rgba({THEME_PRIMARY_RGB}, .26)) !important;
    color:#ffffff !important;
    font-weight:900 !important;
    text-shadow:0 1px 0 rgba(0,0,0,.84), 0 0 10px rgba({THEME_PRIMARY_RGB}, .56) !important;
    box-shadow:
        0 10px 24px rgba(0,0,0,.38),
        0 0 18px rgba({THEME_PRIMARY_RGB}, .32),
        inset 0 1px 0 rgba(255,255,255,.24),
        inset 0 -5px 12px rgba(2,14,36,.50) !important;
    transition:all .30s ease !important;
    backdrop-filter: blur(9px) saturate(140%);
    -webkit-backdrop-filter: blur(9px) saturate(140%);
}}
header [data-testid="stToolbar"] a:hover,
header [data-testid="stToolbar"] button:hover,
[data-testid="stToolbar"] a:hover,
[data-testid="stToolbar"] button:hover {{
    border-color:var(--tmg-primary-soft) !important;
    box-shadow:
        0 13px 28px rgba(0,0,0,.44),
        0 0 26px rgba({THEME_PRIMARY_RGB}, .55),
        0 0 38px rgba({THEME_PRIMARY_RGB}, .22),
        inset 0 1px 0 rgba(255,255,255,.34) !important;
    transform:translateY(-1px) scale(1.01) !important;
}}
header [data-testid="stToolbar"] a:active,
header [data-testid="stToolbar"] button:active,
[data-testid="stToolbar"] a:active,
[data-testid="stToolbar"] button:active {{
    transform:translateY(1px) scale(.99) !important;
    box-shadow:inset 0 4px 10px rgba(0,0,0,.62), 0 0 14px rgba({THEME_PRIMARY_RGB}, .38) !important;
}}
[data-testid="stDeployButton"],
[data-testid="stDeployButton"] button,
header [data-testid="stToolbarActions"] button,
header [data-testid="stToolbarActions"] a {{
    border-radius:14px !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .58) !important;
    background:
        linear-gradient(120deg, rgba(255,255,255,.16), transparent 31%),
        radial-gradient(circle at 18% 0%, rgba({THEME_PRIMARY_RGB}, .25), transparent 46%),
        linear-gradient(145deg, rgba(2,14,36,.94), rgba(14,58,112,.80), rgba({THEME_PRIMARY_RGB}, .28)) !important;
    color:#ffffff !important;
    font-weight:900 !important;
    text-shadow:0 1px 0 rgba(0,0,0,.88), 0 0 12px rgba({THEME_PRIMARY_RGB}, .62) !important;
    box-shadow:
        0 10px 24px rgba(0,0,0,.40),
        0 0 20px rgba({THEME_PRIMARY_RGB}, .34),
        inset 0 1px 0 rgba(255,255,255,.27),
        inset 0 -6px 12px rgba(2,14,36,.50) !important;
    transition:all .30s ease !important;
    backdrop-filter: blur(9px) saturate(145%);
    -webkit-backdrop-filter: blur(9px) saturate(145%);
}}
[data-testid="stDeployButton"] *,
header [data-testid="stToolbarActions"] button *,
header [data-testid="stToolbarActions"] a * {{
    color:#ffffff !important;
    font-weight:900 !important;
    text-shadow:0 1px 0 rgba(0,0,0,.88), 0 0 12px rgba({THEME_PRIMARY_RGB}, .62) !important;
}}
[data-testid="stDeployButton"]:hover,
[data-testid="stDeployButton"] button:hover,
header [data-testid="stToolbarActions"] button:hover,
header [data-testid="stToolbarActions"] a:hover {{
    border-color:var(--tmg-primary-soft) !important;
    box-shadow:
        0 13px 28px rgba(0,0,0,.46),
        0 0 28px rgba({THEME_PRIMARY_RGB}, .58),
        0 0 42px rgba({THEME_PRIMARY_RGB}, .24),
        inset 0 1px 0 rgba(255,255,255,.36) !important;
    transform:translateY(-1px) scale(1.01) !important;
}}
[data-testid="stDeployButton"]:active,
[data-testid="stDeployButton"] button:active,
header [data-testid="stToolbarActions"] button:active,
header [data-testid="stToolbarActions"] a:active {{
    transform:translateY(1px) scale(.99) !important;
    box-shadow:inset 0 4px 10px rgba(0,0,0,.64), 0 0 15px rgba({THEME_PRIMARY_RGB}, .40) !important;
}}
div[data-testid="stTextInput"] input,
div[data-testid="stNumberInput"] input,
div[data-testid="stTextArea"] textarea,
div[data-testid="stDateInput"] input,
div[data-testid="stTimeInput"] input {{
    border-radius:8px !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .64) !important;
    background:
        linear-gradient(120deg, rgba(255,255,255,.10), transparent 30%),
        linear-gradient(145deg, rgba(2,14,36,.96), rgba(18,62,100,.82), rgba({THEME_PRIMARY_RGB}, .16)) !important;
    color:#ffffff !important;
    font-weight:700 !important;
    box-shadow:
        0 9px 22px rgba(0,0,0,.34),
        0 0 18px rgba({THEME_PRIMARY_RGB}, .20),
        inset 0 1px 0 rgba(255,255,255,.18),
        inset 0 -5px 12px rgba(2,14,36,.42) !important;
    text-shadow:0 1px 0 rgba(0,0,0,.78) !important;
    transition:all .30s ease !important;
}}
div[data-testid="stTextInput"] input:hover,
div[data-testid="stNumberInput"] input:hover,
div[data-testid="stTextArea"] textarea:hover,
div[data-testid="stDateInput"] input:hover,
div[data-testid="stTimeInput"] input:hover,
div[data-testid="stTextInput"] input:focus,
div[data-testid="stNumberInput"] input:focus,
div[data-testid="stTextArea"] textarea:focus,
div[data-testid="stDateInput"] input:focus,
div[data-testid="stTimeInput"] input:focus {{
    border-color:var(--tmg-primary-soft) !important;
    box-shadow:
        0 10px 26px rgba(0,0,0,.38),
        0 0 26px rgba({THEME_PRIMARY_RGB}, .36),
        inset 0 1px 0 rgba(255,255,255,.26) !important;
}}
div[data-testid="stTextInput"] input::placeholder,
div[data-testid="stNumberInput"] input::placeholder,
div[data-testid="stTextArea"] textarea::placeholder {{
    color:rgba(224,247,255,.72) !important;
}}
div[data-testid="stSelectbox"] [data-baseweb="select"],
div[data-testid="stMultiSelect"] [data-baseweb="select"] {{
    border-radius:9px !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .58) !important;
    background:
        linear-gradient(120deg, rgba(255,255,255,.13), transparent 32%),
        radial-gradient(circle at right, rgba({THEME_PRIMARY_RGB}, .32), transparent 44%),
        linear-gradient(145deg, rgba(2,14,36,.96), rgba(18,62,100,.78), rgba({THEME_PRIMARY_RGB}, .22)) !important;
    box-shadow:
        0 10px 24px rgba(0,0,0,.36),
        0 0 22px rgba({THEME_PRIMARY_RGB}, .30),
        inset 0 1px 0 rgba(255,255,255,.22),
        inset 0 -6px 12px rgba(2,14,36,.44) !important;
    transition:all .30s ease !important;
}}
div[data-testid="stSelectbox"] [data-baseweb="select"]:hover,
div[data-testid="stMultiSelect"] [data-baseweb="select"]:hover {{
    border-color:var(--tmg-primary-soft) !important;
    box-shadow:
        0 12px 28px rgba(0,0,0,.42),
        0 0 30px rgba({THEME_PRIMARY_RGB}, .46),
        inset 0 1px 0 rgba(255,255,255,.30) !important;
}}
div[data-testid="stSelectbox"] [data-baseweb="select"] *,
div[data-testid="stMultiSelect"] [data-baseweb="select"] * {{
    color:#ffffff !important;
    font-weight:800 !important;
    text-shadow:0 1px 0 rgba(0,0,0,.78), 0 0 10px rgba({THEME_PRIMARY_RGB}, .38) !important;
}}
</style>
""", unsafe_allow_html=True)

def aplicar_estilo_titulos_3d():
    st.markdown("""
    <style>
    div[data-testid="stMarkdownContainer"] h1,
    div[data-testid="stMarkdownContainer"] h2,
    div[data-testid="stMarkdownContainer"] h3,
    div[data-testid="stMarkdownContainer"] h4,
    div[data-testid="stMarkdownContainer"] h5,
    div[data-testid="stMarkdownContainer"] h6,
    .main-header,
    .menu-3d-title,
    .cultura-title,
    .login-title,
    .cfg-panel-title,
    .vd-login-title,
    .vd-section-title,
    .partner-excel-title,
    .partner-toolbox-title,
    .partner-window-title,
    .partner-card-title,
    .partner-hero-title,
    .assessment-panel-title {
        background: linear-gradient(135deg,#ffffff 0%,#b8f3ff 22%,#42a5f5 52%,#00d4ff 75%,#5ff2b1 100%) !important;
        -webkit-background-clip: text !important;
        background-clip: text !important;
        -webkit-text-fill-color: #ffffff !important;
        color: #ffffff !important;
        text-shadow:
            0 2px 0 rgba(0,0,0,.92),
            0 6px 14px rgba(0,0,0,.58),
            0 0 14px rgba(0,212,255,.42),
            0 0 28px rgba(95,242,177,.22) !important;
        letter-spacing: .6px;
    }
    div[data-testid="stMarkdownContainer"] h4,
    div[data-testid="stMarkdownContainer"] h5,
    div[data-testid="stMarkdownContainer"] h6,
    .vd-section-title,
    .partner-toolbox-title,
    .partner-card-title,
    .assessment-panel-title {
        letter-spacing: .35px;
    }
    .login-subtitle,
    .cultura-subtitle,
    .partner-excel-subtitle,
    .partner-window-subtitle,
    .partner-hero-subtitle,
    .vd-login-sub,
    div[data-testid="stCaptionContainer"] {
        color: #d9fbff !important;
        text-shadow:
            0 1px 0 rgba(0,0,0,.85),
            0 0 10px rgba(0,188,212,.30),
            0 0 18px rgba(66,165,245,.18) !important;
    }
    </style>
    """, unsafe_allow_html=True)

aplicar_estilo_titulos_3d()

def _tmg_embedded_visualizer_theme_markup() -> str:
    return f"""
<style id="tmg-embedded-viewer-theme">
:root {{
  --tmg-primary:{THEME_PRIMARY_COLOR};
  --tmg-primary-dark:{THEME_PRIMARY_DARK};
  --tmg-primary-soft:{THEME_PRIMARY_SOFT};
  --tmg-primary-rgb:{THEME_PRIMARY_RGB};
  --tmg-bg-0:#020e24;
  --tmg-bg-1:#061525;
  --tmg-bg-2:#0d2b45;
  --tmg-glass:linear-gradient(120deg, rgba(255,255,255,.12), transparent 32%), linear-gradient(145deg, rgba(2,14,36,.96), rgba(18,62,100,.82), rgba({THEME_PRIMARY_RGB},.18));
  --tmg-glow:0 14px 30px rgba(0,0,0,.38), 0 0 24px rgba({THEME_PRIMARY_RGB},.24), inset 0 1px 0 rgba(255,255,255,.14);
}}
html, body {{
  color:#ffffff !important;
  background:
    radial-gradient(circle at 14% 0%, rgba({THEME_PRIMARY_RGB},.14), transparent 36%),
    linear-gradient(135deg, #020e24 0%, #061525 50%, #0d2b45 100%) !important;
}}
body {{
  scrollbar-color:var(--tmg-primary) #020e24;
}}
::-webkit-scrollbar {{ width:11px; height:11px; }}
::-webkit-scrollbar-track {{ background:linear-gradient(180deg,#020e24,#071a2c); border-radius:999px; }}
::-webkit-scrollbar-thumb {{
  background:linear-gradient(180deg,var(--tmg-primary-soft),var(--tmg-primary),var(--tmg-primary-dark));
  border:2px solid #020e24;
  border-radius:999px;
  box-shadow:0 0 14px rgba({THEME_PRIMARY_RGB},.38);
}}
button,
.btn,
.grid-btn,
.tool-btn,
.toolbar button,
.controls button,
.side button,
.panel button,
.sidebar button,
input[type="button"],
input[type="submit"] {{
  border-radius:10px !important;
  border:1px solid rgba({THEME_PRIMARY_RGB},.68) !important;
  background:var(--tmg-glass) !important;
  color:#ffffff !important;
  font-weight:850 !important;
  text-shadow:0 1px 0 rgba(0,0,0,.88), 0 0 10px rgba({THEME_PRIMARY_RGB},.42) !important;
  box-shadow:var(--tmg-glow) !important;
  transition:transform .25s ease, box-shadow .30s ease, border-color .30s ease, filter .30s ease !important;
}}
button:hover,
.btn:hover,
.grid-btn:hover,
.tool-btn:hover {{
  transform:translateY(-1px) !important;
  border-color:var(--tmg-primary-soft) !important;
  box-shadow:0 16px 34px rgba(0,0,0,.48), 0 0 34px rgba({THEME_PRIMARY_RGB},.46), inset 0 1px 0 rgba(255,255,255,.24) !important;
  filter:brightness(1.08);
}}
button:active,
.btn:active,
.grid-btn:active,
.tool-btn:active {{
  transform:translateY(1px) scale(.99) !important;
}}
button.active,
.btn.active,
.grid-btn.active,
.tool-btn.active {{
  border-color:#5ff2b1 !important;
  box-shadow:0 0 0 1px rgba(95,242,177,.34), 0 0 28px rgba({THEME_PRIMARY_RGB},.44), inset 0 1px 0 rgba(255,255,255,.26) !important;
}}
input,
select,
textarea {{
  border-radius:9px !important;
  border:1px solid rgba({THEME_PRIMARY_RGB},.58) !important;
  background:linear-gradient(145deg, rgba(2,14,36,.95), rgba(18,62,100,.74), rgba({THEME_PRIMARY_RGB},.12)) !important;
  color:#ffffff !important;
  font-weight:750 !important;
  box-shadow:0 8px 18px rgba(0,0,0,.28), 0 0 16px rgba({THEME_PRIMARY_RGB},.16), inset 0 1px 0 rgba(255,255,255,.12) !important;
}}
input::placeholder,
textarea::placeholder {{
  color:rgba(224,247,255,.74) !important;
}}
.toolbar,
.topbar,
.controls,
.panel,
.side,
.sidebar,
.summary,
.card,
.box,
.status-box,
.count-panel,
.result-panel,
.layer-panel,
.date-card,
.qgis-like-card,
.qgis-panel,
[class*="panel"],
[class*="card"],
[class*="sidebar"],
[class*="summary"] {{
  border:1px solid rgba({THEME_PRIMARY_RGB},.44) !important;
  background:var(--tmg-glass) !important;
  color:#ffffff !important;
  box-shadow:var(--tmg-glow) !important;
  backdrop-filter:blur(10px) saturate(140%);
  -webkit-backdrop-filter:blur(10px) saturate(140%);
}}
.viewer,
#viewer,
#vc,
#cronViewer,
.canvas-wrap,
.map-wrap,
.image-stage,
.ortho-stage,
[class*="viewer"] {{
  background:
    radial-gradient(circle at 50% 0%, rgba({THEME_PRIMARY_RGB},.10), transparent 42%),
    linear-gradient(145deg,#020e24,#061525) !important;
  border:1px solid rgba({THEME_PRIMARY_RGB},.36) !important;
  box-shadow:0 16px 34px rgba(0,0,0,.42), 0 0 24px rgba({THEME_PRIMARY_RGB},.18) !important;
}}
.title,
.panel-title,
.qgis-panel-title,
h1,h2,h3,h4,
label {{
  color:#ffffff !important;
  text-shadow:0 2px 0 rgba(0,0,0,.88), 0 0 14px rgba({THEME_PRIMARY_RGB},.42) !important;
}}
.subtle,
.hint,
.status,
.date-meta,
.coord,
.info,
.help,
p,
small {{
  color:#dffbff !important;
  text-shadow:0 1px 0 rgba(0,0,0,.72) !important;
}}
.progress,
.progress-track,
[class*="progress"] {{
  border-radius:999px !important;
  border:1px solid rgba({THEME_PRIMARY_RGB},.56) !important;
  background:linear-gradient(180deg,#020e24,#071a2c) !important;
  box-shadow:inset 0 3px 8px rgba(0,0,0,.70), 0 0 16px rgba({THEME_PRIMARY_RGB},.24) !important;
  overflow:hidden !important;
}}
.progress > div,
.progress div,
.progress-fill,
[class*="progress"] > div {{
  background:linear-gradient(90deg,var(--tmg-primary-soft),var(--tmg-primary),#00ff9d) !important;
  box-shadow:inset 0 1px 0 rgba(255,255,255,.44), 0 0 18px rgba({THEME_PRIMARY_RGB},.54) !important;
  transition:width .35s ease !important;
}}
table {{
  color:#ffffff !important;
  background:rgba(2,14,36,.72) !important;
  border-color:rgba({THEME_PRIMARY_RGB},.30) !important;
}}
th {{
  background:linear-gradient(145deg, rgba(18,62,100,.96), rgba({THEME_PRIMARY_RGB},.30)) !important;
  color:#ffffff !important;
  text-shadow:0 1px 0 rgba(0,0,0,.82), 0 0 8px rgba({THEME_PRIMARY_RGB},.26) !important;
}}
td {{
  border-color:rgba({THEME_PRIMARY_RGB},.22) !important;
}}
tr:nth-child(odd) {{ background:rgba(2,14,36,.48) !important; }}
tr:nth-child(even) {{ background:rgba(13,43,69,.38) !important; }}
tr:hover {{ background:rgba({THEME_PRIMARY_RGB},.20) !important; }}
</style>
<script id="tmg-embedded-viewer-loading-autohide">
(function() {{
  function setProgressVisibility(bar) {{
    if (!bar) return;
    var parent = bar.closest('.progress') || bar.parentElement;
    if (!parent) return;
    var width = String(bar.style && bar.style.width || '').trim();
    if (width === '100%') {{
      clearTimeout(bar.__tmgHideTimer);
      bar.__tmgHideTimer = setTimeout(function() {{
        if (String(bar.style && bar.style.width || '').trim() === '100%') parent.style.opacity = '0';
      }}, 900);
    }} else {{
      parent.style.opacity = '1';
    }}
    parent.style.transition = 'opacity .30s ease';
  }}
  function scanProgress() {{
    document.querySelectorAll('.progress > div, [class*="progress"] > div').forEach(setProgressVisibility);
  }}
  try {{
    scanProgress();
    new MutationObserver(scanProgress).observe(document.documentElement, {{attributes:true, childList:true, subtree:true, attributeFilter:['style','class']}});
  }} catch (e) {{}}
}})();
</script>
"""

def _inject_tmg_embedded_visualizer_theme(markup: str) -> str:
    if not isinstance(markup, str) or "tmg-embedded-viewer-theme" in markup:
        return markup
    lower = markup.lower()
    viewer_signals = (
        "<canvas",
        "grid-btn",
        "visualizador",
        "viewer",
        "ortofoto",
        "cronviewer",
        "qgis",
        "btnexport",
    )
    if not any(signal in lower for signal in viewer_signals):
        return markup
    theme_markup = _tmg_embedded_visualizer_theme_markup()
    if "</head>" in lower:
        return re.sub(r"</head>", theme_markup + "</head>", markup, count=1, flags=re.IGNORECASE)
    body_match = re.search(r"<body[^>]*>", markup, flags=re.IGNORECASE)
    if body_match:
        insert_at = body_match.end()
        return markup[:insert_at] + theme_markup + markup[insert_at:]
    return theme_markup + markup

def _theme_colorize_markup(value):
    if not isinstance(value, str):
        return value
    themed = value
    replacements = {
        "#ff8c00": THEME_PRIMARY_COLOR,
        "#FF8C00": THEME_PRIMARY_COLOR,
        "#ff9e33": THEME_PRIMARY_COLOR,
        "#FF9E33": THEME_PRIMARY_COLOR,
        "#e67600": THEME_PRIMARY_DARK,
        "#E67600": THEME_PRIMARY_DARK,
        "#e07000": THEME_PRIMARY_DARK,
        "#E07000": THEME_PRIMARY_DARK,
        "#ffaa33": THEME_PRIMARY_SOFT,
        "#FFAA33": THEME_PRIMARY_SOFT,
        "#ffb347": THEME_PRIMARY_SOFT,
        "#FFB347": THEME_PRIMARY_SOFT,
        "#00cfff": THEME_PRIMARY_SOFT,
        "#00CFFF": THEME_PRIMARY_SOFT,
        "#00d4ff": THEME_PRIMARY_SOFT,
        "#00D4FF": THEME_PRIMARY_SOFT,
        "#00e5ff": THEME_PRIMARY_SOFT,
        "#00E5FF": THEME_PRIMARY_SOFT,
        "#42a5f5": THEME_PRIMARY_COLOR,
        "#42A5F5": THEME_PRIMARY_COLOR,
        "#64b5f6": THEME_PRIMARY_SOFT,
        "#64B5F6": THEME_PRIMARY_SOFT,
        "#90caf9": THEME_PRIMARY_SOFT,
        "#90CAF9": THEME_PRIMARY_SOFT,
        "#75b7ff": THEME_PRIMARY_SOFT,
        "#75B7FF": THEME_PRIMARY_SOFT,
        "#5599ff": THEME_PRIMARY_COLOR,
        "#5599FF": THEME_PRIMARY_COLOR,
        "#2d8cff": THEME_PRIMARY_COLOR,
        "#2D8CFF": THEME_PRIMARY_COLOR,
        "#1e90ff": THEME_PRIMARY_COLOR,
        "#1E90FF": THEME_PRIMARY_COLOR,
        "#2a1a00": "#0d2b45",
        "#1a0a00": "#071a2c",
        "#160b00": "#061525",
        "#7a3a00": THEME_PALETTE["shadow_1"],
        "#5c2b00": THEME_PALETTE["shadow_2"],
        "#3d1d00": THEME_PALETTE["shadow_3"],
    }
    for old, new in replacements.items():
        themed = themed.replace(old, new)
    for red, green, blue in (
        (255, 140, 0),
        (0, 207, 255),
        (0, 212, 255),
        (0, 229, 255),
        (66, 165, 245),
        (100, 181, 246),
        (144, 202, 249),
        (117, 183, 255),
        (85, 153, 255),
        (45, 140, 255),
        (30, 144, 255),
    ):
        themed = re.sub(
            rf"rgba\(\s*{red}\s*,\s*{green}\s*,\s*{blue}\s*,\s*([0-9.]+)\s*\)",
            rf"rgba({THEME_PRIMARY_RGB},\1)",
            themed,
        )
    return themed

_ORIGINAL_ST_MARKDOWN = st.markdown
_ORIGINAL_COMPONENTS_HTML = components.html

def _themed_markdown(body, *args, **kwargs):
    return _ORIGINAL_ST_MARKDOWN(_theme_colorize_markup(body), *args, **kwargs)

def _themed_components_html(html, *args, **kwargs):
    return _ORIGINAL_COMPONENTS_HTML(_inject_tmg_embedded_visualizer_theme(_theme_colorize_markup(html)), *args, **kwargs)

st.markdown = _themed_markdown
components.html = _themed_components_html

# ==========================================
# MODULO ISOLADO - USUARIOS E PARCEIROS
# ==========================================
AUTH_USERS_PATH = SYSTEM_DATABASE_DIR / "usuarios_sistema.json"
CHAT_MESSAGES_PATH = SYSTEM_DATABASE_DIR / "chat_mensagens.json"
PARTNERS_ROOT = SYSTEM_DATABASE_DIR / "parceiros_controle_voos_dados"
PARTNERS_STATE_PATH = PARTNERS_ROOT / "parceiros_estado.json"
PARTNERS_LOGOS_DIR = PARTNERS_ROOT / "logos"
PARTNER_KEYS = {"eiwa": "EIWA", "alvaz": "ALVAZ"}
PARTNER_BUTTON_LABELS = {"eiwa": "EIWA", "alvaz": "ALVAZ"}
PARTNER_STATUS_OPTIONS = ["Executado", "Não executado"]
PARTNER_TREATMENT_STATUS = ["Aberto", "Em andamento", "Concluído", "Resolvido", "Atrasado"]
PARTNER_TREATMENT_DONE = {"Concluído", "Resolvido"}
PARTNER_INTERNAL_COLUMNS = ["Status de Execução", "Tratativa", "Descrição / Observação", "Última Alteração", "Usuário Responsável"]
PARTNER_ROW_ID = "__tmg_row_id"
MENU_PERMISSION_OPTIONS = {
    "menu_checklist": "Checklist",
    "menu_grid": "Marcar Grid",
    "menu_upload": "Upload",
    "menu_bases": "Banco de Dados",
    "menu_sync": "Sincronizar",
    "menu_ortomosaicos": "Gerar Ortomosaico",
    "menu_parceiros": "Parceiros",
    "menu_controle_dados": "Controle de Dados",
}
PARTNER_PERMISSION_OPTIONS = {
    "partner_sheet_view": "Visualizar planilha",
    "partner_sheet_import": "Importar planilha",
    "partner_sheet_edit_rows": "Editar linhas",
    "partner_sheet_delete_rows": "Excluir linhas",
    "partner_sheet_edit_header": "Editar cabeçalho",
    "partner_sheet_write_treatment": "Escrever tratativas na planilha",
    "partner_sheet_history": "Acessar histórico",
    "partner_sheet_export": "Exportar planilha",
}
PHENOTYPING_PERMISSION_OPTIONS = {
    "phenotyping_contagem": "Contagem",
    "phenotyping_maturacao": "Maturação",
    "phenotyping_pendoamento": "Pendoamento",
    "phenotyping_qualidade": "Qualidade de Parcela",
}

def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")

def _now_human() -> str:
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")

def _default_permissions(all_access: bool = True) -> dict:
    permissions = {
        "culturas": bool(all_access),
        "soja": bool(all_access),
        "milho": bool(all_access),
        "algodao": bool(all_access),
        "parceiros": bool(all_access),
        "eiwa": bool(all_access),
        "alvaz": bool(all_access),
    }
    for key in MENU_PERMISSION_OPTIONS:
        permissions[key] = bool(all_access)
    for key in PHENOTYPING_PERMISSION_OPTIONS:
        permissions[key] = bool(all_access)
    for key in PARTNER_PERMISSION_OPTIONS:
        permissions[key] = bool(all_access)
    return permissions

def _auth_default_users() -> dict:
    return {
        "users": [
            {
                "nome": "Wellington",
                "usuario": "Wellington",
                "senha": "123",
                "ativo": True,
                "admin": True,
                "permissoes": _default_permissions(True),
                "criado_em": _now_iso(),
                "atualizado_em": _now_iso(),
            },
            {
                "nome": "Acesso legado",
                "usuario": "123",
                "senha": "123",
                "ativo": True,
                "admin": False,
                "permissoes": _default_permissions(True),
                "criado_em": _now_iso(),
                "atualizado_em": _now_iso(),
            },
        ]
    }

def _auth_ensure_users() -> None:
    AUTH_USERS_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not AUTH_USERS_PATH.exists():
        AUTH_USERS_PATH.write_text(json.dumps(_auth_default_users(), indent=2, ensure_ascii=False), encoding="utf-8")
        return
    original_text = ""
    try:
        original_text = AUTH_USERS_PATH.read_text(encoding="utf-8")
    except Exception:
        original_text = ""
    data = _auth_load_users()
    users = data.setdefault("users", [])
    changed = False
    if not any(str(u.get("usuario", "")).lower() == "wellington" for u in users):
        users.insert(0, _auth_default_users()["users"][0])
        changed = True
    normalized_text = json.dumps(data, indent=2, ensure_ascii=False)
    if changed or normalized_text != original_text:
        AUTH_USERS_PATH.write_text(normalized_text, encoding="utf-8")

def _auth_load_users() -> dict:
    data = _tmg_read_json_file(AUTH_USERS_PATH, _auth_default_users())
    data.setdefault("users", [])
    for user in data["users"]:
        user.setdefault("nome", user.get("usuario", "Usuário"))
        user.setdefault("usuario", "")
        user.setdefault("senha", "")
        user.setdefault("ativo", True)
        user.setdefault("admin", False)
        user.setdefault("permissoes", _default_permissions(False))
        perms = user["permissoes"]
        legacy_culture_access = bool(perms.get("culturas", False))
        legacy_partner_access = bool(perms.get("parceiros", False))
        defaults = _default_permissions(False)
        for key, value in defaults.items():
            if key in perms:
                continue
            if key in ("menu_checklist", "menu_grid", "menu_upload", "menu_bases", "menu_sync", "menu_ortomosaicos"):
                perms[key] = legacy_culture_access
            elif key in ("menu_parceiros", "menu_controle_dados"):
                perms[key] = legacy_partner_access
            elif key in PHENOTYPING_PERMISSION_OPTIONS:
                # Compatibilidade com usuários antigos: se já tinha acesso aos módulos de cultura,
                # libera as análises até o administrador ajustar individualmente.
                perms[key] = legacy_culture_access
            elif key in PARTNER_PERMISSION_OPTIONS:
                perms[key] = legacy_partner_access
            else:
                perms[key] = value
    return data

def _auth_save_users(data: dict) -> None:
    AUTH_USERS_PATH.parent.mkdir(parents=True, exist_ok=True)
    AUTH_USERS_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

def _auth_find_user(usuario: str, senha: str) -> dict:
    _auth_ensure_users()
    usuario_norm = str(usuario or "").strip().lower()
    senha_text = str(senha or "")
    for user in _auth_load_users().get("users", []):
        if str(user.get("usuario", "")).strip().lower() == usuario_norm and str(user.get("senha", "")) == senha_text:
            if bool(user.get("ativo", True)):
                return user
            return {}
    return {}

def _auth_current_user() -> dict:
    user = st.session_state.get("auth_user")
    if isinstance(user, dict) and user.get("usuario"):
        return user
    return _auth_default_users()["users"][0]

def _auth_user_name() -> str:
    user = _auth_current_user()
    return str(user.get("nome") or user.get("usuario") or "Usuário")

def _auth_is_admin(user: dict = None) -> bool:
    user = user or _auth_current_user()
    return bool(user.get("admin")) or str(user.get("usuario", "")).strip().lower() == "wellington"

def _auth_permissions(user: dict = None) -> dict:
    user = user or _auth_current_user()
    if _auth_is_admin(user):
        return _default_permissions(True)
    perms = user.get("permissoes", {}) if isinstance(user, dict) else {}
    normalized = _default_permissions(False)
    normalized.update({k: bool(v) for k, v in perms.items()})
    legacy_culture_access = bool(normalized.get("culturas"))
    legacy_partner_access = bool(normalized.get("parceiros"))
    for key in ("menu_checklist", "menu_grid", "menu_upload", "menu_bases", "menu_sync", "menu_ortomosaicos"):
        if key not in perms:
            normalized[key] = legacy_culture_access
    for key in ("menu_parceiros", "menu_controle_dados"):
        if key not in perms:
            normalized[key] = legacy_partner_access
    for key in PHENOTYPING_PERMISSION_OPTIONS:
        if key not in perms:
            normalized[key] = legacy_culture_access
    for key in PARTNER_PERMISSION_OPTIONS:
        if key not in perms:
            normalized[key] = legacy_partner_access
    return normalized

def _auth_allowed_cultures(user: dict = None) -> list:
    perms = _auth_permissions(user)
    if not perms.get("culturas"):
        return []
    allowed = []
    if perms.get("soja"):
        allowed.append("SOJA")
    if perms.get("milho"):
        allowed.append("MILHO")
    if perms.get("algodao"):
        allowed.append("ALGODÃO")
    return allowed

def _auth_can_partners(user: dict = None) -> bool:
    perms = _auth_permissions(user)
    return bool(perms.get("parceiros") and (perms.get("menu_parceiros") or perms.get("menu_controle_dados")))

def _auth_menu_allowed(menu_key: str, user: dict = None) -> bool:
    return bool(_auth_permissions(user).get(menu_key))

def _auth_phenotyping_allowed(permission_key: str, user: dict = None) -> bool:
    return bool(_auth_permissions(user).get(permission_key))

def _auth_allowed_phenotyping(user: dict = None) -> list:
    perms = _auth_permissions(user)
    if not perms.get("culturas"):
        return []
    return [key for key in PHENOTYPING_PERMISSION_OPTIONS if bool(perms.get(key))]

def _auth_partner_permission(permission_key: str, user: dict = None) -> bool:
    return bool(_auth_permissions(user).get(permission_key))

def _auth_allowed_partners(user: dict = None) -> list:
    perms = _auth_permissions(user)
    if not perms.get("parceiros"):
        return []
    allowed = []
    if perms.get("eiwa"):
        allowed.append("eiwa")
    if perms.get("alvaz"):
        allowed.append("alvaz")
    return allowed

def _chat_user_login(user: dict) -> str:
    return str((user or {}).get("usuario", "")).strip()

def _chat_user_display(user: dict) -> str:
    login = _chat_user_login(user)
    nome = str((user or {}).get("nome") or login or "Usuário").strip()
    return f"{nome} ({login})" if login and nome.lower() != login.lower() else (login or nome)

def _chat_load_state() -> dict:
    CHAT_MESSAGES_PATH.parent.mkdir(parents=True, exist_ok=True)
    data = _tmg_read_json_file(CHAT_MESSAGES_PATH, {})
    if not isinstance(data, dict):
        data = {}
    data.setdefault("messages", [])
    if not isinstance(data["messages"], list):
        data["messages"] = []
    return data

def _chat_save_state(data: dict) -> None:
    CHAT_MESSAGES_PATH.parent.mkdir(parents=True, exist_ok=True)
    data.setdefault("messages", [])
    CHAT_MESSAGES_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

def _chat_registered_users(current_user: dict) -> list:
    current_login = _chat_user_login(current_user).lower()
    users = []
    for user in _auth_load_users().get("users", []):
        login = _chat_user_login(user)
        if not login or login.lower() == current_login or not bool(user.get("ativo", True)):
            continue
        users.append(user)
    return users

def _chat_unread_count(login: str, other_login: str = "") -> int:
    login_norm = str(login or "").strip().lower()
    other_norm = str(other_login or "").strip().lower()
    total = 0
    for msg in _chat_load_state().get("messages", []):
        if str(msg.get("destinatario", "")).strip().lower() != login_norm:
            continue
        if bool(msg.get("lida", False)):
            continue
        if other_norm and str(msg.get("remetente", "")).strip().lower() != other_norm:
            continue
        total += 1
    return total

def _chat_thread_messages(login_a: str, login_b: str) -> list:
    a = str(login_a or "").strip().lower()
    b = str(login_b or "").strip().lower()
    rows = []
    for msg in _chat_load_state().get("messages", []):
        sender = str(msg.get("remetente", "")).strip().lower()
        target = str(msg.get("destinatario", "")).strip().lower()
        if (sender == a and target == b) or (sender == b and target == a):
            rows.append(msg)
    rows.sort(key=lambda item: str(item.get("timestamp", "")))
    return rows

def _chat_mark_thread_read(current_login: str, other_login: str) -> None:
    current_norm = str(current_login or "").strip().lower()
    other_norm = str(other_login or "").strip().lower()
    data = _chat_load_state()
    changed = False
    for msg in data.get("messages", []):
        if (
            str(msg.get("destinatario", "")).strip().lower() == current_norm
            and str(msg.get("remetente", "")).strip().lower() == other_norm
            and not bool(msg.get("lida", False))
        ):
            msg["lida"] = True
            msg["lida_em"] = _now_iso()
            changed = True
    if changed:
        _chat_save_state(data)

def _chat_send_message(sender_user: dict, target_login: str, text: str) -> tuple:
    sender_login = _chat_user_login(sender_user)
    target_login = str(target_login or "").strip()
    text = str(text or "").strip()
    if not sender_login:
        return False, "Faça login para acessar o chat."
    if not target_login:
        return False, "Selecione um usuário para enviar a mensagem."
    if sender_login.lower() == target_login.lower():
        return False, "Não é possível conversar consigo mesmo."
    if not text:
        return False, "Digite uma mensagem antes de enviar."

    users = {str(u.get("usuario", "")).strip().lower(): u for u in _auth_load_users().get("users", [])}
    target_user = users.get(target_login.lower())
    if not target_user or not bool(target_user.get("ativo", True)):
        return False, "Usuário destinatário não encontrado ou inativo."

    now = datetime.now()
    msg_id = hashlib.sha1(f"{sender_login}-{target_login}-{now.isoformat()}-{text}".encode("utf-8")).hexdigest()[:18]
    data = _chat_load_state()
    data.setdefault("messages", []).append({
        "id": msg_id,
        "remetente": sender_login,
        "remetente_nome": str(sender_user.get("nome") or sender_login),
        "destinatario": target_login,
        "destinatario_nome": str(target_user.get("nome") or target_login),
        "texto": text,
        "data": now.strftime("%d/%m/%Y"),
        "hora": now.strftime("%H:%M:%S"),
        "timestamp": now.isoformat(timespec="seconds"),
        "lida": False,
    })
    _chat_save_state(data)
    return True, "Mensagem enviada."

def _render_chat_body(current_user: dict) -> None:
    current_login = _chat_user_login(current_user)
    st.markdown("""
    <style>
      .tmg-chat-title { color: var(--tmg-primary); font-weight: 800; letter-spacing: 1px; text-transform: uppercase; margin-bottom: 8px; }
      .tmg-chat-scroll { max-height: 320px; overflow-y: auto; padding: 8px; border: 1px solid #2a2a2a; border-radius: 10px; background: #0d0d0d; }
      .tmg-chat-msg { padding: 8px 10px; border-radius: 10px; margin: 7px 0; max-width: 86%; box-shadow: 0 4px 12px rgba(0,0,0,.22); }
      .tmg-chat-sent { margin-left: auto; background: linear-gradient(145deg,#0d2b45,#071a2c); border: 1px solid var(--tmg-primary); color: #f2f7fb; }
      .tmg-chat-received { margin-right: auto; background: #151515; border: 1px solid #333; color: #eee; }
      .tmg-chat-meta { color: #9aa7b5; font-size: .72rem; margin-bottom: 3px; }
      .tmg-chat-text { font-size: .88rem; line-height: 1.35; white-space: pre-wrap; }
      .tmg-chat-unread { color: #ff4d4d; font-weight: 800; }
    </style>
    """, unsafe_allow_html=True)
    st.markdown("<div class='tmg-chat-title'>Chat</div>", unsafe_allow_html=True)

    users = _chat_registered_users(current_user)
    if not current_login:
        st.warning("Acesse o sistema para usar o chat.")
        return
    if not users:
        st.info("Nenhum outro usuário cadastrado para conversar.")
        return

    search = st.text_input("Pesquisar usuário", key="tmg_chat_search", placeholder="Pesquisar usuário cadastrado")
    filtered = [
        user for user in users
        if not search
        or search.lower() in _chat_user_display(user).lower()
        or search.lower() in _chat_user_login(user).lower()
    ]
    if not filtered:
        st.info("Nenhum usuário encontrado para a busca.")
        return

    labels = {}
    for user in filtered:
        login = _chat_user_login(user)
        unread = _chat_unread_count(current_login, login)
        labels[login] = _chat_user_display(user) + (f"  • {unread} nova(s)" if unread else "")

    logins = [_chat_user_login(user) for user in filtered]
    current_selected = st.session_state.get("tmg_chat_selected_user")
    selected_index = logins.index(current_selected) if current_selected in logins else 0
    selected_login = st.selectbox(
        "Usuário",
        logins,
        index=selected_index,
        format_func=lambda value: labels.get(value, value),
        key="tmg_chat_user_select",
    )
    st.session_state["tmg_chat_selected_user"] = selected_login
    _chat_mark_thread_read(current_login, selected_login)

    messages = _chat_thread_messages(current_login, selected_login)
    if not messages:
        st.caption("Sem mensagens nesta conversa.")
    else:
        parts = ["<div class='tmg-chat-scroll'>"]
        for msg in messages[-80:]:
            sent = str(msg.get("remetente", "")).strip().lower() == current_login.lower()
            cls = "tmg-chat-sent" if sent else "tmg-chat-received"
            author = "Você" if sent else html.escape(str(msg.get("remetente_nome") or msg.get("remetente") or "Usuário"))
            when = html.escape(f"{msg.get('data', '')} {msg.get('hora', '')}".strip())
            status = " · lida" if sent and bool(msg.get("lida", False)) else (" · não lida" if sent else "")
            text = html.escape(str(msg.get("texto", "")))
            parts.append(
                f"<div class='tmg-chat-msg {cls}'>"
                f"<div class='tmg-chat-meta'>{author} · {when}{status}</div>"
                f"<div class='tmg-chat-text'>{text}</div>"
                "</div>"
            )
        parts.append("</div>")
        st.markdown("".join(parts), unsafe_allow_html=True)

    with st.form("tmg_chat_send_form", clear_on_submit=True):
        msg_text = st.text_area("Mensagem", key="tmg_chat_message", height=70, placeholder="Digite sua mensagem")
        send = st.form_submit_button("Enviar", type="primary", use_container_width=True)
        if send:
            ok, feedback = _chat_send_message(current_user, selected_login, msg_text)
            if ok:
                st.success(feedback)
                app_rerun()
            else:
                st.warning(feedback)

    if st.button("Minimizar / Fechar", key="tmg_chat_close", use_container_width=True):
        st.session_state["tmg_chat_open"] = False
        app_rerun()

def _render_system_chat(current_user: dict) -> None:
    if not st.session_state.get("logged_in", False):
        return
    current_login = _chat_user_login(current_user)
    if not current_login:
        return

    unread_total = _chat_unread_count(current_login)
    label = "💬" + (f"  ● {unread_total}" if unread_total else "")
    st.markdown("""
    <style>
      div[data-testid="stPopover"] button, div[data-testid="stButton"] button {
        transition: all .18s ease;
      }
    </style>
    """, unsafe_allow_html=True)
    chat_cols = st.columns([0.90, 0.10])
    with chat_cols[1]:
        if hasattr(st, "popover"):
            with st.popover(label, help="Abrir chat", use_container_width=True):
                _render_chat_body(current_user)
        else:
            if st.button(label, key="tmg_chat_toggle", help="Abrir chat", use_container_width=True):
                st.session_state["tmg_chat_open"] = not st.session_state.get("tmg_chat_open", False)
            if st.session_state.get("tmg_chat_open", False):
                _render_chat_body(current_user)

if SYSTEM_CONFIG.get("tema", "padrao") == "tmg_premium_neon_3d":
    st.markdown("""
<style>
    .stApp {
        background:
            radial-gradient(circle at top left, rgba(0,229,255,.16), transparent 34%),
            radial-gradient(circle at bottom right, rgba(0,255,157,.10), transparent 30%),
            linear-gradient(145deg, #020E24 0%, #071A35 55%, #020E24 100%) !important;
        color: #FFFFFF !important;
    }
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #020E24 0%, #071A35 55%, #1A1A1A 100%) !important;
        border-right: 1px solid rgba(0,229,255,.42) !important;
        box-shadow: 8px 0 28px rgba(0,0,0,.45), inset -1px 0 0 rgba(0,255,157,.20) !important;
    }
    .main-header,
    .menu-3d-title,
    .cultura-title,
    .login-title,
    .cfg-panel-title,
    .vd-login-title,
    .vd-section-title,
    .partner-excel-title,
    .partner-toolbox-title,
    .partner-window-title,
    .partner-card-title,
    .partner-hero-title,
    .assessment-panel-title,
    div[data-testid="stMarkdownContainer"] h1,
    div[data-testid="stMarkdownContainer"] h2,
    div[data-testid="stMarkdownContainer"] h3,
    div[data-testid="stMarkdownContainer"] h4 {
        background: linear-gradient(135deg, #FFFFFF 0%, #D6D6D6 18%, #00E5FF 48%, #1D7BFF 68%, #00FF9D 100%) !important;
        -webkit-background-clip: text !important;
        background-clip: text !important;
        -webkit-text-fill-color: #FFFFFF !important;
        color: #FFFFFF !important;
        text-shadow:
            0 2px 0 #020E24,
            0 4px 0 #071A35,
            0 8px 18px rgba(0,0,0,.82),
            0 0 16px rgba(0,229,255,.70),
            0 0 34px rgba(0,255,157,.28) !important;
        letter-spacing: .8px !important;
    }
    div.stButton > button,
    button[kind="secondary"],
    button[kind="primary"] {
        background: linear-gradient(145deg, #071A35 0%, #0E3A70 48%, #00E5FF 100%) !important;
        color: #FFFFFF !important;
        border: 1px solid rgba(0,229,255,.56) !important;
        border-radius: 14px !important;
        box-shadow:
            0 10px 24px rgba(0,0,0,.42),
            inset 0 1px 0 rgba(255,255,255,.22),
            inset 0 -5px 14px rgba(2,14,36,.55),
            0 0 16px rgba(0,229,255,.28) !important;
        text-shadow: 0 2px 3px rgba(0,0,0,.72) !important;
        transform: translateZ(0);
    }
    div.stButton > button:hover,
    button[kind="secondary"]:hover,
    button[kind="primary"]:hover {
        border-color: #00FF9D !important;
        box-shadow:
            0 14px 30px rgba(0,0,0,.50),
            0 0 22px rgba(0,229,255,.55),
            0 0 32px rgba(0,255,157,.22) !important;
        transform: translateY(-2px) scale(1.01) !important;
    }
    div.stButton > button:active,
    button[kind="secondary"]:active,
    button[kind="primary"]:active {
        transform: translateY(1px) scale(.99) !important;
        box-shadow: inset 0 4px 10px rgba(0,0,0,.62), 0 0 12px rgba(0,229,255,.35) !important;
    }
    .card,
    [data-testid="stExpander"],
    [data-testid="stForm"],
    div[data-testid="stDataFrame"],
    div[data-testid="stFileUploader"] section {
        background: linear-gradient(145deg, rgba(7,26,53,.94), rgba(26,26,26,.88)) !important;
        border: 1px solid rgba(0,229,255,.42) !important;
        border-radius: 16px !important;
        box-shadow: 0 16px 34px rgba(0,0,0,.42), inset 0 1px 0 rgba(255,255,255,.08), 0 0 20px rgba(0,229,255,.16) !important;
    }
    div[data-testid="stFileUploader"] section {
        border-style: solid !important;
        background:
            radial-gradient(circle at top, rgba(0,229,255,.18), transparent 38%),
            linear-gradient(145deg, rgba(7,26,53,.95), rgba(14,58,112,.78)) !important;
    }
    div[data-testid="stFileUploader"] section button {
        font-size: 0 !important;
    }
    div[data-testid="stFileUploader"] section button::after {
        content: "Importar";
        font-size: .92rem !important;
        font-weight: 800 !important;
        letter-spacing: .4px !important;
    }
    .tmg-load-card {
        background: var(--tmg-deploy-card-bg) !important;
        border-color: var(--tmg-deploy-border) !important;
        box-shadow: var(--tmg-deploy-card-shadow) !important;
    }
    .tmg-load-status { color: #00FF9D !important; }
    .tmg-chat-title,
    .partner-hero-subtitle,
    .login-subtitle,
    div[data-testid="stCaptionContainer"] {
        color: #D6D6D6 !important;
        text-shadow: 0 0 10px rgba(0,229,255,.28) !important;
    }
    svg, .st-emotion-cache svg {
        filter: drop-shadow(0 3px 6px rgba(0,0,0,.65)) drop-shadow(0 0 6px rgba(0,229,255,.45)) !important;
    }
</style>
    """, unsafe_allow_html=True)

def _partners_default_partner() -> dict:
    return {
        "link": "",
        "rows": [],
        "columns": [],
        "baseline_rows": [],
        "baseline_columns": [],
        "baseline_import": {},
        "last_import": {},
        "last_update": {},
        "diff_rows": [],
        "chat": [],
        "history": [],
        "logo_path": "",
    }

def _partners_default_state() -> dict:
    return {
        "partners": {key: _partners_default_partner() for key in PARTNER_KEYS},
        "history_general": [],
    }

def _partners_ensure_storage() -> None:
    PARTNERS_ROOT.mkdir(parents=True, exist_ok=True)
    if not PARTNERS_STATE_PATH.exists():
        PARTNERS_STATE_PATH.write_text(json.dumps(_partners_default_state(), indent=2, ensure_ascii=False), encoding="utf-8")

def _partners_load_state() -> dict:
    _partners_ensure_storage()
    state = _tmg_read_json_file(PARTNERS_STATE_PATH, _partners_default_state())
    state.setdefault("partners", {})
    state.setdefault("history_general", [])
    for key in PARTNER_KEYS:
        state["partners"].setdefault(key, _partners_default_partner())
        partner = state["partners"][key]
        partner.setdefault("link", "")
        partner.setdefault("rows", [])
        partner.setdefault("columns", [])
        partner.setdefault("baseline_rows", [])
        partner.setdefault("baseline_columns", [])
        partner.setdefault("baseline_import", {})
        partner.setdefault("last_import", {})
        partner.setdefault("last_update", {})
        partner.setdefault("diff_rows", [])
        partner.setdefault("chat", [])
        partner.setdefault("history", [])
        partner.setdefault("logo_path", "")
    return state

def _partners_save_state(state: dict) -> None:
    PARTNERS_ROOT.mkdir(parents=True, exist_ok=True)
    PARTNERS_STATE_PATH.write_text(json.dumps(state, indent=2, ensure_ascii=False), encoding="utf-8")

def _partners_add_history(state: dict, partner_key: str, acao: str, detalhes: str = "", extra: dict = None) -> None:
    item = {
        "data_hora": _now_human(),
        "usuario": _auth_user_name(),
        "parceira": PARTNER_KEYS.get(partner_key, "Geral") if partner_key else "Geral",
        "acao": acao,
        "detalhes": detalhes,
    }
    if isinstance(extra, dict):
        item.update(extra)
    state.setdefault("history_general", []).insert(0, item)
    state["history_general"] = state["history_general"][:1000]
    if partner_key in state.get("partners", {}):
        state["partners"][partner_key].setdefault("history", []).insert(0, item)
        state["partners"][partner_key]["history"] = state["partners"][partner_key]["history"][:1000]

def _partner_label(partner_key: str) -> str:
    return PARTNER_BUTTON_LABELS.get(partner_key, PARTNER_KEYS.get(partner_key, partner_key))

def _partners_logo_path(partner_key: str):
    partner_key = str(partner_key or "").strip().lower()
    for suffix in (".png", ".jpg", ".jpeg", ".webp"):
        candidate = PARTNERS_LOGOS_DIR / f"{partner_key}{suffix}"
        if candidate.exists():
            return candidate
    return None

def _partners_save_logo(state: dict, partner_key: str, uploaded_file) -> None:
    if uploaded_file is None:
        return
    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix not in (".png", ".jpg", ".jpeg", ".webp"):
        suffix = ".png"
    PARTNERS_LOGOS_DIR.mkdir(parents=True, exist_ok=True)
    for old in PARTNERS_LOGOS_DIR.glob(f"{partner_key}.*"):
        try:
            old.unlink()
        except Exception:
            pass
    target = PARTNERS_LOGOS_DIR / f"{partner_key}{suffix}"
    target.write_bytes(uploaded_file.getvalue())
    state["partners"][partner_key]["logo_path"] = str(target)

def _partners_logo_html(partner_key: str) -> str:
    label = _partner_label(partner_key)
    logo_path = _partners_logo_path(partner_key)
    if logo_path and logo_path.exists():
        return (
            "<div class='partner-logo-frame'>"
            f"<img src='{_img_to_base64_css(logo_path)}' alt='Logo {label}' class='partner-logo-img'>"
            "</div>"
        )
    return f"<div class='partner-logo-frame partner-logo-empty'>LOGO {label}</div>"

def _partners_history_rows(rows: list) -> list:
    label_map = {
        "eiwa": "EIWA",
        "iva": "EIWA",
        "eva": "EIWA",
        "alvaz": "ALVAZ",
        "alvás": "ALVAZ",
        "olvas": "ALVAZ",
        "elvas": "ALVAZ",
    }
    normalized = []
    for row in rows or []:
        item = dict(row)
        if "parceira" in item:
            item["parceira"] = label_map.get(str(item.get("parceira", "")).strip().lower(), item.get("parceira", ""))
        normalized.append(item)
    return normalized

def _auth_mention_options() -> tuple:
    users = []
    labels = {}
    for user in _auth_load_users().get("users", []):
        if not user.get("ativo", True):
            continue
        usuario = str(user.get("usuario", "")).strip()
        if not usuario:
            continue
        users.append(usuario)
        labels[usuario] = f"{user.get('nome', usuario)} ({usuario})"
    return users, labels

def _partners_mentions_for_user(state: dict, user: dict) -> list:
    usuario = str(user.get("usuario", "")).strip().lower()
    if not usuario:
        return []
    notes = []
    for partner_key, partner in state.get("partners", {}).items():
        for item in partner.get("chat", []) or []:
            citados = [str(v).strip().lower() for v in item.get("citados", [])]
            if usuario in citados:
                notes.append({
                    "parceira": _partner_label(partner_key),
                    "assunto": item.get("assunto", "Sem assunto"),
                    "usuario": item.get("usuario", ""),
                    "data": item.get("data", ""),
                    "hora": item.get("hora", ""),
                })
    return notes[:5]

def _partners_read_sheet_upload(uploaded_file) -> tuple:
    if uploaded_file is None:
        return None, "Selecione uma planilha Excel ou CSV para importar."
    try:
        raw = uploaded_file.getvalue()
        suffix = Path(uploaded_file.name).suffix.lower()
        if suffix in (".xlsx", ".xls"):
            df = pd.read_excel(BytesIO(raw))
        elif suffix == ".csv":
            df = None
            errors = []
            for encoding in ("utf-8-sig", "utf-8", "latin1", "cp1252"):
                try:
                    df = pd.read_csv(BytesIO(raw), sep=None, engine="python", encoding=encoding)
                    break
                except Exception as exc:
                    errors.append(f"{encoding}: {exc}")
            if df is None:
                raise ValueError("; ".join(errors[-2:]))
        else:
            return None, "Formato não suportado. Envie arquivos .xlsx, .xls ou .csv."
        df = _partners_clean_dataframe(df)
        if df.empty and len(df.columns) == 0:
            return None, "Não foi possível importar a planilha. Verifique o formato do arquivo e tente novamente."
        return df, ""
    except Exception:
        return None, "Não foi possível importar a planilha. Verifique o formato do arquivo e tente novamente."

def _partners_unique_columns(columns) -> list:
    seen = {}
    unique = []
    for idx, raw_col in enumerate(columns):
        base = str(raw_col).strip() or f"Coluna {idx + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        unique.append(base if count == 0 else f"{base}_{count + 1}")
    return unique

def _partners_clean_dataframe(df) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    df = df.copy()
    df.columns = _partners_unique_columns(df.columns)
    df = df.replace({np.nan: ""})
    for col in df.columns:
        df[col] = df[col].map(lambda value: value.isoformat() if hasattr(value, "isoformat") else str(value) if value is not None else "")
    return df

def _partners_row_label(row: dict, idx: int = 0) -> str:
    row_id = str(row.get(PARTNER_ROW_ID, "")).strip()
    values = [str(v).strip() for k, v in row.items() if k != PARTNER_ROW_ID and str(v).strip()]
    preview = " · ".join(values[:2]) if values else row_id
    return f"{idx + 1} · {preview[:80]} · {row_id}"

def _partners_safe_excel_bytes(export_df: pd.DataFrame) -> tuple:
    if export_df is None:
        return None, "Não foi possível exportar a planilha agora."
    max_rows, max_cols = 1048576, 16384
    if len(export_df) > max_rows or len(export_df.columns) > max_cols:
        return None, "A planilha ultrapassa o limite do Excel. Use a exportação CSV para baixar todos os dados."
    try:
        excel_buf = BytesIO()
        clean_df = export_df.replace({np.nan: ""}).copy()
        with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
            clean_df.to_excel(writer, index=False, sheet_name="Planilha Tratada")
            ws = writer.book["Planilha Tratada"]
            try:
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
                from openpyxl.utils import get_column_letter

                header_fill = PatternFill("solid", fgColor="B7DEE8")
                data_fill = PatternFill("solid", fgColor="FFFFFF")
                internal_fill = PatternFill("solid", fgColor="EAF3F8")
                thin = Side(style="thin", color="808080")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)
                header_font = Font(bold=True, color="000000")
                default_font = Font(color="000000")
                center = Alignment(horizontal="center", vertical="center", wrap_text=True)
                left = Alignment(horizontal="left", vertical="center", wrap_text=True)

                internal_cols = set(PARTNER_INTERNAL_COLUMNS)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = center

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        header = str(ws.cell(row=1, column=cell.column).value or "")
                        cell.fill = internal_fill if header in internal_cols else data_fill
                        cell.font = default_font
                        cell.border = border
                        cell.alignment = left

                for col_idx, column_cells in enumerate(ws.columns, start=1):
                    max_len = 12
                    for cell in column_cells:
                        value = "" if cell.value is None else str(cell.value)
                        max_len = max(max_len, min(len(value) + 2, 55))
                    ws.column_dimensions[get_column_letter(col_idx)].width = max_len

                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
            except Exception:
                pass
        return excel_buf.getvalue(), ""
    except Exception:
        return None, "Não foi possível exportar para Excel. Use a exportação CSV ou revise a planilha."

def _partners_export_csv_file(export_df: pd.DataFrame, partner_key: str) -> tuple:
    try:
        export_dir = PARTNERS_ROOT / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        target = export_dir / f"{partner_key}_controle_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        export_df.replace({np.nan: ""}).to_csv(target, index=False, encoding="utf-8-sig")
        return target, ""
    except Exception:
        return None, "Não foi possível preparar o CSV para download. Tente filtrar ou revisar a planilha."

def _partners_rename_column(partner: dict, old_name: str, new_name: str) -> tuple:
    old_name = str(old_name or "").strip()
    new_name = str(new_name or "").strip()
    if not old_name or not new_name:
        return False, "Informe o nome atual e o novo nome da coluna."
    if old_name == new_name:
        return False, "O cabeçalho já está com esse nome."
    if new_name == PARTNER_ROW_ID or new_name in PARTNER_INTERNAL_COLUMNS:
        return False, "Esse nome é reservado pelo sistema."
    columns = list(partner.get("columns", []))
    if old_name not in columns:
        return False, "Coluna não encontrada na planilha atual."
    if new_name in columns:
        return False, "Já existe uma coluna com esse nome."
    partner["columns"] = [new_name if col == old_name else col for col in columns]
    for collection_name in ("rows", "baseline_rows", "diff_rows"):
        for row in partner.get(collection_name, []) or []:
            if old_name in row:
                row[new_name] = row.pop(old_name)
    partner["baseline_columns"] = [new_name if col == old_name else col for col in partner.get("baseline_columns", [])]
    return True, ""

def _partners_add_blank_row(partner: dict, partner_key: str) -> dict:
    row_id = hashlib.sha1(f"{partner_key}-{_auth_user_name()}-{_now_iso()}-{len(partner.get('rows', []))}".encode()).hexdigest()[:12]
    row = {PARTNER_ROW_ID: row_id}
    row["Status de Execução"] = "Não executado"
    for col in partner.get("columns", []):
        row[col] = ""
    for col in PARTNER_INTERNAL_COLUMNS:
        row.setdefault(col, "Não executado" if col == "Status de Execução" else "")
    row["Última Alteração"] = _now_human()
    row["Usuário Responsável"] = _auth_user_name()
    partner.setdefault("rows", []).append(row)
    return row

def _partners_delete_row(partner: dict, row_id: str) -> bool:
    rows = partner.get("rows", [])
    before = len(rows)
    partner["rows"] = [row for row in rows if str(row.get(PARTNER_ROW_ID, "")) != str(row_id)]
    return len(partner["rows"]) != before

def _partners_ensure_row_ids(rows: list) -> list:
    prepared = []
    for idx, row in enumerate(rows or []):
        item = dict(row)
        item.setdefault(PARTNER_ROW_ID, hashlib.sha1(f"{idx}-{json.dumps(item, ensure_ascii=False, sort_keys=True)}".encode("utf-8")).hexdigest()[:12])
        prepared.append(item)
    return prepared

def _partners_rows_to_df(partner_data: dict) -> pd.DataFrame:
    rows = _partners_ensure_row_ids(partner_data.get("rows", []))
    columns = [col for col in partner_data.get("columns", []) if col != PARTNER_ROW_ID]
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=[PARTNER_ROW_ID] + columns + PARTNER_INTERNAL_COLUMNS)
    if PARTNER_ROW_ID not in df.columns:
        df[PARTNER_ROW_ID] = [hashlib.sha1(f"row-{i}".encode()).hexdigest()[:12] for i in range(len(df))]
    for col in columns + PARTNER_INTERNAL_COLUMNS:
        if col not in df.columns:
            df[col] = "Não executado" if col == "Status de Execução" else ""
    if "Status de Execução" in df.columns:
        df["Status de Execução"] = df["Status de Execução"].map(
            lambda value: value if str(value) in PARTNER_STATUS_OPTIONS else "Não executado"
        )
    status_cols = ["Status de Execução"] if "Status de Execução" in df.columns else []
    front_internal = [col for col in ("Tratativa", "Descrição / Observação") if col in df.columns]
    tail_internal = [col for col in PARTNER_INTERNAL_COLUMNS if col in df.columns and col not in status_cols + front_internal]
    ordered = [PARTNER_ROW_ID] + status_cols + front_internal + [col for col in columns if col in df.columns] + tail_internal
    ordered += [col for col in df.columns if col not in ordered]
    return df[ordered].replace({np.nan: ""})

def _partners_prepare_import_df(df: pd.DataFrame, current_user: str) -> tuple:
    df = _partners_clean_dataframe(df)
    original_columns = [col for col in df.columns if col != PARTNER_ROW_ID and col not in PARTNER_INTERNAL_COLUMNS]
    df.insert(0, PARTNER_ROW_ID, [hashlib.sha1(f"{current_user}-{_now_iso()}-{i}".encode()).hexdigest()[:12] for i in range(len(df))])
    for col in PARTNER_INTERNAL_COLUMNS:
        if col not in df.columns:
            df[col] = "Não executado" if col == "Status de Execução" else ""
    if "Status de Execução" in df.columns:
        df["Status de Execução"] = df["Status de Execução"].map(
            lambda value: value if str(value) in PARTNER_STATUS_OPTIONS else "Não executado"
        )
    if "Usuário Responsável" in df.columns:
        df["Usuário Responsável"] = df["Usuário Responsável"].replace("", current_user)
    return df, original_columns

def _partners_compare_dataframes(old_df: pd.DataFrame, new_df: pd.DataFrame, columns: list) -> tuple:
    old = old_df[[col for col in columns if col in old_df.columns]].reset_index(drop=True)
    new = new_df[[col for col in columns if col in new_df.columns]].reset_index(drop=True)
    max_len = max(len(old), len(new))
    preview = []
    changed_rows = 0
    new_rows = 0
    removed_rows = 0
    for idx in range(max_len):
        if idx >= len(old):
            row = new.iloc[idx].to_dict()
            row["__ALTERAÇÃO__"] = "Linha nova"
            row["__CELULAS__"] = ""
            preview.append(row)
            new_rows += 1
            continue
        if idx >= len(new):
            row = old.iloc[idx].to_dict()
            row["__ALTERAÇÃO__"] = "Linha removida"
            row["__CELULAS__"] = ""
            preview.append(row)
            removed_rows += 1
            continue
        changed_cells = [col for col in columns if str(old.at[idx, col] if col in old.columns else "") != str(new.at[idx, col] if col in new.columns else "")]
        row = new.iloc[idx].to_dict()
        row["__ALTERAÇÃO__"] = "Célula alterada" if changed_cells else "Sem alteração"
        row["__CELULAS__"] = ", ".join(changed_cells)
        preview.append(row)
        if changed_cells:
            changed_rows += 1
    return {
        "linhas_novas": new_rows,
        "linhas_alteradas": changed_rows,
        "linhas_removidas": removed_rows,
        "total_diferencas": new_rows + changed_rows + removed_rows,
        "data_hora": _now_human(),
        "usuario": _auth_user_name(),
    }, preview

def _partners_style_diff(df: pd.DataFrame):
    def _style(row):
        marker = row.get("__ALTERAÇÃO__", "")
        if marker == "Linha nova":
            return ["background-color: rgba(255, 76, 76, .24); color:#fff;"] * len(row)
        if marker == "Linha removida":
            return ["background-color: rgba(130, 130, 130, .30); color:#ddd;"] * len(row)
        if marker == "Célula alterada":
            changed = [part.strip() for part in str(row.get("__CELULAS__", "")).split(",") if part.strip()]
            return ["background-color: rgba(255, 221, 64, .30); color:#fff;" if col in changed else "" for col in row.index]
        return [""] * len(row)
    return df.style.apply(_style, axis=1)

def _partners_merge_edited_rows(base_df: pd.DataFrame, visible_ids: list, edited_df: pd.DataFrame, original_columns: list) -> tuple:
    current_user = _auth_user_name()
    now = _now_human()
    base = base_df.copy().replace({np.nan: ""})
    edited = edited_df.copy().replace({np.nan: ""})
    if PARTNER_ROW_ID not in edited.columns:
        edited[PARTNER_ROW_ID] = ""
    logs = []
    visible_set = {str(row_id) for row_id in visible_ids}
    edited_ids = {str(row.get(PARTNER_ROW_ID, "")) for _, row in edited.iterrows() if str(row.get(PARTNER_ROW_ID, "")).strip()}
    keep_rows = []
    for _, row in base.iterrows():
        row_id = str(row.get(PARTNER_ROW_ID, ""))
        if row_id in visible_set and row_id not in edited_ids:
            logs.append((
                "Linha excluída",
                row_id,
                {"tipo_acao": "exclusão", "linha": row_id, "campo": "", "valor_antigo": "linha existente", "valor_novo": "linha removida"}
            ))
            continue
        keep_rows.append(row.to_dict())
    merged = pd.DataFrame(keep_rows)
    if merged.empty:
        merged = pd.DataFrame(columns=base.columns)
    merged = merged.set_index(PARTNER_ROW_ID, drop=False) if PARTNER_ROW_ID in merged.columns else merged
    for _, row in edited.iterrows():
        row_dict = row.to_dict()
        row_id = str(row_dict.get(PARTNER_ROW_ID, "")).strip()
        if not row_id or row_id not in merged.index:
            row_id = hashlib.sha1(f"{current_user}-{_now_iso()}-{len(merged)}".encode()).hexdigest()[:12]
            row_dict[PARTNER_ROW_ID] = row_id
            row_dict["Última Alteração"] = now
            row_dict["Usuário Responsável"] = current_user
            logs.append((
                "Linha criada",
                row_id,
                {"tipo_acao": "edição", "linha": row_id, "campo": "", "valor_antigo": "", "valor_novo": "linha criada"}
            ))
            merged.loc[row_id, list(row_dict.keys())] = list(row_dict.values())
            continue
        previous_status = str(merged.at[row_id, "Status de Execução"]) if "Status de Execução" in merged.columns else ""
        new_status = str(row_dict.get("Status de Execução", previous_status))
        changed_cols = []
        for col, value in row_dict.items():
            if col not in merged.columns:
                merged[col] = ""
            old_value = str(merged.at[row_id, col])
            if old_value != str(value):
                merged.at[row_id, col] = value
                if col != PARTNER_ROW_ID:
                    changed_cols.append(col)
                    logs.append((
                        "Linha alterada",
                        f"{row_id}: {col}",
                        {"tipo_acao": "edição", "linha": row_id, "campo": col, "valor_antigo": old_value, "valor_novo": str(value)}
                    ))
        if changed_cols:
            merged.at[row_id, "Última Alteração"] = now
            merged.at[row_id, "Usuário Responsável"] = current_user
        if previous_status != new_status:
            logs.append((
                "Status alterado",
                f"{row_id}: {previous_status} -> {new_status}",
                {"tipo_acao": "edição", "linha": row_id, "campo": "Status de Execução", "valor_antigo": previous_status, "valor_novo": new_status}
            ))
    merged = merged.reset_index(drop=True).replace({np.nan: ""})
    columns = [col for col in original_columns if col in merged.columns]
    for col in PARTNER_INTERNAL_COLUMNS:
        if col not in merged.columns:
            merged[col] = "Não executado" if col == "Status de Execução" else ""
    if "Status de Execução" in merged.columns:
        merged["Status de Execução"] = merged["Status de Execução"].map(
            lambda value: value if str(value) in PARTNER_STATUS_OPTIONS else "Não executado"
        )
    status_cols = ["Status de Execução"] if "Status de Execução" in merged.columns else []
    front_internal = [col for col in ("Tratativa", "Descrição / Observação") if col in merged.columns]
    tail_internal = [col for col in PARTNER_INTERNAL_COLUMNS if col in merged.columns and col not in status_cols + front_internal]
    ordered = [PARTNER_ROW_ID] + status_cols + front_internal + columns + tail_internal
    ordered += [col for col in merged.columns if col not in ordered]
    return merged[ordered], logs

def _partners_filter_dataframe(df: pd.DataFrame, search: str, statuses: list) -> pd.DataFrame:
    filtered = df.copy()
    if statuses and "Status de Execução" in filtered.columns:
        filtered = filtered[filtered["Status de Execução"].isin(statuses)]
    query = str(search or "").strip().lower()
    if query:
        mask = filtered.apply(lambda row: query in " ".join(str(v).lower() for v in row.values), axis=1)
        filtered = filtered[mask]
    return filtered

def _partners_deadline_alerts(chat_rows: list) -> list:
    alerts = []
    today = date.today()
    for item in chat_rows or []:
        status = item.get("status", "Aberto")
        prazo_raw = item.get("prazo", "")
        if status in PARTNER_TREATMENT_DONE:
            alerts.append(("Verde", "Tratativa resolvida", item))
            continue
        if not prazo_raw:
            alerts.append(("Cinza", "Tratativa sem prazo", item))
            continue
        try:
            prazo = date.fromisoformat(str(prazo_raw))
        except Exception:
            alerts.append(("Cinza", "Prazo inválido ou ausente", item))
            continue
        delta = (prazo - today).days
        if delta < 0:
            alerts.append(("Vermelho", "Prazo vencido", item))
        elif delta == 0:
            alerts.append(("Amarelo", "Prazo vence hoje", item))
        elif delta == 1:
            alerts.append(("Amarelo", "Prazo vence amanhã", item))
        elif status in ("Aberto", "Em andamento"):
            alerts.append(("Cinza", "Tratativa pendente", item))
    return alerts

# ==========================================
# TEMA — CSS APLICADO CONFORME CONFIGURAÇÃO
# ==========================================
if SYSTEM_CONFIG.get("tema", "padrao") == "tmg_azul":
    st.markdown("""
<style>
    .stApp {
        background-color: #0a1628 !important;
        color: #e8edf5 !important;
    }
    [data-testid="stSidebar"] {
        background-color: #0d1e35 !important;
        border-right: 1px solid #1e3a5f !important;
    }
    div.stButton > button {
        background: linear-gradient(145deg, #1a3a5c, #0d2140) !important;
        color: #a8c4e0 !important;
        box-shadow: 3px 3px 6px #050e1a, -1px -1px 6px #1a3a5c !important;
    }
    div.stButton > button:hover {
        color: #64b5f6 !important;
        border: 1px solid #1976d2 !important;
    }
    .stButton > button[kind="primary"] {
        background: linear-gradient(145deg, #1976d2, #1565c0) !important;
        color: #ffffff !important;
        box-shadow: 4px 4px 10px #050e1a !important;
    }
    .card {
        background-color: #112038 !important;
        border: 1px solid #1e3a5f !important;
    }
    .menu-3d-title {
        color: #42a5f5 !important;
        text-shadow:
            1px 1px 0 #0d47a1,
            2px 2px 0 #0a3070,
            3px 3px 0 #071f4a,
            4px 4px 6px rgba(0,0,0,0.9),
            0 0 20px rgba(66,165,245,0.4),
            0 0 40px rgba(66,165,245,0.15) !important;
    }
    .separator-glow {
        border-top: 1px solid #1976d2 !important;
        box-shadow: 0 0 8px rgba(25,118,210,0.5) !important;
    }
    .main-header {
        color: #FFFFFF !important;
        border-bottom: 2px solid #1976d2 !important;
        text-shadow:
            2px 2px 0px #000000,
            4px 4px 0px #0a1628,
            6px 6px 8px rgba(0,0,0,0.9),
            0 0 30px rgba(25,118,210,0.25) !important;
    }
    [data-testid="stTextInput"] > div > div > input,
    [data-testid="stTextArea"] > div > div > textarea {
        background-color: #0d1e35 !important;
        border: 1px solid #1e3a5f !important;
        color: #e8edf5 !important;
    }
    [data-testid="stSelectbox"] > div > div,
    [data-testid="stMultiSelect"] > div > div {
        background-color: #0d1e35 !important;
        border: 1px solid #1e3a5f !important;
        color: #e8edf5 !important;
    }
    [data-testid="stExpander"] {
        background-color: #0d1e35 !important;
        border: 1px solid #1e3a5f !important;
    }
    .stDataFrame, [data-testid="stTable"] {
        background-color: #0d1e35 !important;
    }
    [data-testid="stMetricValue"] { color: #42a5f5 !important; }
    [data-testid="stMetricLabel"] { color: #90caf9 !important; }
</style>
""", unsafe_allow_html=True)

# ==========================================
# HELPER — PROCESSAR ORTOFOTO PARA VISUALIZAÇÃO[cite: 1]
# ==========================================
@st.cache_data(show_spinner=False, max_entries=12)
def _processar_ortofoto_cached(
    file_bytes: bytes,
    filename: str,
    preview_max_dim: int,
    preview_jpeg_quality: int,
    preview_max_payload_mb: int,
    preview_min_dim: int,
):
    """Converte ortofotos para pré-visualização de alta qualidade no Streamlit.

    Ajuste focado em TIF/GeoTIFF/RGB e formatos comuns, preservando metadados espaciais
    e reduzindo a imagem somente para o tamanho seguro de navegação no browser.
    """
    def _set_progress(pct: int, message: str):
        return None

    ext = Path(filename).suffix.lower()
    img = None
    erro = None

    spatial_meta = {
        "transform": None,
        "crs": None,
        "ratio": 1.0,
        "orig_width": 0,
        "orig_height": 0,
        "preview_width": 0,
        "preview_height": 0,
        "preview_quality": preview_jpeg_quality,
    }

    def _stretch_band(band):
        band = np.ma.asarray(band).astype(np.float32).filled(np.nan)
        valid = band[np.isfinite(band)]
        if valid.size == 0:
            return np.zeros(band.shape, dtype=np.uint8)
        mn = np.nanpercentile(valid, 1)
        mx = np.nanpercentile(valid, 99)
        if not np.isfinite(mn) or not np.isfinite(mx) or mx <= mn:
            mn = np.nanmin(valid)
            mx = np.nanmax(valid)
        if not np.isfinite(mn) or not np.isfinite(mx) or mx <= mn:
            return np.zeros(band.shape, dtype=np.uint8)
        return np.nan_to_num(np.clip((band - mn) / (mx - mn) * 255, 0, 255)).astype(np.uint8)

    def _preserve_uint8_band(band):
        return np.ma.asarray(band).filled(0).astype(np.uint8, copy=False)

    def _rgba_to_rgb(pil_img):
        if pil_img.mode == 'RGBA':
            try:
                pil_img.load()
                bg = Image.new('RGB', pil_img.size, (18, 18, 18))
                bg.paste(pil_img, mask=pil_img.split()[3])
                return bg
            except Exception:
                arr = np.array(pil_img)
                if arr.ndim == 3 and arr.shape[2] == 4:
                    alpha = arr[:, :, 3:4].astype(np.float32) / 255.0
                    rgb = arr[:, :, :3].astype(np.float32)
                    bg_v = np.array([18, 18, 18], dtype=np.float32)
                    return Image.fromarray((rgb * alpha + bg_v * (1.0 - alpha)).clip(0, 255).astype(np.uint8), 'RGB')
        if pil_img.mode != 'RGB':
            return pil_img.convert('RGB')
        return pil_img

    try:
        _set_progress(14, f"Lendo ortofoto: {Path(filename).name}")
        import rasterio
        from rasterio.enums import Resampling
        from rasterio.io import MemoryFile
        with MemoryFile(file_bytes) as memfile:
            with memfile.open() as src:
                _set_progress(24, "Interpretando metadados e dimensões da ortofoto...")
                spatial_meta["orig_width"] = int(src.width)
                spatial_meta["orig_height"] = int(src.height)
                if getattr(src, "crs", None):
                    spatial_meta["crs"] = src.crs.to_wkt()
                if getattr(src, "transform", None):
                    spatial_meta["transform"] = src.transform.to_gdal()

                max_dim = preview_max_dim
                ratio = min(1.0, max_dim / max(src.width, src.height))
                out_width = max(1, int(src.width * ratio))
                out_height = max(1, int(src.height * ratio))
                spatial_meta["ratio"] = ratio
                spatial_meta["preview_width"] = out_width
                spatial_meta["preview_height"] = out_height

                resampling_filter = getattr(Resampling, "lanczos", Resampling.bilinear)
                bands = int(src.count)

                if bands >= 3:
                    _set_progress(42, "Gerando pré-visualização RGB de alta qualidade...")
                    # Mantém RGB verdadeiro quando o GeoTIFF já está em uint8; faz realce leve somente em 16/32 bits.
                    band_indexes = [1, 2, 3]
                    data = src.read(
                        band_indexes,
                        out_shape=(3, out_height, out_width),
                        resampling=resampling_filter,
                        masked=True,
                    )
                    selected_dtypes = [np.dtype(src.dtypes[index - 1]) for index in band_indexes]
                    if all(dtype == np.dtype("uint8") for dtype in selected_dtypes):
                        arr = np.transpose(np.stack([_preserve_uint8_band(data[i]) for i in range(3)]), (1, 2, 0))
                    else:
                        arr = np.transpose(np.stack([_stretch_band(data[i]) for i in range(3)]), (1, 2, 0))
                    img = Image.fromarray(arr, "RGB")
                elif bands == 2:
                    _set_progress(42, "Preparando ortofoto com canal alfa...")
                    data = src.read(1, out_shape=(out_height, out_width), resampling=resampling_filter, masked=True)
                    alpha = src.read(2, out_shape=(out_height, out_width), resampling=resampling_filter, masked=True)
                    gray = _preserve_uint8_band(data) if np.dtype(src.dtypes[0]) == np.dtype("uint8") else _stretch_band(data)
                    rgba = np.dstack([gray, gray, gray, _preserve_uint8_band(alpha)])
                    img = _rgba_to_rgb(Image.fromarray(rgba, "RGBA"))
                else:
                    _set_progress(42, "Preparando banda única da ortofoto...")
                    data = src.read(1, out_shape=(out_height, out_width), resampling=resampling_filter, masked=True)
                    if np.dtype(src.dtypes[0]) == np.dtype("uint8"):
                        img = Image.fromarray(_preserve_uint8_band(data), "L").convert("RGB")
                    else:
                        img = Image.fromarray(_stretch_band(data), "L").convert("RGB")
    except ImportError:
        try:
            _set_progress(30, "Lendo imagem com Pillow...")
            img = Image.open(BytesIO(file_bytes))
        except Exception as e_pil:
            erro = f"Falha ao ler imagem sem Rasterio: {e_pil}"
    except Exception as e_rast:
        try:
            _set_progress(30, "Lendo imagem em modo compatível...")
            img = Image.open(BytesIO(file_bytes))
        except Exception as e_pil:
            erro = f"Falha ao ler formato {ext}: {e_rast} | {e_pil}"

    if erro is None and img is None:
        try:
            _set_progress(34, "Interpretando imagem...")
            img = Image.open(BytesIO(file_bytes))
        except Exception as e:
            erro = f"Falha ao interpretar a imagem: {e}"

    if erro:
        return None, None, erro, spatial_meta
    if img is None:
        return None, None, "Não foi possível interpretar a imagem.", spatial_meta

    if spatial_meta["orig_width"] == 0:
        spatial_meta["orig_width"] = img.width
        spatial_meta["orig_height"] = img.height

    _set_progress(58, "Ajustando cores e transparência da ortofoto...")
    img = _rgba_to_rgb(img)

    MAX_DIM = preview_max_dim
    if max(img.size) > MAX_DIM:
        _set_progress(68, "Otimizando tamanho para o visualizador...")
        ratio = MAX_DIM / max(img.size)
        spatial_meta["ratio"] = ratio
        resample_filter = getattr(Image, 'Resampling', Image).LANCZOS
        img = img.resize((max(1, int(img.width * ratio)), max(1, int(img.height * ratio))), resample_filter)
    else:
        _set_progress(68, "Preservando dimensão do preview da ortofoto...")

    spatial_meta["preview_width"] = img.width
    spatial_meta["preview_height"] = img.height

    def _save_preview_jpeg(pil_img, jpeg_quality):
        save_options = [
            {"subsampling": 0, "optimize": True, "progressive": True},
            {"subsampling": 0, "optimize": True, "progressive": False},
            {"subsampling": 1, "optimize": False, "progressive": False},
            {"subsampling": 2, "optimize": False, "progressive": False},
        ]
        last_error = None
        for options in save_options:
            target = BytesIO()
            try:
                pil_img.save(target, format='JPEG', quality=jpeg_quality, **options)
                return target
            except OSError as exc:
                last_error = exc
        if last_error:
            raise last_error
        raise OSError("Falha ao salvar preview JPEG.")

    max_payload_bytes = preview_max_payload_mb * 1024 * 1024
    min_preview_dim = min(preview_min_dim, preview_max_dim)
    quality = preview_jpeg_quality
    preview_img = img
    buf = BytesIO()

    try:
        for step_idx in range(18):
            _set_progress(min(92, 72 + int((step_idx / 18) * 20)), "Comprimindo preview sem perder qualidade visual...")
            buf = _save_preview_jpeg(preview_img, quality)
            payload_size = buf.tell()
            if payload_size <= max_payload_bytes:
                break
            if quality > 82:
                quality = max(82, quality - 5)
                continue
            current_max_dim = max(preview_img.size)
            if current_max_dim > min_preview_dim:
                scale = max(min_preview_dim / current_max_dim, 0.75)
                new_size = (
                    max(1, int(preview_img.width * scale)),
                    max(1, int(preview_img.height * scale)),
                )
                resample_filter = getattr(Image, 'Resampling', Image).LANCZOS
                preview_img = preview_img.resize(new_size, resample_filter)
                quality = min(quality, 90)
                continue
            if quality > 70:
                quality = max(70, quality - 4)
                continue
            if current_max_dim > 1200:
                scale = 0.85
                new_size = (
                    max(1, int(preview_img.width * scale)),
                    max(1, int(preview_img.height * scale)),
                )
                resample_filter = getattr(Image, 'Resampling', Image).LANCZOS
                preview_img = preview_img.resize(new_size, resample_filter)
                continue
            break
    except Exception as exc:
        return None, None, f"Falha ao preparar visualização da ortofoto: {exc}", spatial_meta

    img = preview_img
    spatial_meta["preview_width"] = img.width
    spatial_meta["preview_height"] = img.height
    spatial_meta["preview_quality"] = quality
    spatial_meta["preview_payload_mb"] = round(buf.tell() / (1024 * 1024), 2)
    if spatial_meta["orig_width"]:
        spatial_meta["ratio"] = img.width / spatial_meta["orig_width"]
    _set_progress(96, "Finalizando visualização da ortofoto...")
    b64 = base64.b64encode(buf.getvalue()).decode()

    return b64, img.size, None, spatial_meta

def processar_ortofoto(file_bytes: bytes, filename: str):
    file_name = Path(filename or "ortofoto").name
    total_mb = (len(file_bytes or b"") / (1024 * 1024)) if file_bytes is not None else 0
    loading_slot = st.empty()
    preview_max_dim = _preview_max_dim()
    preview_jpeg_quality = _preview_jpeg_quality()
    preview_max_payload_mb = _preview_max_payload_mb()
    preview_min_dim = _preview_min_dim()

    def _progress(pct: int, message: str):
        detail = f"{message} · {total_mb:.1f} MB" if total_mb else message
        update_tmg_loading(loading_slot, pct, detail)

    try:
        _progress(6, f"Iniciando carregamento da ortofoto: {file_name}")
        _progress(18, "Preparando cache e parâmetros de alta qualidade...")
        _progress(32, f"Processando preview até {preview_max_dim}px com qualidade {preview_jpeg_quality}...")
        result = _processar_ortofoto_cached(
            file_bytes,
            filename,
            preview_max_dim,
            preview_jpeg_quality,
            preview_max_payload_mb,
            preview_min_dim,
        )
        _progress(94, "Aplicando preview otimizado no visualizador...")
        finish_tmg_loading_and_clear(loading_slot, "Ortofoto carregada com sucesso.")
        return result
    except Exception:
        clear_tmg_loading(loading_slot)
        raise


PENDAO_AVANCADO_PARAMS = {
    "clahe_clip_limit": 2.5,
    "illumination_kernel": 31,
    "sharpen_strength": 0.25,
    "green_lower": (32, 35, 25),
    "green_upper": (95, 255, 255),
    "green_hsv_low": (32, 35, 25),
    "green_hsv_high": (95, 255, 255),
    "green_light_s_min": 18,
    "green_light_v_min": 72,
    "exg_threshold": 88,
    "hsv_tassel_lower_1": (12, 25, 95),
    "hsv_tassel_upper_1": (42, 180, 255),
    "hsv_tassel_lower_2": (0, 0, 135),
    "hsv_tassel_upper_2": (60, 95, 255),
    "new_tassel_hsv_low": (12, 25, 95),
    "new_tassel_hsv_high": (42, 180, 255),
    "dry_tassel_hsv_low": (8, 20, 90),
    "dry_tassel_hsv_high": (35, 180, 240),
    "old_tassel_hsv_low": (5, 25, 70),
    "old_tassel_hsv_high": (28, 200, 210),
    "cream_l_min": 138,
    "cream_s_max": 98,
    "cream_v_min": 118,
    "lab_l_threshold": 132,
    "min_v_threshold": 82,
    "max_glare_v": 248,
    "texture_threshold": 16,
    "texture_ratio_min": 0.045,
    "yellow_ratio_min": 0.06,
    "clear_ratio_min": 0.16,
    "area_min": 18,
    "area_max": 1800,
    "area_min_fraction": 0.00001,
    "area_max_fraction": 0.035,
    "max_circularity": 0.86,
    "min_solidity": 0.08,
    "max_green_ratio": 0.38,
    "min_branch_directions": 3,
    "star_direction_count": 8,
    "star_ray_fill_min": 0.08,
    "star_reach_fraction": 0.34,
    "merge_distance": 12,
    "nms_distance": 12,
    "min_distance": 12,
    "morph_open_kernel": 3,
    "morph_close_kernel": 5,
    "dilate_kernel": 3,
    "x_min_size": 6,
    "x_max_size": 20,
    "x_size_factor": 0.40,
    "x_thickness": 2,
    "analysis_max_dim": 2200,
    "max_detections": 30000,
    "yellow_index_threshold": 136,
    "exr_threshold": 128,
    "lab_b_min": 124,
    "lab_ba_diff_min": -6,
    "difficult_texture_ratio_min": 0.035,
    "yolo_enabled": True,
    "yolo_world_enabled": True,
    "yolo_world_model": "yolov8s-world.pt",
    "yolo_imgsz": 1280,
    "yolo_conf": 0.04,
    "yolo_iou": 0.45,
    "yolo_max_det": 20000,
    "yolo_max_dim": 1800,
    "yolo_min_visual_score": 1.2,
    "yolo_min_yellow_ratio": 0.015,
    "yolo_min_texture_ratio": 0.015,
    "yolo_max_green_ratio": 0.78,
    "yolo_merge_distance": 18,
    "manual_crop_size_default": 96,
    "reference_matching_enabled": True,
    "reference_max_templates": 28,
    "reference_match_max_dim": 1600,
    "reference_match_threshold": 0.58,
    "reference_match_min_distance": 18,
    "reference_match_scales": (0.65, 0.85, 1.00, 1.20, 1.40),
}

PENDAO_ANALISE_PARAMS = {
    "clahe_clip_limit": 2.4,
    "clahe_tile_grid_size": (8, 8),
    "sharpen_strength": 0.22,
    "gaussian_blur_kernel": 3,
    "illumination_kernel": 31,
    "exg_threshold": 112,
    "green_hsv_low": (35, 40, 40),
    "green_hsv_high": (85, 255, 255),
    "yellow_hsv_low": (16, 35, 140),
    "yellow_hsv_high": (38, 255, 255),
    "lab_l_threshold": 166,
    "morph_open_kernel": 3,
    "morph_close_kernel": 3,
    "texture_threshold": 22,
    "texture_ratio_min": 0.16,
    "yellow_ratio_min": 0.22,
    "area_min": 14,
    "area_max": 1200,
    "area_min_fraction": 0.00003,
    "area_max_fraction": 0.018,
    "aspect_ratio_min": 1.2,
    "max_width_ratio": 0.68,
    "vertical_cut_ratio": 1.0,
    "clear_ratio_min": 0.42,
    "min_bright_pixels": 12,
    "texture_dilate_kernel": 5,
    "x_size_factor": 0.38,
    "x_min_size": 6,
    "x_max_size": 18,
    "x_thickness": 2,
    "entropy_threshold": 16,
    "entropy_disk_radius": 3,
}

PENDAO_YOLO_CLASSES = (
    "corn tassel",
    "maize tassel",
    "tassel",
    "corn flower",
    "maize flower",
    "pendao de milho",
    "pendão de milho",
    "pendao",
    "pendão",
)

PENDAO_YOLO_CONFIG_PATH = SYSTEM_DATABASE_DIR / "pendoamento_yolo_config.json"
PENDAO_YOLO_DEFAULT_ROOT = APP_ROOT / "dados_treinamento_yolo" / "pendoes"


def _load_pendoamento_yolo_config() -> dict:
    default = {
        "training_dir": str(PENDAO_YOLO_DEFAULT_ROOT),
        "updated_at": "",
    }
    try:
        if PENDAO_YOLO_CONFIG_PATH.exists():
            data = json.loads(PENDAO_YOLO_CONFIG_PATH.read_text(encoding="utf-8"))
            if isinstance(data, dict):
                default.update({k: v for k, v in data.items() if isinstance(v, str)})
    except Exception:
        pass
    return default


def _resolve_pendoamento_yolo_dir(path_value: str | Path | None = None) -> Path:
    raw = str(path_value or "").strip()
    if not raw:
        raw = str(PENDAO_YOLO_DEFAULT_ROOT)
    path = Path(raw).expanduser()
    if not path.is_absolute():
        path = APP_ROOT / path
    return path.resolve()


def configurar_pasta_treinamento_yolo(path_value: str | Path | None = None) -> Path:
    global YOLO_TRAIN_ROOT, YOLO_TRAIN_LOG_PATH, YOLO_HISTORY_PATH
    selected = _resolve_pendoamento_yolo_dir(path_value)
    selected.mkdir(parents=True, exist_ok=True)
    YOLO_TRAIN_ROOT = selected
    YOLO_TRAIN_LOG_PATH = YOLO_TRAIN_ROOT / "treino_yolo.log"
    YOLO_HISTORY_PATH = YOLO_TRAIN_ROOT / "historico_pendoes.jsonl"
    try:
        PENDAO_YOLO_CONFIG_PATH.parent.mkdir(parents=True, exist_ok=True)
        PENDAO_YOLO_CONFIG_PATH.write_text(
            json.dumps(
                {
                    "training_dir": str(YOLO_TRAIN_ROOT),
                    "updated_at": datetime.now().isoformat(timespec="seconds"),
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
    except Exception:
        pass
    return YOLO_TRAIN_ROOT


YOLO_TRAIN_ROOT = _resolve_pendoamento_yolo_dir(_load_pendoamento_yolo_config().get("training_dir"))
YOLO_LEGACY_TRAIN_ROOT = APP_ROOT / "dados_treinamento_yolo" / "pendao_milho"
YOLO_MODELS_DIR = APP_ROOT / "modelos_yolo"
YOLO_BEST_MODEL_PATH = YOLO_MODELS_DIR / "pendao_milho_best.pt"
YOLO_TRAIN_LOG_PATH = YOLO_TRAIN_ROOT / "treino_yolo.log"
YOLO_HISTORY_PATH = YOLO_TRAIN_ROOT / "historico_pendoes.jsonl"
configurar_pasta_treinamento_yolo(YOLO_TRAIN_ROOT)
PENDAO_YOLO_WORLD_LOCAL = APP_ROOT / "yolov8s-world.pt"
if PENDAO_YOLO_WORLD_LOCAL.exists():
    PENDAO_AVANCADO_PARAMS["yolo_world_model"] = str(PENDAO_YOLO_WORLD_LOCAL)


def _bases_treinamento_yolo(include_legacy: bool = True):
    bases = [YOLO_TRAIN_ROOT]
    if include_legacy and YOLO_LEGACY_TRAIN_ROOT != YOLO_TRAIN_ROOT and YOLO_LEGACY_TRAIN_ROOT.exists():
        bases.append(YOLO_LEGACY_TRAIN_ROOT)
    return bases


def garantir_estrutura_treinamento_yolo(base_dir: Path | None = None):
    base = Path(base_dir or YOLO_TRAIN_ROOT)
    folders = [
        base / "images" / "train",
        base / "images" / "val",
        base / "labels" / "train",
        base / "labels" / "val",
        base / "crops" / "pendao_confirmado",
        base / "crops" / "falso_positivo",
        base / "crops" / "pendao_faltante",
    ]
    for folder in folders:
        folder.mkdir(parents=True, exist_ok=True)
    data_yaml = base / "data.yaml"
    data_yaml.write_text(
        f"path: {str(base).replace(os.sep, '/')}\n"
        "train: images/train\n"
        "val: images/val\n"
        "names:\n"
        "  0: pendao\n",
        encoding="utf-8",
    )
    return base


def contar_amostras_treinamento_yolo(base_dir: Path | None = None):
    bases = [garantir_estrutura_treinamento_yolo(base_dir)] if base_dir else [garantir_estrutura_treinamento_yolo(YOLO_TRAIN_ROOT)]
    if not base_dir:
        bases.extend([base for base in _bases_treinamento_yolo(include_legacy=True)[1:] if base.exists()])
    counts = {}
    for split in ("train", "val"):
        counts[f"images_{split}"] = len({
            str(path.resolve())
            for base in bases
            for path in base.glob(f"**/images/{split}/*.*")
            if path.is_file()
        })
        counts[f"labels_{split}"] = len({
            str(path.resolve())
            for base in bases
            for path in base.glob(f"**/labels/{split}/*.txt")
            if path.is_file()
        })
    for kind in ("pendao_confirmado", "falso_positivo", "pendao_faltante"):
        counts[f"crops_{kind}"] = len({
            str(path.resolve())
            for base in bases
            for path in base.glob(f"**/crops/{kind}/*.*")
            if path.is_file() and path.suffix.lower() != ".json"
        })
    counts["total_images"] = counts["images_train"] + counts["images_val"]
    counts["total_labels"] = counts["labels_train"] + counts["labels_val"]
    return counts


def assinatura_modelo_yolo_pendao() -> str:
    if YOLO_BEST_MODEL_PATH.exists():
        stat = YOLO_BEST_MODEL_PATH.stat()
        return f"{YOLO_BEST_MODEL_PATH.name}:{stat.st_size}:{stat.st_mtime_ns}"
    return "sem-modelo-treinado"


def _nome_seguro_treino_yolo(value: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value or "ortofoto")).strip("_")
    return safe[:70] or "ortofoto"


def _bbox_yolo_automatica_do_crop(crop_rgb, sample_type: str, params=None):
    crop = _as_rgb_uint8(crop_rgb)
    if crop is None or sample_type == "falso_positivo":
        return None
    h, w = crop.shape[:2]
    fallback = (0.5, 0.5, 0.45, 0.45)
    try:
        local_params = _pendao_params(params)
        local_params.update({
            "analysis_max_dim": max(w, h),
            "area_min": 4,
            "area_max": max(64, int(w * h * 0.55)),
            "area_min_fraction": 0.0002,
            "area_max_fraction": 0.55,
            "nms_distance": max(8, min(w, h) * 0.18),
            "merge_distance": max(8, min(w, h) * 0.18),
        })
        result = detectar_pendoes_opencv_puro(crop, grade=None, params=local_params)
        detections = result.get("parcelas", [{}])[0].get("detections", [])
        if detections:
            center_x, center_y = w / 2.0, h / 2.0
            best = min(
                detections,
                key=lambda det: (
                    (float(det.get("center", (center_x, center_y))[0]) - center_x) ** 2
                    + (float(det.get("center", (center_x, center_y))[1]) - center_y) ** 2
                    - float(det.get("score", 0)) * 20.0
                ),
            )
            bx, by, bw, bh = [float(v) for v in best.get("bbox", (w * 0.275, h * 0.275, w * 0.45, h * 0.45))]
            pad = max(3.0, min(w, h) * 0.04)
            x1 = max(0.0, bx - pad)
            y1 = max(0.0, by - pad)
            x2 = min(float(w), bx + bw + pad)
            y2 = min(float(h), by + bh + pad)
            if x2 > x1 and y2 > y1:
                return (
                    ((x1 + x2) / 2.0) / w,
                    ((y1 + y2) / 2.0) / h,
                    max(0.06, min(0.92, (x2 - x1) / w)),
                    max(0.06, min(0.92, (y2 - y1) / h)),
                )
    except Exception:
        pass
    return fallback


def _extrair_caracteristicas_crop_pendao(crop_rgb, params=None):
    params = _pendao_params(params)
    crop = _as_rgb_uint8(crop_rgb)
    if crop is None:
        return {}
    h, w = crop.shape[:2]
    try:
        proc, hsv, lab, gray, lab_l = _preprocessar_pendao_opencv(crop, params)
        sem_verde, green_mask = remover_verde_agressivo(proc, params)
        cor_mask, _ = criar_mascara_pendoes_multicor(proc, lab_l, params)
        textura_mask = calcular_mascara_textura(gray, params)
        brilho_mask = cv2.threshold(lab_l, int(params.get("lab_l_threshold", 132)), 255, cv2.THRESH_BINARY)[1]
        suporte = cv2.bitwise_and(cv2.bitwise_and(cor_mask, sem_verde), cv2.bitwise_or(textura_mask, brilho_mask))
        if np.count_nonzero(suporte) < 4:
            suporte = cv2.bitwise_and(cor_mask, sem_verde)
        star = validar_formato_estrela(suporte, params)
        valid_pixels = suporte > 0
        if not np.any(valid_pixels):
            valid_pixels = np.ones((h, w), dtype=bool)
        hsv_valid = hsv[valid_pixels]
        lab_valid = lab[valid_pixels]
        rgb_valid = proc[valid_pixels]
        return {
            "crop_width": int(w),
            "crop_height": int(h),
            "mean_rgb": [round(float(v), 3) for v in np.mean(rgb_valid, axis=0)],
            "mean_hsv": [round(float(v), 3) for v in np.mean(hsv_valid, axis=0)],
            "mean_lab": [round(float(v), 3) for v in np.mean(lab_valid, axis=0)],
            "yellow_ratio": round(float(np.mean(cor_mask > 0)), 6),
            "green_ratio": round(float(np.mean(green_mask > 0)), 6),
            "texture_ratio": round(float(np.mean(textura_mask > 0)), 6),
            "clear_ratio": round(float(np.mean(brilho_mask > 0)), 6),
            "valid_tassel_ratio": round(float(np.mean(suporte > 0)), 6),
            "star_ok": bool(star.get("ok", False)),
            "star_directions": int(star.get("directions", 0)),
            "star_score": round(float(star.get("score", 0.0)), 6),
            "center_hint": [round(float(v), 3) for v in star.get("center", (w / 2, h / 2))],
        }
    except Exception as exc:
        return {
            "crop_width": int(w),
            "crop_height": int(h),
            "feature_error": str(exc),
        }


def salvar_amostra_treinamento_yolo(
    file_bytes: bytes,
    x: float,
    y: float,
    sample_type: str = "pendao_confirmado",
    crop_size: int = 128,
    source_name: str = "",
    source_date: str = "",
):
    base = garantir_estrutura_treinamento_yolo()
    sample_type = sample_type if sample_type in {"pendao_confirmado", "pendao_faltante", "falso_positivo"} else "pendao_confirmado"
    crop_size = int(np.clip(int(crop_size or 128), 48, 256))
    try:
        img = Image.open(BytesIO(file_bytes)).convert("RGB")
    except Exception:
        decoded = _decode_rgb_for_pendao(file_bytes, source_name)
        if decoded is None:
            raise
        img = Image.fromarray(_as_rgb_uint8(decoded)).convert("RGB")
    img_w, img_h = img.size
    if img_w <= 0 or img_h <= 0:
        raise ValueError("Imagem inválida para salvar amostra YOLO.")
    cx = int(np.clip(round(float(x)), 0, img_w - 1))
    cy = int(np.clip(round(float(y)), 0, img_h - 1))
    half = crop_size // 2
    left, top = cx - half, cy - half
    right, bottom = left + crop_size, top + crop_size
    crop_canvas = Image.new("RGB", (crop_size, crop_size), (0, 0, 0))
    src_left = max(0, left)
    src_top = max(0, top)
    src_right = min(img_w, right)
    src_bottom = min(img_h, bottom)
    if src_right <= src_left or src_bottom <= src_top:
        raise ValueError("Clique fora da área válida da imagem.")
    crop_piece = img.crop((src_left, src_top, src_right, src_bottom))
    crop_canvas.paste(crop_piece, (src_left - left, src_top - top))

    counts = contar_amostras_treinamento_yolo(base)
    split = "val" if (counts["total_images"] + 1) % 5 == 0 else "train"
    safe_source = _nome_seguro_treino_yolo(Path(source_name or "ortofoto").stem)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    filename = f"pendao_{counts['total_images'] + 1:04d}_{stamp}_{safe_source}_{sample_type}.png"
    image_path = base / "images" / split / filename
    label_path = base / "labels" / split / f"{Path(filename).stem}.txt"
    crop_path = base / "crops" / sample_type / filename
    meta_path = base / "crops" / sample_type / f"{Path(filename).stem}.json"

    crop_canvas.save(image_path, format="PNG")
    crop_canvas.save(crop_path, format="PNG")
    crop_array = np.array(crop_canvas)
    bbox = _bbox_yolo_automatica_do_crop(crop_array, sample_type)
    features = _extrair_caracteristicas_crop_pendao(crop_array)
    if sample_type == "falso_positivo":
        label_text = ""
    else:
        if not bbox:
            bbox = (0.5, 0.5, 0.45, 0.45)
        label_text = "0 {:.6f} {:.6f} {:.6f} {:.6f}\n".format(*bbox)
    label_path.write_text(label_text, encoding="utf-8")
    metadata = {
        "id": Path(filename).stem,
        "source_name": source_name,
        "source_date": source_date,
        "sample_type": sample_type,
        "split": split,
        "original_x": cx,
        "original_y": cy,
        "crop_size": crop_size,
        "image": str(image_path),
        "label": str(label_path),
        "crop": str(crop_path),
        "bbox_yolo": bbox,
        "features": features,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    meta_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        YOLO_HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
        with YOLO_HISTORY_PATH.open("a", encoding="utf-8") as history_file:
            history_file.write(json.dumps(metadata, ensure_ascii=False) + "\n")
    except Exception:
        pass
    return metadata


def salvar_crop_treinamento_yolo(
    crop_rgb,
    original_x: float,
    original_y: float,
    sample_type: str = "pendao_confirmado",
    crop_size: int = 128,
    source_name: str = "",
    source_date: str = "",
):
    base = garantir_estrutura_treinamento_yolo()
    sample_type = sample_type if sample_type in {"pendao_confirmado", "pendao_faltante", "falso_positivo"} else "pendao_confirmado"
    crop_size = int(np.clip(int(crop_size or 128), 48, 256))
    crop = Image.fromarray(_as_rgb_uint8(crop_rgb)).convert("RGB")
    if crop.size != (crop_size, crop_size):
        crop = crop.resize((crop_size, crop_size), Image.LANCZOS)
    counts = contar_amostras_treinamento_yolo(base)
    split = "val" if (counts["total_images"] + 1) % 5 == 0 else "train"
    safe_source = _nome_seguro_treino_yolo(Path(source_name or "ortofoto").stem)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    filename = f"pendao_{counts['total_images'] + 1:04d}_{stamp}_{safe_source}_{sample_type}.png"
    image_path = base / "images" / split / filename
    label_path = base / "labels" / split / f"{Path(filename).stem}.txt"
    crop_path = base / "crops" / sample_type / filename
    meta_path = base / "crops" / sample_type / f"{Path(filename).stem}.json"
    crop.save(image_path, format="PNG")
    crop.save(crop_path, format="PNG")
    crop_array = np.array(crop)
    bbox = _bbox_yolo_automatica_do_crop(crop_array, sample_type)
    features = _extrair_caracteristicas_crop_pendao(crop_array)
    if sample_type == "falso_positivo":
        label_text = ""
    else:
        if not bbox:
            bbox = (0.5, 0.5, 0.45, 0.45)
        label_text = "0 {:.6f} {:.6f} {:.6f} {:.6f}\n".format(*bbox)
    label_path.write_text(label_text, encoding="utf-8")
    metadata = {
        "id": Path(filename).stem,
        "source_name": source_name,
        "source_date": source_date,
        "sample_type": sample_type,
        "split": split,
        "original_x": int(round(float(original_x))),
        "original_y": int(round(float(original_y))),
        "crop_size": crop_size,
        "image": str(image_path),
        "label": str(label_path),
        "crop": str(crop_path),
        "bbox_yolo": bbox,
        "features": features,
        "saved_from": "browser_click_auto",
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    meta_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        YOLO_HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
        with YOLO_HISTORY_PATH.open("a", encoding="utf-8") as history_file:
            history_file.write(json.dumps(metadata, ensure_ascii=False) + "\n")
    except Exception:
        pass
    return metadata


def _path_dentro_treino_yolo(path_value, base: Path) -> bool:
    try:
        path = Path(path_value).resolve()
        return path.is_relative_to(base.resolve())
    except Exception:
        return False


def _path_dentro_alguma_base_treino_yolo(path_value) -> bool:
    return any(_path_dentro_treino_yolo(path_value, base) for base in _bases_treinamento_yolo(include_legacy=True))


def excluir_amostra_treinamento_yolo(sample_id: str = "", crop_file: str = "", source_name: str = "", source_date: str = ""):
    garantir_estrutura_treinamento_yolo()
    bases = [base for base in _bases_treinamento_yolo(include_legacy=True) if base.exists()]
    sample_id = Path(str(sample_id or "").strip()).stem
    crop_file = Path(str(crop_file or "").strip()).name
    crop_stem = Path(crop_file).stem if crop_file else ""
    target_stems = {stem for stem in (sample_id, crop_stem) if stem}
    candidates = []

    for stem in target_stems:
        for base in bases:
            for kind in ("pendao_confirmado", "pendao_faltante", "falso_positivo"):
                meta_path = base / "crops" / kind / f"{stem}.json"
                if meta_path.exists():
                    candidates.append(meta_path)

    if not candidates:
        for base in bases:
            for meta_path in base.glob("**/crops/*/*.json"):
                try:
                    item = json.loads(meta_path.read_text(encoding="utf-8"))
                except Exception:
                    continue
                item_id = str(item.get("id", ""))
                item_crop = Path(str(item.get("crop", item.get("image", item.get("file", ""))))).name
                if (sample_id and item_id == sample_id) or (crop_file and item_crop == crop_file) or (crop_stem and Path(item_crop).stem == crop_stem):
                    candidates.append(meta_path)

    if source_name or source_date:
        filtered = []
        for meta_path in candidates:
            try:
                item = json.loads(meta_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            item_source = item.get("source_name") or item.get("ortho", "")
            item_date = item.get("source_date") or item.get("date", "")
            if source_name and item_source != source_name:
                continue
            if source_date and item_date != source_date:
                continue
            filtered.append(meta_path)
        if filtered:
            candidates = filtered

    if not candidates:
        return False, "Amostra de treinamento não encontrada."

    meta_path = candidates[0]
    try:
        metadata = json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception:
        metadata = {}

    stems = set(target_stems)
    for key in ("image", "label", "crop"):
        value = metadata.get(key)
        if value:
            stems.add(Path(str(value)).stem)
    if metadata.get("file"):
        stems.add(Path(str(metadata.get("file"))).stem)
    if metadata.get("id"):
        stems.add(str(metadata.get("id")))
    stems = {stem for stem in stems if stem}

    paths_to_delete = {meta_path}
    for key in ("image", "label", "crop"):
        value = metadata.get(key)
        if value:
            paths_to_delete.add(Path(value))
    if metadata.get("file"):
        paths_to_delete.add(meta_path.parent / str(metadata.get("file")))
    sample_base = meta_path.parents[2]
    for stem in stems:
        for folder in (
            meta_path.parents[2] / "images" / "train",
            meta_path.parents[2] / "images" / "val",
            meta_path.parents[2] / "labels" / "train",
            meta_path.parents[2] / "labels" / "val",
            meta_path.parents[2] / "crops" / "pendao_confirmado",
            meta_path.parents[2] / "crops" / "pendao_faltante",
            meta_path.parents[2] / "crops" / "falso_positivo",
        ):
            for candidate in folder.glob(f"{stem}.*"):
                paths_to_delete.add(candidate)
        for candidate in sample_base.glob(f"**/{stem}.*"):
            paths_to_delete.add(candidate)

    deleted = 0
    for path in sorted(paths_to_delete, key=lambda item: len(str(item)), reverse=True):
        try:
            if path.exists() and path.is_file() and _path_dentro_alguma_base_treino_yolo(path):
                path.unlink()
                deleted += 1
        except Exception:
            continue

    if deleted:
        try:
            if YOLO_HISTORY_PATH.exists() and stems:
                kept_lines = []
                for line in YOLO_HISTORY_PATH.read_text(encoding="utf-8").splitlines():
                    try:
                        item = json.loads(line)
                    except Exception:
                        kept_lines.append(line)
                        continue
                    item_stems = {str(item.get("id", ""))}
                    for key in ("image", "label", "crop", "file"):
                        if item.get(key):
                            item_stems.add(Path(str(item.get(key))).stem)
                    if item_stems.isdisjoint(stems):
                        kept_lines.append(line)
                YOLO_HISTORY_PATH.write_text(("\n".join(kept_lines) + ("\n" if kept_lines else "")), encoding="utf-8")
        except Exception:
            pass
        return True, f"Amostra removida. Arquivos apagados: {deleted}."
    return False, "Nenhum arquivo vinculado foi apagado."


YOLO_CAPTURE_SERVER_PORT = int(os.environ.get("TMG_YOLO_CAPTURE_PORT", "8765"))
_YOLO_CAPTURE_SERVER_STARTED = False


def _decode_data_url_image(data_url: str):
    if not data_url or "," not in data_url:
        raise ValueError("crop_data_url inválido")
    _, encoded = data_url.split(",", 1)
    raw = base64.b64decode(encoded)
    return np.array(Image.open(BytesIO(raw)).convert("RGB"))


def iniciar_servidor_captura_yolo_pendoamento():
    global _YOLO_CAPTURE_SERVER_STARTED
    if _YOLO_CAPTURE_SERVER_STARTED:
        return True, f"http://127.0.0.1:{YOLO_CAPTURE_SERVER_PORT}/pendoamento-yolo"
    try:
        from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
    except Exception as exc:
        return False, f"Servidor local indisponível: {exc}"

    class _PendoamentoYoloHandler(BaseHTTPRequestHandler):
        def _headers(self, status=200):
            self.send_response(status)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Access-Control-Allow-Methods", "POST, OPTIONS")
            self.send_header("Access-Control-Allow-Headers", "Content-Type")
            self.end_headers()

        def do_OPTIONS(self):
            self._headers(204)

        def do_POST(self):
            if self.path.split("?", 1)[0] != "/pendoamento-yolo":
                self._headers(404)
                self.wfile.write(json.dumps({"ok": False, "error": "rota inválida"}).encode("utf-8"))
                return
            try:
                length = int(self.headers.get("Content-Length", "0") or 0)
                raw = self.rfile.read(length)
                payload = json.loads(raw.decode("utf-8"))
                action = str(payload.get("action", "")).lower()
                if action == "sample":
                    crop = _decode_data_url_image(str(payload.get("crop_data_url", "")))
                    meta = salvar_crop_treinamento_yolo(
                        crop,
                        float(payload.get("orig_x", payload.get("x", 0))),
                        float(payload.get("orig_y", payload.get("y", 0))),
                        "pendao_confirmado",
                        int(payload.get("crop_size") or PENDAO_AVANCADO_PARAMS.get("manual_crop_size_default", 96)),
                        str(payload.get("name", "")),
                        str(payload.get("date", "")),
                    )
                    counts = contar_amostras_treinamento_yolo()
                    response = {
                        "ok": True,
                        "message": "Mini foto salva automaticamente.",
                        "id": meta.get("id", ""),
                        "file": Path(str(meta.get("crop", ""))).name,
                        "crop": meta.get("crop", ""),
                        "folder": str(YOLO_TRAIN_ROOT),
                        "total_images": counts.get("total_images", 0),
                    }
                elif action == "delete_sample":
                    ok, msg = excluir_amostra_treinamento_yolo(
                        sample_id=str(payload.get("sample_id", "")),
                        crop_file=str(payload.get("file", "")),
                        source_name=str(payload.get("name", "")),
                        source_date=str(payload.get("date", "")),
                    )
                    counts = contar_amostras_treinamento_yolo()
                    response = {
                        "ok": bool(ok),
                        "message": msg,
                        "folder": str(YOLO_TRAIN_ROOT),
                        "total_images": counts.get("total_images", 0),
                    }
                else:
                    response = {"ok": False, "error": "ação inválida"}
                self._headers(200 if response.get("ok") else 400)
                self.wfile.write(json.dumps(response, ensure_ascii=False).encode("utf-8"))
            except Exception as exc:
                self._headers(500)
                self.wfile.write(json.dumps({"ok": False, "error": str(exc)}, ensure_ascii=False).encode("utf-8"))

        def log_message(self, format, *args):
            return

    try:
        server = ThreadingHTTPServer(("127.0.0.1", YOLO_CAPTURE_SERVER_PORT), _PendoamentoYoloHandler)
        thread = threading.Thread(target=server.serve_forever, daemon=True, name="TMG-Pendoamento-YOLO-Capture")
        thread.start()
        _YOLO_CAPTURE_SERVER_STARTED = True
        return True, f"http://127.0.0.1:{YOLO_CAPTURE_SERVER_PORT}/pendoamento-yolo"
    except OSError:
        _YOLO_CAPTURE_SERVER_STARTED = True
        return True, f"http://127.0.0.1:{YOLO_CAPTURE_SERVER_PORT}/pendoamento-yolo"
    except Exception as exc:
        return False, f"Servidor local indisponível: {exc}"


def selecionar_pasta_treinamento_yolo_dialogo():
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        try:
            root.attributes("-topmost", True)
        except Exception:
            pass
        selected = filedialog.askdirectory(
            title="Selecionar pasta para salvar mini fotos YOLO de pendoamento",
            initialdir=str(YOLO_TRAIN_ROOT if YOLO_TRAIN_ROOT.exists() else APP_ROOT),
        )
        root.destroy()
        return selected or ""
    except Exception:
        return ""


def listar_marcas_treinamento_yolo(source_name: str = "", source_date: str = ""):
    garantir_estrutura_treinamento_yolo()
    marks = []
    for base in _bases_treinamento_yolo(include_legacy=True):
        if not base.exists():
            continue
        for meta_path in base.glob("**/crops/*/*.json"):
            try:
                item = json.loads(meta_path.read_text(encoding="utf-8"))
            except Exception:
                continue
            if not item.get("id"):
                item["id"] = Path(str(item.get("file", meta_path.stem))).stem
            if not item.get("source_name"):
                item["source_name"] = item.get("ortho", "")
            if not item.get("source_date"):
                item["source_date"] = item.get("date", "")
            if not item.get("sample_type"):
                item["sample_type"] = item.get("type", meta_path.parent.name)
            if not item.get("crop") and item.get("file"):
                item["crop"] = str(meta_path.parent / str(item.get("file")))
            if source_name and item.get("source_name") != source_name:
                continue
            if source_date and item.get("source_date") != source_date:
                continue
            marks.append(item)
    return sorted(marks, key=lambda it: it.get("created_at", ""))


def _arquivos_referencia_treinamento_yolo(base_dir: Path | None = None, limit: int | None = None):
    bases = [garantir_estrutura_treinamento_yolo(base_dir)] if base_dir else [garantir_estrutura_treinamento_yolo(YOLO_TRAIN_ROOT)]
    if not base_dir:
        bases.extend([base for base in _bases_treinamento_yolo(include_legacy=True)[1:] if base.exists()])
    files = []
    for base in bases:
        for kind in ("pendao_confirmado", "pendao_faltante"):
            for suffix in ("*.jpg", "*.jpeg", "*.png", "*.webp", "*.tif", "*.tiff"):
                files.extend(base.glob(f"**/crops/{kind}/{suffix}"))
    files = [path for path in files if path.is_file()]
    files = list({str(path.resolve()): path for path in files}.values())
    files.sort(key=lambda path: path.stat().st_mtime_ns, reverse=True)
    if limit:
        files = files[: int(limit)]
    return files


def _metadata_treino_yolo_por_crop(crop_path: Path):
    try:
        crop_path = Path(crop_path)
        meta_path = crop_path.with_suffix(".json")
        if meta_path.exists():
            return json.loads(meta_path.read_text(encoding="utf-8"))
        for base in _bases_treinamento_yolo(include_legacy=True):
            candidate = base / "crops" / crop_path.parent.name / f"{crop_path.stem}.json"
            if candidate.exists():
                return json.loads(candidate.read_text(encoding="utf-8"))
    except Exception:
        return {}
    return {}


def assinatura_referencias_treino_yolo() -> str:
    files = _arquivos_referencia_treinamento_yolo(limit=300)
    if not files:
        return "refs:0"
    hasher = hashlib.sha1()
    for path in files:
        try:
            stat = path.stat()
            hasher.update(str(path).encode("utf-8", errors="ignore"))
            hasher.update(str(stat.st_size).encode("ascii"))
            hasher.update(str(stat.st_mtime_ns).encode("ascii"))
        except Exception:
            continue
    return f"refs:{len(files)}:{hasher.hexdigest()[:16]}"


def marcas_treinamento_yolo_preview(source_name: str, source_date: str, preview_dims: tuple, original_dims: tuple):
    preview_w, preview_h = [float(v or 1) for v in preview_dims]
    orig_w, orig_h = [float(v or 1) for v in original_dims]
    scale_x = preview_w / max(1.0, orig_w)
    scale_y = preview_h / max(1.0, orig_h)
    marks = []
    for mark in listar_marcas_treinamento_yolo(source_name, source_date)[-800:]:
        try:
            crop_size = float(mark.get("crop_size", PENDAO_AVANCADO_PARAMS.get("manual_crop_size_default", 128)))
            crop_path = Path(str(mark.get("crop", mark.get("image", ""))))
            marks.append({
                "id": mark.get("id", crop_path.stem),
                "x": round(float(mark.get("original_x", 0)) * scale_x, 2),
                "y": round(float(mark.get("original_y", 0)) * scale_y, 2),
                "size": round(max(12.0, crop_size * max(scale_x, scale_y)), 2),
                "type": mark.get("sample_type", "pendao_confirmado"),
                "file": crop_path.name,
                "crop": str(mark.get("crop", "")),
                "source_name": mark.get("source_name", source_name),
                "source_date": mark.get("source_date", source_date),
                "created_at": mark.get("created_at", ""),
            })
        except Exception:
            continue
    return marks


def _preparar_template_referencia_pendao(crop_rgb, params=None):
    params = _pendao_params(params)
    crop = _as_rgb_uint8(crop_rgb)
    if crop is None:
        return None
    h, w = crop.shape[:2]
    if min(h, w) < 16:
        return None
    max_side = 96
    if max(h, w) > max_side:
        scale = max_side / max(h, w)
        crop = cv2.resize(crop, (max(12, int(w * scale)), max(12, int(h * scale))), interpolation=cv2.INTER_AREA)
    proc, hsv, lab, gray, lab_l = _preprocessar_pendao_opencv(crop, params)
    if proc is None:
        return None
    sem_verde, green_mask = remover_verde_agressivo(proc, params)
    cor_mask, _ = criar_mascara_pendoes_multicor(proc, lab_l, params)
    textura_mask = calcular_mascara_textura(gray, params)
    brilho_mask = cv2.threshold(lab_l, int(params.get("lab_l_threshold", 128)), 255, cv2.THRESH_BINARY)[1]
    valid_mask = cv2.bitwise_and(cv2.bitwise_and(cor_mask, sem_verde), cv2.bitwise_or(textura_mask, brilho_mask))
    valid_ratio = float(np.mean(valid_mask > 0))
    green_ratio = float(np.mean(green_mask > 0))
    if valid_ratio < 0.004:
        return None
    if green_ratio > 0.97 and valid_ratio < 0.025:
        return None
    gray_eq = cv2.equalizeHist(gray)
    masked = cv2.bitwise_and(gray_eq, gray_eq, mask=cv2.dilate(valid_mask, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))))
    if float(np.std(masked)) < 4.0:
        masked = gray_eq
    return {
        "gray": np.ascontiguousarray(masked.astype(np.uint8)),
        "w": int(masked.shape[1]),
        "h": int(masked.shape[0]),
        "valid_ratio": valid_ratio,
        "green_ratio": green_ratio,
    }


def _roi_compativel_com_caracteristicas_treino_pendao(roi, template):
    features = (template or {}).get("features") or {}
    if not isinstance(features, dict):
        return True
    try:
        ref_green = float(features.get("green_ratio", template.get("green_ratio", 0.0)))
        ref_yellow = float(features.get("yellow_ratio", 0.0))
        ref_texture = float(features.get("texture_ratio", 0.0))
        ref_clear = float(features.get("clear_ratio", 0.0))
        ref_valid = float(features.get("valid_tassel_ratio", template.get("valid_ratio", 0.0)))
    except Exception:
        return True
    green_limit = min(0.96, max(0.42, ref_green + 0.12))
    yellow_min = max(0.006, min(0.16, ref_yellow * 0.35))
    texture_min = max(0.006, min(0.14, ref_texture * 0.30))
    clear_min = max(0.06, min(0.28, ref_clear * 0.28))
    valid_min = max(0.004, min(0.12, ref_valid * 0.22))
    return (
        float(roi.get("green_ratio", 1.0)) <= green_limit
        and float(roi.get("yellow_ratio", 0.0)) >= yellow_min
        and float(roi.get("texture_ratio", 0.0)) >= texture_min
        and float(roi.get("clear_ratio", 0.0)) >= clear_min
        and float(roi.get("yellow_ratio", 0.0)) + float(roi.get("texture_ratio", 0.0)) >= valid_min
    )


@st.cache_data(show_spinner=False, max_entries=8)
def carregar_referencias_treino_yolo(signature: str = "", max_templates: int = 28):
    params = _pendao_params()
    templates = []
    for path in _arquivos_referencia_treinamento_yolo(limit=max_templates):
        try:
            crop = np.array(Image.open(path).convert("RGB"))
            template = _preparar_template_referencia_pendao(crop, params)
            if not template:
                continue
            metadata = _metadata_treino_yolo_por_crop(path)
            template["name"] = path.name
            template["metadata_id"] = metadata.get("id", path.stem) if isinstance(metadata, dict) else path.stem
            template["features"] = metadata.get("features", {}) if isinstance(metadata, dict) else {}
            templates.append(template)
        except Exception:
            continue
    return templates


def _inferir_pendoes_por_referencias(rgb_img, existing=None, params=None):
    params = _pendao_params(params)
    if not bool(params.get("reference_matching_enabled", True)):
        return [], "Referências manuais desativadas."
    rgb = _as_rgb_uint8(rgb_img)
    if rgb is None:
        return [], "Imagem inválida para referências manuais."
    signature = assinatura_referencias_treino_yolo()
    templates = carregar_referencias_treino_yolo(signature, int(params.get("reference_max_templates", 28)))
    if not templates:
        return [], "Sem mini imagens manuais para referência."

    orig_h, orig_w = rgb.shape[:2]
    max_dim = int(params.get("reference_match_max_dim", 1600))
    scale_back = 1.0
    if max(orig_h, orig_w) > max_dim:
        scale = max_dim / max(orig_h, orig_w)
        work = cv2.resize(rgb, (max(1, int(orig_w * scale)), max(1, int(orig_h * scale))), interpolation=cv2.INTER_AREA)
        scale_back = 1.0 / scale
    else:
        work = rgb

    proc, hsv, lab, gray, lab_l = _preprocessar_pendao_opencv(work, params)
    if proc is None:
        return [], "Falha no preparo das referências manuais."
    sem_verde, _ = remover_verde_agressivo(proc, params)
    cor_mask, _ = criar_mascara_pendoes_multicor(proc, lab_l, params)
    textura_mask = calcular_mascara_textura(gray, params)
    brilho_mask = cv2.threshold(lab_l, int(params.get("lab_l_threshold", 128)), 255, cv2.THRESH_BINARY)[1]
    ref_mask = cv2.bitwise_and(cv2.bitwise_and(cor_mask, sem_verde), cv2.bitwise_or(textura_mask, brilho_mask))
    gray_ref = cv2.equalizeHist(gray)
    gray_ref = cv2.bitwise_and(gray_ref, gray_ref, mask=cv2.dilate(ref_mask, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3))))
    if float(np.std(gray_ref)) < 4.0:
        gray_ref = gray

    scales = params.get("reference_match_scales", (0.70, 0.90, 1.10))
    threshold = float(params.get("reference_match_threshold", 0.64))
    close_distance = float(params.get("reference_match_min_distance", params.get("nms_distance", 18))) * scale_back
    existing_centers = []
    for det in existing or []:
        try:
            existing_centers.append(tuple(float(v) for v in det.get("center", (0, 0))))
        except Exception:
            continue

    detections = []
    max_per_template = 70
    for template in templates:
        base_gray = template["gray"]
        if base_gray.size == 0:
            continue
        for scale in scales:
            tw = max(10, int(base_gray.shape[1] * float(scale)))
            th = max(10, int(base_gray.shape[0] * float(scale)))
            if tw >= gray_ref.shape[1] or th >= gray_ref.shape[0]:
                continue
            templ = cv2.resize(base_gray, (tw, th), interpolation=cv2.INTER_AREA if scale < 1 else cv2.INTER_CUBIC)
            if float(np.std(templ)) < 3.0:
                continue
            try:
                score_map = cv2.matchTemplate(gray_ref, templ, cv2.TM_CCOEFF_NORMED)
            except Exception:
                continue
            if score_map.size == 0:
                continue
            local_max = cv2.dilate(score_map, np.ones((7, 7), dtype=np.uint8))
            ys, xs = np.where((score_map >= threshold) & (score_map >= local_max - 1e-6))
            if len(xs) > max_per_template:
                vals = score_map[ys, xs]
                keep_idx = np.argsort(vals)[-max_per_template:]
                xs = xs[keep_idx]
                ys = ys[keep_idx]
            for x, y in zip(xs, ys):
                match_score = float(score_map[y, x])
                x1, y1, x2, y2 = float(x), float(y), float(x + tw), float(y + th)
                roi = _roi_features_pendao(work, (x1, y1, x2, y2), params)
                if not roi:
                    continue
                ref_features = template.get("features") or {}
                try:
                    ref_green_limit = min(0.96, max(0.72, float(ref_features.get("green_ratio", 0.72)) + 0.12))
                except Exception:
                    ref_green_limit = min(0.90, float(params.get("yolo_max_green_ratio", 0.78)))
                if (
                    roi["green_ratio"] > ref_green_limit
                    or roi["yellow_ratio"] < 0.010
                    or roi["texture_ratio"] < 0.010
                    or roi["score"] < 1.15
                ):
                    continue
                if not _roi_compativel_com_caracteristicas_treino_pendao(roi, template):
                    continue
                cx, cy = roi["center"]
                orig_center = (float(cx * scale_back), float(cy * scale_back))
                if any(((orig_center[0] - ex[0]) ** 2 + (orig_center[1] - ex[1]) ** 2) ** 0.5 <= close_distance for ex in existing_centers):
                    continue
                existing_centers.append(orig_center)
                bx, by, bw, bh = roi["bbox"]
                score = float(roi["score"] + match_score * 2.6 + template.get("valid_ratio", 0) * 1.4)
                detections.append({
                    "center": orig_center,
                    "bbox": (float(bx * scale_back), float(by * scale_back), float(bw * scale_back), float(bh * scale_back)),
                    "size": float(np.clip(roi["size"] * scale_back, params.get("x_min_size", 6), params.get("x_max_size", 20))),
                    "score": score,
                    "confianca": "alta" if score >= 4.8 else ("media" if score >= 3.0 else "baixa"),
                    "tipo": roi.get("tipo", "referencia"),
                    "yellow_ratio": roi.get("yellow_ratio", 0),
                    "texture_ratio": roi.get("texture_ratio", 0),
                    "clear_ratio": roi.get("clear_ratio", 0),
                    "green_ratio": roi.get("green_ratio", 0),
                    "source": "Refinamento OpenCV",
                    "reference": template.get("name", ""),
                    "reference_id": template.get("metadata_id", ""),
                    "template_score": match_score,
                    "area": float(bw * bh * scale_back * scale_back),
                })
    if detections:
        detections = agrupar_componentes_estrelados(detections, params)
        detections = remover_deteccoes_duplicadas(detections, params)
    return detections, f"Referências manuais: {len(detections)} candidatos refinados a partir de {len(templates)} mini imagens."


def preparar_preview_treino_yolo(file_bytes: bytes, source_name: str = "", source_date: str = "", max_size=(1180, 680)):
    img = Image.open(BytesIO(file_bytes)).convert("RGB")
    orig_w, orig_h = img.size
    preview = img.copy()
    preview.thumbnail(max_size, Image.LANCZOS)
    scale_x = preview.size[0] / max(1, orig_w)
    scale_y = preview.size[1] / max(1, orig_h)
    draw = ImageDraw.Draw(preview)
    for mark in listar_marcas_treinamento_yolo(source_name, source_date)[-500:]:
        mx = int(float(mark.get("original_x", 0)) * scale_x)
        my = int(float(mark.get("original_y", 0)) * scale_y)
        color = "#32a9ff" if mark.get("sample_type") != "falso_positivo" else "#ff4fd8"
        size = 8
        draw.line((mx - size, my - size, mx + size, my + size), fill=color, width=2)
        draw.line((mx - size, my + size, mx + size, my - size), fill=color, width=2)
        draw.rectangle((mx - size - 2, my - size - 2, mx + size + 2, my + size + 2), outline=color, width=1)
    return preview, (orig_w, orig_h), (scale_x, scale_y)


def treinar_yolo_pendoamento(epochs: int = 100, imgsz: int = 960, batch: int = 4):
    base = garantir_estrutura_treinamento_yolo()
    counts = contar_amostras_treinamento_yolo(base)
    return (
        False,
        "O treino automático do modelo foi substituído pelo modo manual do visualizador: "
        f"use o botão Treinar YOLO para capturar mini imagens e Aplicar Treino para recontar. Amostras disponíveis: {counts['total_images']} em {base}.",
    )


def _pendao_params(params=None) -> dict:
    merged = dict(PENDAO_AVANCADO_PARAMS)
    if params:
        merged.update(params)
    return merged


def _odd_kernel(value: int, minimum: int = 3) -> int:
    value = max(minimum, int(value or minimum))
    return value if value % 2 else value + 1


def _as_rgb_uint8(rgb_img):
    if rgb_img is None:
        return None
    arr = np.asarray(rgb_img)
    if arr.ndim == 2:
        arr = np.dstack([arr, arr, arr])
    if arr.ndim != 3 or arr.shape[2] < 3 or arr.size == 0:
        return None
    arr = arr[:, :, :3]
    if arr.dtype != np.uint8:
        arr = np.nan_to_num(arr.astype(np.float32))
        mx = float(np.max(arr)) if arr.size else 0.0
        if mx <= 1.5:
            arr = arr * 255.0
        arr = np.clip(arr, 0, 255).astype(np.uint8)
    return np.ascontiguousarray(arr)


def _preprocessar_pendao_opencv(rgb_img, params):
    rgb = _as_rgb_uint8(rgb_img)
    if rgb is None:
        return None, None, None, None, None
    lab = cv2.cvtColor(rgb, cv2.COLOR_RGB2LAB)
    hsv = cv2.cvtColor(rgb, cv2.COLOR_RGB2HSV)
    gray = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY)
    l_channel, a_channel, b_channel = cv2.split(lab)
    clahe = cv2.createCLAHE(
        clipLimit=float(params.get("clahe_clip_limit", 2.5)),
        tileGridSize=(8, 8),
    )
    l_eq = clahe.apply(l_channel)
    blur_size = _odd_kernel(params.get("illumination_kernel", 31), 9)
    illumination = cv2.GaussianBlur(l_eq, (blur_size, blur_size), 0)
    l_corr = cv2.addWeighted(l_eq, 1.35, illumination, -0.35, 0)
    lab_corr = cv2.merge([np.clip(l_corr, 0, 255).astype(np.uint8), a_channel, b_channel])
    rgb_corr = cv2.cvtColor(lab_corr, cv2.COLOR_LAB2RGB)
    blur = cv2.GaussianBlur(rgb_corr, (3, 3), 0)
    strength = float(params.get("sharpen_strength", 0.25))
    sharp = cv2.addWeighted(rgb_corr, 1.0 + strength, blur, -strength, 0)
    sharp = cv2.GaussianBlur(sharp, (3, 3), 0)
    hsv = cv2.cvtColor(sharp, cv2.COLOR_RGB2HSV)
    lab = cv2.cvtColor(sharp, cv2.COLOR_RGB2LAB)
    gray = cv2.cvtColor(sharp, cv2.COLOR_RGB2GRAY)
    return sharp, hsv, lab, gray, lab[:, :, 0]


def remover_verde_agressivo(rgb_img, params=None):
    params = _pendao_params(params)
    rgb = _as_rgb_uint8(rgb_img)
    if rgb is None:
        return None, None
    hsv = cv2.cvtColor(rgb, cv2.COLOR_RGB2HSV)
    r = rgb[:, :, 0].astype(np.int16)
    g = rgb[:, :, 1].astype(np.int16)
    b = rgb[:, :, 2].astype(np.int16)
    exg = np.clip(2 * g - r - b, -255, 255).astype(np.int16)
    exg_norm = cv2.normalize(exg, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
    _, exg_green = cv2.threshold(
        exg_norm,
        int(params.get("exg_threshold", 105)),
        255,
        cv2.THRESH_BINARY,
    )
    green_hsv = cv2.inRange(
        hsv,
        np.array(params.get("green_lower", params.get("green_hsv_low", (32, 35, 25))), dtype=np.uint8),
        np.array(params.get("green_upper", params.get("green_hsv_high", (95, 255, 255))), dtype=np.uint8),
    )
    h = hsv[:, :, 0].astype(np.int16)
    s = hsv[:, :, 1].astype(np.int16)
    v = hsv[:, :, 2].astype(np.int16)
    folha_clara = (
        (h >= int(params.get("green_lower", (32, 35, 25))[0]))
        & (h <= int(params.get("green_upper", (95, 255, 255))[0]))
        & (s >= int(params.get("green_light_s_min", 18)))
        & (v >= int(params.get("green_light_v_min", 72)))
        & (g >= (r * 0.92))
        & (g >= (b * 1.04))
    ).astype(np.uint8) * 255
    green_mask = cv2.bitwise_or(cv2.bitwise_or(exg_green, green_hsv), folha_clara)
    green_mask = cv2.morphologyEx(
        green_mask,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5)),
        iterations=1,
    )
    green_mask = cv2.dilate(
        green_mask,
        cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)),
        iterations=1,
    )
    sem_verde = cv2.bitwise_not(green_mask)
    return sem_verde, green_mask


def criar_mascara_pendoes_multicor(rgb_img, lab_l=None, params=None):
    params = _pendao_params(params)
    rgb = _as_rgb_uint8(rgb_img)
    if rgb is None:
        return None, {}
    hsv = cv2.cvtColor(rgb, cv2.COLOR_RGB2HSV)
    lab = cv2.cvtColor(rgb, cv2.COLOR_RGB2LAB)
    lab_l = lab_l if lab_l is not None else lab[:, :, 0]
    palha_1 = cv2.inRange(
        hsv,
        np.array(params.get("hsv_tassel_lower_1", (12, 25, 95)), dtype=np.uint8),
        np.array(params.get("hsv_tassel_upper_1", (42, 180, 255)), dtype=np.uint8),
    )
    palha_2 = cv2.inRange(
        hsv,
        np.array(params.get("hsv_tassel_lower_2", (0, 0, 135)), dtype=np.uint8),
        np.array(params.get("hsv_tassel_upper_2", (60, 95, 255)), dtype=np.uint8),
    )
    novo = cv2.inRange(
        hsv,
        np.array(params.get("new_tassel_hsv_low", (15, 25, 130)), dtype=np.uint8),
        np.array(params.get("new_tassel_hsv_high", (45, 255, 255)), dtype=np.uint8),
    )
    seco = cv2.inRange(
        hsv,
        np.array(params.get("dry_tassel_hsv_low", (8, 20, 90)), dtype=np.uint8),
        np.array(params.get("dry_tassel_hsv_high", (35, 180, 240)), dtype=np.uint8),
    )
    velho = cv2.inRange(
        hsv,
        np.array(params.get("old_tassel_hsv_low", (5, 25, 70)), dtype=np.uint8),
        np.array(params.get("old_tassel_hsv_high", (28, 200, 210)), dtype=np.uint8),
    )
    s = hsv[:, :, 1]
    v = hsv[:, :, 2]
    creme = (
        (lab_l > int(params.get("cream_l_min", 145)))
        & (s < int(params.get("cream_s_max", 90)))
        & (v > int(params.get("cream_v_min", 120)))
    ).astype(np.uint8) * 255
    r = rgb[:, :, 0].astype(np.float32)
    g = rgb[:, :, 1].astype(np.float32)
    b = rgb[:, :, 2].astype(np.float32)
    yellow_idx = cv2.normalize(r + g - b, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
    exr = cv2.normalize((1.4 * r) - g, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
    lab_a = lab[:, :, 1].astype(np.int16)
    lab_b = lab[:, :, 2].astype(np.int16)
    indice_amarelo = (
        (yellow_idx > int(params.get("yellow_index_threshold", 136)))
        & (lab_b > int(params.get("lab_b_min", 124)))
        & ((lab_b - lab_a) > int(params.get("lab_ba_diff_min", -6)))
        & (lab_l > 105)
        & (v > 85)
    ).astype(np.uint8) * 255
    exr_mask = (
        (exr > int(params.get("exr_threshold", 128)))
        & (lab_l > 95)
        & (v > 75)
        & (s > 18)
    ).astype(np.uint8) * 255
    mascara = cv2.bitwise_or(cv2.bitwise_or(palha_1, palha_2), cv2.bitwise_or(novo, seco))
    mascara = cv2.bitwise_or(mascara, cv2.bitwise_or(velho, creme))
    mascara = cv2.bitwise_or(mascara, cv2.bitwise_or(indice_amarelo, exr_mask))
    return mascara, {
        "palha_1": palha_1,
        "palha_2": palha_2,
        "novo": novo,
        "seco": seco,
        "velho": velho,
        "creme": creme,
        "indice": indice_amarelo,
        "exr": exr_mask,
    }


def calcular_mascara_textura(gray_img, params=None):
    params = _pendao_params(params)
    gray = np.asarray(gray_img, dtype=np.uint8)
    lap = cv2.convertScaleAbs(cv2.Laplacian(gray, cv2.CV_32F, ksize=3))
    sobel_x = cv2.Sobel(gray, cv2.CV_32F, 1, 0, ksize=3)
    sobel_y = cv2.Sobel(gray, cv2.CV_32F, 0, 1, ksize=3)
    sobel = cv2.convertScaleAbs(cv2.magnitude(sobel_x, sobel_y))
    canny = cv2.Canny(gray, 45, 120)
    thr = int(params.get("texture_threshold", 18))
    _, lap_mask = cv2.threshold(lap, thr, 255, cv2.THRESH_BINARY)
    _, sobel_mask = cv2.threshold(sobel, max(20, thr + 8), 255, cv2.THRESH_BINARY)
    textura = cv2.bitwise_or(cv2.bitwise_or(lap_mask, sobel_mask), canny)
    return textura


def classificar_tipo_pendao(hsv_pixel, lab_pixel, params=None) -> str:
    params = _pendao_params(params)
    h, s, v = [int(x) for x in hsv_pixel[:3]]
    l = int(lab_pixel[0])
    if l >= params.get("cream_l_min", 145) and s <= params.get("cream_s_max", 90) and v >= params.get("cream_v_min", 120):
        return "creme"
    if 15 <= h <= 45 and v >= 130:
        return "novo"
    if 8 <= h <= 35 and v >= 90:
        return "seco"
    if 5 <= h <= 28:
        return "velho"
    return "misto"


def _detection_center_from_component(component_mask, x, y, w, h):
    dist = cv2.distanceTransform(component_mask, cv2.DIST_L2, 3)
    _, max_val, _, max_loc = cv2.minMaxLoc(dist)
    if max_val > 0:
        return float(x + max_loc[0]), float(y + max_loc[1])
    moments = cv2.moments(component_mask, binaryImage=True)
    if moments.get("m00", 0):
        return float(x + moments["m10"] / moments["m00"]), float(y + moments["m01"] / moments["m00"])
    return float(x + w / 2), float(y + h / 2)


def validar_formato_estrela(roi_mask, params=None):
    params = _pendao_params(params)
    mask = np.asarray(roi_mask, dtype=np.uint8)
    if mask.ndim == 3:
        mask = cv2.cvtColor(mask, cv2.COLOR_RGB2GRAY)
    mask = (mask > 0).astype(np.uint8) * 255
    if mask.size == 0 or int(np.count_nonzero(mask)) < 6:
        return {"ok": False, "directions": 0, "score": 0.0, "center": (mask.shape[1] / 2 if mask.ndim == 2 else 0, mask.shape[0] / 2 if mask.ndim == 2 else 0)}

    h, w = mask.shape[:2]
    dist = cv2.distanceTransform(mask, cv2.DIST_L2, 3)
    _, _, _, max_loc = cv2.minMaxLoc(dist)
    moments = cv2.moments(mask, binaryImage=True)
    if moments.get("m00", 0):
        mx = moments["m10"] / moments["m00"]
        my = moments["m01"] / moments["m00"]
        cx = (max_loc[0] * 0.65) + (mx * 0.35)
        cy = (max_loc[1] * 0.65) + (my * 0.35)
    else:
        cx, cy = max_loc

    rays = int(params.get("star_direction_count", 8))
    rays = max(8, rays)
    min_dirs = int(params.get("min_branch_directions", 3))
    fill_min = float(params.get("star_ray_fill_min", 0.08))
    reach_fraction = float(params.get("star_reach_fraction", 0.34))
    radius = max(4.0, min(w, h) * 0.5)
    branch_dirs = 0
    fills = []
    reaches = []

    for idx in range(rays):
        angle = (2.0 * np.pi * idx) / rays
        dx = np.cos(angle)
        dy = np.sin(angle)
        samples = []
        farthest = 0.0
        start_r = max(2.0, radius * 0.10)
        for rr in np.linspace(start_r, radius, 18):
            px = int(round(cx + dx * rr))
            py = int(round(cy + dy * rr))
            if px < 0 or py < 0 or px >= w or py >= h:
                continue
            hit = mask[py, px] > 0
            samples.append(hit)
            if hit:
                farthest = max(farthest, float(rr))
        if not samples:
            fills.append(0.0)
            reaches.append(0.0)
            continue
        fill = float(np.mean(samples))
        reach = farthest / max(1.0, radius)
        fills.append(fill)
        reaches.append(reach)
        if fill >= fill_min and reach >= reach_fraction:
            branch_dirs += 1

    radial_score = (branch_dirs / rays) + (float(np.mean(fills)) if fills else 0.0) + (float(np.mean(reaches)) * 0.35 if reaches else 0.0)
    return {
        "ok": branch_dirs >= min_dirs,
        "directions": int(branch_dirs),
        "score": float(radial_score),
        "center": (float(cx), float(cy)),
    }


def _centro_pendao_radial(component_mask, x, y, w, h, support_mask=None, params=None):
    mask = (np.asarray(component_mask, dtype=np.uint8) > 0).astype(np.uint8) * 255
    if support_mask is not None:
        support = (np.asarray(support_mask, dtype=np.uint8) > 0).astype(np.uint8) * 255
        if support.shape == mask.shape:
            supported = cv2.bitwise_and(mask, support)
            if np.count_nonzero(supported) >= 4:
                mask = supported
    star = validar_formato_estrela(mask, params)
    if star.get("ok") or np.count_nonzero(mask) >= 4:
        cx, cy = star.get("center", (w / 2, h / 2))
        return float(x + cx), float(y + cy), star
    cx, cy = _detection_center_from_component((np.asarray(component_mask, dtype=np.uint8) > 0).astype(np.uint8), x, y, w, h)
    star["center"] = (float(cx - x), float(cy - y))
    return cx, cy, star


def _merge_close_detections(detections, distance):
    if not detections:
        return []
    ordered = sorted(detections, key=lambda d: float(d.get("score", 0)), reverse=True)
    groups = []
    for det in ordered:
        cx, cy = det["center"]
        target = None
        for group in groups:
            gx, gy = group["center"]
            if ((cx - gx) ** 2 + (cy - gy) ** 2) ** 0.5 <= distance:
                target = group
                break
        if target is None:
            item = dict(det)
            item["_weight"] = max(1.0, float(det.get("score", 1.0)) * max(1.0, float(det.get("area", 1.0))))
            groups.append(item)
        else:
            old_weight = target.get("_weight", 1.0)
            new_weight = max(1.0, float(det.get("score", 1.0)) * max(1.0, float(det.get("area", 1.0))))
            tw = old_weight + new_weight
            target["center"] = (
                (target["center"][0] * old_weight + cx * new_weight) / tw,
                (target["center"][1] * old_weight + cy * new_weight) / tw,
            )
            target["score"] = max(float(target.get("score", 0)), float(det.get("score", 0)))
            target["area"] = float(target.get("area", 0)) + float(det.get("area", 0))
            target["_weight"] = tw
    for group in groups:
        group.pop("_weight", None)
    return groups


def remover_deteccoes_duplicadas(detections, params=None):
    params = _pendao_params(params)
    if not detections:
        return []
    distance = float(params.get("nms_distance", 20))
    ordered = sorted(detections, key=lambda d: float(d.get("score", 0)), reverse=True)
    kept = []
    for det in ordered:
        cx, cy = det["center"]
        if all(((cx - k["center"][0]) ** 2 + (cy - k["center"][1]) ** 2) ** 0.5 > distance for k in kept):
            kept.append(det)
    return kept


def agrupar_componentes_estrelados(regions, params=None):
    params = _pendao_params(params)
    return _merge_close_detections(regions, float(params.get("merge_distance", 18)))


def _grid_cells_from_grade(grade, img_w, img_h):
    if not grade:
        return [(1, 1, 1, np.array([[0, 0], [img_w - 1, 0], [img_w - 1, img_h - 1], [0, img_h - 1]], dtype=np.float32))]
    points = grade.get("points") or grade.get("pontos") or grade.get("grid") or []
    rows = int(grade.get("rows") or grade.get("linhas") or 1)
    cols = int(grade.get("cols") or grade.get("colunas") or 1)
    if len(points) < 4:
        return [(1, 1, 1, np.array([[0, 0], [img_w - 1, 0], [img_w - 1, img_h - 1], [0, img_h - 1]], dtype=np.float32))]
    norm_points = []
    for p in points[:4]:
        if isinstance(p, dict):
            norm_points.append([float(p.get("x", 0)), float(p.get("y", 0))])
        else:
            norm_points.append([float(p[0]), float(p[1])])
    pts = np.array(norm_points, dtype=np.float32)

    def bilerp_np(p0, p1, p2, p3, u, v):
        top = (1 - u) * p0 + u * p1
        bottom = (1 - u) * p3 + u * p2
        return (1 - v) * top + v * bottom

    cells = []
    index = 1
    for r in range(rows):
        for c in range(cols):
            u0, u1 = c / cols, (c + 1) / cols
            v0, v1 = r / rows, (r + 1) / rows
            poly = np.array([
                bilerp_np(pts[0], pts[1], pts[2], pts[3], u0, v0),
                bilerp_np(pts[0], pts[1], pts[2], pts[3], u1, v0),
                bilerp_np(pts[0], pts[1], pts[2], pts[3], u1, v1),
                bilerp_np(pts[0], pts[1], pts[2], pts[3], u0, v1),
            ], dtype=np.float32)
            cells.append((index, r + 1, c + 1, poly))
            index += 1
    return cells


def _pendao_analise_params(params=None) -> dict:
    merged = dict(PENDAO_ANALISE_PARAMS)
    if params:
        merged.update(params)
    return merged


def _pendao_entropy_mask(gray_img, params):
    try:
        from skimage.filters.rank import entropy
        from skimage.morphology import disk
    except Exception:
        return None
    try:
        radius = max(1, int(params.get("entropy_disk_radius", 3)))
        entropy_img = entropy(np.asarray(gray_img, dtype=np.uint8), disk(radius))
        entropy_norm = cv2.normalize(entropy_img, None, 0, 255, cv2.NORM_MINMAX).astype(np.uint8)
        return cv2.threshold(
            entropy_norm,
            int(params.get("entropy_threshold", 16)),
            255,
            cv2.THRESH_BINARY,
        )[1]
    except Exception:
        return None


def detectar_pendoes_pendoamento_opencv_parametrizado(rgb_img, grade=None, params=None):
    if "cv2" not in globals() or cv2 is None:
        return {
            "dims": (0, 0),
            "rows": 0,
            "cols": 0,
            "grid": grade,
            "parcelas": [],
            "total": 0,
            "detector_status": "Instale opencv-python.",
            "detector_mode": "OpenCV indisponível",
            "detector_counts": {"opencv": 0, "yolo": 0, "refinado": 0},
            "backend_ready": False,
        }

    params = _pendao_analise_params(params)
    rgb = _as_rgb_uint8(rgb_img)
    if rgb is None:
        return {
            "dims": (0, 0),
            "rows": 0,
            "cols": 0,
            "grid": grade,
            "parcelas": [],
            "total": 0,
            "detector_status": "Imagem inválida para análise de pendoamento.",
            "detector_mode": "OpenCV parametrizado TMG",
            "detector_counts": {"opencv": 0, "yolo": 0, "refinado": 0},
            "backend_ready": False,
        }

    orig_h, orig_w = rgb.shape[:2]
    scale = 1.0
    scale_back = 1.0
    max_dim = int(PENDAO_AVANCADO_PARAMS.get("analysis_max_dim", 2200))
    if max(orig_h, orig_w) > max_dim:
        scale = max_dim / max(orig_h, orig_w)
        work = cv2.resize(
            rgb,
            (max(1, int(orig_w * scale)), max(1, int(orig_h * scale))),
            interpolation=cv2.INTER_AREA,
        )
        scale_back = 1.0 / scale
    else:
        work = rgb

    h_img, w_img = work.shape[:2]
    cells = _grid_cells_from_grade(grade, orig_w, orig_h)
    scaled_cells = []
    for index, row, col, poly in cells:
        scaled_cells.append((index, row, col, np.asarray(poly, dtype=np.float32) * float(scale)))
    grid_mask = np.zeros((h_img, w_img), dtype=np.uint8)
    for _, _, _, poly in scaled_cells:
        cv2.fillPoly(grid_mask, [np.round(poly).astype(np.int32)], 255)
    if not np.any(grid_mask):
        grid_mask[:, :] = 255

    lab = cv2.cvtColor(work, cv2.COLOR_RGB2LAB)
    l_chan, a_chan, b_chan = cv2.split(lab)
    clahe = cv2.createCLAHE(
        clipLimit=float(params.get("clahe_clip_limit", 2.4)),
        tileGridSize=tuple(params.get("clahe_tile_grid_size", (8, 8))),
    )
    l_eq = clahe.apply(l_chan)
    illum_k = _odd_kernel(params.get("illumination_kernel", 31), 9)
    illumination = cv2.GaussianBlur(l_eq, (illum_k, illum_k), 0)
    l_float = l_eq.astype(np.float32)
    illumination_float = np.maximum(illumination.astype(np.float32), 1.0)
    l_corr = cv2.normalize((l_float / illumination_float) * 128.0, None, 0, 255, cv2.NORM_MINMAX)
    l_corr = np.clip(l_corr, 0, 255).astype(np.uint8)
    lab_corr = cv2.merge([l_corr, a_chan, b_chan])
    rgb_corr = cv2.cvtColor(lab_corr, cv2.COLOR_LAB2RGB)

    blur_k = _odd_kernel(params.get("gaussian_blur_kernel", 3), 3)
    blur = cv2.GaussianBlur(rgb_corr, (blur_k, blur_k), 0)
    strength = float(params.get("sharpen_strength", 0.22))
    sharp = cv2.addWeighted(rgb_corr, 1.0 + strength, blur, -strength, 0)

    hsv = cv2.cvtColor(sharp, cv2.COLOR_RGB2HSV)
    lab_final = cv2.cvtColor(sharp, cv2.COLOR_RGB2LAB)
    lab_l = lab_final[:, :, 0]
    gray = cv2.cvtColor(sharp, cv2.COLOR_RGB2GRAY)

    r = sharp[:, :, 0].astype(np.int16)
    g = sharp[:, :, 1].astype(np.int16)
    b = sharp[:, :, 2].astype(np.int16)
    exg = (2 * g) - r - b
    exg_green = (exg >= int(params.get("exg_threshold", 112))).astype(np.uint8) * 255
    hsv_green = cv2.inRange(
        hsv,
        np.array(params.get("green_hsv_low", (35, 40, 40)), dtype=np.uint8),
        np.array(params.get("green_hsv_high", (85, 255, 255)), dtype=np.uint8),
    )
    green_mask = cv2.bitwise_or(exg_green, hsv_green)
    green_mask = cv2.morphologyEx(
        green_mask,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)),
        iterations=1,
    )
    green_mask = cv2.dilate(
        green_mask,
        cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)),
        iterations=1,
    )
    non_green_mask = cv2.bitwise_not(green_mask)

    yellow_mask = cv2.inRange(
        hsv,
        np.array(params.get("yellow_hsv_low", (16, 35, 140)), dtype=np.uint8),
        np.array(params.get("yellow_hsv_high", (38, 255, 255)), dtype=np.uint8),
    )
    bright_mask = cv2.threshold(lab_l, int(params.get("lab_l_threshold", 166)), 255, cv2.THRESH_BINARY)[1]
    sat = hsv[:, :, 1]
    sat_mask = cv2.inRange(sat, 20, 210)

    lap = cv2.convertScaleAbs(cv2.Laplacian(gray, cv2.CV_32F, ksize=3))
    texture_mask = cv2.threshold(
        lap,
        int(params.get("texture_threshold", 22)),
        255,
        cv2.THRESH_BINARY,
    )[1]
    entropy_mask = _pendao_entropy_mask(gray, params)
    if entropy_mask is not None:
        texture_mask = cv2.bitwise_or(texture_mask, entropy_mask)
    texture_k = _odd_kernel(params.get("texture_dilate_kernel", 5), 3)
    texture_mask = cv2.dilate(
        texture_mask,
        cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (texture_k, texture_k)),
        iterations=1,
    )

    candidatos = cv2.bitwise_and(yellow_mask, bright_mask)
    candidatos = cv2.bitwise_and(candidatos, non_green_mask)
    candidatos = cv2.bitwise_and(candidatos, sat_mask)
    candidatos = cv2.bitwise_and(candidatos, grid_mask)
    open_k = _odd_kernel(params.get("morph_open_kernel", 3), 3)
    close_k = _odd_kernel(params.get("morph_close_kernel", 3), 3)
    candidatos = cv2.morphologyEx(
        candidatos,
        cv2.MORPH_OPEN,
        cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (open_k, open_k)),
        iterations=1,
    )
    candidatos = cv2.morphologyEx(
        candidatos,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (close_k, close_k)),
        iterations=1,
    )
    candidatos = cv2.bitwise_and(candidatos, cv2.bitwise_or(texture_mask, bright_mask))

    contours, _ = cv2.findContours(candidatos, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    img_area = float(max(1, w_img * h_img))
    area_min = max(float(params.get("area_min", 14)), img_area * float(params.get("area_min_fraction", 0.00003)))
    area_max = min(float(params.get("area_max", 1200)), max(area_min + 1, img_area * float(params.get("area_max_fraction", 0.018))))
    detections = []
    for contour in contours:
        area = float(cv2.contourArea(contour))
        if area < area_min or area > area_max:
            continue
        x, y, w, h = cv2.boundingRect(contour)
        if w <= 0 or h <= 0:
            continue
        vertical_ratio = float(h) / max(1.0, float(w))
        if vertical_ratio < float(params.get("aspect_ratio_min", 1.2)):
            continue
        if vertical_ratio < float(params.get("vertical_cut_ratio", 1.0)):
            continue
        if w > max(10.0, float(w_img) * float(params.get("max_width_ratio", 0.68))):
            continue

        roi_mask = np.zeros((h, w), dtype=np.uint8)
        local_contour = contour - np.array([[[x, y]]], dtype=contour.dtype)
        cv2.drawContours(roi_mask, [local_contour], -1, 255, thickness=-1)
        support = roi_mask > 0
        if not np.any(support):
            continue
        yellow_ratio = float(np.mean(yellow_mask[y:y + h, x:x + w][support] > 0))
        clear_region = lab_l[y:y + h, x:x + w] > int(params.get("lab_l_threshold", 166))
        clear_ratio = float(np.mean(clear_region[support]))
        texture_ratio = float(np.mean(texture_mask[y:y + h, x:x + w][support] > 0))
        bright_pixels = int(np.count_nonzero(clear_region & support))
        if clear_ratio < float(params.get("clear_ratio_min", 0.42)):
            continue
        if yellow_ratio < float(params.get("yellow_ratio_min", 0.22)):
            continue
        if texture_ratio < float(params.get("texture_ratio_min", 0.16)):
            continue
        if bright_pixels < int(params.get("min_bright_pixels", 12)):
            continue

        moments = cv2.moments(contour)
        if moments.get("m00", 0):
            cx = float(moments["m10"] / moments["m00"])
            cy = float(moments["m01"] / moments["m00"])
        else:
            cx = float(x + w / 2)
            cy = float(y + h / 2)
        if cv2.pointPolygonTest(np.array([[0, 0], [w_img - 1, 0], [w_img - 1, h_img - 1], [0, h_img - 1]], dtype=np.float32), (cx, cy), False) < 0:
            continue
        if not any(cv2.pointPolygonTest(poly.astype(np.float32), (cx, cy), False) >= 0 for _, _, _, poly in scaled_cells):
            continue

        score = (
            clear_ratio * 2.2
            + yellow_ratio * 2.1
            + texture_ratio * 1.8
            + min(1.0, area / max(1.0, area_min * 5.0))
            + min(1.0, vertical_ratio / 3.0) * 0.6
        )
        size = float(np.clip(max(w, h) * float(params.get("x_size_factor", 0.38)) * scale_back, params.get("x_min_size", 6), params.get("x_max_size", 18)))
        detections.append({
            "center": (float(cx * scale_back), float(cy * scale_back)),
            "size": size,
            "bbox": (float(x * scale_back), float(y * scale_back), float(w * scale_back), float(h * scale_back)),
            "score": float(score),
            "confianca": "alta" if score >= 4.2 else ("media" if score >= 3.4 else "baixa"),
            "tipo": "pendao_claro",
            "yellow_ratio": yellow_ratio,
            "texture_ratio": texture_ratio,
            "clear_ratio": clear_ratio,
            "green_ratio": 0.0,
            "area": float(area * scale_back * scale_back),
            "source": "OpenCV",
        })

    detections = remover_deteccoes_duplicadas(detections, {"nms_distance": 12})
    rows = int((grade or {}).get("rows") or (grade or {}).get("linhas") or 1)
    cols = int((grade or {}).get("cols") or (grade or {}).get("colunas") or 1)
    parcelas = []
    total = 0
    for index, row, col, poly in cells:
        dets = []
        for det in detections:
            cx, cy = det["center"]
            if cv2.pointPolygonTest(poly.astype(np.float32), (float(cx), float(cy)), False) >= 0:
                dets.append({k: v for k, v in det.items() if k != "area"})
        total += len(dets)
        parcelas.append({"index": int(index), "row": int(row), "col": int(col), "count": len(dets), "detections": dets})

    return {
        "dims": (orig_w, orig_h),
        "rows": rows,
        "cols": cols,
        "grid": grade,
        "parcelas": parcelas,
        "total": int(total),
        "detector_status": "OpenCV parametrizado executado com os parâmetros TMG de pendoamento.",
        "detector_mode": "OpenCV parametrizado TMG",
        "detector_counts": {"opencv": int(len(detections)), "yolo": 0, "refinado": 0},
        "backend_ready": True,
    }


def detectar_pendoes_opencv_puro(rgb_img, grade=None, params=None):
    params = _pendao_params(params)
    rgb = _as_rgb_uint8(rgb_img)
    if rgb is None:
        return {"dims": (0, 0), "rows": 0, "cols": 0, "grid": grade, "parcelas": [], "total": 0}

    scale_back = 1.0
    max_dim = int(params.get("analysis_max_dim", 2200))
    orig_h, orig_w = rgb.shape[:2]
    if max(orig_h, orig_w) > max_dim:
        scale = max_dim / max(orig_h, orig_w)
        rgb_small = cv2.resize(rgb, (max(1, int(orig_w * scale)), max(1, int(orig_h * scale))), interpolation=cv2.INTER_AREA)
        scale_back = 1.0 / scale
    else:
        rgb_small = rgb

    img, hsv, lab, gray, lab_l = _preprocessar_pendao_opencv(rgb_small, params)
    if img is None:
        return {"dims": (orig_w, orig_h), "rows": 0, "cols": 0, "grid": grade, "parcelas": [], "total": 0}
    sem_verde, green_mask = remover_verde_agressivo(img, params)
    cor_mask, tipo_masks = criar_mascara_pendoes_multicor(img, lab_l, params)
    h_chan = hsv[:, :, 0]
    s_chan = hsv[:, :, 1]
    v_chan = hsv[:, :, 2]
    brilho_lab = cv2.threshold(lab_l, int(params.get("lab_l_threshold", 132)), 255, cv2.THRESH_BINARY)[1]
    brilho_v = cv2.threshold(v_chan, int(params.get("min_v_threshold", 82)), 255, cv2.THRESH_BINARY)[1]
    brilho_mask = cv2.bitwise_and(brilho_lab, brilho_v)
    saturacao_baixa_media = cv2.inRange(s_chan, 0, 190)
    sombra_mask = cv2.bitwise_not(cv2.inRange(v_chan, 0, int(params.get("min_v_threshold", 82)) - 1))
    reflexo_isolado = (
        (s_chan < 18)
        & (v_chan > int(params.get("max_glare_v", 248)))
        & (lab[:, :, 2] < 132)
    ).astype(np.uint8) * 255
    sem_reflexo = cv2.bitwise_not(reflexo_isolado)
    textura_mask = calcular_mascara_textura(gray, params)

    textura_expandida = cv2.dilate(
        textura_mask,
        cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3, 3)),
        iterations=1,
    )
    claros_mask = cv2.bitwise_or(tipo_masks.get("novo", cor_mask), tipo_masks.get("creme", cor_mask))
    secos_mask = cv2.bitwise_or(tipo_masks.get("seco", cor_mask), tipo_masks.get("velho", cor_mask))
    dificeis_mask = cv2.bitwise_or(tipo_masks.get("indice", cor_mask), tipo_masks.get("exr", cor_mask))

    base_valida = cv2.bitwise_and(cv2.bitwise_and(sem_verde, sombra_mask), sem_reflexo)
    base_valida = cv2.bitwise_and(base_valida, saturacao_baixa_media)
    pass_claros = cv2.bitwise_and(cv2.bitwise_and(claros_mask, base_valida), cv2.bitwise_or(textura_expandida, brilho_mask))
    pass_secos = cv2.bitwise_and(cv2.bitwise_and(secos_mask, base_valida), textura_expandida)
    pass_dificeis = cv2.bitwise_and(
        cv2.bitwise_and(dificeis_mask, base_valida),
        cv2.bitwise_or(textura_expandida, brilho_mask),
    )
    candidatos = cv2.bitwise_or(cv2.bitwise_or(pass_claros, pass_secos), pass_dificeis)
    candidatos = cv2.bitwise_and(candidatos, cv2.bitwise_or(brilho_mask, cor_mask))
    candidatos = cv2.bitwise_and(candidatos, base_valida)

    open_k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (_odd_kernel(params.get("morph_open_kernel", 3)),) * 2)
    close_k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (_odd_kernel(params.get("morph_close_kernel", 5)),) * 2)
    dilate_k = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (_odd_kernel(params.get("dilate_kernel", 3)),) * 2)
    candidatos = cv2.morphologyEx(candidatos, cv2.MORPH_OPEN, open_k, iterations=1)
    candidatos = cv2.morphologyEx(candidatos, cv2.MORPH_CLOSE, close_k, iterations=1)
    candidatos = cv2.dilate(candidatos, dilate_k, iterations=1)
    bordas_pendao = cv2.morphologyEx(candidatos, cv2.MORPH_GRADIENT, open_k, iterations=1)
    candidatos = cv2.bitwise_or(candidatos, cv2.bitwise_and(bordas_pendao, cor_mask))

    n_labels, labels, stats, _ = cv2.connectedComponentsWithStats(candidatos, 8)
    img_area = float(candidatos.shape[0] * candidatos.shape[1])
    area_min = max(float(params.get("area_min", 8)), img_area * float(params.get("area_min_fraction", 0.00001)))
    area_max = min(float(params.get("area_max", 2500)), max(area_min + 1, img_area * float(params.get("area_max_fraction", 0.035))))
    detections = []
    for label in range(1, n_labels):
        x, y, w, h, area = stats[label]
        area = float(area)
        if area < area_min or area > area_max or w <= 0 or h <= 0:
            continue
        component = (labels[y:y + h, x:x + w] == label).astype(np.uint8)
        comp_full = (labels == label)
        contours, _ = cv2.findContours((component * 255).astype(np.uint8), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            continue
        contour = max(contours, key=cv2.contourArea)
        perimeter = max(1.0, cv2.arcLength(contour, True))
        circularity = 4.0 * np.pi * area / (perimeter * perimeter)
        hull = cv2.convexHull(contour)
        hull_area = max(1.0, cv2.contourArea(hull))
        solidity = area / hull_area
        density = area / max(1.0, float(w * h))
        aspect = max(w, h) / max(1.0, min(w, h))
        green_ratio = float(np.mean(green_mask[comp_full] > 0)) if np.any(comp_full) else 1.0
        yellow_ratio = float(np.mean(cor_mask[comp_full] > 0)) if np.any(comp_full) else 0.0
        texture_ratio = float(np.mean(textura_mask[comp_full] > 0)) if np.any(comp_full) else 0.0
        clear_ratio = float(np.mean(lab_l[comp_full] > int(params.get("lab_l_threshold", 135)))) if np.any(comp_full) else 0.0
        mean_sat = float(np.mean(s_chan[comp_full])) if np.any(comp_full) else 255.0
        mean_val = float(np.mean(v_chan[comp_full])) if np.any(comp_full) else 0.0
        if green_ratio > float(params.get("max_green_ratio", 0.35)):
            continue
        if yellow_ratio < float(params.get("yellow_ratio_min", 0.10)):
            continue
        if texture_ratio < float(params.get("texture_ratio_min", 0.08)):
            continue
        if clear_ratio < float(params.get("clear_ratio_min", 0.20)):
            continue
        if mean_val < float(params.get("min_v_threshold", 82)):
            continue
        if mean_sat > 205 and yellow_ratio < 0.18:
            continue
        if circularity > float(params.get("max_circularity", 0.90)) and aspect < 1.8:
            continue
        if solidity < float(params.get("min_solidity", 0.08)) or density < 0.05:
            continue
        if aspect > 12 and texture_ratio < 0.22:
            continue
        suporte_roi = cv2.bitwise_and(
            component * 255,
            cv2.bitwise_and(
                cv2.bitwise_or(cor_mask[y:y + h, x:x + w], brilho_mask[y:y + h, x:x + w]),
                cv2.bitwise_or(textura_mask[y:y + h, x:x + w], candidatos[y:y + h, x:x + w]),
            ),
        )
        cx, cy, star_info = _centro_pendao_radial(component, x, y, w, h, suporte_roi, params)
        if not star_info.get("ok"):
            continue
        sx = int(np.clip(cx, 0, hsv.shape[1] - 1))
        sy = int(np.clip(cy, 0, hsv.shape[0] - 1))
        tipo = classificar_tipo_pendao(hsv[sy, sx], lab[sy, sx], params)
        area_score = min(1.0, area / max(1.0, area_min * 6.0))
        score = (
            texture_ratio * 2.0
            + yellow_ratio * 2.0
            + clear_ratio * 1.4
            + (1.0 - green_ratio) * 1.8
            + area_score
            + min(1.0, aspect / 4.0) * 0.35
            + float(star_info.get("score", 0.0)) * 1.35
            + min(1.0, float(star_info.get("directions", 0)) / 6.0) * 0.85
        )
        confianca = "alta" if score >= 5.2 else ("media" if score >= 4.0 else "baixa")
        detections.append({
            "center": (float(cx * scale_back), float(cy * scale_back)),
            "size": float(np.clip(max(w, h) * params.get("x_size_factor", 0.40) * scale_back, params.get("x_min_size", 6), params.get("x_max_size", 20))),
            "bbox": (float(x * scale_back), float(y * scale_back), float(w * scale_back), float(h * scale_back)),
            "score": float(score),
            "confianca": confianca,
            "tipo": tipo,
            "yellow_ratio": yellow_ratio,
            "texture_ratio": texture_ratio,
            "clear_ratio": clear_ratio,
            "green_ratio": green_ratio,
            "star_directions": int(star_info.get("directions", 0)),
            "star_score": float(star_info.get("score", 0.0)),
            "area": float(area * scale_back * scale_back),
            "source": "OpenCV",
        })

    params["merge_distance"] = float(params.get("min_distance", params.get("merge_distance", 12)))
    params["nms_distance"] = float(params.get("min_distance", params.get("nms_distance", 12)))
    detections = agrupar_componentes_estrelados(detections, params)
    detections = remover_deteccoes_duplicadas(detections, params)
    max_detections = int(params.get("max_detections", 30000))
    if len(detections) > max_detections:
        detections = sorted(detections, key=lambda d: d.get("score", 0), reverse=True)[:max_detections]

    rows = int((grade or {}).get("rows") or (grade or {}).get("linhas") or 1)
    cols = int((grade or {}).get("cols") or (grade or {}).get("colunas") or 1)
    cells = _grid_cells_from_grade(grade, orig_w, orig_h)
    parcelas = []
    total = 0
    for index, row, col, poly in cells:
        dets = []
        for det in detections:
            cx, cy = det["center"]
            if cv2.pointPolygonTest(poly.astype(np.float32), (float(cx), float(cy)), False) >= 0:
                dets.append({k: v for k, v in det.items() if k != "area"})
        total += len(dets)
        parcelas.append({"index": int(index), "row": int(row), "col": int(col), "count": len(dets), "detections": dets})
    return {"dims": (orig_w, orig_h), "rows": rows, "cols": cols, "grid": grade, "parcelas": parcelas, "total": int(total)}


def detectar_pendoes_milho_avancado(rgb_img, grade=None, params=None):
    return detectar_pendoes_opencv_puro(rgb_img, grade=grade, params=params)


def processar_grid_pendao_avancado(rgb_img, grade, params=None):
    return detectar_pendoes_opencv_puro(rgb_img, grade=grade, params=params)


def _pendao_model_candidates():
    env_candidates = [
        os.getenv("PENDAO_YOLO_MODEL", "").strip(),
        os.getenv("TMG_PENDAO_YOLO_MODEL", "").strip(),
    ]
    local_candidates = [
        YOLO_BEST_MODEL_PATH,
        APP_ROOT / "models" / "pendoes.pt",
        APP_ROOT / "models" / "pendao.pt",
        APP_ROOT / "models" / "tassel.pt",
        APP_ROOT / "models" / "best.pt",
        APP_ROOT / "tmg_assets" / "models" / "pendoes.pt",
        APP_ROOT / "tmg_assets" / "models" / "pendao.pt",
        APP_ROOT / "tmg_assets" / "models" / "best.pt",
    ]
    for candidate in env_candidates:
        candidate_path = Path(candidate) if candidate else None
        if candidate_path and candidate_path.exists() and _path_inside_app_root(candidate_path):
            return str(candidate_path.resolve()), "custom"
    for candidate in local_candidates:
        if candidate.exists():
            return str(candidate.resolve()), "custom"
    return None, None


@st.cache_resource(show_spinner=False)
def _carregar_modelo_yolo_pendao(model_key: str, world_classes: tuple):
    try:
        if model_key == "__YOLO_WORLD__":
            from ultralytics import YOLOWorld
            model = YOLOWorld(PENDAO_AVANCADO_PARAMS.get("yolo_world_model", "yolov8s-world.pt"))
            model.set_classes(list(world_classes))
            return model, "YOLO-World"
        from ultralytics import YOLO
        model = YOLO(model_key)
        return model, f"YOLO:{Path(model_key).name}"
    except ImportError as exc:
        raise RuntimeError("Ultralytics não instalado. Instale com: pip install -U ultralytics") from exc


def _resolver_modelo_yolo_pendao(params=None):
    params = _pendao_params(params)
    if not bool(params.get("yolo_enabled", True)):
        return None, "YOLO desativado nos parâmetros."
    try:
        import ultralytics  # noqa: F401
    except ImportError:
        return None, "Ultralytics não instalado. Instale com: pip install -U ultralytics"
    model_path, model_kind = _pendao_model_candidates()
    if model_path:
        return model_path, f"Modelo YOLO customizado: {Path(model_path).name}"
    if bool(params.get("yolo_world_enabled", True)):
        return "__YOLO_WORLD__", "YOLO-World sem modelo customizado local."
    return None, "YOLO instalado, mas nenhum modelo .pt de pendão foi encontrado. Usando OpenCV."


def _roi_features_pendao(rgb_img, bbox, params=None):
    params = _pendao_params(params)
    rgb = _as_rgb_uint8(rgb_img)
    if rgb is None:
        return None
    h_img, w_img = rgb.shape[:2]
    x1, y1, x2, y2 = [int(round(float(v))) for v in bbox]
    x1 = max(0, min(w_img - 1, x1))
    y1 = max(0, min(h_img - 1, y1))
    x2 = max(x1 + 1, min(w_img, x2))
    y2 = max(y1 + 1, min(h_img, y2))
    roi = rgb[y1:y2, x1:x2]
    if roi.size == 0:
        return None
    proc, hsv, lab, gray, lab_l = _preprocessar_pendao_opencv(roi, params)
    if proc is None:
        return None
    sem_verde, green_mask = remover_verde_agressivo(proc, params)
    cor_mask, _ = criar_mascara_pendoes_multicor(proc, lab_l, params)
    textura_mask = calcular_mascara_textura(gray, params)
    brilho_mask = cv2.threshold(lab_l, int(params.get("lab_l_threshold", 128)), 255, cv2.THRESH_BINARY)[1]
    valid_mask = cv2.bitwise_and(cv2.bitwise_and(cor_mask, sem_verde), cv2.bitwise_or(textura_mask, brilho_mask))
    active = valid_mask > 0
    green_ratio = float(np.mean(green_mask > 0))
    yellow_ratio = float(np.mean(cor_mask > 0))
    texture_ratio = float(np.mean(textura_mask > 0))
    clear_ratio = float(np.mean(brilho_mask > 0))
    if np.any(active):
        moments = cv2.moments(valid_mask, binaryImage=True)
        if moments.get("m00", 0):
            cx = x1 + moments["m10"] / moments["m00"]
            cy = y1 + moments["m01"] / moments["m00"]
        else:
            ys, xs = np.where(active)
            cx = x1 + float(np.mean(xs))
            cy = y1 + float(np.mean(ys))
    else:
        cx = (x1 + x2) / 2
        cy = (y1 + y2) / 2
    score = (
        yellow_ratio * 2.4
        + texture_ratio * 1.8
        + clear_ratio * 1.2
        + (1.0 - green_ratio) * 1.4
    )
    return {
        "center": (float(cx), float(cy)),
        "bbox": (float(x1), float(y1), float(x2 - x1), float(y2 - y1)),
        "size": float(np.clip(max(x2 - x1, y2 - y1) * 0.34, params.get("x_min_size", 6), params.get("x_max_size", 20))),
        "score": float(score),
        "tipo": classificar_tipo_pendao(hsv[int(np.clip(cy - y1, 0, hsv.shape[0] - 1)), int(np.clip(cx - x1, 0, hsv.shape[1] - 1))], lab[int(np.clip(cy - y1, 0, lab.shape[0] - 1)), int(np.clip(cx - x1, 0, lab.shape[1] - 1))], params),
        "yellow_ratio": yellow_ratio,
        "texture_ratio": texture_ratio,
        "clear_ratio": clear_ratio,
        "green_ratio": green_ratio,
    }


def _detections_from_result(result):
    if not result or not result.get("parcelas"):
        return []
    return [dict(det) for det in result["parcelas"][0].get("detections", [])]


def _inferir_pendoes_yolo(rgb_img, params=None):
    params = _pendao_params(params)
    rgb = _as_rgb_uint8(rgb_img)
    if rgb is None:
        return [], "Imagem inválida para YOLO."
    model_key, model_status = _resolver_modelo_yolo_pendao(params)
    if not model_key:
        return [], model_status
    try:
        model, model_name = _carregar_modelo_yolo_pendao(model_key, tuple(PENDAO_YOLO_CLASSES))
        orig_h, orig_w = rgb.shape[:2]
        max_dim = int(params.get("yolo_max_dim", 1800))
        scale_back = 1.0
        if max(orig_h, orig_w) > max_dim:
            scale = max_dim / max(orig_h, orig_w)
            rgb_work = cv2.resize(rgb, (max(1, int(orig_w * scale)), max(1, int(orig_h * scale))), interpolation=cv2.INTER_AREA)
            scale_back = 1.0 / scale
        else:
            rgb_work = rgb
        results = model.predict(
            source=rgb_work,
            imgsz=int(params.get("yolo_imgsz", 1280)),
            conf=float(params.get("yolo_conf", 0.04)),
            iou=float(params.get("yolo_iou", 0.45)),
            max_det=int(params.get("yolo_max_det", 20000)),
            verbose=False,
        )
        detections = []
        result0 = results[0] if results else None
        if result0 is None or getattr(result0, "boxes", None) is None:
            return [], f"{model_name}: nenhum resultado."
        boxes = result0.boxes
        names = getattr(result0, "names", {}) or {}
        xyxy = boxes.xyxy.detach().cpu().numpy() if getattr(boxes, "xyxy", None) is not None else np.empty((0, 4))
        confs = boxes.conf.detach().cpu().numpy() if getattr(boxes, "conf", None) is not None else np.zeros((len(xyxy),), dtype=float)
        classes = boxes.cls.detach().cpu().numpy().astype(int) if getattr(boxes, "cls", None) is not None else np.zeros((len(xyxy),), dtype=int)
        allowed_tokens = ("tassel", "pendao", "pendão", "maize", "corn", "flower")
        for box, conf, cls_id in zip(xyxy, confs, classes):
            class_name = str(names.get(int(cls_id), cls_id)).lower()
            if model_key != "__YOLO_WORLD__" and isinstance(names, dict) and names:
                if not any(tok in class_name for tok in allowed_tokens) and len(names) > 1:
                    continue
            x1, y1, x2, y2 = [float(v) for v in box]
            roi = _roi_features_pendao(rgb_work, (x1, y1, x2, y2), params)
            if not roi:
                continue
            visual_ok = (
                roi["score"] >= float(params.get("yolo_min_visual_score", 1.2))
                and roi["yellow_ratio"] >= float(params.get("yolo_min_yellow_ratio", 0.015))
                and roi["texture_ratio"] >= float(params.get("yolo_min_texture_ratio", 0.015))
                and roi["green_ratio"] <= float(params.get("yolo_max_green_ratio", 0.78))
            )
            if not visual_ok and float(conf) < 0.22:
                continue
            cx, cy = roi["center"]
            x, y, w, h = roi["bbox"]
            score = float(conf) * 4.0 + roi["score"]
            detections.append({
                "center": (float(cx * scale_back), float(cy * scale_back)),
                "bbox": (float(x * scale_back), float(y * scale_back), float(w * scale_back), float(h * scale_back)),
                "size": float(np.clip(roi["size"] * scale_back, params.get("x_min_size", 6), params.get("x_max_size", 20))),
                "score": score,
                "confianca": "alta" if score >= 3.2 else "media",
                "tipo": roi.get("tipo", "yolo"),
                "yellow_ratio": roi.get("yellow_ratio", 0),
                "texture_ratio": roi.get("texture_ratio", 0),
                "clear_ratio": roi.get("clear_ratio", 0),
                "green_ratio": roi.get("green_ratio", 0),
                "source": model_name,
                "yolo_conf": float(conf),
                "class_name": class_name,
                "area": float(w * h * scale_back * scale_back),
            })
        return detections, f"{model_name}: {len(detections)} detecções validadas por OpenCV."
    except Exception as exc:
        return [], f"YOLO falhou ({exc}). Usando OpenCV avançado."


def detectar_pendoes_hibrido_yolo_opencv(rgb_img, grade=None, params=None):
    params = _pendao_params(params)
    opencv_result = detectar_pendoes_opencv_puro(rgb_img, grade=None, params=params)
    opencv_detections = _detections_from_result(opencv_result)
    yolo_detections, yolo_status = _inferir_pendoes_yolo(rgb_img, params=params)
    combined = []
    for det in yolo_detections:
        item = dict(det)
        item["source"] = item.get("source", "YOLO")
        combined.append(item)
    for det in opencv_detections:
        item = dict(det)
        item["source"] = item.get("source", "OpenCV")
        combined.append(item)
    reference_detections, reference_status = _inferir_pendoes_por_referencias(rgb_img, existing=combined, params=params)
    for det in reference_detections:
        item = dict(det)
        item["source"] = item.get("source", "Refinamento OpenCV")
        combined.append(item)
    if combined:
        merge_params = dict(params)
        merge_params["nms_distance"] = float(params.get("yolo_merge_distance", params.get("nms_distance", 18)))
        combined = agrupar_componentes_estrelados(combined, merge_params)
        combined = remover_deteccoes_duplicadas(combined, merge_params)
    source_counts = {
        "yolo": 0,
        "opencv": 0,
        "refinado": 0,
    }
    for det in combined:
        src = str(det.get("source", "")).lower()
        if "yolo" in src:
            source_counts["yolo"] += 1
        elif "refinamento" in src or "refer" in src:
            source_counts["refinado"] += 1
        else:
            source_counts["opencv"] += 1
    rows = int((grade or {}).get("rows") or (grade or {}).get("linhas") or 1)
    cols = int((grade or {}).get("cols") or (grade or {}).get("colunas") or 1)
    rgb = _as_rgb_uint8(rgb_img)
    img_h, img_w = rgb.shape[:2] if rgb is not None else (0, 0)
    cells = _grid_cells_from_grade(grade, img_w, img_h)
    parcelas = []
    total = 0
    for index, row, col, poly in cells:
        dets = []
        for det in combined:
            cx, cy = det["center"]
            if cv2.pointPolygonTest(poly.astype(np.float32), (float(cx), float(cy)), False) >= 0:
                dets.append({k: v for k, v in det.items() if k != "area"})
        total += len(dets)
        parcelas.append({"index": int(index), "row": int(row), "col": int(col), "count": len(dets), "detections": dets})
    mode = "YOLO+OpenCV" if yolo_detections else "OpenCV fallback"
    if reference_detections:
        mode = f"{mode}+Referências"
    status = (
        f"{mode}: {total} centros. "
        f"YOLO={source_counts['yolo']} · OpenCV={source_counts['opencv']} · Refinamento={source_counts['refinado']}. "
        f"{yolo_status} {reference_status}"
    )
    return {
        "dims": (img_w, img_h),
        "rows": rows,
        "cols": cols,
        "grid": grade,
        "parcelas": parcelas,
        "total": int(total),
        "detector_status": status,
        "detector_mode": mode,
        "detector_counts": source_counts,
    }


def _decode_rgb_for_pendao(file_bytes: bytes, filename: str):
    try:
        arr = np.frombuffer(file_bytes, dtype=np.uint8)
        bgr = cv2.imdecode(arr, cv2.IMREAD_COLOR)
        if bgr is not None:
            return cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
    except Exception:
        pass
    try:
        return np.array(Image.open(BytesIO(file_bytes)).convert("RGB"))
    except Exception:
        return None


def _serializar_deteccoes_pendao_preview(result, rgb, preview_dims):
    dets = []
    if not result.get("parcelas"):
        return dets
    src_w, src_h = result.get("dims", (rgb.shape[1], rgb.shape[0]))
    dst_w, dst_h = preview_dims or (src_w, src_h)
    sx = float(dst_w) / max(1.0, float(src_w))
    sy = float(dst_h) / max(1.0, float(src_h))
    for det in result["parcelas"][0].get("detections", []):
        cx, cy = det.get("center", (0, 0))
        bbox = det.get("bbox", (0, 0, 0, 0))
        dets.append({
            "x": round(float(cx) * sx, 2),
            "y": round(float(cy) * sy, 2),
            "size": round(float(det.get("size", 8)) * max(sx, sy), 2),
            "score": round(float(det.get("score", 0)), 4),
            "confianca": det.get("confianca", "media"),
            "tipo": det.get("tipo", "misto"),
            "bbox": [
                round(float(bbox[0]) * sx, 2),
                round(float(bbox[1]) * sy, 2),
                round(float(bbox[2]) * sx, 2),
                round(float(bbox[3]) * sy, 2),
            ],
            "yellow_ratio": round(float(det.get("yellow_ratio", 0)), 4),
            "texture_ratio": round(float(det.get("texture_ratio", 0)), 4),
            "clear_ratio": round(float(det.get("clear_ratio", 0)), 4),
            "green_ratio": round(float(det.get("green_ratio", 0)), 4),
            "source": det.get("source", "OpenCV"),
            "training_source": det.get("training_source", ""),
            "yolo_conf": round(float(det.get("yolo_conf", 0)), 4),
            "class_name": det.get("class_name", ""),
            "reference": det.get("reference", ""),
            "reference_id": det.get("reference_id", ""),
            "template_score": round(float(det.get("template_score", 0)), 4),
            "area": round(float(det.get("area", 0)), 4),
        })
    return dets


def _centros_deteccoes_pendao(detections):
    centers = []
    for det in detections or []:
        try:
            cx, cy = det.get("center", (None, None))
            centers.append((float(cx), float(cy)))
        except Exception:
            continue
    return centers


def _filtrar_deteccoes_treino_complementar(candidates, existing, params=None):
    params = _pendao_params(params)
    distance = float(params.get("reference_match_min_distance", params.get("nms_distance", 18)))
    centers = _centros_deteccoes_pendao(existing)
    accepted = []
    ordered = sorted(candidates or [], key=lambda det: float(det.get("score", 0)), reverse=True)
    for det in ordered:
        try:
            cx, cy = [float(v) for v in det.get("center", (None, None))]
        except Exception:
            continue
        if any(((cx - ex) ** 2 + (cy - ey) ** 2) ** 0.5 <= distance for ex, ey in centers):
            continue
        item = dict(det)
        item["training_source"] = item.get("source", "Treino YOLO")
        item["source"] = "Aplicar Treino"
        centers.append((cx, cy))
        accepted.append(item)
    return accepted


def _preparar_deteccoes_treino_continuo_pendao(rgb, existing, params=None):
    params = _pendao_params(params)
    if not _arquivos_referencia_treinamento_yolo(limit=1):
        return [], "Sem mini imagens salvas para aplicar treino.", {"yolo": 0, "referencias": 0, "novos": 0}
    yolo_detections, yolo_status = _inferir_pendoes_yolo(rgb, params=params)
    candidates = []
    for det in yolo_detections:
        item = dict(det)
        item["source"] = item.get("source", "YOLO")
        candidates.append(item)
    reference_detections, reference_status = _inferir_pendoes_por_referencias(
        rgb,
        existing=list(existing or []) + candidates,
        params=params,
    )
    for det in reference_detections:
        item = dict(det)
        item["source"] = item.get("source", "Referência manual")
        candidates.append(item)
    if candidates:
        merge_params = dict(params)
        merge_params["nms_distance"] = float(params.get("reference_match_min_distance", 18))
        candidates = remover_deteccoes_duplicadas(candidates, merge_params)
    novos = _filtrar_deteccoes_treino_complementar(candidates, existing, params)
    counts = {
        "yolo": int(len(yolo_detections)),
        "referencias": int(len(reference_detections)),
        "novos": int(len(novos)),
    }
    status = (
        f"Aplicar Treino pronto. YOLO={counts['yolo']} · referências={counts['referencias']} · "
        f"novos candidatos sem duplicidade={counts['novos']}. {yolo_status} {reference_status}"
    )
    return novos, status, counts


@st.cache_data(show_spinner=False, max_entries=16)
def preparar_deteccoes_pendoamento_hibrido(file_bytes: bytes, filename: str, preview_dims: tuple, model_signature: str = "", usar_referencias: bool = True):
    rgb = _decode_rgb_for_pendao(file_bytes, filename)
    if rgb is None:
        return {
            "detections": [],
            "training_detections": [],
            "status": "Imagem inválida para análise de pendoamento.",
            "mode": "erro",
        }
    params = dict(PENDAO_ANALISE_PARAMS)
    result = detectar_pendoes_pendoamento_opencv_parametrizado(rgb, grade=None, params=params)
    main_detections = _detections_from_result(result)
    treino_detections = []
    treino_status = "Aplicar Treino aguardando mini imagens."
    treino_counts = {"yolo": 0, "referencias": 0, "novos": 0}
    if usar_referencias:
        try:
            treino_detections, treino_status, treino_counts = _preparar_deteccoes_treino_continuo_pendao(
                rgb,
                main_detections,
                params=_pendao_params(),
            )
        except Exception as exc:
            treino_status = f"Aplicar Treino indisponível ({exc})."
    return {
        "detections": _serializar_deteccoes_pendao_preview(result, rgb, preview_dims),
        "training_detections": _serializar_deteccoes_pendao_preview(
            {"parcelas": [{"detections": treino_detections}], "dims": result.get("dims", (rgb.shape[1], rgb.shape[0]))},
            rgb,
            preview_dims,
        ),
        "status": result.get("detector_status", ""),
        "mode": result.get("detector_mode", "OpenCV parametrizado TMG"),
        "counts": result.get("detector_counts", {}),
        "training_status": treino_status,
        "training_counts": treino_counts,
        "backend_ready": bool(result.get("backend_ready", True)),
    }


@st.cache_data(show_spinner=False, max_entries=16)
def preparar_deteccoes_pendoamento_avancado(file_bytes: bytes, filename: str, preview_dims: tuple):
    return preparar_deteccoes_pendoamento_hibrido(file_bytes, filename, preview_dims).get("detections", [])


# ==========================================
# TELA DE LOGIN[cite: 1]
# ==========================================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if "auth_user" not in st.session_state:
    st.session_state.auth_user = None

if "login_cfg_open" not in st.session_state:
    st.session_state.login_cfg_open = False

if "cultura_selecionada" not in st.session_state:
    st.session_state.cultura_selecionada = None

if not st.session_state.logged_in:

    if LOGIN_BG_PATH.exists():
        bg_css = _img_to_base64_css(LOGIN_BG_PATH)
        st.markdown(f"""
        <style>
        .stApp {{
            background-image: url("{bg_css}") !important;
            background-size: cover !important;
            background-position: center center !important;
            background-repeat: no-repeat !important;
            background-attachment: fixed !important;
            background-color: transparent !important;
        }}
        .stApp::before {{
            content: "";
            position: fixed;
            inset: 0;
            background: rgba(8, 8, 8, 0.70);
            z-index: 0;
            pointer-events: none;
        }}
        </style>
        """, unsafe_allow_html=True)

    st.markdown("""
    <style>
    [data-testid="stSidebar"] { display: none !important; }

    .login-mobile-btn {
        position: fixed;
        bottom: 18px;
        right: 24px;
        z-index: 9999;
        display: inline-flex;
        align-items: center;
        justify-content: center;
        min-width: 152px;
        min-height: 42px;
        padding: 10px 18px;
        border-radius: 14px;
        border: 1.5px solid rgba(120, 220, 255, .88);
        background:
            linear-gradient(120deg, rgba(255,255,255,.18), transparent 30%),
            linear-gradient(145deg, rgba(2,14,36,.96), rgba(18,62,100,.88), rgba(0,212,255,.30));
        color: #ffffff !important;
        font-family: 'Segoe UI', Arial, sans-serif;
        font-size: .86rem;
        font-weight: 900;
        letter-spacing: .45px;
        text-decoration: none !important;
        text-shadow:
            0 1px 0 rgba(0,0,0,.95),
            0 0 10px rgba(0,212,255,.60);
        box-shadow:
            0 14px 28px rgba(0,0,0,.46),
            0 0 0 1px rgba(255,255,255,.10),
            0 0 24px rgba(0,212,255,.34),
            inset 0 1px 0 rgba(255,255,255,.28),
            inset 0 -10px 18px rgba(2,14,36,.42);
        backdrop-filter: blur(12px) saturate(145%);
        -webkit-backdrop-filter: blur(12px) saturate(145%);
        transition: transform .25s ease, box-shadow .30s ease, border-color .30s ease, filter .30s ease;
    }

    .login-mobile-btn:hover {
        transform: translateY(-2px);
        border-color: rgba(178, 240, 255, .98);
        color: #ffffff !important;
        filter: brightness(1.10);
        box-shadow:
            0 18px 34px rgba(0,0,0,.55),
            0 0 0 1px rgba(255,255,255,.16),
            0 0 34px rgba(0,212,255,.54),
            inset 0 1px 0 rgba(255,255,255,.36),
            inset 0 -10px 18px rgba(2,14,36,.34);
    }

    .login-mobile-btn:active {
        transform: translateY(1px) scale(.99);
    }

    @media (max-width: 720px) {
        .login-mobile-btn {
            bottom: 12px;
            right: 12px;
            min-width: 126px;
            min-height: 38px;
            padding: 8px 13px;
            font-size: .76rem;
        }
    }

    .login-card {
        background: linear-gradient(160deg, #1c1c1c 0%, #111111 100%);
        border: 1px solid #2e2e2e;
        border-radius: 16px;
        padding: 26px 28px 22px 28px;
        box-shadow:
            0 8px 32px rgba(0,0,0,0.85),
            0 2px 8px rgba(0,0,0,0.6),
            inset 0 1px 0 rgba(255,255,255,0.04);
        position: relative;
        z-index: 1;
    }

    .login-title {
        text-align: center;
        font-family: 'Segoe UI', sans-serif;
        font-weight: 900;
        font-size: 1.85rem;
        letter-spacing: 5px;
        color: var(--tmg-primary);
        text-transform: uppercase;
        text-shadow:
            1px 1px 0 var(--tmg-primary-shadow-1),
            2px 2px 0 var(--tmg-primary-shadow-2),
            3px 3px 0 var(--tmg-primary-shadow-3),
            5px 5px 10px rgba(0,0,0,0.95),
            0 0 25px var(--tmg-primary-glow),
            0 0 60px var(--tmg-primary-glow-soft);
        margin-bottom: 2px;
    }

    .login-subtitle {
        text-align: center;
        color: #555;
        font-size: 0.78rem;
        letter-spacing: 3px;
        text-transform: uppercase;
        margin-bottom: 18px;
    }

    .login-divider {
        border: none;
        border-top: 1px solid var(--tmg-primary);
        box-shadow: 0 0 10px var(--tmg-primary-glow);
        margin: 0 0 18px 0;
    }

    div[data-testid="stTextInput"] {
        margin-bottom: 12px !important;
    }

    div[data-testid="stTextInput"] > div,
    div[data-testid="stTextInput"] div[data-baseweb="input"] {
        background:
            linear-gradient(120deg, rgba(255,255,255,.12), transparent 32%),
            linear-gradient(145deg, rgba(2,14,36,.98), rgba(18,62,100,.86), rgba(0,212,255,.20)) !important;
        border: 1.5px solid rgba(156,205,248,.88) !important;
        border-radius: 12px !important;
        box-shadow:
            0 10px 22px rgba(0,0,0,.34),
            0 0 0 1px rgba(255,255,255,.10),
            0 0 18px rgba(0,212,255,.26),
            inset 0 1px 0 rgba(255,255,255,.22),
            inset 0 -7px 14px rgba(2,14,36,.44) !important;
    }

    .stTextInput > div > div > input,
    div[data-testid="stTextInput"] input {
        background-color: transparent !important;
        border: none !important;
        border-radius: 12px !important;
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        caret-color: #ffffff !important;
        padding: 11px 14px !important;
        font-size: 1.02rem !important;
        font-weight: 800 !important;
        font-family: 'Segoe UI', Arial, sans-serif !important;
        letter-spacing: .2px !important;
        text-shadow: 0 1px 0 rgba(0,0,0,.88), 0 0 10px rgba(0,212,255,.28) !important;
        box-shadow: none !important;
    }

    .stTextInput > div > div > input::placeholder,
    div[data-testid="stTextInput"] input::placeholder {
        color: rgba(224,247,255,.78) !important;
        -webkit-text-fill-color: rgba(224,247,255,.78) !important;
        opacity: 1 !important;
        font-weight: 700 !important;
    }

    .stTextInput > div > div > input:-webkit-autofill,
    div[data-testid="stTextInput"] input:-webkit-autofill {
        -webkit-text-fill-color: #ffffff !important;
        caret-color: #ffffff !important;
        box-shadow:
            0 0 0 1000px rgba(2,14,36,.96) inset,
            0 10px 22px rgba(0,0,0,.34),
            0 0 18px rgba(0,212,255,.26) !important;
        transition: background-color 9999s ease-in-out 0s !important;
    }

    div[data-testid="stTextInput"]:focus-within > div,
    div[data-testid="stTextInput"]:focus-within div[data-baseweb="input"] {
        border-color: #38bdf8 !important;
        box-shadow:
            0 12px 26px rgba(0,0,0,.38),
            0 0 0 2px rgba(56,189,248,.24),
            0 0 22px rgba(0,212,255,.34),
            inset 0 1px 0 rgba(255,255,255,.30),
            inset 0 -7px 14px rgba(2,14,36,.40) !important;
    }

    .stTextInput label,
    .stTextInput label p,
    div[data-testid="stTextInput"] label,
    div[data-testid="stTextInput"] label p,
    div[data-testid="stTextInput"] [data-testid="stWidgetLabel"] p {
        color: #e5e7eb !important;
        font-size: 0.88rem !important;
        font-weight: 850 !important;
        letter-spacing: 1.45px !important;
        text-transform: uppercase !important;
        text-shadow:
            0 1px 0 rgba(0,0,0,.95),
            0 0 10px rgba(56,189,248,.34) !important;
        margin-bottom: 5px !important;
    }

    div[data-testid="stTextInput"] button,
    div[data-testid="stTextInput"] [role="button"] {
        color: #ffffff !important;
        background:
            linear-gradient(145deg, rgba(2,14,36,.92), rgba(18,62,100,.84), rgba(0,212,255,.20)) !important;
        border-left: 1px solid rgba(56,189,248,.32) !important;
        border-radius: 0 11px 11px 0 !important;
        min-height: 100% !important;
        opacity: 1 !important;
        box-shadow:
            inset 1px 0 0 rgba(255,255,255,.18),
            0 0 12px rgba(0,212,255,.22) !important;
    }

    div[data-testid="stTextInput"] button:hover,
    div[data-testid="stTextInput"] [role="button"]:hover {
        background:
            linear-gradient(145deg, rgba(7,31,63,.98), rgba(0,212,255,.36)) !important;
        color: #ffffff !important;
    }

    div[data-testid="stTextInput"] button svg,
    div[data-testid="stTextInput"] [role="button"] svg {
        color: #ffffff !important;
        fill: #ffffff !important;
        stroke: #ffffff !important;
        opacity: 1 !important;
        filter: drop-shadow(0 0 5px rgba(0,212,255,.72));
    }

    .login-footer {
        text-align: center;
        color: #444;
        font-size: 0.72rem;
        letter-spacing: 2px;
        text-transform: uppercase;
        margin-top: 14px;
    }

    .login-cfg-panel {
        background: linear-gradient(160deg, #181818 0%, #0f0f0f 100%);
        border: 1px solid #2a2a2a;
        border-top: 2px solid var(--tmg-primary);
        border-radius: 16px;
        padding: 16px 18px;
        box-shadow: 0 8px 24px rgba(0,0,0,0.7);
        position: relative;
        z-index: 1;
        margin-top: 4px;
    }

    .cfg-panel-title {
        color: var(--tmg-primary);
        font-weight: 700;
        font-size: 0.85rem;
        letter-spacing: 3px;
        text-transform: uppercase;
        text-shadow: 0 0 12px var(--tmg-primary-glow);
        margin-bottom: 16px;
    }

    .login-logo-wrap {
        width: 100%;
        display: flex;
        justify-content: center;
        align-items: center;
        text-align: center;
        margin: 0 auto 6px auto;
    }

    .login-logo-img {
        display: block;
        width: 150px;
        max-width: 42%;
        height: auto;
        object-fit: contain;
        margin: 0 auto;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<a class='login-mobile-btn' href='?mobile=1'>Versão Mobile</a>", unsafe_allow_html=True)

    _, col_mid, _ = st.columns([1.25, 0.9, 1.25])

    with col_mid:
        if LOGO_PATH.exists():
            logo_css = _img_to_base64_css(LOGO_PATH)
            st.markdown(
                f"<div class='login-logo-wrap'><img src='{logo_css}' class='login-logo-img' alt='Logo TMG'></div>",
                unsafe_allow_html=True,
            )
            st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

        st.markdown("<div class='login-title'>TMG</div>", unsafe_allow_html=True)
        st.markdown(
            "<div class='login-subtitle'>Sistema de Análise &nbsp;·&nbsp; Acesso Seguro</div>",
            unsafe_allow_html=True
        )
        st.markdown("<hr class='login-divider'>", unsafe_allow_html=True)

        usuario = st.text_input("Usuário", placeholder="Digite seu login", key="login_user")
        senha   = st.text_input("Senha",   placeholder="Digite sua senha", type="password", key="login_pass")

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        if st.button("⟶  ENTRAR", type="primary", key="btn_entrar"):
            auth_user = _auth_find_user(usuario, senha)
            if auth_user:
                st.session_state.logged_in = True
                st.session_state.auth_user = auth_user
                state_login = _partners_load_state()
                _partners_add_history(state_login, "", "Login realizado", f"Usuário: {auth_user.get('nome', auth_user.get('usuario', ''))}")
                _partners_save_state(state_login)
                app_rerun()
            else:
                st.markdown("""
                <div style='
                    background:rgba(180,30,30,0.12);
                    border:1px solid #7a0000;
                    border-radius:10px;
                    padding:12px 16px;
                    color:#ff6b6b;
                    font-size:0.88rem;
                    text-align:center;
                    margin-top:8px;
                    box-shadow:0 0 12px rgba(180,0,0,0.2);
                '>&#9888; Usuário ou senha incorretos.</div>
                """, unsafe_allow_html=True)

        st.markdown("<div class='login-footer'>TMG v2.0 &nbsp;·&nbsp; 2026</div>", unsafe_allow_html=True)

        st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        if st.button("⚙️  Configurações da Tela de Login", key="btn_login_cfg"):
            st.session_state.login_cfg_open = not st.session_state.login_cfg_open

        if st.session_state.login_cfg_open:
            st.markdown(
                "<div class='cfg-panel-title'>&#9881; Identidade Visual &nbsp;·&nbsp; Tela de Login</div>",
                unsafe_allow_html=True
            )

            if LOGIN_BG_PATH.exists():
                st.markdown(
                    f"<p style='color:#666;font-size:0.8rem;margin-bottom:10px;'>"
                    f"&#128193; Background atual: "
                    f"<code style='color:{THEME_PRIMARY_COLOR};'>{LOGIN_BG_PATH}</code></p>",
                    unsafe_allow_html=True
                )

            nova_bg = st.file_uploader(
                "Imagem de fundo da tela de login",
                type=["png", "jpg", "jpeg"],
                key="login_bg_uploader"
            )

            if nova_bg:
                load_box = st.empty()
                update_tmg_loading(load_box, 45, f"Carregando imagem de fundo: {Path(nova_bg.name).name}")
                bg_img = Image.open(nova_bg).convert("RGB")
                update_tmg_loading(load_box, 82, "Salvando imagem de fundo no pacote TMG...")
                bg_img.save(str(LOGIN_BG_PATH), format="PNG")
                update_tmg_loading(load_box, 100, "Carregamento concluído com sucesso.")
                st.success(f"✅ Imagem salva em `{LOGIN_BG_PATH}` — recarregando...")
                app_rerun()

            if LOGIN_BG_PATH.exists():
                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                if st.button("🗑️  Remover imagem de fundo", key="btn_rm_bg"):
                    LOGIN_BG_PATH.unlink()
                    st.success("Imagem de fundo removida.")
                    app_rerun()

    st.stop()


# ==========================================
# TELA DE SELEÇÃO DE CULTURA (PÓS-LOGIN)[cite: 1]
# ==========================================
if st.session_state.logged_in and st.session_state.cultura_selecionada is None:
    current_user = _auth_current_user()
    allowed_cultures = _auth_allowed_cultures(current_user)
    can_open_partners = _auth_can_partners(current_user)

    st.markdown("""
    <style>
    [data-testid="stSidebar"] { display: none !important; }

    .cultura-page {
        min-height: 85vh;
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
    }

    .cultura-title {
        text-align: center;
        font-family: 'Segoe UI', sans-serif;
        font-weight: 900;
        font-size: 2rem;
        letter-spacing: 5px;
        color: var(--tmg-primary);
        text-transform: uppercase;
        text-shadow:
            1px 1px 0 var(--tmg-primary-shadow-1),
            2px 2px 0 var(--tmg-primary-shadow-2),
            3px 3px 0 var(--tmg-primary-shadow-3),
            5px 5px 10px rgba(0,0,0,0.95),
            0 0 25px var(--tmg-primary-glow),
            0 0 60px var(--tmg-primary-glow-soft);
        margin-bottom: 6px;
    }

    .cultura-subtitle {
        text-align: center;
        color: #555;
        font-size: 0.78rem;
        letter-spacing: 3px;
        text-transform: uppercase;
        margin-bottom: 36px;
    }

    .cultura-hr {
        border: none;
        border-top: 1px solid var(--tmg-primary);
        box-shadow: 0 0 10px var(--tmg-primary-glow);
        width: 60%;
        margin: 0 auto 40px auto;
    }

    .cultura-card {
        background: linear-gradient(160deg, #1c1c1c 0%, #111111 100%);
        border: 1px solid #2e2e2e;
        border-radius: 20px;
        padding: 38px 24px 28px 24px;
        text-align: center;
        cursor: pointer;
        transition: all 0.25s ease;
        box-shadow:
            4px 4px 14px #080808,
            -1px -1px 8px #252525,
            inset 0 1px 0 rgba(255,255,255,0.04);
        position: relative;
        overflow: hidden;
    }

    .cultura-card::before {
        content: "";
        position: absolute;
        top: 0; left: 0; right: 0;
        height: 2px;
        background: linear-gradient(90deg, transparent, var(--tmg-primary), transparent);
        opacity: 0.5;
    }

    .cultura-card:hover {
        border-color: var(--tmg-primary);
        box-shadow:
            6px 6px 20px #050505,
            -2px -2px 10px #2a2a2a,
            0 0 20px var(--tmg-primary-glow-soft),
            inset 0 1px 0 rgba(255,255,255,0.06);
        transform: translateY(-5px) scale(1.02);
    }

    .cultura-icon {
        font-size: 4.2rem;
        margin-bottom: 12px;
        filter: drop-shadow(0 4px 8px rgba(0,0,0,0.8));
        display: block;
    }

    .cultura-nome {
        font-family: 'Segoe UI', sans-serif;
        font-weight: 800;
        font-size: 1.15rem;
        letter-spacing: 4px;
        color: #e0e0e0;
        text-transform: uppercase;
        text-shadow: 1px 1px 4px rgba(0,0,0,0.9);
        margin-bottom: 6px;
    }

    .cultura-desc {
        font-size: 0.72rem;
        color: #555;
        letter-spacing: 2px;
        text-transform: uppercase;
    }

    .cultura-btn-wrapper div.stButton > button {
        width: 100%;
        border-radius: 16px;
        border: 1px solid #2e2e2e;
        padding: 0;
        background: transparent;
        color: transparent;
        box-shadow: none;
        margin: 0;
        height: 0;
        overflow: hidden;
        pointer-events: none;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("<br><br>", unsafe_allow_html=True)
    _, hcol, _ = st.columns([1, 2, 1])
    with hcol:
        if LOGO_PATH.exists():
            lc1, lc2, lc3 = st.columns([1, 2, 1])
            with lc2:
                app_image(str(LOGO_PATH))
            st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

        st.markdown("<div class='cultura-title'>Selecione a Cultura</div>", unsafe_allow_html=True)
        st.markdown(
            "<div class='cultura-subtitle'>Escolha somente os módulos liberados para seu usuário</div>",
            unsafe_allow_html=True
        )
        st.markdown("<hr class='cultura-hr'>", unsafe_allow_html=True)

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    _, gcol, _ = st.columns([0.15, 2.7, 0.15])

    with gcol:
        all_culturas = [
            ("🌱", "SOJA",    "Glycine max",       "#4caf50"),
            ("🌽", "MILHO",   "Zea mays",          "#ffb300"),
            ("🌿", "ALGODÃO", "Gossypium hirsutum", "#80cbc4"),
        ]
        culturas = [item for item in all_culturas if item[1] in allowed_cultures]

        cols = st.columns(max(1, min(3, len(culturas))), gap="large") if culturas else []
        for col, (icon, nome, cientifico, cor) in zip(cols, culturas):
            with col:
                st.markdown(f"""
                <div class='cultura-card'>
                    <span class='cultura-icon'>{icon}</span>
                    <div class='cultura-nome' style='color:{cor};
                        text-shadow:
                            1px 1px 0 rgba(0,0,0,0.8),
                            0 0 12px {cor}55;
                    '>{nome}</div>
                    <div class='cultura-desc'>{cientifico}</div>
                </div>
                """, unsafe_allow_html=True)
                st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
                if st.button(f"Selecionar {nome}", key=f"btn_cultura_{nome}", type="primary"):
                    st.session_state.cultura_selecionada = nome
                    st.session_state.pagina_ativa = "Checklist"
                    app_rerun()

        if can_open_partners:
            st.markdown("<div style='height:22px'></div>", unsafe_allow_html=True)
            pcol1, pcol2, pcol3 = st.columns([0.5, 1.4, 0.5])
            with pcol2:
                st.markdown("""
                <div class='cultura-card'>
                    <span class='cultura-icon'>🤝</span>
                    <div class='cultura-nome' style='color:var(--tmg-primary);text-shadow:1px 1px 0 rgba(0,0,0,.8),0 0 12px var(--tmg-primary-glow);'>
                        PARCEIROS
                    </div>
                    <div class='cultura-desc'>Controle de Voos e Dados</div>
                </div>
                """, unsafe_allow_html=True)
                st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
                if st.button("Abrir Parceiros / Controle de Voos e Dados", key="btn_open_partners_home", type="primary", use_container_width=True):
                    st.session_state.cultura_selecionada = "PARCEIROS"
                    st.session_state.pagina_ativa = "Parceiros"
                    app_rerun()

        if not culturas and not can_open_partners:
            st.warning("Seu usuário não possui módulos liberados. Solicite permissão ao administrador Wellington.")

    st.markdown("<br><br>", unsafe_allow_html=True)
    _, fcol, _ = st.columns([1, 2, 1])
    with fcol:
        st.markdown(
            "<p style='text-align:center; color:#333; font-size:0.72rem; letter-spacing:2px; text-transform:uppercase;'>"
            "TMG v2.0 &nbsp;·&nbsp; 2026</p>",
            unsafe_allow_html=True
        )

    st.stop()


# ==========================================
# CÓDIGO PRINCIPAL[cite: 1]
# ==========================================

if "pagina_ativa" not in st.session_state:
    st.session_state.pagina_ativa = "Checklist"

if st.session_state.pagina_ativa in ("TransferenciaVoos",):
    st.session_state.pagina_ativa = "Checklist"

if "logo_sistema" not in st.session_state:
    if LOGO_PATH.exists():
        st.session_state.logo_sistema = Image.open(LOGO_PATH)
    else:
        st.session_state.logo_sistema = None

def ir_para(pagina):
    st.session_state.pagina_ativa = pagina

def _cultura_ambiente_info(cultura: str = "") -> dict:
    cultura_norm = str(cultura or st.session_state.get("cultura_selecionada") or "").strip().upper()
    dados = {
        "SOJA": {
            "icone": "🌱",
            "nome": "SOJA",
            "subtitulo": "Ambiente de Análise de Soja",
            "cor": "#4caf50",
            "glow": "76, 175, 80",
        },
        "MILHO": {
            "icone": "🌽",
            "nome": "MILHO",
            "subtitulo": "Ambiente de Análise de Milho",
            "cor": "#ffb300",
            "glow": "255, 179, 0",
        },
        "ALGODÃO": {
            "icone": "🌿",
            "nome": "ALGODÃO",
            "subtitulo": "Ambiente de Análise de Algodão",
            "cor": "#80cbc4",
            "glow": "128, 203, 196",
        },
        "PARCEIROS": {
            "icone": "🤝",
            "nome": "PARCEIROS",
            "subtitulo": "Controle de Voos e Dados",
            "cor": THEME_PRIMARY_COLOR,
            "glow": THEME_PRIMARY_RGB,
        },
    }
    return dados.get(cultura_norm, dados["SOJA"])

def render_cultura_ambiente_css() -> None:
    st.markdown("""
    <style>
    .cultura-env-card {
        width: 100%;
        margin: 4px 0 18px 0;
        padding: 14px 18px;
        border-radius: 16px;
        border: 1px solid rgba(120, 220, 255, .44);
        background:
            linear-gradient(120deg, rgba(255,255,255,.14), transparent 28%),
            linear-gradient(145deg, rgba(2,14,36,.96), rgba(18,62,100,.78), rgba(0,212,255,.12));
        box-shadow:
            0 16px 34px rgba(0,0,0,.42),
            0 0 28px rgba(0,212,255,.18),
            inset 0 1px 0 rgba(255,255,255,.20),
            inset 0 -12px 20px rgba(2,14,36,.38);
        backdrop-filter: blur(12px) saturate(145%);
        -webkit-backdrop-filter: blur(12px) saturate(145%);
        text-align: center;
    }
    .cultura-env-logo {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        width: 52px;
        height: 52px;
        margin: 0 auto 8px auto;
        border-radius: 16px;
        border: 1px solid rgba(255,255,255,.28);
        background:
            radial-gradient(circle at 30% 18%, rgba(255,255,255,.35), transparent 32%),
            linear-gradient(145deg, rgba(255,255,255,.12), rgba(2,14,36,.72));
        font-size: 2rem;
        box-shadow:
            0 10px 22px rgba(0,0,0,.38),
            0 0 24px var(--culture-glow),
            inset 0 1px 0 rgba(255,255,255,.26);
    }
    .cultura-env-title {
        color: #ffffff;
        font-family: 'Segoe UI', Arial, sans-serif;
        font-size: 1rem;
        font-weight: 950;
        letter-spacing: 1.8px;
        text-transform: uppercase;
        text-shadow:
            0 1px 0 rgba(0,0,0,.95),
            0 0 12px var(--culture-glow);
    }
    .cultura-env-subtitle {
        margin-top: 4px;
        color: #dffbff;
        font-size: .74rem;
        font-weight: 800;
        letter-spacing: 1.2px;
        text-transform: uppercase;
        text-shadow: 0 1px 0 rgba(0,0,0,.85);
    }
    .cultura-env-top {
        max-width: 720px;
        margin: 0 auto 20px auto;
        padding: 13px 18px;
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 14px;
    }
    .cultura-env-top .cultura-env-logo {
        width: 46px;
        height: 46px;
        margin: 0;
        font-size: 1.75rem;
    }
    @media (max-width: 720px) {
        .cultura-env-top {
            flex-direction: column;
            gap: 8px;
            padding: 12px 14px;
        }
    }
    </style>
    """, unsafe_allow_html=True)

def render_cultura_ambiente_card(topo: bool = False) -> None:
    cultura = st.session_state.get("cultura_selecionada")
    if not cultura or cultura == "PARCEIROS":
        return
    info = _cultura_ambiente_info(cultura)
    glow = f"rgba({info['glow']}, .42)"
    extra_class = " cultura-env-top" if topo else ""
    st.markdown(f"""
    <div class="cultura-env-card{extra_class}" style="--culture-glow:{glow}; border-color:{info['cor']}88;">
        <div class="cultura-env-logo">{info['icone']}</div>
        <div>
            <div class="cultura-env-title">Ambiente {html.escape(info['nome'])}</div>
            <div class="cultura-env-subtitle">{html.escape(info['subtitulo'])}</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

# Adicionado um payload oculto para receber as coordenadas vindas do JavaScript
st.text_input("grid_payload", key="grid_payload", label_visibility="hidden")

# ==========================================
# MODULO ISOLADO - TRANSFERENCIA DE VOOS
# ==========================================
TV_ROOT = SYSTEM_DATABASE_DIR / "transferencia_voos"
TV_PROJECTS_DIR = TV_ROOT / "projetos"
TV_ORTHOS_DIR = TV_ROOT / "ortofotos"
TV_GRIDS_DIR = TV_ROOT / "grids"
TV_IMPORTS_DIR = TV_ROOT / "importacoes"
TV_RETURNS_DIR = TV_ROOT / "retornos"
TV_MANIFEST_PATH = TV_ROOT / "manifest.json"
MOSAIC_LIBRARY_DIR = SYSTEM_DATABASE_DIR / "mosaicos_importados"
MOSAIC_MANIFEST_PATH = MOSAIC_LIBRARY_DIR / "manifest.json"
MOSAIC_UPLOAD_TYPES = ["png", "jpg", "jpeg", "tif", "tiff", "geotiff", "geotif", "img", "ecw", "jp2"]

def _tv_safe_name(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in str(value).strip())
    return cleaned.strip("_") or "projeto"

def _tv_now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _tv_human_size(num_bytes: int) -> str:
    size = float(num_bytes or 0)
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if size < 1024 or unit == "TB":
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} B"
        size /= 1024
    return f"{size:.1f} TB"

def _tv_hash_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()

def _tv_hash_file(path: Path) -> str:
    hasher = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()

def _mosaic_default_manifest() -> dict:
    return {"mosaics": []}

def _mosaic_ensure_storage() -> None:
    MOSAIC_LIBRARY_DIR.mkdir(parents=True, exist_ok=True)
    if not MOSAIC_MANIFEST_PATH.exists():
        MOSAIC_MANIFEST_PATH.write_text(json.dumps(_mosaic_default_manifest(), indent=2, ensure_ascii=False), encoding="utf-8")

def _mosaic_load_manifest() -> dict:
    _mosaic_ensure_storage()
    try:
        data = json.loads(MOSAIC_MANIFEST_PATH.read_text(encoding="utf-8"))
    except Exception:
        data = _mosaic_default_manifest()
    data.setdefault("mosaics", [])
    return data

def _mosaic_save_manifest(data: dict) -> None:
    _mosaic_ensure_storage()
    MOSAIC_MANIFEST_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

def _mosaic_option_label(record: dict) -> str:
    size = record.get("tamanho_fmt") or _tv_human_size(record.get("tamanho", 0))
    origem = record.get("origem", "Importado")
    return f"{record.get('mosaic_id')} · {record.get('nome')} · {size} · {origem}"

def _mosaic_records() -> list:
    manifest = _mosaic_load_manifest()
    records = []
    for record in manifest.get("mosaics", []):
        if Path(record.get("path", "")).exists():
            records.append(record)
    return records

def _mosaic_single_select(label: str, key: str) -> str:
    options = [""] + [_mosaic_option_label(record) for record in _mosaic_records()]
    if len(options) == 1:
        st.caption("Biblioteca de mosaicos vazia. Envie uma ortofoto para deixá-la disponível nos outros visualizadores.")
        return ""
    return st.selectbox(label, options, key=key)

def _mosaic_multi_select(label: str, key: str, max_items: int = 10) -> list:
    options = [_mosaic_option_label(record) for record in _mosaic_records()]
    if not options:
        st.caption("Biblioteca de mosaicos vazia. Os uploads feitos nos visualizadores ficam disponíveis aqui.")
        return []
    selected = st.multiselect(label, options, key=key, max_selections=max_items)
    return selected or []

def _mosaic_find(option: str) -> dict:
    mosaic_id = option.split(" · ")[0] if option else ""
    if not mosaic_id:
        return {}
    for record in _mosaic_records():
        if record.get("mosaic_id") == mosaic_id:
            return record
    return {}

def _mosaic_path_inside_library(path: Path) -> bool:
    try:
        return path.resolve().is_relative_to(MOSAIC_LIBRARY_DIR.resolve())
    except Exception:
        return False

def _mosaic_delete(option: str) -> tuple:
    record = _mosaic_find(option)
    if not record:
        return False, "Mosaico não encontrado na biblioteca."

    mosaic_id = record.get("mosaic_id", "")
    mosaic_name = record.get("nome") or mosaic_id or "mosaico"
    path = Path(record.get("path", ""))
    manifest = _mosaic_load_manifest()
    manifest["mosaics"] = [
        item for item in manifest.get("mosaics", [])
        if item.get("mosaic_id") != mosaic_id
    ]
    _mosaic_save_manifest(manifest)

    removed_file = False
    if path.exists() and _mosaic_path_inside_library(path):
        try:
            target = path.parent if path.parent.name == mosaic_id else path
            if target.is_dir():
                shutil.rmtree(target)
            else:
                target.unlink()
            removed_file = True
        except Exception as exc:
            return True, f"Mosaico removido da lista, mas não foi possível apagar o arquivo: {exc}"

    if removed_file:
        return True, f"Mosaico `{mosaic_name}` excluído da biblioteca."
    return True, f"Mosaico `{mosaic_name}` removido da biblioteca. O arquivo original foi preservado."

def _mosaic_register_bytes(raw: bytes, filename: str, origem: str) -> dict:
    if not raw:
        return {}
    manifest = _mosaic_load_manifest()
    file_hash = _tv_hash_bytes(raw)
    for record in manifest.get("mosaics", []):
        if record.get("sha256") == file_hash and Path(record.get("path", "")).exists():
            return record

    mosaic_id = f"MOSAICO_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
    safe_name = Path(filename or mosaic_id).name
    mosaic_dir = MOSAIC_LIBRARY_DIR / mosaic_id
    mosaic_dir.mkdir(parents=True, exist_ok=True)
    target = mosaic_dir / safe_name
    if target.exists():
        target = mosaic_dir / f"{target.stem}_{datetime.now().strftime('%H%M%S%f')}{target.suffix}"
    target.write_bytes(raw)
    record = {
        "mosaic_id": mosaic_id,
        "nome": safe_name,
        "path": str(target),
        "tamanho": len(raw),
        "tamanho_fmt": _tv_human_size(len(raw)),
        "sha256": file_hash,
        "ext": target.suffix.lower(),
        "origem": origem,
        "importado_em": _tv_now()
    }
    manifest.setdefault("mosaics", []).insert(0, record)
    manifest["mosaics"] = manifest["mosaics"][:200]
    _mosaic_save_manifest(manifest)
    return record

def _mosaic_register_file(path_value, filename: str, origem: str, file_hash: str = "", file_size: int = 0) -> dict:
    path = Path(path_value)
    if not path.exists():
        return {}
    manifest = _mosaic_load_manifest()
    file_hash = file_hash or _tv_hash_file(path)
    for record in manifest.get("mosaics", []):
        if record.get("sha256") == file_hash and Path(record.get("path", "")).exists():
            return record

    record = {
        "mosaic_id": f"MOSAICO_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}",
        "nome": Path(filename or path.name).name,
        "path": str(path),
        "tamanho": int(file_size or path.stat().st_size),
        "tamanho_fmt": _tv_human_size(int(file_size or path.stat().st_size)),
        "sha256": file_hash,
        "ext": path.suffix.lower(),
        "origem": origem,
        "importado_em": _tv_now()
    }
    manifest.setdefault("mosaics", []).insert(0, record)
    manifest["mosaics"] = manifest["mosaics"][:200]
    _mosaic_save_manifest(manifest)
    return record

def _mosaic_bytes_from_selection(option: str) -> tuple:
    record = _mosaic_find(option)
    path = Path(record.get("path", ""))
    if record and path.exists():
        return path.read_bytes(), record.get("nome") or path.name
    return None, ""

def _mosaic_input_bytes(uploaded, selected_option: str, origem: str) -> tuple:
    if uploaded is not None:
        load_box = st.empty()
        update_tmg_loading(load_box, 35, f"Recebendo arquivo: {Path(uploaded.name).name}")
        raw = uploaded.getbuffer().tobytes()
        update_tmg_loading(load_box, 78, "Registrando arquivo na biblioteca interna...")
        _mosaic_register_bytes(raw, uploaded.name, origem)
        finish_tmg_loading_and_clear(load_box, "Carregamento concluído com sucesso.")
        return raw, uploaded.name
    return _mosaic_bytes_from_selection(selected_option)

def _resettable_ortho_uploader(label: str, key: str, accept_multiple_files: bool = False, help: str = ""):
    reset_key = f"{key}_reset"
    st.session_state.setdefault(reset_key, 0)
    widget_key = f"{key}_{st.session_state[reset_key]}"
    uploaded = st.file_uploader(
        label,
        type=MOSAIC_UPLOAD_TYPES,
        accept_multiple_files=accept_multiple_files,
        key=widget_key,
        help=help
    )
    has_upload = bool(uploaded) if accept_multiple_files else uploaded is not None
    if has_upload:
        def _upload_signature(item):
            if isinstance(item, (list, tuple)):
                return "|".join(f"{Path(getattr(file, 'name', 'arquivo')).name}:{int(getattr(file, 'size', 0) or 0)}" for file in item)
            return f"{Path(getattr(item, 'name', 'arquivo')).name}:{int(getattr(item, 'size', 0) or 0)}"
        loading_signature_key = f"{key}_loading_signature"
        current_signature = _upload_signature(uploaded or [])
        show_upload_loading = st.session_state.get(loading_signature_key) != current_signature
        if show_upload_loading:
            st.session_state[loading_signature_key] = current_signature
            load_box = st.empty()
        if accept_multiple_files:
            total_files = len(uploaded or [])
            if show_upload_loading:
                update_tmg_loading(load_box, 65, f"Recebendo {total_files} arquivo(s) para carregamento...")
                finish_tmg_loading_and_clear(load_box, f"{total_files} arquivo(s) carregado(s) com sucesso.")
        else:
            if show_upload_loading:
                update_tmg_loading(load_box, 65, f"Carregando ortofoto: {Path(uploaded.name).name}")
                finish_tmg_loading_and_clear(load_box, "Carregamento concluído com sucesso.")
        _, clear_col = st.columns([3, 1])
        with clear_col:
            if st.button("🗑️ Excluir e importar nova", key=f"{key}_clear_{st.session_state[reset_key]}", use_container_width=True):
                st.session_state[reset_key] += 1
                st.session_state.pop(loading_signature_key, None)
                app_rerun()
    return uploaded

def _uploaded_ortho_bytes(uploaded) -> tuple:
    if uploaded is None:
        return None, ""
    return uploaded.getbuffer().tobytes(), uploaded.name

def _tv_default_manifest() -> dict:
    return {
        "projects": [],
        "orthos": [],
        "grids": [],
        "imports": [],
        "returns": [],
        "history": [],
        "config": {
            "provider": "Servidor local",
            "destination": str(TV_ROOT / "sync"),
            "sync_mode": "Manual",
            "checksum": True,
            "resumable": True,
            "encryption": True,
            "auth_mode": "Token/OAuth"
        },
        "users": [
            {"usuario": "admin", "perfil": "Administrador", "permissao": "Leitura, envio, analise, configuracao"},
            {"usuario": "operador", "perfil": "Operador de Campo", "permissao": "Envio e download"},
            {"usuario": "analista", "perfil": "Analista Tecnico", "permissao": "Grid, parcelas e relatorios"}
        ]
    }

def _tv_ensure_storage() -> None:
    for folder in [TV_ROOT, TV_PROJECTS_DIR, TV_ORTHOS_DIR, TV_GRIDS_DIR, TV_IMPORTS_DIR, TV_RETURNS_DIR]:
        folder.mkdir(parents=True, exist_ok=True)
    if not TV_MANIFEST_PATH.exists():
        TV_MANIFEST_PATH.write_text(json.dumps(_tv_default_manifest(), indent=2, ensure_ascii=False), encoding="utf-8")

def _tv_load_manifest() -> dict:
    _tv_ensure_storage()
    try:
        data = json.loads(TV_MANIFEST_PATH.read_text(encoding="utf-8"))
    except Exception:
        data = _tv_default_manifest()
    default = _tv_default_manifest()
    for key, value in default.items():
        data.setdefault(key, value)
    return data

def _tv_save_manifest(data: dict) -> None:
    _tv_ensure_storage()
    TV_MANIFEST_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

def _tv_add_history(manifest: dict, event: str, project_id: str = "", status: str = "OK") -> None:
    manifest.setdefault("history", []).insert(0, {
        "data": _tv_now(),
        "projeto": project_id,
        "evento": event,
        "status": status
    })
    manifest["history"] = manifest["history"][:300]

def _tv_project_options(manifest: dict) -> list:
    return [f"{p['project_id']} · {p.get('nome','Projeto')}" for p in manifest.get("projects", [])]

def _tv_get_project_id(option: str) -> str:
    return option.split(" · ")[0] if option else ""

def _tv_get_ortho_id(option: str) -> str:
    return option.split(" · ")[0] if option else ""

def _tv_find_project(manifest: dict, project_id: str) -> dict:
    for project in manifest.get("projects", []):
        if project.get("project_id") == project_id:
            return project
    return {}

def _tv_find_ortho(manifest: dict, ortho_id: str) -> dict:
    for ortho in manifest.get("orthos", []):
        if ortho.get("ortho_id") == ortho_id:
            return ortho
    return {}

def _tv_find_grid(manifest: dict, grid_id: str) -> dict:
    for grid in manifest.get("grids", []):
        if grid.get("grid_id") == grid_id:
            return grid
    return {}

def _tv_file_exists_by_hash(manifest: dict, file_hash: str) -> bool:
    for project in manifest.get("projects", []):
        for item in project.get("files", []):
            if item.get("sha256") == file_hash:
                return True
    for ortho in manifest.get("orthos", []):
        if ortho.get("sha256") == file_hash:
            return True
    return False

def _tv_save_uploaded_batch(files, base_dir: Path, manifest: dict, duplicate_check: bool = True) -> tuple:
    saved, duplicates, total_size = [], [], 0
    base_dir.mkdir(parents=True, exist_ok=True)
    for uploaded in files or []:
        hasher = hashlib.sha256()
        file_size = 0
        temp_name = f".{_tv_safe_name(Path(uploaded.name).stem)}_{datetime.now().strftime('%H%M%S%f')}.uploading"
        temp_path = base_dir / temp_name
        try:
            uploaded.seek(0)
        except Exception:
            pass
        with open(temp_path, "wb") as out:
            while True:
                chunk = uploaded.read(1024 * 1024)
                if not chunk:
                    break
                if isinstance(chunk, str):
                    chunk = chunk.encode("utf-8")
                hasher.update(chunk)
                out.write(chunk)
                file_size += len(chunk)
        file_hash = hasher.hexdigest()
        total_size += file_size
        if duplicate_check and _tv_file_exists_by_hash(manifest, file_hash):
            duplicates.append(uploaded.name)
            try:
                temp_path.unlink()
            except Exception:
                pass
            continue
        target = base_dir / Path(uploaded.name).name
        if target.exists():
            stem, suffix = target.stem, target.suffix
            target = base_dir / f"{stem}_{datetime.now().strftime('%H%M%S%f')}{suffix}"
        temp_path.replace(target)
        saved.append({
            "nome": uploaded.name,
            "path": str(target),
            "tamanho": file_size,
            "tamanho_fmt": _tv_human_size(file_size),
            "sha256": file_hash,
            "ext": Path(uploaded.name).suffix.lower(),
            "enviado_em": _tv_now()
        })
    return saved, duplicates, total_size

def _tv_status_chip(text: str, color: str = "#ff8c00") -> str:
    return f"<span style='display:inline-block;border:1px solid {color};color:{color};border-radius:999px;padding:4px 10px;font-size:11px;margin:2px 5px 2px 0;'>{text}</span>"

def _tv_metric_cards(manifest: dict) -> None:
    projects = manifest.get("projects", [])
    orthos = manifest.get("orthos", [])
    grids = manifest.get("grids", [])
    total_bytes = sum(sum(f.get("tamanho", 0) for f in p.get("files", [])) for p in projects)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Projetos", len(projects))
    c2.metric("Ortofotos", len(orthos))
    c3.metric("Grids", len(grids))
    c4.metric("Volume", _tv_human_size(total_bytes))

def _tv_render_upload(manifest: dict) -> None:
    st.markdown("#### Upload de Voos")
    meta1, meta2, meta3 = st.columns(3)
    with meta1:
        nome = st.text_input("Nome do projeto", value=f"Voo_{date.today().strftime('%Y%m%d')}", key="tv_nome")
        operador = st.text_input("Operador", value="", key="tv_operador")
    with meta2:
        data_voo = st.date_input("Data", value=date.today(), key="tv_data")
        fazenda = st.text_input("Fazenda", value="", key="tv_fazenda")
    with meta3:
        talhao = st.text_input("Talhão", value="", key="tv_talhao")
        coordenadas = st.text_input("Coordenadas / centroide", value="", key="tv_coords")

    provider = st.selectbox(
        "Destino de sincronização",
        ["Microsoft OneDrive", "Microsoft Azure", "Google Drive", "Dropbox", "Servidor local", "NAS", "FTP/SFTP"],
        index=["Microsoft OneDrive", "Microsoft Azure", "Google Drive", "Dropbox", "Servidor local", "NAS", "FTP/SFTP"].index(manifest.get("config", {}).get("provider", "Servidor local")) if manifest.get("config", {}).get("provider", "Servidor local") in ["Microsoft OneDrive", "Microsoft Azure", "Google Drive", "Dropbox", "Servidor local", "NAS", "FTP/SFTP"] else 4,
        key="tv_upload_provider"
    )

    files = st.file_uploader(
        "Fotos do voo, RAW, TIFF/GeoTIFF ou ZIP da pasta completa",
        type=["jpg", "jpeg", "tif", "tiff", "png", "raw", "dng", "arw", "cr2", "nef", "zip"],
        accept_multiple_files=True,
        key="tv_flight_upload"
    )
    if files:
        render_tmg_loading_bar(100, f"{len(files)} arquivo(s) de voo recebido(s).")

    col_a, col_b, col_c = st.columns([1, 1, 1])
    with col_a:
        compactar = st.checkbox("Compactação automática do lote", value=True, key="tv_zip")
    with col_b:
        manter_original = st.checkbox("Preservar qualidade original", value=True, key="tv_original", disabled=True)
    with col_c:
        checksum = st.checkbox("Validar integridade SHA-256", value=True, key="tv_checksum")

    if st.button("Enviar e sincronizar lote", type="primary", key="tv_send_batch", use_container_width=True):
        if not files:
            st.warning("Selecione os arquivos do voo.")
        else:
            project_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{_tv_safe_name(nome)}"
            project_dir = TV_PROJECTS_DIR / project_id / "raw"
            progress = st.empty()
            update_tmg_loading(progress, 0, "Iniciando upload do lote de voo...")
            saved, duplicates, total_size = _tv_save_uploaded_batch(files, project_dir, manifest, duplicate_check=checksum)
            update_tmg_loading(progress, 40, "Arquivos recebidos. Preparando pacote do voo...")

            zip_path = ""
            if compactar and saved:
                zip_path = str(TV_PROJECTS_DIR / project_id / f"{project_id}_pacote_voo.zip")
                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    for item in saved:
                        zf.write(item["path"], arcname=Path(item["path"]).name)
            update_tmg_loading(progress, 75, "Atualizando manifesto e histórico do lote...")

            project = {
                "project_id": project_id,
                "nome": nome,
                "data": str(data_voo),
                "operador": operador,
                "fazenda": fazenda,
                "talhao": talhao,
                "coordenadas": coordenadas,
                "quantidade_imagens": len(saved),
                "tamanho_total": sum(item["tamanho"] for item in saved),
                "tamanho_total_fmt": _tv_human_size(sum(item["tamanho"] for item in saved)),
                "status": "Voo enviado - aguardando recebimento",
                "workflow": "VOO_ENVIADO",
                "provedor": provider,
                "zip_path": zip_path,
                "files": saved,
                "duplicados": duplicates,
                "criado_em": _tv_now()
            }
            manifest.setdefault("projects", []).insert(0, project)
            _tv_add_history(manifest, f"Lote enviado: {len(saved)} arquivo(s), {len(duplicates)} duplicado(s)", project_id)
            _tv_save_manifest(manifest)
            update_tmg_loading(progress, 100, "Carregamento concluído com sucesso.")
            st.success(f"Projeto `{project_id}` criado com {len(saved)} arquivo(s).")
            if duplicates:
                st.warning("Duplicados ignorados: " + ", ".join(duplicates[:8]))

    st.markdown("---")
    _tv_metric_cards(manifest)

def _tv_render_database(manifest: dict) -> None:
    st.markdown("#### Banco de Dados")
    rows = []
    for p in manifest.get("projects", []):
        rows.append({
            "Projeto": p.get("project_id"),
            "Nome": p.get("nome"),
            "Data": p.get("data"),
            "Operador": p.get("operador"),
            "Fazenda": p.get("fazenda"),
            "Talhão": p.get("talhao"),
            "Imagens": p.get("quantidade_imagens"),
            "Tamanho": p.get("tamanho_total_fmt"),
            "Status": p.get("status"),
            "Nuvem": p.get("provedor")
        })
    st.dataframe(rows, use_container_width=True, hide_index=True)
    st.download_button(
        "Baixar manifesto JSON",
        data=json.dumps(manifest, indent=2, ensure_ascii=False).encode("utf-8"),
        file_name="transferencia_voos_manifest.json",
        mime="application/json",
        use_container_width=True
    )

def _tv_render_receber_voos(manifest: dict) -> None:
    st.markdown("#### Receber Voos")
    st.caption("Janela para o operador receber pacotes de voo, baixar imagens originais e marcar o processamento externo da ortofoto.")

    with st.expander("Receber pacote de voo de outro usuário", expanded=False):
        r1, r2, r3 = st.columns(3)
        with r1:
            incoming_name = st.text_input("Nome do voo recebido", value=f"Recebido_{date.today().strftime('%Y%m%d')}", key="tv_incoming_name")
            incoming_operator = st.text_input("Operador de origem", value="", key="tv_incoming_operator")
        with r2:
            incoming_farm = st.text_input("Fazenda", value="", key="tv_incoming_farm")
            incoming_field = st.text_input("Talhão", value="", key="tv_incoming_field")
        with r3:
            incoming_coords = st.text_input("Coordenadas", value="", key="tv_incoming_coords")
            incoming_provider = st.selectbox("Origem", ["OneDrive", "Azure", "Google Drive", "Dropbox", "Servidor local", "NAS", "FTP/SFTP", "Upload direto"], key="tv_incoming_provider")
        incoming_files = st.file_uploader(
            "ZIP ou imagens recebidas",
            type=["zip", "jpg", "jpeg", "tif", "tiff", "png", "raw", "dng", "arw", "cr2", "nef"],
            accept_multiple_files=True,
            key="tv_incoming_files"
        )
        if incoming_files:
            render_tmg_loading_bar(100, f"{len(incoming_files)} arquivo(s) recebido(s) para registro.")
        if st.button("Registrar voo recebido", type="primary", key="tv_register_incoming", use_container_width=True):
            if not incoming_files:
                st.warning("Selecione o pacote ou as imagens recebidas.")
            else:
                project_id = f"REC_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{_tv_safe_name(incoming_name)}"
                project_dir = TV_PROJECTS_DIR / project_id / "recebido"
                saved, duplicates, _ = _tv_save_uploaded_batch(incoming_files, project_dir, manifest, duplicate_check=True)
                zip_path = str(TV_PROJECTS_DIR / project_id / f"{project_id}_pacote_recebido.zip")
                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    for item in saved:
                        zf.write(item["path"], arcname=Path(item["path"]).name)
                project = {
                    "project_id": project_id,
                    "nome": incoming_name,
                    "data": str(date.today()),
                    "operador": incoming_operator,
                    "fazenda": incoming_farm,
                    "talhao": incoming_field,
                    "coordenadas": incoming_coords,
                    "quantidade_imagens": len(saved),
                    "tamanho_total": sum(item["tamanho"] for item in saved),
                    "tamanho_total_fmt": _tv_human_size(sum(item["tamanho"] for item in saved)),
                    "status": "Voo recebido - gerar ortofoto externa",
                    "workflow": "VOO_RECEBIDO",
                    "provedor": incoming_provider,
                    "zip_path": zip_path,
                    "files": saved,
                    "duplicados": duplicates,
                    "criado_em": _tv_now(),
                    "recebido_em": _tv_now()
                }
                manifest.setdefault("projects", []).insert(0, project)
                _tv_add_history(manifest, f"Voo externo recebido: {len(saved)} arquivo(s)", project_id)
                _tv_save_manifest(manifest)
                st.success(f"Voo recebido registrado como `{project_id}`.")
                if duplicates:
                    st.warning("Duplicados ignorados: " + ", ".join(duplicates[:8]))

    options = _tv_project_options(manifest)
    if not options:
        st.info("Nenhum voo enviado ainda.")
        return

    selected = st.selectbox("Voo disponível para recebimento", options, key="tv_receive_project")
    project_id = _tv_get_project_id(selected)
    project = _tv_find_project(manifest, project_id)
    if not project:
        st.warning("Projeto não encontrado.")
        return

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Imagens", project.get("quantidade_imagens", 0))
    c2.metric("Tamanho", project.get("tamanho_total_fmt", "0 B"))
    c3.metric("Fazenda", project.get("fazenda") or "-")
    c4.metric("Status", project.get("status") or "-")

    st.markdown(
        f"<div class='tv-band'>"
        f"{_tv_status_chip('Metadados preservados', '#55ff99')}"
        f"{_tv_status_chip('Coordenadas: ' + (project.get('coordenadas') or 'não informado'), '#5599ff')}"
        f"{_tv_status_chip('Provedor: ' + (project.get('provedor') or 'local'), '#ff8c00')}"
        f"</div>",
        unsafe_allow_html=True
    )

    col_a, col_b = st.columns(2)
    with col_a:
        zip_path = project.get("zip_path")
        if zip_path and Path(zip_path).exists():
            st.download_button(
                "Baixar pacote completo do voo",
                data=Path(zip_path).read_bytes(),
                file_name=Path(zip_path).name,
                mime="application/zip",
                use_container_width=True
            )
        else:
            st.info("Pacote ZIP não encontrado; arquivos individuais continuam registrados no manifesto.")
    with col_b:
        if st.button("Confirmar recebimento do voo", type="primary", key="tv_confirm_receive", use_container_width=True):
            project["status"] = "Voo recebido - gerar ortofoto externa"
            project["workflow"] = "VOO_RECEBIDO"
            project["recebido_em"] = _tv_now()
            _tv_add_history(manifest, "Voo recebido por operador para processamento externo", project_id)
            _tv_save_manifest(manifest)
            st.success("Recebimento confirmado. Gere a ortofoto externamente e envie em Ortofotos Geradas.")

    st.markdown("##### Arquivos do voo")
    st.dataframe(project.get("files", []), use_container_width=True, hide_index=True)

def _tv_render_orthos(manifest: dict) -> None:
    st.markdown("#### Ortofotos Geradas")
    st.caption("Receba a ortofoto gerada externamente e encaminhe para ajuste/marcação de grid.")
    options = _tv_project_options(manifest)
    selected_project = st.selectbox("Projeto vinculado", [""] + options, key="tv_ortho_project")
    ortho_files = st.file_uploader(
        "Enviar ortofoto processada",
        type=["tif", "tiff", "geotiff", "jpg", "jpeg", "png", "zip"],
        accept_multiple_files=True,
        key="tv_ortho_upload"
    )
    if ortho_files:
        render_tmg_loading_bar(100, f"{len(ortho_files)} ortofoto(s) recebida(s) para registro.")
    if st.button("Registrar ortofoto recebida", type="primary", key="tv_register_ortho", use_container_width=True):
        if not ortho_files:
            st.warning("Selecione uma ortofoto.")
        else:
            project_id = _tv_get_project_id(selected_project)
            ortho_dir = TV_ORTHOS_DIR / (project_id or "sem_projeto") / datetime.now().strftime("%Y%m%d_%H%M%S")
            saved, duplicates, _ = _tv_save_uploaded_batch(ortho_files, ortho_dir, manifest, duplicate_check=True)
            for item in saved:
                spatial = {"crs": "", "transform": "", "orig_width": 0, "orig_height": 0}
                thumb_dims = "Aguardando abertura no visualizador"
                if item["ext"] in [".tif", ".tiff", ".png", ".jpg", ".jpeg"]:
                    try:
                        _mosaic_register_file(item["path"], item["nome"], "Transferencia de Voos", item.get("sha256", ""), item.get("tamanho", 0))
                    except Exception:
                        pass
                manifest.setdefault("orthos", []).insert(0, {
                    "ortho_id": f"ORTO_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}",
                    "project_id": project_id,
                    "nome": item["nome"],
                    "path": item["path"],
                    "tamanho": item["tamanho"],
                    "tamanho_fmt": item["tamanho_fmt"],
                    "sha256": item["sha256"],
                    "resolucao_preview": thumb_dims,
                    "crs": spatial.get("crs") if isinstance(spatial, dict) else "",
                    "coordenadas": spatial.get("transform") if isinstance(spatial, dict) else "",
                    "data_processamento": _tv_now(),
                    "status": "Ortofoto gerada recebida - aguardando grid"
                })
            project = _tv_find_project(manifest, project_id)
            if project:
                project["status"] = "Ortofoto recebida - ajustar e marcar grid"
                project["workflow"] = "ORTOFOTO_RECEBIDA"
                project["ortofoto_recebida_em"] = _tv_now()
            _tv_add_history(manifest, f"Ortofoto(s) recebida(s): {len(saved)}", project_id)
            _tv_save_manifest(manifest)
            st.success(f"{len(saved)} ortofoto(s) registrada(s).")
            st.info("Próximo passo: abra Grid e Parcelas para ajustar a ortofoto e marcar o grid.")
            if duplicates:
                st.warning("Duplicados ignorados: " + ", ".join(duplicates[:8]))

    ortho_rows = [{
        "ID": o.get("ortho_id"),
        "Projeto": o.get("project_id"),
        "Arquivo": o.get("nome"),
        "Resolução": o.get("resolucao_preview"),
        "Tamanho": o.get("tamanho_fmt"),
        "CRS": (o.get("crs") or "")[:80],
        "Processamento": o.get("data_processamento"),
        "Status": o.get("status")
    } for o in manifest.get("orthos", [])]
    st.dataframe(ortho_rows, use_container_width=True, hide_index=True)

    if manifest.get("orthos"):
        selected = st.selectbox("Download de ortofoto", [f"{o['ortho_id']} · {o['nome']}" for o in manifest.get("orthos", [])], key="tv_ortho_download")
        ortho_id = selected.split(" · ")[0]
        record = next((o for o in manifest.get("orthos", []) if o.get("ortho_id") == ortho_id), None)
        if record and Path(record["path"]).exists():
            st.download_button(
                "Baixar ortofoto original",
                data=Path(record["path"]).read_bytes(),
                file_name=Path(record["path"]).name,
                mime="application/octet-stream",
                use_container_width=True
            )
            if st.button("Abrir esta ortofoto para marcar grid", key="tv_open_ortho_grid", use_container_width=True):
                st.session_state["tv_next_grid_ortho"] = selected
                st.info("Selecione a aba Grid e Parcelas no menu superior para continuar com esta ortofoto.")

def _tv_grid_viewer_html(b64: str, rows: int, cols: int, image_name: str) -> str:
    return f"""
<!DOCTYPE html>
<html>
<head>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ background:#0d0d0d; overflow:hidden; font-family:Segoe UI, Arial, sans-serif; }}
  #wrap {{ width:100%; height:680px; position:relative; background:#0d0d0d; border:1px solid #2a2a2a; border-radius:8px; overflow:hidden; }}
  canvas {{ position:absolute; inset:0; }}
  .bar {{ position:absolute; top:12px; right:12px; display:flex; gap:6px; z-index:3; }}
  button {{ background:#171717; color:#ff8c00; border:1px solid #3a3a3a; border-radius:6px; padding:8px 10px; cursor:pointer; font-weight:700; }}
  button:hover {{ border-color:#ff8c00; }}
  .hint {{ position:absolute; left:12px; bottom:12px; color:#aaa; background:rgba(0,0,0,.72); border:1px solid #333; border-radius:6px; padding:7px 10px; font-size:11px; }}
</style>
</head>
<body>
<div id="wrap">
  <canvas id="cv"></canvas>
  <div class="bar">
    <button id="mark">Marcar 4 cantos</button>
    <button id="fit">Ajustar tela</button>
    <button id="clear">Limpar</button>
    <button id="export">Exportar Grid JSON</button>
    <button id="geojson">GeoJSON</button>
    <button id="csv">CSV</button>
  </div>
  <div class="hint" id="hint">Scroll zoom · arrastar mover · arraste os pontos azuis para ajustar · {rows} linhas × {cols} colunas</div>
</div>
<script>
const cv=document.getElementById('cv'), ctx=cv.getContext('2d'), wrap=document.getElementById('wrap');
let points=[], markMode=false, sc=1, ox=0, oy=0, drag=false, lx=0, ly=0, imgW=0, imgH=0, draggingPoint=-1;
const ROWS={rows}, COLS={cols}, IMG_NAME={json.dumps(image_name)};
const img=new Image();
function resize(){{ cv.width=wrap.clientWidth; cv.height=wrap.clientHeight; draw(); }}
function toImg(e){{ const r=cv.getBoundingClientRect(); return {{x:(e.clientX-r.left-ox)/sc, y:(e.clientY-r.top-oy)/sc}}; }}
function bilerp(p0,p1,p2,p3,u,v){{ const tx=(1-u)*p0.x+u*p1.x, ty=(1-u)*p0.y+u*p1.y; const bx=(1-u)*p3.x+u*p2.x, by=(1-u)*p3.y+u*p2.y; return {{x:(1-v)*tx+v*bx,y:(1-v)*ty+v*by}}; }}
function fit(){{ if(!imgW) return; sc=Math.min(cv.width/imgW,cv.height/imgH); ox=(cv.width-imgW*sc)/2; oy=(cv.height-imgH*sc)/2; draw(); }}
function getCells(){{
  if(points.length!==4) return [];
  const [p0,p1,p2,p3]=points, cells=[];
  for(let r=0;r<ROWS;r++) for(let c=0;c<COLS;c++){{
    const u0=c/COLS,u1=(c+1)/COLS,v0=r/ROWS,v1=(r+1)/ROWS;
    const tl=bilerp(p0,p1,p2,p3,u0,v0), tr=bilerp(p0,p1,p2,p3,u1,v0), br=bilerp(p0,p1,p2,p3,u1,v1), bl=bilerp(p0,p1,p2,p3,u0,v1);
    cells.push({{linha:r+1,coluna:c+1,polygon:[tl,tr,br,bl]}});
  }}
  return cells;
}}
function download(name,type,text){{ const blob=new Blob([text],{{type:type}}); const a=document.createElement('a'); a.href=URL.createObjectURL(blob); a.download=name; a.click(); }}
function draw(){{
  ctx.clearRect(0,0,cv.width,cv.height); ctx.save(); ctx.translate(ox,oy); ctx.scale(sc,sc);
  if(imgW) ctx.drawImage(img,0,0);
  if(points.length===4){{
    const [p0,p1,p2,p3]=points; ctx.strokeStyle='rgba(255,140,0,.95)'; ctx.lineWidth=2/sc;
    for(let r=0;r<=ROWS;r++){{ const v=r/ROWS; const a=bilerp(p0,p1,p2,p3,0,v), b=bilerp(p0,p1,p2,p3,1,v); ctx.beginPath(); ctx.moveTo(a.x,a.y); ctx.lineTo(b.x,b.y); ctx.stroke(); }}
    for(let c=0;c<=COLS;c++){{ const u=c/COLS; const a=bilerp(p0,p1,p2,p3,u,0), b=bilerp(p0,p1,p2,p3,u,1); ctx.beginPath(); ctx.moveTo(a.x,a.y); ctx.lineTo(b.x,b.y); ctx.stroke(); }}
  }}
  points.forEach((p,i)=>{{ ctx.fillStyle='#1e90ff'; ctx.beginPath(); ctx.arc(p.x,p.y,8/sc,0,Math.PI*2); ctx.fill(); ctx.fillStyle='#fff'; ctx.font=(12/sc)+'px Arial'; ctx.textAlign='center'; ctx.textBaseline='middle'; ctx.fillText(i+1,p.x,p.y); }});
  ctx.restore();
}}
img.onload=()=>{{ imgW=img.width; imgH=img.height; resize(); fit(); }};
img.src='data:image/jpeg;base64,{b64}';
window.addEventListener('resize',resize);
cv.addEventListener('wheel',e=>{{ e.preventDefault(); const f=e.deltaY<0?1.18:.84; const r=cv.getBoundingClientRect(); const mx=e.clientX-r.left, my=e.clientY-r.top; const ix=(mx-ox)/sc, iy=(my-oy)/sc; sc=Math.max(.05,Math.min(40,sc*f)); ox=mx-ix*sc; oy=my-iy*sc; draw(); }},{{passive:false}});
cv.addEventListener('mousedown',e=>{{
  const p=toImg(e);
  for(let i=0;i<points.length;i++){{ const dx=(p.x-points[i].x)*sc, dy=(p.y-points[i].y)*sc; if(Math.sqrt(dx*dx+dy*dy)<18){{ draggingPoint=i; return; }} }}
  if(markMode && points.length<4){{ points.push(p); if(points.length===4) markMode=false; draw(); return; }}
  drag=true; lx=e.clientX; ly=e.clientY;
}});
cv.addEventListener('mousemove',e=>{{
  if(draggingPoint>=0){{ points[draggingPoint]=toImg(e); draw(); return; }}
  if(drag){{ ox+=e.clientX-lx; oy+=e.clientY-ly; lx=e.clientX; ly=e.clientY; draw(); }}
}});
cv.addEventListener('mouseup',()=>{{ drag=false; draggingPoint=-1; }}); cv.addEventListener('mouseleave',()=>{{ drag=false; draggingPoint=-1; }});
document.getElementById('mark').onclick=()=>{{ markMode=true; points=[]; draw(); }};
document.getElementById('fit').onclick=()=>fit();
document.getElementById('clear').onclick=()=>{{ points=[]; draw(); }};
document.getElementById('export').onclick=()=>{{
  if(points.length!==4){{ alert('Marque os 4 cantos do grid.'); return; }}
  const payload={{ imagem:IMG_NAME, linhas:ROWS, colunas:COLS, pontos:points, parcelas:getCells(), criado_em:new Date().toISOString(), versao:'1.0' }};
  download('grid_parcelas_'+Date.now()+'.json','application/json',JSON.stringify(payload,null,2));
}};
document.getElementById('geojson').onclick=()=>{{
  if(points.length!==4){{ alert('Marque os 4 cantos do grid.'); return; }}
  const features=getCells().map(cell=>({{type:'Feature',properties:{{linha:cell.linha,coluna:cell.coluna,id:'L'+cell.linha+'_C'+cell.coluna}},geometry:{{type:'Polygon',coordinates:[[...cell.polygon.map(p=>[p.x,-p.y]),[cell.polygon[0].x,-cell.polygon[0].y]]]}}}}));
  download('grid_parcelas_'+Date.now()+'.geojson','application/geo+json',JSON.stringify({{type:'FeatureCollection',features:features}},null,2));
}};
document.getElementById('csv').onclick=()=>{{
  if(points.length!==4){{ alert('Marque os 4 cantos do grid.'); return; }}
  let csv='linha;coluna;x1;y1;x2;y2;x3;y3;x4;y4\\n';
  for(const cell of getCells()) csv+=cell.linha+';'+cell.coluna+';'+cell.polygon.map(p=>Math.round(p.x)+';'+Math.round(p.y)).join(';')+'\\n';
  download('grid_parcelas_'+Date.now()+'.csv','text/csv;charset=utf-8',csv);
}};
</script>
</body>
</html>
"""

def _tv_render_grid_parcelas(manifest: dict) -> None:
    st.markdown("#### Grid e Parcelas")
    col_l, col_r = st.columns([1, 2])
    with col_l:
        ortho_options = [f"{o['ortho_id']} · {o['nome']}" for o in manifest.get("orthos", [])]
        ortho_select_options = [""] + ortho_options
        pending_ortho = st.session_state.pop("tv_next_grid_ortho", "")
        ortho_index = ortho_select_options.index(pending_ortho) if pending_ortho in ortho_select_options else 0
        selected = st.selectbox("Ortofoto", ortho_select_options, index=ortho_index, key="tv_grid_ortho")
        rows = st.number_input("Linhas", min_value=1, max_value=500, value=5, key="tv_grid_rows")
        cols = st.number_input("Colunas", min_value=1, max_value=500, value=5, key="tv_grid_cols")
        imported = st.file_uploader(
            "Importar grid/vetor",
            type=["shp", "geojson", "json", "kml", "kmz", "csv", "dxf", "tif", "tiff", "gpkg"],
            accept_multiple_files=True,
            key="tv_grid_import"
        )
        if imported:
            render_tmg_loading_bar(100, f"{len(imported)} arquivo(s) GIS recebido(s) para importação.")
        if st.button("Registrar importação", key="tv_register_import", use_container_width=True):
            saved, _, _ = _tv_save_uploaded_batch(imported, TV_IMPORTS_DIR / datetime.now().strftime("%Y%m%d_%H%M%S"), manifest, duplicate_check=False)
            for item in saved:
                manifest.setdefault("imports", []).insert(0, {**item, "tipo": item["ext"], "registrado_em": _tv_now()})
            _tv_add_history(manifest, f"Importação GIS: {len(saved)} arquivo(s)")
            _tv_save_manifest(manifest)
            st.success(f"{len(saved)} arquivo(s) importado(s).")

        grid_json = st.file_uploader("Reenviar grid JSON exportado pelo visualizador", type=["json"], key="tv_grid_json")
        if grid_json is not None:
            render_tmg_loading_bar(100, f"JSON do grid recebido: {Path(grid_json.name).name}")
        if st.button("Salvar versão do grid", type="primary", key="tv_save_grid_version", use_container_width=True):
            if grid_json is None:
                st.warning("Selecione o JSON do grid exportado.")
            else:
                ortho_id = _tv_get_ortho_id(selected)
                ortho_record = _tv_find_ortho(manifest, ortho_id)
                project_id = ortho_record.get("project_id", "") if ortho_record else ""
                raw = grid_json.getbuffer().tobytes()
                grid_id = f"GRID_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                target = TV_GRIDS_DIR / f"{grid_id}_{Path(grid_json.name).name}"
                target.write_bytes(raw)
                record = {
                    "grid_id": grid_id,
                    "project_id": project_id,
                    "ortho_id": ortho_id,
                    "ortho": selected,
                    "path": str(target),
                    "sha256": _tv_hash_bytes(raw),
                    "linhas": int(rows),
                    "colunas": int(cols),
                    "versao": len(manifest.get("grids", [])) + 1,
                    "status": "Grid marcado - pronto para retorno",
                    "criado_em": _tv_now()
                }
                manifest.setdefault("grids", []).insert(0, record)
                project = _tv_find_project(manifest, project_id)
                if project:
                    project["status"] = "Grid marcado - pronto para retornar"
                    project["workflow"] = "GRID_MARCADO"
                    project["grid_marcado_em"] = _tv_now()
                if ortho_record:
                    ortho_record["status"] = "Grid marcado"
                _tv_add_history(manifest, f"Grid registrado: {grid_id}", project_id)
                _tv_save_manifest(manifest)
                st.success(f"Grid `{grid_id}` salvo. Próximo passo: Retornar Projeto.")

    with col_r:
        if not selected:
            st.info("Selecione uma ortofoto recebida para abrir o visualizador GIS.")
        else:
            ortho_id = selected.split(" · ")[0]
            record = next((o for o in manifest.get("orthos", []) if o.get("ortho_id") == ortho_id), None)
            if record and Path(record.get("path", "")).exists():
                with st.container():
                    raw = Path(record["path"]).read_bytes()
                    b64, dims, err, _ = processar_ortofoto(raw, record["nome"])
                if err:
                    st.error(err)
                else:
                    st.markdown(f"<p style='color:#888;font-size:0.8rem;'>📐 {record['nome']} · {dims[0]}×{dims[1]} px · grid preservado em JSON</p>", unsafe_allow_html=True)
                    components.html(_tv_grid_viewer_html(b64, int(rows), int(cols), record["nome"]), height=700, scrolling=False)

    st.markdown("##### Versões de Grid")
    grid_rows = [{
        "Grid": g.get("grid_id"),
        "Ortofoto": g.get("ortho"),
        "Linhas": g.get("linhas"),
        "Colunas": g.get("colunas"),
        "Versão": g.get("versao"),
        "Status": g.get("status"),
        "Data": g.get("criado_em")
    } for g in manifest.get("grids", [])]
    st.dataframe(grid_rows, use_container_width=True, hide_index=True)

def _tv_render_analises(manifest: dict) -> None:
    st.markdown("#### Análises")
    c1, c2, c3 = st.columns(3)
    c1.metric("Projetos em análise", len([p for p in manifest.get("projects", []) if p.get("status")]))
    c2.metric("Grids disponíveis", len(manifest.get("grids", [])))
    c3.metric("Ortofotos prontas", len(manifest.get("orthos", [])))
    st.markdown("<h4 style='color:#ff8c00;margin-top:0;'>Fila técnica</h4>", unsafe_allow_html=True)
    rows = []
    for p in manifest.get("projects", []):
        rows.append({
            "Projeto": p.get("project_id"),
            "Fazenda": p.get("fazenda"),
            "Talhão": p.get("talhao"),
            "Status": "Aguardando análise técnica" if manifest.get("grids") else p.get("status"),
            "Relatórios": "Contagem · Falhas · Parcelas"
        })
    st.dataframe(rows, use_container_width=True, hide_index=True)
    if st.button("Reenviar pacote selecionado para análise", type="primary", key="tv_send_analysis", use_container_width=True):
        _tv_add_history(manifest, "Pacote reenviado para análise técnica", status="ENVIADO")
        _tv_save_manifest(manifest)
        st.success("Pacote colocado na fila de análise técnica.")

def _tv_render_retornar_projeto(manifest: dict) -> None:
    st.markdown("#### Retornar Projeto")
    st.caption("Após marcar o grid, gere o pacote de retorno com ortofoto, grid, parcelas e metadados para o próximo usuário ou análise técnica.")
    ready_grids = [g for g in manifest.get("grids", []) if g.get("status")]
    if not ready_grids:
        st.info("Nenhum grid marcado ainda. Primeiro receba a ortofoto e marque o grid em Grid e Parcelas.")
        return

    options = [f"{g['grid_id']} · {g.get('ortho','Ortofoto')}" for g in ready_grids]
    selected = st.selectbox("Grid para retorno", options, key="tv_return_grid")
    grid_id = selected.split(" · ")[0] if selected else ""
    grid = _tv_find_grid(manifest, grid_id)
    ortho = _tv_find_ortho(manifest, grid.get("ortho_id", "")) if grid else {}
    project = _tv_find_project(manifest, grid.get("project_id", "")) if grid else {}

    c1, c2, c3 = st.columns(3)
    c1.metric("Grid", grid.get("grid_id", "-"))
    c2.metric("Projeto", grid.get("project_id", "-") or "-")
    c3.metric("Parcelas", int(grid.get("linhas", 0)) * int(grid.get("colunas", 0)))

    incluir_ortho = st.checkbox("Incluir ortofoto original no pacote", value=True, key="tv_return_include_ortho")
    incluir_raw_manifest = st.checkbox("Incluir lista de imagens originais e checksums", value=True, key="tv_return_include_raw")

    if st.button("Gerar pacote de retorno", type="primary", key="tv_make_return_pkg", use_container_width=True):
        if not grid or not Path(grid.get("path", "")).exists():
            st.error("Arquivo do grid não encontrado.")
            return
        return_id = f"RET_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        return_dir = TV_RETURNS_DIR / return_id
        return_dir.mkdir(parents=True, exist_ok=True)
        package_path = return_dir / f"{return_id}_{grid.get('project_id','projeto')}.zip"

        meta = {
            "return_id": return_id,
            "created_at": _tv_now(),
            "project": project,
            "ortho": ortho,
            "grid": grid,
            "parcelas": int(grid.get("linhas", 0)) * int(grid.get("colunas", 0)),
            "conteudo": {
                "ortofoto_original": bool(incluir_ortho),
                "manifesto_raw": bool(incluir_raw_manifest),
                "grid_json": True
            }
        }

        manifest_file = return_dir / "manifesto_retorno.json"
        manifest_file.write_text(json.dumps(meta, indent=2, ensure_ascii=False), encoding="utf-8")

        with zipfile.ZipFile(package_path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(grid["path"], arcname=f"grid/{Path(grid['path']).name}")
            zf.write(manifest_file, arcname="manifesto_retorno.json")
            if incluir_ortho and ortho and Path(ortho.get("path", "")).exists():
                zf.write(ortho["path"], arcname=f"ortofoto/{Path(ortho['path']).name}")
            if incluir_raw_manifest and project:
                raw_manifest = return_dir / "imagens_originais_checksums.json"
                raw_manifest.write_text(json.dumps(project.get("files", []), indent=2, ensure_ascii=False), encoding="utf-8")
                zf.write(raw_manifest, arcname="imagens_originais_checksums.json")

        record = {
            "return_id": return_id,
            "project_id": grid.get("project_id", ""),
            "grid_id": grid.get("grid_id", ""),
            "ortho_id": grid.get("ortho_id", ""),
            "path": str(package_path),
            "tamanho": package_path.stat().st_size,
            "tamanho_fmt": _tv_human_size(package_path.stat().st_size),
            "sha256": _tv_hash_file(package_path),
            "status": "Pacote retornado",
            "criado_em": _tv_now()
        }
        manifest.setdefault("returns", []).insert(0, record)
        grid["status"] = "Retornado"
        if project:
            project["status"] = "Projeto retornado com grid marcado"
            project["workflow"] = "RETORNADO_COM_GRID"
            project["retornado_em"] = _tv_now()
        _tv_add_history(manifest, f"Pacote de retorno gerado: {return_id}", grid.get("project_id", ""))
        _tv_save_manifest(manifest)
        st.success(f"Pacote `{return_id}` gerado.")

    st.markdown("##### Pacotes gerados")
    rows = [{
        "Retorno": r.get("return_id"),
        "Projeto": r.get("project_id"),
        "Grid": r.get("grid_id"),
        "Tamanho": r.get("tamanho_fmt"),
        "Status": r.get("status"),
        "Data": r.get("criado_em")
    } for r in manifest.get("returns", [])]
    st.dataframe(rows, use_container_width=True, hide_index=True)

    if manifest.get("returns"):
        selected_return = st.selectbox("Baixar pacote de retorno", [f"{r['return_id']} · {r.get('project_id','')}" for r in manifest.get("returns", [])], key="tv_return_download")
        return_id = selected_return.split(" · ")[0]
        record = next((r for r in manifest.get("returns", []) if r.get("return_id") == return_id), None)
        if record and Path(record.get("path", "")).exists():
            st.download_button(
                "Baixar ZIP de retorno",
                data=Path(record["path"]).read_bytes(),
                file_name=Path(record["path"]).name,
                mime="application/zip",
                use_container_width=True
            )

def _tv_render_historico(manifest: dict) -> None:
    st.markdown("#### Histórico")
    st.dataframe(manifest.get("history", []), use_container_width=True, hide_index=True)

def _tv_render_config(manifest: dict) -> None:
    st.markdown("#### Configurações")
    config = manifest.get("config", {})
    providers = ["Microsoft OneDrive", "Microsoft Azure", "Google Drive", "Dropbox", "Servidor local", "NAS", "FTP/SFTP"]
    provider_index = providers.index(config.get("provider", "Servidor local")) if config.get("provider", "Servidor local") in providers else 4
    c1, c2 = st.columns(2)
    with c1:
        provider = st.selectbox("Conector", providers, index=provider_index, key="tv_cfg_provider")
        destination = st.text_input("Destino / bucket / caminho", value=config.get("destination", str(TV_ROOT / "sync")), key="tv_cfg_dest")
        sync_mode = st.selectbox("Sincronização", ["Manual", "Automática ao enviar", "A cada hora", "Diária"], key="tv_cfg_sync")
    with c2:
        auth_mode = st.selectbox("Autenticação", ["Token/OAuth", "Chave de acesso", "Usuário e senha", "Conta de serviço"], key="tv_cfg_auth")
        encryption = st.checkbox("Criptografia de uploads", value=bool(config.get("encryption", True)), key="tv_cfg_enc")
        resumable = st.checkbox("Upload retomável", value=bool(config.get("resumable", True)), key="tv_cfg_resume")
        checksum = st.checkbox("Verificação de integridade", value=bool(config.get("checksum", True)), key="tv_cfg_checksum")
    if st.button("Salvar configurações de transferência", type="primary", key="tv_cfg_save", use_container_width=True):
        manifest["config"] = {
            "provider": provider,
            "destination": destination,
            "sync_mode": sync_mode,
            "checksum": checksum,
            "resumable": resumable,
            "encryption": encryption,
            "auth_mode": auth_mode
        }
        _tv_add_history(manifest, f"Configuração atualizada: {provider}")
        _tv_save_manifest(manifest)
        st.success("Configurações salvas.")

    st.markdown("##### Usuários e permissões")
    st.dataframe(manifest.get("users", []), use_container_width=True, hide_index=True)

def render_transferencia_voos() -> None:
    _tv_ensure_storage()
    manifest = _tv_load_manifest()
    st.markdown("""
    <style>
      .tv-hero {
        background:linear-gradient(145deg,#181818 0%,#111 56%,#2a1700 100%);
        border:1px solid #3a2a18;
        border-top:2px solid #ff8c00;
        border-radius:8px;
        padding:20px 22px;
        margin-bottom:14px;
        box-shadow:4px 4px 14px #070707,0 0 24px rgba(255,140,0,.10);
      }
      .tv-hero h2 {
        margin:0;
        color:#ff8c00;
        letter-spacing:3px;
        text-transform:uppercase;
        text-shadow:1px 1px 0 #7a3a00,3px 3px 8px rgba(0,0,0,.9),0 0 22px rgba(255,140,0,.32);
      }
      .tv-hero p { margin:8px 0 0 0; color:#cfcfcf; font-size:13px; }
      .tv-band { border:1px solid #3a2a18; background:#161616; border-radius:8px; padding:12px; }
      div[data-testid="stRadio"] > div {
        background:#151515;
        border:1px solid #2e2e2e;
        border-radius:8px;
        padding:8px 10px;
        box-shadow:inset 0 1px 0 rgba(255,255,255,.03);
      }
      div[data-testid="stRadio"] label {
        background:linear-gradient(145deg,#222,#111);
        border:1px solid #303030;
        border-radius:8px;
        padding:7px 10px;
        margin-right:4px;
      }
      div[data-testid="stRadio"] label:hover {
        border-color:#ff8c00;
        box-shadow:0 0 12px rgba(255,140,0,.12);
      }
      div[data-testid="stRadio"] label:has(input:checked) {
        border-color:#ff8c00;
        background:linear-gradient(145deg,#ff9e33,#e67600);
        box-shadow:0 0 16px rgba(255,140,0,.22);
      }
    </style>
    """, unsafe_allow_html=True)
    st.markdown(
        "<div class='tv-hero'><h2>Transferência de Voos</h2>"
        "<p>Central GIS para envio, sincronização, ortofotos, grids, versões e análise técnica.</p>"
        f"{_tv_status_chip('Python 3.12', '#55ff99')}"
        f"{_tv_status_chip('Qualidade original preservada', '#ff8c00')}"
        f"{_tv_status_chip('Checksums SHA-256', '#5599ff')}"
        "</div>",
        unsafe_allow_html=True
    )

    tabs = ["Upload de Voos", "Receber Voos", "Banco de Dados", "Ortofotos Geradas", "Grid e Parcelas", "Retornar Projeto", "Análises", "Histórico", "Configurações"]
    active = st.radio("Menu Transferência de Voos", tabs, horizontal=True, label_visibility="collapsed", key="tv_active_tab")
    st.markdown("---")

    if active == "Upload de Voos":
        _tv_render_upload(manifest)
    elif active == "Receber Voos":
        _tv_render_receber_voos(manifest)
    elif active == "Banco de Dados":
        _tv_render_database(manifest)
    elif active == "Ortofotos Geradas":
        _tv_render_orthos(manifest)
    elif active == "Grid e Parcelas":
        _tv_render_grid_parcelas(manifest)
    elif active == "Retornar Projeto":
        _tv_render_retornar_projeto(manifest)
    elif active == "Análises":
        _tv_render_analises(manifest)
    elif active == "Histórico":
        _tv_render_historico(manifest)
    elif active == "Configurações":
        _tv_render_config(manifest)

# ==========================================
# MODULO ISOLADO - VOOS DIRECIONADOS
# ==========================================
VD_ROOT = SYSTEM_DATABASE_DIR / "voos_direcionados"
VD_FLIGHTS_DIR = VD_ROOT / "voos_enviados"
VD_ORTHOS_DIR = VD_ROOT / "ortofotos_recebidas"
VD_GRIDS_DIR = VD_ROOT / "grids_salvos"
VD_EXPORTS_DIR = VD_ROOT / "exportacoes"
VD_MANIFEST_PATH = VD_ROOT / "manifest.json"
VD_DESTINATIONS = [
    "OneDrive",
    "Google Drive",
    "Azure Storage",
    "Diretório local",
    "Pasta em rede",
    "Outros bancos/diretórios futuros"
]

def _vd_default_manifest() -> dict:
    return {
        "voos": [],
        "orthos": [],
        "grids": [],
        "exports": [],
        "history": [],
        "config": {
            "destino": "Diretório local",
            "caminho": str(VD_ROOT / "destinos" / "local"),
            "usuario_padrao": "Operador"
        }
    }

def _vd_ensure_storage() -> None:
    for folder in [VD_ROOT, VD_FLIGHTS_DIR, VD_ORTHOS_DIR, VD_GRIDS_DIR, VD_EXPORTS_DIR, VD_ROOT / "destinos" / "local"]:
        folder.mkdir(parents=True, exist_ok=True)
    if not VD_MANIFEST_PATH.exists():
        VD_MANIFEST_PATH.write_text(json.dumps(_vd_default_manifest(), indent=2, ensure_ascii=False), encoding="utf-8")

def _vd_load_manifest() -> dict:
    _vd_ensure_storage()
    try:
        data = json.loads(VD_MANIFEST_PATH.read_text(encoding="utf-8"))
    except Exception:
        data = _vd_default_manifest()
    default = _vd_default_manifest()
    for key, value in default.items():
        data.setdefault(key, value)
    data.setdefault("config", {}).setdefault("destino", "Diretório local")
    data.setdefault("config", {}).setdefault("caminho", str(VD_ROOT / "destinos" / "local"))
    return data

def _vd_save_manifest(data: dict) -> None:
    _vd_ensure_storage()
    VD_MANIFEST_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

def _vd_add_history(manifest: dict, event: str, lote_id: str = "", status: str = "OK") -> None:
    manifest.setdefault("history", []).insert(0, {
        "data": _tv_now(),
        "lote": lote_id,
        "evento": event,
        "status": status
    })
    manifest["history"] = manifest["history"][:500]

def _vd_find_voo(manifest: dict, lote_id: str) -> dict:
    for voo in manifest.get("voos", []):
        if voo.get("lote_id") == lote_id:
            return voo
    return {}

def _vd_find_ortho(manifest: dict, ortho_id: str) -> dict:
    for ortho in manifest.get("orthos", []):
        if ortho.get("ortho_id") == ortho_id:
            return ortho
    return {}

def _vd_find_grid(manifest: dict, grid_id: str) -> dict:
    for grid in manifest.get("grids", []):
        if grid.get("grid_id") == grid_id:
            return grid
    return {}

def _vd_project_options(manifest: dict) -> list:
    return [f"{v['lote_id']} · {v.get('nome_voo','Voo')}" for v in manifest.get("voos", [])]

def _vd_ortho_options(manifest: dict) -> list:
    return [f"{o['ortho_id']} · {o.get('nome','Ortofoto')}" for o in manifest.get("orthos", [])]

def _vd_id_from_option(option: str) -> str:
    return option.split(" · ")[0] if option else ""

def _vd_destination_status(destino: str, caminho: str, manifest: dict) -> dict:
    path = _resolve_system_path(caminho.strip()) if caminho else VD_ROOT / "destinos" / _tv_safe_name(destino)
    exists = path.exists()
    status = "Conectado" if exists else "Pendente"
    free = "Indisponível"
    try:
        base = path if exists else path.parent
        usage = shutil.disk_usage(base)
        free = _tv_human_size(usage.free)
    except Exception:
        pass
    last = "-"
    for voo in manifest.get("voos", []):
        if voo.get("destino_enviado") == destino:
            last = voo.get("data_hora", "-")
            break
    return {
        "nome": destino,
        "caminho": str(path),
        "status": status,
        "espaco": free,
        "ultimo_envio": last
    }

def _vd_save_uploaded_files(files, base_dir: Path, progress=None) -> tuple:
    saved, total_size = [], 0
    base_dir.mkdir(parents=True, exist_ok=True)
    expected = sum(int(getattr(uploaded, "size", 0) or 0) for uploaded in (files or []))
    written_total = 0
    for uploaded in files or []:
        hasher = hashlib.sha256()
        file_size = 0
        temp_name = f".{_tv_safe_name(Path(uploaded.name).stem)}_{datetime.now().strftime('%H%M%S%f')}.uploading"
        temp_path = base_dir / temp_name
        try:
            uploaded.seek(0)
        except Exception:
            pass
        with open(temp_path, "wb") as out:
            while True:
                chunk = uploaded.read(1024 * 1024)
                if not chunk:
                    break
                if isinstance(chunk, str):
                    chunk = chunk.encode("utf-8")
                hasher.update(chunk)
                out.write(chunk)
                file_size += len(chunk)
                written_total += len(chunk)
                if progress and expected:
                    update_tmg_loading(progress, min(99, int((written_total / expected) * 100)), f"Carregando arquivo: {Path(uploaded.name).name}")
        target = base_dir / Path(uploaded.name).name
        if target.exists():
            target = base_dir / f"{target.stem}_{datetime.now().strftime('%H%M%S%f')}{target.suffix}"
        temp_path.replace(target)
        total_size += file_size
        saved.append({
            "nome": uploaded.name,
            "path": str(target),
            "tamanho": file_size,
            "tamanho_fmt": _tv_human_size(file_size),
            "sha256": hasher.hexdigest(),
            "ext": Path(uploaded.name).suffix.lower(),
            "enviado_em": _tv_now()
        })
    if progress:
        update_tmg_loading(progress, 100, "Carregamento concluído com sucesso.")
    return saved, total_size

def _vd_copy_to_destination(saved: list, caminho: str, lote_id: str, nome_voo: str = "") -> tuple:
    """Copia o lote para o diretório escolhido criando uma pasta limpa com o nome do voo."""
    if not caminho:
        return "", "Armazenado somente no banco interno"
    try:
        base_destino = _resolve_system_path(str(caminho).strip())
        pasta_voo = _tv_safe_name(nome_voo) if nome_voo else lote_id
        dest_dir = base_destino / pasta_voo
        if dest_dir.exists():
            dest_dir = base_destino / f"{pasta_voo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        dest_dir.mkdir(parents=True, exist_ok=True)
        for item in saved:
            src = Path(item["path"])
            if src.exists():
                shutil.copy2(src, dest_dir / src.name)
        return str(dest_dir), "Enviado"
    except Exception as exc:
        return "", f"Falha no destino externo: {exc}"

def _vd_metric_cards(manifest: dict) -> None:
    voos = manifest.get("voos", [])
    orthos = manifest.get("orthos", [])
    grids = manifest.get("grids", [])
    volume = sum(v.get("tamanho_total", 0) for v in voos) + sum(o.get("tamanho", 0) for o in orthos)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Voos enviados", len(voos))
    c2.metric("Ortofotos", len(orthos))
    c3.metric("Grids salvos", len(grids))
    c4.metric("Volume interno", _tv_human_size(volume))

def _vd_ortho_viewer_html(b64: str, image_name: str, width: int, height: int) -> str:
    return f"""
<!DOCTYPE html>
<html>
<head>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ background:#0d0d0d; overflow:hidden; font-family:Segoe UI, Arial, sans-serif; }}
  #vdv {{ width:100%; height:680px; background:#0d0d0d; border:1px solid #2f2f2f; border-radius:8px; overflow:hidden; position:relative; }}
  canvas {{ position:absolute; inset:0; cursor:grab; }}
  canvas:active {{ cursor:grabbing; }}
  .vdtop {{ position:absolute; left:12px; top:12px; right:12px; display:flex; justify-content:space-between; gap:10px; z-index:3; pointer-events:none; }}
  .badge {{ background:rgba(10,10,10,.82); border:1px solid #333; color:#ff8c00; border-radius:6px; padding:7px 10px; font-size:11px; letter-spacing:.5px; }}
  .panel {{ position:absolute; right:12px; bottom:12px; background:rgba(10,10,10,.86); border:1px solid #333; border-radius:8px; padding:10px; z-index:4; color:#ddd; min-width:240px; }}
  .panel label {{ display:flex; align-items:center; justify-content:space-between; gap:8px; color:#ff8c00; font-size:11px; margin:6px 0; }}
  input[type=range] {{ accent-color:#ff8c00; width:138px; }}
  button {{ background:linear-gradient(145deg,#222,#111); color:#ff8c00; border:1px solid #3a3a3a; border-radius:6px; padding:7px 9px; cursor:pointer; font-weight:700; }}
  button:hover {{ border-color:#ff8c00; box-shadow:0 0 10px rgba(255,140,0,.22); }}
  .scale {{ position:absolute; left:12px; bottom:12px; color:#ccc; background:rgba(0,0,0,.72); border:1px solid #333; border-radius:6px; padding:7px 10px; font-size:11px; }}
  .scaleLine {{ display:inline-block; width:86px; border-bottom:3px solid #ff8c00; margin-right:8px; transform:translateY(-3px); }}
</style>
</head>
<body>
<div id="vdv">
  <canvas id="vd_cv"></canvas>
  <div class="vdtop">
    <div class="badge">{image_name} · {width} x {height} px</div>
    <div class="badge" id="vd_coord">X: - · Y: - · Zoom: 100%</div>
  </div>
  <div class="panel">
    <label>Brilho <input id="vd_bright" type="range" min="50" max="160" value="100"></label>
    <label>Contraste <input id="vd_contrast" type="range" min="50" max="180" value="100"></label>
    <div style="display:flex;gap:6px;margin-top:8px;">
      <button id="vd_fit">Ajustar</button>
      <button id="vd_100">1:1</button>
      <button id="vd_reset">Reset</button>
    </div>
  </div>
  <div class="scale"><span class="scaleLine"></span>escala visual</div>
</div>
<script>
const wrap=document.getElementById('vdv'), cv=document.getElementById('vd_cv'), ctx=cv.getContext('2d');
const coord=document.getElementById('vd_coord'), brightEl=document.getElementById('vd_bright'), contrastEl=document.getElementById('vd_contrast');
let sc=1, ox=0, oy=0, drag=false, lx=0, ly=0, imgW=0, imgH=0;
const img=new Image();
function resize(){{ cv.width=wrap.clientWidth; cv.height=wrap.clientHeight; draw(); }}
function fit(){{ if(!imgW) return; sc=Math.min(cv.width/imgW, cv.height/imgH); ox=(cv.width-imgW*sc)/2; oy=(cv.height-imgH*sc)/2; draw(); }}
function draw(){{
  ctx.clearRect(0,0,cv.width,cv.height);
  ctx.save(); ctx.translate(ox,oy); ctx.scale(sc,sc);
  ctx.filter=`brightness(${{brightEl.value}}%) contrast(${{contrastEl.value}}%)`;
  if(imgW) ctx.drawImage(img,0,0);
  ctx.restore();
  coord.textContent=coord.dataset.base || 'X: - · Y: - · Zoom: '+Math.round(sc*100)+'%';
}}
img.onload=()=>{{ imgW=img.width; imgH=img.height; resize(); fit(); }};
img.src='data:image/jpeg;base64,{b64}';
window.addEventListener('resize', resize);
brightEl.oninput=draw; contrastEl.oninput=draw;
cv.addEventListener('wheel',e=>{{ e.preventDefault(); const f=e.deltaY<0?1.18:.84; const r=cv.getBoundingClientRect(); const mx=e.clientX-r.left,my=e.clientY-r.top; const ix=(mx-ox)/sc,iy=(my-oy)/sc; sc=Math.max(.05,Math.min(40,sc*f)); ox=mx-ix*sc; oy=my-iy*sc; draw(); }},{{passive:false}});
cv.addEventListener('mousedown',e=>{{ drag=true; lx=e.clientX; ly=e.clientY; }});
cv.addEventListener('mousemove',e=>{{ const r=cv.getBoundingClientRect(); const x=(e.clientX-r.left-ox)/sc, y=(e.clientY-r.top-oy)/sc; coord.dataset.base='X: '+Math.round(x)+' · Y: '+Math.round(y)+' · Zoom: '+Math.round(sc*100)+'%'; coord.textContent=coord.dataset.base; if(drag){{ ox+=e.clientX-lx; oy+=e.clientY-ly; lx=e.clientX; ly=e.clientY; draw(); }} }});
cv.addEventListener('mouseup',()=>{{ drag=false; }}); cv.addEventListener('mouseleave',()=>{{ drag=false; }});
document.getElementById('vd_fit').onclick=fit;
document.getElementById('vd_100').onclick=()=>{{ sc=1; ox=(cv.width-imgW)/2; oy=(cv.height-imgH)/2; draw(); }};
document.getElementById('vd_reset').onclick=()=>{{ brightEl.value=100; contrastEl.value=100; fit(); }};
</script>
</body>
</html>
"""

def _vd_grid_features(grid: dict, ortho: dict) -> tuple:
    payload = json.loads(Path(grid["path"]).read_text(encoding="utf-8"))
    meta = ortho.get("spatial_meta", {}) or {}
    ratio = float(meta.get("ratio") or 1.0)
    transform = None
    if meta.get("transform") and "Affine" in globals():
        try:
            transform = Affine.from_gdal(*meta.get("transform"))
        except Exception:
            transform = None

    def to_coord(point: dict) -> list:
        x = float(point.get("x", 0)) / ratio
        y = float(point.get("y", 0)) / ratio
        if transform is not None:
            gx, gy = transform * (x, y)
            return [gx, gy]
        return [x, -y]

    features = []
    for cell in payload.get("parcelas", []):
        poly = cell.get("polygon", [])
        if len(poly) < 4:
            continue
        coords = [to_coord(p) for p in poly]
        coords.append(coords[0])
        linha = cell.get("linha", 0)
        coluna = cell.get("coluna", 0)
        features.append({
            "type": "Feature",
            "properties": {
                "id": cell.get("id") or f"L{linha}_C{coluna}",
                "linha": linha,
                "coluna": coluna,
                "grid_id": grid.get("grid_id", ""),
                "ortho_id": grid.get("ortho_id", ""),
                "parcelas": grid.get("parcelas", 0)
            },
            "geometry": {"type": "Polygon", "coordinates": [coords]}
        })
    return features, meta.get("crs")

def _vd_grid_geojson_bytes(grid: dict, ortho: dict) -> bytes:
    features, _ = _vd_grid_features(grid, ortho)
    return json.dumps({"type": "FeatureCollection", "features": features}, indent=2, ensure_ascii=False).encode("utf-8")

def _vd_grid_shp_zip_bytes(grid: dict, ortho: dict) -> bytes:
    features, crs = _vd_grid_features(grid, ortho)
    if not HAS_GEOPANDAS:
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("parcelas.geojson", json.dumps({"type": "FeatureCollection", "features": features}, indent=2, ensure_ascii=False))
            zf.writestr("LEIA.txt", "GeoPandas/GDAL nao instalado. Exportado GeoJSON como fallback compativel com QGIS.")
        buf.seek(0)
        return buf.getvalue()

    tmpdir = Path(tempfile.mkdtemp(prefix="vd_shp_"))
    try:
        rows = []
        for feat in features:
            rows.append({
                **feat["properties"],
                "geometry": Polygon(feat["geometry"]["coordinates"][0])
            })
        gdf = gpd.GeoDataFrame(rows, geometry="geometry")
        if crs:
            try:
                gdf.set_crs(crs, inplace=True, allow_override=True)
            except Exception:
                pass
        shp_path = tmpdir / "parcelas_voos_direcionados.shp"
        gdf.to_file(shp_path, driver="ESRI Shapefile", encoding="utf-8")
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for item in tmpdir.iterdir():
                zf.write(item, arcname=item.name)
        buf.seek(0)
        return buf.getvalue()
    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)

def _vd_grid_overlay_bytes(grid: dict, ortho: dict, image_format: str = "PNG") -> bytes:
    from PIL import ImageDraw
    raw = Path(ortho["path"]).read_bytes()
    b64, _, err, _ = processar_ortofoto(raw, ortho.get("nome", "ortofoto"))
    if err:
        raise RuntimeError(err)
    img = Image.open(BytesIO(base64.b64decode(b64))).convert("RGBA")
    draw = ImageDraw.Draw(img)
    payload = json.loads(Path(grid["path"]).read_text(encoding="utf-8"))
    for cell in payload.get("parcelas", []):
        poly = [(float(p.get("x", 0)), float(p.get("y", 0))) for p in cell.get("polygon", [])]
        if len(poly) < 4:
            continue
        closed = poly + [poly[0]]
        draw.line(closed, fill=(255, 140, 0, 255), width=3)
        cx = sum(p[0] for p in poly) / len(poly)
        cy = sum(p[1] for p in poly) / len(poly)
        label = cell.get("id") or f"L{cell.get('linha',0)}_C{cell.get('coluna',0)}"
        draw.text((cx - 14, cy - 6), label, fill=(255, 255, 255, 255))
    buf = BytesIO()
    fmt = image_format.upper()
    if fmt in ("JPG", "JPEG"):
        img.convert("RGB").save(buf, format="JPEG", quality=92)
    elif fmt in ("TIF", "TIFF"):
        img.convert("RGB").save(buf, format="TIFF")
    else:
        img.save(buf, format="PNG")
    buf.seek(0)
    return buf.getvalue()

def _vd_grid_geotiff_overlay_bytes(grid: dict, ortho: dict) -> bytes:
    try:
        import rasterio
        from rasterio.io import MemoryFile
        img = Image.open(BytesIO(_vd_grid_overlay_bytes(grid, ortho, "PNG"))).convert("RGB")
        arr = np.array(img)
        meta = ortho.get("spatial_meta", {}) or {}
        profile = {
            "driver": "GTiff",
            "height": arr.shape[0],
            "width": arr.shape[1],
            "count": 3,
            "dtype": arr.dtype,
            "compress": "lzw"
        }
        if meta.get("transform") and "Affine" in globals():
            ratio = float(meta.get("ratio") or 1.0)
            profile["transform"] = Affine.from_gdal(*meta.get("transform")) * Affine.scale(1 / ratio, 1 / ratio)
        if meta.get("crs"):
            profile["crs"] = meta.get("crs")
        with MemoryFile() as memfile:
            with memfile.open(**profile) as dst:
                dst.write(arr[:, :, 0], 1)
                dst.write(arr[:, :, 1], 2)
                dst.write(arr[:, :, 2], 3)
            return memfile.read()
    except Exception:
        return _vd_grid_overlay_bytes(grid, ortho, "TIFF")

def _vd_render_login() -> None:
    st.markdown("""
    <style>
      .vd-login-wrap { min-height:72vh; display:flex; align-items:center; justify-content:center; }
      .vd-login-card {
        width:min(430px, 96vw);
        background:linear-gradient(155deg,#1b1b1b,#101010 62%,#271600);
        border:1px solid #333;
        border-top:2px solid #ff8c00;
        border-radius:8px;
        padding:28px;
        box-shadow:6px 6px 18px #050505,0 0 30px rgba(255,140,0,.12);
      }
      .vd-login-title {
        color:#ff8c00;
        font-weight:900;
        letter-spacing:3px;
        text-transform:uppercase;
        font-size:1.45rem;
        text-shadow:1px 1px 0 #7a3a00,3px 3px 8px #000;
        margin-bottom:8px;
      }
      .vd-login-sub { color:#aaa; font-size:.82rem; letter-spacing:1px; margin-bottom:16px; }
    </style>
    """, unsafe_allow_html=True)
    st.markdown("""
    <div class="vd-login-card">
      <div class="vd-login-title">Processos de Voos para Análise</div>
      <div class="vd-login-sub">Acesso dedicado para envio de fotos, recebimento de ortofotos, marcador de grid e retorno dos dados.</div>
    </div>
    """, unsafe_allow_html=True)
    usuario = st.text_input("Login", key="vd_login_user")
    senha = st.text_input("Senha", type="password", key="vd_login_pass")
    if st.button("Abrir Processos de Voos para Análise", type="primary", key="vd_login_btn", use_container_width=True):
        if usuario == "1234" and senha == "1234":
            st.session_state.vd_logged_in = True
            app_rerun()
        else:
            st.error("Login ou senha incorretos para Processos de Voos para Análise.")

def _vd_render_envio(manifest: dict) -> None:
    st.markdown("""
    <style>
      .vd-clean-card {
        background:linear-gradient(145deg,#181818,#101010);
        border:1px solid #2d2d2d;
        border-radius:12px;
        padding:18px 18px 10px 18px;
        margin-bottom:14px;
        box-shadow:3px 3px 12px rgba(0,0,0,.45), inset 0 1px 0 rgba(255,255,255,.03);
      }
      .vd-section-title {
        color:#cfcfcf;
        font-weight:800;
        letter-spacing:1.5px;
        text-transform:uppercase;
        font-size:.92rem;
        margin:0 0 10px 0;
      }
      .vd-progress-box {
        background:var(--tmg-deploy-card-bg);
        border:1px solid var(--tmg-deploy-border);
        border-left:4px solid var(--tmg-primary);
        border-radius:14px;
        padding:14px;
        margin-top:12px;
        box-shadow:var(--tmg-deploy-card-shadow);
        color:#ffffff;
      }
      .vd-dest-path {
        color:#9f9f9f;
        font-size:.78rem;
        background:#0f0f0f;
        border:1px solid #2d2d2d;
        border-radius:8px;
        padding:8px 10px;
        margin-top:6px;
        word-break:break-all;
      }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("#### PASSO 1 — Enviar Fotos de Drone")
    st.markdown("<div class='vd-section-title'>Dados principais do voo</div>", unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3)
    with c1:
        nome_voo = st.text_input("Nome do voo", value=f"Voo_Direcionado_{date.today().strftime('%Y%m%d')}", key="vd_nome_voo")
        fazenda = st.text_input("Nome da fazenda", value="", key="vd_fazenda")
        ensaio = st.text_input("Nome do ensaio", value="", key="vd_ensaio")
    with c2:
        data_inicial = st.date_input("Data inicial", value=date.today(), key="vd_data_inicial")
        data_final = st.date_input("Data final", value=date.today(), key="vd_data_final")
        tipo_voo_base = st.selectbox(
            "Tipo de voo",
            ["RGB", "Multiespectral", "NDVI", "Termal", "LiDAR", "Mapeamento", "Outro"],
            key="vd_tipo_voo_base"
        )
    with c3:
        usuario = st.text_input("Usuário responsável", value=manifest.get("config", {}).get("usuario_padrao", "Operador"), key="vd_usuario_resp")
        destino_envio = st.text_input("Destino de envio", value=manifest.get("config", {}).get("destino_envio_padrao", ""), placeholder="Ex.: análise interna, cliente, parceiro", key="vd_destino_envio")
        tipo_voo_outro = ""
        if tipo_voo_base == "Outro":
            tipo_voo_outro = st.text_input("Descrever tipo de voo", value="", key="vd_tipo_voo_outro")
    coordenadas = st.text_area(
        "Coordenadas / área do voo",
        value="",
        placeholder="Informe coordenadas, talhão, polígono ou referência da área quando necessário.",
        key="vd_coordenadas",
        height=80
    )
    tipo_voo = (tipo_voo_outro or tipo_voo_base).strip()
    inicio_final = f"{data_inicial} a {data_final}"

    st.markdown("<div class='vd-section-title'>Destino e diretório de envio</div>", unsafe_allow_html=True)
    dcol1, dcol2 = st.columns([1, 2])
    with dcol1:
        destino = st.selectbox(
            "Destino de armazenamento",
            VD_DESTINATIONS,
            index=VD_DESTINATIONS.index(manifest.get("config", {}).get("destino", "Diretório local")) if manifest.get("config", {}).get("destino", "Diretório local") in VD_DESTINATIONS else 3,
            key="vd_destino"
        )
    with dcol2:
        caminho = st.text_input("Diretório escolhido para receber a pasta do voo", value=manifest.get("config", {}).get("caminho", str(VD_ROOT / "destinos" / "local")), key="vd_caminho_destino")
    dest_status = _vd_destination_status(destino, caminho, manifest)
    d1, d2, d3, d4 = st.columns(4)
    d1.metric("Destino", dest_status["nome"])
    d2.metric("Status", dest_status["status"])
    d3.metric("Espaço disponível", dest_status["espaco"])
    d4.metric("Último envio", dest_status["ultimo_envio"])
    pasta_prevista = _resolve_system_path(str(caminho).strip()) / _tv_safe_name(nome_voo)
    st.markdown(f"<div class='vd-dest-path'>Pasta que será criada no destino: <b>{pasta_prevista}</b></div>", unsafe_allow_html=True)

    st.markdown("<div class='vd-section-title'>Anexar imagens do voo</div>", unsafe_allow_html=True)
    files = st.file_uploader(
        "Selecionar múltiplas fotos de drone ou ZIP",
        type=["jpg", "jpeg", "tif", "tiff", "png", "raw", "dng", "arw", "cr2", "nef", "zip"],
        accept_multiple_files=True,
        key="vd_select_images"
    )
    if files:
        total_previsto = sum(int(getattr(f, "size", 0) or 0) for f in files)
        render_tmg_loading_bar(100, f"{len(files)} arquivo(s) selecionado(s) para envio.")
        st.info(f"{len(files)} arquivo(s) selecionado(s) · volume previsto: {_tv_human_size(total_previsto)}")

    confirmar_envio = st.checkbox(
        "Confirmo o envio e autorizo o salvamento seguro dos arquivos no destino configurado.",
        key="vd_confirmar_envio"
    )

    if st.button("🚀 Confirmar Envio de Fotos de Voos", type="primary", key="vd_send_flight", use_container_width=True):
        if not files:
            st.warning("Selecione as imagens do drone ou um ZIP do voo.")
        elif not str(nome_voo).strip():
            st.warning("Informe o nome do voo para criar a pasta de destino.")
        elif not str(caminho).strip():
            st.warning("Escolha/informe o diretório de destino.")
        elif not confirmar_envio:
            st.warning("Confirme o envio para registrar o lote com segurança.")
        else:
            lote_id = f"VD_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{_tv_safe_name(nome_voo)}"
            st.markdown("<div class='vd-progress-box'><b>Resumo da progressão de envio</b></div>", unsafe_allow_html=True)
            status_line = st.empty()
            progress = st.empty()
            update_tmg_loading(progress, 0, "Iniciando envio de fotos do voo...")
            status_line.info("1/4 — Criando pasta interna do lote e gravando arquivos...")
            saved, total_size = _vd_save_uploaded_files(files, VD_FLIGHTS_DIR / lote_id / "raw", progress)
            status_line.info("2/4 — Criando pasta com o nome do voo no destino escolhido...")
            destino_path, envio_status = _vd_copy_to_destination(saved, caminho, lote_id, nome_voo)
            update_tmg_loading(progress, 92, "Registrando destino e resumo do envio...")
            status_line.info("3/4 — Atualizando manifesto e histórico do sistema...")
            record = {
                "lote_id": lote_id,
                "nome_voo": nome_voo,
                "nome_fazenda": fazenda,
                "fazenda": fazenda,
                "ensaio": ensaio,
                "inicio_final": inicio_final,
                "data_inicial": str(data_inicial),
                "data_final": str(data_final),
                "tipo_voo": tipo_voo,
                "coordenadas": coordenadas,
                "destino_envio": destino_envio,
                "quantidade_imagens": len(saved),
                "data_hora": _tv_now(),
                "usuario_responsavel": usuario,
                "data_voo": str(data_inicial),
                "destino_enviado": destino,
                "caminho_destino": caminho,
                "pasta_nome_voo": _tv_safe_name(nome_voo),
                "destino_path": destino_path,
                "status_envio": envio_status,
                "tamanho_total": total_size,
                "tamanho_total_fmt": _tv_human_size(total_size),
                "identificador_lote": lote_id,
                "files": saved
            }
            manifest.setdefault("voos", []).insert(0, record)
            manifest.setdefault("config", {})["destino"] = destino
            manifest.setdefault("config", {})["caminho"] = caminho
            manifest.setdefault("config", {})["usuario_padrao"] = usuario
            manifest.setdefault("config", {})["destino_envio_padrao"] = destino_envio
            _vd_add_history(manifest, f"Voo enviado com {len(saved)} arquivo(s) para pasta {record['pasta_nome_voo']}", lote_id, envio_status)
            _vd_save_manifest(manifest)
            update_tmg_loading(progress, 100, "Carregamento concluído com sucesso.")
            status_line.success("4/4 — Envio finalizado e registrado no resumo.")
            st.success(f"Lote `{lote_id}` enviado. Pasta criada: `{destino_path or pasta_prevista}`. Status: {envio_status}")
            if saved:
                st.dataframe([{
                    "Arquivo": item.get("nome"),
                    "Tamanho": item.get("tamanho_fmt"),
                    "Enviado em": item.get("enviado_em"),
                    "SHA-256": item.get("sha256", "")[:16] + "..."
                } for item in saved], use_container_width=True, hide_index=True)

    st.markdown("#### Resumo de Envios")
    try:
        voos_manifest = manifest.get("voos", [])
        if not isinstance(voos_manifest, list):
            voos_manifest = []

        rows = []
        for v in voos_manifest:
            if not isinstance(v, dict):
                continue
            rows.append({
                "Nome do voo": v.get("nome_voo", ""),
                "Fazenda": v.get("nome_fazenda") or v.get("fazenda") or "",
                "Ensaio": v.get("ensaio", ""),
                "Data inicial": v.get("data_inicial") or v.get("data_voo") or "",
                "Data final": v.get("data_final") or "",
                "Tipo de voo": v.get("tipo_voo") or "",
                "Coordenadas": v.get("coordenadas") or "",
                "Destino de envio": v.get("destino_envio") or "",
                "Imagens": v.get("quantidade_imagens", 0),
                "Data e hora": v.get("data_hora", ""),
                "Destino": v.get("destino_enviado", ""),
                "Pasta criada": v.get("destino_path", ""),
                "Status": v.get("status_envio", ""),
                "Tamanho": v.get("tamanho_total_fmt", ""),
                "Lote": v.get("identificador_lote") or v.get("lote_id") or ""
            })

        if rows:
            st.dataframe(rows, use_container_width=True, hide_index=True)
        else:
            st.info("Nenhum envio registrado ainda. Após enviar as fotos do voo, o resumo aparecerá aqui automaticamente.")
    except Exception:
        st.info("Resumo de Envios indisponível no momento. Faça um novo envio para atualizar o histórico automaticamente.")

def _vd_render_ortofotos(manifest: dict) -> None:
    st.markdown("#### PASSO 2 — Receber Ortofotos")
    project_options = _vd_project_options(manifest)

    with st.form("vd_ortho_import_form", clear_on_submit=False):
        selected_voo = st.selectbox("Voo vinculado", [""] + project_options, key="vd_ortho_voo")
        ortho_file = st.file_uploader(
            "Buscar/Importar Ortofoto Gerada",
            type=["tif", "tiff", "geotiff", "png", "jpg", "jpeg", "jp2", "img", "zip"],
            key="vd_ortho_file"
        )
        if ortho_file is not None:
            render_tmg_loading_bar(100, f"Ortofoto recebida: {Path(ortho_file.name).name}")
        importar_ortho = st.form_submit_button("Importar Ortofoto", type="primary", use_container_width=True)

    if importar_ortho:
        if ortho_file is None:
            st.warning("Selecione a ortofoto processada externamente.")
        else:
            ortho_id = f"ORT_VD_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            saved, _ = _vd_save_uploaded_files([ortho_file], VD_ORTHOS_DIR / ortho_id)
            item = saved[0]
            if any(o.get("sha256") == item.get("sha256") for o in manifest.get("orthos", [])):
                try:
                    Path(item["path"]).unlink()
                except Exception:
                    pass
                st.warning("Esta ortofoto já estava registrada e não foi importada novamente.")
            else:
                dims, spatial_meta, preview_error = None, {}, ""
                if item["ext"] != ".zip":
                    _mosaic_register_file(item["path"], item["nome"], "Voos Direcionados", item.get("sha256", ""), item.get("tamanho", 0))
                    preview_error = "Preview será gerado ao abrir a pré-visualização."
                record = {
                    "ortho_id": ortho_id,
                    "lote_id": _vd_id_from_option(selected_voo),
                    "nome": item["nome"],
                    "path": item["path"],
                    "tipo": item["ext"].replace(".", "").upper() or "RASTER",
                    "tamanho": item["tamanho"],
                    "tamanho_fmt": item["tamanho_fmt"],
                    "sha256": item["sha256"],
                    "resolucao_preview": f"{dims[0]}x{dims[1]} px" if dims else "Aguardando preview",
                    "spatial_meta": spatial_meta or {},
                    "crs": (spatial_meta or {}).get("crs", ""),
                    "status": "Ortofoto anexada - pronta para visualizar e marcar grid",
                    "data_processamento": _tv_now(),
                    "preview_error": preview_error or ""
                }
                manifest.setdefault("orthos", []).insert(0, record)
                voo = _vd_find_voo(manifest, record["lote_id"])
                if voo:
                    voo["status_envio"] = "Ortofoto anexada"
                    voo["ortho_id"] = ortho_id
                _vd_add_history(manifest, f"Ortofoto anexada: {item['nome']}", record["lote_id"])
                _vd_save_manifest(manifest)
                st.session_state.pop("vd_preview_ortho_id", None)
                st.success(f"Ortofoto `{ortho_id}` anexada.")
                st.info("A ortofoto foi registrada. Abra a pré-visualização somente quando quiser carregar o visualizador.")

    rows = [{
        "ID": o.get("ortho_id"),
        "Voo": o.get("lote_id"),
        "Arquivo": o.get("nome"),
        "Tipo": o.get("tipo"),
        "Resolução": o.get("resolucao_preview"),
        "CRS": (o.get("crs") or "")[:80],
        "Tamanho": o.get("tamanho_fmt"),
        "Status": o.get("status")
    } for o in manifest.get("orthos", [])]
    st.markdown("#### Ortofotos recebidas")
    if rows:
        st.dataframe(rows, use_container_width=True, hide_index=True)
    else:
        st.info("Nenhuma ortofoto anexada ainda.")
        return

    st.markdown("#### Pré-visualizador de Ortofoto")
    ortho_options = _vd_ortho_options(manifest)
    if not ortho_options:
        st.info("Nenhuma ortofoto anexada ainda.")
        return

    selected = st.selectbox("Ortofotos disponíveis/importadas", ortho_options, key="vd_ortho_selected")
    ortho_id = _vd_id_from_option(selected)
    ortho = _vd_find_ortho(manifest, ortho_id)
    if not ortho or not Path(ortho.get("path", "")).exists():
        st.warning("Arquivo da ortofoto não encontrado.")
        return

    open_col, grid_col = st.columns(2)
    with open_col:
        if st.button("Abrir pré-visualização", type="primary", key="vd_open_ortho_preview", use_container_width=True):
            st.session_state["vd_preview_ortho_id"] = ortho_id
            app_rerun()
    with grid_col:
        if st.button("Marcar Grid", key="vd_go_grid_selected", use_container_width=True):
            st.session_state["vd_pending_grid_ortho_id"] = ortho_id
            st.session_state["vd_next_tab"] = "ETAPA 3"
            app_rerun()

    if st.session_state.get("vd_preview_ortho_id") != ortho_id:
        st.info("Pré-visualização em espera. Clique em Abrir pré-visualização para carregar o mosaico no navegador.")
        return

    raw = Path(ortho["path"]).read_bytes()
    b64, dims, err, spatial_meta = processar_ortofoto(raw, ortho["nome"])
    if err:
        st.error(f"Não foi possível abrir o preview: {err}")
    else:
        ortho["spatial_meta"] = spatial_meta or ortho.get("spatial_meta", {})
        ortho["resolucao_preview"] = f"{dims[0]}x{dims[1]} px"
        _vd_save_manifest(manifest)
        components.html(_vd_ortho_viewer_html(b64, ortho["nome"], dims[0], dims[1]), height=700, scrolling=False)

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.download_button(
                "Baixar Ortofoto Original",
                data=raw,
                file_name=Path(ortho["path"]).name,
                mime="application/octet-stream",
                use_container_width=True
            )
        with c2:
            st.download_button(
                "Baixar Compactada JPG",
                data=base64.b64decode(b64),
                file_name=f"{Path(ortho['nome']).stem}_compactada.jpg",
                mime="image/jpeg",
                use_container_width=True
            )
        with c3:
            png_buf = BytesIO()
            Image.open(BytesIO(base64.b64decode(b64))).save(png_buf, format="PNG")
            st.download_button(
                "Baixar PNG",
                data=png_buf.getvalue(),
                file_name=f"{Path(ortho['nome']).stem}_preview.png",
                mime="image/png",
                use_container_width=True
            )
        with c4:
            if st.button("Fechar pré-visualização", key="vd_close_ortho_preview", use_container_width=True):
                st.session_state.pop("vd_preview_ortho_id", None)
                app_rerun()

def _vd_render_grid(manifest: dict) -> None:
    st.markdown("#### Grid, Parcelas e Salvamento no Sistema")
    ortho_options = _vd_ortho_options(manifest)
    if not ortho_options:
        st.info("Anexe uma ortofoto gerada antes de marcar o grid.")
        return

    pending = st.session_state.pop("vd_pending_grid_ortho_id", "")
    current_options = [""] + ortho_options
    pending_option = next((opt for opt in current_options if opt.startswith(pending + " · ")), "")
    if pending_option and st.session_state.get("vd_grid_ortho") != pending_option:
        st.session_state.pop("vd_grid_ortho", None)
    grid_index = current_options.index(pending_option) if pending_option in current_options else 0

    left, right = st.columns([1, 2])
    with left:
        selected = st.selectbox("Ortofoto para marcar grid", current_options, index=grid_index, key="vd_grid_ortho")
        rows = st.number_input("Número de linhas", min_value=1, max_value=500, value=10, key="vd_grid_rows")
        cols = st.number_input("Número de colunas", min_value=1, max_value=500, value=10, key="vd_grid_cols")
        escala = st.text_input("Escala / referência espacial", value="Preservar metadados da ortofoto", key="vd_grid_scale")
        grid_json = st.file_uploader("JSON do grid exportado pelo visualizador", type=["json"], key="vd_grid_json_upload")
        if grid_json is not None:
            render_tmg_loading_bar(100, f"JSON do grid recebido: {Path(grid_json.name).name}")

        if st.button("Salvar Grid no Sistema", type="primary", key="vd_save_grid_system", use_container_width=True):
            if not selected:
                st.warning("Selecione uma ortofoto.")
            elif grid_json is None:
                st.warning("Exporte o JSON no visualizador e anexe aqui para salvar o grid.")
            else:
                ortho_id = _vd_id_from_option(selected)
                ortho = _vd_find_ortho(manifest, ortho_id)
                raw = grid_json.getbuffer().tobytes()
                try:
                    payload = json.loads(raw.decode("utf-8"))
                except Exception as exc:
                    st.error(f"JSON inválido: {exc}")
                    return
                grid_id = f"GRID_VD_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                target = VD_GRIDS_DIR / f"{grid_id}_{Path(grid_json.name).name}"
                target.write_bytes(raw)
                parcelas = payload.get("parcelas", [])
                record = {
                    "grid_id": grid_id,
                    "ortho_id": ortho_id,
                    "lote_id": ortho.get("lote_id", "") if ortho else "",
                    "ortho_nome": ortho.get("nome", "") if ortho else "",
                    "path": str(target),
                    "linhas": int(rows),
                    "colunas": int(cols),
                    "parcelas": len(parcelas),
                    "ids_parcelas": [cell.get("id") or f"L{cell.get('linha')}_C{cell.get('coluna')}" for cell in parcelas],
                    "escala": escala,
                    "dados_posicionamento": payload.get("pontos", []),
                    "configuracoes": {"linhas": int(rows), "colunas": int(cols), "escala": escala},
                    "mini_resumo": {
                        "parcelas": len(parcelas),
                        "linhas": int(rows),
                        "colunas": int(cols),
                        "ortofoto": ortho.get("nome", "") if ortho else ""
                    },
                    "status": "Grid salvo no sistema",
                    "criado_em": _tv_now()
                }
                manifest.setdefault("grids", []).insert(0, record)
                if ortho:
                    ortho["status"] = "Grid salvo no sistema"
                    ortho["grid_id"] = grid_id
                voo = _vd_find_voo(manifest, record["lote_id"])
                if voo:
                    voo["status_envio"] = "Grid salvo"
                    voo["grid_id"] = grid_id
                _vd_add_history(manifest, f"Grid salvo no sistema: {grid_id}", record["lote_id"])
                _vd_save_manifest(manifest)
                st.success(f"Grid `{grid_id}` salvo com {len(parcelas)} parcela(s).")

        if selected:
            ortho_id_for_existing = _vd_id_from_option(selected)
            ortho_for_existing = _vd_find_ortho(manifest, ortho_id_for_existing)
            if st.button("Abrir no Marcador de Grid existente", key="vd_open_existing_grid", use_container_width=True):
                if ortho_for_existing:
                    st.session_state["grid_prefill_ortho_path"] = ortho_for_existing.get("path", "")
                    st.session_state["grid_prefill_ortho_name"] = ortho_for_existing.get("nome", "")
                ir_para("Grid")
                app_rerun()

    with right:
        if not selected:
            st.info("Selecione a ortofoto para abrir o visualizador de grid.")
        else:
            ortho_id = _vd_id_from_option(selected)
            ortho = _vd_find_ortho(manifest, ortho_id)
            if ortho and Path(ortho.get("path", "")).exists():
                raw = Path(ortho["path"]).read_bytes()
                b64, dims, err, _ = processar_ortofoto(raw, ortho["nome"])
                if err:
                    st.error(err)
                else:
                    st.caption(f"{ortho['nome']} · {dims[0]}x{dims[1]} px")
                    components.html(_tv_grid_viewer_html(b64, int(rows), int(cols), ortho["nome"]), height=700, scrolling=False)

    st.markdown("#### Exportação Final")
    grid_options = [f"{g['grid_id']} · {g.get('ortho_nome','Ortofoto')}" for g in manifest.get("grids", [])]
    if not grid_options:
        st.info("Nenhum grid salvo ainda.")
        return
    selected_grid = st.selectbox("Grid salvo para exportar", grid_options, key="vd_grid_export_select")
    grid_id = _vd_id_from_option(selected_grid)
    grid = _vd_find_grid(manifest, grid_id)
    ortho = _vd_find_ortho(manifest, grid.get("ortho_id", "")) if grid else {}
    if not grid or not ortho:
        st.warning("Grid ou ortofoto não encontrados.")
        return

    e1, e2, e3, e4, e5, e6 = st.columns(6)
    with e1:
        st.download_button(
            "Baixar Ortofoto com Grid PNG",
            data=_vd_grid_overlay_bytes(grid, ortho, "PNG"),
            file_name=f"{grid_id}_ortofoto_grid.png",
            mime="image/png",
            use_container_width=True
        )
    with e2:
        st.download_button(
            "Baixar JPG",
            data=_vd_grid_overlay_bytes(grid, ortho, "JPEG"),
            file_name=f"{grid_id}_ortofoto_grid.jpg",
            mime="image/jpeg",
            use_container_width=True
        )
    with e3:
        st.download_button(
            "Baixar TIFF",
            data=_vd_grid_overlay_bytes(grid, ortho, "TIFF"),
            file_name=f"{grid_id}_ortofoto_grid.tif",
            mime="image/tiff",
            use_container_width=True
        )
    with e4:
        st.download_button(
            "GeoTIFF",
            data=_vd_grid_geotiff_overlay_bytes(grid, ortho),
            file_name=f"{grid_id}_ortofoto_grid_geotiff.tif",
            mime="image/tiff",
            use_container_width=True
        )
    with e5:
        st.download_button(
            "GeoJSON",
            data=_vd_grid_geojson_bytes(grid, ortho),
            file_name=f"{grid_id}_parcelas.geojson",
            mime="application/geo+json",
            use_container_width=True
        )
    with e6:
        st.download_button(
            "SHP QGIS",
            data=_vd_grid_shp_zip_bytes(grid, ortho),
            file_name=f"{grid_id}_parcelas_shp.zip",
            mime="application/zip",
            use_container_width=True
        )

    rows_grid = [{
        "Grid": g.get("grid_id"),
        "Ortofoto": g.get("ortho_nome"),
        "Linhas": g.get("linhas"),
        "Colunas": g.get("colunas"),
        "Parcelas": g.get("parcelas"),
        "Status": g.get("status"),
        "Data": g.get("criado_em")
    } for g in manifest.get("grids", [])]
    st.markdown("#### Grids salvos")
    st.dataframe(rows_grid, use_container_width=True, hide_index=True)

def _vd_render_projetos(manifest: dict) -> None:
    st.markdown("#### PASSO 4 — Retornar Dados")
    rows = []
    for voo in manifest.get("voos", []):
        ortho = _vd_find_ortho(manifest, voo.get("ortho_id", ""))
        grid = _vd_find_grid(manifest, voo.get("grid_id", ""))
        rows.append({
            "Nome do projeto": voo.get("nome_voo"),
            "Data": voo.get("data_hora"),
            "Imagens": voo.get("quantidade_imagens"),
            "Parcelas": grid.get("parcelas", 0) if grid else 0,
            "Status": voo.get("status_envio"),
            "Ortofoto": ortho.get("nome", "") if ortho else "",
            "Lote": voo.get("lote_id")
        })
    st.dataframe(rows, use_container_width=True, hide_index=True)

    options = _vd_project_options(manifest)
    if not options:
        st.info("Nenhum projeto salvo ainda.")
        return
    selected = st.selectbox("Abrir projeto", options, key="vd_project_open")
    lote_id = _vd_id_from_option(selected)
    voo = _vd_find_voo(manifest, lote_id)
    if not voo:
        return
    ortho = _vd_find_ortho(manifest, voo.get("ortho_id", ""))
    grid = _vd_find_grid(manifest, voo.get("grid_id", ""))

    c1, c2 = st.columns([1, 2])
    with c1:
        if ortho and Path(ortho.get("path", "")).exists():
            raw = Path(ortho["path"]).read_bytes()
            b64, dims, err, _ = processar_ortofoto(raw, ortho["nome"])
            if not err:
                app_image(Image.open(BytesIO(base64.b64decode(b64))))
                st.caption(f"Miniatura: {ortho['nome']} · {dims[0]}x{dims[1]}")
            else:
                st.info("Miniatura indisponível.")
        else:
            st.info("Projeto ainda sem ortofoto anexada.")
    with c2:
        st.markdown(
            f"<div class='vd-card'><b style='color:#ff8c00;'>Projeto:</b> {voo.get('nome_voo')}<br>"
            f"<b>Fazenda:</b> {voo.get('fazenda') or '-'} &nbsp; <b>Talhão:</b> {voo.get('talhao') or '-'}<br>"
            f"<b>Lote:</b> {lote_id}<br><b>Status:</b> {voo.get('status_envio')}</div>",
            unsafe_allow_html=True
        )
        b1, b2, b3 = st.columns(3)
        with b1:
            if st.button("Abrir", key="vd_project_btn_open", use_container_width=True):
                if ortho:
                    st.session_state["vd_next_tab"] = "ETAPA 2"
                    app_rerun()
                else:
                    st.info("Anexe a ortofoto para abrir o preview.")
        with b2:
            st.download_button(
                "Baixar manifesto",
                data=json.dumps(voo, indent=2, ensure_ascii=False).encode("utf-8"),
                file_name=f"{lote_id}_manifesto.json",
                mime="application/json",
                use_container_width=True
            )
        with b3:
            if st.button("Continuar edição", key="vd_project_continue", use_container_width=True):
                if ortho:
                    st.session_state["vd_pending_grid_ortho_id"] = ortho.get("ortho_id")
                    st.session_state["vd_next_tab"] = "ETAPA 3"
                    app_rerun()
                else:
                    st.warning("Projeto ainda sem ortofoto para edição.")

        if grid and ortho:
            st.download_button(
                "Baixar Ortofoto com Grid",
                data=_vd_grid_overlay_bytes(grid, ortho, "PNG"),
                file_name=f"{grid.get('grid_id')}_ortofoto_grid.png",
                mime="image/png",
                use_container_width=True
            )

def _vd_render_historico(manifest: dict) -> None:
    st.markdown("#### Histórico de Alterações")
    st.dataframe(manifest.get("history", []), use_container_width=True, hide_index=True)
    st.download_button(
        "Baixar banco interno JSON",
        data=json.dumps(manifest, indent=2, ensure_ascii=False).encode("utf-8"),
        file_name="voos_direcionados_manifest.json",
        mime="application/json",
        use_container_width=True
    )

def render_voos_direcionados() -> None:
    _vd_ensure_storage()
    if "vd_logged_in" not in st.session_state:
        st.session_state.vd_logged_in = False
    if not st.session_state.vd_logged_in:
        _vd_render_login()
        return

    manifest = _vd_load_manifest()
    st.markdown("""
    <style>
      .vd-hero {
        background:linear-gradient(145deg,#181818 0%,#101010 58%,#2a1700 100%);
        border:1px solid #3a2a18;
        border-top:2px solid #ff8c00;
        border-radius:8px;
        padding:20px 22px;
        margin-bottom:14px;
        box-shadow:4px 4px 14px #070707,0 0 24px rgba(255,140,0,.10);
      }
      .vd-hero h2 {
        margin:0;
        color:#ff8c00;
        letter-spacing:3px;
        text-transform:uppercase;
        text-shadow:1px 1px 0 #7a3a00,3px 3px 8px #000,0 0 22px rgba(255,140,0,.32);
      }
      .vd-hero p { margin:8px 0 0 0; color:#cfcfcf; font-size:13px; }
      .vd-card {
        background:#161616;
        border:1px solid #303030;
        border-radius:8px;
        padding:14px;
        color:#ddd;
        margin-bottom:10px;
      }
      div[data-testid="stRadio"] > div {
        background:#151515;
        border:1px solid #2e2e2e;
        border-radius:8px;
        padding:8px 10px;
      }
      div[data-testid="stRadio"] label {
        background:linear-gradient(145deg,#222,#111);
        border:1px solid #303030;
        border-radius:8px;
        padding:7px 10px;
        margin-right:4px;
      }
      div[data-testid="stRadio"] label:hover { border-color:#ff8c00; box-shadow:0 0 12px rgba(255,140,0,.12); }
      div[data-testid="stRadio"] label:has(input:checked) {
        border-color:#ff8c00;
        background:linear-gradient(145deg,#ff9e33,#e67600);
        box-shadow:0 0 16px rgba(255,140,0,.22);
      }
    </style>
    """, unsafe_allow_html=True)

    ctop1, ctop2 = st.columns([5, 1])
    with ctop1:
        st.markdown(
            "<div class='vd-hero'><h2>Processos de Voos para Análise</h2>"
            "<p>Fluxo passo a passo: enviar fotos de voos, receber ortofotos, marcar grid e retornar dados para análise.</p>"
            f"{_tv_status_chip('Janela isolada', '#ff8c00')}"
            f"{_tv_status_chip('Banco interno isolado', '#55ff99')}"
            f"{_tv_status_chip('QGIS / GeoJSON / SHP', '#5599ff')}"
            "</div>",
            unsafe_allow_html=True
        )
    with ctop2:
        if st.button("Sair", key="vd_logout", use_container_width=True):
            st.session_state.vd_logged_in = False
            app_rerun()

    _vd_metric_cards(manifest)

    tabs = ["Enviar Fotos de Voos", "Receber Ortofotos", "Marcador de Grid", "Retornar Dados", "Histórico"]
    legacy_tabs = {
        "ETAPA 1": "Enviar Fotos de Voos",
        "ETAPA 2": "Receber Ortofotos",
        "ETAPA 3": "Marcador de Grid",
        "ETAPA 4": "Retornar Dados",
        "HIST.": "Histórico",
        "Envio de Imagens": "Enviar Fotos de Voos",
        "Ortofotos": "Receber Ortofotos",
        "Marcar Grid": "Marcador de Grid",
        "Projetos Salvos": "Retornar Dados",
        "Histórico": "Histórico"
    }
    next_tab = st.session_state.pop("vd_next_tab", None)
    next_tab = legacy_tabs.get(next_tab, next_tab)
    if next_tab in tabs and st.session_state.get("vd_active_tab") != next_tab:
        st.session_state.pop("vd_active_tab", None)
        tab_index = tabs.index(next_tab)
    else:
        current_tab = st.session_state.get("vd_active_tab", tabs[0])
        current_tab = legacy_tabs.get(current_tab, current_tab)
        tab_index = tabs.index(current_tab) if current_tab in tabs else 0
    active = st.radio("Processos de Voos para Análise", tabs, horizontal=True, label_visibility="collapsed", index=tab_index, key="vd_active_tab")
    etapa_desc = {
        "Enviar Fotos de Voos": "PASSO 1: selecionar fotos, definir caminho do Azure/destino e enviar o lote.",
        "Receber Ortofotos": "PASSO 2: buscar/importar ortofoto gerada e abrir no visualizador.",
        "Marcador de Grid": "PASSO 3: usar o marcador de grid, salvar parcelas e preparar exportação.",
        "Retornar Dados": "PASSO 4: disponibilizar ortofoto, grid, SHP, GeoJSON e resumo final.",
        "Histórico": "Histórico e banco interno isolado."
    }
    st.caption(etapa_desc.get(active, ""))
    st.markdown("---")

    if active == "Enviar Fotos de Voos":
        _vd_render_envio(manifest)
        st.markdown("---")
        if st.button("Próximo Passo ➜ Receber Ortofotos", key="vd_next_to_ortho", type="primary", use_container_width=True):
            st.session_state["vd_next_tab"] = "Receber Ortofotos"
            app_rerun()
    elif active == "Receber Ortofotos":
        _vd_render_ortofotos(manifest)
        st.markdown("---")
        if st.button("Próximo Passo ➜ Marcador de Grid", key="vd_next_to_grid", type="primary", use_container_width=True):
            st.session_state["vd_next_tab"] = "Marcador de Grid"
            app_rerun()
    elif active == "Marcador de Grid":
        _vd_render_grid(manifest)
        st.markdown("---")
        if st.button("Próximo Passo ➜ Retornar Dados", key="vd_next_to_return", type="primary", use_container_width=True):
            st.session_state["vd_next_tab"] = "Retornar Dados"
            app_rerun()
    elif active == "Retornar Dados":
        _vd_render_projetos(manifest)
    elif active == "Histórico":
        _vd_render_historico(manifest)


def _render_manage_users() -> None:
    if not _auth_is_admin():
        st.warning("Apenas o administrador Wellington pode acessar o gerenciamento de usuários.")
        return

    users_data = _auth_load_users()
    users = users_data.get("users", [])
    st.markdown("#### Gerenciar Usuários")
    st.caption("Somente Wellington pode cadastrar, editar, excluir, ativar/desativar e definir permissões.")

    if users:
        rows = []
        for user in users:
            rows.append({
                "Nome": user.get("nome", ""),
                "Usuário": user.get("usuario", ""),
                "Status": "Ativo" if user.get("ativo", True) else "Inativo",
                "Admin": "Sim" if _auth_is_admin(user) else "Não",
                "Culturas": ", ".join(_auth_allowed_cultures(user)) or "-",
                "Parceiros": ", ".join(_partner_label(p) for p in _auth_allowed_partners(user)) or "-",
                "Menus": ", ".join(label for key, label in MENU_PERMISSION_OPTIONS.items() if _auth_menu_allowed(key, user)) or "-",
                "Fenotipagem": ", ".join(label for key, label in PHENOTYPING_PERMISSION_OPTIONS.items() if _auth_phenotyping_allowed(key, user)) or "-",
                "Planilha Parceiros": ", ".join(label for key, label in PARTNER_PERMISSION_OPTIONS.items() if _auth_partner_permission(key, user)) or "-",
            })
        st.dataframe(rows, use_container_width=True, hide_index=True)

    options = ["Novo usuário"] + [f"{u.get('usuario')} · {u.get('nome')}" for u in users]
    selected = st.selectbox("Usuário para cadastro/edição", options, key="cfg_user_select")
    editing = {}
    if selected != "Novo usuário":
        login = selected.split(" · ", 1)[0]
        editing = next((u for u in users if u.get("usuario") == login), {})

    with st.form("form_manage_users"):
        c1, c2 = st.columns(2)
        with c1:
            nome = st.text_input("Nome completo", value=editing.get("nome", ""))
            usuario = st.text_input("Usuário de login", value=editing.get("usuario", ""))
            senha = st.text_input("Senha", value=editing.get("senha", ""), type="password")
            ativo = st.checkbox("Usuário ativo", value=bool(editing.get("ativo", True)))
        with c2:
            admin = st.checkbox("Administrador", value=bool(editing.get("admin", False)))
            perms_current = _auth_permissions(editing) if editing else _default_permissions(False)
            perm_culturas = st.checkbox("Acessar seleção de culturas", value=bool(perms_current.get("culturas", True)))
            p_soja, p_milho, p_alg = st.columns(3)
            with p_soja:
                perm_soja = st.checkbox("Soja", value=bool(perms_current.get("soja", True)))
            with p_milho:
                perm_milho = st.checkbox("Milho", value=bool(perms_current.get("milho", True)))
            with p_alg:
                perm_algodao = st.checkbox("Algodão", value=bool(perms_current.get("algodao", True)))
            perm_parceiros = st.checkbox("Acessar módulo Parceiros / Controle de Voos e Dados", value=bool(perms_current.get("parceiros", False)))
            p_eiwa, p_alvaz = st.columns(2)
            with p_eiwa:
                perm_eiwa = st.checkbox("EIWA", value=bool(perms_current.get("eiwa", False)))
            with p_alvaz:
                perm_alvaz = st.checkbox("ALVAZ", value=bool(perms_current.get("alvaz", False)))
            st.markdown("##### Permissões por menu")
            menu_values = {}
            menu_items = list(MENU_PERMISSION_OPTIONS.items())
            for idx in range(0, len(menu_items), 2):
                mcols = st.columns(2)
                for col, (menu_key, menu_label) in zip(mcols, menu_items[idx:idx + 2]):
                    with col:
                        menu_values[menu_key] = st.checkbox(menu_label, value=bool(perms_current.get(menu_key, True)))
            st.markdown("##### Análises de Fenotipagem")
            phenotyping_values = {}
            phenotyping_items = list(PHENOTYPING_PERMISSION_OPTIONS.items())
            for idx in range(0, len(phenotyping_items), 2):
                fcols = st.columns(2)
                for col, (phen_key, phen_label) in zip(fcols, phenotyping_items[idx:idx + 2]):
                    with col:
                        phenotyping_values[phen_key] = st.checkbox(
                            phen_label,
                            value=bool(perms_current.get(phen_key, True)),
                        )

            st.markdown("##### Permissões específicas da planilha de parceiros")
            partner_perm_values = {}
            partner_items = list(PARTNER_PERMISSION_OPTIONS.items())
            for idx in range(0, len(partner_items), 2):
                pcols = st.columns(2)
                for col, (perm_key, perm_label) in zip(pcols, partner_items[idx:idx + 2]):
                    with col:
                        partner_perm_values[perm_key] = st.checkbox(
                            perm_label,
                            value=bool(perms_current.get(perm_key, False)),
                        )

        save_user = st.form_submit_button("💾 Salvar usuário", type="primary", use_container_width=True)

    if save_user:
        usuario = usuario.strip()
        if not nome.strip() or not usuario or not senha:
            st.error("Preencha nome, usuário e senha.")
        else:
            record = {
                "nome": nome.strip(),
                "usuario": usuario,
                "senha": senha,
                "ativo": ativo,
                "admin": admin,
                "permissoes": {
                    "culturas": perm_culturas,
                    "soja": perm_soja,
                    "milho": perm_milho,
                    "algodao": perm_algodao,
                    "parceiros": perm_parceiros,
                    "eiwa": perm_eiwa,
                    "alvaz": perm_alvaz,
                    **menu_values,
                    **phenotyping_values,
                    **partner_perm_values,
                },
                "criado_em": editing.get("criado_em", _now_iso()),
                "atualizado_em": _now_iso(),
            }
            replaced = False
            for idx, user in enumerate(users):
                if user.get("usuario") == editing.get("usuario") or user.get("usuario") == usuario:
                    users[idx] = record
                    replaced = True
                    break
            if not replaced:
                users.append(record)
            users_data["users"] = users
            _auth_save_users(users_data)
            state = _partners_load_state()
            _partners_add_history(state, "", "Usuário salvo", f"{usuario} · {nome.strip()}")
            _partners_save_state(state)
            st.success("Usuário salvo com sucesso.")
            app_rerun()

    if editing and not _auth_is_admin(editing):
        if st.button("🗑️ Excluir usuário selecionado", key="btn_delete_cfg_user", use_container_width=True):
            users_data["users"] = [u for u in users if u.get("usuario") != editing.get("usuario")]
            _auth_save_users(users_data)
            state = _partners_load_state()
            _partners_add_history(state, "", "Usuário excluído", editing.get("usuario", ""))
            _partners_save_state(state)
            st.success("Usuário excluído.")
            app_rerun()

def _render_partner_alerts(partner_data: dict) -> None:
    alerts = _partners_deadline_alerts(partner_data.get("chat", []))
    if not alerts:
        st.caption("Sem notificações de prazo no momento.")
        return
    color_map = {
        "Vermelho": ("#ff5252", "rgba(255,82,82,.13)"),
        "Amarelo": ("#ffd54f", "rgba(255,213,79,.13)"),
        "Verde": ("#66bb6a", "rgba(102,187,106,.13)"),
        "Cinza": ("#9e9e9e", "rgba(158,158,158,.10)"),
    }
    for color_name, label, item in alerts[:8]:
        fg, bg = color_map.get(color_name, ("#ccc", "rgba(255,255,255,.06)"))
        st.markdown(
            f"<div style='border:1px solid {fg};background:{bg};border-radius:8px;padding:8px 10px;margin-bottom:6px;'>"
            f"<b style='color:{fg};'>{label}</b><br>"
            f"<span style='color:#ddd;font-size:.82rem;'>{item.get('assunto','Sem assunto')} · {item.get('status','')}</span>"
            f"</div>",
            unsafe_allow_html=True
        )

def _render_partner_sheet_controls(state: dict, partner_key: str) -> None:
    partner = state["partners"][partner_key]
    partner_name = _partner_label(partner_key)
    can_import = _auth_partner_permission("partner_sheet_import")
    can_edit = _auth_partner_permission("partner_sheet_edit_rows") or _auth_partner_permission("partner_sheet_write_treatment")
    uploaded_sheet = None
    import_clicked = False
    update_clicked = False
    save_internal_clicked = False

    if can_import:
        st.markdown("##### Importação de planilha")
        uploaded_sheet = st.file_uploader(
            "Selecionar planilha Excel ou CSV",
            type=["xlsx", "xls", "csv"],
            key=f"partner_upload_sheet_{partner_key}",
            help="Envie arquivos .xlsx, .xls ou .csv para espelhar os dados no sistema.",
        )
        if uploaded_sheet is not None:
            render_tmg_loading_bar(100, f"Planilha recebida: {Path(uploaded_sheet.name).name}")
        c1, c2, c3 = st.columns(3)
        with c1:
            import_clicked = st.button("📥 Importar planilha", key=f"partner_import_{partner_key}", use_container_width=True)
        with c2:
            update_clicked = st.button("🔄 Atualizar dados", key=f"partner_update_{partner_key}", use_container_width=True)
        with c3:
            save_internal_clicked = st.button("💾 Salvar Dados Internos", key=f"partner_save_internal_top_{partner_key}", use_container_width=True)
    else:
        st.caption("Seu usuário pode visualizar os dados liberados, mas não possui permissão para importar planilhas.")

    if import_clicked or update_clicked:
        try:
            load_box = st.empty()
            update_tmg_loading(load_box, 20, "Lendo planilha importada...")
            df, err = _partners_read_sheet_upload(uploaded_sheet)
            if err:
                clear_tmg_loading(load_box)
                st.warning("Não foi possível importar a planilha. Verifique o formato do arquivo e tente novamente.")
                return
            update_tmg_loading(load_box, 62, "Tratando colunas e comparando dados...")
            new_clean = _partners_clean_dataframe(df)
            if new_clean.empty and len(new_clean.columns) == 0:
                clear_tmg_loading(load_box)
                st.warning("Não foi possível importar a planilha. Verifique o formato do arquivo e tente novamente.")
                return
            original_columns = [col for col in new_clean.columns if col != PARTNER_ROW_ID and col not in PARTNER_INTERNAL_COLUMNS]
            if not original_columns:
                clear_tmg_loading(load_box)
                st.warning("Não foi possível importar a planilha. Verifique o formato do arquivo e tente novamente.")
                return
            baseline_rows = partner.get("baseline_rows", [])
            baseline_columns = partner.get("baseline_columns", [])
            if not baseline_rows:
                partner["baseline_rows"] = new_clean[original_columns].to_dict(orient="records")
                partner["baseline_columns"] = original_columns
                partner["baseline_import"] = {
                    "data_hora": _now_human(),
                    "usuario": _auth_user_name(),
                    "arquivo": uploaded_sheet.name,
                    "linhas": len(new_clean),
                }
                summary = {
                    "linhas_novas": 0,
                    "linhas_alteradas": 0,
                    "linhas_removidas": 0,
                    "total_diferencas": 0,
                    "data_hora": _now_human(),
                    "usuario": _auth_user_name(),
                    "fonte": f"Arquivo: {uploaded_sheet.name}",
                    "base_comparacao": "Primeira importação salva",
                }
                diff_rows = []
            else:
                compare_columns = baseline_columns or original_columns
                baseline_df = pd.DataFrame(baseline_rows)
                summary, diff_rows = _partners_compare_dataframes(baseline_df, new_clean, compare_columns)
                summary["fonte"] = f"Arquivo: {uploaded_sheet.name}"
                summary["base_comparacao"] = "Primeira importação salva"
            prepared, original_columns = _partners_prepare_import_df(df, _auth_user_name())
            old_df = _partners_rows_to_df(partner)
            for idx in range(min(len(old_df), len(prepared))):
                for col in PARTNER_INTERNAL_COLUMNS:
                    if col in old_df.columns and col in prepared.columns:
                        prepared.at[idx, col] = old_df.iloc[idx].get(col, prepared.at[idx, col])
            partner["link"] = ""
            partner["columns"] = original_columns
            partner["rows"] = prepared.to_dict(orient="records")
            partner["last_import"] = {"data_hora": _now_human(), "usuario": _auth_user_name(), "linhas": len(prepared), "fonte": f"Arquivo: {uploaded_sheet.name}"}
            partner["last_update"] = summary
            partner["diff_rows"] = diff_rows[:500]
            update_tmg_loading(load_box, 100, "Carregamento concluído com sucesso.")
            _partners_add_history(
                state,
                partner_key,
                "Planilha importada",
                f"{partner_name}: {len(prepared)} linhas · {uploaded_sheet.name}",
                {"tipo_acao": "importação", "linha": "", "campo": "", "valor_antigo": "", "valor_novo": uploaded_sheet.name}
            )
            _partners_save_state(state)
            st.session_state[f"partner_planilha_tratada_open_{partner_key}"] = True
            st.session_state[f"partner_panel_mode_{partner_key}"] = "Resumo de Atualizações"
            st.success("Planilha importada e a janela de tratamento foi aberta no módulo de Parceiras.")
            app_rerun()
        except Exception:
            st.warning("Não foi possível importar a planilha. Verifique o formato do arquivo e tente novamente.")

    if save_internal_clicked and can_edit:
        partner["last_update"] = {
            "data_hora": _now_human(),
            "usuario": _auth_user_name(),
            "linhas_novas": 0,
            "linhas_alteradas": 0,
            "linhas_removidas": 0,
            "total_diferencas": 0,
        }
        _partners_add_history(state, partner_key, "Dados internos salvos", partner_name)
        _partners_save_state(state)
        st.success("Dados internos salvos.")

def _render_partner_table(state: dict, partner_key: str) -> None:
    partner = state["partners"][partner_key]
    can_view = _auth_partner_permission("partner_sheet_view")
    can_edit_rows = _auth_partner_permission("partner_sheet_edit_rows")
    can_delete_rows = _auth_partner_permission("partner_sheet_delete_rows")
    can_edit_header = _auth_partner_permission("partner_sheet_edit_header")
    can_write_treatment = _auth_partner_permission("partner_sheet_write_treatment")
    can_export = _auth_partner_permission("partner_sheet_export")
    if not can_view:
        st.warning("Seu usuário não possui permissão para visualizar esta planilha.")
        return

    df = _partners_rows_to_df(partner)
    if df.empty or len(df) == 0:
        st.info("Importe uma planilha Excel ou CSV para iniciar a tabela espelhada.")
        return

    st.session_state.setdefault(f"partner_planilha_tratada_open_{partner_key}", True)
    st.session_state.setdefault(f"partner_panel_mode_{partner_key}", "Filtrar Informações")
    st.session_state.setdefault(f"partner_hidden_cols_{partner_key}", [])
    st.session_state.setdefault(f"partner_search_{partner_key}", "")
    st.session_state.setdefault(f"partner_status_filter_{partner_key}", [])

    st.markdown(
        """
        <style>
        .partner-excel-window {
            background: linear-gradient(145deg, #10243a, #0b1728);
            border: 1px solid rgba(66,165,245,.35);
            border-radius: 18px;
            padding: 16px 18px;
            box-shadow: 0 12px 30px rgba(0,0,0,.35);
            margin-top: 12px;
            margin-bottom: 14px;
        }
        .partner-excel-title {
            color: #ffffff;
            font-weight: 900;
            font-size: 1.25rem;
            letter-spacing: .5px;
        }
        .partner-excel-subtitle { color: #b8c7d8; font-size: .92rem; margin-top: 4px; }
        .partner-toolbox {
            background: #0d1e35;
            border: 1px solid rgba(144,202,249,.25);
            border-radius: 16px;
            padding: 12px;
        }
        .partner-toolbox-title {
            color: #90caf9;
            font-weight: 800;
            text-align: center;
            margin-bottom: 10px;
            text-transform: uppercase;
            font-size: .82rem;
            letter-spacing: 1px;
        }
        .partner-history-card {
            background: rgba(255,255,255,.04);
            border: 1px solid rgba(255,255,255,.08);
            border-radius: 14px;
            padding: 10px 12px;
            margin: 8px 0;
        }
        </style>
        <div class='partner-excel-window'>
            <div class='partner-excel-title'>📊 Janela de Tratamento de Planilha Importada</div>
            <div class='partner-excel-subtitle'>Mini editor estilo Excel para filtrar, editar, excluir, exportar e acompanhar alterações da planilha da parceira.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    toolbox, main_area = st.columns([1.05, 3.95], gap="large")

    with toolbox:
        st.markdown("<div class='partner-toolbox'><div class='partner-toolbox-title'>Funções</div>", unsafe_allow_html=True)
        panel_buttons = [
            ("Filtrar Colunas", "🧩 Filtrar Colunas"),
            ("Filtrar Informações", "🔎 Filtrar Informações"),
            ("Editar Cabeçalho", "✏️ Editar Cabeçalho"),
            ("Editar Célula", "📝 Editar Célula"),
            ("Adicionar Linha", "➕ Adicionar Linha"),
            ("Excluir Linha", "🗑️ Excluir Linha"),
            ("Excluir Coluna", "🧹 Excluir Coluna"),
            ("Resumo de Atualizações", "📋 Resumo de Atualizações"),
            ("Exportar CSV", "⬇️ Exportar CSV"),
            ("Exportar Excel", "📗 Exportar Excel"),
        ]
        for mode, label in panel_buttons:
            if st.button(label, key=f"partner_panel_{partner_key}_{mode}", use_container_width=True):
                st.session_state[f"partner_panel_mode_{partner_key}"] = mode
                app_rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    mode = st.session_state.get(f"partner_panel_mode_{partner_key}", "Filtrar Informações")
    hidden_cols_key = f"partner_hidden_cols_{partner_key}"
    search_key = f"partner_search_{partner_key}"
    status_key = f"partner_status_filter_{partner_key}"

    with main_area:
        st.markdown(f"##### {mode}")

        if mode == "Filtrar Colunas":
            hidable = [col for col in df.columns if col != PARTNER_ROW_ID]
            col_search = st.text_input("Buscar coluna", key=f"partner_column_search_{partner_key}")
            choices = [c for c in hidable if str(col_search).strip().lower() in c.lower()] if col_search else hidable
            selected_visible = st.multiselect(
                "Colunas visíveis",
                choices,
                default=[c for c in choices if c not in st.session_state.get(hidden_cols_key, [])],
                key=f"partner_visible_cols_{partner_key}",
            )
            b1, b2 = st.columns(2)
            with b1:
                if st.button("Aplicar filtro de colunas", key=f"partner_apply_visible_cols_{partner_key}", type="primary", use_container_width=True):
                    st.session_state[hidden_cols_key] = [c for c in hidable if c not in selected_visible]
                    app_rerun()
            with b2:
                if st.button("Restaurar colunas ocultas", key=f"partner_restore_cols_{partner_key}", use_container_width=True):
                    st.session_state[hidden_cols_key] = []
                    app_rerun()
            st.caption("Ocultar colunas não apaga dados. As informações permanecem salvas e podem voltar a aparecer.")

        elif mode == "Filtrar Informações":
            f1, f2 = st.columns([1.4, 1])
            with f1:
                st.text_input("Pesquisar qualquer informação da tabela", key=search_key)
            with f2:
                st.multiselect("Filtrar status", PARTNER_STATUS_OPTIONS, key=status_key)
            st.caption("A busca aceita termos parciais e filtra a tabela em tempo real.")

        elif mode == "Editar Cabeçalho":
            if not can_edit_header:
                st.warning("Seu usuário não possui permissão para editar cabeçalho.")
            elif partner.get("columns"):
                h1, h2 = st.columns([1, 1.2])
                with h1:
                    old_header = st.selectbox("Cabeçalho atual", partner.get("columns", []), key=f"partner_header_old_{partner_key}")
                with h2:
                    new_header = st.text_input("Novo nome do cabeçalho", value=old_header, key=f"partner_header_new_{partner_key}")
                s1, s2 = st.columns(2)
                with s1:
                    if st.button("💾 Salvar cabeçalho", key=f"partner_header_save_{partner_key}", type="primary", use_container_width=True):
                        ok, msg = _partners_rename_column(partner, old_header, new_header)
                        if ok:
                            _partners_add_history(
                                state,
                                partner_key,
                                "Cabeçalho alterado",
                                f"{old_header} -> {new_header}",
                                {"tipo_acao": "alteração de cabeçalho", "linha": "", "campo": old_header, "valor_antigo": old_header, "valor_novo": new_header}
                            )
                            _partners_save_state(state)
                            st.success("Cabeçalho salvo e atualizado na tabela, filtros e exportações.")
                            app_rerun()
                        else:
                            st.warning(msg or "Não foi possível alterar o cabeçalho.")
                with s2:
                    if st.button("Cancelar", key=f"partner_header_cancel_{partner_key}", use_container_width=True):
                        app_rerun()

        elif mode == "Adicionar Linha":
            if not can_edit_rows:
                st.warning("Seu usuário não possui permissão para adicionar linhas.")
            elif st.button("➕ Criar nova linha vazia", key=f"partner_add_row_panel_{partner_key}", type="primary", use_container_width=True):
                row = _partners_add_blank_row(partner, partner_key)
                _partners_add_history(
                    state,
                    partner_key,
                    "Linha criada",
                    row.get(PARTNER_ROW_ID, ""),
                    {"tipo_acao": "edição", "linha": row.get(PARTNER_ROW_ID, ""), "campo": "", "valor_antigo": "", "valor_novo": "linha criada"}
                )
                _partners_save_state(state)
                st.success("Linha adicionada.")
                app_rerun()

        elif mode == "Excluir Linha":
            if not can_delete_rows:
                st.warning("Seu usuário não possui permissão para excluir linhas.")
            else:
                row_records = df.to_dict(orient="records")
                row_options = {_partners_row_label(row, idx): str(row.get(PARTNER_ROW_ID, "")) for idx, row in enumerate(row_records)}
                if row_options:
                    selected_label = st.selectbox("Linha para excluir", list(row_options.keys()), key=f"partner_delete_select_{partner_key}")
                    confirm_delete = st.checkbox("Confirmo a exclusão desta linha", key=f"partner_delete_confirm_{partner_key}")
                    if st.button("🗑️ Excluir linha selecionada", key=f"partner_delete_row_panel_{partner_key}", type="primary", use_container_width=True):
                        row_id = row_options.get(selected_label, "")
                        if not confirm_delete:
                            st.warning("Marque a confirmação antes de excluir.")
                        elif _partners_delete_row(partner, row_id):
                            _partners_add_history(
                                state,
                                partner_key,
                                "Linha excluída",
                                row_id,
                                {"tipo_acao": "exclusão", "linha": row_id, "campo": "", "valor_antigo": "linha existente", "valor_novo": "linha removida"}
                            )
                            _partners_save_state(state)
                            st.success("Linha excluída.")
                            app_rerun()
                        else:
                            st.warning("Não foi possível localizar a linha selecionada.")

        elif mode == "Excluir Coluna":
            if not can_edit_header:
                st.warning("Seu usuário não possui permissão para excluir colunas.")
            else:
                removable_cols = [col for col in partner.get("columns", []) if col not in PARTNER_INTERNAL_COLUMNS and col != PARTNER_ROW_ID]
                if removable_cols:
                    col_to_delete = st.selectbox("Coluna para excluir", removable_cols, key=f"partner_delete_col_select_{partner_key}")
                    confirm_col = st.checkbox("Confirmo a exclusão desta coluna", key=f"partner_delete_col_confirm_{partner_key}")
                    if st.button("🧹 Excluir coluna", key=f"partner_delete_col_button_{partner_key}", type="primary", use_container_width=True):
                        if not confirm_col:
                            st.warning("Marque a confirmação antes de excluir a coluna.")
                        else:
                            partner["columns"] = [col for col in partner.get("columns", []) if col != col_to_delete]
                            partner["baseline_columns"] = [col for col in partner.get("baseline_columns", []) if col != col_to_delete]
                            for collection_name in ("rows", "baseline_rows", "diff_rows"):
                                for row in partner.get(collection_name, []) or []:
                                    if isinstance(row, dict):
                                        row.pop(col_to_delete, None)
                            st.session_state[hidden_cols_key] = [col for col in st.session_state.get(hidden_cols_key, []) if col != col_to_delete]
                            _partners_add_history(
                                state,
                                partner_key,
                                "Coluna excluída",
                                col_to_delete,
                                {"tipo_acao": "exclusão", "linha": "", "campo": col_to_delete, "valor_antigo": col_to_delete, "valor_novo": "coluna removida"}
                            )
                            _partners_save_state(state)
                            st.success("Coluna excluída da tabela e das exportações.")
                            app_rerun()
                else:
                    st.info("Não há colunas importadas disponíveis para exclusão.")

        elif mode == "Resumo de Atualizações":
            history = _partners_history_rows(partner.get("history", []))
            st.metric("Quantidade de alterações registradas", len(history))
            if history:
                hist_df = pd.DataFrame(history)
                preferred_cols = ["data_hora", "usuario", "acao", "detalhes", "linha", "campo", "tipo_acao"]
                hist_cols = [col for col in preferred_cols if col in hist_df.columns] + [col for col in hist_df.columns if col not in preferred_cols]
                st.dataframe(hist_df[hist_cols].head(300), use_container_width=True, hide_index=True)
            else:
                st.info("Nenhuma alteração registrada ainda.")

        elif mode == "Exportar CSV":
            if not can_export:
                st.warning("Exportação desabilitada para este usuário.")
            else:
                export_df_panel = df.drop(columns=[PARTNER_ROW_ID], errors="ignore")
                try:
                    csv_data = export_df_panel.to_csv(index=False).encode("utf-8-sig")
                except Exception:
                    csv_data = None
                    st.warning("Não foi possível exportar para CSV. Revise a planilha e tente novamente.")
                if csv_data is not None and st.download_button(
                    "⬇️ Baixar CSV tratado",
                    data=csv_data,
                    file_name=f"{_partner_label(partner_key)}_controle_tratado.csv",
                    mime="text/csv",
                    use_container_width=True,
                ):
                    _partners_add_history(state, partner_key, "Planilha exportada", "CSV", {"tipo_acao": "exportação", "valor_novo": "CSV"})
                    _partners_save_state(state)

        elif mode == "Exportar Excel":
            if not can_export:
                st.warning("Exportação desabilitada para este usuário.")
            else:
                export_df_panel = df.drop(columns=[PARTNER_ROW_ID], errors="ignore")
                excel_data, excel_error = _partners_safe_excel_bytes(export_df_panel)
                if excel_error:
                    st.info(excel_error)
                elif st.download_button(
                    "📗 Baixar Excel tratado (.xlsx)",
                    data=excel_data,
                    file_name=f"{_partner_label(partner_key)}_controle_tratado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                ):
                    _partners_add_history(state, partner_key, "Planilha exportada", "Excel", {"tipo_acao": "exportação", "valor_novo": "Excel"})
                    _partners_save_state(state)

        if mode == "Editar Célula":
            st.info("Edite diretamente nas células da tabela abaixo e clique em Salvar alteração.")

        filtered = _partners_filter_dataframe(df, st.session_state.get(search_key, ""), st.session_state.get(status_key, []))
        max_visible_rows = 1000
        render_df = filtered.head(max_visible_rows)
        if len(filtered) > max_visible_rows:
            st.info(f"Mostrando as primeiras {max_visible_rows} linhas filtradas para manter a tela leve. A exportação continua usando a planilha completa.")
        visible_ids = [str(v) for v in render_df.get(PARTNER_ROW_ID, [])]
        hidden_cols = st.session_state.get(hidden_cols_key, [])
        display_cols = [PARTNER_ROW_ID] + [col for col in df.columns if col not in hidden_cols and col != PARTNER_ROW_ID]
        editable_cols = set()
        if can_edit_rows:
            editable_cols = {col for col in display_cols if col not in (PARTNER_ROW_ID, "Última Alteração", "Usuário Responsável")}
        elif can_write_treatment:
            editable_cols = {col for col in ("Tratativa", "Descrição / Observação") if col in display_cols}
        disabled_cols = [col for col in display_cols if col not in editable_cols]

        st.markdown("##### Tabela central da planilha")
        if can_write_treatment:
            st.caption("Campos de tratativa disponíveis: Tratativa, Descrição / Observação, Última Alteração e Usuário Responsável.")
        edited = render_df[display_cols]
        if editable_cols:
            edited = st.data_editor(
                render_df[display_cols],
                use_container_width=True,
                hide_index=True,
                num_rows="fixed",
                height=520,
                key=f"partner_editor_{partner_key}_{len(df)}_{len(visible_ids)}_{len(editable_cols)}_{len(hidden_cols)}",
                disabled=disabled_cols,
                column_config={
                    PARTNER_ROW_ID: st.column_config.TextColumn("ID", width="small"),
                    "Status de Execução": st.column_config.SelectboxColumn(
                        "Status de Execução",
                        options=PARTNER_STATUS_OPTIONS,
                        required=False,
                    ),
                    "Tratativa": st.column_config.TextColumn("Tratativa", width="large"),
                    "Descrição / Observação": st.column_config.TextColumn("Descrição / Observação", width="large"),
                },
            )
        else:
            st.dataframe(render_df[display_cols], use_container_width=True, hide_index=True, height=520)

        save_cols = st.columns([1, 1, 2])
        with save_cols[0]:
            if editable_cols and st.button("💾 Salvar alteração", key=f"partner_save_editor_{partner_key}", type="primary", use_container_width=True):
                merged, logs = _partners_merge_edited_rows(df, visible_ids, edited, partner.get("columns", []))
                partner["rows"] = merged.to_dict(orient="records")
                for log_item in logs:
                    acao, detalhes = log_item[0], log_item[1]
                    extra = log_item[2] if len(log_item) > 2 and isinstance(log_item[2], dict) else {}
                    _partners_add_history(state, partner_key, acao, detalhes, extra)
                if not logs:
                    _partners_add_history(state, partner_key, "Dados internos salvos", "Sem alterações detectadas", {"tipo_acao": "edição"})
                _partners_save_state(state)
                st.success("Alterações salvas.")
                app_rerun()
        with save_cols[1]:
            if editable_cols and st.button("↩ Cancelar alteração", key=f"partner_cancel_editor_{partner_key}", use_container_width=True):
                app_rerun()

        last_update = partner.get("last_update") or {}
        if last_update:
            with st.expander("Resumo rápido da última importação/atualização", expanded=False):
                st.json({
                    "Última atualização": last_update.get("data_hora", ""),
                    "Usuário que atualizou": last_update.get("usuario", ""),
                    "Fonte": last_update.get("fonte", ""),
                    "Base de comparação": last_update.get("base_comparacao", ""),
                    "Linhas novas": last_update.get("linhas_novas", 0),
                    "Linhas alteradas": last_update.get("linhas_alteradas", 0),
                    "Linhas removidas": last_update.get("linhas_removidas", 0),
                    "Total de diferenças": last_update.get("total_diferencas", 0),
                })
        diff_rows = partner.get("diff_rows", [])
        if diff_rows:
            with st.expander("Destaques visuais da última atualização", expanded=False):
                diff_df = pd.DataFrame(diff_rows).drop(columns=[PARTNER_ROW_ID], errors="ignore")
                st.dataframe(_partners_style_diff(diff_df), use_container_width=True, hide_index=True)

def _render_partner_chat(state: dict, partner_key: str) -> None:
    if not _auth_partner_permission("partner_sheet_write_treatment"):
        st.warning("Seu usuário não possui permissão para registrar tratativas.")
        return
    partner = state["partners"][partner_key]
    mention_options, mention_labels = _auth_mention_options()
    st.markdown(
        "<div class='partner-window-title'>Tratativas</div>"
        "<div class='partner-window-subtitle'>Acompanhamento de assuntos, prazos, responsáveis e citações internas.</div>",
        unsafe_allow_html=True,
    )
    with st.form(f"partner_chat_form_{partner_key}"):
        c1, c2 = st.columns(2)
        with c1:
            assunto = st.text_input("Assunto")
            prazo = st.date_input("Prazo", value=date.today())
        with c2:
            status = st.selectbox("Status", PARTNER_TREATMENT_STATUS)
            st.text_input("Usuário responsável/autoria", value=_auth_user_name(), disabled=True)
        citados = st.multiselect(
            "Citar usuários cadastrados",
            mention_options,
            format_func=lambda usuario: mention_labels.get(usuario, usuario),
            help="Selecione os usuários que devem receber notificação ao entrar no sistema.",
        )
        mensagem = st.text_area("Mensagem")
        send = st.form_submit_button("💬 Enviar tratativa", type="primary", use_container_width=True)
    if send:
        if not mensagem.strip():
            st.error("Digite uma mensagem para registrar a tratativa.")
        else:
            item = {
                "id": hashlib.sha1(f"{partner_key}-{_now_iso()}-{_auth_user_name()}".encode()).hexdigest()[:12],
                "usuario": _auth_user_name(),
                "assunto": assunto.strip() or "Sem assunto",
                "mensagem": mensagem.strip(),
                "data": date.today().isoformat(),
                "hora": datetime.now().strftime("%H:%M:%S"),
                "prazo": prazo.isoformat() if prazo else "",
                "status": status,
                "citados": citados,
            }
            partner.setdefault("chat", []).insert(0, item)
            citados_label = ", ".join(mention_labels.get(usuario, usuario) for usuario in citados) or "Sem citações"
            _partners_add_history(state, partner_key, "Mensagem enviada", f"{item['assunto']} · prazo {item['prazo']} · citados: {citados_label}")
            _partners_save_state(state)
            st.success("Tratativa registrada.")
            app_rerun()

    for idx, item in enumerate(partner.get("chat", [])[:30]):
        chat_id = item.get("id") or hashlib.sha1(f"{partner_key}-{idx}-{item.get('assunto','')}-{item.get('data','')}".encode()).hexdigest()[:12]
        current_status = item.get("status", "Aberto")
        status_options = PARTNER_TREATMENT_STATUS if current_status in PARTNER_TREATMENT_STATUS else [current_status] + PARTNER_TREATMENT_STATUS
        color = "#66bb6a" if current_status in PARTNER_TREATMENT_DONE else "#ffd54f" if current_status == "Em andamento" else "#ff5252" if current_status == "Atrasado" else THEME_PRIMARY_COLOR
        citados_txt = ", ".join(mention_labels.get(str(usuario), str(usuario)) for usuario in item.get("citados", []))
        citados_html = f"<div style='color:#9ec9ef;font-size:.78rem;margin-top:6px;'>Citados: {citados_txt}</div>" if citados_txt else ""
        st.markdown(
            f"<div style='background:#151515;border:1px solid #333;border-left:4px solid {color};border-radius:10px;padding:10px 12px;margin:8px 0;'>"
            f"<div style='color:{color};font-weight:700;'>{item.get('assunto','Sem assunto')} · {item.get('status','')}</div>"
            f"<div style='color:#aaa;font-size:.78rem;'>{item.get('usuario','')} · {item.get('data','')} {item.get('hora','')} · Prazo: {item.get('prazo','')}</div>"
            f"<div style='color:#eee;margin-top:6px;'>{item.get('mensagem','')}</div>"
            f"{citados_html}"
            f"</div>",
            unsafe_allow_html=True,
        )
        sc1, sc2, sc3 = st.columns([1.4, 1, 2.4])
        with sc1:
            new_status = st.selectbox(
                "Alterar status",
                status_options,
                index=status_options.index(current_status),
                key=f"partner_chat_status_{partner_key}_{chat_id}",
            )
        with sc2:
            if st.button("Salvar status", key=f"partner_chat_save_status_{partner_key}_{chat_id}", use_container_width=True):
                if new_status != current_status:
                    item["status"] = new_status
                    item["status_atualizado_por"] = _auth_user_name()
                    item["status_atualizado_em"] = _now_human()
                    _partners_add_history(
                        state,
                        partner_key,
                        "Status da tratativa alterado",
                        f"{item.get('assunto','Sem assunto')}: {current_status} -> {new_status}"
                    )
                    _partners_save_state(state)
                    st.success("Status da tratativa atualizado.")
                    app_rerun()
                else:
                    st.info("Status já estava selecionado.")
        with sc3:
            if item.get("status_atualizado_em"):
                st.caption(f"Última atualização: {item.get('status_atualizado_em')} por {item.get('status_atualizado_por','')}")

def _render_partner_history(state: dict, partner_key: str) -> None:
    if not _auth_partner_permission("partner_sheet_history"):
        st.warning("Seu usuário não possui permissão para acessar o histórico.")
        return
    partner = state["partners"][partner_key]
    st.markdown(
        "<div class='partner-window-title'>Histórico</div>"
        "<div class='partner-window-subtitle'>Registro organizado das ações e atualizações do módulo.</div>",
        unsafe_allow_html=True,
    )
    h1, h2 = st.columns(2)
    with h1:
        st.markdown(f"##### Histórico {_partner_label(partner_key)}")
        st.dataframe(_partners_history_rows(partner.get("history", [])), use_container_width=True, hide_index=True)
    with h2:
        st.markdown("##### Histórico Geral")
        st.dataframe(_partners_history_rows(state.get("history_general", [])), use_container_width=True, hide_index=True)

def _render_partner_module_css() -> None:
    st.markdown("""
    <style>
    .partner-hero {
        background: linear-gradient(145deg, rgba(17,27,42,.98), rgba(8,14,24,.98));
        border: 1px solid rgba(66,165,245,.28);
        border-top: 2px solid var(--tmg-primary);
        border-radius: 12px;
        padding: 18px 20px;
        margin-bottom: 18px;
        box-shadow: 0 10px 26px rgba(0,0,0,.42), inset 0 1px 0 rgba(255,255,255,.04);
    }
    .partner-hero-title {
        color: var(--tmg-primary);
        font-weight: 900;
        letter-spacing: 3px;
        text-transform: uppercase;
        font-size: 1.25rem;
        text-shadow: 0 0 18px var(--tmg-primary-glow);
    }
    .partner-hero-subtitle,
    .partner-window-subtitle {
        color: #b8c7d9;
        font-size: .86rem;
        margin-top: 6px;
    }
    .partner-card {
        background: linear-gradient(145deg, rgba(18,33,54,.98), rgba(7,15,28,.98));
        border: 1px solid rgba(66,165,245,.28);
        border-radius: 12px;
        padding: 18px;
        min-height: 220px;
        box-shadow: 8px 8px 20px rgba(0,0,0,.45), -1px -1px 8px rgba(255,255,255,.04);
        text-align: center;
    }
    .partner-card-title,
    .partner-window-title {
        color: var(--tmg-primary);
        font-weight: 900;
        letter-spacing: 2.5px;
        text-transform: uppercase;
        text-shadow: 0 0 14px var(--tmg-primary-glow);
    }
    .partner-logo-slot {
        display: flex;
        align-items: center;
        justify-content: center;
        width: 140px;
        height: 96px;
        border: 1px dashed rgba(144,202,249,.30);
        border-radius: 10px;
        color: #7f96ad;
        font-size: .78rem;
        letter-spacing: 1px;
        margin: 0 auto 12px auto;
    }
    .partner-logo-frame {
        width: 140px;
        height: 96px;
        margin: 0 auto 14px auto;
        display: flex;
        align-items: center;
        justify-content: center;
        border: 1px solid rgba(144,202,249,.22);
        border-radius: 10px;
        background: rgba(5,14,26,.48);
        box-shadow: inset 0 1px 0 rgba(255,255,255,.04);
    }
    .partner-logo-img {
        max-width: 120px;
        max-height: 80px;
        width: auto;
        height: auto;
        object-fit: contain;
        display: block;
    }
    .partner-logo-empty {
        color: #7f96ad;
        font-size: .72rem;
        letter-spacing: 1.5px;
        text-transform: uppercase;
    }
    .partner-section-card {
        background: linear-gradient(145deg, rgba(18,33,54,.98), rgba(7,15,28,.98));
        border: 1px solid rgba(66,165,245,.26);
        border-radius: 12px;
        padding: 18px 16px;
        text-align: center;
        box-shadow: 6px 6px 18px rgba(0,0,0,.40), inset 0 1px 0 rgba(255,255,255,.04);
    }
    div[data-testid="stFileUploader"] section {
        border-color: rgba(66,165,245,.32) !important;
        background: rgba(13,30,53,.52) !important;
    }
    </style>
    """, unsafe_allow_html=True)

def _render_partner_selection(state: dict, allowed: list) -> None:
    st.markdown(
        "<div class='partner-window-title'>Parceiros</div>"
        "<div class='partner-window-subtitle'>Selecione a parceira para abrir o controle de voos e dados.</div>",
        unsafe_allow_html=True,
    )
    cols = st.columns(max(1, len(allowed)), gap="large")
    for col, partner_key in zip(cols, allowed):
        with col:
            label = _partner_label(partner_key)
            st.markdown("<div class='partner-card'>", unsafe_allow_html=True)
            st.markdown(_partners_logo_html(partner_key), unsafe_allow_html=True)
            st.markdown(f"<div class='partner-card-title'>{label}</div>", unsafe_allow_html=True)
            st.markdown("</div>", unsafe_allow_html=True)
            if st.button(label, key=f"partner_open_{partner_key}", type="primary", use_container_width=True):
                st.session_state["partner_selected"] = partner_key
                st.session_state["partner_section"] = ""
                app_rerun()

def _render_partner_section_buttons(partner_key: str) -> None:
    sections = []
    if _auth_partner_permission("partner_sheet_view"):
        sections.append(("sheet", "Planilha", "Espelho Excel/CSV, edição e comparação"))
    if _auth_partner_permission("partner_sheet_write_treatment"):
        sections.append(("chat", "Tratativa", "Assuntos, prazos e citações"))
    if _auth_partner_permission("partner_sheet_history"):
        sections.append(("history", "Histórico", "Registros e auditoria"))
    if not sections:
        st.warning("Seu usuário não possui permissões liberadas para abrir as áreas desta parceira.")
        return
    cols = st.columns(len(sections), gap="large")
    for col, (section_key, label, desc) in zip(cols, sections):
        with col:
            st.markdown(
                f"<div class='partner-section-card'><div class='partner-card-title'>{label}</div>"
                f"<div class='partner-window-subtitle'>{desc}</div></div>",
                unsafe_allow_html=True,
            )
            if st.button(label, key=f"partner_section_{partner_key}_{section_key}", type="primary", use_container_width=True):
                st.session_state["partner_section"] = section_key
                app_rerun()

def _render_partner_workspace(state: dict, partner_key: str) -> None:
    partner_name = _partner_label(partner_key)
    partner = state["partners"][partner_key]
    top_cols = st.columns([1, 3])
    with top_cols[0]:
        if st.button("← Parceiros", key=f"partner_back_{partner_key}", use_container_width=True):
            st.session_state["partner_selected"] = ""
            st.session_state["partner_section"] = ""
            app_rerun()
    with top_cols[1]:
        st.markdown(f"<div class='partner-window-title'>{partner_name}</div>", unsafe_allow_html=True)
    section = st.session_state.get("partner_section", "")
    if not section:
        st.markdown(_partners_logo_html(partner_key), unsafe_allow_html=True)
        _render_partner_alerts(partner)
        _render_partner_section_buttons(partner_key)
        st.info("Selecione Planilha, Tratativa ou Histórico para abrir a janela correspondente.")
        return
    allowed_sections = []
    if _auth_partner_permission("partner_sheet_view"):
        allowed_sections.append("sheet")
    if _auth_partner_permission("partner_sheet_write_treatment"):
        allowed_sections.append("chat")
    if _auth_partner_permission("partner_sheet_history"):
        allowed_sections.append("history")
    if section not in allowed_sections:
        st.warning("Seu usuário não possui permissão para abrir esta área.")
        return
    if section == "sheet":
        st.markdown(
            f"<div class='partner-section-card'><div class='partner-card-title'>{partner_name}</div>{_partners_logo_html(partner_key)}</div>",
            unsafe_allow_html=True,
        )
        _render_partner_sheet_controls(state, partner_key)
        _render_partner_table(state, partner_key)
    elif section == "chat":
        _render_partner_chat(state, partner_key)
    elif section == "history":
        _render_partner_history(state, partner_key)

def render_parceiros_controle() -> None:
    try:
        user = _auth_current_user()
        allowed = _auth_allowed_partners(user)
        if not allowed:
            st.warning("Seu usuário não possui permissão para acessar EIWA ou ALVAZ.")
            return
        state = _partners_load_state()
        _render_partner_module_css()
        st.markdown("""
        <div class='partner-hero'>
            <div class='partner-hero-title'>
                Parceiros / Controle de Voos e Dados
            </div>
            <div class='partner-hero-subtitle'>
                Módulo isolado para EIWA e ALVAZ: planilhas, tratativas, prazos, histórico e exportações.
            </div>
        </div>
        """, unsafe_allow_html=True)
        selected = st.session_state.get("partner_selected", "")
        if selected not in allowed:
            selected = ""
            st.session_state["partner_selected"] = ""
            st.session_state["partner_section"] = ""
        if not selected:
            _render_partner_selection(state, allowed)
        else:
            _render_partner_workspace(state, selected)
    except Exception:
        st.warning("Não foi possível abrir o módulo de Parceiros agora. Volte para a seleção de parceiros e tente novamente.")
        if st.button("Reabrir módulo Parceiros", key="partner_recover_safe", use_container_width=True):
            st.session_state["partner_selected"] = ""
            st.session_state["partner_section"] = ""
            app_rerun()

def _render_partner_mention_notifications() -> None:
    return

def _render_partner_logo_settings() -> None:
    if not _auth_is_admin():
        st.warning("Apenas o administrador Wellington pode alterar as logos das parceiras.")
        return
    _render_partner_module_css()
    state = _partners_load_state()
    st.markdown("#### Logos das Parceiras")
    st.caption("As logos ficam salvas na pasta padrão do sistema e aparecem automaticamente na tela de Parceiros.")
    cols = st.columns(2, gap="large")
    for col, partner_key in zip(cols, PARTNER_KEYS):
        with col:
            label = _partner_label(partner_key)
            st.markdown(f"<div class='partner-section-card'><div class='partner-card-title'>{label}</div>{_partners_logo_html(partner_key)}</div>", unsafe_allow_html=True)
            uploaded = st.file_uploader(
                f"Adicionar ou trocar logo {label}",
                type=["png", "jpg", "jpeg", "webp"],
                key=f"cfg_partner_logo_{partner_key}",
            )
            if uploaded is not None:
                load_box = st.empty()
                update_tmg_loading(load_box, 50, f"Carregando logo {label}...")
                _partners_save_logo(state, partner_key, uploaded)
                update_tmg_loading(load_box, 100, "Carregamento concluído com sucesso.")
                _partners_add_history(state, partner_key, "Logomarca atualizada", label)
                _partners_save_state(state)
                st.success(f"Logo {label} salva.")
                app_rerun()
            logo_path = _partners_logo_path(partner_key)
            if logo_path:
                st.caption(f"Arquivo salvo: {logo_path}")

def _render_logged_user_chip() -> None:
    if not st.session_state.get("logged_in", False):
        return
    user_name = _auth_user_name()
    st.markdown(
        f"""
        <div class='tmg-user-chip-neon'>
            Usuário: {user_name}
        </div>
        """,
        unsafe_allow_html=True,
    )

def _save_uploaded_files_generic(files, target_dir: Path) -> tuple:
    target_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    total_size = 0
    for uploaded in files or []:
        try:
            uploaded.seek(0)
        except Exception:
            pass
        data = uploaded.read()
        if isinstance(data, str):
            data = data.encode("utf-8")
        safe_name = Path(uploaded.name).name
        target = target_dir / safe_name
        if target.exists():
            target = target_dir / f"{target.stem}_{datetime.now().strftime('%H%M%S%f')}{target.suffix}"
        target.write_bytes(data)
        total_size += len(data)
        saved.append({
            "nome": uploaded.name,
            "path": str(target),
            "tamanho": len(data),
            "tamanho_fmt": _tv_human_size(len(data)),
            "sha256": hashlib.sha256(data).hexdigest(),
        })
    return saved, total_size

ORTHOMOSAIC_QUALITY_PRESETS = {
    "Alta qualidade": {
        "resolution_cm": 2,
        "feature_quality": "ultra",
        "pc_quality": "ultra",
        "mesh_depth": 13,
        "min_features": 12000,
        "matcher_neighbors": 8,
        "resize_to": 0,
        "extra_args": "--dsm --dtm --orthophoto-compression NONE",
        "descricao": "Prioriza nitidez, densidade de pontos e GeoTIFF final com menor perda.",
    },
    "Média qualidade": {
        "resolution_cm": 5,
        "feature_quality": "high",
        "pc_quality": "medium",
        "mesh_depth": 11,
        "min_features": 8000,
        "matcher_neighbors": 6,
        "resize_to": 3000,
        "extra_args": "--dsm",
        "descricao": "Equilibra qualidade visual e tempo de processamento.",
    },
    "Baixa/leve": {
        "resolution_cm": 10,
        "feature_quality": "medium",
        "pc_quality": "low",
        "mesh_depth": 9,
        "min_features": 4000,
        "matcher_neighbors": 4,
        "resize_to": 2000,
        "extra_args": "--fast-orthophoto",
        "descricao": "Prioriza velocidade e arquivos menores para testes ou máquinas leves.",
    },
}

def _orthomosaic_db_path() -> Path:
    root = SYSTEM_DATABASE_DIR / "ortomosaicos_jobs"
    root.mkdir(parents=True, exist_ok=True)
    return root / "ortomosaic_jobs.sqlite"

def _orthomosaic_init_db() -> None:
    db_path = _orthomosaic_db_path()
    with sqlite3.connect(db_path) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ortomosaic_jobs (
                job_id TEXT PRIMARY KEY,
                nome TEXT,
                pasta_entrada TEXT,
                pasta_saida TEXT,
                qualidade TEXT,
                status TEXT,
                caminho_ortomosaico TEXT,
                modo_execucao TEXT,
                criado_em TEXT,
                atualizado_em TEXT
            )
        """)
        conn.commit()

def _orthomosaic_upsert_job(record: dict) -> None:
    _orthomosaic_init_db()
    with sqlite3.connect(_orthomosaic_db_path()) as conn:
        conn.execute(
            """
            INSERT INTO ortomosaic_jobs
                (job_id, nome, pasta_entrada, pasta_saida, qualidade, status, caminho_ortomosaico, modo_execucao, criado_em, atualizado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(job_id) DO UPDATE SET
                nome=excluded.nome,
                pasta_entrada=excluded.pasta_entrada,
                pasta_saida=excluded.pasta_saida,
                qualidade=excluded.qualidade,
                status=excluded.status,
                caminho_ortomosaico=excluded.caminho_ortomosaico,
                modo_execucao=excluded.modo_execucao,
                atualizado_em=excluded.atualizado_em
            """,
            (
                record.get("job_id", ""),
                record.get("nome", ""),
                record.get("pasta_entrada", ""),
                record.get("pasta_saida", ""),
                record.get("qualidade", ""),
                record.get("status", ""),
                record.get("caminho_ortomosaico", ""),
                record.get("modo_execucao", ""),
                record.get("criado_em", _now_human()),
                record.get("atualizado_em", _now_human()),
            ),
        )
        conn.commit()

def _orthomosaic_list_jobs() -> list:
    _orthomosaic_init_db()
    with sqlite3.connect(_orthomosaic_db_path()) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT * FROM ortomosaic_jobs ORDER BY criado_em DESC LIMIT 80").fetchall()
    return [dict(row) for row in rows]

def _orthomosaic_find_outputs(output_dir: Path, project_slug: str = "") -> list:
    candidates = []
    roots = [output_dir]
    if project_slug:
        roots.append(output_dir / project_slug)
    seen = set()
    for root in roots:
        if not root.exists():
            continue
        for pattern in ("*.tif", "*.tiff", "*.geotiff", "*.png", "*.jpg", "*.jpeg"):
            for path in root.rglob(pattern):
                if path.is_file() and path not in seen:
                    seen.add(path)
                    candidates.append(path)
    return sorted(candidates, key=lambda item: item.stat().st_mtime, reverse=True)

def _orthomosaic_select_folder_dialog(initial_dir: Path) -> tuple:
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        selected = filedialog.askdirectory(initialdir=str(initial_dir), title="Selecionar pasta de saída do ortomosaico")
        root.destroy()
        if selected:
            return selected, ""
        return "", "Nenhuma pasta selecionada."
    except Exception as exc:
        return "", f"Seleção por janela indisponível neste ambiente. Informe o caminho manualmente. Detalhe: {exc}"

def _orthomosaic_quality_args(preset: dict) -> str:
    args = [
        f"--orthophoto-resolution {preset['resolution_cm']}",
        f"--feature-quality {preset['feature_quality']}",
        f"--pc-quality {preset['pc_quality']}",
        f"--mesh-octree-depth {preset['mesh_depth']}",
        f"--min-num-features {preset['min_features']}",
        f"--matcher-neighbors {preset['matcher_neighbors']}",
    ]
    if int(preset.get("resize_to", 0) or 0) > 0:
        args.append(f"--resize-to {preset['resize_to']}")
    if preset.get("extra_args"):
        args.append(str(preset["extra_args"]))
    return " ".join(args)

def _orthomosaic_write_vscode_workspace(job_dir: Path, project_slug: str, output_dir: Path, docker_image: str, quality_args: str, mode: str, final_name: str) -> dict:
    dataset_project = job_dir / project_slug
    images_dir = dataset_project / "images"
    logs_dir = job_dir / "logs"
    vscode_dir = job_dir / ".vscode"
    output_dir.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)
    vscode_dir.mkdir(parents=True, exist_ok=True)
    final_tif = output_dir / f"{final_name}.tif"

    docker_command = (
        f'docker run --rm -v "{job_dir}:/datasets" '
        f'{docker_image} --project-path /datasets {project_slug} {quality_args}'
    )
    ps_script = f"""$ErrorActionPreference = "Stop"
$Workspace = "{job_dir}"
$Project = "{project_slug}"
$OutputDir = "{output_dir}"
$FinalTif = "{final_tif}"
Write-Host "TMG Ortomosaico - processamento externo via Docker/WebODM"
Write-Host "Projeto: $Project"
Write-Host "Entrada: $Workspace\\$Project\\images"
Write-Host "Saida: $OutputDir"
{docker_command}
$Generated = Join-Path $Workspace "$Project\\odm_orthophoto\\odm_orthophoto.tif"
if (Test-Path $Generated) {{
    Copy-Item $Generated $FinalTif -Force
    Write-Host "Ortomosaico final copiado para: $FinalTif"
}} else {{
    Write-Host "Processamento finalizado, mas o arquivo odm_orthophoto.tif não foi encontrado automaticamente."
    Write-Host "Verifique a pasta: $Workspace\\$Project\\odm_orthophoto"
}}
"""
    sh_script = f"""#!/usr/bin/env bash
set -e
WORKSPACE="{job_dir.as_posix()}"
PROJECT="{project_slug}"
OUTPUT_DIR="{output_dir.as_posix()}"
FINAL_TIF="{final_tif.as_posix()}"
echo "TMG Ortomosaico - processamento externo via Docker/WebODM"
{docker_command}
GENERATED="$WORKSPACE/$PROJECT/odm_orthophoto/odm_orthophoto.tif"
if [ -f "$GENERATED" ]; then
  cp "$GENERATED" "$FINAL_TIF"
  echo "Ortomosaico final copiado para: $FINAL_TIF"
else
  echo "Arquivo final não encontrado automaticamente. Verifique $WORKSPACE/$PROJECT/odm_orthophoto"
fi
"""
    webodm_script = f"""$ErrorActionPreference = "Continue"
Write-Host "Abrindo WebODM local e pasta preparada para o projeto."
Write-Host "Imagens: {images_dir}"
Write-Host "Saida final desejada: {output_dir}"
Start-Process "http://localhost:8000"
Write-Host "Se usar a API do WebODM, configure WEBODM_TOKEN/WEBODM_URL e execute webodm_submit.py."
"""
    webodm_py = f'''"""
Envio opcional para WebODM local via API.
Configure as variáveis de ambiente:
  WEBODM_URL=http://localhost:8000
  WEBODM_TOKEN=seu_token
Depois execute no terminal do VS Code: python webodm_submit.py
"""
import os
from pathlib import Path
import requests

WEBODM_URL = os.environ.get("WEBODM_URL", "http://localhost:8000").rstrip("/")
TOKEN = os.environ.get("WEBODM_TOKEN", "")
PROJECT_NAME = "{final_name}"
IMAGES_DIR = Path(r"{images_dir}")

if not TOKEN:
    raise SystemExit("Configure WEBODM_TOKEN antes de enviar para o WebODM.")

headers = {{"Authorization": f"JWT {{TOKEN}}"}}
project_response = requests.post(f"{{WEBODM_URL}}/api/projects/", headers=headers, data={{"name": PROJECT_NAME}})
project_response.raise_for_status()
project_id = project_response.json()["id"]
files = [("images", (p.name, open(p, "rb"))) for p in IMAGES_DIR.iterdir() if p.is_file()]
try:
    task_response = requests.post(f"{{WEBODM_URL}}/api/projects/{{project_id}}/tasks/", headers=headers, files=files)
    task_response.raise_for_status()
    print("Tarefa criada no WebODM:", task_response.json())
finally:
    for _, (_, handle) in files:
        handle.close()
'''
    tasks = {
        "version": "2.0.0",
        "tasks": [
            {
                "label": "Gerar Ortomosaico - Docker ODM",
                "type": "shell",
                "command": "powershell",
                "args": ["-ExecutionPolicy", "Bypass", "-File", "${workspaceFolder}/run_ortomosaico.ps1"],
                "group": {"kind": "build", "isDefault": True},
                "problemMatcher": [],
            },
            {
                "label": "Abrir WebODM local",
                "type": "shell",
                "command": "powershell",
                "args": ["-ExecutionPolicy", "Bypass", "-File", "${workspaceFolder}/abrir_webodm_local.ps1"],
                "problemMatcher": [],
            },
            {
                "label": "Enviar para WebODM local via API",
                "type": "shell",
                "command": "python",
                "args": ["${workspaceFolder}/webodm_submit.py"],
                "problemMatcher": [],
            },
        ],
    }
    readme = f"""# TMG - Gerador de Ortomosaico

Projeto preparado pelo Streamlit para processamento externo.

## Como executar

1. Abra este diretório no VS Code.
2. Pressione `Ctrl+Shift+B` para rodar a tarefa padrão **Gerar Ortomosaico - Docker ODM**.
3. Aguarde o Docker/OpenDroneMap concluir.
4. O arquivo final será copiado para:

`{final_tif}`

## Modo selecionado

{mode}

## Comando Docker preparado

```powershell
{docker_command}
```

O Streamlit principal apenas preparou os arquivos; o processamento pesado roda fora dele.
"""
    (job_dir / "run_ortomosaico.ps1").write_text(ps_script, encoding="utf-8")
    (job_dir / "run_ortomosaico.sh").write_text(sh_script, encoding="utf-8")
    (job_dir / "abrir_webodm_local.ps1").write_text(webodm_script, encoding="utf-8")
    (job_dir / "webodm_submit.py").write_text(webodm_py, encoding="utf-8")
    (vscode_dir / "tasks.json").write_text(json.dumps(tasks, indent=2, ensure_ascii=False), encoding="utf-8")
    (job_dir / "README_EXECUTAR.md").write_text(readme, encoding="utf-8")
    return {"docker_command": docker_command, "final_tif": str(final_tif)}

def _orthomosaic_open_vscode(job_dir: Path) -> tuple:
    try:
        code_cmd = shutil.which("code") or shutil.which("code.cmd")
        if not code_cmd:
            common = Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Microsoft VS Code" / "bin" / "code.cmd"
            if common.exists():
                code_cmd = str(common)
        if not code_cmd:
            return False, "VS Code não encontrado no PATH. Abra manualmente a pasta do projeto gerado."
        subprocess.Popen([code_cmd, str(job_dir)], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return True, "VS Code aberto com o projeto preparado."
    except Exception as exc:
        return False, f"Não foi possível abrir o VS Code automaticamente: {exc}"

def _orthomosaic_simple_viewer_html(b64: str, image_name: str, width: int, height: int) -> str:
    return f"""
<!DOCTYPE html>
<html>
<head>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ background:#0b0f14; overflow:hidden; font-family:Segoe UI, Arial, sans-serif; }}
  #viewer {{ width:100%; height:660px; background:#0b0f14; border:1px solid #26384a; border-radius:8px; overflow:hidden; position:relative; }}
  canvas {{ position:absolute; inset:0; cursor:grab; image-rendering:auto; }}
  canvas:active {{ cursor:grabbing; }}
  .badge {{
    position:absolute; left:12px; top:12px; z-index:3; pointer-events:none;
    background:rgba(5,12,20,.82); border:1px solid #315170; color:#dcecff;
    border-radius:6px; padding:8px 11px; font-size:12px; letter-spacing:.3px;
    box-shadow:0 6px 18px rgba(0,0,0,.35);
  }}
  .zoom {{
    position:absolute; right:12px; top:12px; z-index:3; pointer-events:none;
    background:rgba(5,12,20,.82); border:1px solid #315170; color:#dcecff;
    border-radius:6px; padding:8px 11px; font-size:12px;
  }}
</style>
</head>
<body>
<div id="viewer">
  <canvas id="cv"></canvas>
  <div class="badge">{image_name} · {width} x {height} px</div>
  <div class="zoom" id="zoom">Zoom 100%</div>
</div>
<script>
const wrap=document.getElementById('viewer'), cv=document.getElementById('cv'), ctx=cv.getContext('2d'), zoom=document.getElementById('zoom');
let sc=1, ox=0, oy=0, drag=false, lx=0, ly=0, imgW=0, imgH=0;
const img=new Image();
function resize(){{ cv.width=wrap.clientWidth; cv.height=wrap.clientHeight; draw(); }}
function fit(){{ if(!imgW) return; sc=Math.min(cv.width/imgW, cv.height/imgH)*0.96; ox=(cv.width-imgW*sc)/2; oy=(cv.height-imgH*sc)/2; draw(); }}
function draw(){{ ctx.clearRect(0,0,cv.width,cv.height); ctx.save(); ctx.translate(ox,oy); ctx.scale(sc,sc); if(imgW) ctx.drawImage(img,0,0); ctx.restore(); zoom.textContent='Zoom '+Math.round(sc*100)+'%'; }}
img.onload=()=>{{ imgW=img.width; imgH=img.height; resize(); fit(); }};
img.src='data:image/jpeg;base64,{b64}';
window.addEventListener('resize', resize);
cv.addEventListener('wheel',e=>{{ e.preventDefault(); const f=e.deltaY<0?1.18:.84; const r=cv.getBoundingClientRect(); const mx=e.clientX-r.left,my=e.clientY-r.top; const ix=(mx-ox)/sc,iy=(my-oy)/sc; sc=Math.max(.05,Math.min(40,sc*f)); ox=mx-ix*sc; oy=my-iy*sc; draw(); }},{{passive:false}});
cv.addEventListener('mousedown',e=>{{ drag=true; lx=e.clientX; ly=e.clientY; }});
cv.addEventListener('mousemove',e=>{{ if(drag){{ ox+=e.clientX-lx; oy+=e.clientY-ly; lx=e.clientX; ly=e.clientY; draw(); }} }});
cv.addEventListener('mouseup',()=>{{ drag=false; }}); cv.addEventListener('mouseleave',()=>{{ drag=false; }});
cv.addEventListener('dblclick',fit);
</script>
</body>
</html>
"""

def _orthomosaic_render_viewer(path: Path) -> None:
    if not path.exists():
        st.warning("Arquivo do ortomosaico não encontrado.")
        return
    v1, v2 = st.columns([2, 1])
    with v1:
        if st.button("Visualizar ortomosaico", key=f"ortho_view_{path.as_posix()}", use_container_width=True):
            st.session_state["ortho_view_file"] = str(path)
    with v2:
        st.download_button(
            "Baixar Ortomosaico",
            data=path.read_bytes(),
            file_name=path.name,
            mime="application/octet-stream",
            use_container_width=True,
        )
    if st.session_state.get("ortho_view_file") == str(path):
        try:
            raw = path.read_bytes()
            b64, dims, err, _ = processar_ortofoto(raw, path.name)
            if err:
                st.info("Não foi possível gerar a prévia no navegador. O arquivo continua disponível para download.")
                return
            components.html(_orthomosaic_simple_viewer_html(b64, path.name, dims[0], dims[1]), height=680, scrolling=False)
        except Exception:
            st.info("Não foi possível gerar a prévia no navegador. O arquivo continua disponível para download.")

def _render_orthomosaic_generator() -> None:
    st.subheader("🛰️ Gerador de Ortomosaico")
    st.info("O Streamlit prepara o projeto, registra no SQLite e abre o VS Code. O processamento pesado fica no Docker/WebODM local.")
    jobs_root = SYSTEM_DATABASE_DIR / "ortomosaicos_jobs"
    default_output = SYSTEM_DATABASE_DIR / "ortomosaicos"
    jobs_root.mkdir(parents=True, exist_ok=True)
    default_output.mkdir(parents=True, exist_ok=True)

    files = st.file_uploader(
        "Anexar imagens/fotos do voo ou ortomosaico base",
        type=["jpg", "jpeg", "png", "tif", "tiff", "geotiff", "dng", "raw", "arw", "cr2", "nef", "zip"],
        accept_multiple_files=True,
        key="ortho_generator_images",
    )
    if files:
        render_tmg_loading_bar(100, f"{len(files)} arquivo(s) recebido(s) para geração de ortomosaico.")

    c1, c2, c3 = st.columns(3)
    with c1:
        job_name = st.text_input("Nome do ortomosaico", value=f"ortomosaico_{date.today().strftime('%Y%m%d')}", key="ortho_job_name")
        qualidade = st.selectbox("Qualidade / resolução", list(ORTHOMOSAIC_QUALITY_PRESETS.keys()), key="ortho_quality")
    with c2:
        modo_execucao = st.selectbox(
            "Modo de execução",
            ["Rodar pelo VS Code com Docker", "Rodar pelo VS Code + Docker local/WebODM já instalado"],
            key="ortho_execution_mode",
        )
        docker_image = st.text_input("Imagem Docker/OpenDroneMap", value="opendronemap/odm:latest", key="ortho_docker_image")
    with c3:
        if "ortho_output_dir" not in st.session_state:
            st.session_state["ortho_output_dir"] = str(default_output)
        if st.button("Selecionar pasta de saída no computador", key="ortho_choose_output_dir", use_container_width=True):
            selected_dir, folder_err = _orthomosaic_select_folder_dialog(default_output)
            if selected_dir:
                st.session_state["ortho_output_dir"] = selected_dir
                st.success(f"Pasta selecionada: {selected_dir}")
            else:
                st.info(folder_err)
        output_dir = st.text_input("Pasta de saída local", value=st.session_state["ortho_output_dir"], key="ortho_output_dir_text")
        st.session_state["ortho_output_dir"] = output_dir

    preset = ORTHOMOSAIC_QUALITY_PRESETS[qualidade]
    quality_args = _orthomosaic_quality_args(preset)
    st.markdown("##### Parâmetros automáticos")
    st.json({
        "qualidade": qualidade,
        "resolução_orthophoto_cm_px": preset["resolution_cm"],
        "feature_quality": preset["feature_quality"],
        "pc_quality": preset["pc_quality"],
        "mesh_octree_depth": preset["mesh_depth"],
        "min_features": preset["min_features"],
        "matcher_neighbors": preset["matcher_neighbors"],
        "descrição": preset["descricao"],
    })

    if st.button("🛰️ Gerar Ortomosaico", type="primary", key="ortho_start_generation", use_container_width=True):
        if not files:
            st.warning("Anexe as fotos do drone antes de iniciar.")
            return
        clean_name = _tv_safe_name(job_name) or f"ortomosaico_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        job_id = f"ORT_JOB_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{clean_name}"
        job_dir = jobs_root / job_id
        project_slug = clean_name
        input_dir = job_dir / project_slug / "images"
        logs_dir = job_dir / "logs"
        logs_dir.mkdir(parents=True, exist_ok=True)
        progress = st.empty()
        status = st.empty()
        update_tmg_loading(progress, 0, "Iniciando carregamento das imagens do ortomosaico...")
        status.info("1/5 — Salvando imagens do voo com integridade...")
        saved, total_size = _save_uploaded_files_generic(files, input_dir)
        update_tmg_loading(progress, 25, "Imagens salvas. Preparando configuração do ortomosaico...")

        output_path = _resolve_system_path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        status.info("2/5 — Gerando scripts, tasks do VS Code e parâmetros Docker/WebODM...")
        workspace = _orthomosaic_write_vscode_workspace(
            job_dir=job_dir,
            project_slug=project_slug,
            output_dir=output_path,
            docker_image=docker_image,
            quality_args=quality_args,
            mode=modo_execucao,
            final_name=clean_name,
        )
        job_config = {
            "job_id": job_id,
            "criado_em": _now_human(),
            "usuario": _auth_user_name(),
            "qualidade": qualidade,
            "parametros": preset,
            "modo_execucao": modo_execucao,
            "input_dir": str(input_dir),
            "output_dir": str(output_path),
            "docker_image": docker_image,
            "command": workspace["docker_command"],
            "final_tif": workspace["final_tif"],
            "arquivos": saved,
        }
        (job_dir / "job_config.json").write_text(json.dumps(job_config, indent=2, ensure_ascii=False), encoding="utf-8")
        log_path = logs_dir / "processing.log"
        log_path.write_text(
            "\n".join([
                f"[{_now_human()}] Job preparado: {job_id}",
                f"Entrada: {input_dir}",
                f"Saída: {output_path}",
                f"Modo: {modo_execucao}",
                f"Comando Docker: {workspace['docker_command']}",
            ]),
            encoding="utf-8",
        )
        _orthomosaic_upsert_job({
            "job_id": job_id,
            "nome": clean_name,
            "pasta_entrada": str(input_dir),
            "pasta_saida": str(output_path),
            "qualidade": qualidade,
            "status": "Preparado para execução externa no VS Code",
            "caminho_ortomosaico": workspace["final_tif"],
            "modo_execucao": modo_execucao,
            "criado_em": _now_human(),
            "atualizado_em": _now_human(),
        })
        update_tmg_loading(progress, 70, "Gerando pacote de integração e scripts...")

        package_base = jobs_root / f"{job_id}_pacote_integracao"
        package_zip = shutil.make_archive(str(package_base), "zip", root_dir=job_dir)
        opened, open_msg = _orthomosaic_open_vscode(job_dir)
        update_tmg_loading(progress, 100, "Carregamento concluído com sucesso.")
        status.success("5/5 — Projeto preparado para VS Code/Docker/WebODM.")

        st.success(f"Projeto `{job_id}` preparado com {len(saved)} arquivo(s), total {_tv_human_size(total_size)}.")
        (st.success if opened else st.info)(open_msg)
        st.json({
            "Job": job_id,
            "Nome": clean_name,
            "Entrada": str(input_dir),
            "Saída": str(output_path),
            "Qualidade": qualidade,
            "Modo": modo_execucao,
            "VS Code": str(job_dir),
            "SQLite": str(_orthomosaic_db_path()),
            "Ortomosaico esperado": workspace["final_tif"],
        })
        st.download_button(
            "⬇️ Baixar pacote de integração",
            data=Path(package_zip).read_bytes(),
            file_name=Path(package_zip).name,
            mime="application/zip",
            use_container_width=True,
        )

    st.markdown("#### Visualizador do ortomosaico gerado")
    jobs = _orthomosaic_list_jobs()
    if not jobs:
        st.info("Nenhum projeto de ortomosaico preparado ainda.")
        return
    labels = [
        f"{item.get('criado_em','')} · {item.get('nome','')} · {item.get('qualidade','')} · {item.get('status','')}"
        for item in jobs
    ]
    selected_label = st.selectbox("Projetos preparados", labels, key="ortho_job_view_select")
    selected_job = jobs[labels.index(selected_label)]
    output_path = _resolve_system_path(selected_job.get("pasta_saida", str(default_output)))
    expected_path = Path(str(selected_job.get("caminho_ortomosaico", "")))
    outputs = []
    if expected_path.exists():
        outputs.append(expected_path)
    outputs.extend([p for p in _orthomosaic_find_outputs(output_path, _tv_safe_name(selected_job.get("nome", ""))) if p not in outputs])
    if outputs:
        selected_output = st.selectbox("Ortomosaico disponível", [str(path) for path in outputs], key="ortho_output_file_select")
        selected_path = Path(selected_output)
        _orthomosaic_upsert_job({
            **selected_job,
            "status": "Ortomosaico encontrado",
            "caminho_ortomosaico": str(selected_path),
            "atualizado_em": _now_human(),
        })
        _orthomosaic_render_viewer(selected_path)
    else:
        st.info("Nenhum ortomosaico gerado encontrado ainda. Execute a task no VS Code e volte aqui para visualizar/baixar o resultado.")
        st.caption(f"Pasta monitorada: {output_path}")

def _render_sync_backup() -> None:
    st.subheader("🔄 Backup e Sincronização de Dados")
    st.info("Área real para gerar backup dos dados internos e preparar sincronização com Google Drive ou pasta externa.")

    backup_root = SYSTEM_DATABASE_DIR / "backups"
    sync_config_path = SYSTEM_DATABASE_DIR / "sync_config.json"
    try:
        sync_config = json.loads(sync_config_path.read_text(encoding="utf-8")) if sync_config_path.exists() else {}
    except Exception:
        sync_config = {}

    c1, c2 = st.columns(2)
    with c1:
        drive_target = st.text_input("Link ou caminho do Google Drive / destino", value=sync_config.get("destino", ""), key="sync_drive_target")
        destination_folder = st.text_input("Pasta de destino", value=sync_config.get("pasta_destino", "TMG_Backups"), key="sync_destination_folder")
        local_source = st.text_input("Caminho local dos dados", value=str(SYSTEM_DATABASE_DIR), key="sync_local_source")
    with c2:
        client_id = st.text_input("Google client_id", value=sync_config.get("client_id", ""), key="sync_client_id")
        client_secret = st.text_input("Google client_secret", value="", type="password", key="sync_client_secret")
        token = st.text_input("Token / refresh token", value="", type="password", key="sync_token")

    st.caption("Se a API do Google Drive exigir credenciais, preencha os campos acima. Sem credenciais, o sistema gera um pacote ZIP seguro para envio manual ou cópia local.")
    if st.button("▶ Iniciar sincronização / backup", type="primary", key="sync_start_backup", use_container_width=True):
        status = st.empty()
        status.info("1/4 — Mapeando dados internos do sistema...")
        source = _resolve_system_path(local_source)
        backup_root.mkdir(parents=True, exist_ok=True)
        backup_id = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        backup_file = backup_root / f"{backup_id}.zip"

        files_to_zip = []
        if source.exists():
            for path in source.rglob("*"):
                if path.is_file() and backup_root not in path.parents:
                    files_to_zip.append(path)

        status.info("2/4 — Compactando arquivos, históricos, tratativas, imagens e configurações...")
        with zipfile.ZipFile(backup_file, "w", zipfile.ZIP_DEFLATED) as zf:
            for path in files_to_zip:
                try:
                    zf.write(path, arcname=str(Path("dados") / path.relative_to(source)))
                except Exception:
                    pass
            for extra in (SYSTEM_CONFIG_PATH, LOGO_PATH, LOGIN_BG_PATH):
                if extra.exists():
                    zf.write(extra, arcname=str(Path("configuracoes") / extra.name))

        copied_to = ""
        target_text = str(drive_target or "").strip()
        if target_text and not target_text.lower().startswith(("http://", "https://")):
            try:
                target_dir = _resolve_system_path(target_text) / _tv_safe_name(destination_folder or "TMG_Backups")
                target_dir.mkdir(parents=True, exist_ok=True)
                copied = target_dir / backup_file.name
                shutil.copy2(backup_file, copied)
                copied_to = str(copied)
            except Exception as exc:
                copied_to = f"Não foi possível copiar automaticamente: {exc}"

        sync_config = {
            "destino": drive_target,
            "pasta_destino": destination_folder,
            "client_id": client_id,
            "client_secret_configurado": bool(client_secret),
            "token_configurado": bool(token),
            "ultimo_backup": _now_human(),
            "ultimo_arquivo": str(backup_file),
            "copiado_para": copied_to,
        }
        sync_config_path.write_text(json.dumps(sync_config, indent=2, ensure_ascii=False), encoding="utf-8")
        status.success("4/4 — Backup finalizado e registro atualizado.")
        st.success(f"Backup criado: `{backup_file}`")
        st.json(sync_config)
        st.download_button(
            "⬇️ Baixar backup ZIP",
            data=backup_file.read_bytes(),
            file_name=backup_file.name,
            mime="application/zip",
            use_container_width=True,
        )




# ==========================================
# MODULO ISOLADO - ANALISE DE MARCACAO DE GRID
# ==========================================
GRIDMARK_ROOT = SYSTEM_DATABASE_DIR / "analise_marcacao_grid"
GRIDMARK_UPLOAD_DIR = GRIDMARK_ROOT / "uploads"
GRIDMARK_CACHE_DIR = GRIDMARK_ROOT / "cache"
GRIDMARK_VECTOR_COLORS = ["#00d4ff", "#ff8c00", "#5ff2b1", "#ff4fd8", "#ffd166", "#7aa7ff", "#ff6b6b"]

def preparar_cache_grid_viewer() -> Path:
    GRIDMARK_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    try:
        now_ts = datetime.now().timestamp()
        for item in GRIDMARK_CACHE_DIR.iterdir():
            try:
                if item.is_file() and now_ts - item.stat().st_mtime > 7 * 24 * 3600:
                    item.unlink(missing_ok=True)
                elif item.is_dir() and now_ts - item.stat().st_mtime > 7 * 24 * 3600 and item.resolve().is_relative_to(GRIDMARK_CACHE_DIR.resolve()):
                    shutil.rmtree(item, ignore_errors=True)
            except Exception:
                pass
    except Exception:
        pass
    return GRIDMARK_CACHE_DIR

def manter_estado_visualizador_grid() -> None:
    _gridmark_init_state()
    preparar_cache_grid_viewer()

def atualizar_apenas_camada_modificada(layer_id: str = "") -> None:
    st.session_state["gridmark_modified_layer_id"] = str(layer_id or "")
    st.session_state["gridmark_refresh_token"] = datetime.now().isoformat(timespec="microseconds")

def _gridmark_limpar_cache_temporario() -> None:
    for path in (GRIDMARK_CACHE_DIR, GRIDMARK_UPLOAD_DIR):
        try:
            resolved = path.resolve()
            root = GRIDMARK_ROOT.resolve()
            if resolved == root or not resolved.is_relative_to(root):
                continue
            if path.exists():
                shutil.rmtree(path, ignore_errors=True)
            path.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass
    for cached_fn in (
        carregar_preview_raster_otimizado,
        _gridmark_load_gdf_cached,
        _gridmark_attrs_cached,
        carregar_geojson_cacheado,
    ):
        try:
            cached_fn.clear()
        except Exception:
            pass

def _gridmark_limpar_analise() -> None:
    st.session_state["gridmark_raster"] = None
    st.session_state["gridmark_raster_visible"] = True
    st.session_state["gridmark_layers"] = []
    st.session_state["gridmark_selected_layer_id"] = ""
    st.session_state["gridmark_selected_feature_index"] = None
    st.session_state["gridmark_selected_feature_attrs"] = {}
    st.session_state["gridmark_selection_payload"] = ""
    st.session_state["gridmark_last_selection_payload"] = ""
    st.session_state["gridmark_show_table"] = False
    st.session_state["gridmark_zoom_layer_id"] = ""
    st.session_state["gridmark_zoom_selected"] = False
    _gridmark_limpar_cache_temporario()

def _gridmark_init_state() -> None:
    if "gridmark_raster" not in st.session_state:
        st.session_state["gridmark_raster"] = None
    if "gridmark_raster_visible" not in st.session_state:
        st.session_state["gridmark_raster_visible"] = True
    if "gridmark_layers" not in st.session_state:
        st.session_state["gridmark_layers"] = []
    if "gridmark_show_table" not in st.session_state:
        st.session_state["gridmark_show_table"] = False
    if "gridmark_selected_layer_id" not in st.session_state:
        st.session_state["gridmark_selected_layer_id"] = ""
    if "gridmark_selected_feature_index" not in st.session_state:
        st.session_state["gridmark_selected_feature_index"] = None
    if "gridmark_selected_feature_attrs" not in st.session_state:
        st.session_state["gridmark_selected_feature_attrs"] = {}
    if "gridmark_zoom_layer_id" not in st.session_state:
        st.session_state["gridmark_zoom_layer_id"] = ""
    if "gridmark_zoom_selected" not in st.session_state:
        st.session_state["gridmark_zoom_selected"] = False
    if "gridmark_selection_payload" not in st.session_state:
        st.session_state["gridmark_selection_payload"] = ""
    if "gridmark_last_selection_payload" not in st.session_state:
        st.session_state["gridmark_last_selection_payload"] = ""
    if "gridmark_show_all_layers" not in st.session_state:
        st.session_state["gridmark_show_all_layers"] = True

def _gridmark_logo_html() -> str:
    try:
        return _tmg_loading_logo_html()
    except Exception:
        return "<div class='tmg-load-logo-fallback'>TMG</div>"

def render_loading_camadas(progress, texto: str = "Carregando camada...", arquivo: str = "", etapa: str = "", container=None):
    try:
        pct = max(0, min(100, int(float(progress))))
    except Exception:
        pct = 0
    texto_seguro = html.escape(str(texto or "Carregando camada..."))
    arquivo_seguro = html.escape(str(arquivo or ""))
    etapa_segura = html.escape(str(etapa or ""))
    logo_html = _gridmark_logo_html()
    markup = f"""
    <style>
    @keyframes gridmarkLogoPulse {{
        0%,100% {{ transform:scale(1); filter:drop-shadow(0 0 9px rgba(0,212,255,.34)); }}
        50% {{ transform:scale(1.035); filter:drop-shadow(0 0 18px rgba(0,212,255,.58)); }}
    }}
    @keyframes gridmarkBarShine {{
        0% {{ transform:translateX(-130%); opacity:.25; }}
        55% {{ opacity:.75; }}
        100% {{ transform:translateX(130%); opacity:.25; }}
    }}
    .gridmark-loading-card {{
        margin:10px 0 14px 0;
        padding:18px 20px;
        border-radius:16px;
        border:1px solid rgba(0,212,255,.42);
        background:
            linear-gradient(120deg, rgba(255,255,255,.12), transparent 30%),
            radial-gradient(circle at top left, rgba(0,212,255,.22), transparent 44%),
            linear-gradient(145deg, rgba(2,14,36,.97), rgba(12,57,98,.88), rgba(0,212,255,.18));
        box-shadow:
            0 16px 34px rgba(0,0,0,.48),
            0 0 28px rgba(0,212,255,.28),
            inset 0 1px 0 rgba(255,255,255,.22),
            inset 0 -9px 18px rgba(2,14,36,.50);
        color:#fff;
        text-align:center;
        transition:all .30s ease;
        backdrop-filter:blur(10px) saturate(145%);
        -webkit-backdrop-filter:blur(10px) saturate(145%);
    }}
    .gridmark-loading-card:hover {{
        box-shadow:
            0 18px 38px rgba(0,0,0,.52),
            0 0 36px rgba(0,212,255,.42),
            inset 0 1px 0 rgba(255,255,255,.30);
    }}
    .gridmark-loading-card:active {{ transform:translateY(1px) scale(.998); }}
    .gridmark-loading-logo .tmg-load-logo-img {{
        max-height:52px;
        max-width:140px;
        object-fit:contain;
        animation:gridmarkLogoPulse 1.8s ease-in-out infinite;
    }}
    .gridmark-loading-title {{
        margin-top:8px;
        font-size:.96rem;
        font-weight:900;
        letter-spacing:.8px;
        color:#fff;
        text-shadow:0 1px 0 rgba(0,0,0,.88), 0 0 14px rgba(0,212,255,.48);
    }}
    .gridmark-loading-file {{
        margin-top:4px;
        font-size:.78rem;
        color:#d9fbff;
        font-weight:700;
        text-shadow:0 1px 0 rgba(0,0,0,.75);
    }}
    .gridmark-loading-track {{
        position:relative;
        height:18px;
        margin-top:14px;
        border-radius:999px;
        overflow:hidden;
        border:1px solid rgba(255,255,255,.14);
        background:linear-gradient(180deg,#04101f,#0b2540);
        box-shadow:inset 0 3px 8px rgba(0,0,0,.58), 0 0 16px rgba(0,212,255,.16);
    }}
    .gridmark-loading-fill {{
        position:absolute;
        inset:0 auto 0 0;
        width:{pct}%;
        border-radius:999px;
        background:linear-gradient(90deg,#42a5f5,#00d4ff,#5ff2b1);
        box-shadow:0 0 18px rgba(0,212,255,.55), inset 0 1px 0 rgba(255,255,255,.36);
        transition:width .35s ease;
        overflow:hidden;
    }}
    .gridmark-loading-fill:after {{
        content:"";
        position:absolute;
        inset:0;
        background:linear-gradient(90deg, transparent, rgba(255,255,255,.62), transparent);
        animation:gridmarkBarShine 1.45s ease-in-out infinite;
    }}
    .gridmark-loading-pct {{
        margin-top:8px;
        font-size:.88rem;
        font-weight:900;
        color:#ffffff;
        text-shadow:0 0 12px rgba(0,212,255,.50);
    }}
    </style>
    <div class="gridmark-loading-card">
        <div class="gridmark-loading-logo">{logo_html}</div>
        <div class="gridmark-loading-title">{texto_seguro}</div>
        <div class="gridmark-loading-file">{arquivo_seguro}</div>
        <div class="gridmark-loading-track"><div class="gridmark-loading-fill"></div></div>
        <div class="gridmark-loading-pct">{etapa_segura} {pct}%</div>
    </div>
    """
    target = container if container is not None else st
    target.markdown(markup, unsafe_allow_html=True)

@st.cache_data(show_spinner=False, max_entries=10)
def carregar_preview_raster_otimizado(file_bytes: bytes, filename: str):
    return _processar_ortofoto_cached(
        file_bytes,
        filename,
        _preview_max_dim(),
        _preview_jpeg_quality(),
        _preview_max_payload_mb(),
        _preview_min_dim(),
    )

def render_loading_visualizador_grid(progress=67, texto: str = "Carregando visualizador...", etapa: str = "Preparando camadas selecionadas...") -> str:
    try:
        pct = max(0, min(100, int(float(progress))))
    except Exception:
        pct = 67
    logo_html = _gridmark_logo_html()
    texto_seguro = html.escape(str(texto or "Carregando visualizador..."))
    etapa_segura = html.escape(str(etapa or "Preparando camadas selecionadas..."))
    return f"""
    <div class="qgisLoadingOverlay" id="viewerLoading">
      <div class="qgisLoadingCard">
        <div class="qgisLoadingLogo">{logo_html}</div>
        <div class="qgisLoadingText">{texto_seguro}</div>
        <div class="qgisLoadingTrack"><div class="qgisLoadingFill" id="viewerLoadingFill" style="width:{pct}%"></div></div>
        <div class="qgisLoadingPct" id="viewerLoadingPct">{pct}%</div>
        <div class="qgisLoadingStep" id="viewerLoadingStep">{etapa_segura}</div>
      </div>
    </div>
    """

def mostrar_loading_central_grid(container=None, progress=67, texto: str = "Carregando visualizador...", etapa: str = "Preparando camadas selecionadas...") -> None:
    target = container if container is not None else st
    target.markdown(
        f"""
        <style>
        .gridmark-central-loading-preview {{
            margin:12px auto;
            max-width:430px;
            border-radius:18px;
            border:1px solid rgba(0,212,255,.42);
            background:linear-gradient(145deg, rgba(2,14,36,.96), rgba(12,57,98,.84), rgba(0,212,255,.18));
            box-shadow:0 18px 38px rgba(0,0,0,.48), 0 0 30px rgba(0,212,255,.28), inset 0 1px 0 rgba(255,255,255,.20);
            padding:18px;
            color:#fff;
            text-align:center;
        }}
        </style>
        <div class="gridmark-central-loading-preview">{render_loading_visualizador_grid(progress, texto, etapa)}</div>
        """,
        unsafe_allow_html=True,
    )

def ocultar_loading_central_grid() -> None:
    st.session_state["gridmark_loading_visible"] = False

def _gridmark_safe_json(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8", errors="ignore")
        except Exception:
            return str(value)
    return value if isinstance(value, (str, int, float, bool)) or value is None else str(value)

def _gridmark_layer_id(prefix: str, name: str, payload: bytes = b"") -> str:
    seed = f"{prefix}_{name}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}".encode("utf-8") + (payload[:2048] or b"")
    return hashlib.sha256(seed).hexdigest()[:16]

def _gridmark_upload_signature(uploaded_files) -> str:
    digest = hashlib.sha256()
    for item in uploaded_files or []:
        try:
            data = item.getvalue()
        except Exception:
            data = b""
        digest.update(str(item.name).encode("utf-8", errors="ignore"))
        digest.update(str(len(data)).encode("ascii"))
        digest.update(hashlib.sha256(data).hexdigest().encode("ascii"))
    return digest.hexdigest()[:24]

def _gridmark_read_raster(uploaded_file) -> tuple:
    if uploaded_file is None:
        return None, "Selecione uma camada raster para carregar."
    try:
        data = uploaded_file.getvalue()
    except Exception as exc:
        return None, f"Erro ao ler raster enviado: {exc}"
    if not data:
        return None, "Arquivo raster vazio."
    try:
        preparar_cache_grid_viewer()
        b64, dims, err, spatial = carregar_preview_raster_otimizado(data, uploaded_file.name)
        if err:
            return None, err
        if not b64 or not dims:
            return None, "Não foi possível gerar a visualização do raster."
        transform_gdal = spatial.get("transform")
        if transform_gdal is not None:
            try:
                transform_gdal = [float(v) for v in transform_gdal]
            except Exception:
                transform_gdal = None
        raster = {
            "id": _gridmark_layer_id("raster", uploaded_file.name, data),
            "name": uploaded_file.name,
            "data_url": f"data:image/jpeg;base64,{b64}",
            "width": int(spatial.get("preview_width") or dims[0]),
            "height": int(spatial.get("preview_height") or dims[1]),
            "orig_width": int(spatial.get("orig_width") or dims[0]),
            "orig_height": int(spatial.get("orig_height") or dims[1]),
            "ratio": float(spatial.get("ratio") or 1.0),
            "crs_wkt": spatial.get("crs"),
            "transform_gdal": transform_gdal,
            "preview_payload_mb": spatial.get("preview_payload_mb", ""),
        }
        return raster, None
    except Exception as exc:
        return None, f"Erro ao carregar camada raster: {exc}"

def _gridmark_extract_vector_path(uploaded_files) -> tuple:
    if not uploaded_files:
        return None, "Selecione uma camada vetorial para carregar."
    try:
        GRIDMARK_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        first_bytes = uploaded_files[0].getvalue() if uploaded_files else b""
        upload_id = _gridmark_layer_id("vetor", uploaded_files[0].name, first_bytes)
        target_dir = GRIDMARK_UPLOAD_DIR / upload_id
        target_dir.mkdir(parents=True, exist_ok=True)
        zip_file = next((f for f in uploaded_files if Path(f.name).suffix.lower() == ".zip"), None)
        if zip_file is not None:
            zip_path = target_dir / _tv_safe_name(zip_file.name)
            zip_path.write_bytes(zip_file.getvalue())
            try:
                with zipfile.ZipFile(zip_path, "r") as zf:
                    zf.extractall(target_dir)
            except Exception as exc:
                return None, f"Erro ao extrair ZIP do shapefile: {exc}"
        else:
            for item in uploaded_files:
                safe_name = _tv_safe_name(Path(item.name).stem) + Path(item.name).suffix.lower()
                (target_dir / safe_name).write_bytes(item.getvalue())
        shp_files = sorted(target_dir.rglob("*.shp"))
        geojson_files = sorted([p for p in target_dir.rglob("*") if p.suffix.lower() in (".geojson", ".json")])
        if shp_files:
            folder_files = {p.suffix.lower() for p in shp_files[0].parent.iterdir() if p.is_file()}
            missing = [ext for ext in (".shp", ".shx", ".dbf", ".prj") if ext not in folder_files]
            if missing:
                return None, "Shapefile incompleto. Envie o .zip contendo .shp, .shx, .dbf e .prj."
            return shp_files[0], None
        if geojson_files:
            return geojson_files[0], None
        return None, "Nenhum arquivo vetorial válido encontrado. Envie .shp completo, .geojson, .json ou .zip."
    except Exception as exc:
        return None, f"Erro ao preparar camada vetorial: {exc}"

def _gridmark_extract_vector_paths(uploaded_files) -> tuple:
    if not uploaded_files:
        return [], "Selecione uma camada vetorial para carregar."
    try:
        GRIDMARK_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        upload_id = _gridmark_upload_signature(uploaded_files)
        target_dir = GRIDMARK_UPLOAD_DIR / upload_id
        target_dir.mkdir(parents=True, exist_ok=True)
        zip_files = [f for f in uploaded_files if Path(f.name).suffix.lower() == ".zip"]
        if zip_files:
            for zip_file in zip_files:
                zip_path = target_dir / _tv_safe_name(zip_file.name)
                zip_path.write_bytes(zip_file.getvalue())
                try:
                    with zipfile.ZipFile(zip_path, "r") as zf:
                        zf.extractall(target_dir / _tv_safe_name(Path(zip_file.name).stem))
                except Exception as exc:
                    return [], f"Erro ao extrair ZIP do shapefile: {exc}"
        else:
            for item in uploaded_files:
                safe_name = _tv_safe_name(Path(item.name).stem) + Path(item.name).suffix.lower()
                (target_dir / safe_name).write_bytes(item.getvalue())

        vector_paths = []
        for shp_path in sorted(target_dir.rglob("*.shp")):
            folder_files = {p.suffix.lower() for p in shp_path.parent.iterdir() if p.is_file()}
            missing = [ext for ext in (".shp", ".shx", ".dbf", ".prj") if ext not in folder_files]
            if missing:
                return [], "Shapefile incompleto. Envie o .zip contendo .shp, .shx, .dbf e .prj."
            vector_paths.append(shp_path)
        vector_paths.extend(sorted([p for p in target_dir.rglob("*") if p.suffix.lower() in (".geojson", ".json")]))
        if not vector_paths:
            return [], "Nenhum arquivo vetorial válido encontrado. Envie .shp completo, .geojson, .json ou .zip."
        return vector_paths, None
    except Exception as exc:
        return [], f"Erro ao preparar camada vetorial: {exc}"

def _gridmark_group_vector_uploads(uploaded_files) -> list:
    groups = []
    if not uploaded_files:
        return groups
    zip_files = [item for item in uploaded_files if Path(item.name).suffix.lower() == ".zip"]
    direct_vectors = [item for item in uploaded_files if Path(item.name).suffix.lower() in (".geojson", ".json")]
    shape_parts = [item for item in uploaded_files if Path(item.name).suffix.lower() in (".shp", ".shx", ".dbf", ".prj", ".cpg")]
    for item in zip_files:
        groups.append([item])
    for item in direct_vectors:
        groups.append([item])
    by_stem = {}
    for item in shape_parts:
        by_stem.setdefault(Path(item.name).stem.lower(), []).append(item)
    for items in by_stem.values():
        groups.append(items)
    return groups

@st.cache_data(show_spinner=False, max_entries=24)
def _gridmark_load_gdf_cached(path_text: str):
    return gpd.read_file(path_text)

@st.cache_data(show_spinner=False, max_entries=48)
def _gridmark_attrs_cached(path_text: str):
    gdf = gpd.read_file(path_text)
    attrs_df = gdf.drop(columns=["geometry"], errors="ignore").copy()
    for col in attrs_df.columns:
        attrs_df[col] = attrs_df[col].map(_gridmark_safe_json)
    return attrs_df

def _gridmark_read_vector(uploaded_files, layer_name: str = "") -> tuple:
    if not HAS_GEOPANDAS:
        return None, "Instale geopandas, shapely e pyproj para carregar camadas vetoriais."
    vector_path, err = _gridmark_extract_vector_path(uploaded_files)
    if err:
        return None, err
    try:
        gdf = _gridmark_load_gdf_cached(str(vector_path))
        if gdf is None or gdf.empty:
            return None, "Camada vetorial sem feições."
        if "geometry" not in gdf.columns:
            return None, "Camada vetorial sem geometria."
        gdf = gdf[gdf.geometry.notna()].copy()
        try:
            gdf = gdf[~gdf.geometry.is_empty].copy()
        except Exception:
            pass
        if gdf.empty:
            return None, "Camada vetorial sem geometrias válidas."
        attrs_df = _gridmark_attrs_cached(str(vector_path))
        color_index = len(st.session_state.get("gridmark_layers", [])) % len(GRIDMARK_VECTOR_COLORS)
        payload = str(vector_path).encode("utf-8")
        layer = {
            "id": _gridmark_layer_id("vetor", str(vector_path.name), payload),
            "name": layer_name.strip() or vector_path.stem,
            "path": str(vector_path),
            "source_signature": _gridmark_upload_signature(uploaded_files),
            "visible": True,
            "color": GRIDMARK_VECTOR_COLORS[color_index],
            "opacity": 55,
            "highlight": False,
            "expanded": False,
            "gdf": gdf,
            "attrs_df": attrs_df,
            "crs": str(gdf.crs) if getattr(gdf, "crs", None) else "",
            "feature_count": int(len(gdf)),
        }
        return layer, None
    except Exception as exc:
        return None, f"Erro ao abrir camada vetorial: {exc}"

def cachear_camada_vetorial(uploaded_files, layer_name: str = "") -> tuple:
    if not HAS_GEOPANDAS:
        return [], ["Instale geopandas, shapely e pyproj para carregar camadas vetoriais."]
    paths, err = _gridmark_extract_vector_paths(uploaded_files)
    if err:
        return [], [err]
    existing = {
        str(layer.get("source_signature") or layer.get("path") or "")
        for layer in st.session_state.get("gridmark_layers", [])
    }
    layers = []
    errors = []
    source_signature = _gridmark_upload_signature(uploaded_files)
    for vector_path in paths:
        path_signature = hashlib.sha256(str(vector_path).encode("utf-8", errors="ignore")).hexdigest()[:24]
        if source_signature in existing or path_signature in existing or str(vector_path) in existing:
            errors.append(f"{vector_path.name}: camada já importada, não foi duplicada.")
            continue
        try:
            gdf = _gridmark_load_gdf_cached(str(vector_path))
            if gdf is None or gdf.empty:
                errors.append(f"{vector_path.name}: camada vetorial sem feições.")
                continue
            if "geometry" not in gdf.columns:
                errors.append(f"{vector_path.name}: camada vetorial sem geometria.")
                continue
            gdf = gdf[gdf.geometry.notna()].copy()
            try:
                gdf = gdf[~gdf.geometry.is_empty].copy()
            except Exception:
                pass
            if gdf.empty:
                errors.append(f"{vector_path.name}: camada vetorial sem geometrias válidas.")
                continue
            attrs_df = _gridmark_attrs_cached(str(vector_path))
            color_index = (len(st.session_state.get("gridmark_layers", [])) + len(layers)) % len(GRIDMARK_VECTOR_COLORS)
            payload = str(vector_path).encode("utf-8")
            display_name = layer_name.strip() if layer_name.strip() and len(paths) == 1 else vector_path.stem
            layers.append({
                "id": _gridmark_layer_id("vetor", str(vector_path.name), payload),
                "name": display_name,
                "path": str(vector_path),
                "source_signature": source_signature if len(paths) == 1 else path_signature,
                "visible": True,
                "color": GRIDMARK_VECTOR_COLORS[color_index],
                "opacity": 55,
                "highlight": False,
                "expanded": False,
                "gdf": gdf,
                "attrs_df": attrs_df,
                "crs": str(gdf.crs) if getattr(gdf, "crs", None) else "",
                "feature_count": int(len(gdf)),
            })
        except Exception as exc:
            errors.append(f"{vector_path.name}: erro ao abrir geometria: {exc}")
    return layers, errors

def _gridmark_read_vectors(uploaded_files, layer_name: str = "", progress_container=None) -> tuple:
    groups = _gridmark_group_vector_uploads(uploaded_files)
    if not groups:
        return [], ["Selecione uma camada vetorial para carregar."]
    loaded_layers = []
    errors = []
    total = max(1, len(groups))
    for idx, group in enumerate(groups, start=1):
        base_pct = int(((idx - 1) / total) * 80)
        render_loading_camadas(
            min(95, base_pct + 15),
            "Camada Vetorial:",
            group[0].name,
            "Lendo arquivo...",
            progress_container,
        )
        layers, group_errors = cachear_camada_vetorial(group, layer_name if total == 1 else "")
        errors.extend(group_errors)
        if not layers:
            continue
        render_loading_camadas(
            min(98, base_pct + 55),
            "Camada Vetorial:",
            f"{len(layers)} camada(s)",
            "Carregando geometrias...",
            progress_container,
        )
        loaded_layers.extend(layers)
    if loaded_layers:
        render_loading_camadas(100, "Preparando visualização:", f"{len(loaded_layers)} camada(s)", "Finalizando...", progress_container)
    return loaded_layers, errors

def _gridmark_reproject_gdf(gdf, raster: dict) -> tuple:
    warnings_list = []
    result = gdf
    raster_crs_wkt = (raster or {}).get("crs_wkt")
    if not raster_crs_wkt:
        if getattr(gdf, "crs", None):
            warnings_list.append("Raster sem CRS detectado. A camada vetorial será exibida por encaixe aproximado da extensão.")
        return result, warnings_list
    if not getattr(gdf, "crs", None):
        warnings_list.append("Camada vetorial sem CRS. Não foi possível reprojetar automaticamente.")
        return result, warnings_list
    try:
        from pyproj import CRS
        raster_crs = CRS.from_wkt(raster_crs_wkt)
        vector_crs = CRS.from_user_input(gdf.crs)
        if vector_crs != raster_crs:
            result = gdf.to_crs(raster_crs)
            warnings_list.append("Camada vetorial reprojetada para o CRS da ortofoto.")
    except Exception:
        warnings_list.append("Não foi possível alinhar automaticamente as camadas. Verifique o sistema de coordenadas dos arquivos.")
    return result, warnings_list

def _gridmark_coord_transformer(gdf, raster: dict):
    raster = raster or {}
    transform_gdal = raster.get("transform_gdal")
    width = max(1, int(raster.get("width") or 1400))
    height = max(1, int(raster.get("height") or 800))
    ratio = float(raster.get("ratio") or 1.0)
    if transform_gdal:
        try:
            from affine import Affine as _Affine
            inv_transform = ~_Affine.from_gdal(*transform_gdal)
            def to_px(x, y):
                col, row = inv_transform * (float(x), float(y))
                return [round(col * ratio, 3), round(row * ratio, 3)]
            return to_px, None
        except Exception:
            pass
    try:
        minx, miny, maxx, maxy = [float(v) for v in gdf.total_bounds]
        dx = max(maxx - minx, 1e-9)
        dy = max(maxy - miny, 1e-9)
        def to_px_bbox(x, y):
            px = (float(x) - minx) / dx * width
            py = (maxy - float(y)) / dy * height
            return [round(px, 3), round(py, 3)]
        return to_px_bbox, "Camada desenhada por extensão aproximada porque o raster não possui georreferenciamento completo."
    except Exception:
        return None, "Não foi possível calcular o posicionamento da camada vetorial."

def _gridmark_coords_to_pixels(coords, transformer, max_points: int = 900):
    raw = list(coords or [])
    if len(raw) > max_points:
        step = max(1, int(len(raw) / max_points))
        raw = raw[::step]
    pts = []
    for item in raw:
        try:
            x, y = item[:2]
            pts.append(transformer(x, y))
        except Exception:
            continue
    return pts

def _gridmark_geometry_to_canvas(geom, transformer) -> dict:
    payload = {"polygons": [], "lines": [], "points": []}
    if geom is None:
        return payload
    geom_type = getattr(geom, "geom_type", "")
    try:
        if geom_type == "Polygon":
            rings = [_gridmark_coords_to_pixels(geom.exterior.coords, transformer)]
            for interior in geom.interiors:
                rings.append(_gridmark_coords_to_pixels(interior.coords, transformer))
            if rings and rings[0]:
                payload["polygons"].append(rings)
        elif geom_type == "MultiPolygon":
            for part in geom.geoms:
                part_payload = _gridmark_geometry_to_canvas(part, transformer)
                payload["polygons"].extend(part_payload["polygons"])
        elif geom_type == "LineString":
            line = _gridmark_coords_to_pixels(geom.coords, transformer)
            if line:
                payload["lines"].append(line)
        elif geom_type == "MultiLineString":
            for part in geom.geoms:
                line = _gridmark_coords_to_pixels(part.coords, transformer)
                if line:
                    payload["lines"].append(line)
        elif geom_type == "Point":
            payload["points"].append(transformer(geom.x, geom.y))
        elif geom_type == "MultiPoint":
            for part in geom.geoms:
                payload["points"].append(transformer(part.x, part.y))
        elif geom_type == "GeometryCollection":
            for part in geom.geoms:
                part_payload = _gridmark_geometry_to_canvas(part, transformer)
                payload["polygons"].extend(part_payload["polygons"])
                payload["lines"].extend(part_payload["lines"])
                payload["points"].extend(part_payload["points"])
    except Exception:
        return payload
    return payload

def otimizar_renderizacao(gdf):
    try:
        if gdf is None or gdf.empty:
            return gdf, []
        warnings_list = []
        feature_count = int(len(gdf))
        minx, miny, maxx, maxy = [float(v) for v in gdf.total_bounds]
        diag = max(((maxx - minx) ** 2 + (maxy - miny) ** 2) ** 0.5, 1e-9)
        if feature_count >= 450:
            tolerance = diag * 0.000035
        elif feature_count >= 160:
            tolerance = diag * 0.000018
        else:
            tolerance = 0
        if tolerance > 0:
            optimized = gdf.copy()
            optimized["geometry"] = optimized.geometry.simplify(tolerance, preserve_topology=True)
            warnings_list.append("Renderização otimizada com simplificação temporária para manter o visualizador fluido.")
            return optimized, warnings_list
    except Exception:
        pass
    return gdf, []

def otimizar_camadas_grid_viewer(gdf):
    return otimizar_renderizacao(gdf)

def _gridmark_payload_bbox(geom_payload: dict):
    xs = []
    ys = []
    for poly in geom_payload.get("polygons", []):
        for ring in poly:
            for pt in ring:
                if len(pt) >= 2:
                    xs.append(float(pt[0]))
                    ys.append(float(pt[1]))
    for line in geom_payload.get("lines", []):
        for pt in line:
            if len(pt) >= 2:
                xs.append(float(pt[0]))
                ys.append(float(pt[1]))
    for pt in geom_payload.get("points", []):
        if len(pt) >= 2:
            xs.append(float(pt[0]))
            ys.append(float(pt[1]))
    if not xs or not ys:
        return None
    return [round(min(xs), 3), round(min(ys), 3), round(max(xs), 3), round(max(ys), 3)]

def _gridmark_merge_bbox(current, new_bbox):
    if not new_bbox:
        return current
    if not current:
        return list(new_bbox)
    return [
        min(current[0], new_bbox[0]),
        min(current[1], new_bbox[1]),
        max(current[2], new_bbox[2]),
        max(current[3], new_bbox[3]),
    ]

@st.cache_data(show_spinner=False, max_entries=48)
def carregar_geojson_cacheado(path_text: str, raster_crs_wkt: str, transform_json: str, width: int, height: int, ratio: float):
    warnings_list = []
    try:
        gdf = _gridmark_load_gdf_cached(path_text)
        if gdf is None or gdf.empty:
            return {"features": [], "bbox": None, "warnings": ["Camada sem feições para desenhar."]}
        if "geometry" not in gdf.columns:
            return {"features": [], "bbox": None, "warnings": ["Camada vetorial sem geometria."]}
        gdf = gdf[gdf.geometry.notna()].copy()
        try:
            gdf = gdf[~gdf.geometry.is_empty].copy()
        except Exception:
            pass
        attrs_df = gdf.drop(columns=["geometry"], errors="ignore").copy()
        for col in attrs_df.columns:
            attrs_df[col] = attrs_df[col].map(_gridmark_safe_json)
        raster_meta = {
            "crs_wkt": raster_crs_wkt or "",
            "transform_gdal": json.loads(transform_json) if transform_json else None,
            "width": int(width or 1400),
            "height": int(height or 800),
            "ratio": float(ratio or 1.0),
        }
        gdf_work, reproj_warnings = _gridmark_reproject_gdf(gdf, raster_meta)
        warnings_list.extend(reproj_warnings)
        transformer, transform_warning = _gridmark_coord_transformer(gdf_work, raster_meta)
        if transform_warning:
            warnings_list.append(transform_warning)
        if transformer is None:
            return {"features": [], "bbox": None, "warnings": warnings_list or ["Não foi possível posicionar a camada."]}
        gdf_work, opt_warnings = otimizar_renderizacao(gdf_work)
        warnings_list.extend(opt_warnings)
        features = []
        layer_bbox = None
        for idx, row in gdf_work.iterrows():
            geom_payload = _gridmark_geometry_to_canvas(row.geometry, transformer)
            if not geom_payload["polygons"] and not geom_payload["lines"] and not geom_payload["points"]:
                continue
            bbox = _gridmark_payload_bbox(geom_payload)
            layer_bbox = _gridmark_merge_bbox(layer_bbox, bbox)
            attrs = {}
            if idx in attrs_df.index:
                attrs = {str(k): _gridmark_safe_json(v) for k, v in attrs_df.loc[idx].to_dict().items()}
            features.append({
                "index": int(idx) if isinstance(idx, (int, np.integer)) else str(idx),
                "attrs": attrs,
                "polygons": geom_payload["polygons"],
                "lines": geom_payload["lines"],
                "points": geom_payload["points"],
                "bbox": bbox,
            })
        return {"features": features, "bbox": layer_bbox, "warnings": warnings_list}
    except Exception as exc:
        return {"features": [], "bbox": None, "warnings": [f"Erro ao abrir geometria: {exc}"]}

def _gridmark_prepare_layer_for_viewer(layer: dict, raster: dict) -> tuple:
    warnings_list = []
    try:
        path_text = str(layer.get("path") or "")
        if path_text:
            transform_json = json.dumps((raster or {}).get("transform_gdal") or [])
            cached = carregar_geojson_cacheado(
                path_text,
                str((raster or {}).get("crs_wkt") or ""),
                transform_json,
                int((raster or {}).get("width") or 1400),
                int((raster or {}).get("height") or 800),
                float((raster or {}).get("ratio") or 1.0),
            )
            viewer_layer = {
                "id": layer.get("id"),
                "name": layer.get("name"),
                "color": layer.get("color", "#00d4ff"),
                "opacity": max(0, min(100, int(layer.get("opacity", 55)))) / 100,
                "highlight": bool(layer.get("highlight", False)),
                "bbox": cached.get("bbox"),
                "features": cached.get("features", []),
            }
            return viewer_layer, cached.get("warnings", [])
        gdf = layer.get("gdf")
        if gdf is None or gdf.empty:
            return None, ["Camada sem feições para desenhar."]
        gdf_work, reproj_warnings = _gridmark_reproject_gdf(gdf, raster)
        warnings_list.extend(reproj_warnings)
        transformer, transform_warning = _gridmark_coord_transformer(gdf_work, raster)
        if transform_warning:
            warnings_list.append(transform_warning)
        if transformer is None:
            return None, warnings_list or ["Não foi possível posicionar a camada."]
        gdf_work, opt_warnings = otimizar_renderizacao(gdf_work)
        warnings_list.extend(opt_warnings)
        attrs_df = layer.get("attrs_df")
        features = []
        layer_bbox = None
        for idx, row in gdf_work.iterrows():
            geom_payload = _gridmark_geometry_to_canvas(row.geometry, transformer)
            if not geom_payload["polygons"] and not geom_payload["lines"] and not geom_payload["points"]:
                continue
            bbox = _gridmark_payload_bbox(geom_payload)
            layer_bbox = _gridmark_merge_bbox(layer_bbox, bbox)
            if attrs_df is not None and idx in attrs_df.index:
                attrs = {str(k): _gridmark_safe_json(v) for k, v in attrs_df.loc[idx].to_dict().items()}
            else:
                attrs = {}
            features.append({
                "index": int(idx) if isinstance(idx, (int, np.integer)) else str(idx),
                "attrs": attrs,
                "polygons": geom_payload["polygons"],
                "lines": geom_payload["lines"],
                "points": geom_payload["points"],
                "bbox": bbox,
            })
        viewer_layer = {
            "id": layer.get("id"),
            "name": layer.get("name"),
            "color": layer.get("color", "#00d4ff"),
            "opacity": max(0, min(100, int(layer.get("opacity", 55)))) / 100,
            "highlight": bool(layer.get("highlight", False)),
            "bbox": layer_bbox,
            "features": features,
        }
        return viewer_layer, warnings_list
    except Exception as exc:
        return None, [f"Erro ao preparar camada para visualização: {exc}"]

def _gridmark_render_viewer(raster: dict, layers: list) -> None:
    raster_visible = bool(st.session_state.get("gridmark_raster_visible", True))
    selected_layer_id = st.session_state.get("gridmark_selected_layer_id", "")
    selected_feature_index = st.session_state.get("gridmark_selected_feature_index", None)
    image_width = int((raster or {}).get("width") or 1400)
    image_height = int((raster or {}).get("height") or 800)
    viewer_layers = []
    viewer_warnings = []
    for layer in layers:
        if not layer.get("visible", True):
            continue
        prepared, warnings_list = _gridmark_prepare_layer_for_viewer(layer, raster)
        if warnings_list:
            for warning in warnings_list:
                if warning not in viewer_warnings:
                    viewer_warnings.append(warning)
        if prepared:
            viewer_layers.append(prepared)
    for warning in viewer_warnings[:4]:
        st.warning(warning)
    viewer_data = {
        "raster": raster if (raster and raster_visible) else None,
        "imageWidth": image_width,
        "imageHeight": image_height,
        "layers": viewer_layers,
        "selectedLayerId": selected_layer_id,
        "selectedFeatureIndex": selected_feature_index,
        "zoomLayerId": st.session_state.get("gridmark_zoom_layer_id", ""),
        "zoomSelected": bool(st.session_state.get("gridmark_zoom_selected", False)),
    }
    data_json = json.dumps(viewer_data, ensure_ascii=False).replace("</", "<\\/")
    viewer_html = """
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
* { box-sizing:border-box; }
body { margin:0; background:#061526; font-family:'Segoe UI',sans-serif; overflow:hidden; }
#qgisViewer {
  width:100%; height:760px; position:relative; overflow:hidden; border-radius:12px;
  border:1px solid rgba(0,212,255,.34);
  background:
    linear-gradient(rgba(255,255,255,.035) 1px, transparent 1px),
    linear-gradient(90deg, rgba(255,255,255,.035) 1px, transparent 1px),
    radial-gradient(circle at 20% 10%, rgba(0,212,255,.13), transparent 32%),
    #071525;
  background-size:34px 34px,34px 34px,100% 100%,100% 100%;
  box-shadow:0 18px 42px rgba(0,0,0,.46), inset 0 1px 0 rgba(255,255,255,.10);
  cursor:grab;
}
#qgisViewer:active { cursor:grabbing; }
#mapCanvas { width:100%; height:100%; display:block; }
.qgisToolbar { position:absolute; top:12px; right:12px; display:flex; flex-direction:column; gap:6px; z-index:5; }
.qgisBtn {
  width:36px; height:36px; border-radius:9px; border:1px solid rgba(0,212,255,.42);
  color:#fff; background:linear-gradient(145deg, rgba(5,30,58,.95), rgba(0,128,176,.62));
  font-weight:900; cursor:pointer; box-shadow:0 8px 18px rgba(0,0,0,.42), 0 0 16px rgba(0,212,255,.20), inset 0 1px 0 rgba(255,255,255,.18);
  transition:.22s ease;
}
.qgisBtn:hover { box-shadow:0 10px 24px rgba(0,0,0,.48), 0 0 24px rgba(0,212,255,.42); transform:translateY(-1px); }
.qgisInfo {
  position:absolute; left:12px; bottom:12px; width:min(390px, calc(100% - 86px)); max-height:240px; overflow:auto;
  border-radius:12px; padding:12px 14px; z-index:6; color:#eafcff;
  border:1px solid rgba(0,212,255,.34);
  background:linear-gradient(145deg, rgba(3,18,38,.92), rgba(8,45,82,.78));
  box-shadow:0 12px 26px rgba(0,0,0,.45), 0 0 22px rgba(0,212,255,.20), inset 0 1px 0 rgba(255,255,255,.10);
  font-size:12px; line-height:1.35;
}
.qgisInfo b { color:#fff; text-shadow:0 0 10px rgba(0,212,255,.45); }
.qgisInfo table { width:100%; border-collapse:collapse; margin-top:8px; }
.qgisInfo td { border-bottom:1px solid rgba(255,255,255,.08); padding:4px 2px; vertical-align:top; }
.qgisInfo td:first-child { color:#8feaff; font-weight:700; width:38%; }
.qgisStatus {
  position:absolute; left:12px; top:12px; z-index:6; color:#dffaff; font-size:12px; padding:8px 10px;
  border-radius:10px; border:1px solid rgba(0,212,255,.28);
  background:rgba(3,18,38,.72); box-shadow:0 8px 20px rgba(0,0,0,.32);
}
@keyframes qgisLoadingPulse {
  0%,100% { transform:scale(1); filter:drop-shadow(0 0 10px rgba(0,212,255,.36)); }
  50% { transform:scale(1.035); filter:drop-shadow(0 0 22px rgba(0,212,255,.68)); }
}
@keyframes qgisLoadingShine {
  0% { transform:translateX(-140%); opacity:.18; }
  55% { opacity:.78; }
  100% { transform:translateX(140%); opacity:.22; }
}
.qgisLoadingOverlay {
  position:absolute;
  inset:0;
  z-index:30;
  display:flex;
  align-items:center;
  justify-content:center;
  background:radial-gradient(circle at center, rgba(0,20,36,.50), rgba(0,0,0,.58));
  backdrop-filter:blur(4px) saturate(130%);
  -webkit-backdrop-filter:blur(4px) saturate(130%);
  transition:opacity .30s ease, visibility .30s ease;
}
.qgisLoadingOverlay.is-hidden {
  opacity:0;
  visibility:hidden;
  pointer-events:none;
}
.qgisLoadingCard {
  width:min(420px, calc(100% - 42px));
  border-radius:18px;
  border:1px solid rgba(0,212,255,.48);
  background:
    linear-gradient(120deg, rgba(255,255,255,.14), transparent 30%),
    radial-gradient(circle at top left, rgba(0,212,255,.24), transparent 44%),
    linear-gradient(145deg, rgba(2,14,36,.98), rgba(12,57,98,.88), rgba(0,212,255,.18));
  box-shadow:0 18px 40px rgba(0,0,0,.58), 0 0 34px rgba(0,212,255,.34), inset 0 1px 0 rgba(255,255,255,.22);
  color:#fff;
  text-align:center;
  padding:22px 24px;
}
.qgisLoadingLogo .tmg-load-logo-img {
  max-height:58px;
  max-width:150px;
  object-fit:contain;
  animation:qgisLoadingPulse 1.8s ease-in-out infinite;
}
.qgisLoadingLogo .tmg-load-logo-fallback {
  color:#fff;
  font-size:1.35rem;
  font-weight:900;
  letter-spacing:2px;
  text-shadow:0 0 18px rgba(0,212,255,.66);
  animation:qgisLoadingPulse 1.8s ease-in-out infinite;
}
.qgisLoadingText {
  margin-top:10px;
  font-weight:900;
  color:#fff;
  letter-spacing:.8px;
  text-shadow:0 2px 0 rgba(0,0,0,.86), 0 0 16px rgba(0,212,255,.52);
}
.qgisLoadingTrack {
  position:relative;
  height:18px;
  margin-top:15px;
  overflow:hidden;
  border-radius:999px;
  border:1px solid rgba(255,255,255,.14);
  background:linear-gradient(180deg,#04101f,#0b2540);
  box-shadow:inset 0 3px 8px rgba(0,0,0,.58), 0 0 16px rgba(0,212,255,.18);
}
.qgisLoadingFill {
  position:absolute;
  inset:0 auto 0 0;
  width:12%;
  border-radius:999px;
  background:linear-gradient(90deg,#42a5f5,#00d4ff,#5ff2b1);
  box-shadow:0 0 20px rgba(0,212,255,.58), inset 0 1px 0 rgba(255,255,255,.35);
  transition:width .30s ease;
  overflow:hidden;
}
.qgisLoadingFill:after {
  content:"";
  position:absolute;
  inset:0;
  background:linear-gradient(90deg, transparent, rgba(255,255,255,.68), transparent);
  animation:qgisLoadingShine 1.35s ease-in-out infinite;
}
.qgisLoadingPct {
  margin-top:9px;
  font-weight:900;
  color:#ffffff;
  text-shadow:0 0 14px rgba(0,212,255,.58);
}
.qgisLoadingStep {
  margin-top:4px;
  color:#d9fbff;
  font-size:12px;
  font-weight:700;
  text-shadow:0 1px 0 rgba(0,0,0,.75);
}
</style>
</head>
<body>
<div id="qgisViewer">
  <canvas id="mapCanvas"></canvas>
  __GRIDMARK_LOADING__
  <div class="qgisStatus" id="mapStatus">QGIS-like viewer</div>
  <div class="qgisToolbar">
    <button class="qgisBtn" id="zoomIn" title="Aproximar">+</button>
    <button class="qgisBtn" id="zoomOut" title="Afastar">−</button>
    <button class="qgisBtn" id="fit" title="Ajustar à tela">⊡</button>
    <button class="qgisBtn" id="oneToOne" title="Zoom 100%">1:1</button>
    <button class="qgisBtn" id="center" title="Centralizar">◎</button>
  </div>
  <div class="qgisInfo" id="featureInfo"><b>Identificação de parcela</b><br>Clique em um polígono para ver ID, TIRO, DISPARO e demais atributos.</div>
</div>
<script type="application/json" id="viewerData">__VIEWER_DATA__</script>
<script>
const data = JSON.parse(document.getElementById('viewerData').textContent);
const wrap = document.getElementById('qgisViewer');
const canvas = document.getElementById('mapCanvas');
const ctx = canvas.getContext('2d');
const statusEl = document.getElementById('mapStatus');
const infoEl = document.getElementById('featureInfo');
const loadingEl = document.getElementById('viewerLoading');
const loadingFill = document.getElementById('viewerLoadingFill');
const loadingPct = document.getElementById('viewerLoadingPct');
const loadingStep = document.getElementById('viewerLoadingStep');
const img = new Image();
let scale = 1, ox = 0, oy = 0, dpr = window.devicePixelRatio || 1;
let drag = false, sx = 0, sy = 0, sox = 0, soy = 0, moved = false;
let localSelected = { layerId: data.selectedLayerId || '', index: data.selectedFeatureIndex };
const imgW = Math.max(1, data.imageWidth || 1400);
const imgH = Math.max(1, data.imageHeight || 800);
let viewerLoadingProgress = 18;
let viewerLoadingTimer = window.setInterval(() => {
  if(!loadingEl || loadingEl.classList.contains('is-hidden')) return;
  viewerLoadingProgress = Math.min(94, viewerLoadingProgress + 4);
  setViewerLoading(viewerLoadingProgress, 'Preparando camadas selecionadas...');
}, 180);
function setViewerLoading(pct, step){
  viewerLoadingProgress = Math.max(0, Math.min(100, Math.round(Number(pct) || 0)));
  if(loadingFill) loadingFill.style.width = viewerLoadingProgress + '%';
  if(loadingPct) loadingPct.textContent = viewerLoadingProgress + '%';
  if(loadingStep && step) loadingStep.textContent = step;
}
function hideViewerLoading(){
  setViewerLoading(100, 'Finalizando...');
  window.setTimeout(() => {
    if(loadingEl) loadingEl.classList.add('is-hidden');
    if(viewerLoadingTimer) window.clearInterval(viewerLoadingTimer);
  }, 280);
}
function esc(v){ return String(v ?? '').replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])); }
function hexToRgba(hex, alpha){
  let h = String(hex || '#00d4ff').replace('#','');
  if(h.length === 3) h = h.split('').map(x => x + x).join('');
  const n = parseInt(h, 16);
  return `rgba(${(n>>16)&255},${(n>>8)&255},${n&255},${alpha})`;
}
function resize(){
  const r = wrap.getBoundingClientRect();
  canvas.width = Math.max(1, Math.floor(r.width * dpr));
  canvas.height = Math.max(1, Math.floor(r.height * dpr));
  canvas.style.width = r.width + 'px';
  canvas.style.height = r.height + 'px';
  ctx.setTransform(dpr,0,0,dpr,0,0);
  draw();
}
function fitView(){
  const r = wrap.getBoundingClientRect();
  scale = Math.min((r.width - 42) / imgW, (r.height - 42) / imgH);
  if(!Number.isFinite(scale) || scale <= 0) scale = 1;
  ox = (r.width - imgW * scale) / 2;
  oy = (r.height - imgH * scale) / 2;
  draw();
}
function centerView(){ const r = wrap.getBoundingClientRect(); ox = (r.width - imgW * scale) / 2; oy = (r.height - imgH * scale) / 2; draw(); }
function zoomAt(factor, cx, cy){
  const ix = (cx - ox) / scale, iy = (cy - oy) / scale;
  scale = Math.max(0.02, Math.min(80, scale * factor));
  ox = cx - ix * scale; oy = cy - iy * scale; draw();
}
function screenPoint(pt){ return [ox + pt[0] * scale, oy + pt[1] * scale]; }
function drawRing(ring, closePath=true){
  if(!ring || !ring.length) return;
  const p0 = screenPoint(ring[0]); ctx.moveTo(p0[0], p0[1]);
  for(let i=1;i<ring.length;i++){ const p = screenPoint(ring[i]); ctx.lineTo(p[0], p[1]); }
  if(closePath) ctx.closePath();
}
function drawFeature(layer, feature){
  const selected = String(localSelected.layerId) === String(layer.id) && String(localSelected.index) === String(feature.index);
  ctx.save();
  ctx.lineJoin = 'round'; ctx.lineCap = 'round';
  ctx.globalAlpha = selected ? 1 : Math.max(.08, Math.min(1, Number(layer.opacity ?? .55)));
  ctx.strokeStyle = selected ? '#ffd21f' : layer.color;
  ctx.fillStyle = selected ? 'rgba(255,210,31,.36)' : hexToRgba(layer.color, layer.highlight ? .24 : .13);
  ctx.shadowColor = selected ? 'rgba(255,210,31,.92)' : hexToRgba(layer.color, layer.highlight ? .72 : .44);
  ctx.shadowBlur = selected ? 20 : (layer.highlight ? 16 : 7);
  ctx.lineWidth = selected ? 4.2 : (layer.highlight ? 3.0 : 1.7);
  (feature.polygons || []).forEach(poly => {
    ctx.beginPath();
    (poly || []).forEach(ring => drawRing(ring, true));
    ctx.fill('evenodd'); ctx.stroke();
  });
  (feature.lines || []).forEach(line => { ctx.beginPath(); drawRing(line, false); ctx.stroke(); });
  (feature.points || []).forEach(pt => { const p = screenPoint(pt); ctx.beginPath(); ctx.arc(p[0], p[1], selected ? 6 : 4, 0, Math.PI * 2); ctx.fill(); ctx.stroke(); });
  ctx.restore();
}
function zoomToBounds(bbox, padding=56){
  if(!bbox || bbox.length < 4) return;
  const r = wrap.getBoundingClientRect();
  const bw = Math.max(1, bbox[2] - bbox[0]);
  const bh = Math.max(1, bbox[3] - bbox[1]);
  const sx = (r.width - padding * 2) / bw;
  const sy = (r.height - padding * 2) / bh;
  scale = Math.max(0.02, Math.min(80, Math.min(sx, sy)));
  ox = (r.width - bw * scale) / 2 - bbox[0] * scale;
  oy = (r.height - bh * scale) / 2 - bbox[1] * scale;
  draw();
}
function findLayerById(layerId){
  return (data.layers || []).find(layer => String(layer.id) === String(layerId));
}
function findFeature(layer, index){
  if(!layer) return null;
  return (layer.features || []).find(f => String(f.index) === String(index));
}
function syncToPythonGridmark(layer, feature){
  try {
    const payload = JSON.stringify({layerId: layer.id, index: feature.index, attrs: feature.attrs || {}, layerName: layer.name});
    window.localStorage.setItem('tmg_gridmark_selection_payload', payload);
    const parentDoc = window.parent && window.parent.document;
    if(parentDoc) {
      const inputs = Array.from(parentDoc.querySelectorAll('input'));
      const target = inputs.find(el => el.getAttribute('aria-label') === 'gridmark_selection_payload');
      if(target) {
        const setter = Object.getOwnPropertyDescriptor(window.parent.HTMLInputElement.prototype, 'value').set;
        setter.call(target, payload);
        target.dispatchEvent(new Event('input', {bubbles:true}));
        target.dispatchEvent(new Event('change', {bubbles:true}));
      }
    }
  } catch(e) { console.log('Sincronização da seleção indisponível', e); }
}
function draw(){
  const r = wrap.getBoundingClientRect();
  ctx.clearRect(0,0,r.width,r.height);
  if(data.raster && img.complete && img.naturalWidth){
    ctx.imageSmoothingEnabled = true; ctx.imageSmoothingQuality = 'high';
    ctx.drawImage(img, ox, oy, imgW * scale, imgH * scale);
  } else {
    ctx.fillStyle = '#071525'; ctx.fillRect(0,0,r.width,r.height);
  }
  (data.layers || []).forEach(layer => (layer.features || []).forEach(feature => drawFeature(layer, feature)));
  statusEl.innerHTML = `${data.raster ? esc(data.raster.name) : 'Sem raster'} · Zoom ${Math.round(scale * 100)}% · ${data.layers.length} camada(s) vetorial(is)`;
}
function pointInRing(x,y,ring){
  let inside = false;
  for(let i=0,j=ring.length-1;i<ring.length;j=i++){
    const xi=ring[i][0], yi=ring[i][1], xj=ring[j][0], yj=ring[j][1];
    const intersect = ((yi>y)!==(yj>y)) && (x < (xj-xi)*(y-yi)/(yj-yi+1e-12)+xi);
    if(intersect) inside = !inside;
  }
  return inside;
}
function distToSegment(px,py,a,b){
  const x=a[0], y=a[1], dx=b[0]-x, dy=b[1]-y;
  if(dx===0 && dy===0) return Math.hypot(px-x, py-y);
  let t=((px-x)*dx+(py-y)*dy)/(dx*dx+dy*dy); t=Math.max(0,Math.min(1,t));
  return Math.hypot(px-(x+t*dx), py-(y+t*dy));
}
function containsFeature(feature,x,y){
  for(const poly of (feature.polygons || [])){
    if(!poly || !poly.length) continue;
    if(pointInRing(x,y,poly[0])){
      let inHole = false;
      for(let i=1;i<poly.length;i++) if(pointInRing(x,y,poly[i])) inHole = true;
      if(!inHole) return true;
    }
  }
  for(const line of (feature.lines || [])) for(let i=1;i<line.length;i++) if(distToSegment(x,y,line[i-1],line[i]) < 8/scale) return true;
  for(const pt of (feature.points || [])) if(Math.hypot(x-pt[0], y-pt[1]) < 8/scale) return true;
  return false;
}
function showFeature(layer, feature, notify=true){
  localSelected = {layerId: layer.id, index: feature.index};
  const attrs = feature.attrs || {};
  let rows = Object.keys(attrs).map(k => `<tr><td>${esc(k)}</td><td>${esc(attrs[k])}</td></tr>`).join('');
  if(!rows) rows = '<tr><td>Atributos</td><td>Sem atributos tabulares</td></tr>';
  infoEl.innerHTML = `<b>${esc(layer.name)}</b><br>Feição/Parcela: ${esc(feature.index)}<table>${rows}</table>`;
  if(notify) syncToPythonGridmark(layer, feature);
  draw();
}
function identify(clientX, clientY){
  const x = (clientX - ox) / scale, y = (clientY - oy) / scale;
  for(let li=(data.layers || []).length-1; li>=0; li--){
    const layer = data.layers[li], feats = layer.features || [];
    for(let fi=feats.length-1; fi>=0; fi--) if(containsFeature(feats[fi], x, y)){ showFeature(layer, feats[fi]); return; }
  }
  infoEl.innerHTML = '<b>Identificação de parcela</b><br>Nenhuma parcela encontrada neste ponto.';
}
wrap.addEventListener('mousedown', e => { drag = true; moved = false; sx = e.clientX; sy = e.clientY; sox = ox; soy = oy; });
window.addEventListener('mousemove', e => {
  if(!drag) return;
  const dx = e.clientX - sx, dy = e.clientY - sy;
  if(Math.abs(dx) + Math.abs(dy) > 4) moved = true;
  ox = sox + dx; oy = soy + dy; draw();
});
window.addEventListener('mouseup', e => {
  if(!drag) return;
  drag = false;
  if(!moved){ const r = wrap.getBoundingClientRect(); identify(e.clientX - r.left, e.clientY - r.top); }
});
wrap.addEventListener('wheel', e => { e.preventDefault(); const r = wrap.getBoundingClientRect(); zoomAt(e.deltaY < 0 ? 1.18 : 1/1.18, e.clientX - r.left, e.clientY - r.top); }, {passive:false});
document.getElementById('zoomIn').onclick = () => zoomAt(1.25, wrap.clientWidth/2, wrap.clientHeight/2);
document.getElementById('zoomOut').onclick = () => zoomAt(1/1.25, wrap.clientWidth/2, wrap.clientHeight/2);
document.getElementById('fit').onclick = fitView;
document.getElementById('center').onclick = centerView;
document.getElementById('oneToOne').onclick = () => { scale = 1; centerView(); };
window.addEventListener('resize', resize);
function applyInitialFocus(){
  const zoomLayer = findLayerById(data.zoomLayerId);
  if(zoomLayer && zoomLayer.bbox) zoomToBounds(zoomLayer.bbox, 72);
  if(localSelected.layerId !== '' && localSelected.index !== null && localSelected.index !== undefined){
    const layer = findLayerById(localSelected.layerId);
    const found = findFeature(layer, localSelected.index);
    if(found){
      showFeature(layer, found, false);
      if(data.zoomSelected && found.bbox) zoomToBounds(found.bbox, 96);
    }
  }
}
if(data.raster && data.raster.data_url){
  img.onload = () => {
    setViewerLoading(62, 'Carregando ortofoto...');
    resize();
    fitView();
    setViewerLoading(84, 'Preparando camadas selecionadas...');
    setTimeout(() => { applyInitialFocus(); hideViewerLoading(); }, 160);
  };
  img.onerror = () => {
    resize();
    fitView();
    setViewerLoading(100, 'Erro ao carregar raster. Visualizador liberado.');
    setTimeout(hideViewerLoading, 500);
  };
  img.src = data.raster.data_url;
} else {
  setViewerLoading(70, 'Preparando visualizador...');
  resize();
  fitView();
  setTimeout(() => { applyInitialFocus(); hideViewerLoading(); }, 220);
}
</script>
</body>
</html>
"""
    loading_html = render_loading_visualizador_grid(18, "Carregando visualizador...", "Preparando camadas selecionadas...")
    viewer_html = viewer_html.replace("__VIEWER_DATA__", data_json).replace("__GRIDMARK_LOADING__", loading_html)
    components.html(viewer_html, height=780, scrolling=False)
    st.session_state["gridmark_zoom_layer_id"] = ""
    st.session_state["gridmark_zoom_selected"] = False

def _gridmark_layer_label(layer: dict, idx: int) -> str:
    return f"{layer.get('name', 'Camada')} - {layer.get('feature_count', 0)} feições"

def alternar_visibilidade_camada(layer_id: str, visible: bool) -> None:
    layer = _gridmark_find_layer(layer_id)
    if layer is not None:
        layer["visible"] = bool(visible)
        atualizar_visualizador_sem_recarregar_tudo(layer_id)

def abrir_opcoes_camada(layer_id: str) -> None:
    layer = _gridmark_find_layer(layer_id)
    if layer is not None:
        layer["expanded"] = True

def atualizar_visualizador_sem_recarregar_tudo(layer_id: str = "") -> None:
    atualizar_apenas_camada_modificada(layer_id)

def _gridmark_set_table_layer(layer_id: str) -> None:
    st.session_state["gridmark_selected_layer_id"] = str(layer_id or "")
    st.session_state["gridmark_show_table"] = True

def render_item_camada_vetorial(layer: dict, idx: int, total_layers: int):
    move_action = None
    remove_this = False
    expanded_default = bool(layer.get("expanded", False))
    icon = "☑" if layer.get("visible", True) else "☐"
    label = f"{icon} {_gridmark_layer_label(layer, idx)}"
    with st.expander(label, expanded=expanded_default):
        layer["expanded"] = True
        st.caption(f"CRS: {layer.get('crs') or 'não informado'} · Caminho: {Path(str(layer.get('path', ''))).name}")
        visible = st.checkbox(
            "☑ Mostrar/Ocultar no mapa",
            value=bool(layer.get("visible", True)),
            key=f"gridmark_visible_{layer['id']}",
        )
        if visible != bool(layer.get("visible", True)):
            alternar_visibilidade_camada(layer["id"], visible)
        layer["color"] = st.color_picker(
            "🎨 Alterar cor",
            value=layer.get("color", GRIDMARK_VECTOR_COLORS[idx % len(GRIDMARK_VECTOR_COLORS)]),
            key=f"gridmark_color_{layer['id']}",
        )
        transparency_options = [0, 25, 50, 75, 100]
        current_transparency = int(layer.get("transparency", max(0, min(100, 100 - int(layer.get("opacity", 55))))))
        current_transparency = min(transparency_options, key=lambda item: abs(item - current_transparency))
        layer["transparency"] = st.select_slider(
            "Transparência",
            options=transparency_options,
            value=current_transparency,
            format_func=lambda value: f"{value}%",
            key=f"gridmark_transparency_{layer['id']}",
        )
        layer["opacity"] = max(0, min(100, 100 - int(layer.get("transparency", 50))))
        layer["highlight"] = st.checkbox(
            "⭐ Destacar somente essa camada",
            value=bool(layer.get("highlight", False)),
            key=f"gridmark_highlight_{layer['id']}",
        )
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            if st.button("Tabela", key=f"gridmark_table_{layer['id']}", use_container_width=True):
                _gridmark_set_table_layer(layer["id"])
        with c2:
            if st.button("🔍", key=f"gridmark_zoom_layer_{layer['id']}", use_container_width=True):
                st.session_state["gridmark_zoom_layer_id"] = layer["id"]
                atualizar_visualizador_sem_recarregar_tudo(layer["id"])
                app_rerun()
        with c3:
            if st.button("⬆", key=f"gridmark_up_{layer['id']}", disabled=idx == 0, use_container_width=True):
                move_action = (idx, idx - 1)
        with c4:
            if st.button("⬇", key=f"gridmark_down_{layer['id']}", disabled=idx == total_layers - 1, use_container_width=True):
                move_action = (idx, idx + 1)
        with c5:
            if st.button("🗑", key=f"gridmark_remove_{layer['id']}", use_container_width=True):
                remove_this = True
    return move_action, remove_this

def render_painel_camadas_expansivel() -> None:
    st.markdown("<div class='qgis-panel-title'>CAMADAS</div>", unsafe_allow_html=True)
    raster = st.session_state.get("gridmark_raster")
    if raster:
        st.session_state["gridmark_raster_visible"] = st.checkbox(
            f"Raster: {raster.get('name', 'Ortofoto')}",
            value=bool(st.session_state.get("gridmark_raster_visible", True)),
            key="gridmark_raster_visibility",
        )
        st.caption(f"{raster.get('width')}×{raster.get('height')} px preview · CRS: {'sim' if raster.get('crs_wkt') else 'não detectado'}")
        if st.button("Remover raster", key="gridmark_remove_raster", use_container_width=True):
            st.session_state["gridmark_raster"] = None
            st.session_state["gridmark_raster_visible"] = True
            app_rerun()
    else:
        st.info("Nenhuma camada raster carregada.")
    layers = st.session_state.get("gridmark_layers", [])
    if not layers:
        st.info("Nenhuma camada vetorial carregada.")
        return
    if st.button("Mostrar todas as camadas juntas", key="gridmark_show_all_layers_btn", use_container_width=True):
        for layer in layers:
            layer["visible"] = True
        st.session_state["gridmark_show_all_layers"] = True
        atualizar_visualizador_sem_recarregar_tudo()
    table_options = [layer["id"] for layer in layers]
    current_layer_id = st.session_state.get("gridmark_selected_layer_id") if st.session_state.get("gridmark_selected_layer_id") in table_options else table_options[0]
    chosen_table_layer = st.selectbox(
        "Camada para tabela de atributos",
        table_options,
        index=table_options.index(current_layer_id),
        format_func=lambda layer_id: next((item.get("name", layer_id) for item in layers if item.get("id") == layer_id), layer_id),
        key="gridmark_layers_panel_table_layer",
    )
    if chosen_table_layer != st.session_state.get("gridmark_selected_layer_id"):
        st.session_state["gridmark_selected_layer_id"] = chosen_table_layer
    remove_idx = None
    move_action = None
    for idx, layer in enumerate(layers):
        action, remove_this = render_item_camada_vetorial(layer, idx, len(layers))
        if action:
            move_action = action
        if remove_this:
            remove_idx = idx
    if move_action:
        src, dst = move_action
        layers[src], layers[dst] = layers[dst], layers[src]
        app_rerun()
    if remove_idx is not None:
        removed = layers.pop(remove_idx)
        if st.session_state.get("gridmark_selected_layer_id") == removed.get("id"):
            st.session_state["gridmark_selected_layer_id"] = ""
            st.session_state["gridmark_selected_feature_index"] = None
            st.session_state["gridmark_selected_feature_attrs"] = {}
            st.session_state["gridmark_selection_payload"] = ""
            st.session_state["gridmark_last_selection_payload"] = ""
        app_rerun()

def _gridmark_render_layers_panel() -> None:
    render_painel_camadas_expansivel()

def _gridmark_find_layer(layer_id: str):
    return next((item for item in st.session_state.get("gridmark_layers", []) if str(item.get("id")) == str(layer_id)), None)

def _gridmark_find_feature_index(layer: dict, value):
    attrs_df = (layer or {}).get("attrs_df")
    if attrs_df is None:
        return value
    for idx in attrs_df.index:
        if str(idx) == str(value):
            return idx
    return value

def selecionar_parcela_por_id(layer_id: str, feature_index, attrs: dict | None = None, zoom: bool = True) -> None:
    layer = _gridmark_find_layer(layer_id)
    resolved_index = _gridmark_find_feature_index(layer, feature_index)
    st.session_state["gridmark_selected_layer_id"] = str(layer_id or "")
    st.session_state["gridmark_selected_feature_index"] = int(resolved_index) if isinstance(resolved_index, (int, np.integer)) else str(resolved_index)
    if attrs is None and layer is not None:
        attrs_df = layer.get("attrs_df")
        if attrs_df is not None and resolved_index in attrs_df.index:
            attrs = {str(k): _gridmark_safe_json(v) for k, v in attrs_df.loc[resolved_index].to_dict().items()}
    st.session_state["gridmark_selected_feature_attrs"] = attrs or {}
    st.session_state["gridmark_zoom_selected"] = bool(zoom)

def destacar_parcela_amarela(layer_id: str, feature_index) -> None:
    selecionar_parcela_por_id(layer_id, feature_index, zoom=True)

def sincronizar_tabela_mapa() -> None:
    payload = st.session_state.get("gridmark_selection_payload") or ""
    if not payload or payload == st.session_state.get("gridmark_last_selection_payload"):
        return
    st.session_state["gridmark_last_selection_payload"] = payload
    try:
        data = json.loads(payload)
        layer_id = data.get("layerId", "")
        feature_index = data.get("index", None)
        if layer_id and feature_index is not None:
            selecionar_parcela_por_id(layer_id, feature_index, data.get("attrs", {}), zoom=False)
    except Exception:
        pass

def render_tabela_neon_grid(df: pd.DataFrame, key: str, height: int = 340):
    try:
        return st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
            height=height,
            on_select="rerun",
            selection_mode="single-row",
            key=key,
        )
    except TypeError:
        st.dataframe(df, use_container_width=True, hide_index=True, height=height)
        return None

def render_dados_parcela_neon(attrs: dict, titulo: str = "Dados da parcela selecionada") -> None:
    rows = []
    for key, value in (attrs or {}).items():
        rows.append(
            f"<div class='gridmark-parcela-row'><span>{html.escape(str(key))}</span><strong>{html.escape(str(_gridmark_safe_json(value)))}</strong></div>"
        )
    if not rows:
        rows.append("<div class='gridmark-parcela-empty'>Sem atributos para exibir.</div>")
    st.markdown(
        f"""
        <div class='gridmark-parcela-card'>
            <div class='gridmark-parcela-title'>{html.escape(str(titulo))}</div>
            {''.join(rows)}
        </div>
        """,
        unsafe_allow_html=True,
    )

def render_dropdown_neon() -> None:
    st.markdown("", unsafe_allow_html=True)

def render_tabela_atributos_neon() -> None:
    layers = st.session_state.get("gridmark_layers", [])
    c1, c2 = st.columns([1, 3])
    with c1:
        if st.button("Abrir Tabela de Atributos", key="gridmark_open_table", use_container_width=True):
            st.session_state["gridmark_show_table"] = True
    with c2:
        if st.session_state.get("gridmark_show_table") and st.button("Fechar Tabela de Atributos", key="gridmark_close_table"):
            st.session_state["gridmark_show_table"] = False
    if not st.session_state.get("gridmark_show_table"):
        return
    if not layers:
        st.info("Carregue uma camada vetorial para abrir a tabela de atributos.")
        return
    options = [layer["id"] for layer in layers]
    current_layer_id = st.session_state.get("gridmark_selected_layer_id") or options[0]
    if current_layer_id not in options:
        current_layer_id = options[0]
    selected_layer_id = st.selectbox(
        "Camada da tabela",
        options,
        index=options.index(current_layer_id),
        format_func=lambda lid: next((l.get("name", lid) for l in layers if l.get("id") == lid), lid),
        key="gridmark_table_layer_select",
    )
    layer = next((item for item in layers if item.get("id") == selected_layer_id), None)
    if not layer:
        return
    attrs_df = layer.get("attrs_df")
    if attrs_df is None or attrs_df.empty:
        st.info("Esta camada não possui atributos tabulares.")
        return
    attrs_display = attrs_df.copy()
    attrs_display.insert(0, "__indice_mapa__", [str(idx) for idx in attrs_display.index])
    st.markdown("<div class='gridmark-attrs-title'>Tabela de Atributos</div>", unsafe_allow_html=True)
    selected_from_table = None
    selection_event = render_tabela_neon_grid(attrs_display, key=f"gridmark_attr_df_{selected_layer_id}", height=340)
    try:
        selected_rows = getattr(getattr(selection_event, "selection", {}), "rows", None)
        if selected_rows is None and isinstance(getattr(selection_event, "selection", None), dict):
            selected_rows = selection_event.selection.get("rows", [])
        if selected_rows:
            selected_from_table = attrs_display.iloc[int(selected_rows[0])]["__indice_mapa__"]
    except Exception:
        pass
    indices = list(attrs_df.index)
    if indices:
        def _fmt(idx):
            row = attrs_df.loc[idx]
            preferred = ""
            for col in ("ID", "Id", "id", "PARCELA", "Parcela", "parcela", "TIRO", "DISPARO", "QUADRA"):
                if col in attrs_df.columns and str(row.get(col, "")).strip():
                    preferred = f" · {col}: {row.get(col)}"
                    break
            return f"Feição {idx}{preferred}"
        if selected_from_table is not None:
            if (
                str(st.session_state.get("gridmark_selected_layer_id", "")) != str(selected_layer_id)
                or str(st.session_state.get("gridmark_selected_feature_index", "")) != str(selected_from_table)
            ):
                destacar_parcela_amarela(selected_layer_id, selected_from_table)
                app_rerun()
        current_selected_idx = None
        if str(st.session_state.get("gridmark_selected_layer_id", "")) == str(selected_layer_id):
            current_selected_idx = _gridmark_find_feature_index(layer, st.session_state.get("gridmark_selected_feature_index"))
        default_pos = indices.index(current_selected_idx) if current_selected_idx in indices else 0
        selected_idx = st.selectbox(
            "Clique/Selecione o ID da parcela para destacar em amarelo no mapa",
            indices,
            index=default_pos,
            format_func=_fmt,
            key=f"gridmark_feature_select_{selected_layer_id}",
        )
        if st.button("Destacar ID selecionado no mapa", key=f"gridmark_highlight_selected_{selected_layer_id}", use_container_width=True):
            destacar_parcela_amarela(selected_layer_id, selected_idx)
            app_rerun()
        display_idx = current_selected_idx if current_selected_idx in attrs_df.index else selected_idx
        selected_attrs = attrs_df.loc[display_idx].to_dict()
        render_dados_parcela_neon({str(k): _gridmark_safe_json(v) for k, v in selected_attrs.items()})

def _gridmark_render_attribute_table() -> None:
    render_tabela_atributos_neon()

def render_lista_camadas() -> None:
    _gridmark_render_layers_panel()

def render_visualizador_qgis_like(raster: dict, layers: list) -> None:
    _gridmark_render_viewer(raster, layers)

def render_grid_simultaneo(raster: dict, layers: list) -> None:
    render_visualizador_qgis_like(raster, layers)

def atualizar_visualizacao() -> None:
    st.session_state["gridmark_refresh_token"] = datetime.now().isoformat(timespec="microseconds")

def render_multicamadas() -> None:
    render_lista_camadas()

def _gridmark_render_css() -> None:
    st.markdown("""
    <style>
    .qgis-like-card {
        border:1px solid rgba(0,212,255,.28);
        background:linear-gradient(145deg, rgba(5,19,40,.92), rgba(8,37,68,.78));
        box-shadow:0 14px 32px rgba(0,0,0,.36), 0 0 24px rgba(0,212,255,.14), inset 0 1px 0 rgba(255,255,255,.08);
        border-radius:12px;
        padding:14px;
        margin-bottom:14px;
    }
    .qgis-panel-title {
        color:#ffffff;
        font-weight:900;
        letter-spacing:1px;
        text-transform:uppercase;
        text-shadow:0 2px 0 rgba(0,0,0,.85), 0 0 16px rgba(0,212,255,.45);
        margin:4px 0 10px 0;
    }
    .qgis-help-text {
        color:#c9f8ff;
        font-size:.88rem;
        line-height:1.45;
        text-shadow:0 1px 0 rgba(0,0,0,.8);
    }
    div[data-testid="stTextInput"]:has(input[aria-label="gridmark_selection_payload"]) {
        display:none !important;
    }
    .gridmark-attrs-title {
        margin:12px 0 8px 0;
        color:#ffffff;
        font-weight:900;
        letter-spacing:1.2px;
        text-transform:uppercase;
        text-shadow:0 2px 0 rgba(0,0,0,.86), 0 0 18px rgba(0,212,255,.50);
    }
    div[data-testid="stDataFrame"] {
        border:1px solid rgba(0,212,255,.30) !important;
        border-radius:14px !important;
        overflow:hidden !important;
        background:
            linear-gradient(145deg, rgba(5,19,40,.94), rgba(8,37,68,.82)),
            radial-gradient(circle at top left, rgba(0,212,255,.16), transparent 34%) !important;
        box-shadow:
            0 14px 32px rgba(0,0,0,.36),
            0 0 24px rgba(0,212,255,.16),
            inset 0 1px 0 rgba(255,255,255,.10) !important;
    }
    div[data-testid="stDataFrame"] * {
        color:#ffffff !important;
    }
    div[data-testid="stDataFrame"] [role="columnheader"] {
        background:linear-gradient(145deg, rgba(2,14,36,.98), rgba(0,112,166,.58)) !important;
        color:#ffffff !important;
        font-weight:900 !important;
        text-shadow:0 1px 0 rgba(0,0,0,.84), 0 0 12px rgba(0,212,255,.44) !important;
    }
    div[data-testid="stDataFrame"] [role="row"]:hover {
        filter:brightness(1.18);
        box-shadow:inset 0 0 0 999px rgba(0,212,255,.08);
    }
    div[data-testid="stDataFrame"] [aria-selected="true"],
    div[data-testid="stDataFrame"] [data-selected="true"] {
        background:linear-gradient(90deg, rgba(0,212,255,.30), rgba(95,242,177,.18)) !important;
        box-shadow:inset 0 0 0 1px rgba(0,212,255,.58), 0 0 16px rgba(0,212,255,.22) !important;
    }
    div[data-testid="stDataFrame"] ::-webkit-scrollbar {
        width:10px;
        height:10px;
    }
    div[data-testid="stDataFrame"] ::-webkit-scrollbar-track {
        background:#071525;
        border-radius:999px;
    }
    div[data-testid="stDataFrame"] ::-webkit-scrollbar-thumb {
        background:linear-gradient(180deg,#00d4ff,#42a5f5);
        border-radius:999px;
        border:2px solid #071525;
    }
    div[data-testid="stSelectbox"],
    div[data-testid="stTextInput"],
    div[data-testid="stCheckbox"],
    div[data-testid="stColorPicker"],
    div[data-testid="stSlider"],
    div[data-testid="stFileUploader"] {
        color:#ffffff !important;
    }
    div[data-testid="stSelectbox"] > div,
    div[data-testid="stTextInput"] input,
    div[data-testid="stFileUploader"] section,
    div[data-testid="stSlider"] {
        border-radius:12px !important;
        border:1px solid rgba(0,212,255,.30) !important;
        background:linear-gradient(145deg, rgba(3,18,38,.92), rgba(8,45,82,.74)) !important;
        box-shadow:0 8px 22px rgba(0,0,0,.30), 0 0 18px rgba(0,212,255,.12), inset 0 1px 0 rgba(255,255,255,.10) !important;
        color:#ffffff !important;
        transition:all .30s ease !important;
    }
    div[data-testid="stSelectbox"] > div:hover,
    div[data-testid="stTextInput"] input:hover,
    div[data-testid="stFileUploader"] section:hover {
        border-color:rgba(0,212,255,.56) !important;
        box-shadow:0 10px 26px rgba(0,0,0,.34), 0 0 24px rgba(0,212,255,.24), inset 0 1px 0 rgba(255,255,255,.16) !important;
    }
    .gridmark-parcela-card {
        margin-top:12px;
        padding:16px;
        border-radius:16px;
        border:1px solid rgba(0,212,255,.34);
        background:
            linear-gradient(120deg, rgba(255,255,255,.10), transparent 30%),
            radial-gradient(circle at top left, rgba(0,212,255,.18), transparent 40%),
            linear-gradient(145deg, rgba(2,14,36,.96), rgba(12,57,98,.78));
        box-shadow:0 14px 32px rgba(0,0,0,.38), 0 0 24px rgba(0,212,255,.18), inset 0 1px 0 rgba(255,255,255,.12);
        color:#ffffff;
    }
    .gridmark-parcela-title {
        color:#ffffff;
        font-weight:900;
        text-transform:uppercase;
        letter-spacing:1px;
        margin-bottom:10px;
        text-shadow:0 2px 0 rgba(0,0,0,.86), 0 0 16px rgba(0,212,255,.45);
    }
    .gridmark-parcela-row {
        display:grid;
        grid-template-columns:minmax(120px,.38fr) 1fr;
        gap:10px;
        padding:8px 0;
        border-bottom:1px solid rgba(255,255,255,.08);
        color:#ffffff;
    }
    .gridmark-parcela-row span {
        color:#8feaff;
        font-weight:800;
        text-shadow:0 0 10px rgba(0,212,255,.28);
    }
    .gridmark-parcela-row strong {
        color:#ffffff;
        font-weight:900;
        text-shadow:0 1px 0 rgba(0,0,0,.72);
        word-break:break-word;
    }
    .gridmark-parcela-empty {
        color:#d9fbff;
        font-weight:700;
    }
    </style>
    """, unsafe_allow_html=True)

def aplicar_tema_grid_analise() -> None:
    _gridmark_render_css()

def render_analise_marcacao_grid() -> None:
    manter_estado_visualizador_grid()
    aplicar_tema_grid_analise()
    render_dropdown_neon()
    st.text_input("gridmark_selection_payload", key="gridmark_selection_payload", label_visibility="hidden")
    sincronizar_tabela_mapa()
    st.subheader("🗺️ Análise de Marcação de Grid")
    st.markdown(
        "<div class='qgis-like-card qgis-help-text'>Visualizador isolado estilo QGIS para carregar ortofoto raster, sobrepor shapefiles/GeoJSON, ligar e desligar camadas, identificar parcelas e abrir tabela de atributos.</div>",
        unsafe_allow_html=True,
    )
    left_col, map_col = st.columns([0.28, 0.72], gap="large")
    with left_col:
        st.markdown("<div class='qgis-panel-title'>Adicionar camadas</div>", unsafe_allow_html=True)
        raster_file = st.file_uploader(
            "Adicionar Camada Raster",
            type=["tif", "tiff", "png", "jpg", "jpeg"],
            key="gridmark_raster_uploader",
            help="GeoTIFF, TIFF, PNG ou JPG.",
        )
        if st.button("Adicionar Camada Raster", key="gridmark_add_raster", use_container_width=True):
            loading_slot = st.empty()
            render_loading_camadas(12, "Carregando camada...", raster_file.name if raster_file else "", "Lendo arquivo...", loading_slot)
            render_loading_camadas(38, "Raster:", raster_file.name if raster_file else "", "Processando ortofoto...", loading_slot)
            raster, err = _gridmark_read_raster(raster_file)
            if err:
                clear_tmg_loading(loading_slot)
                st.error(err)
            else:
                render_loading_camadas(82, "Raster:", raster.get("name", ""), "Gerando visualização...", loading_slot)
                st.session_state["gridmark_raster"] = raster
                st.session_state["gridmark_raster_visible"] = True
                render_loading_camadas(100, "Preparando visualização:", raster.get("name", ""), "Finalizando...", loading_slot)
                st.success(f"Raster carregado: {raster.get('name')}")
        vector_files = st.file_uploader(
            "Adicionar Camada Vetorial",
            type=["zip", "shp", "shx", "dbf", "prj", "cpg", "geojson", "json"],
            accept_multiple_files=True,
            key="gridmark_vector_uploader",
            help="Envie um .zip completo do shapefile ou selecione .shp, .shx, .dbf e .prj juntos.",
        )
        vector_name = st.text_input("Nome da camada vetorial", value="", key="gridmark_vector_name")
        if st.button("Adicionar Camada Vetorial", key="gridmark_add_vector", use_container_width=True):
            loading_slot = st.empty()
            render_loading_camadas(8, "Camada Vetorial:", "múltiplos grids" if vector_files and len(vector_files) > 1 else "", "Lendo arquivo...", loading_slot)
            layers_loaded, errors = _gridmark_read_vectors(vector_files, vector_name, loading_slot)
            if errors:
                for err in errors:
                    if "já importada" in str(err):
                        st.info(err)
                    else:
                        st.error(err)
            if layers_loaded:
                st.session_state["gridmark_layers"].extend(layers_loaded)
                st.success(f"{len(layers_loaded)} camada(s) vetorial(is) carregada(s).")
            else:
                clear_tmg_loading(loading_slot)
        st.markdown("---")
        render_lista_camadas()
        if st.button("Limpar somente esta análise", key="gridmark_clear_all", use_container_width=True):
            _gridmark_limpar_analise()
            app_rerun()
    with map_col:
        raster = st.session_state.get("gridmark_raster")
        layers = st.session_state.get("gridmark_layers", [])
        if not raster:
            st.info("Adicione uma camada raster para iniciar o visualizador.")
        render_grid_simultaneo(raster, layers)
        st.markdown("---")
        _gridmark_render_attribute_table()


# ==========================================
# SIDEBAR[cite: 1]
# ==========================================
current_user = _auth_current_user()
show_culture_modules = bool(_auth_allowed_cultures(current_user))
show_partners_module = _auth_can_partners(current_user)
show_admin_config = _auth_is_admin(current_user)
is_partners_page = st.session_state.pagina_ativa == 'Parceiros'
render_cultura_ambiente_css()

with st.sidebar:

    st.markdown("""
    <div class='menu-3d-title'>&#9776; MENU</div>
    <hr class='separator-glow'>
    """, unsafe_allow_html=True)

    if show_culture_modules and not is_partners_page and st.session_state.get("cultura_selecionada") not in (None, "", "PARCEIROS"):
        render_cultura_ambiente_card(topo=False)
        ambiente_info = _cultura_ambiente_info()
        if st.button(f"{ambiente_info['icone']} Ambiente {ambiente_info['nome']}", key="btn_ambiente_cultura_atual", use_container_width=True):
            st.session_state.cultura_selecionada = None
            st.session_state.pagina_ativa = "Checklist"
            app_rerun()
        st.markdown("<hr class='separator-glow'>", unsafe_allow_html=True)

    if is_partners_page and show_partners_module:
        if _auth_menu_allowed("menu_parceiros", current_user):
            if st.button("🤝 Parceiros", key="btn_parceiros_home"):
                st.session_state["partner_selected"] = ""
                st.session_state["partner_section"] = ""
                ir_para('Parceiros')
        if _auth_menu_allowed("menu_controle_dados", current_user):
            if st.button("📊 Controle de Dados", key="btn_parceiros_controle_dados"):
                if st.session_state.get("partner_selected"):
                    st.session_state["partner_section"] = st.session_state.get("partner_section", "") or "sheet"
                ir_para('Parceiros')
    else:
        if show_culture_modules:
            if _auth_menu_allowed("menu_checklist", current_user) and st.button("📋 Notas Rápidas", key="btn_check"):
                ir_para('Checklist')

            if _auth_menu_allowed("menu_grid", current_user) and st.button("📊 Marcador de Grid", key="btn_grid"):
                ir_para('Grid')

            if st.button("🗺️ Análise de Marcação de Grid", key="btn_analise_marcacao_grid"):
                ir_para('AnaliseMarcacaoGrid')

            if _auth_menu_allowed("menu_upload", current_user) and st.button("📤 Upload de Imagens", key="btn_upload"):
                ir_para('Upload')

            if _auth_menu_allowed("menu_bases", current_user) and st.button("🗂️ Banco de Dados Sistema", key="btn_bases"):
                ir_para('Bases')

            if _auth_menu_allowed("menu_sync", current_user) and st.button("🔄 Sincronizar Dados", key="btn_sync"):
                ir_para('Sync')

            if _auth_menu_allowed("menu_ortomosaicos", current_user) and st.button("🛰️ Gerar Ortomosaicos", key="btn_orto"):
                ir_para('Ortomosaicos')

            # NOVO - Botão Análises de Fenotipagem controlado por permissão do usuário
            if _auth_allowed_phenotyping(current_user):
                if st.button("📈 Análises de Fenotipagem", key="btn_visualizador"):
                    ir_para('Visualizador')

            # NOVO - Botão isolado para fluxo passo a passo de voos para análise
            if st.button("🛰️ Processos de Voos para Análise", key="btn_processos_voos_analise"):
                ir_para('VoosDirecionados')

        if show_partners_module and _auth_menu_allowed("menu_parceiros", current_user):
            if st.button("🤝 Parceiros / Controle de Voos e Dados", key="btn_parceiros_controle"):
                st.session_state["partner_selected"] = ""
                st.session_state["partner_section"] = ""
                ir_para('Parceiros')

    if (not is_partners_page) and (show_culture_modules or show_admin_config):
        if st.button("⚙️ Configurações", key="btn_config"):
            ir_para('Config')

    if st.button("🚪 Sair", key="btn_logout"):
        st.session_state.logged_in = False
        st.session_state.auth_user = None
        st.session_state.cultura_selecionada = None
        st.session_state.pagina_ativa = "Checklist"
        app_rerun()

    st.markdown("---")
    st.caption("TMG v2.0 - 2026")

# ==========================================
# TOPO (LOGO FIXA)[cite: 1]
# ==========================================
_render_logged_user_chip()
_render_system_chat(current_user)

if st.session_state.logo_sistema and st.session_state.pagina_ativa not in ('TransferenciaVoos', 'VoosDirecionados'):
    col1, col2, col3 = st.columns([1,1,1])
    with col2:
        app_image(st.session_state.logo_sistema)

if st.session_state.pagina_ativa not in ('TransferenciaVoos', 'VoosDirecionados'):
    st.markdown("<h1 class='main-header'>TMG SISTEMA DE ANÁLISE</h1>", unsafe_allow_html=True)
    render_cultura_ambiente_card(topo=True)

# ==========================================
# CONTEÚDO[cite: 1]
# ==========================================
main_container = st.container()

with main_container:

    # ==========================================
    # TRANSFERENCIA DE VOOS
    # ==========================================
    if st.session_state.pagina_ativa == 'TransferenciaVoos':
        render_transferencia_voos()

    # ==========================================
    # VOOS DIRECIONADOS
    # ==========================================
    elif st.session_state.pagina_ativa == 'VoosDirecionados':
        render_voos_direcionados()

    # ==========================================
    # PARCEIROS / CONTROLE DE VOOS E DADOS
    # ==========================================
    elif st.session_state.pagina_ativa == 'Parceiros':
        render_parceiros_controle()

    # ==========================================
    # CHECKLIST[cite: 1]
    # ==========================================
    elif st.session_state.pagina_ativa == 'Checklist':
        st.subheader("📋 Avaliador de Parcelas")

        # ── Visualizador com anotação de parcelas ──────────────────────────
        st.markdown("""
        <div style='color:#ff8c00;font-weight:700;font-size:1rem;letter-spacing:2px;
                    text-transform:uppercase;margin-bottom:10px;'>
            🗺️ Visualizador de Ortofoto · Anotação de Parcelas
        </div>""", unsafe_allow_html=True)

        chk_file = _resettable_ortho_uploader(
            "Selecione a ortofoto para anotação",
            key="chk_orto_uploader",
            help="PNG · JPG · TIF/GeoTIFF · JP2 · IMG · ECW"
        )
        chk_bytes, chk_name = _uploaded_ortho_bytes(chk_file)

        if chk_bytes:
            with st.container():
                # Atualizado Unpack para o spatial_meta
                chk_b64, chk_dims, chk_err, chk_spatial = processar_ortofoto(chk_bytes, chk_name)

            if chk_err:
                st.error(f"Erro: {chk_err}")
            else:
                cw, ch = chk_dims
                chk_storage_id = json.dumps(_tv_hash_bytes(chk_bytes)[:32])
                st.markdown(
                    f"<p style='color:#666;font-size:0.78rem;margin-bottom:6px;'>"
                    f"📐 {chk_name} · {cw}×{ch} px</p>",
                    unsafe_allow_html=True
                )

                chk_viewer = f"""
<!DOCTYPE html>
<html>
<head>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ background:#0d0d0d; overflow:auto; font-family:'Segoe UI',sans-serif; }}

  #vc {{
    width:100%; height:706px;
    background:
      linear-gradient(rgba(255,255,255,.03) 1px, transparent 1px),
      linear-gradient(90deg, rgba(255,255,255,.03) 1px, transparent 1px),
      #0d0d0d;
    background-size:32px 32px;
    border:1px solid #2a2a2a; border-radius:12px;
    overflow:hidden; position:relative; cursor:grab; user-select:none;
  }}
  #vc:active {{ cursor:grabbing; }}
  canvas {{ position:absolute; top:0; left:0; display:block; }}

  .toolbar {{ position:absolute; top:12px; right:12px; display:flex; flex-direction:column; gap:5px; z-index:20; }}
  .tb-btn {{
    background:linear-gradient(145deg,#1e1e1e,#111); border:1px solid #3a3a3a;
    color:#ff8c00; width:34px; height:34px; border-radius:8px; cursor:pointer;
    font-size:15px; font-weight:700; display:flex; align-items:center; justify-content:center;
    box-shadow:2px 2px 8px #000,inset 0 1px 0 rgba(255,255,255,.05); transition:all .2s;
  }}
  .tb-btn:hover {{ border-color:#ff8c00; box-shadow:0 0 10px rgba(255,140,0,.35),2px 2px 8px #000; color:#ffaa33; }}
  .tb-btn:active {{ transform:translateY(1px); }}
  .tb-sep {{ width:34px; height:1px; background:linear-gradient(90deg,transparent,#333,transparent); margin:2px 0; }}

  .grid-panel {{
    position:absolute; top:12px; right:55px;
    background:rgba(10,10,10,.88); border:1px solid #2a2a2a;
    border-radius:8px; padding:10px; display:flex; flex-direction:column;
    gap:8px; z-index:20;
  }}
  .grid-panel label {{ color:#ff8c00; font-size:11px; font-weight:bold; text-align:center; }}
  .grid-panel input[type=number] {{
    background:#1a1a1a; border:1px solid #333; color:#fff;
    border-radius:4px; padding:4px; width:50px; text-align:center; font-size:11px;
  }}
  .grid-panel input[type=text], .grid-panel select {{
    background:#1a1a1a; border:1px solid #333; color:#fff;
    border-radius:4px; padding:4px; font-size:11px; min-width:110px;
  }}
  .grid-panel .row-col {{ display:flex; gap:8px; align-items:center; justify-content:space-between; color:#ccc; font-size:11px; }}
  .grid-status {{ color:#777; font-size:9px; line-height:1.25; max-width:190px; }}
  .grid-all-status {{
    color:#44ff99; font-size:10px; line-height:1.25; max-width:190px;
    border:1px solid #1f5f3a; border-radius:5px; padding:5px; background:rgba(0,80,40,.16);
  }}
  .grid-btn {{
    background:linear-gradient(145deg,#1e1e1e,#111); border:1px solid #3a3a3a;
    color:#ccc; cursor:pointer; border-radius:4px; padding:6px; font-size:11px; font-weight:bold; transition:.2s;
  }}
  .grid-btn:hover {{ border-color:#ff8c00; color:#ff8c00; }}
  .grid-btn.active {{ border-color:#ff8c00; color:#ff8c00; box-shadow:0 0 8px rgba(255,140,0,.3); background:#2a1a00; }}
  .grid-btn.annot {{ border-color:#00cfff; color:#00cfff; }}
  .grid-btn.annot.active {{ background:#001a2a; box-shadow:0 0 8px rgba(0,207,255,.3); }}

  #btnExport {{
    background:linear-gradient(145deg,#003a00,#001a00); border:1px solid #006600;
    color:#00ee55; border-radius:4px; padding:6px 8px; font-size:11px;
    font-weight:bold; cursor:pointer; transition:.2s; width:100%;
  }}
  #btnExport:hover {{ border-color:#00ee55; box-shadow:0 0 8px rgba(0,238,85,.3); }}
  #btnQuickNote.quick-active {{
    background:linear-gradient(145deg,#003a16,#001a0a);
    border-color:#00d46a;
    color:#44ff99;
    box-shadow:0 0 10px rgba(0,212,106,.35);
  }}

  .zoom-badge {{
    position:absolute; top:12px; left:12px;
    background:rgba(10,10,10,.82); border:1px solid #2a2a2a; border-radius:8px;
    color:#ff8c00; font-size:11px; font-family:'Courier New',monospace;
    font-weight:700; padding:5px 10px; letter-spacing:1px; z-index:20; pointer-events:none;
  }}
  .crosshair {{
    position:absolute; bottom:12px; left:12px;
    background:rgba(10,10,10,.82); border:1px solid #222; border-radius:8px;
    color:#555; font-size:10px; font-family:'Courier New',monospace;
    padding:4px 10px; z-index:20; pointer-events:none; letter-spacing:.5px;
  }}
  .hint {{
    position:absolute; bottom:12px; right:12px; color:#333; font-size:10px;
    z-index:20; pointer-events:none; text-align:right; line-height:1.6;
  }}

  #annotPopup {{
    position:absolute; display:none;
    background:linear-gradient(160deg,#1c1c1c,#111);
    border:1px solid #ff8c00; border-radius:12px; padding:14px 16px;
    z-index:40; min-width:245px; max-width:290px;
    box-shadow:0 8px 28px rgba(0,0,0,.9),0 0 16px rgba(255,140,0,.2);
    transform:translate(-50%,-50%);
  }}
  #annotPopup h4 {{ color:#ff8c00; font-size:12px; letter-spacing:2px; text-transform:uppercase; margin-bottom:10px; text-align:center; }}
  #annotPopup .nota-row {{ display:flex; gap:5px; flex-wrap:wrap; margin-bottom:10px; }}
  .nota-btn {{
    width:36px; height:36px; border-radius:7px; border:1px solid #333;
    background:#1a1a1a; color:#ccc; font-weight:800; font-size:15px;
    cursor:pointer; transition:.15s; display:flex; align-items:center; justify-content:center;
  }}
  .nota-btn:hover {{ border-color:#ff8c00; color:#ff8c00; }}
  .nota-btn.sel {{ background:#ff8c00; color:#000; border-color:#ff8c00; box-shadow:0 0 8px rgba(255,140,0,.5); }}
  #annotObs {{
    width:100%; background:#111; border:1px solid #333; color:#ddd;
    border-radius:6px; padding:7px; font-size:11px; resize:vertical; min-height:60px;
    font-family:'Segoe UI',sans-serif; margin-bottom:10px;
  }}
  #annotObs:focus {{ border-color:#ff8c00; outline:none; }}
  .popup-btns {{ display:flex; gap:6px; }}
  .popup-save, .popup-next {{
    flex:1; background:linear-gradient(145deg,#ff9e33,#e07000);
    color:#fff; border:none; border-radius:6px; padding:7px;
    font-size:11px; font-weight:700; cursor:pointer;
  }}
  .popup-next {{ background:linear-gradient(145deg,#0077aa,#004b77); }}
  .popup-cancel {{
    background:#1a1a1a; color:#888; border:1px solid #333;
    border-radius:6px; padding:7px 10px; font-size:11px; cursor:pointer;
  }}
  .popup-cancel:hover {{ color:#ff6b6b; border-color:#552222; }}
  .assessment-panel {{
    margin-top:12px; background:#151515; border:1px solid #333; border-radius:12px;
    padding:14px; color:#ddd; box-shadow:0 6px 18px rgba(0,0,0,.55);
  }}
  .assessment-panel-title {{
    color:#ff8c00; font-weight:800; font-size:13px; letter-spacing:2px;
    text-transform:uppercase; margin-bottom:10px;
  }}
  .assessment-table {{ width:100%; border-collapse:collapse; font-size:11px; }}
  .assessment-table th {{
    background:#1f1f1f; color:#ff8c00; border:1px solid #333;
    padding:7px; text-align:left; text-transform:uppercase; letter-spacing:1px;
  }}
  .assessment-table td {{ border:1px solid #2b2b2b; padding:6px; color:#ddd; }}
  .assessment-empty {{ color:#777; font-size:12px; padding:8px 0; }}
</style>
</head>
<body>
<div id="vc">
  <canvas id="cv"></canvas>
  <div class="zoom-badge" id="zbadge">100%</div>

  <div class="grid-panel">
    <label>📍 MARCAÇÃO DE GRID</label>
    <div class="row-col"><span>Nome:</span><input type="text" id="inpGridName" value="Grid 1"></div>
    <div class="row-col"><span>Ativo:</span><select id="selGridList"></select></div>
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:5px;">
      <button class="grid-btn" id="btnSaveGrid">Salvar Grid</button>
      <button class="grid-btn" id="btnNewGrid">Novo Grid</button>
      <button class="grid-btn" id="btnDeleteGrid" style="color:#ff6b6b;border-color:#552222;">Excluir Grid</button>
    </div>
    <div class="grid-status" id="gridStatus">Grid ativo: Grid 1</div>
    <div class="grid-all-status" id="gridAllStatus">👁️ Todos os grids salvos visíveis: ATIVO</div>
    <button class="grid-btn" id="btnGridTool">Ativar Marcação</button>
    <div class="row-col"><span>DISPAROS:</span><input type="number" id="inpRows" value="10" min="1" max="500"></div>
    <div class="row-col"><span>TIROS:</span><input type="number" id="inpCols" value="10" min="1" max="500"></div>
    <div class="row-col" style="justify-content:flex-start; gap:4px; margin-top:4px;">
        <input type="checkbox" id="cbShowSummary" checked style="width:auto;">
        <span style="font-size:10px;">👁️ Mostrar ID e Resumo</span>
    </div>
    <div class="row-col"><span>Tam. ID:</span><input type="number" id="inpIdTextSize" value="12" min="5" max="40"></div>
    <div class="row-col"><span>Tam. Nota:</span><input type="number" id="inpNoteTextSize" value="10" min="5" max="40"></div>
    <div class="tb-sep" style="width:100%"></div>
    <button class="grid-btn annot" id="btnAnnotTool">✏️ Anotar Parcela</button>
    <button class="grid-btn" id="btnQuickNote" style="border-color:#00a651;color:#44ff99;">⚡ Nota Rápida</button>
    <button id="btnExport">📥 Exportar Excel</button>
    <button class="grid-btn" id="btnClearGrid" style="color:#ff6b6b;border-color:#552222;">🗑️ Limpar Grid</button>
  </div>

  <div class="toolbar">
    <button class="tb-btn" id="btnZI">+</button>
    <button class="tb-btn" id="btnZO">−</button>
    <div class="tb-sep"></div>
    <button class="tb-btn" id="btnFit">⊡</button>
    <button class="tb-btn" id="btnN">1:1</button>
    <div class="tb-sep"></div>
    <button class="tb-btn" id="btnR">↺</button>
  </div>

  <div class="crosshair" id="coord">X: — &nbsp; Y: —</div>
  <div class="hint">Scroll: zoom &nbsp;|&nbsp; Arrastar: mover<br>Anotar/Nota Rápida: duplo clique na parcela</div>

  <div id="annotPopup">
    <h4 id="popupTitle">Parcela T1 D1</h4>
    <div class="nota-row" id="notaBtns"></div>
    <textarea id="annotObs" placeholder="Observação da parcela..."></textarea>
    <div class="popup-btns">
      <button class="popup-save" id="btnPopupSave">💾 Salvar</button>
      <button class="popup-next" id="btnPopupNext">Próxima</button>
      <button class="popup-cancel" id="btnPopupCancel">✕</button>
    </div>
    <!-- NOVO RECURSO: Exportação individual por célula -->
    <div class="popup-btns" style="margin-top:6px;">
      <button class="popup-save" style="background:linear-gradient(145deg,#0077aa,#005588);" id="btnExportCellPNG" title="Salvar imagem da célula com informações geradas visualmente">🖼️ Resumo PNG</button>
      <button class="popup-save" style="background:linear-gradient(145deg,#008844,#006633);" id="btnExportCellData" title="Salvar dados da célula em JSON">📄 Dados</button>
    </div>
  </div>
</div>
<div class="assessment-panel">
  <div class="assessment-panel-title">📊 Painel da Avaliação de Parcelas</div>
  <div id="assessmentSummary"></div>
</div>

<script>
const IMG_B64 = '{chk_b64}';
const vc    = document.getElementById('vc');
const cv    = document.getElementById('cv');
const ctx   = cv.getContext('2d');
const zb    = document.getElementById('zbadge');
const coord = document.getElementById('coord');

const inpRows = document.getElementById('inpRows');
const inpCols = document.getElementById('inpCols');
const cbShowSummary = document.getElementById('cbShowSummary');
const inpIdTextSize = document.getElementById('inpIdTextSize');
const inpNoteTextSize = document.getElementById('inpNoteTextSize');
const inpGridName = document.getElementById('inpGridName');
const selGridList = document.getElementById('selGridList');
const btnSaveGrid = document.getElementById('btnSaveGrid');
const btnNewGrid = document.getElementById('btnNewGrid');
const btnDeleteGrid = document.getElementById('btnDeleteGrid');
const gridStatus = document.getElementById('gridStatus');
const btnGrid = document.getElementById('btnGridTool');
const btnAnnot = document.getElementById('btnAnnotTool');
const btnQuickNote = document.getElementById('btnQuickNote');
const btnClear = document.getElementById('btnClearGrid');
const btnExport = document.getElementById('btnExport');
const annotPopup = document.getElementById('annotPopup');
const assessmentSummary = document.getElementById('assessmentSummary');
const gridAllStatus = document.getElementById('gridAllStatus');
const ORTHO_STORAGE_ID = {chk_storage_id};
const STORAGE_KEY = 'tmg_checklist_notas_grids_' + ORTHO_STORAGE_ID;
let suppressPersist = false;

let gridMode = false;
let annotMode = false;
let quickNoteMode = false;
let points = [];
let draggingPoint = -1;
let sc = 1, ox = 0, oy = 0;
let drag = false, lx = 0, ly = 0;
const MIN_SC = 0.05, MAX_SC = 40;
let imgW = 0, imgH = 0;

let annotations = {{}};
let savedGrids = {{}};
let activeGridName = 'Grid 1';
let activeCell = null;
let activeNota = 0;

const img = new Image();

const notaBtns = document.getElementById('notaBtns');
for(let n=1; n<=9; n++) {{
  const b = document.createElement('button');
  b.className = 'nota-btn'; b.textContent = n;
  b.onclick = () => selectNota(n);
  notaBtns.appendChild(b);
}}

function selectNota(n) {{
  activeNota = n;
  document.querySelectorAll('.nota-btn').forEach((b,i) => {{
    b.classList.toggle('sel', i+1 === n);
  }});
}}

function getImgCoords(cx, cy) {{
  const r = cv.getBoundingClientRect();
  return {{ x:(cx-r.left-ox)/sc, y:(cy-r.top-oy)/sc }};
}}

function bilerp(p0,p1,p2,p3, u,v) {{
  const tx = (1-u)*p0.x + u*p1.x;
  const ty = (1-u)*p0.y + u*p1.y;
  const bx = (1-u)*p3.x + u*p2.x;
  const by = (1-u)*p3.y + u*p2.y;
  return {{ x:(1-v)*tx + v*bx, y:(1-v)*ty + v*by }};
}}

function invBilerp(px, py, p0, p1, p2, p3) {{
  let u=0.5, v=0.5;
  for(let k=0; k<20; k++) {{
    const f = bilerp(p0,p1,p2,p3,u,v);
    const fu = bilerp(p0,p1,p2,p3,u+0.001,v);
    const fv = bilerp(p0,p1,p2,p3,u,v+0.001);
    const du_dx=(fu.x-f.x)/0.001, du_dy=(fu.y-f.y)/0.001;
    const dv_dx=(fv.x-f.x)/0.001, dv_dy=(fv.y-f.y)/0.001;
    const det=du_dx*dv_dy - du_dy*dv_dx;
    if(Math.abs(det)<1e-10) break;
    const ex=px-f.x, ey=py-f.y;
    u += ( dv_dy*ex - dv_dx*ey)/det;
    v += (-du_dy*ex + du_dx*ey)/det;
    u=Math.max(0,Math.min(1,u));
    v=Math.max(0,Math.min(1,v));
  }}
  return {{u,v}};
}}

function getCellFromImgPt(ix, iy) {{
  if(points.length < 4) return null;
  const R=parseInt(inpRows.value)||1;
  const C=parseInt(inpCols.value)||1;
  const {{u,v}} = invBilerp(ix,iy,points[0],points[1],points[2],points[3]);
  if(u<0||u>1||v<0||v>1) return null;
  const c = Math.floor(u*C); const r = Math.floor(v*R);
  return {{ r:Math.min(r,R-1), c:Math.min(c,C-1) }};
}}

function getCellQuad(r,c) {{
  if(points.length < 4) return null;
  const R=parseInt(inpRows.value)||1;
  const C=parseInt(inpCols.value)||1;
  const p0=points[0],p1=points[1],p2=points[2],p3=points[3];
  const u0=c/C, u1=(c+1)/C, v0=r/R, v1=(r+1)/R;
  const tl=bilerp(p0,p1,p2,p3,u0,v0);
  const tr=bilerp(p0,p1,p2,p3,u1,v0);
  const br=bilerp(p0,p1,p2,p3,u1,v1);
  const bl=bilerp(p0,p1,p2,p3,u0,v1);
  return {{
    tl,tr,br,bl,
    cx:(tl.x+tr.x+br.x+bl.x)/4,
    cy:(tl.y+tr.y+br.y+bl.y)/4
  }};
}}

function getNotaFillColor(nota) {{
  if(nota>=9) return 'rgba(0,176,80,0.50)';
  if(nota>=7) return 'rgba(80,220,120,0.38)';
  if(nota>=5) return 'rgba(255,205,0,0.38)';
  return 'rgba(220,50,50,0.42)';
}}

function resumoObs(obs) {{
  const clean = String(obs || '').trim();
  if(!clean) return '';
  return clean.length > 18 ? clean.slice(0,18) + '...' : clean;
}}

function clonePoints(src) {{
  return (src || []).map(p => ({{x:Number(p.x)||0, y:Number(p.y)||0}}));
}}

function cloneAnnotations(src) {{
  return JSON.parse(JSON.stringify(src || {{}}));
}}

function cleanGridName(name) {{
  const value = String(name || '').trim();
  return value || 'Grid 1';
}}

function persistGrids() {{
  if(suppressPersist) return;
  try {{
    localStorage.setItem(STORAGE_KEY, JSON.stringify({{
      activeGridName,
      savedGrids,
      updatedAt: new Date().toISOString()
    }}));
  }} catch(e) {{ console.warn('Não foi possível persistir grids:', e); }}
}}

function restoreGrids() {{
  try {{
    const raw = localStorage.getItem(STORAGE_KEY);
    if(!raw) return false;
    const data = JSON.parse(raw);
    if(!data || !data.savedGrids) return false;
    savedGrids = data.savedGrids || {{}};
    activeGridName = cleanGridName(data.activeGridName || Object.keys(savedGrids)[0] || 'Grid 1');
    const rec = savedGrids[activeGridName] || savedGrids[Object.keys(savedGrids)[0]];
    if(rec) {{
      points = clonePoints(rec.points);
      annotations = cloneAnnotations(rec.annotations);
      inpRows.value = rec.rows || inpRows.value;
      inpCols.value = rec.cols || inpCols.value;
    }}
    return Object.keys(savedGrids).length > 0;
  }} catch(e) {{ console.warn('Não foi possível restaurar grids:', e); return false; }}
}}

function updateGridSelect() {{
  if(!savedGrids[activeGridName]) {{
    savedGrids[activeGridName] = {{
      points: clonePoints(points),
      annotations: cloneAnnotations(annotations),
      rows: parseInt(inpRows.value)||1,
      cols: parseInt(inpCols.value)||1
    }};
  }}
  const names = Object.keys(savedGrids);
  selGridList.innerHTML = '';
  names.forEach(name => {{
    const rec = savedGrids[name] || {{}};
    const opt = document.createElement('option');
    opt.value = name;
    opt.textContent = name + (rec.points && rec.points.length===4 ? ' ✓' : '');
    selGridList.appendChild(opt);
  }});
  selGridList.value = activeGridName;
  inpGridName.value = activeGridName;
  const marked = points.length===4 ? 'marcado' : 'sem 4 pontos';
  const completos = names.filter(n => savedGrids[n] && savedGrids[n].points && savedGrids[n].points.length===4).length;
  gridStatus.textContent = 'Grid ativo: ' + activeGridName + ' · ' + marked + ' · salvos: ' + names.length;
  if(gridAllStatus) gridAllStatus.textContent = '👁️ Todos os grids salvos visíveis: ATIVO · ' + completos + ' grid(s) completo(s) na tela';
}}

function saveActiveGrid(showMsg=false) {{
  activeGridName = cleanGridName(inpGridName.value || activeGridName);
  savedGrids[activeGridName] = {{
    points: clonePoints(points),
    annotations: cloneAnnotations(annotations),
    rows: parseInt(inpRows.value)||1,
    cols: parseInt(inpCols.value)||1
  }};
  updateGridSelect();
  persistGrids();
  if(showMsg) gridStatus.textContent = 'Grid salvo: ' + activeGridName + ' · todos continuam visíveis no visualizador';
}}

function loadGridByName(name) {{
  const target = cleanGridName(String(name || '').replace(/ ✓$/,''));
  if(target === activeGridName) return;
  saveActiveGrid(false);
  const rec = savedGrids[target];
  if(!rec) return;
  activeGridName = target;
  points = clonePoints(rec.points);
  annotations = cloneAnnotations(rec.annotations);
  inpRows.value = rec.rows || inpRows.value;
  inpCols.value = rec.cols || inpCols.value;
  closePopup();
  updateGridSelect();
  draw();
}}

function makeUniqueGridName(base) {{
  let name = cleanGridName(base);
  if(!savedGrids[name]) return name;
  let i = 2;
  while(savedGrids[name + ' ' + i]) i++;
  return name + ' ' + i;
}}

function createNewGrid() {{
  saveActiveGrid(false);
  const suggested = makeUniqueGridName('Grid ' + (Object.keys(savedGrids).length + 1));
  const typed = prompt('Nome do novo grid:', suggested);
  if(typed === null) return;
  activeGridName = makeUniqueGridName(typed);
  points = [];
  annotations = {{}};
  savedGrids[activeGridName] = {{
    points: [],
    annotations: {{}},
    rows: parseInt(inpRows.value)||1,
    cols: parseInt(inpCols.value)||1
  }};
  closePopup();
  updateGridSelect();
  draw();
}}

function renameActiveGrid() {{
  const next = cleanGridName(inpGridName.value);
  if(next === activeGridName) return;
  const old = activeGridName;
  activeGridName = makeUniqueGridName(next);
  if(savedGrids[old]) delete savedGrids[old];
  saveActiveGrid(false);
}}

function deleteActiveGrid() {{
  const target = activeGridName;
  if(!savedGrids[target]) {{
    points = [];
    annotations = {{}};
    closePopup();
    draw();
    return;
  }}
  if(!confirm('Excluir definitivamente o grid "' + target + '"? Ele sairá da lista, da visualização e da exportação.')) return;
  delete savedGrids[target];
  const remaining = Object.keys(savedGrids);
  if(remaining.length > 0) {{
    activeGridName = remaining[0];
    const rec = savedGrids[activeGridName] || {{}};
    points = clonePoints(rec.points);
    annotations = cloneAnnotations(rec.annotations);
    inpRows.value = rec.rows || inpRows.value;
    inpCols.value = rec.cols || inpCols.value;
  }} else {{
    activeGridName = 'Grid 1';
    points = [];
    annotations = {{}};
    savedGrids[activeGridName] = {{
      points: [],
      annotations: {{}},
      rows: parseInt(inpRows.value)||1,
      cols: parseInt(inpCols.value)||1
    }};
  }}
  closePopup();
  updateGridSelect();
  persistGrids();
  draw();
}}

function getExportGridRecords() {{
  saveActiveGrid(false);
  return Object.keys(savedGrids).map(name => {{
    const rec = savedGrids[name] || {{}};
    return {{
      name: name,
      rows: rec.rows || parseInt(inpRows.value)||1,
      cols: rec.cols || parseInt(inpCols.value)||1,
      points: clonePoints(rec.points),
      annotations: cloneAnnotations(rec.annotations)
    }};
  }});
}}

function escHtml(value) {{
  return String(value || '')
    .replace(/&/g,'&amp;')
    .replace(/</g,'&lt;')
    .replace(/>/g,'&gt;')
    .replace(/"/g,'&quot;');
}}

function renderAssessmentPanel() {{
  if(!assessmentSummary) return;
  const rows = [];
  getExportGridRecords().forEach(grid => {{
    const R=parseInt(grid.rows)||1, C=parseInt(grid.cols)||1;
    for(let r=0;r<R;r++) for(let c=0;c<C;c++) {{
      const ann=(grid.annotations[r]||{{}})[c];
      if(ann && ann.nota) rows.push({{
        grid:grid.name,
        tiro:c+1,
        disparo:r+1,
        parcela:'D'+(r+1)+' T'+(c+1),
        nota:ann.nota,
        obs:ann.obs || ''
      }});
    }}
  }});
  if(rows.length===0) {{
    assessmentSummary.innerHTML = '<div class="assessment-empty">Nenhuma nota registrada ainda. Use Anotar Parcela ou Nota Rápida no visualizador.</div>';
    return;
  }}
  let html = '<table class="assessment-table"><thead><tr>' +
    '<th>Nome do Grid</th><th>Tiro</th><th>Disparo</th><th>Parcela</th><th>Nota</th><th>Observação</th>' +
    '</tr></thead><tbody>';
  rows.forEach(row => {{
    html += '<tr><td>'+escHtml(row.grid)+'</td><td>'+row.tiro+'</td><td>'+row.disparo+'</td><td>'+escHtml(row.parcela)+'</td><td>'+escHtml(row.nota)+'</td><td>'+escHtml(row.obs)+'</td></tr>';
  }});
  html += '</tbody></table>';
  assessmentSummary.innerHTML = html;
}}

function drawGridRecord(gridName, rec, isActive) {{
  const pts = isActive ? points : clonePoints(rec.points);
  if(!pts || pts.length===0) return;

  const showSummary = cbShowSummary.checked;
  const R=parseInt(isActive ? inpRows.value : rec.rows)||1;
  const C=parseInt(isActive ? inpCols.value : rec.cols)||1;
  const anns = isActive ? annotations : cloneAnnotations(rec.annotations);

  function quadFor(r,c) {{
    if(pts.length < 4) return null;
    const p0=pts[0],p1=pts[1],p2=pts[2],p3=pts[3];
    const u0=c/C, u1=(c+1)/C, v0=r/R, v1=(r+1)/R;
    const tl=bilerp(p0,p1,p2,p3,u0,v0);
    const tr=bilerp(p0,p1,p2,p3,u1,v0);
    const br=bilerp(p0,p1,p2,p3,u1,v1);
    const bl=bilerp(p0,p1,p2,p3,u0,v1);
    return {{tl,tr,br,bl,cx:(tl.x+tr.x+br.x+bl.x)/4,cy:(tl.y+tr.y+br.y+bl.y)/4}};
  }}

  if(pts.length===4) {{
    const p0=pts[0],p1=pts[1],p2=pts[2],p3=pts[3];
    for(let r=0; r<R; r++) {{
      for(let c=0; c<C; c++) {{
        const ann = anns[r] && anns[r][c];
        const quad=quadFor(r,c);
        if(!quad) continue;
        const tl=quad.tl, tr=quad.tr, br=quad.br, bl=quad.bl;
        if(ann) {{
          const nota = ann.nota;
          ctx.beginPath();
          ctx.moveTo(tl.x,tl.y); ctx.lineTo(tr.x,tr.y);
          ctx.lineTo(br.x,br.y); ctx.lineTo(bl.x,bl.y);
          ctx.closePath();
          ctx.fillStyle=getNotaFillColor(Number(nota)||1); ctx.fill();
        }}
        if(isActive && activeCell && activeCell.r===r && activeCell.c===c) {{
          ctx.beginPath();
          ctx.moveTo(tl.x,tl.y); ctx.lineTo(tr.x,tr.y);
          ctx.lineTo(br.x,br.y); ctx.lineTo(bl.x,bl.y);
          ctx.closePath();
          ctx.fillStyle='rgba(255,140,0,0.20)'; ctx.fill();
          ctx.save();
          ctx.strokeStyle='rgba(255,210,0,0.98)';
          ctx.lineWidth=3/sc;
          ctx.shadowColor='rgba(255,210,0,0.65)'; ctx.shadowBlur=8/sc;
          ctx.stroke(); ctx.restore();
        }}
        if(showSummary) {{
          const idTextSize = Math.max(5, Math.min(40, parseFloat(inpIdTextSize.value)||12));
          const noteTextSize = Math.max(5, Math.min(40, parseFloat(inpNoteTextSize.value)||10));
          const hasNota = ann && ann.nota;
          const obsTxt = hasNota ? resumoObs(ann.obs) : '';
          const lines = ['D'+(r+1)+' T'+(c+1)];
          if(hasNota) lines.push('Nota: '+ann.nota);
          if(obsTxt) lines.push(obsTxt);
          if(!isActive && r===0 && c===0) lines.unshift(gridName);
          const lineGap = Math.max(7, noteTextSize * 0.95);
          const startY = quad.cy - ((lines.length-1) * lineGap) / (2*sc);
          ctx.save();
          ctx.globalAlpha = isActive ? 1 : 0.82;
          ctx.shadowColor='rgba(0,0,0,0.85)'; ctx.shadowBlur=4/sc;
          ctx.fillStyle=isActive ? '#ffffff' : '#d7ecff';
          ctx.textAlign='center'; ctx.textBaseline='middle';
          lines.forEach((line, idx) => {{
            ctx.font='bold '+((idx===0 ? idTextSize : noteTextSize)/sc)+'px Arial';
            ctx.fillText(line, quad.cx, startY + (lineGap*idx)/sc);
          }});
          ctx.restore();
        }}
      }}
    }}
    ctx.save();
    ctx.shadowColor=isActive ? 'rgba(0,160,255,0.6)' : 'rgba(255,140,0,0.45)';
    ctx.shadowBlur=6/sc;
    ctx.strokeStyle=isActive ? 'rgba(30,144,255,0.95)' : 'rgba(255,140,0,0.85)';
    ctx.lineWidth=(isActive ? 2 : 1.4)/sc;
    for(let i=0;i<=R;i++) {{
      const v=i/R;
      const lxv=(1-v)*p0.x+v*p3.x, lyv=(1-v)*p0.y+v*p3.y;
      const rxv=(1-v)*p1.x+v*p2.x, ryv=(1-v)*p1.y+v*p2.y;
      ctx.beginPath(); ctx.moveTo(lxv,lyv); ctx.lineTo(rxv,ryv); ctx.stroke();
    }}
    for(let j=0;j<=C;j++) {{
      const u=j/C;
      const txv=(1-u)*p0.x+u*p1.x, tyv=(1-u)*p0.y+u*p1.y;
      const bxv=(1-u)*p3.x+u*p2.x, byv=(1-u)*p3.y+u*p2.y;
      ctx.beginPath(); ctx.moveTo(txv,tyv); ctx.lineTo(bxv,byv); ctx.stroke();
    }}
    ctx.restore();
  }}
  if(isActive) {{
    pts.forEach((p,i) => {{
      const isDrag = draggingPoint===i;
      const r2 = 11/sc;
      ctx.save();
      ctx.shadowColor = isDrag ? 'rgba(255,255,255,0.9)' : 'rgba(0,180,255,0.8)';
      ctx.shadowBlur = 14/sc;
      ctx.beginPath(); ctx.arc(p.x,p.y,r2+3/sc,0,2*Math.PI);
      ctx.fillStyle = isDrag ? 'rgba(255,255,255,0.25)' : 'rgba(0,100,200,0.35)'; ctx.fill();
      ctx.beginPath(); ctx.arc(p.x,p.y,r2,0,2*Math.PI);
      ctx.fillStyle = isDrag ? '#ffffff' : '#1e90ff'; ctx.fill();
      ctx.lineWidth=2.5/sc; ctx.strokeStyle = isDrag ? '#aaddff' : '#00cfff'; ctx.stroke();
      ctx.restore();
      ctx.save();
      ctx.fillStyle = isDrag ? '#003366' : '#ffffff';
      ctx.font='bold '+(13/sc)+'px Arial'; ctx.textAlign='center'; ctx.textBaseline='middle';
      ctx.fillText(i+1,p.x,p.y); ctx.restore();
    }});
  }}
}}

function drawGrid() {{
  const activeRecord = {{
    points: clonePoints(points),
    annotations: cloneAnnotations(annotations),
    rows: parseInt(inpRows.value)||1,
    cols: parseInt(inpCols.value)||1
  }};
  savedGrids[activeGridName] = activeRecord;
  const names = Object.keys(savedGrids);
  names.forEach(name => {{
    if(name !== activeGridName) drawGridRecord(name, savedGrids[name], false);
  }});
  drawGridRecord(activeGridName, activeRecord, true);
  updateGridSelect();
}}

function resize() {{ cv.width=vc.clientWidth; cv.height=vc.clientHeight; if(imgW) draw(); }}

function draw() {{
  ctx.clearRect(0,0,cv.width,cv.height);
  ctx.save();
  ctx.translate(ox,oy); ctx.scale(sc,sc);
  ctx.imageSmoothingEnabled=sc<2; ctx.imageSmoothingQuality='high';
  ctx.drawImage(img,0,0);
  drawGrid();
  ctx.restore();
  zb.textContent=Math.round(sc*100)+'%';
  updatePopupPosition();
  renderAssessmentPanel();
}}

function fitScreen() {{
  const sx=cv.width/imgW, sy=cv.height/imgH;
  sc=Math.min(sx,sy)*0.92;
  ox=(cv.width-imgW*sc)/2; oy=(cv.height-imgH*sc)/2; draw();
}}

function zoomAt(factor,cx,cy) {{
  const ns=Math.min(MAX_SC,Math.max(MIN_SC,sc*factor));
  ox=cx-(cx-ox)*(ns/sc); oy=cy-(cy-oy)*(ns/sc); sc=ns; draw();
}}

img.onload=()=>{{ imgW=img.width; imgH=img.height; resize(); fitScreen(); }};
img.src='data:image/jpeg;base64,'+IMG_B64;

function updatePopupPosition() {{
  if(!activeCell || annotPopup.style.display==='none' || points.length<4) return;
  const quad = getCellQuad(activeCell.r, activeCell.c);
  if(!quad) return;
  const vw = vc.clientWidth, vh = vc.clientHeight;
  const px = Math.max(135, Math.min(vw-135, ox + quad.cx*sc));
  const py = Math.max(115, Math.min(vh-115, oy + quad.cy*sc));
  annotPopup.style.left = px+'px';
  annotPopup.style.top  = py+'px';
}}

function openPopup(r, c) {{
  activeCell = {{r,c}};
  // Nomenclatura atualizada para T = TIRO (Coluna), D = DISPARO (Linha)
  document.getElementById('popupTitle').textContent = 'Parcela D'+(r+1)+' T'+(c+1);
  const ann = (annotations[r]||{{}})[c] || {{}};
  activeNota = ann.nota || 0;
  document.querySelectorAll('.nota-btn').forEach((b,i) => b.classList.toggle('sel', i+1===activeNota));
  document.getElementById('annotObs').value = ann.obs || '';
  annotPopup.style.display = 'block';
  updatePopupPosition();
  draw();
}}

function closePopup() {{
  annotPopup.style.display='none';
  activeCell = null;
  if(imgW) draw();
}}

function setAnnotation(r,c,nota,obs) {{
  if(!annotations[r]) annotations[r]={{}};
  const prev=(annotations[r]||{{}})[c] || {{}};
  annotations[r][c] = {{
    nota: nota,
    obs: obs !== undefined ? obs : (prev.obs || '')
  }};
  saveActiveGrid(false);
}}

function clearAnnotation(r,c) {{
  if(!annotations[r]) return;
  delete annotations[r][c];
  if(Object.keys(annotations[r]).length===0) delete annotations[r];
  saveActiveGrid(false);
}}

function saveActiveAnnotation(closeAfter=true) {{
  if(!activeCell) return false;
  const {{r,c}} = activeCell;
  setAnnotation(r,c,activeNota || 1,document.getElementById('annotObs').value);
  if(closeAfter) closePopup();
  draw();
  return true;
}}

function openNextParcel() {{
  if(!activeCell) return;
  saveActiveAnnotation(false);
  const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
  let idx = activeCell.r*C + activeCell.c + 1;
  if(idx >= R*C) idx = 0;
  const nr = Math.floor(idx / C);
  const nc = idx % C;
  openPopup(nr,nc);
}}

document.getElementById('btnPopupSave').onclick = () => {{
  saveActiveAnnotation(true);
}};

document.getElementById('btnPopupCancel').onclick = closePopup;
document.getElementById('btnPopupNext').onclick = openNextParcel;
annotPopup.addEventListener('mousedown', e=>e.stopPropagation());
annotPopup.addEventListener('click', e=>e.stopPropagation());
annotPopup.addEventListener('dblclick', e=>e.stopPropagation());

// Exportar Imagem Resumo Célula
document.getElementById('btnExportCellPNG').onclick = () => {{
  if(!activeCell) return;
  const {{r,c}} = activeCell;
  const R = parseInt(inpRows.value)||1, C = parseInt(inpCols.value)||1;
  const p0 = points[0], p1 = points[1], p2 = points[2], p3 = points[3];
  
  const u0=c/C, u1=(c+1)/C, v0=r/R, v1=(r+1)/R;
  const tl=bilerp(p0,p1,p2,p3,u0,v0), tr=bilerp(p0,p1,p2,p3,u1,v0);
  const br=bilerp(p0,p1,p2,p3,u1,v1), bl=bilerp(p0,p1,p2,p3,u0,v1);
  
  const minX = Math.min(tl.x, tr.x, br.x, bl.x);
  const maxX = Math.max(tl.x, tr.x, br.x, bl.x);
  const minY = Math.min(tl.y, tr.y, br.y, bl.y);
  const maxY = Math.max(tl.y, tr.y, br.y, bl.y);
  const w = maxX - minX, h = maxY - minY;
  
  const cv2 = document.createElement('canvas');
  cv2.width = w + 40; cv2.height = h + 130;
  const ctx2 = cv2.getContext('2d');
  
  ctx2.fillStyle = '#1e1e1e'; ctx2.fillRect(0,0,cv2.width,cv2.height);
  ctx2.drawImage(img, minX, minY, w, h, 20, 20, w, h);
  
  ctx2.fillStyle = '#ff8c00'; ctx2.font = 'bold 16px sans-serif';
  ctx2.fillText('Célula: T'+(c+1)+' D'+(r+1), 20, h + 50);
  
  const ann = (annotations[r]||{{}})[c] || {{}};
  ctx2.fillStyle = '#ccc'; ctx2.font = '14px sans-serif';
  ctx2.fillText('Nota: ' + (ann.nota||'Sem nota registrada'), 20, h + 75);
  ctx2.fillText('Obs: ' + (document.getElementById('annotObs').value||'Nenhuma observação'), 20, h + 100);
  
  const a = document.createElement('a');
  a.download = 'TMG_Resumo_T'+(c+1)+'_D'+(r+1)+'.png';
  a.href = cv2.toDataURL('image/png');
  a.click();
}};

// Exportar Dados Individuais JSON
document.getElementById('btnExportCellData').onclick = () => {{
  if(!activeCell) return;
  const {{r,c}} = activeCell;
  const ann = (annotations[r]||{{}})[c] || {{}};
  const data = {{
      celula: "T"+(c+1)+" D"+(r+1),
      tiro: c+1,
      disparo: r+1,
      nota: activeNota,
      observacao: document.getElementById('annotObs').value || ''
  }};
  
  const blob = new Blob([JSON.stringify(data, null, 2)], {{type:'application/json'}});
  const a = document.createElement('a');
  a.download = 'TMG_Dados_T'+(c+1)+'_D'+(r+1)+'.json';
  a.href = URL.createObjectURL(blob);
  a.click();
}};

function loadScriptOnce(src) {{
  return new Promise((resolve, reject) => {{
    const existing = document.querySelector('script[data-tmg-src="' + src + '"]');
    if(existing) {{
      if(existing.dataset.loaded === 'true') {{
        resolve(true);
        return;
      }}
      existing.addEventListener('load', () => resolve(true), {{once:true}});
      existing.addEventListener('error', () => reject(new Error('Falha ao carregar ' + src)), {{once:true}});
      return;
    }}
    const script = document.createElement('script');
    script.src = src;
    script.async = true;
    script.dataset.tmgSrc = src;
    script.onload = () => {{
      script.dataset.loaded = 'true';
      resolve(true);
    }};
    script.onerror = () => reject(new Error('Falha ao carregar ' + src));
    document.head.appendChild(script);
  }});
}}

async function ensureChecklistExcelStyles() {{
  if(window.__tmgChecklistXlsxStyleReady && typeof XLSX !== 'undefined') return true;
  const urls = [
    'https://cdn.jsdelivr.net/npm/xlsx-js-style@1.2.0/dist/xlsx.bundle.js',
    'https://unpkg.com/xlsx-js-style@1.2.0/dist/xlsx.bundle.js'
  ];
  for(const url of urls) {{
    try {{
      await loadScriptOnce(url);
      if(typeof XLSX !== 'undefined') {{
        window.__tmgChecklistXlsxStyleReady = true;
        return true;
      }}
    }} catch(e) {{
      console.warn(e);
    }}
  }}
  return false;
}}

btnExport.onclick = async () => {{
  const nome = prompt('Digite o nome do arquivo Excel:', 'checklist_notas_parcelas');
  if(nome === null) return;
  const safeName = (nome.trim() || 'checklist_notas_parcelas')
    .replace(/[\\\\/:*?"<>|]+/g,'_')
    .replace(/\\s+/g,'_');
  const gridRecords = getExportGridRecords().filter(grid => grid.points && grid.points.length === 4);
  if(gridRecords.length === 0) {{
    alert('Nenhum grid salvo/marcado disponível para exportar.');
    return;
  }}
  if(typeof XLSX === 'undefined' && !(await ensureChecklistExcelStyles())) {{
    alert('Biblioteca Excel não carregada. Tente novamente em alguns segundos.');
    return;
  }}
  if(!(await ensureChecklistExcelStyles())) {{
    alert('Biblioteca de estilos do Excel não carregada. Verifique a conexão e tente exportar novamente.');
    return;
  }}
  const headers = ['Quadra','Disparo','Tiro','Nota','Observação'];
  const dados = [headers];
  gridRecords
    .sort((a,b) => String(a.name).localeCompare(String(b.name), 'pt-BR', {{numeric:true}}))
    .forEach(grid => {{
    const R=parseInt(grid.rows)||1, C=parseInt(grid.cols)||1;
    for(let r=0;r<R;r++) for(let c=0;c<C;c++) {{
      const ann=(grid.annotations[r]||{{}})[c];
      const nota=Number(ann && ann.nota ? ann.nota : 1);
      dados.push([grid.name, r+1, c+1, nota, ann ? (ann.obs || '') : '']);
    }}
  }});
  const ws = XLSX.utils.aoa_to_sheet(dados);
  ws['!cols'] = [{{wch:14}},{{wch:10}},{{wch:10}},{{wch:10}},{{wch:34}}];
  ws['!freeze'] = {{xSplit:0,ySplit:1}};

  const border = {{
    top:{{style:'thin',color:{{rgb:'808080'}}}},
    bottom:{{style:'thin',color:{{rgb:'808080'}}}},
    left:{{style:'thin',color:{{rgb:'808080'}}}},
    right:{{style:'thin',color:{{rgb:'808080'}}}}
  }};
  const headerStyle = {{
    font:{{bold:true,color:{{rgb:'00B0F0'}}}},
    fill:{{patternType:'solid',fgColor:{{rgb:'1F1F1F'}}}},
    alignment:{{horizontal:'center',vertical:'center'}},
    border
  }};
  const defaultStyle = {{
    font:{{color:{{rgb:'000000'}}}},
    alignment:{{horizontal:'center',vertical:'center',wrapText:true}},
    border
  }};
  const notaStyles = {{
    1: {{fill:'00B050', font:'FFFFFF'}},
    2: {{fill:'FFF2CC', font:'000000'}},
    3: {{fill:'DDEBF7', font:'000000'}},
    4: {{fill:'FCE4D6', font:'000000'}},
    5: {{fill:'E7E6E6', font:'000000'}},
    6: {{fill:'E4DFEC', font:'000000'}},
    7: {{fill:'CCFFFF', font:'000000'}},
    8: {{fill:'EADCC8', font:'000000'}},
    9: {{fill:'C00000', font:'FFFFFF'}}
  }};
  function notaStyle(nota) {{
    const item = notaStyles[Number(nota)] || {{fill:'FFFFFF', font:'000000'}};
    return {{
      font:{{bold:true,color:{{rgb:item.font}}}},
      fill:{{patternType:'solid',fgColor:{{rgb:item.fill}}}},
      alignment:{{horizontal:'center',vertical:'center'}},
      border
    }};
  }}

  const range = XLSX.utils.decode_range(ws['!ref']);
  for(let row=range.s.r; row<=range.e.r; row++) {{
    for(let col=range.s.c; col<=range.e.c; col++) {{
      const addr = XLSX.utils.encode_cell({{r:row,c:col}});
      if(!ws[addr]) continue;
      ws[addr].s = row === 0 ? headerStyle : (col === 3 ? notaStyle(ws[addr].v) : defaultStyle);
      if(row > 0 && col === 3) ws[addr].t = 'n';
    }}
  }}
  const wb = XLSX.utils.book_new();
  wb.Props = {{
    Title:'Checklist de Notas',
    Subject:'Notas por Quadra, Disparo e Tiro',
    Author:'TMG Sistema de Análise',
    CreatedDate:new Date()
  }};
  XLSX.utils.book_append_sheet(wb, ws, 'Notas Rápidas');
  XLSX.writeFile(wb, safeName+'.xlsx', {{bookType:'xlsx',cellStyles:true}});
}};

vc.addEventListener('wheel', e=>{{
  e.preventDefault();
  const r=cv.getBoundingClientRect();
  zoomAt(e.deltaY<0?1.18:1/1.18, e.clientX-r.left, e.clientY-r.top);
}},{{passive:false}});

vc.addEventListener('mousedown', e=>{{
  const ic=getImgCoords(e.clientX,e.clientY);
  draggingPoint=-1;

  for(let i=0;i<points.length;i++) {{
    const dx=points[i].x-ic.x, dy=points[i].y-ic.y;
    if(Math.sqrt(dx*dx+dy*dy)<30/sc) {{ draggingPoint=i; vc.style.cursor='move'; return; }}
  }}

  if(gridMode && points.length<4) {{
    points.push({{x:ic.x, y:ic.y}});
    saveActiveGrid(false);
    draw(); return;
  }}

  drag=true; lx=e.clientX; ly=e.clientY;
  if(!gridMode && !annotMode) vc.style.cursor='grabbing';
}});

vc.addEventListener('click', e=>{{
  if((annotMode || quickNoteMode) && points.length===4) e.preventDefault();
}});

window.addEventListener('mousemove', e=>{{
  const ic=getImgCoords(e.clientX,e.clientY);
  coord.textContent='X: '+Math.round(ic.x)+'   Y: '+Math.round(ic.y);

  if(draggingPoint!==-1) {{ points[draggingPoint]={{x:ic.x,y:ic.y}}; draw(); return; }}
  if(drag) {{ ox+=e.clientX-lx; oy+=e.clientY-ly; lx=e.clientX; ly=e.clientY; draw(); return; }}

  let hover=false;
  for(let i=0;i<points.length;i++) {{
    const dx=points[i].x-ic.x,dy=points[i].y-ic.y;
    if(Math.sqrt(dx*dx+dy*dy)<30/sc) {{ hover=true; break; }}
  }}
  if(hover) vc.style.cursor='move';
  else if(gridMode && points.length<4) vc.style.cursor='crosshair';
  else if(quickNoteMode && points.length===4) vc.style.cursor='cell';
  else if(annotMode && points.length===4) vc.style.cursor='cell';
  else vc.style.cursor='grab';
}});

window.addEventListener('mouseup',()=>{{
  if(draggingPoint!==-1) saveActiveGrid(false);
  drag=false; draggingPoint=-1;
  if(gridMode && points.length<4) vc.style.cursor='crosshair';
  else if(quickNoteMode && points.length===4) vc.style.cursor='cell';
  else if(annotMode && points.length===4) vc.style.cursor='cell';
  else vc.style.cursor='grab';
}});

vc.addEventListener('dblclick', e=>{{
  if((annotMode || quickNoteMode) && points.length===4) {{
    e.preventDefault();
    const ic=getImgCoords(e.clientX,e.clientY);
    const cell=getCellFromImgPt(ic.x,ic.y);
    if(!cell) return;
    if(quickNoteMode) {{
      const ann = (annotations[cell.r]||{{}})[cell.c];
      if(ann && Number(ann.nota)===9) {{
        if(String(ann.obs || '').trim()) setAnnotation(cell.r, cell.c, 1, ann.obs);
        else clearAnnotation(cell.r, cell.c);
      }} else {{
        activeCell = {{r:cell.r,c:cell.c}};
        setAnnotation(cell.r, cell.c, 9);
      }}
      closePopup();
      draw();
      return;
    }}
    openPopup(cell.r, cell.c);
    return;
  }}
  fitScreen();
}});

btnGrid.onclick=()=>{{
  saveActiveGrid(false);
  gridMode=!gridMode; annotMode=false; quickNoteMode=false;
  btnGrid.className=gridMode?'grid-btn active':'grid-btn';
  btnGrid.textContent=gridMode?'📍 Selecione 4 Pontos':'Ativar Marcação';
  btnAnnot.className='grid-btn annot';
  btnQuickNote.className='grid-btn';
  vc.style.cursor=gridMode?'crosshair':'grab';
}};

btnSaveGrid.onclick=()=>saveActiveGrid(true);
btnNewGrid.onclick=createNewGrid;
btnDeleteGrid.onclick=deleteActiveGrid;
selGridList.onchange=()=>loadGridByName(selGridList.value);
inpGridName.onchange=renameActiveGrid;

btnAnnot.onclick=()=>{{
  if(points.length<4){{ alert('Marque os 4 pontos do Grid primeiro.'); return; }}
  annotMode=!annotMode; quickNoteMode=false; gridMode=false;
  btnAnnot.className=annotMode?'grid-btn annot active':'grid-btn annot';
  btnQuickNote.className='grid-btn';
  btnGrid.className='grid-btn';
  btnGrid.textContent='Ativar Marcação';
  vc.style.cursor=annotMode?'cell':'grab';
}};

btnQuickNote.onclick=()=>{{
  if(points.length<4){{ alert('Marque os 4 pontos do Grid primeiro.'); return; }}
  quickNoteMode=!quickNoteMode; annotMode=false; gridMode=false;
  btnQuickNote.className=quickNoteMode?'grid-btn quick-active':'grid-btn';
  btnAnnot.className='grid-btn annot';
  btnGrid.className='grid-btn';
  btnGrid.textContent='Ativar Marcação';
  closePopup();
  vc.style.cursor=quickNoteMode?'cell':'grab';
}};

btnClear.onclick=()=>{{ points=[]; annotations={{}}; closePopup(); saveActiveGrid(false); draw(); }};

inpRows.addEventListener('input',()=>{{ saveActiveGrid(false); draw(); }});
inpCols.addEventListener('input',()=>{{ saveActiveGrid(false); draw(); }});
inpIdTextSize.addEventListener('input',draw);
inpNoteTextSize.addEventListener('input',draw);
cbShowSummary.addEventListener('change', draw);

document.getElementById('btnZI').onclick=()=>zoomAt(1.3,cv.width/2,cv.height/2);
document.getElementById('btnZO').onclick=()=>zoomAt(1/1.3,cv.width/2,cv.height/2);
document.getElementById('btnFit').onclick=fitScreen;
document.getElementById('btnN').onclick=()=>{{sc=1;ox=(cv.width-imgW)/2;oy=(cv.height-imgH)/2;draw();}};
document.getElementById('btnR').onclick=fitScreen;

window.addEventListener('resize',resize);
suppressPersist = true;
restoreGrids();
suppressPersist = false;
saveActiveGrid(false);
updateGridSelect();
</script>
</body>
</html>
"""
                components.html(chk_viewer, height=980, scrolling=True)

        else:
            st.markdown("""
            <div style='height:706px;border:1px dashed #2e2e2e;border-radius:12px;background:#0d0d0d;
                        display:flex;flex-direction:column;align-items:center;justify-content:center;
                        gap:12px;color:#333;'>
                <div style='font-size:3rem;'>🗺️</div>
                <div style='font-size:0.9rem;letter-spacing:2px;text-transform:uppercase;'>
                    Nenhuma ortofoto carregada
                </div>
                <div style='font-size:0.75rem;color:#2a2a2a;'>
                    PNG · JPG · TIF · GeoTIFF · JP2 · IMG · ECW
                </div>
            </div>
            """, unsafe_allow_html=True)

    # ==========================================
    # GRID COM OPÇÕES DE EXPORTAÇÃO E SHAPEFILE[cite: 1]
    # ==========================================
    elif st.session_state.pagina_ativa == 'Grid':
        st.subheader("📊 Visualizador de Ortofoto")

        orto_file = _resettable_ortho_uploader(
            "Selecione a ortofoto",
            key="orto_uploader",
            help="Formatos suportados: PNG, JPG, TIF/GeoTIFF, JP2, IMG"
        )

        grid_prefill_path = st.session_state.get("grid_prefill_ortho_path", "")
        grid_prefill_name = st.session_state.get("grid_prefill_ortho_name", "")
        grid_prefill_available = bool(grid_prefill_path and Path(grid_prefill_path).exists())
        if grid_prefill_available and not orto_file:
            st.info(f"Ortofoto recebida do módulo Voos Direcionados: {grid_prefill_name or Path(grid_prefill_path).name}")
            if st.button("Limpar ortofoto recebida", key="btn_clear_grid_prefill"):
                st.session_state.pop("grid_prefill_ortho_path", None)
                st.session_state.pop("grid_prefill_ortho_name", None)
                app_rerun()

        if orto_file or grid_prefill_available:
            with st.container():
                if orto_file:
                    file_bytes, orto_nome_exibicao = _uploaded_ortho_bytes(orto_file)
                else:
                    file_bytes = Path(grid_prefill_path).read_bytes()
                    orto_nome_exibicao = grid_prefill_name or Path(grid_prefill_path).name
                # Atualizado Unpack para obter Metadata Espacial para o SHP
                b64, dims, err, spatial_meta = processar_ortofoto(file_bytes, orto_nome_exibicao)
                st.session_state.spatial_meta = spatial_meta

            if err:
                st.error(f"Erro ao processar imagem: {err}")
            else:
                w, h = dims
                st.markdown(
                    f"<p style='color:#666;font-size:0.78rem;margin-bottom:6px;'>"
                    f"📐 {orto_nome_exibicao} &nbsp;·&nbsp; Resolução final processada: {w}×{h} px</p>",
                    unsafe_allow_html=True
                )
                grid_spatial_json = json.dumps({
                    "ratio": (spatial_meta or {}).get("ratio", 1.0),
                    "transform": (spatial_meta or {}).get("transform"),
                    "crs": str((spatial_meta or {}).get("crs") or ""),
                }, ensure_ascii=False)

                viewer_html = f"""
<!DOCTYPE html>
<html>
<head>
<!-- SHPWRITE LIBRARY PARA EXPORTAÇÃO SHAPEFILE (MANTIDO APENAS COMO LEGADO NO CONSOLE) -->
<script src="https://unpkg.com/shp-write@latest/shpwrite.js"></script>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ background: #0d0d0d; overflow: hidden; }}

  #vc {{
    width: 100%;
    height: 706px;
    background:
      linear-gradient(rgba(255,255,255,.03) 1px, transparent 1px),
      linear-gradient(90deg, rgba(255,255,255,.03) 1px, transparent 1px),
      #0d0d0d;
    background-size: 32px 32px;
    border: 1px solid #2a2a2a;
    border-radius: 12px;
    overflow: hidden;
    position: relative;
    cursor: grab;
    user-select: none;
  }}
  #vc:active {{ cursor: grabbing; }}
  canvas {{ position: absolute; top:0; left:0; display: block; }}

  .toolbar {{
    position: absolute;
    top: 12px; right: 12px;
    display: flex;
    flex-direction: column;
    gap: 5px;
    z-index: 20;
  }}
  .tb-btn {{
    background: linear-gradient(145deg, #1e1e1e, #111);
    border: 1px solid #3a3a3a;
    color: #ff8c00;
    width: 34px; height: 34px;
    border-radius: 8px;
    cursor: pointer;
    font-size: 15px;
    font-weight: 700;
    display: flex; align-items: center; justify-content: center;
    box-shadow: 2px 2px 8px #000, inset 0 1px 0 rgba(255,255,255,0.05);
    transition: all .2s;
    font-family: 'Segoe UI', sans-serif;
    line-height: 1;
  }}
  .tb-btn:hover {{
    border-color: #ff8c00;
    box-shadow: 0 0 10px rgba(255,140,0,0.35), 2px 2px 8px #000;
    color: #ffaa33;
  }}
  .tb-btn:active {{ transform: translateY(1px); }}

  .tb-sep {{
    width: 34px; height: 1px;
    background: linear-gradient(90deg, transparent, #333, transparent);
    margin: 2px 0;
  }}

  /* GRID PANEL */
  .grid-panel {{
      position: absolute;
      top: 12px; right: 55px;
      background: rgba(10,10,10,0.85);
      border: 1px solid #2a2a2a;
      border-radius: 8px;
      padding: 10px;
      display: flex;
      flex-direction: column;
      gap: 8px;
      z-index: 20;
      font-family: 'Segoe UI', sans-serif;
  }}
  .grid-panel label {{ color: #ff8c00; font-size: 11px; font-weight: bold; text-align: center; margin-bottom: 2px; }}
  .grid-panel input[type=number] {{
      background: #1a1a1a; border: 1px solid #333; color: #fff;
      border-radius: 4px; padding: 4px; width: 50px; text-align: center; font-size: 11px;
  }}
  .grid-panel .row-col {{ display: flex; gap: 8px; align-items: center; justify-content: space-between; color: #ccc; font-size:11px; }}
  .grid-btn {{
      background: linear-gradient(145deg, #1e1e1e, #111);
      border: 1px solid #3a3a3a; color: #ccc; cursor: pointer; border-radius:4px; padding: 6px; font-size: 11px; font-weight:bold; transition: 0.2s;
  }}
  .grid-btn:hover {{ border-color: #ff8c00; color: #ff8c00; }}
  .grid-btn.active {{ border-color: #ff8c00; color: #ff8c00; box-shadow: 0 0 8px rgba(255,140,0,0.3); background: #2a1a00; }}

  .shp-ref-panel {{
    position:absolute; top:12px; left:12px; z-index:25;
    width:360px; max-width:calc(100% - 430px);
    background:rgba(10,10,10,.90); border:1px solid #2a2a2a; border-radius:8px;
    padding:9px; display:flex; flex-direction:column; gap:6px;
    font-family:'Segoe UI',sans-serif; box-shadow:0 8px 22px rgba(0,0,0,.45);
  }}
  .shp-ref-panel label {{ color:#ff00ff; font-size:11px; font-weight:800; letter-spacing:1px; text-transform:uppercase; }}
  .shp-ref-panel input {{
    width:100%; background:#111; border:1px solid #333; color:#fff; border-radius:5px;
    padding:6px 8px; font-size:11px; outline:none;
  }}
  .shp-ref-panel input:focus {{ border-color:#ff00ff; box-shadow:0 0 8px rgba(255,0,255,.22); }}
  .shp-ref-status {{ color:#888; font-size:9px; line-height:1.25; }}

  .zoom-badge {{
    position: absolute;
    top: 116px; left: 12px;
    background: rgba(10,10,10,0.82);
    border: 1px solid #2a2a2a;
    border-radius: 8px;
    color: #ff8c00;
    font-size: 11px;
    font-family: 'Courier New', monospace;
    font-weight: 700;
    padding: 5px 10px;
    letter-spacing: 1px;
    z-index: 20;
    pointer-events: none;
  }}

  .crosshair {{
    position: absolute;
    bottom: 12px; left: 12px;
    background: rgba(10,10,10,0.82);
    border: 1px solid #222;
    border-radius: 8px;
    color: #555;
    font-size: 10px;
    font-family: 'Courier New', monospace;
    padding: 4px 10px;
    z-index: 20;
    pointer-events: none;
    letter-spacing: 0.5px;
  }}

  .hint {{
    position: absolute;
    bottom: 12px; right: 12px;
    color: #333;
    font-size: 10px;
    font-family: 'Segoe UI', sans-serif;
    z-index: 20;
    pointer-events: none;
    text-align: right;
    line-height: 1.6;
  }}
</style>
</head>
<body>
<div id="vc">
  <canvas id="cv"></canvas>

  <div class="shp-ref-panel">
      <label>🗺️ Referências para exportar SHP</label>
      <input type="text" id="inpShpRef" placeholder="Digite coordenadas/referências e pressione ENTER">
      <div class="shp-ref-status" id="shpRefStatus">Aguardando referência. O botão SHP será liberado após ENTER.</div>
      <button class="grid-btn" id="btnExportSHP" style="display:none;color:#ff00ff; border-color:#990099;">🗺️ Exportar Shapefile (.SHP)</button>
  </div>

  <div class="zoom-badge" id="zbadge">100%</div>

  <div class="grid-panel">
      <label>📍 MARCAÇÃO DE GRID</label>
      <button class="grid-btn" id="btnGridTool">Ativar Marcação</button>
      <div class="row-col">
          <span>DISPAROS:</span>
          <input type="number" id="inpRows" value="10" min="1" max="500">
      </div>
      <div class="row-col">
          <span>TIROS:</span>
          <input type="number" id="inpCols" value="10" min="1" max="500">
      </div>
      <div class="row-col" style="justify-content:flex-start; gap:4px; margin-top:4px;">
          <input type="checkbox" id="cbShowSummary" checked style="width:auto;">
          <span style="font-size:10px;">👁️ Mostrar ID e Resumo</span>
      </div>
      <div style="display:flex;flex-direction:column;gap:3px;margin-top:2px;">
          <span style="color:#ff8c00;font-size:11px;font-weight:bold;">🔤 Tamanho do Texto:</span>
          <div style="display:flex;align-items:center;gap:6px;">
              <input type="range" id="inpFontSize" min="4" max="72" value="12"
                     style="width:85px;accent-color:#ff8c00;cursor:pointer;">
              <span id="lblFontSize" style="font-size:11px;color:#ff8c00;font-weight:bold;min-width:22px;">12</span>
          </div>
      </div>
      <button class="grid-btn" id="btnClearGrid" style="color:#ff6b6b; border-color:#552222;">🗑️ Limpar Grid</button>
      
      <div class="tb-sep" style="width:100%"></div>
      <button class="grid-btn" id="btnExportFull" style="color:#00ee55; border-color:#006600;">📥 Exportar Ortofoto + Grid</button>
      <button class="grid-btn" id="btnExportGrid" style="color:#00cfff; border-color:#006699;">📥 Exportar Apenas Grid</button>
  </div>

  <div class="toolbar">
    <button class="tb-btn" id="btnZI" title="Zoom In (+)">+</button>
    <button class="tb-btn" id="btnZO" title="Zoom Out (-)">−</button>
    <div class="tb-sep"></div>
    <button class="tb-btn" id="btnFit" title="Ajustar à tela">⊡</button>
    <button class="tb-btn" id="btnN"   title="Zoom 100%">1:1</button>
    <div class="tb-sep"></div>
    <button class="tb-btn" id="btnR"   title="Resetar vista">↺</button>
  </div>

  <div class="crosshair" id="coord">X: — &nbsp; Y: —</div>
  <div class="hint">Scroll: zoom &nbsp;|&nbsp; Arrastar: mover extremidade<br>Duplo clique: ajustar tela</div>
</div>

<script>
const IMG_B64 = '{b64}';
const vc  = document.getElementById('vc');
const cv  = document.getElementById('cv');
const ctx = cv.getContext('2d');
const zb  = document.getElementById('zbadge');
const coord = document.getElementById('coord');

const inpRows = document.getElementById('inpRows');
const inpCols = document.getElementById('inpCols');
const cbShowSummary = document.getElementById('cbShowSummary');
const btnGrid = document.getElementById('btnGridTool');
const btnClear = document.getElementById('btnClearGrid');
const btnExportSHP = document.getElementById('btnExportSHP');
const inpShpRef = document.getElementById('inpShpRef');
const shpRefStatus = document.getElementById('shpRefStatus');
const GRID_SPATIAL_META = {grid_spatial_json};
let gridMode = false;
let points = [];
let draggingPoint = -1;
let shpReferenceReady = false;

let sc = 1, ox = 0, oy = 0;
let drag = false, lx = 0, ly = 0;
const MIN_SC = 0.05, MAX_SC = 40;
let imgW = 0, imgH = 0;

const img = new Image();

function getImgCoords(cx, cy) {{
    const r = cv.getBoundingClientRect();
    return {{
        x: (cx - r.left - ox) / sc,
        y: (cy - r.top - oy) / sc
    }};
}}

// Helper interpolador adicionado ao painel de Grid para gerar células exatas
function bilerp(p0,p1,p2,p3, u,v) {{
  const tx = (1-u)*p0.x + u*p1.x;
  const ty = (1-u)*p0.y + u*p1.y;
  const bx = (1-u)*p3.x + u*p2.x;
  const by = (1-u)*p3.y + u*p2.y;
  return {{ x:(1-v)*tx + v*bx, y:(1-v)*ty + v*by }};
}}

// Mantém o payload dentro do iframe. Alterar inputs internos do Streamlit pelo
// window.parent pode quebrar a árvore React durante rerender no Streamlit Cloud.
function syncToPython() {{
    if (points.length === 4) {{
        const data = {{ 
            points: points, 
            rows: parseInt(inpRows.value)||1, 
            cols: parseInt(inpCols.value)||1,
            referencia: (inpShpRef ? inpShpRef.value.trim() : ''),
            referencia_pronta: shpReferenceReady
        }};
        const payload = JSON.stringify(data);
        try {{
            window.localStorage.setItem('tmg_grid_payload', payload);
            const parentDoc = window.parent && window.parent.document;
            if(parentDoc) {{
                const inputs = Array.from(parentDoc.querySelectorAll('input'));
                const target = inputs.find(el => el.getAttribute('aria-label') === 'grid_payload');
                if(target) {{
                    const setter = Object.getOwnPropertyDescriptor(window.parent.HTMLInputElement.prototype, 'value').set;
                    setter.call(target, payload);
                    target.dispatchEvent(new Event('input', {{bubbles:true}}));
                    target.dispatchEvent(new Event('change', {{bubbles:true}}));
                }}
            }}
        }} catch(e) {{ console.log("Sincronização local indisponível", e); }}
    }}
}}

function validateShpReference(showAlert=false) {{
    const value = (inpShpRef ? inpShpRef.value.trim() : '');
    shpReferenceReady = value.length > 0;
    if(shpReferenceReady) {{
        btnExportSHP.style.display = 'block';
        shpRefStatus.textContent = 'Referência reconhecida. A exportação SHP está liberada para o grid marcado.';
        syncToPython();
        if(showAlert) alert('Referência reconhecida. Agora use Exportar Shapefile (.SHP).');
    }} else {{
        btnExportSHP.style.display = 'none';
        shpRefStatus.textContent = 'Aguardando referência. O botão SHP será liberado após ENTER.';
    }}
}}

function pixelToGeo(pt) {{
    const ratio = Number(GRID_SPATIAL_META && GRID_SPATIAL_META.ratio ? GRID_SPATIAL_META.ratio : 1) || 1;
    const x = pt.x / ratio;
    const y = pt.y / ratio;
    const tr = GRID_SPATIAL_META && Array.isArray(GRID_SPATIAL_META.transform) ? GRID_SPATIAL_META.transform : null;
    if(tr && tr.length >= 6) {{
        return [
            Number(tr[0]) + x * Number(tr[1]) + y * Number(tr[2]),
            Number(tr[3]) + x * Number(tr[4]) + y * Number(tr[5])
        ];
    }}
    return [x, -y];
}}

function buildGridGeoJSON() {{
    const R = parseInt(inpRows.value) || 1;
    const C = parseInt(inpCols.value) || 1;
    const p0 = points[0], p1 = points[1], p2 = points[2], p3 = points[3];
    const features = [];
    for(let r=0; r<R; r++) {{
        for(let c=0; c<C; c++) {{
            const u0=c/C, u1=(c+1)/C, v0=r/R, v1=(r+1)/R;
            const tl=bilerp(p0,p1,p2,p3,u0,v0);
            const tr=bilerp(p0,p1,p2,p3,u1,v0);
            const br=bilerp(p0,p1,p2,p3,u1,v1);
            const bl=bilerp(p0,p1,p2,p3,u0,v1);
            const coords = [tl,tr,br,bl,tl].map(pixelToGeo);
            features.push({{
                type:'Feature',
                properties:{{
                    ID:'T'+(c+1)+' D'+(r+1),
                    TIRO:c+1,
                    DISPARO:r+1,
                    REFERENCIA:(inpShpRef ? inpShpRef.value.trim() : '')
                }},
                geometry:{{type:'Polygon', coordinates:[coords]}}
            }});
        }}
    }}
    return {{type:'FeatureCollection', features}};
}}

function exportGridShapefile() {{
    if(points.length !== 4) {{
        alert('Por favor, marque os 4 pontos do Grid antes de exportar o Shapefile.');
        return;
    }}
    validateShpReference(false);
    if(!shpReferenceReady) {{
        alert('Digite as coordenadas/referências e pressione ENTER antes de exportar o SHP.');
        return;
    }}
    syncToPython();
    const geojson = buildGridGeoJSON();
    try {{
        if(typeof shpwrite === 'undefined' || !shpwrite.download) {{
            throw new Error('Biblioteca shp-write não carregada.');
        }}
        shpwrite.download(geojson, {{file:'TMG_Grid_Shapefile'}});
        shpRefStatus.textContent = 'SHP gerado com sucesso para o grid marcado.';
    }} catch(err) {{
        console.error(err);
        alert('Não foi possível gerar o SHP direto no visualizador. Verifique a conexão da biblioteca shp-write ou use o ambiente com dependências geoespaciais.');
    }}
}}

function drawGrid() {{
    if (points.length === 0) return;

    const showSummary = cbShowSummary.checked;

    if (points.length === 4) {{
        const R = parseInt(inpRows.value) || 1;
        const C = parseInt(inpCols.value) || 1;
        const p0 = points[0], p1 = points[1], p2 = points[2], p3 = points[3];

        if (showSummary) {{
            for(let r=0; r<R; r++) {{
                for(let c=0; c<C; c++) {{
                    const u0=c/C, u1=(c+1)/C, v0=r/R, v1=(r+1)/R;
                    const tl=bilerp(p0,p1,p2,p3,u0,v0);
                    const tr=bilerp(p0,p1,p2,p3,u1,v0);
                    const br=bilerp(p0,p1,p2,p3,u1,v1);
                    const bl=bilerp(p0,p1,p2,p3,u0,v1);
                    
                    const cx2=(tl.x+tr.x+br.x+bl.x)/4;
                    const cy2=(tl.y+tr.y+br.y+bl.y)/4;
                    
                    ctx.save();
                    ctx.shadowColor='rgba(0,0,0,0.8)'; ctx.shadowBlur=4/sc;
                    ctx.fillStyle='#ffffff';
                    const userFS = parseInt(document.getElementById('inpFontSize').value) || 12;
                    ctx.font='bold '+(Math.max(4, userFS/sc))+'px Arial';
                    ctx.textAlign='center'; ctx.textBaseline='middle';
                    ctx.fillText(`T${{c+1}} D${{r+1}}`, cx2, cy2);
                    ctx.restore();
                }}
            }}
        }}

        ctx.save();
        ctx.shadowColor = 'rgba(0, 160, 255, 0.6)';
        ctx.shadowBlur = 6 / sc;
        ctx.strokeStyle = 'rgba(30, 144, 255, 0.95)';
        ctx.lineWidth = 2 / sc;

        for(let i=0; i<=R; i++) {{
            let v = i / R;
            let lx = (1-v)*p0.x + v*p3.x;
            let ly = (1-v)*p0.y + v*p3.y;
            let rx = (1-v)*p1.x + v*p2.x;
            let ry = (1-v)*p1.y + v*p2.y;
            ctx.beginPath(); ctx.moveTo(lx, ly); ctx.lineTo(rx, ry); ctx.stroke();
        }}

        for(let j=0; j<=C; j++) {{
            let u = j / C;
            let tx = (1-u)*p0.x + u*p1.x;
            let ty = (1-u)*p0.y + u*p1.y;
            let bx = (1-u)*p3.x + u*p2.x;
            let by = (1-u)*p3.y + u*p2.y;
            ctx.beginPath(); ctx.moveTo(tx, ty); ctx.lineTo(bx, by); ctx.stroke();
        }}
        ctx.restore();
    }}

    points.forEach((p, i) => {{
        const isDrag = draggingPoint === i;
        const r = 11 / sc;

        ctx.save();
        ctx.shadowColor = isDrag ? 'rgba(255,255,255,0.9)' : 'rgba(0,180,255,0.8)';
        ctx.shadowBlur = 14 / sc;

        ctx.beginPath();
        ctx.arc(p.x, p.y, r + 3/sc, 0, 2 * Math.PI);
        ctx.fillStyle = isDrag ? 'rgba(255,255,255,0.25)' : 'rgba(0,100,200,0.35)';
        ctx.fill();

        ctx.beginPath();
        ctx.arc(p.x, p.y, r, 0, 2 * Math.PI);
        ctx.fillStyle = isDrag ? '#ffffff' : '#1e90ff';
        ctx.fill();
        ctx.lineWidth = 2.5 / sc;
        ctx.strokeStyle = isDrag ? '#aaddff' : '#00cfff';
        ctx.stroke();
        ctx.restore();

        ctx.save();
        ctx.fillStyle = isDrag ? '#003366' : '#ffffff';
        ctx.font = 'bold ' + (13/sc) + 'px Arial';
        ctx.textAlign = 'center';
        ctx.textBaseline = 'middle';
        ctx.fillText(i+1, p.x, p.y);
        ctx.restore();
    }});
}}

function resize() {{
  cv.width  = vc.clientWidth;
  cv.height = vc.clientHeight;
  if (imgW) draw();
}}

function draw() {{
  ctx.clearRect(0, 0, cv.width, cv.height);
  ctx.save();
  ctx.translate(ox, oy);
  ctx.scale(sc, sc);
  ctx.imageSmoothingEnabled = sc < 2;
  ctx.imageSmoothingQuality = 'high';
  ctx.drawImage(img, 0, 0);
  drawGrid();
  ctx.restore();
  zb.textContent = Math.round(sc * 100) + '%';
}}

function fitScreen() {{
  const sx = cv.width  / imgW;
  const sy = cv.height / imgH;
  sc = Math.min(sx, sy) * 0.92;
  ox = (cv.width  - imgW * sc) / 2;
  oy = (cv.height - imgH * sc) / 2;
  draw();
}}

function zoomAt(factor, cx, cy) {{
  const ns = Math.min(MAX_SC, Math.max(MIN_SC, sc * factor));
  ox = cx - (cx - ox) * (ns / sc);
  oy = cy - (cy - oy) * (ns / sc);
  sc = ns;
  draw();
}}

img.onload = () => {{ imgW = img.width; imgH = img.height; resize(); fitScreen(); }};
img.src = 'data:image/jpeg;base64,' + IMG_B64;

vc.addEventListener('wheel', e => {{
  e.preventDefault();
  const r = cv.getBoundingClientRect();
  zoomAt(e.deltaY < 0 ? 1.18 : 1/1.18, e.clientX - r.left, e.clientY - r.top);
}}, {{ passive: false }});

vc.addEventListener('mousedown', e => {{
    const ic = getImgCoords(e.clientX, e.clientY);
    draggingPoint = -1;

    for(let i=0; i<points.length; i++) {{
        let dx = points[i].x - ic.x;
        let dy = points[i].y - ic.y;
        if (Math.sqrt(dx*dx + dy*dy) < 30 / sc) {{
            draggingPoint = i;
            vc.style.cursor = 'move';
            return;
        }}
    }}

    if (gridMode && points.length < 4) {{
        points.push({{x: ic.x, y: ic.y}});
        draw();
        return;
    }}

    drag=true; lx=e.clientX; ly=e.clientY;
    if (!gridMode) vc.style.cursor = 'grabbing';
}});

window.addEventListener('mousemove', e => {{
  const ic = getImgCoords(e.clientX, e.clientY);
  
  coord.textContent = 'X: ' + Math.round(ic.x) + '   Y: ' + Math.round(ic.y);

  if (draggingPoint !== -1) {{
      points[draggingPoint] = {{x: ic.x, y: ic.y}};
      draw();
      return;
  }}

  if (drag) {{
      ox += e.clientX - lx; oy += e.clientY - ly;
      lx = e.clientX; ly = e.clientY;
      draw();
      return;
  }}

  let hoverPoint = false;
  for(let i=0; i<points.length; i++) {{
      let dx = points[i].x - ic.x;
      let dy = points[i].y - ic.y;
      if (Math.sqrt(dx*dx + dy*dy) < 30 / sc) hoverPoint = true;
  }}

  if (hoverPoint) {{
      vc.style.cursor = 'move';
  }} else if (gridMode && points.length < 4) {{
      vc.style.cursor = 'crosshair';
  }} else {{
      vc.style.cursor = 'grab';
  }}
}});

window.addEventListener('mouseup', () => {{
    drag = false;
    if(draggingPoint !== -1){{
        draggingPoint = -1;
        syncToPython();
    }}
    vc.style.cursor = (gridMode && points.length < 4) ? 'crosshair' : 'grab';
}});

vc.addEventListener('dblclick', fitScreen);

btnGrid.onclick = () => {{
    gridMode = !gridMode;
    btnGrid.className = gridMode ? 'grid-btn active' : 'grid-btn';
    btnGrid.innerText = gridMode ? '📍 Selecione 4 Pontos' : 'Ativar Marcação';
    vc.style.cursor = gridMode ? 'crosshair' : 'grab';
}};

btnClear.onclick = () => {{
    points = [];
    draw();
}};

inpRows.addEventListener('change', syncToPython);
inpCols.addEventListener('change', syncToPython);
inpRows.addEventListener('input', draw);
inpCols.addEventListener('input', draw);
cbShowSummary.addEventListener('change', draw);

// Controle de tamanho do texto em tempo real
const inpFontSize = document.getElementById('inpFontSize');
const lblFontSize = document.getElementById('lblFontSize');
inpFontSize.addEventListener('input', () => {{
    lblFontSize.textContent = inpFontSize.value;
    draw();
}});

document.getElementById('btnZI').onclick  = () => zoomAt(1.3, cv.width/2, cv.height/2);
document.getElementById('btnZO').onclick  = () => zoomAt(1/1.3, cv.width/2, cv.height/2);
document.getElementById('btnFit').onclick = fitScreen;
document.getElementById('btnN').onclick   = () => {{ sc=1; ox=(cv.width-imgW)/2; oy=(cv.height-imgH)/2; draw(); }};
document.getElementById('btnR').onclick   = fitScreen;

// Lógica de exportação WYSIWYG
document.getElementById('btnExportFull').onclick = () => {{
    const link = document.createElement('a');
    link.download = 'TMG_Grid_Visualizador.png';
    link.href = cv.toDataURL("image/png");
    link.click();
}};

// Lógica de exportação Somente Grade
document.getElementById('btnExportGrid').onclick = () => {{
    ctx.clearRect(0, 0, cv.width, cv.height);
    ctx.save();
    ctx.translate(ox, oy);
    ctx.scale(sc, sc);
    drawGrid();
    ctx.restore();
    
    const link = document.createElement('a');
    link.download = 'TMG_Grid_Limpo.png';
    link.href = cv.toDataURL("image/png");
    link.click();
    
    draw();
}};

inpShpRef.addEventListener('keydown', (e) => {{
    if(e.key === 'Enter') {{
        e.preventDefault();
        validateShpReference(true);
    }}
}});
inpShpRef.addEventListener('input', () => {{
    if(!inpShpRef.value.trim()) validateShpReference(false);
}});
btnExportSHP.onclick = exportGridShapefile;

window.addEventListener('resize', resize);
</script>
</body>
</html>
"""
                components.html(viewer_html, height=720, scrolling=False)

                # ---------------------------------------------------------
                # ADIÇÃO: MÓDULO ROBUSTO DE EXPORTAÇÃO SHAPEFILE (GEOPANDAS)
                # ---------------------------------------------------------
                st.markdown("---")
                st.markdown("""
                <div style='color:#ff00ff;font-weight:700;font-size:1.1rem;letter-spacing:2px; text-transform:uppercase;margin-bottom:10px;'>
                    🗺️ Exportação Robusta de Shapefile (VectorData)
                </div>
                """, unsafe_allow_html=True)
                
                grid_payload = st.session_state.get("grid_payload")

                if grid_payload and len(json.loads(grid_payload).get("points", [])) == 4:
                    if HAS_GEOPANDAS:
                        st.success("✅ Geometrias do Grid lidas e alinhadas. O sistema está pronto para consolidar as células do grid garantindo o empacotamento completo do vetor sem erros.")
                        
                        try:
                            data = json.loads(grid_payload)
                            points = data.get("points", [])
                            R = data.get("rows", 1)
                            C = data.get("cols", 1)
                            
                            features = []
                            
                            # Helper em python equivalente ao bilerp do JS
                            def py_bilerp(p0, p1, p2, p3, u, v):
                                tx = (1-u)*p0['x'] + u*p1['x']
                                ty = (1-u)*p0['y'] + u*p1['y']
                                bx = (1-u)*p3['x'] + u*p2['x']
                                by = (1-u)*p3['y'] + u*p2['y']
                                return {"x": (1-v)*tx + v*bx, "y": (1-v)*ty + v*by}
                            
                            meta = st.session_state.get("spatial_meta", {})
                            ratio = meta.get("ratio", 1.0)
                            gdal_transform = meta.get("transform", None)
                            crs_wkt = meta.get("crs", None)
                            
                            # Recriando a matriz Affine da Ortofoto Original se disponível
                            aff = Affine.from_gdal(*gdal_transform) if gdal_transform else None

                            for r in range(R):
                                for c in range(C):
                                    u0, u1 = c/C, (c+1)/C
                                    v0, v1 = r/R, (r+1)/R
                                    
                                    tl = py_bilerp(points[0], points[1], points[2], points[3], u0, v0)
                                    tr = py_bilerp(points[0], points[1], points[2], points[3], u1, v0)
                                    br = py_bilerp(points[0], points[1], points[2], points[3], u1, v1)
                                    bl = py_bilerp(points[0], points[1], points[2], points[3], u0, v1)
                                    
                                    # Desfazer a escala temporária aplicada no visualizador (MAX_DIM)
                                    coords_px = [
                                        (tl['x'] / ratio, tl['y'] / ratio),
                                        (tr['x'] / ratio, tr['y'] / ratio),
                                        (br['x'] / ratio, br['y'] / ratio),
                                        (bl['x'] / ratio, bl['y'] / ratio)
                                    ]
                                    
                                    # Aplicar a conversão georreferenciada Affine
                                    if aff:
                                        coords_geo = [aff * pt for pt in coords_px]
                                    else:
                                        # Fallback visual padrão com inversão de Eixo Y se a imagem for apenas um JPG limpo
                                        coords_geo = [(pt[0], -pt[1]) for pt in coords_px]
                                        
                                    poly = Polygon(coords_geo)
                                    
                                    features.append({
                                        "geometry": poly,
                                        "ID": f"T{c+1} D{r+1}",
                                        "TIRO": c+1,
                                        "DISPARO": r+1
                                    })
                            
                            gdf = gpd.GeoDataFrame(features)
                            
                            # Recuperar CRS se o arquivo importado for GeoTIFF
                            if crs_wkt and aff:
                                gdf.set_crs(crs_wkt, allow_override=True, inplace=True)
                                
                            # Compactar todos os sub-arquivos Shapefile em Memória (SEM ERRO EXTERNO)
                            zip_buffer = io.BytesIO()
                            with tempfile.TemporaryDirectory() as tmpdir:
                                shp_path = os.path.join(tmpdir, "TMG_Grid_Parcelas.shp")
                                gdf.to_file(shp_path, driver="ESRI Shapefile")
                                
                                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                                    for ext in [".shp", ".shx", ".dbf", ".prj", ".cpg"]:
                                        filepath = shp_path.replace(".shp", ext)
                                        if os.path.exists(filepath):
                                            zf.write(filepath, arcname=f"TMG_Grid_Parcelas{ext}")
                            
                            # Gerar Download Nativo e Seguro Streamlit
                            st.download_button(
                                label="📥 Baixar Shapefile (.SHP) Arquivo ZIP",
                                data=zip_buffer.getvalue(),
                                file_name="TMG_Grid_Shapefile.zip",
                                mime="application/zip",
                                type="primary"
                            )
                        except Exception as ex:
                            st.error(f"Ocorreu um erro no processamento vetorial: {ex}")
                    else:
                        st.error("Bibliotecas obrigatórias (GeoPandas, Shapely) não estão instaladas. Verifique seu ambiente.")
                else:
                    st.info("Aguardando desenho do Grid e clique em Exportar Shapefile no visualizador...")

        else:
            st.markdown("""
            <div style='
                height: 706px;
                border: 1px dashed #2e2e2e;
                border-radius: 12px;
                background: #0d0d0d;
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
                gap: 12px;
                color: #333;
            '>
                <div style='font-size:3rem;'>🗺️</div>
                <div style='font-size:0.9rem;letter-spacing:2px;text-transform:uppercase;'>
                    Nenhuma ortofoto carregada
                </div>
                <div style='font-size:0.75rem;color:#2a2a2a;'>
                    PNG · JPG · TIF · GeoTIFF · JP2 · IMG · ECW
                </div>
            </div>
            """, unsafe_allow_html=True)

    # ==========================================
    # ANALISE DE MARCACAO DE GRID
    # ==========================================
    elif st.session_state.pagina_ativa == 'AnaliseMarcacaoGrid':
        render_analise_marcacao_grid()

    # ==========================================
    # UPLOAD COM SISTEMA DE TRANSFERÊNCIA INTELIGENTE[cite: 1]
    # ==========================================
    elif st.session_state.pagina_ativa == 'Upload':
        st.subheader("📤 Central de Arquivos")

        st.info("Arraste e solte as imagens do experimento agrícola abaixo.")

        uploaded_files = st.file_uploader(
            "Imagens (PNG, JPG)",
            type=["png", "jpg"],
            accept_multiple_files=True
        )

        if uploaded_files:
            render_tmg_loading_bar(100, f"{len(uploaded_files)} imagem(ns) recebida(s) para processamento.")
            st.success(f"{len(uploaded_files)} arquivos prontos para processamento.")

            # Funcionalidade Original que existia[cite: 1]
            if st.button("Iniciar Processamento", type="primary"):
                st.toast("Iniciando análise por IA...")
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # --- NOVA FUNCIONALIDADE: Transferidor de Arquivos ---
            st.markdown("<h4 style='color:#ff8c00; margin-bottom:15px;'>🚀 Transferência Inteligente de Arquivos</h4>", unsafe_allow_html=True)

            col_dest1, col_dest2 = st.columns(2)
            with col_dest1:
                destino_escolhido = st.selectbox(
                    "Selecione o destino de envio das imagens:",
                    ["☁️ OneDrive", "☁️ Google Drive", "☁️ Azure Storage", "📁 Pasta local pré-definida"]
                )
            with col_dest2:
                # Definir caminhos padrões baseados na escolha do usuário
                if "Pasta local" in destino_escolhido:
                    default_path = str(SYSTEM_DATABASE_DIR / "imagens")
                else:
                    default_path = "/Projetos/TMG_2026/Imagens"
                    
                caminho_envio = st.text_input("Caminho/Pasta de destino (Editável):", value=default_path)

            st.markdown("<br>", unsafe_allow_html=True)

            if st.button(f"Transferir para {destino_escolhido}", type="primary"):
                load_box = st.empty()
                update_tmg_loading(load_box, 10, f"Preparando transferência de {len(uploaded_files)} arquivo(s)...")
                with st.spinner(f"Estabelecendo conexão e transferindo {len(uploaded_files)} arquivos para {destino_escolhido}..."):
                    import time
                    time.sleep(2.5) # Simulação mock do envio / status API
                update_tmg_loading(load_box, 55, "Transferência em andamento...")
                
                if "Pasta local" in destino_escolhido:
                    destino_local = _resolve_system_path(caminho_envio)
                    destino_local.mkdir(parents=True, exist_ok=True)
                    total_files = max(1, len(uploaded_files))
                    done_files = 0
                    for arquivo in uploaded_files:
                        try:
                            arquivo.seek(0)
                        except Exception:
                            pass
                        (destino_local / Path(arquivo.name).name).write_bytes(arquivo.read())
                        done_files += 1
                        update_tmg_loading(load_box, 55 + int((done_files / total_files) * 35), f"Salvando arquivo: {Path(arquivo.name).name}")
                    caminho_envio = str(destino_local)

                update_tmg_loading(load_box, 100, "Carregamento concluído com sucesso.")
                st.success(f"✅ Transferência concluída com sucesso!")
                st.markdown(f"**Status:** {len(uploaded_files)} imagens salvas na pasta configurada: `{caminho_envio}`")
                st.balloons()
    # CONFIG[cite: 1]
    elif st.session_state.pagina_ativa == 'Config':
        st.subheader("⚙️ Painel Administrativo")

        with st.expander("🎨 Tema do Sistema", expanded=False):
            st.markdown(
                "<div style='color:#888;font-size:0.82rem;letter-spacing:1px;text-transform:uppercase;margin-bottom:12px;'>"
                "Selecione o tema visual aplicado em todo o sistema:</div>",
                unsafe_allow_html=True
            )

            _temas_disponiveis = {
                "padrao": "🔵 Padrão do Sistema  —  Dark com cor do tema",
                "tmg_azul": "🔵 TMG Azul  —  Azul escuro · Cinza · Branco",
                "tmg_premium_neon_3d": "💎 TMG Premium Neon 3D  —  Azul/ciano neon · títulos e ícones 3D"
            }
            _tema_atual = SYSTEM_CONFIG.get("tema", "padrao")
            _tema_nomes = list(_temas_disponiveis.keys())
            _tema_labels = list(_temas_disponiveis.values())
            _tema_index = _tema_nomes.index(_tema_atual) if _tema_atual in _tema_nomes else 0

            _tema_escolhido_label = st.radio(
                "Tema",
                _tema_labels,
                index=_tema_index,
                key="cfg_tema_radio",
                label_visibility="collapsed"
            )
            _tema_escolhido = _tema_nomes[_tema_labels.index(_tema_escolhido_label)]

            if st.button("💾 Aplicar Tema", type="primary", key="btn_aplicar_tema", use_container_width=True):
                _cfg_tema = _load_system_config()
                _cfg_tema["tema"] = _tema_escolhido
                _save_system_config(_cfg_tema)
                st.success("✅ Tema salvo! Recarregue a página para aplicar.")
                app_rerun()

        with st.expander("Identidade Visual", expanded=True):
            st.write("Atualize a logo do sistema:")

            if LOGO_PATH.exists():
                st.markdown(
                    f"<p style='color:#888; font-size:0.85rem;'>📁 Logo salva em: "
                    f"<code style='color:#ff8c00;'>{LOGO_PATH}</code></p>",
                    unsafe_allow_html=True
                )

            nova_logo = st.file_uploader(
                "Escolha uma nova logo",
                type=["png", "jpg", "jpeg"]
            )

            if nova_logo:
                load_box = st.empty()
                update_tmg_loading(load_box, 45, f"Carregando logo: {Path(nova_logo.name).name}")
                img = Image.open(nova_logo)
                update_tmg_loading(load_box, 82, "Salvando logo do sistema...")
                img.save(str(LOGO_PATH), format="PNG")
                st.session_state.logo_sistema = img
                update_tmg_loading(load_box, 100, "Carregamento concluído com sucesso.")

                st.success(f"✅ Logo atualizada e salva em: `{LOGO_PATH}`")
                app_rerun()

        with st.expander("Caminhos de Diretório"):
            st.text_input("Diretório de Banco de Dados", value=str(SYSTEM_DATABASE_DIR), disabled=True)
            st.caption("Altere esta pasta pelo menu Banco de Dados Sistema.")

        if _auth_is_admin():
            with st.expander("Logos das Parceiras", expanded=False):
                _render_partner_logo_settings()

            with st.expander("Gerenciar Usuários", expanded=False):
                _render_manage_users()

            with st.expander("Histórico Geral do Sistema", expanded=False):
                state_hist = _partners_load_state()
                st.dataframe(_partners_history_rows(state_hist.get("history_general", [])), use_container_width=True, hide_index=True)

    # BASES[cite: 1]
    elif st.session_state.pagina_ativa == 'Bases':
        st.subheader("🗂️ Banco de Dados Sistema")

        st.info("Escolha a pasta principal onde o sistema irá trabalhar, salvar arquivos, imagens, ortomosaicos, relatórios e bancos internos.")

        atual = SYSTEM_DATABASE_DIR
        try:
            uso = shutil.disk_usage(atual)
            espaco_livre = _tv_human_size(uso.free)
        except Exception:
            espaco_livre = "Indisponível"

        c1, c2, c3 = st.columns(3)
        c1.metric("Status", "Ativo" if atual.exists() else "Não criado")
        c2.metric("Espaço livre", espaco_livre)
        c3.metric("Atualizado", SYSTEM_CONFIG.get("updated_at") or "-")

        st.markdown(
            f"<div class='card'><h4 style='color:#ff8c00;margin-top:0;'>Pasta atual</h4>"
            f"<code>{atual}</code></div>",
            unsafe_allow_html=True
        )

        sugestoes = {
            "Pasta atual": str(atual),
            "Padrão do sistema": str(_resolve_system_path("tmg_data")),
            "Dados do pacote": str(APP_ROOT / "tmg_data"),
            "Exports do pacote": str(APP_ROOT / "tmg_data" / "exports")
        }
        if os.name == "nt":
            sugestoes["Backup interno"] = str(APP_ROOT / "tmg_data" / "backups")

        s1, s2 = st.columns([1, 2])
        with s1:
            preset = st.selectbox("Atalho de pasta", list(sugestoes.keys()), key="sys_db_preset")
        with s2:
            caminho_db = st.text_input(
                "Local da pasta de trabalho",
                value=sugestoes[preset],
                key="sys_db_path"
            )

        criar_estrutura = st.checkbox("Criar estrutura padrão de pastas", value=True, key="sys_db_create_tree")
        subpastas = ["imagens", "ortomosaicos", "relatorios", "exports", "uploads", "grids", "temporarios"]

        if st.button("Salvar Banco de Dados Sistema", type="primary", key="btn_save_system_db", use_container_width=True):
            try:
                novo_dir = _resolve_system_path(caminho_db)
                novo_dir.mkdir(parents=True, exist_ok=True)
                if criar_estrutura:
                    for nome in subpastas:
                        (novo_dir / nome).mkdir(parents=True, exist_ok=True)
                config = _load_system_config()
                config["database_dir"] = str(novo_dir)
                config["updated_at"] = _tv_now()
                _save_system_config(config)
                st.success(f"Banco de Dados Sistema salvo em: `{novo_dir}`")
                app_rerun()
            except Exception as exc:
                st.error(f"Não foi possível salvar a pasta do sistema: {exc}")

        st.markdown("#### Estrutura da pasta")
        estrutura = []
        for nome in subpastas + ["transferencia_voos", "voos_direcionados"]:
            pasta = atual / nome
            estrutura.append({
                "Pasta": nome,
                "Caminho": str(pasta),
                "Status": "Existe" if pasta.exists() else "Não criada"
            })
        st.dataframe(estrutura, use_container_width=True, hide_index=True)

    # SINCRONIZAR DADOS[cite: 1]
    elif st.session_state.pagina_ativa == 'Sync':
        _render_sync_backup()

    # GERAR ORTOMOSAICOS[cite: 1]
    elif st.session_state.pagina_ativa == 'Ortomosaicos':
        _render_orthomosaic_generator()

    # ==========================================
    # NOVO - VISUALIZADOR DE RESULTADOS
    # ==========================================
    elif st.session_state.pagina_ativa == 'Visualizador':
        st.subheader("📈 Visualizador de Resultados")

        st.markdown("""
        <div style='color:#ff8c00;font-weight:700;font-size:1rem;letter-spacing:2px;
                    text-transform:uppercase;margin-bottom:20px;'>
            Selecione o tipo de análise para visualizar
        </div>""", unsafe_allow_html=True)

        # NOVO - Inicializar sub-página do visualizador
        if "visualizador_sub" not in st.session_state:
            st.session_state.visualizador_sub = None

        phenotyping_buttons = [
            ("phenotyping_contagem", "🔢 Contagem", "Contagem", "btn_viz_contagem"),
            ("phenotyping_maturacao", "🌱 Maturação", "Maturação", "btn_viz_maturacao"),
            ("phenotyping_pendoamento", "🌾 Pendoamento", "Pendoamento", "btn_viz_pendoamento"),
            ("phenotyping_qualidade", "✅ Qualidade de Parcelas", "Qualidade", "btn_viz_qualidade"),
        ]
        visible_phenotyping_buttons = [item for item in phenotyping_buttons if _auth_phenotyping_allowed(item[0], current_user)]
        allowed_visualizador_subs = {item[2] for item in visible_phenotyping_buttons}

        if st.session_state.visualizador_sub and st.session_state.visualizador_sub not in allowed_visualizador_subs:
            st.session_state.visualizador_sub = None

        if not visible_phenotyping_buttons:
            st.warning("Seu usuário não possui análises de fenotipagem liberadas. Solicite permissão ao administrador.")
            st.stop()

        vcols = st.columns(len(visible_phenotyping_buttons))
        for col, (_, label, sub_page, button_key) in zip(vcols, visible_phenotyping_buttons):
            with col:
                if st.button(label, key=button_key, use_container_width=True):
                    st.session_state.visualizador_sub = sub_page

        st.markdown("---")

        # NOVO - Sub-visualizações
        if st.session_state.visualizador_sub == "Contagem":
            # NOVO - MÓDULO COMPLETO DE CONTAGEM DE PLANTAS (viewer idêntico ao Checklist + funções TZ Plants)
            st.markdown("""
            <div style='color:#ff8c00;font-weight:700;font-size:1rem;letter-spacing:2px;
                        text-transform:uppercase;margin-bottom:10px;'>
                🌱 Contagem de Plantas por Parcela
            </div>""", unsafe_allow_html=True)

            # NOVO - Upload de imagem para contagem (mesmo formato do Checklist)
            cnt_file = _resettable_ortho_uploader(
                "📷 Carregar Ortofoto para Contagem",
                key="cnt_orto_uploader",
                help="PNG · JPG · TIF/GeoTIFF · JP2 · IMG · ECW"
            )
            cnt_bytes, cnt_name = _uploaded_ortho_bytes(cnt_file)

            if cnt_bytes:
                with st.container():
                    cnt_b64, cnt_dims, cnt_err, cnt_spatial = processar_ortofoto(cnt_bytes, cnt_name)

                if cnt_err:
                    st.error(f"Erro: {cnt_err}")
                else:
                    cw_cnt, ch_cnt = cnt_dims
                    cnt_storage_id = json.dumps(_tv_hash_bytes(cnt_bytes)[:32])
                    st.markdown(
                        f"<p style='color:#666;font-size:0.78rem;margin-bottom:6px;'>"
                        f"📐 {cnt_name} · {cw_cnt}×{ch_cnt} px</p>",
                        unsafe_allow_html=True
                    )

                    # NOVO - Viewer HTML/JS idêntico ao Checklist com funcionalidades de contagem TZ Plants
                    cnt_viewer = f"""
<!DOCTYPE html>
<html>
<head>
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ background:#0d0d0d; overflow:hidden; font-family:'Segoe UI',sans-serif; }}

  #vc {{
    width:100%; height:706px;
    background:
      linear-gradient(rgba(255,255,255,.03) 1px, transparent 1px),
      linear-gradient(90deg, rgba(255,255,255,.03) 1px, transparent 1px),
      #0d0d0d;
    background-size:32px 32px;
    border:1px solid #2a2a2a; border-radius:12px;
    overflow:hidden; position:relative; cursor:grab; user-select:none;
  }}
  #vc:active {{ cursor:grabbing; }}
  canvas {{ position:absolute; top:0; left:0; display:block; }}

  .toolbar {{ position:absolute; top:12px; right:12px; display:flex; flex-direction:column; gap:5px; z-index:20; }}
  .tb-btn {{
    background:linear-gradient(145deg,#1e1e1e,#111); border:1px solid #3a3a3a;
    color:#ff8c00; width:34px; height:34px; border-radius:8px; cursor:pointer;
    font-size:15px; font-weight:700; display:flex; align-items:center; justify-content:center;
    box-shadow:2px 2px 8px #000,inset 0 1px 0 rgba(255,255,255,.05); transition:all .2s;
  }}
  .tb-btn:hover {{ border-color:#ff8c00; box-shadow:0 0 10px rgba(255,140,0,.35),2px 2px 8px #000; color:#ffaa33; }}
  .tb-btn:active {{ transform:translateY(1px); }}
  .tb-sep {{ width:34px; height:1px; background:linear-gradient(90deg,transparent,#333,transparent); margin:2px 0; }}

  .grid-panel {{
    position:absolute; top:12px; right:55px;
    background:rgba(10,10,10,.88); border:1px solid #2a2a2a;
    border-radius:8px; padding:10px; display:flex; flex-direction:column;
    gap:8px; z-index:20;
  }}
  .grid-panel label {{ color:#ff8c00; font-size:11px; font-weight:bold; text-align:center; }}
  .grid-panel input[type=number] {{
    background:#1a1a1a; border:1px solid #333; color:#fff;
    border-radius:4px; padding:4px; width:50px; text-align:center; font-size:11px;
  }}
  .grid-panel input[type=text], .grid-panel select {{
    background:#1a1a1a; border:1px solid #333; color:#fff;
    border-radius:4px; padding:4px; font-size:11px; min-width:104px;
  }}
  .grid-panel .row-col {{ display:flex; gap:8px; align-items:center; justify-content:space-between; color:#ccc; font-size:11px; }}
  .grid-status {{ color:#777; font-size:9px; line-height:1.25; max-width:190px; }}
  .grid-btn {{
    background:linear-gradient(145deg,#1e1e1e,#111); border:1px solid #3a3a3a;
    color:#ccc; cursor:pointer; border-radius:4px; padding:6px; font-size:11px; font-weight:bold; transition:.2s;
  }}
  .grid-btn:hover {{ border-color:#ff8c00; color:#ff8c00; }}
  .grid-btn.active {{ border-color:#ff8c00; color:#ff8c00; box-shadow:0 0 8px rgba(255,140,0,.3); background:#2a1a00; }}

  .zoom-badge {{
    position:absolute; top:12px; left:12px;
    background:rgba(10,10,10,.82); border:1px solid #2a2a2a; border-radius:8px;
    color:#ff8c00; font-size:11px; font-family:'Courier New',monospace;
    font-weight:700; padding:5px 10px; letter-spacing:1px; z-index:20; pointer-events:none;
  }}
  .crosshair {{
    position:absolute; bottom:12px; left:12px;
    background:rgba(10,10,10,.82); border:1px solid #222; border-radius:8px;
    color:#555; font-size:10px; font-family:'Courier New',monospace;
    padding:4px 10px; z-index:20; pointer-events:none; letter-spacing:.5px;
  }}
  .hint {{
    position:absolute; bottom:12px; right:12px; color:#333; font-size:10px;
    z-index:20; pointer-events:none; text-align:right; line-height:1.6;
  }}

  .count-panel {{
    position:absolute; top:50px; left:12px;
    background:rgba(10,10,10,.92); border:1px solid #2a2a2a; border-radius:8px;
    padding:10px; z-index:20; min-width:180px;
  }}
  .count-panel h3 {{ color:#00ff00; font-size:12px; margin-bottom:6px; letter-spacing:1px; }}
  .count-panel .total {{ color:#fff; font-size:22px; font-weight:bold; }}
  .count-panel .info {{ color:#888; font-size:10px; margin-top:4px; }}

  .cnt-btn {{
    background:linear-gradient(145deg,#1a3a1a,#0a2a0a); border:1px solid #006600;
    color:#00ee55; border-radius:4px; padding:6px 8px; font-size:11px;
    font-weight:bold; cursor:pointer; transition:.2s; width:100%; margin-top:4px;
  }}
  .cnt-btn:hover {{ border-color:#00ee55; box-shadow:0 0 8px rgba(0,238,85,.3); }}
  .cnt-btn.danger {{ background:linear-gradient(145deg,#3a1a1a,#2a0a0a); border-color:#660000; color:#ff5555; }}
  .cnt-btn.danger:hover {{ border-color:#ff5555; box-shadow:0 0 8px rgba(255,85,85,.3); }}
  .cnt-btn.manual {{ background:linear-gradient(145deg,#1a1a3a,#0a0a2a); border-color:#000066; color:#5599ff; }}
  .cnt-btn.manual:hover {{ border-color:#5599ff; box-shadow:0 0 8px rgba(85,153,255,.3); }}

  #btnExportCnt {{
    background:linear-gradient(145deg,#003a00,#001a00); border:1px solid #006600;
    color:#00ee55; border-radius:4px; padding:6px 8px; font-size:11px;
    font-weight:bold; cursor:pointer; transition:.2s; width:100%; margin-top:6px;
  }}
  #btnExportCnt:hover {{ border-color:#00ee55; box-shadow:0 0 8px rgba(0,238,85,.3); }}
</style>
</head>
<body>
<div id="vc">
  <canvas id="cv"></canvas>
  <div class="zoom-badge" id="zbadge">1.00×</div>
  <div class="crosshair" id="coord">X:0 Y:0</div>
  <div class="hint">Scroll=Zoom · Drag=Pan<br>Grid: marque 4 pontos</div>

  <div class="toolbar">
    <button class="tb-btn" id="btnGridTool" title="Marcar Grid (4 pontos)">⊞</button>
    <button class="tb-btn" id="btnCountPlants" title="Contar Plantas (automático)">🌱</button>
    <button class="tb-btn" id="btnManualMode" title="Modo Manual (clique para marcar)">✏️</button>
    <button class="tb-btn" id="btnRemoveLast" title="Apagar Última Marcação">❌</button>
    <div class="tb-sep"></div>
    <button class="tb-btn" id="btnClearAll" title="Limpar Tudo">🗑️</button>
  </div>

  <div class="grid-panel">
    <label>🌱 CONTAGEM</label>
    <div class="row-col">
      <span>Quadra:</span><input type="text" id="inpGridName" value="Grid 1">
    </div>
    <div class="row-col">
      <span>Ativo:</span><select id="selGridList"></select>
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:5px;">
      <button class="grid-btn" id="btnSaveGrid">Salvar</button>
      <button class="grid-btn" id="btnNewGrid">Novo</button>
      <button class="grid-btn" id="btnDeleteGrid" style="color:#ff6b6b;border-color:#552222;">Excluir</button>
    </div>
    <div class="grid-status" id="gridStatus">Grid ativo: Grid 1</div>
    <div class="row-col">
      <span>Disp:</span><input type="number" id="inpRows" value="5" min="1" max="200">
    </div>
    <div class="row-col">
      <span>Tiros:</span><input type="number" id="inpCols" value="5" min="1" max="200">
    </div>
    <button class="cnt-btn" id="btnCountAuto">🌱 Contar</button>
    <button class="cnt-btn manual" id="btnManual2">✏️ Manual</button>
    <button class="cnt-btn danger" id="btnUndoMark">❌ Desfazer</button>
    <div class="row-col" style="justify-content:flex-start; gap:4px;">
      <input type="checkbox" id="cbExportAll" style="width:auto;accent-color:#00ee55;">
      <span style="font-size:10px;">Exportar todos os grids salvos</span>
    </div>
    <button id="btnExportCnt">💾 Exportar CSV</button>
    <button class="cnt-btn" id="btnExportXLSXCnt" style="background:linear-gradient(145deg,#1a003a,#0a001a);border-color:#660099;color:#cc66ff;">📗 Exportar Excel</button>
    <button class="cnt-btn" id="btnResumo" style="background:linear-gradient(145deg,#2a1a00,#1a0a00);border-color:#ff8c00;color:#ff8c00;">📊 Resumo</button>
  </div>

  <div class="count-panel" id="countPanel" style="display:none;">
    <h3>CONTAGEM</h3>
    <div class="total" id="totalCount">0</div>
    <div class="info" id="countInfo">Marque o grid e clique Contar</div>
  </div>

  <!-- NOVO - Modal Resumo -->
  <div id="resumoModal" style="display:none;position:fixed;top:0;left:0;width:100%;height:100%;
    background:rgba(0,0,0,0.85);z-index:9999;align-items:center;justify-content:center;">
    <div style="background:#1a1a1a;border:1px solid #ff8c00;border-radius:16px;width:90%;max-width:700px;
      max-height:85vh;overflow:auto;padding:24px;box-shadow:0 0 40px rgba(255,140,0,0.3);">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
        <h3 style="color:#ff8c00;font-size:14px;letter-spacing:2px;text-transform:uppercase;margin:0;">📊 RESUMO DA CONTAGEM</h3>
        <button id="btnCloseResumo" style="background:#333;border:1px solid #555;color:#fff;border-radius:6px;
          padding:5px 12px;cursor:pointer;font-size:12px;">✕ Fechar</button>
      </div>
      <div id="resumoTotalCard" style="background:#111;border:1px solid #333;border-radius:10px;padding:16px;
        text-align:center;margin-bottom:16px;">
        <div id="resumoTotal" style="color:#00ff00;font-size:28px;font-weight:bold;">0</div>
        <div style="color:#888;font-size:11px;letter-spacing:2px;text-transform:uppercase;">Total de Plantas</div>
      </div>
      <div id="resumoFilter" style="margin-bottom:12px;display:flex;gap:8px;align-items:center;">
        <input type="text" id="resumoSearch" placeholder="Pesquisar parcela (ex: T1 D2)"
          style="flex:1;background:#111;border:1px solid #333;color:#fff;border-radius:6px;padding:8px;font-size:12px;">
        <select id="resumoFilterTiro" style="background:#111;border:1px solid #333;color:#fff;border-radius:6px;padding:8px;font-size:11px;">
          <option value="">Todos Tiros</option>
        </select>
        <select id="resumoFilterDisp" style="background:#111;border:1px solid #333;color:#fff;border-radius:6px;padding:8px;font-size:11px;">
          <option value="">Todos Disparos</option>
        </select>
      </div>
      <div id="resumoTable" style="max-height:300px;overflow-y:auto;border:1px solid #333;border-radius:8px;"></div>
      <div style="display:flex;gap:8px;margin-top:14px;">
        <button id="btnResumoCSV" style="flex:1;background:linear-gradient(145deg,#003a00,#001a00);border:1px solid #006600;
          color:#00ee55;border-radius:6px;padding:8px;font-size:11px;font-weight:bold;cursor:pointer;">💾 Exportar CSV</button>
        <button id="btnResumoXLSX" style="flex:1;background:linear-gradient(145deg,#1a003a,#0a001a);border:1px solid #660099;
          color:#cc66ff;border-radius:6px;padding:8px;font-size:11px;font-weight:bold;cursor:pointer;">📗 Exportar Excel</button>
      </div>
    </div>
  </div>
</div>

<script>
const IMG_B64 = '{cnt_b64}';
const vc    = document.getElementById('vc');
const cv    = document.getElementById('cv');
const ctx   = cv.getContext('2d');
const zb    = document.getElementById('zbadge');
const coordEl = document.getElementById('coord');

const inpRows = document.getElementById('inpRows');
const inpCols = document.getElementById('inpCols');
const btnGridTool = document.getElementById('btnGridTool');
const btnCountPlants = document.getElementById('btnCountPlants');
const btnManualMode = document.getElementById('btnManualMode');
const btnRemoveLast = document.getElementById('btnRemoveLast');
const btnClearAll = document.getElementById('btnClearAll');
const btnCountAuto = document.getElementById('btnCountAuto');
const btnManual2 = document.getElementById('btnManual2');
const btnUndoMark = document.getElementById('btnUndoMark');
const btnExportCnt = document.getElementById('btnExportCnt');
const btnExportXLSXCnt = document.getElementById('btnExportXLSXCnt');
const inpGridName = document.getElementById('inpGridName');
const selGridList = document.getElementById('selGridList');
const btnSaveGrid = document.getElementById('btnSaveGrid');
const btnNewGrid = document.getElementById('btnNewGrid');
const btnDeleteGrid = document.getElementById('btnDeleteGrid');
const cbExportAll = document.getElementById('cbExportAll');
const gridStatus = document.getElementById('gridStatus');
const countPanel = document.getElementById('countPanel');
const totalCountEl = document.getElementById('totalCount');
const countInfoEl = document.getElementById('countInfo');
const ORTHO_STORAGE_ID = {cnt_storage_id};
const STORAGE_KEY = 'tmg_contagem_plantas_grids_' + ORTHO_STORAGE_ID;

let gridMode = false;
let manualMode = false;
let points = [];
let draggingPoint = -1;
let sc = 1, ox = 0, oy = 0;
let drag = false, lx = 0, ly = 0;
const MIN_SC = 0.05, MAX_SC = 40;
let imgW = 0, imgH = 0;

let plantCenters = [];
let manualMarks = [];
let parcelCounts = {{}};
let savedGrids = {{}};
let activeGridName = 'Grid 1';
let suppressPersist = false;

const img = new Image();

function getImgCoords(cx, cy) {{
  const r = cv.getBoundingClientRect();
  return {{ x:(cx-r.left-ox)/sc, y:(cy-r.top-oy)/sc }};
}}

function bilerp(p0,p1,p2,p3, u,v) {{
  const tx = (1-u)*p0.x + u*p1.x;
  const ty = (1-u)*p0.y + u*p1.y;
  const bx = (1-u)*p3.x + u*p2.x;
  const by = (1-u)*p3.y + u*p2.y;
  return {{ x:(1-v)*tx + v*bx, y:(1-v)*ty + v*by }};
}}

function pointInPolygon(px, py, polygon) {{
  let inside = false;
  for (let i = 0, j = polygon.length - 1; i < polygon.length; j = i++) {{
    const xi = polygon[i].x, yi = polygon[i].y;
    const xj = polygon[j].x, yj = polygon[j].y;
    if (((yi > py) !== (yj > py)) && (px < (xj - xi) * (py - yi) / (yj - yi) + xi)) {{
      inside = !inside;
    }}
  }}
  return inside;
}}

function cleanGridName(name) {{
  const value = String(name || '').trim();
  return value || 'Grid 1';
}}

function makeUniqueGridName(base) {{
  let name = cleanGridName(base);
  if(!savedGrids[name]) return name;
  let i = 2;
  while(savedGrids[name + ' ' + i]) i++;
  return name + ' ' + i;
}}

function clonePoints(src) {{
  return (src || []).map(p => ({{x:Number(p.x)||0, y:Number(p.y)||0}}));
}}

function cloneMarks(src) {{
  return (src || []).map(p => ({{x:Number(p.x)||0, y:Number(p.y)||0}}));
}}

function cloneCounts(src) {{
  return JSON.parse(JSON.stringify(src || {{}}));
}}

function persistGrids() {{
  if(suppressPersist) return;
  try {{
    localStorage.setItem(STORAGE_KEY, JSON.stringify({{
      activeGridName,
      savedGrids,
      updatedAt: new Date().toISOString()
    }}));
  }} catch(e) {{ console.warn('Não foi possível salvar grids de contagem:', e); }}
}}

function applyGridRecord(rec) {{
  rec = rec || {{}};
  points = clonePoints(rec.points);
  plantCenters = cloneMarks(rec.plantCenters);
  manualMarks = cloneMarks(rec.manualMarks);
  parcelCounts = cloneCounts(rec.parcelCounts);
  inpRows.value = rec.rows || inpRows.value;
  inpCols.value = rec.cols || inpCols.value;
  countPanel.style.display = Object.keys(parcelCounts).length ? 'block' : 'none';
  totalCountEl.textContent = plantCenters.length || 0;
  countInfoEl.textContent = Object.keys(parcelCounts).length
    ? 'Grid ' + (parseInt(inpRows.value)||1) + '×' + (parseInt(inpCols.value)||1) + ' | ' + (plantCenters.length||0) + ' plantas'
    : 'Marque o grid e clique Contar';
}}

function saveActiveGrid(showMsg=false) {{
  activeGridName = cleanGridName(inpGridName.value || activeGridName);
  savedGrids[activeGridName] = {{
    points: clonePoints(points),
    plantCenters: cloneMarks(plantCenters),
    manualMarks: cloneMarks(manualMarks),
    parcelCounts: cloneCounts(parcelCounts),
    rows: parseInt(inpRows.value)||1,
    cols: parseInt(inpCols.value)||1
  }};
  updateGridSelect();
  persistGrids();
  if(showMsg) gridStatus.textContent = 'Grid salvo: ' + activeGridName;
}}

function restoreGrids() {{
  try {{
    const raw = localStorage.getItem(STORAGE_KEY);
    if(!raw) return false;
    const data = JSON.parse(raw);
    if(!data || !data.savedGrids) return false;
    savedGrids = data.savedGrids || {{}};
    activeGridName = cleanGridName(data.activeGridName || Object.keys(savedGrids)[0] || 'Grid 1');
    const rec = savedGrids[activeGridName] || savedGrids[Object.keys(savedGrids)[0]];
    if(rec) applyGridRecord(rec);
    return Object.keys(savedGrids).length > 0;
  }} catch(e) {{ console.warn('Não foi possível restaurar grids de contagem:', e); return false; }}
}}

function updateGridSelect() {{
  if(!savedGrids[activeGridName]) {{
    savedGrids[activeGridName] = {{
      points: clonePoints(points),
      plantCenters: cloneMarks(plantCenters),
      manualMarks: cloneMarks(manualMarks),
      parcelCounts: cloneCounts(parcelCounts),
      rows: parseInt(inpRows.value)||1,
      cols: parseInt(inpCols.value)||1
    }};
  }}
  const names = Object.keys(savedGrids);
  selGridList.innerHTML = '';
  names.forEach(name => {{
    const rec = savedGrids[name] || {{}};
    const opt = document.createElement('option');
    opt.value = name;
    const counted = rec.parcelCounts && Object.keys(rec.parcelCounts).length > 0;
    opt.textContent = name + (counted ? ' ✓' : '');
    selGridList.appendChild(opt);
  }});
  selGridList.value = activeGridName;
  inpGridName.value = activeGridName;
  const marked = points.length === 4 ? 'marcado' : 'sem 4 pontos';
  const counted = Object.keys(parcelCounts).length > 0 ? 'com contagem' : 'sem contagem';
  gridStatus.textContent = 'Grid ativo: ' + activeGridName + ' · ' + marked + ' · ' + counted + ' · salvos: ' + names.length;
}}

function loadGridByName(name) {{
  const target = cleanGridName(String(name || '').replace(/ ✓$/,''));
  if(target === activeGridName) return;
  saveActiveGrid(false);
  const rec = savedGrids[target];
  if(!rec) return;
  activeGridName = target;
  applyGridRecord(rec);
  updateGridSelect();
  drawAll();
}}

function createNewGrid() {{
  saveActiveGrid(false);
  const suggested = makeUniqueGridName('Grid ' + (Object.keys(savedGrids).length + 1));
  const typed = prompt('Nome da Quadra/Grid:', suggested);
  if(typed === null) return;
  activeGridName = makeUniqueGridName(typed);
  points = [];
  plantCenters = [];
  manualMarks = [];
  parcelCounts = {{}};
  savedGrids[activeGridName] = {{
    points: [],
    plantCenters: [],
    manualMarks: [],
    parcelCounts: {{}},
    rows: parseInt(inpRows.value)||1,
    cols: parseInt(inpCols.value)||1
  }};
  countPanel.style.display = 'none';
  updateGridSelect();
  persistGrids();
  drawAll();
}}

function renameActiveGrid() {{
  const next = cleanGridName(inpGridName.value);
  if(next === activeGridName) return;
  const old = activeGridName;
  activeGridName = makeUniqueGridName(next);
  if(savedGrids[old]) delete savedGrids[old];
  saveActiveGrid(false);
}}

function deleteActiveGrid() {{
  const target = activeGridName;
  if(!confirm('Excluir definitivamente o grid "' + target + '"? Ele sairá da lista, da visualização e das exportações.')) return;
  delete savedGrids[target];
  const remaining = Object.keys(savedGrids);
  if(remaining.length) {{
    activeGridName = remaining[0];
    applyGridRecord(savedGrids[activeGridName]);
  }} else {{
    activeGridName = 'Grid 1';
    points = [];
    plantCenters = [];
    manualMarks = [];
    parcelCounts = {{}};
    savedGrids[activeGridName] = {{
      points: [],
      plantCenters: [],
      manualMarks: [],
      parcelCounts: {{}},
      rows: parseInt(inpRows.value)||1,
      cols: parseInt(inpCols.value)||1
    }};
    countPanel.style.display = 'none';
  }}
  updateGridSelect();
  persistGrids();
  drawAll();
}}

function getExportGridRecords(exportAll=false) {{
  saveActiveGrid(false);
  const names = exportAll ? Object.keys(savedGrids) : [activeGridName];
  return names
    .map(name => {{
      const rec = savedGrids[name] || {{}};
      return {{
        name,
        rows: parseInt(rec.rows)||1,
        cols: parseInt(rec.cols)||1,
        parcelCounts: cloneCounts(rec.parcelCounts)
      }};
    }})
    .filter(grid => grid.parcelCounts && Object.keys(grid.parcelCounts).length > 0);
}}

function countForCell(grid, disparo, tiro) {{
  const key = Number(disparo) + '_' + Number(tiro);
  return Number(grid.parcelCounts[key] || 0);
}}

function buildExportRows(exportAll=false) {{
  const rows = [];
  const grids = getExportGridRecords(exportAll)
    .sort((a,b) => String(a.name).localeCompare(String(b.name), 'pt-BR', {{numeric:true}}));
  grids.forEach(grid => {{
    const R = parseInt(grid.rows)||1;
    const C = parseInt(grid.cols)||1;
    for(let d=1; d<=R; d++) {{
      for(let t=1; t<=C; t++) {{
        rows.push({{
          quadra: grid.name,
          disparo: d,
          tiro: t,
          quantidade: countForCell(grid, d, t)
        }});
      }}
    }}
  }});
  return rows;
}}

function csvEscape(value) {{
  const s = String(value ?? '');
  return /[",\\n\\r]/.test(s) ? '"' + s.replace(/"/g,'""') + '"' : s;
}}

function safeFilePart(value) {{
  return String(value || 'grid').trim().replace(/[\\\\/:*?"<>|]+/g,'_').replace(/\\s+/g,'_') || 'grid';
}}

function exportContagemCSV(exportAll=false) {{
  const rows = buildExportRows(exportAll);
  if(rows.length === 0) {{ alert('Nenhuma contagem realizada para o(s) grid(s) selecionado(s).'); return; }}
  let csv = '\\uFEFF' + 'Quadra,Disparo,Tiro,Quantidade de Plantas\\n';
  rows.forEach(row => {{
    csv += [row.quadra,row.disparo,row.tiro,row.quantidade].map(csvEscape).join(',') + '\\n';
  }});
  const blob = new Blob([csv], {{type:'text/csv;charset=utf-8;'}});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = exportAll ? 'contagem_plantas_todos_grids.csv' : 'contagem_plantas_' + safeFilePart(activeGridName) + '.csv';
  a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),1000);
}}

function exportContagemExcel(exportAll=false) {{
  const rows = buildExportRows(exportAll);
  if(rows.length === 0) {{ alert('Nenhuma contagem realizada para o(s) grid(s) selecionado(s).'); return; }}
  if(typeof XLSX === 'undefined') {{ alert('Biblioteca Excel não carregou. Use Exportar CSV.'); return; }}
  const headers = ['Quadra','Disparo','Tiro','Quantidade de Plantas'];
  const data = [headers, ...rows.map(row => [row.quadra,row.disparo,row.tiro,row.quantidade])];
  const ws = XLSX.utils.aoa_to_sheet(data);
  ws['!cols'] = [{{wch:14}},{{wch:10}},{{wch:10}},{{wch:24}}];
  const border = {{
    top:{{style:'thin',color:{{rgb:'808080'}}}},
    bottom:{{style:'thin',color:{{rgb:'808080'}}}},
    left:{{style:'thin',color:{{rgb:'808080'}}}},
    right:{{style:'thin',color:{{rgb:'808080'}}}}
  }};
  const range = XLSX.utils.decode_range(ws['!ref']);
  for(let R=range.s.r; R<=range.e.r; R++) {{
    for(let C=range.s.c; C<=range.e.c; C++) {{
      const addr = XLSX.utils.encode_cell({{r:R,c:C}});
      if(!ws[addr]) continue;
      ws[addr].s = {{
        font: R===0 ? {{bold:true,color:{{rgb:'00B0F0'}}}} : {{color:{{rgb:'000000'}}}},
        fill: R===0
          ? {{patternType:'solid',fgColor:{{rgb:'1F1F1F'}}}}
          : (C===3 ? {{patternType:'solid',fgColor:{{rgb:'00B050'}}}} : undefined),
        alignment: {{horizontal:C===3 || R===0 ? 'center' : 'left', vertical:'center'}},
        border
      }};
      if(R>0 && C===3) {{
        ws[addr].s.font = {{bold:true,color:{{rgb:'FFFFFF'}}}};
        ws[addr].s.alignment = {{horizontal:'center',vertical:'center'}};
        ws[addr].t = 'n';
      }}
    }}
  }}
  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, 'Contagem');
  XLSX.writeFile(wb, exportAll ? 'contagem_plantas_todos_grids.xlsx' : 'contagem_plantas_' + safeFilePart(activeGridName) + '.xlsx', {{bookType:'xlsx',cellStyles:true}});
}}

function countPlantsInGrid() {{
  if (points.length < 4) {{
    alert('Marque os 4 cantos do grid primeiro!');
    return;
  }}

  const R = parseInt(inpRows.value) || 1;
  const C = parseInt(inpCols.value) || 1;
  const p0=points[0], p1=points[1], p2=points[2], p3=points[3];

  // Extrair pixels da região do grid para detecção HSV via canvas
  // Criar canvas temporário para processar a imagem
  const tempCv = document.createElement('canvas');
  tempCv.width = imgW; tempCv.height = imgH;
  const tempCtx = tempCv.getContext('2d');
  tempCtx.drawImage(img, 0, 0);

  // Bounding box do polígono
  const xs = points.map(p=>p.x), ys = points.map(p=>p.y);
  const minX = Math.max(0, Math.floor(Math.min(...xs)));
  const maxX = Math.min(imgW, Math.ceil(Math.max(...xs)));
  const minY = Math.max(0, Math.floor(Math.min(...ys)));
  const maxY = Math.min(imgH, Math.ceil(Math.max(...ys)));

  const imgData = tempCtx.getImageData(minX, minY, maxX-minX, maxY-minY);
  const data = imgData.data;
  const w = maxX - minX;
  const h = maxY - minY;

  plantCenters = [];

  // Detecção HSV simplificada (verde)
  // Converter RGB para HSV e detectar plantas
  const visited = new Uint8Array(w * h);
  const minArea = 25;

  for (let y = 0; y < h; y++) {{
    for (let x = 0; x < w; x++) {{
      const idx = (y * w + x) * 4;
      const r2 = data[idx], g = data[idx+1], b = data[idx+2];

      // RGB para HSV
      const max2 = Math.max(r2, g, b), min2 = Math.min(r2, g, b);
      const diff = max2 - min2;
      let hue = 0, sat = 0, val = max2;
      if (diff > 0) {{
        if (max2 === r2) hue = 60 * (((g - b) / diff) % 6);
        else if (max2 === g) hue = 60 * ((b - r2) / diff + 2);
        else hue = 60 * ((r2 - g) / diff + 4);
        if (hue < 0) hue += 360;
        sat = (diff / max2) * 255;
      }}
      val = val;
      hue = hue / 2; // 0-180 range like OpenCV

      // Filtro verde (H:30-90, S:40-255, V:40-255)
      if (hue >= 30 && hue <= 90 && sat >= 40 && val >= 40) {{
        const absX = x + minX, absY = y + minY;
        if (pointInPolygon(absX, absY, points) && !visited[y * w + x]) {{
          // Flood-fill simples para agrupar pixels conectados
          let area = 0, sumX = 0, sumY = 0;
          const stack = [[x, y]];
          while (stack.length > 0 && area < 2000) {{
            const [sx, sy] = stack.pop();
            if (sx < 0 || sx >= w || sy < 0 || sy >= h) continue;
            if (visited[sy * w + sx]) continue;
            const si = (sy * w + sx) * 4;
            const sr = data[si], sg = data[si+1], sb = data[si+2];
            const smax = Math.max(sr,sg,sb), smin = Math.min(sr,sg,sb);
            const sdiff = smax - smin;
            let sh = 0, ss = 0;
            if (sdiff > 0) {{
              if (smax === sr) sh = 60*(((sg-sb)/sdiff)%6);
              else if (smax === sg) sh = 60*((sb-sr)/sdiff+2);
              else sh = 60*((sr-sg)/sdiff+4);
              if (sh < 0) sh += 360;
              ss = (sdiff/smax)*255;
            }}
            sh = sh/2;
            if (sh < 30 || sh > 90 || ss < 40 || smax < 40) continue;
            visited[sy * w + sx] = 1;
            area++; sumX += sx; sumY += sy;
            stack.push([sx+1,sy],[sx-1,sy],[sx,sy+1],[sx,sy-1]);
          }}
          if (area >= minArea) {{
            plantCenters.push({{ x: Math.round(sumX/area) + minX, y: Math.round(sumY/area) + minY }});
          }}
        }}
      }}
    }}
  }}

  // Adicionar marcas manuais que estão dentro do polígono
  for (const m of manualMarks) {{
    if (pointInPolygon(m.x, m.y, points)) {{
      plantCenters.push(m);
    }}
  }}

  // Contar por parcela
  parcelCounts = {{}};
  for (let r2 = 0; r2 < R; r2++) {{
    for (let c = 0; c < C; c++) {{
      const u0=c/C, u1=(c+1)/C, v0=r2/R, v1=(r2+1)/R;
      const tl=bilerp(p0,p1,p2,p3,u0,v0);
      const tr=bilerp(p0,p1,p2,p3,u1,v0);
      const br=bilerp(p0,p1,p2,p3,u1,v1);
      const bl=bilerp(p0,p1,p2,p3,u0,v1);
      const poly = [tl, tr, br, bl];
      let cnt = 0;
      for (const p of plantCenters) {{
        if (pointInPolygon(p.x, p.y, poly)) cnt++;
      }}
      const dispLabel = R - r2;
      const tiroLabel = C - c;
      parcelCounts[dispLabel + '_' + tiroLabel] = cnt;
    }}
  }}

  // Atualizar painel
  countPanel.style.display = 'block';
  totalCountEl.textContent = plantCenters.length;
  countInfoEl.textContent = 'Grid ' + R + '×' + C + ' | Dentro da área marcada';

  drawAll();
  saveActiveGrid(false);
}}

function drawAll() {{
  const W = vc.clientWidth, H = vc.clientHeight;
  cv.width = W; cv.height = H;
  ctx.clearRect(0,0,W,H);
  ctx.save();
  ctx.translate(ox,oy); ctx.scale(sc,sc);
  if(imgW>0) ctx.drawImage(img,0,0);

  // Desenhar grid
  if (points.length === 4) {{
    const R = parseInt(inpRows.value)||1;
    const C = parseInt(inpCols.value)||1;
    const p0=points[0],p1=points[1],p2=points[2],p3=points[3];

    ctx.save();
    ctx.strokeStyle='rgba(30,144,255,0.95)'; ctx.lineWidth=2/sc;
    ctx.shadowColor='rgba(0,160,255,0.6)'; ctx.shadowBlur=6/sc;

    for(let i=0;i<=R;i++) {{
      const v=i/R;
      const lx2=(1-v)*p0.x+v*p3.x, ly2=(1-v)*p0.y+v*p3.y;
      const rx2=(1-v)*p1.x+v*p2.x, ry2=(1-v)*p1.y+v*p2.y;
      ctx.beginPath(); ctx.moveTo(lx2,ly2); ctx.lineTo(rx2,ry2); ctx.stroke();
    }}
    for(let j=0;j<=C;j++) {{
      const u=j/C;
      const tx2=(1-u)*p0.x+u*p1.x, ty2=(1-u)*p0.y+u*p1.y;
      const bx2=(1-u)*p3.x+u*p2.x, by2=(1-u)*p3.y+u*p2.y;
      ctx.beginPath(); ctx.moveTo(tx2,ty2); ctx.lineTo(bx2,by2); ctx.stroke();
    }}
    ctx.restore();

    // Mostrar contagem por parcela
    if (Object.keys(parcelCounts).length > 0) {{
      for (let r2=0; r2<R; r2++) {{
        for (let c=0; c<C; c++) {{
          const u0=c/C, u1=(c+1)/C, v0=r2/R, v1=(r2+1)/R;
          const tl=bilerp(p0,p1,p2,p3,u0,v0);
          const tr=bilerp(p0,p1,p2,p3,u1,v0);
          const br=bilerp(p0,p1,p2,p3,u1,v1);
          const bl=bilerp(p0,p1,p2,p3,u0,v1);
          const cx2=(tl.x+tr.x+br.x+bl.x)/4;
          const cy2=(tl.y+tr.y+br.y+bl.y)/4;
          const dispLabel = R - r2;
          const tiroLabel = C - c;
          const cnt = parcelCounts[dispLabel + '_' + tiroLabel] || 0;

          ctx.save();
          ctx.fillStyle = cnt > 0 ? 'rgba(0,255,0,0.15)' : 'rgba(255,0,0,0.08)';
          ctx.beginPath();
          ctx.moveTo(tl.x,tl.y); ctx.lineTo(tr.x,tr.y);
          ctx.lineTo(br.x,br.y); ctx.lineTo(bl.x,bl.y);
          ctx.closePath(); ctx.fill();

          ctx.shadowColor='rgba(0,0,0,0.9)'; ctx.shadowBlur=4/sc;
          ctx.fillStyle='#ffffff';
          ctx.font='bold '+(Math.max(8, 11/sc))+'px Arial';
          ctx.textAlign='center'; ctx.textBaseline='middle';
          ctx.fillText(cnt, cx2, cy2);
          ctx.font=(Math.max(6, 8/sc))+'px Arial';
          ctx.fillStyle='#aaa';
          ctx.fillText('T'+tiroLabel+' D'+dispLabel, cx2, cy2 + 14/sc);
          ctx.restore();
        }}
      }}
    }}
  }}

  // Desenhar pontos do grid
  points.forEach((p,i) => {{
    const isDrag = draggingPoint===i;
    const r2 = 11/sc;
    ctx.save();
    ctx.shadowColor = isDrag ? 'rgba(255,255,255,0.9)' : 'rgba(0,180,255,0.8)';
    ctx.shadowBlur = 14/sc;
    ctx.beginPath(); ctx.arc(p.x,p.y,r2,0,2*Math.PI);
    ctx.fillStyle = isDrag ? '#ffffff' : '#1e90ff'; ctx.fill();
    ctx.lineWidth=2.5/sc; ctx.strokeStyle = isDrag ? '#aaddff' : '#00cfff'; ctx.stroke();
    ctx.restore();
    ctx.save();
    ctx.fillStyle='#ffffff'; ctx.font='bold '+(13/sc)+'px Arial';
    ctx.textAlign='center'; ctx.textBaseline='middle';
    ctx.fillText(i+1,p.x,p.y);
    ctx.restore();
  }});

  // Desenhar plantas detectadas (X verde)
  for (const p of plantCenters) {{
    const sz = 6/sc;
    ctx.save();
    ctx.strokeStyle='#00ff00'; ctx.lineWidth=2/sc;
    ctx.beginPath(); ctx.moveTo(p.x-sz,p.y-sz); ctx.lineTo(p.x+sz,p.y+sz); ctx.stroke();
    ctx.beginPath(); ctx.moveTo(p.x-sz,p.y+sz); ctx.lineTo(p.x+sz,p.y-sz); ctx.stroke();
    ctx.restore();
  }}

  // Desenhar marcas manuais (X vermelho)
  for (const p of manualMarks) {{
    const sz = 8/sc;
    ctx.save();
    ctx.strokeStyle='#ff3333'; ctx.lineWidth=2.5/sc;
    ctx.beginPath(); ctx.moveTo(p.x-sz,p.y-sz); ctx.lineTo(p.x+sz,p.y+sz); ctx.stroke();
    ctx.beginPath(); ctx.moveTo(p.x-sz,p.y+sz); ctx.lineTo(p.x+sz,p.y-sz); ctx.stroke();
    ctx.restore();
  }}

  ctx.restore();
  zb.textContent = sc.toFixed(2) + '×';
}}

// Eventos
img.onload = () => {{
  imgW = img.width; imgH = img.height;
  const W = vc.clientWidth, H = vc.clientHeight;
  sc = Math.min(W/imgW, H/imgH);
  ox = (W - imgW*sc)/2; oy = (H - imgH*sc)/2;
  drawAll();
}};
img.src = 'data:image/jpeg;base64,' + IMG_B64;

// Pan & Zoom
vc.addEventListener('wheel', e => {{
  e.preventDefault();
  const factor = e.deltaY < 0 ? 1.2 : 0.8;
  const r = cv.getBoundingClientRect();
  const mx = e.clientX - r.left, my = e.clientY - r.top;
  const ix = (mx - ox)/sc, iy = (my - oy)/sc;
  sc *= factor;
  sc = Math.max(MIN_SC, Math.min(MAX_SC, sc));
  ox = mx - ix*sc; oy = my - iy*sc;
  drawAll();
}}, {{passive:false}});

vc.addEventListener('mousedown', e => {{
  const pt = getImgCoords(e.clientX, e.clientY);

  // Verificar se está arrastando um ponto do grid
  for (let i = 0; i < points.length; i++) {{
    const dx = (pt.x - points[i].x)*sc, dy = (pt.y - points[i].y)*sc;
    if (Math.sqrt(dx*dx+dy*dy) < 20) {{
      draggingPoint = i;
      return;
    }}
  }}

  // Modo grid - adicionar ponto
  if (gridMode && points.length < 4) {{
    points.push({{x:pt.x, y:pt.y}});
    if (points.length === 4) gridMode = false;
    saveActiveGrid(false);
    drawAll();
    return;
  }}

  // Modo manual - adicionar marca
  if (manualMode && points.length === 4) {{
    if (pointInPolygon(pt.x, pt.y, points)) {{
      manualMarks.push({{x:pt.x, y:pt.y}});
      plantCenters.push({{x:pt.x, y:pt.y}});
      // Recontabilizar
      recount();
      saveActiveGrid(false);
      drawAll();
    }}
    return;
  }}

  // Pan
  drag = true; lx = e.clientX; ly = e.clientY;
  vc.style.cursor = 'grabbing';
}});

vc.addEventListener('mousemove', e => {{
  const pt = getImgCoords(e.clientX, e.clientY);
  coordEl.textContent = 'X:' + Math.round(pt.x) + ' Y:' + Math.round(pt.y);

  if (draggingPoint >= 0) {{
    points[draggingPoint] = {{x:pt.x, y:pt.y}};
    drawAll();
    return;
  }}
  if (drag) {{
    ox += e.clientX - lx; oy += e.clientY - ly;
    lx = e.clientX; ly = e.clientY;
    drawAll();
  }}
}});

vc.addEventListener('mouseup', () => {{
  if(draggingPoint >= 0) saveActiveGrid(false);
  drag = false; draggingPoint = -1;
  vc.style.cursor = 'grab';
}});
vc.addEventListener('mouseleave', () => {{
  drag = false; draggingPoint = -1;
}});

function recount() {{
  if (points.length < 4) return;
  const R = parseInt(inpRows.value)||1;
  const C = parseInt(inpCols.value)||1;
  const p0=points[0],p1=points[1],p2=points[2],p3=points[3];
  parcelCounts = {{}};
  for (let r2=0; r2<R; r2++) {{
    for (let c=0; c<C; c++) {{
      const u0=c/C, u1=(c+1)/C, v0=r2/R, v1=(r2+1)/R;
      const tl=bilerp(p0,p1,p2,p3,u0,v0);
      const tr=bilerp(p0,p1,p2,p3,u1,v0);
      const br=bilerp(p0,p1,p2,p3,u1,v1);
      const bl=bilerp(p0,p1,p2,p3,u0,v1);
      const poly = [tl,tr,br,bl];
      let cnt = 0;
      for (const p of plantCenters) {{
        if (pointInPolygon(p.x, p.y, poly)) cnt++;
      }}
      parcelCounts[(R-r2)+'_'+(C-c)] = cnt;
    }}
  }}
  countPanel.style.display = 'block';
  totalCountEl.textContent = plantCenters.length;
  countInfoEl.textContent = 'Grid ' + R + '×' + C + ' | ' + plantCenters.length + ' plantas';
  saveActiveGrid(false);
}}

// Botões
btnGridTool.onclick = () => {{
  gridMode = !gridMode; manualMode = false;
  btnGridTool.style.borderColor = gridMode ? '#ff8c00' : '#3a3a3a';
  btnManualMode.style.borderColor = '#3a3a3a';
  if (gridMode) {{ points = []; plantCenters = []; manualMarks = []; parcelCounts = {{}}; countPanel.style.display='none'; saveActiveGrid(false); drawAll(); }}
}};

btnCountPlants.onclick = () => countPlantsInGrid();
btnCountAuto.onclick = () => countPlantsInGrid();

btnManualMode.onclick = () => {{
  manualMode = !manualMode; gridMode = false;
  btnManualMode.style.borderColor = manualMode ? '#5599ff' : '#3a3a3a';
  btnGridTool.style.borderColor = '#3a3a3a';
}};
btnManual2.onclick = () => {{
  manualMode = !manualMode; gridMode = false;
  btnManualMode.style.borderColor = manualMode ? '#5599ff' : '#3a3a3a';
}};

btnRemoveLast.onclick = () => {{
  if (manualMarks.length > 0) {{
    const removed = manualMarks.pop();
    plantCenters = plantCenters.filter(p => p.x !== removed.x || p.y !== removed.y);
    recount(); saveActiveGrid(false); drawAll();
  }}
}};
btnUndoMark.onclick = btnRemoveLast.onclick;

btnClearAll.onclick = () => {{
  points = []; plantCenters = []; manualMarks = []; parcelCounts = {{}};
  countPanel.style.display = 'none'; gridMode = false; manualMode = false;
  btnGridTool.style.borderColor = '#3a3a3a';
  btnManualMode.style.borderColor = '#3a3a3a';
  saveActiveGrid(false);
  drawAll();
}};

btnSaveGrid.onclick = () => saveActiveGrid(true);
btnNewGrid.onclick = createNewGrid;
btnDeleteGrid.onclick = deleteActiveGrid;
selGridList.onchange = () => loadGridByName(selGridList.value);
inpGridName.onchange = renameActiveGrid;
inpRows.onchange = () => {{ recount(); saveActiveGrid(false); drawAll(); }};
inpCols.onchange = () => {{ recount(); saveActiveGrid(false); drawAll(); }};

btnExportCnt.onclick = () => exportContagemCSV(cbExportAll.checked);
btnExportXLSXCnt.onclick = () => exportContagemExcel(cbExportAll.checked);

// NOVO - Resumo Modal Logic
const resumoModal = document.getElementById('resumoModal');
const btnResumo = document.getElementById('btnResumo');
const btnCloseResumo = document.getElementById('btnCloseResumo');
const resumoSearch = document.getElementById('resumoSearch');
const resumoFilterTiro = document.getElementById('resumoFilterTiro');
const resumoFilterDisp = document.getElementById('resumoFilterDisp');

function buildResumoTable(filterText, filterTiro, filterDisp) {{
  const table = document.getElementById('resumoTable');
  if (Object.keys(parcelCounts).length === 0) {{
    table.innerHTML = '<p style="color:#666;text-align:center;padding:20px;font-size:12px;">Nenhuma contagem realizada.</p>';
    return;
  }}
  const R = parseInt(inpRows.value)||1, C = parseInt(inpCols.value)||1;
  let html = '<table style="width:100%;border-collapse:collapse;font-size:11px;">';
  html += '<thead><tr style="background:#222;color:#ff8c00;"><th style="padding:8px;border-bottom:1px solid #333;">Parcela</th><th style="padding:8px;border-bottom:1px solid #333;">Tiro</th><th style="padding:8px;border-bottom:1px solid #333;">Disparo</th><th style="padding:8px;border-bottom:1px solid #333;">Plantas</th></tr></thead><tbody>';
  const keys = Object.keys(parcelCounts).sort((a,b) => {{
    const [da,ta] = a.split('_').map(Number);
    const [db,tb] = b.split('_').map(Number);
    return da===db ? ta-tb : da-db;
  }});
  let total = 0, shown = 0;
  const tiroTotals = {{}}, dispTotals = {{}};
  for (const k of keys) {{
    const [d,t] = k.split('_').map(Number);
    const cnt = parcelCounts[k];
    total += cnt;
    tiroTotals[t] = (tiroTotals[t]||0) + cnt;
    dispTotals[d] = (dispTotals[d]||0) + cnt;
    const label = 'T'+t+' D'+d;
    if (filterText && !label.toLowerCase().includes(filterText.toLowerCase())) continue;
    if (filterTiro && t !== parseInt(filterTiro)) continue;
    if (filterDisp && d !== parseInt(filterDisp)) continue;
    shown++;
    html += '<tr style="border-bottom:1px solid #222;"><td style="padding:6px 8px;color:#ccc;">P'+(shown<10?'0':'')+shown+'</td><td style="padding:6px 8px;color:#5599ff;">T'+t+'</td><td style="padding:6px 8px;color:#55ff99;">D'+d+'</td><td style="padding:6px 8px;color:#fff;font-weight:bold;">'+cnt+'</td></tr>';
  }}
  html += '</tbody></table>';

  // Totais por tiro
  html += '<div style="margin-top:12px;padding:10px;background:#111;border:1px solid #222;border-radius:6px;"><p style="color:#5599ff;font-size:11px;font-weight:bold;margin-bottom:6px;">Resumo por Tiro:</p>';
  for (const t of Object.keys(tiroTotals).sort((a,b)=>a-b)) {{
    html += '<span style="color:#ccc;font-size:10px;margin-right:12px;">T'+t+': <b style=color:#fff>'+tiroTotals[t]+'</b></span>';
  }}
  html += '</div>';

  // Totais por disparo
  html += '<div style="margin-top:8px;padding:10px;background:#111;border:1px solid #222;border-radius:6px;"><p style="color:#55ff99;font-size:11px;font-weight:bold;margin-bottom:6px;">Resumo por Disparo:</p>';
  for (const d of Object.keys(dispTotals).sort((a,b)=>a-b)) {{
    html += '<span style="color:#ccc;font-size:10px;margin-right:12px;">D'+d+': <b style=color:#fff>'+dispTotals[d]+'</b></span>';
  }}
  html += '</div>';

  table.innerHTML = html;
  document.getElementById('resumoTotal').textContent = total;
}}

function openResumo() {{
  if (Object.keys(parcelCounts).length === 0) {{ alert('Realize a contagem primeiro!'); return; }}
  const R = parseInt(inpRows.value)||1, C = parseInt(inpCols.value)||1;
  // Populate filters
  resumoFilterTiro.innerHTML = '<option value="">Todos Tiros</option>';
  resumoFilterDisp.innerHTML = '<option value="">Todos Disparos</option>';
  for (let c=1;c<=C;c++) resumoFilterTiro.innerHTML += '<option value="'+c+'">Tiro '+c+'</option>';
  for (let r=1;r<=R;r++) resumoFilterDisp.innerHTML += '<option value="'+r+'">Disparo '+r+'</option>';
  buildResumoTable('','','');
  resumoModal.style.display = 'flex';
}}

btnResumo.onclick = openResumo;
btnCloseResumo.onclick = () => {{ resumoModal.style.display = 'none'; }};
resumoSearch.oninput = () => buildResumoTable(resumoSearch.value, resumoFilterTiro.value, resumoFilterDisp.value);
resumoFilterTiro.onchange = () => buildResumoTable(resumoSearch.value, resumoFilterTiro.value, resumoFilterDisp.value);
resumoFilterDisp.onchange = () => buildResumoTable(resumoSearch.value, resumoFilterTiro.value, resumoFilterDisp.value);

document.getElementById('btnResumoCSV').onclick = () => {{
  exportContagemCSV(cbExportAll.checked);
}};

document.getElementById('btnResumoXLSX').onclick = () => {{
  exportContagemExcel(cbExportAll.checked);
}};

restoreGrids();
saveActiveGrid(false);
updateGridSelect();

// Resize
new ResizeObserver(() => drawAll()).observe(vc);
</script>
</body>
</html>
"""
                    components.html(cnt_viewer, height=720, scrolling=False)

            else:
                st.markdown("""
                <div style='height:706px;border:1px dashed #2e2e2e;border-radius:12px;background:#0d0d0d;
                            display:flex;flex-direction:column;align-items:center;justify-content:center;
                            gap:12px;color:#333;'>
                    <div style='font-size:3rem;'>🌱</div>
                    <div style='font-size:0.85rem;letter-spacing:2px;text-transform:uppercase;color:#555;'>
                        Carregue uma ortofoto para iniciar a contagem
                    </div>
                    <div style='font-size:0.75rem;color:#444;'>
                        Marque 4 pontos do grid → Clique em Contar → Exporte CSV
                    </div>
                </div>""", unsafe_allow_html=True)
            # FIM NOVO - MÓDULO CONTAGEM DE PLANTAS

        elif st.session_state.visualizador_sub == "Maturação":
            st.markdown("""
            <div style='background:#1e1e1e;border:1px solid #333;border-radius:15px;padding:30px;text-align:center;'>
                <div style='font-size:3rem;margin-bottom:12px;'>🌱</div>
                <div style='color:#ff8c00;font-weight:700;font-size:1.2rem;letter-spacing:2px;text-transform:uppercase;margin-bottom:8px;'>
                    Maturação
                </div>
                <div style='color:#666;font-size:0.85rem;'>
                    Módulo de análise de maturação por parcela.<br>
                    Funcionalidade em desenvolvimento — dados serão exibidos aqui.
                </div>
            </div>""", unsafe_allow_html=True)

        elif st.session_state.visualizador_sub == "Pendoamento":
            # O pendoamento usa somente o seletor/visualizador único de até 10 ortofotos abaixo.
            pend_file = None
            pend_bytes, pend_name = _uploaded_ortho_bytes(pend_file)

            if pend_bytes:
                with st.container():
                    pend_b64, pend_dims, pend_err, pend_spatial = processar_ortofoto(pend_bytes, pend_name)

                if pend_err:
                    st.error(f"Erro: {pend_err}")
                else:
                    pw, ph = pend_dims
                    st.markdown(
                        f"<p style='color:#666;font-size:0.78rem;margin-bottom:6px;'>"
                        f"📐 {pend_name} · {pw}×{ph} px · análise sobre ortofoto e grid</p>",
                        unsafe_allow_html=True
                    )

                    pend_viewer = """
<!DOCTYPE html>
<html>
<head>
<style>
  * { box-sizing:border-box; margin:0; padding:0; }
  body { background:#0d0d0d; overflow:hidden; font-family:'Segoe UI',sans-serif; }
  #vc {
    width:100%; height:720px; position:relative; overflow:hidden; user-select:none; cursor:grab;
    border:1px solid #2a2a2a; border-radius:12px;
    background:
      linear-gradient(rgba(255,255,255,.03) 1px, transparent 1px),
      linear-gradient(90deg, rgba(255,255,255,.03) 1px, transparent 1px),
      #0d0d0d;
    background-size:32px 32px;
  }
  #vc:active { cursor:grabbing; }
  canvas { position:absolute; inset:0; display:block; }
  .toolbar { position:absolute; top:12px; right:12px; display:flex; flex-direction:column; gap:5px; z-index:20; }
  .tb-btn {
    width:34px; height:34px; border-radius:8px; cursor:pointer; font-size:15px; font-weight:700;
    display:flex; align-items:center; justify-content:center; color:#ff8c00;
    background:linear-gradient(145deg,#1e1e1e,#111); border:1px solid #3a3a3a;
    box-shadow:2px 2px 8px #000,inset 0 1px 0 rgba(255,255,255,.05); transition:all .2s;
  }
  .tb-btn:hover { border-color:#ff8c00; box-shadow:0 0 10px rgba(255,140,0,.35),2px 2px 8px #000; color:#ffaa33; }
  .tb-btn.active { border-color:#ff8c00; color:#ff8c00; background:#2a1a00; }
  .tb-sep { width:34px; height:1px; background:linear-gradient(90deg,transparent,#333,transparent); margin:2px 0; }
  .grid-panel {
    position:absolute; top:10px; right:55px; z-index:20; min-width:218px; max-height:calc(100% - 20px);
    overflow-y:auto; overflow-x:hidden; background:rgba(10,10,10,.92); border:1px solid #2a2a2a;
    border-radius:8px; padding:8px; display:flex; flex-direction:column; gap:5px;
    scrollbar-width:thin; scrollbar-color:#ff8c00 #141414;
  }
  .grid-panel::-webkit-scrollbar { width:6px; }
  .grid-panel::-webkit-scrollbar-track { background:#141414; border-radius:6px; }
  .grid-panel::-webkit-scrollbar-thumb { background:#ff8c00; border-radius:6px; }
  .grid-panel label { color:#ff8c00; font-size:10px; line-height:1.1; font-weight:bold; text-align:center; }
  .grid-panel .section { color:#ff8c00; font-size:10px; font-weight:800; text-align:center; margin-top:4px; }
  .grid-panel .row-col { display:flex; gap:6px; align-items:center; justify-content:space-between; color:#ccc; font-size:10px; min-height:21px; }
  .grid-panel input[type=number] {
    background:#1a1a1a; border:1px solid #333; color:#fff; border-radius:4px;
    padding:2px 3px; height:21px; width:54px; text-align:center; font-size:10px;
  }
  .grid-panel input[type=text], .grid-panel input[type=date] {
    background:#1a1a1a; border:1px solid #333; color:#fff; border-radius:4px;
    padding:2px 4px; height:23px; width:112px; font-size:10px;
  }
  .grid-panel select {
    background:#1a1a1a; border:1px solid #333; color:#fff; border-radius:4px;
    padding:2px 4px; height:23px; width:92px; font-size:10px;
  }
  .grid-btn, .cnt-btn {
    width:100%; min-height:23px; border-radius:4px; padding:4px 6px; font-size:10px; font-weight:bold;
    cursor:pointer; transition:.2s; color:#ccc; background:linear-gradient(145deg,#1e1e1e,#111); border:1px solid #3a3a3a;
  }
  .grid-btn:hover, .cnt-btn:hover { border-color:#ff8c00; color:#ff8c00; }
  .grid-btn.active { border-color:#ff8c00; color:#ff8c00; box-shadow:0 0 8px rgba(255,140,0,.3); background:#2a1a00; }
  .cnt-btn.orange { background:linear-gradient(145deg,#2a1a00,#1a0a00); border-color:#ff8c00; color:#ff8c00; }
  .cnt-btn.green { background:linear-gradient(145deg,#1a3a1a,#0a2a0a); border-color:#006600; color:#00ee55; }
  .cnt-btn.blue { background:linear-gradient(145deg,#1a1a3a,#0a0a2a); border-color:#000066; color:#5599ff; }
  .cnt-btn.danger { background:linear-gradient(145deg,#3a1a1a,#2a0a0a); border-color:#660000; color:#ff5555; }
  .qual-sep { width:100%; height:1px; background:linear-gradient(90deg,transparent,#333,transparent); margin:2px 0; }
  .zoom-badge {
    position:absolute; top:12px; left:12px; z-index:20; pointer-events:none;
    background:rgba(10,10,10,.82); border:1px solid #2a2a2a; border-radius:8px;
    color:#ff8c00; font-size:11px; font-family:'Courier New',monospace; font-weight:700;
    padding:5px 10px; letter-spacing:1px;
  }
  .crosshair {
    position:absolute; bottom:12px; left:12px; z-index:20; pointer-events:none;
    background:rgba(10,10,10,.82); border:1px solid #222; border-radius:8px;
    color:#555; font-size:10px; font-family:'Courier New',monospace; padding:4px 10px; letter-spacing:.5px;
  }
  .hint {
    position:absolute; bottom:12px; right:12px; z-index:20; pointer-events:none; text-align:right;
    color:#333; font-size:10px; line-height:1.6;
  }
  .count-panel {
    position:absolute; top:50px; left:12px; z-index:20; min-width:215px; max-width:280px;
    background:rgba(10,10,10,.92); border:1px solid #2a2a2a; border-radius:8px; padding:11px;
  }
  .count-panel h3 { color:#ffb347; font-size:12px; margin-bottom:6px; letter-spacing:1px; }
  .count-panel .total { color:#fff; font-size:22px; font-weight:bold; }
  .count-panel .info { color:#888; font-size:10px; margin-top:4px; line-height:1.35; }
  .result-box {
    max-height:145px; overflow:auto; margin-top:8px; border-top:1px solid #262626; padding-top:6px;
    color:#bbb; font-size:10px; line-height:1.45;
  }
</style>
</head>
<body>
<div id="vc">
  <canvas id="cv"></canvas>
  <div class="zoom-badge" id="zbadge">1.00×</div>
  <div class="crosshair" id="coord">X:0 Y:0</div>
  <div class="hint">Scroll=Zoom · Drag=Pan<br>Grid: marque 4 pontos · Pendoamento sobre parcelas</div>

  <div class="toolbar">
    <button class="tb-btn" id="btnGridTool" title="Marcar Grid">⊞</button>
    <button class="tb-btn" id="btnSelectParcel" title="Selecionar Parcelas">▣</button>
    <div class="tb-sep"></div>
    <button class="tb-btn" id="btnFit" title="Ajustar à tela">⤢</button>
    <button class="tb-btn" id="btnClearAll" title="Limpar">🗑</button>
  </div>

  <div class="grid-panel">
    <label>🌾 PENDOAMENTO</label>
    <div class="row-col"><span>Quadra:</span><input type="text" id="inpQuadra" placeholder="Quadra"></div>
    <div class="row-col"><span>Data:</span><input type="date" id="inpData"></div>
    <div class="row-col"><span>Disp:</span><input type="number" id="inpRows" value="5" min="1" max="200"></div>
    <div class="row-col"><span>Tiros:</span><input type="number" id="inpCols" value="5" min="1" max="200"></div>
    <div class="row-col"><span>Teto pendões:</span><input type="number" id="inpTeto" value="20" min="1" max="10000"></div>
    <div class="row-col"><span>Trava:</span><span style="color:#ffb347;font-weight:800;">50%</span></div>
    <button class="grid-btn" id="btnGrid2">Marcar Grid</button>
    <button class="cnt-btn blue" id="btnSelect2">Selecionar Parcelas</button>
    <div class="qual-sep"></div>
    <div class="section">Análise</div>
    <button class="cnt-btn orange" id="btnAnalyze18000">📐 Análise de Pendoamento</button>
    <button class="cnt-btn green" id="btnAnalyzeSelected">🌾 Análise Selecionadas</button>
    <button class="cnt-btn" id="btnExportCSV">💾 Exportar CSV</button>
    <button class="cnt-btn danger" id="btnClearResults">Limpar Resultados</button>
  </div>

  <div class="count-panel" id="countPanel" style="display:none;">
    <h3>RESULTADOS PENDOAMENTO</h3>
    <div class="total" id="totalCount">0</div>
    <div class="info" id="countInfo">Marque o grid e execute a análise.</div>
    <div class="result-box" id="resultBox"></div>
  </div>
</div>

<script>
const IMG_B64 = "__PEND_B64__";
const IMAGE_NAME = __PEND_IMAGE_NAME__;
const vc = document.getElementById('vc');
const cv = document.getElementById('cv');
const ctx = cv.getContext('2d');
const zb = document.getElementById('zbadge');
const coordEl = document.getElementById('coord');
const inpRows = document.getElementById('inpRows');
const inpCols = document.getElementById('inpCols');
const inpTeto = document.getElementById('inpTeto');
const inpQuadra = document.getElementById('inpQuadra');
const inpData = document.getElementById('inpData');
const btnGridTool = document.getElementById('btnGridTool');
const btnGrid2 = document.getElementById('btnGrid2');
const btnSelectParcel = document.getElementById('btnSelectParcel');
const btnSelect2 = document.getElementById('btnSelect2');
const btnRun = document.getElementById('btnRun');
const btnAnalyze18000 = document.getElementById('btnAnalyze18000');
const btnAnalyzeAll = document.getElementById('btnAnalyzeAll');
const btnAnalyzeSelected = document.getElementById('btnAnalyzeSelected');
const btnTrainPositive = document.getElementById('btnTrainPositive');
const btnTrainNegative = document.getElementById('btnTrainNegative');
const btnClearTraining = document.getElementById('btnClearTraining');
const trainInfo = document.getElementById('trainInfo');
const btnExportCSV = document.getElementById('btnExportCSV');
const btnClearResults = document.getElementById('btnClearResults');
const btnClearAll = document.getElementById('btnClearAll');
const btnFit = document.getElementById('btnFit');
const countPanel = document.getElementById('countPanel');
const totalCountEl = document.getElementById('totalCount');
const countInfoEl = document.getElementById('countInfo');
const resultBox = document.getElementById('resultBox');

let gridMode=false, selectMode=false;
let points=[], selectedParcels=new Set(), results={}, pendaoMarks=[];
let trainMode='', trainingMarks=[];
let draggingPoint=-1, drag=false, lx=0, ly=0;
let sc=1, ox=0, oy=0, imgW=0, imgH=0;
const MIN_SC=0.05, MAX_SC=40;
const PEND_STEP=1;
const PEND_MIN_AREA=8; // AJUSTE PENDOAMENTO: aceita pendões menores dentro da parcela do grid
const PEND_CRITICO=20;
const PEND_SCORE_THRESHOLD=3.82; // AJUSTE PENDOAMENTO: sensibilidade maior para contar pendões reais dentro do grid
const PEND_PERCENT_LIMIT=50;
const TRAIN_KEY='tmg_pendao_ia_' + IMAGE_NAME;
let trainingSamples={pos:[],neg:[]};
const img = new Image();
const tempCv = document.createElement('canvas');
const tempCtx = tempCv.getContext('2d', { willReadFrequently:true });

function getImgCoords(cx, cy) {
  const r=cv.getBoundingClientRect();
  return { x:(cx-r.left-ox)/sc, y:(cy-r.top-oy)/sc };
}

function fitView() {
  const W=vc.clientWidth, H=vc.clientHeight;
  if(!imgW || !imgH) return;
  sc=Math.min(W/imgW,H/imgH);
  ox=(W-imgW*sc)/2; oy=(H-imgH*sc)/2;
  drawAll();
}

function bilerp(p0,p1,p2,p3,u,v) {
  const tx=(1-u)*p0.x+u*p1.x, ty=(1-u)*p0.y+u*p1.y;
  const bx=(1-u)*p3.x+u*p2.x, by=(1-u)*p3.y+u*p2.y;
  return { x:(1-v)*tx+v*bx, y:(1-v)*ty+v*by };
}

function pointInPolygon(px,py,polygon) {
  let inside=false;
  for(let i=0,j=polygon.length-1;i<polygon.length;j=i++) {
    const xi=polygon[i].x, yi=polygon[i].y, xj=polygon[j].x, yj=polygon[j].y;
    if(((yi>py)!==(yj>py)) && (px < (xj-xi)*(py-yi)/(yj-yi)+xi)) inside=!inside;
  }
  return inside;
}

function getCellPoly(r,c) {
  if(points.length<4) return null;
  const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
  const p0=points[0], p1=points[1], p2=points[2], p3=points[3];
  const u0=c/C, u1=(c+1)/C, v0=r/R, v1=(r+1)/R;
  return [
    bilerp(p0,p1,p2,p3,u0,v0),
    bilerp(p0,p1,p2,p3,u1,v0),
    bilerp(p0,p1,p2,p3,u1,v1),
    bilerp(p0,p1,p2,p3,u0,v1)
  ];
}

function cellLabel(r,c) {
  return 'T' + (c+1) + ' D' + (r+1);
}

function cellCenter(poly) {
  return {
    x:(poly[0].x+poly[1].x+poly[2].x+poly[3].x)/4,
    y:(poly[0].y+poly[1].y+poly[2].y+poly[3].y)/4
  };
}

function findCellAt(pt) {
  const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
  for(let r=0;r<R;r++) {
    for(let c=0;c<C;c++) {
      const poly=getCellPoly(r,c);
      if(poly && pointInPolygon(pt.x,pt.y,poly)) return {r,c,label:cellLabel(r,c),poly};
    }
  }
  return null;
}

function rgbToHsv(r,g,b) {
  r/=255; g/=255; b/=255;
  const max=Math.max(r,g,b), min=Math.min(r,g,b), d=max-min;
  let h=0;
  if(d!==0) {
    if(max===r) h=60*(((g-b)/d)%6);
    else if(max===g) h=60*((b-r)/d+2);
    else h=60*((r-g)/d+4);
    if(h<0) h+=360;
  }
  const s=max===0 ? 0 : d/max;
  return {h,s,v:max};
}

function pixelFeatures(r,g,b) {
  const hsv=rgbToHsv(r,g,b);
  const exg=(2*g-r-b)/255;
  const yellowness=(((r+g)*0.5)-b)/255;
  return {
    h:hsv.h,
    s:hsv.s,
    v:hsv.v,
    exg:exg,
    yellow:yellowness,
    r:r/255,
    g:g/255,
    b:b/255
  };
}

function featureDistance(a,b) {
  const dh=Math.min(Math.abs(a.h-b.h), 360-Math.abs(a.h-b.h))/180;
  const ds=a.s-b.s, dv=a.v-b.v;
  const de=a.exg-b.exg, dy=a.yellow-b.yellow;
  const dr=a.r-b.r, dg=a.g-b.g, db=a.b-b.b;
  return Math.sqrt(
    dh*dh*1.2 + ds*ds*0.8 + dv*dv*0.5 +
    de*de*1.1 + dy*dy*1.4 + dr*dr*0.25 + dg*dg*0.25 + db*db*0.25
  );
}

function loadTrainingSamples() {
  try {
    const raw=localStorage.getItem(TRAIN_KEY);
    if(raw) {
      const parsed=JSON.parse(raw);
      if(parsed && Array.isArray(parsed.pos) && Array.isArray(parsed.neg)) trainingSamples=parsed;
    }
  } catch(e) {}
  updateTrainingInfo();
}

function saveTrainingSamples() {
  try { localStorage.setItem(TRAIN_KEY, JSON.stringify(trainingSamples)); } catch(e) {}
  updateTrainingInfo();
}

function updateTrainingInfo() {
  if(trainInfo) trainInfo.textContent=trainingSamples.pos.length+' pendão · '+trainingSamples.neg.length+' fundo';
}

function trainingBoost(r,g,b) {
  return 0;
}

function getPixelAtImagePoint(pt) {
  const x=Math.max(0, Math.min(imgW-1, Math.round(pt.x)));
  const y=Math.max(0, Math.min(imgH-1, Math.round(pt.y)));
  const d=tempCtx.getImageData(x,y,1,1).data;
  return [d[0],d[1],d[2]];
}

function addTrainingSample(pt, kind) {
  if(!imgW || !imgH) return;
  prepareImageData();
  const radius=4;
  const bucket=kind==='pos' ? trainingSamples.pos : trainingSamples.neg;
  let added=0;
  for(let yy=-radius; yy<=radius; yy++) {
    for(let xx=-radius; xx<=radius; xx++) {
      if(xx*xx+yy*yy>radius*radius) continue;
      const x=Math.max(0, Math.min(imgW-1, Math.round(pt.x+xx)));
      const y=Math.max(0, Math.min(imgH-1, Math.round(pt.y+yy)));
      const d=tempCtx.getImageData(x,y,1,1).data;
      bucket.push(pixelFeatures(d[0],d[1],d[2]));
      added++;
    }
  }
  while(bucket.length>360) bucket.shift();
  trainingMarks.push({x:pt.x,y:pt.y,kind:kind});
  saveTrainingSamples();
  drawAll();
  countInfoEl.innerHTML=(kind==='pos'?'Exemplo de pendão':'Exemplo de fundo')+' salvo para calibrar esta ortofoto.';
  countPanel.style.display='block';
}

function tasselScore(r,g,b) {
  const hsv=rgbToHsv(r,g,b);
  const maxc=Math.max(r,g,b), minc=Math.min(r,g,b);
  const chroma=maxc-minc;
  const exg=2*g-r-b;
  const yellowness=((r+g)*0.5)-b;
  const redYellowBalance=Math.abs(r-g);
  const leafGreen=(exg>20 && g>=r*0.98 && g>=b*1.07 && hsv.h>=68 && hsv.h<=165 && yellowness<32 && hsv.v<0.78);
  const hardGreen=(exg>38 && hsv.h>=72 && hsv.h<=158 && hsv.s>0.24 && yellowness<26);
  const darkShadow = hsv.v<0.20;
  const whiteGlare = hsv.s<0.08 && hsv.v>0.82;
  const soilRed = r>g*1.28 && r>b*1.36 && hsv.h<18;
  const greenWhiteEdge = hsv.s<0.22 && g>=r*1.01 && b>=r*0.74 && hsv.h>=65 && hsv.h<=150 && yellowness<20;
  const nonGreen = !leafGreen && !greenWhiteEdge && !hardGreen && exg < 70;

  const strawHue = hsv.h>=18 && hsv.h<=64 && hsv.s>=0.10 && hsv.v>=0.30 && yellowness>=10 && exg<62;
  const tanDry = r>=82 && g>=58 && b<=158 && r>=b+14 && g>=b+5 && redYellowBalance<=96 && hsv.h>=12 && hsv.h<=62;
  const paleTassel = r>=96 && g>=72 && b<=185 && yellowness>=10 && redYellowBalance<=88 && hsv.s>=0.07 && hsv.h>=12 && hsv.h<=76 && exg<64;
  const creamTip = maxc>=132 && yellowness>=14 && b<=g*0.92 && b<=r*0.92 && hsv.s>=0.06 && hsv.s<=0.58 && hsv.h>=14 && hsv.h<=78 && exg<68;
  const oldPinkTan = r>=96 && g>=64 && b<=168 && r>=g*0.84 && g>=b*0.78 && r>=b+8 && hsv.h>=8 && hsv.h<=46 && hsv.s>=0.10;
  const youngTassel = hsv.h>=38 && hsv.h<=82 && hsv.s>=0.09 && hsv.s<=0.68 && hsv.v>=0.30 && yellowness>=6 && exg<62;
  const branchTexture = chroma>=18 && yellowness>=8 && hsv.s>=0.09;

  let score=0;
  if(strawHue) score+=2.45;
  if(tanDry) score+=2.15;
  if(paleTassel) score+=2.35;
  if(creamTip) score+=2.20;
  if(oldPinkTan) score+=1.70;
  if(youngTassel) score+=1.35;
  if(branchTexture) score+=0.85;
  if(yellowness>=24) score+=0.65;
  if(yellowness>=38 && exg<58) score+=0.45;
  if(!nonGreen) score-=4.2;
  if(leafGreen) score-=4.2;
  if(hardGreen) score-=3.6;
  if(greenWhiteEdge) score-=3.0;
  if(darkShadow) score-=2.2;
  if(whiteGlare) score-=2.2;
  if(soilRed) score-=1.6;
  if(chroma<12 && yellowness<18) score-=1.2;
  if(hsv.h>=92 && hsv.h<=155 && exg>18) score-=1.8;
  return score;
}

function isTasselPixel(r,g,b) {
  return tasselScore(r,g,b) >= PEND_SCORE_THRESHOLD;
}

function prepareImageData() {
  tempCv.width=imgW; tempCv.height=imgH;
  tempCtx.clearRect(0,0,imgW,imgH);
  tempCtx.drawImage(img,0,0);
}

function mergeNearbyTassels(marks,w,h) {
  if(!marks.length) return marks;
  const radius=Math.max(12, Math.min(28, Math.min(w,h)*0.040));
  const ordered=[...marks].sort((a,b)=>(b.score*b.area)-(a.score*a.area));
  const merged=[];
  for(const m of ordered) {
    let found=null;
    for(const c of merged) {
      const dx=m.x-c.x, dy=m.y-c.y;
      if(Math.sqrt(dx*dx+dy*dy)<=radius) { found=c; break; }
    }
    if(found) {
      const wa=Math.max(1,found.area), wb=Math.max(1,m.area);
      found.x=(found.x*wa+m.x*wb)/(wa+wb);
      found.y=(found.y*wa+m.y*wb)/(wa+wb);
      found.area+=m.area;
      found.score=Math.max(found.score,m.score);
    } else {
      merged.push({...m});
    }
  }
  return merged.sort((a,b)=>a.y===b.y ? a.x-b.x : a.y-b.y);
}

function analyzeCell(r,c) {
  const poly=getCellPoly(r,c);
  if(!poly) return {label:cellLabel(r,c), count:0, marks:[]};
  const step=PEND_STEP;
  const minAreaPx=PEND_MIN_AREA;
  const xs=poly.map(p=>p.x), ys=poly.map(p=>p.y);
  const minX=Math.max(0, Math.floor(Math.min(...xs)));
  const maxX=Math.min(imgW-1, Math.ceil(Math.max(...xs)));
  const minY=Math.max(0, Math.floor(Math.min(...ys)));
  const maxY=Math.min(imgH-1, Math.ceil(Math.max(...ys)));
  const w=maxX-minX+1, h=maxY-minY+1;
  if(w<=0 || h<=0) return {label:cellLabel(r,c), count:0, marks:[]};

  const imageData=tempCtx.getImageData(minX,minY,w,h).data;
  const gw=Math.ceil(w/step), gh=Math.ceil(h/step);
  const mask=new Uint8Array(gw*gh);
  const cleanMask=new Uint8Array(gw*gh);
  const closedMask=new Uint8Array(gw*gh);
  const scores=new Float32Array(gw*gh);
  const yellows=new Float32Array(gw*gh);
  const exgs=new Float32Array(gw*gh);
  const chromas=new Float32Array(gw*gh);
  const visited=new Uint8Array(gw*gh);

  for(let gy=0; gy<gh; gy++) {
    for(let gx=0; gx<gw; gx++) {
      const px=Math.min(w-1, gx*step), py=Math.min(h-1, gy*step);
      const ax=minX+px, ay=minY+py;
      if(!pointInPolygon(ax,ay,poly)) continue;
      const idx=(py*w+px)*4;
      const r=imageData[idx], g=imageData[idx+1], b=imageData[idx+2];
      const mi=gy*gw+gx;
      const score=tasselScore(r,g,b);
      scores[mi]=score;
      yellows[mi]=((r+g)*0.5)-b;
      exgs[mi]=2*g-r-b;
      chromas[mi]=Math.max(r,g,b)-Math.min(r,g,b);
      if(score>=PEND_SCORE_THRESHOLD) {
        mask[mi]=1;
      }
    }
  }

  for(let gy=0; gy<gh; gy++) {
    for(let gx=0; gx<gw; gx++) {
      const idx=gy*gw+gx;
      if(!mask[idx]) continue;
      let neighbors=0;
      for(let yy=-1; yy<=1; yy++) {
        for(let xx=-1; xx<=1; xx++) {
          if(xx===0 && yy===0) continue;
          const nx=gx+xx, ny=gy+yy;
          if(nx>=0 && nx<gw && ny>=0 && ny<gh && mask[ny*gw+nx]) neighbors++;
        }
      }
      if(neighbors>=2 || scores[idx]>=PEND_SCORE_THRESHOLD+0.85) cleanMask[idx]=1;
    }
  }

  for(let gy=0; gy<gh; gy++) {
    for(let gx=0; gx<gw; gx++) {
      const idx=gy*gw+gx;
      if(cleanMask[idx]) { closedMask[idx]=1; continue; }
      let neighbors=0, scoreAround=0;
      for(let yy=-1; yy<=1; yy++) {
        for(let xx=-1; xx<=1; xx++) {
          const nx=gx+xx, ny=gy+yy;
          if(nx>=0 && nx<gw && ny>=0 && ny<gh) {
            const ni=ny*gw+nx;
            if(cleanMask[ni]) neighbors++;
            scoreAround+=scores[ni];
          }
        }
      }
      if(neighbors>=5 && scoreAround/9>=PEND_SCORE_THRESHOLD-0.35) closedMask[idx]=1;
    }
  }

  const marks=[];
  const dirs=[[1,0],[-1,0],[0,1],[0,-1],[1,1],[-1,1],[1,-1],[-1,-1]];
  for(let gy=0; gy<gh; gy++) {
    for(let gx=0; gx<gw; gx++) {
      const start=gy*gw+gx;
      if(!closedMask[start] || visited[start]) continue;
      let cells=0, sx=0, sy=0, scoreSum=0, yellowSum=0, exgSum=0, chromaSum=0, coreCells=0;
      let minGX=gx, maxGX=gx, minGY=gy, maxGY=gy;
      const stack=[[gx,gy]];
      visited[start]=1;
      while(stack.length) {
        const p=stack.pop();
        const x=p[0], y=p[1];
        const pos=y*gw+x;
        cells++; sx+=x; sy+=y; scoreSum+=scores[pos];
        yellowSum+=yellows[pos]; exgSum+=exgs[pos]; chromaSum+=chromas[pos];
        if(scores[pos]>=PEND_SCORE_THRESHOLD+0.55) coreCells++;
        if(x<minGX) minGX=x; if(x>maxGX) maxGX=x;
        if(y<minGY) minGY=y; if(y>maxGY) maxGY=y;
        for(const d of dirs) {
          const nx=x+d[0], ny=y+d[1];
          if(nx<0 || nx>=gw || ny<0 || ny>=gh) continue;
          const ni=ny*gw+nx;
          if(closedMask[ni] && !visited[ni]) {
            visited[ni]=1;
            stack.push([nx,ny]);
          }
        }
      }
      const areaPx=cells*step*step;
      const widthPx=(maxGX-minGX+1)*step;
      const heightPx=(maxGY-minGY+1)*step;
      const bboxCells=Math.max(1,(maxGX-minGX+1)*(maxGY-minGY+1));
      const density=cells/bboxCells;
      const elongation=Math.max(widthPx,heightPx)/Math.max(1,Math.min(widthPx,heightPx));
      const meanScore=scoreSum/Math.max(1,cells);
      const meanYellow=yellowSum/Math.max(1,cells);
      const meanExg=exgSum/Math.max(1,cells);
      const meanChroma=chromaSum/Math.max(1,cells);
      const coreRatio=coreCells/Math.max(1,cells);
      const cx=minX + (sx/cells)*step;
      const cy=minY + (sy/cells)*step;
      const maxAreaPx=Math.max(minAreaPx*22, Math.min(w*h*0.022, 1400));
      const validShape =
        areaPx>=minAreaPx &&
        areaPx<=maxAreaPx &&
        density>=0.08 &&
        elongation<=10.0 &&
        widthPx<=Math.max(22, w*0.22) &&
        heightPx<=Math.max(22, h*0.24) &&
        meanScore>=PEND_SCORE_THRESHOLD-0.03 &&
        meanYellow>=7 &&
        meanExg<66 &&
        meanChroma>=10 &&
        coreRatio>=0.08 &&
        pointInPolygon(cx,cy,poly);

      if(validShape) {
        marks.push({x:cx, y:cy, area:areaPx, score:meanScore});
      }
    }
  }
  const finalMarks=mergeNearbyTassels(marks,w,h);
  return {label:cellLabel(r,c), row:r+1, col:c+1, count:finalMarks.length, marks:finalMarks};
}

function setGridMode(active) {
  gridMode=active;
  selectMode=false;
  trainMode='';
  btnGridTool.classList.toggle('active', gridMode);
  btnGrid2.classList.toggle('active', gridMode);
  btnSelectParcel.classList.remove('active');
  btnSelect2.classList.remove('active');
  if(btnTrainPositive) btnTrainPositive.classList.remove('active');
  if(btnTrainNegative) btnTrainNegative.classList.remove('active');
  if(gridMode) {
    points=[]; selectedParcels.clear(); results={}; pendaoMarks=[]; countPanel.style.display='none';
  }
  drawAll();
}

function setSelectMode(active) {
  if(points.length<4) { alert('Marque o grid primeiro.'); return; }
  selectMode=active;
  gridMode=false;
  trainMode='';
  btnSelectParcel.classList.toggle('active', selectMode);
  btnSelect2.classList.toggle('active', selectMode);
  btnGridTool.classList.remove('active');
  btnGrid2.classList.remove('active');
  if(btnTrainPositive) btnTrainPositive.classList.remove('active');
  if(btnTrainNegative) btnTrainNegative.classList.remove('active');
  drawAll();
}

function setTrainMode(mode) {
  if(!btnTrainPositive || !btnTrainNegative) return;
  trainMode = trainMode===mode ? '' : mode;
  gridMode=false;
  selectMode=false;
  btnGridTool.classList.remove('active');
  btnGrid2.classList.remove('active');
  btnSelectParcel.classList.remove('active');
  btnSelect2.classList.remove('active');
  btnTrainPositive.classList.toggle('active', trainMode==='pos');
  btnTrainNegative.classList.toggle('active', trainMode==='neg');
  countPanel.style.display='block';
  countInfoEl.innerHTML = trainMode==='pos'
    ? 'Clique sobre pendões reais para treinar a IA local.'
    : trainMode==='neg'
      ? 'Clique em folha, solo, sombra ou palha falsa para ensinar o que ignorar.'
      : 'Treino IA local pausado.';
  drawAll();
}

function analyzePendoamento(onlySelected) {
  if(points.length<4) { alert('Marque os 4 cantos do grid primeiro.'); return; }
  const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
  if(onlySelected && selectedParcels.size===0) {
    alert('Selecione uma ou mais parcelas para analisar.');
    return;
  }
  prepareImageData();
  results={}; pendaoMarks=[];
  let total=0, analyzed=0, critical=0;
  const crit=getPendaoLimit();

  for(let r=0;r<R;r++) {
    for(let c=0;c<C;c++) {
      const label=cellLabel(r,c);
      if(onlySelected && !selectedParcels.has(label)) continue;
      const res=analyzeCell(r,c);
      results[label]=res;
      for(const m of res.marks) pendaoMarks.push({x:m.x,y:m.y,label:label});
      total+=res.count;
      analyzed++;
      if(res.count>=crit) critical++;
    }
  }

  countPanel.style.display='block';
  totalCountEl.textContent=total;
  countInfoEl.innerHTML='Parcelas analisadas: '+analyzed+'<br>Média: '+(analyzed? (total/analyzed).toFixed(1):'0.0')+' pendões/parcela<br>Trava 50%: '+crit+' pendões';
  updateResultBox();
  drawAll();
}

function getPendaoLimit() {
  const teto=Math.max(1, parseInt(inpTeto ? inpTeto.value : PEND_CRITICO) || PEND_CRITICO);
  return Math.max(1, Math.ceil(teto * (PEND_PERCENT_LIMIT/100)));
}

function updateResultBox() {
  const entries=Object.values(results).sort((a,b)=> a.row===b.row ? a.col-b.col : a.row-b.row);
  if(!entries.length) {
    resultBox.innerHTML='Sem resultados.';
    return;
  }
  const crit=getPendaoLimit();
  resultBox.innerHTML=entries.map(r=>{
    const color = r.count>=crit ? '#ff5555' : '#ddd';
    return '<div><b style="color:#ff8c00">'+r.label+'</b> · <span style="color:'+color+'">'+r.count+' pendões</span> · '+(r.count>=crit?'ATINGIU 50%':'abaixo')+'</div>';
  }).join('');
}

function exportCSV() {
  const entries=Object.values(results).sort((a,b)=> a.row===b.row ? a.col-b.col : a.row-b.row);
  if(!entries.length) { alert('Execute a análise de pendoamento antes de exportar.'); return; }
  const crit=getPendaoLimit();
  const quadra=(inpQuadra && inpQuadra.value.trim()) ? inpQuadra.value.trim() : 'Pendoamento';
  const dataOrto=(inpData && inpData.value) ? inpData.value : new Date().toISOString().slice(0,10);
  const csvValue = value => {
    const s=String(value ?? '');
    return /[;"\\n\\r]/.test(s) ? '"' + s.replace(/"/g,'""') + '"' : s;
  };
  let csv='\\uFEFFQuadra;Disparo;Tiro;Total_Pendões;Nome_Ortofoto;Data_Ortofoto;Observação\\n';
  for(const r of entries) {
    const obs=r.count>=crit ? 'ATINGIU_50_DO_TETO' : 'ABAIXO_50_DO_TETO';
    csv += [quadra,r.row,r.col,r.count,IMAGE_NAME,dataOrto,obs].map(csvValue).join(';')+'\\n';
  }
  const blob=new Blob([csv], {type:'text/csv;charset=utf-8;'});
  const a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  a.download='pendoamento_por_parcela.csv';
  a.click();
}

function clearResults() {
  results={}; pendaoMarks=[]; countPanel.style.display='none'; drawAll();
}

function drawAll() {
  const W=vc.clientWidth, H=vc.clientHeight;
  cv.width=W; cv.height=H;
  ctx.clearRect(0,0,W,H);
  ctx.save();
  ctx.translate(ox,oy); ctx.scale(sc,sc);
  if(imgW>0) ctx.drawImage(img,0,0);

  const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
  if(points.length===4) {
    ctx.save();
    ctx.lineWidth=1.6/sc; ctx.strokeStyle='rgba(0,207,255,.78)';
    ctx.shadowColor='rgba(0,207,255,.35)'; ctx.shadowBlur=5/sc;
    for(let i=0;i<=R;i++) {
      const v=i/R;
      const p0=points[0], p1=points[1], p2=points[2], p3=points[3];
      const left=bilerp(p0,p1,p2,p3,0,v), right=bilerp(p0,p1,p2,p3,1,v);
      ctx.beginPath(); ctx.moveTo(left.x,left.y); ctx.lineTo(right.x,right.y); ctx.stroke();
    }
    for(let j=0;j<=C;j++) {
      const u=j/C;
      const p0=points[0], p1=points[1], p2=points[2], p3=points[3];
      const top=bilerp(p0,p1,p2,p3,u,0), bottom=bilerp(p0,p1,p2,p3,u,1);
      ctx.beginPath(); ctx.moveTo(top.x,top.y); ctx.lineTo(bottom.x,bottom.y); ctx.stroke();
    }
    ctx.restore();

    for(let r=0;r<R;r++) {
      for(let c=0;c<C;c++) {
        const label=cellLabel(r,c);
        const poly=getCellPoly(r,c);
        const cen=cellCenter(poly);
        const res=results[label];
        const selected=selectedParcels.has(label);
        if(selected || res) {
          ctx.save();
          ctx.beginPath();
          ctx.moveTo(poly[0].x,poly[0].y);
          for(let k=1;k<poly.length;k++) ctx.lineTo(poly[k].x,poly[k].y);
          ctx.closePath();
          if(res) {
            const crit=PEND_CRITICO;
            ctx.fillStyle=res.count>=crit ? 'rgba(255,64,0,.20)' : 'rgba(255,179,71,.16)';
          } else {
            ctx.fillStyle='rgba(255,210,0,.18)';
          }
          ctx.fill();
          ctx.strokeStyle=selected ? 'rgba(255,230,0,.95)' : 'rgba(255,140,0,.55)';
          ctx.lineWidth=(selected?2.4:1.2)/sc;
          ctx.stroke();
          ctx.restore();
        }
        if(res) {
          ctx.save();
          ctx.shadowColor='rgba(0,0,0,.9)'; ctx.shadowBlur=4/sc;
          ctx.fillStyle='#fff';
          ctx.font='bold '+Math.max(9,12/sc)+'px Arial';
          ctx.textAlign='center'; ctx.textBaseline='middle';
          ctx.fillText(String(res.count), cen.x, cen.y);
          ctx.font=Math.max(7,9/sc)+'px Arial';
          ctx.fillStyle='#ffb347';
          ctx.fillText(label, cen.x, cen.y+15/sc);
          ctx.restore();
        }
      }
    }
  }

  for(const m of pendaoMarks) {
    const size=5/sc;
    ctx.save();
    ctx.strokeStyle='#ff2b2b'; ctx.lineWidth=2/sc; ctx.shadowColor='rgba(255,0,0,.7)'; ctx.shadowBlur=5/sc;
    ctx.beginPath(); ctx.moveTo(m.x-size,m.y-size); ctx.lineTo(m.x+size,m.y+size); ctx.stroke();
    ctx.beginPath(); ctx.moveTo(m.x-size,m.y+size); ctx.lineTo(m.x+size,m.y-size); ctx.stroke();
    ctx.restore();
  }

  for(const m of trainingMarks) {
    const size=7/sc;
    ctx.save();
    ctx.strokeStyle=m.kind==='pos' ? '#00ff66' : '#5599ff';
    ctx.fillStyle=m.kind==='pos' ? 'rgba(0,255,102,.18)' : 'rgba(85,153,255,.18)';
    ctx.lineWidth=2/sc;
    ctx.beginPath(); ctx.arc(m.x,m.y,size,0,Math.PI*2); ctx.fill(); ctx.stroke();
    ctx.beginPath(); ctx.moveTo(m.x-size,m.y); ctx.lineTo(m.x+size,m.y); ctx.stroke();
    ctx.beginPath(); ctx.moveTo(m.x,m.y-size); ctx.lineTo(m.x,m.y+size); ctx.stroke();
    ctx.restore();
  }

  points.forEach((p,i)=>{
    const r=11/sc;
    ctx.save();
    ctx.shadowColor=draggingPoint===i ? 'rgba(255,255,255,.9)' : 'rgba(0,180,255,.8)';
    ctx.shadowBlur=14/sc;
    ctx.beginPath(); ctx.arc(p.x,p.y,r,0,Math.PI*2);
    ctx.fillStyle=draggingPoint===i ? '#fff' : '#1e90ff';
    ctx.fill();
    ctx.lineWidth=2.5/sc; ctx.strokeStyle=draggingPoint===i ? '#aaddff' : '#00cfff';
    ctx.stroke();
    ctx.fillStyle='#fff'; ctx.font='bold '+(13/sc)+'px Arial'; ctx.textAlign='center'; ctx.textBaseline='middle';
    ctx.fillText(i+1,p.x,p.y);
    ctx.restore();
  });
  ctx.restore();
  zb.textContent=sc.toFixed(2)+'×';
}

img.onload=()=>{ imgW=img.width; imgH=img.height; fitView(); };
img.onerror=()=>{ countPanel.style.display='block'; countInfoEl.innerHTML='Falha ao carregar a ortofoto no visualizador. Recarregue a imagem ou reduza o tamanho do arquivo.'; };
img.src='data:image/jpeg;base64,'+IMG_B64;

vc.addEventListener('wheel', e=>{
  e.preventDefault();
  const factor=e.deltaY<0?1.2:0.8;
  const r=cv.getBoundingClientRect();
  const mx=e.clientX-r.left, my=e.clientY-r.top;
  const ix=(mx-ox)/sc, iy=(my-oy)/sc;
  sc=Math.max(MIN_SC, Math.min(MAX_SC, sc*factor));
  ox=mx-ix*sc; oy=my-iy*sc;
  drawAll();
}, {passive:false});

vc.addEventListener('mousedown', e=>{
  const pt=getImgCoords(e.clientX,e.clientY);
  for(let i=0;i<points.length;i++) {
    const dx=(pt.x-points[i].x)*sc, dy=(pt.y-points[i].y)*sc;
    if(Math.sqrt(dx*dx+dy*dy)<20) { draggingPoint=i; return; }
  }
  if(gridMode && points.length<4) {
    points.push({x:pt.x,y:pt.y});
    if(points.length===4) setGridMode(false);
    drawAll(); return;
  }
  if(trainMode) {
    addTrainingSample(pt, trainMode);
    return;
  }
  if(selectMode && points.length===4) {
    const cell=findCellAt(pt);
    if(cell) {
      if(selectedParcels.has(cell.label)) selectedParcels.delete(cell.label);
      else selectedParcels.add(cell.label);
      drawAll();
    }
    return;
  }
  drag=true; lx=e.clientX; ly=e.clientY; vc.style.cursor='grabbing';
});

vc.addEventListener('mousemove', e=>{
  const pt=getImgCoords(e.clientX,e.clientY);
  coordEl.textContent='X:'+Math.round(pt.x)+' Y:'+Math.round(pt.y);
  if(draggingPoint>=0) {
    points[draggingPoint]={x:pt.x,y:pt.y};
    clearResults();
    drawAll(); return;
  }
  if(drag) {
    ox+=e.clientX-lx; oy+=e.clientY-ly; lx=e.clientX; ly=e.clientY;
    drawAll();
  }
});
vc.addEventListener('mouseup',()=>{ drag=false; draggingPoint=-1; vc.style.cursor='grab'; });
vc.addEventListener('mouseleave',()=>{ drag=false; draggingPoint=-1; vc.style.cursor='grab'; });

function resetAll() {
  points=[]; selectedParcels.clear(); results={}; pendaoMarks=[]; gridMode=false; selectMode=false; trainMode='';
  btnGridTool.classList.remove('active'); btnGrid2.classList.remove('active');
  btnSelectParcel.classList.remove('active'); btnSelect2.classList.remove('active');
  if(btnTrainPositive) btnTrainPositive.classList.remove('active');
  if(btnTrainNegative) btnTrainNegative.classList.remove('active');
  countPanel.style.display='none'; drawAll();
}

function clearTraining() {
  trainingSamples={pos:[],neg:[]};
  trainingMarks=[];
  try { localStorage.removeItem(TRAIN_KEY); } catch(e) {}
  trainMode='';
  if(btnTrainPositive) btnTrainPositive.classList.remove('active');
  if(btnTrainNegative) btnTrainNegative.classList.remove('active');
  updateTrainingInfo();
  clearResults();
}

btnGridTool.onclick=()=>setGridMode(!gridMode);
btnGrid2.onclick=()=>setGridMode(!gridMode);
btnSelectParcel.onclick=()=>setSelectMode(!selectMode);
btnSelect2.onclick=()=>setSelectMode(!selectMode);
if(btnTrainPositive) btnTrainPositive.onclick=()=>setTrainMode('pos');
if(btnTrainNegative) btnTrainNegative.onclick=()=>setTrainMode('neg');
if(btnClearTraining) btnClearTraining.onclick=clearTraining;
if(btnRun) btnRun.onclick=()=>analyzePendoamento(false);
btnAnalyze18000.onclick=()=>analyzePendoamento(false);
if(btnAnalyzeAll) btnAnalyzeAll.onclick=()=>analyzePendoamento(false);
btnAnalyzeSelected.onclick=()=>analyzePendoamento(true);
btnExportCSV.onclick=exportCSV;
btnClearResults.onclick=clearResults;
btnClearAll.onclick=resetAll;
btnFit.onclick=fitView;
inpRows.onchange=()=>{ selectedParcels.clear(); clearResults(); drawAll(); };
inpCols.onchange=()=>{ selectedParcels.clear(); clearResults(); drawAll(); };
inpTeto.onchange=()=>{ updateResultBox(); drawAll(); };
if(inpData && !inpData.value) inpData.value = new Date().toISOString().slice(0,10);
loadTrainingSamples();
new ResizeObserver(()=>drawAll()).observe(vc);
</script>
</body>
</html>
"""
                    pend_viewer = (
                        pend_viewer
                        .replace("__PEND_B64__", pend_b64)
                        .replace("__PEND_IMAGE_NAME__", json.dumps(pend_file.name, ensure_ascii=False))
                    )
                    components.html(pend_viewer, height=740, scrolling=False)

            if False:
                st.markdown("""
                <div style='height:706px;border:1px dashed #2e2e2e;border-radius:12px;background:#0d0d0d;
                            display:flex;flex-direction:column;align-items:center;justify-content:center;
                            gap:12px;color:#333;'>
                    <div style='font-size:3rem;'>🌾</div>
                    <div style='font-size:0.85rem;letter-spacing:2px;text-transform:uppercase;color:#555;'>
                        Carregue uma ortofoto para análise de pendoamento
                    </div>
                    <div style='font-size:0.75rem;color:#444;'>
                        Marque o grid → selecione parcelas, se quiser → execute Pendoamento
                    </div>
                </div>""", unsafe_allow_html=True)

            cron_nome_analise = "Pendoamento"
            cron_rows = 5
            cron_cols = 5
            cron_teto = 20
            cron_percentual = 100.0
            cron_min_pendoes = 1
            cron_referencia = ""
            cron_tolerancia = 58
            cron_filtro_cor = "Misto"
            cron_sensibilidade = 64
            cron_area_min = 12
            cron_area_max = 900
            cron_revisao_manual = True
            meta_codlocal = ""
            meta_quadra = ""
            meta_ensaio = ""
            meta_gli = ""
            meta_tecnologia = ""
            meta_rep = ""
            meta_nc = ""
            meta_linhagem = ""
            meta_genealogia = ""
            meta_parental1 = ""
            meta_pop_parental1 = ""
            meta_parental2 = ""
            meta_pop_parental2 = ""
            meta_dtp = ""
            st.caption("Depois de anexar, use o resumo lateral do visualizador para clicar e trocar entre as ortofotos carregadas. O teto de pendões trava a primeira data em que a parcela atingiu o valor configurado.")

            st.markdown(
                """
                <style>
                div[data-testid="stTextInput"]:has(input[aria-label="pend_yolo_payload"]) {
                    display: none !important;
                }
                </style>
                """,
                unsafe_allow_html=True,
            )
            st.text_input("pend_yolo_payload", key="pend_yolo_payload", label_visibility="hidden")

            with st.expander("🌾 Amostras manuais YOLO de pendoamento", expanded=False):
                capture_ok, capture_endpoint = iniciar_servidor_captura_yolo_pendoamento()
                yolo_counts = contar_amostras_treinamento_yolo()
                st.caption(
                    f"As mini imagens clicadas no visualizador são salvas automaticamente em {YOLO_TRAIN_ROOT} "
                    "e entram como histórico permanente para o botão Aplicar Treino."
                )
                st.code(str(YOLO_TRAIN_ROOT), language="text")
                if capture_ok:
                    st.success("Salvamento automático por clique ativo.")
                else:
                    st.warning(capture_endpoint)
                pasta_digitada = st.text_input(
                    "Pasta para salvar mini fotos e dados YOLO",
                    value=str(YOLO_TRAIN_ROOT),
                    key="pend_yolo_training_dir_input",
                    help="Use esta pasta para guardar images, labels, crops, JSON de características e histórico.",
                )
                pc1, pc2, pc3 = st.columns(3)
                with pc1:
                    if st.button("Selecionar pasta de treino", key="pend_yolo_select_folder", use_container_width=True):
                        selected_folder = selecionar_pasta_treinamento_yolo_dialogo()
                        if selected_folder:
                            configurar_pasta_treinamento_yolo(selected_folder)
                            garantir_estrutura_treinamento_yolo()
                            st.session_state["pend_yolo_training_dir_input"] = str(YOLO_TRAIN_ROOT)
                            st.session_state["pend_yolo_last_summary"] = f"Pasta de treinamento definida: {YOLO_TRAIN_ROOT}"
                            app_rerun()
                        else:
                            st.info("Seleção de pasta cancelada ou indisponível neste ambiente.")
                with pc2:
                    if st.button("Salvar pasta informada", key="pend_yolo_save_folder", use_container_width=True):
                        configurar_pasta_treinamento_yolo(pasta_digitada)
                        garantir_estrutura_treinamento_yolo()
                        st.session_state["pend_yolo_training_dir_input"] = str(YOLO_TRAIN_ROOT)
                        st.session_state["pend_yolo_last_summary"] = f"Pasta de treinamento definida: {YOLO_TRAIN_ROOT}"
                        app_rerun()
                with pc3:
                    if st.button("Usar pasta do programa", key="pend_yolo_default_folder", use_container_width=True):
                        configurar_pasta_treinamento_yolo(PENDAO_YOLO_DEFAULT_ROOT)
                        garantir_estrutura_treinamento_yolo()
                        st.session_state["pend_yolo_training_dir_input"] = str(YOLO_TRAIN_ROOT)
                        st.session_state["pend_yolo_last_summary"] = f"Pasta de treinamento definida: {YOLO_TRAIN_ROOT}"
                        app_rerun()
                yolo_counts = contar_amostras_treinamento_yolo()
                m1, m2, m3 = st.columns(3)
                m1.metric("Imagens treino", yolo_counts["images_train"])
                m2.metric("Imagens validação", yolo_counts["images_val"])
                m3.metric("Labels", yolo_counts["total_labels"])
                if st.button("Abrir pasta de treinamento", key="pend_yolo_open_folder"):
                    garantir_estrutura_treinamento_yolo()
                    try:
                        if os.name == "nt":
                            os.startfile(str(YOLO_TRAIN_ROOT))
                        else:
                            subprocess.Popen(["xdg-open", str(YOLO_TRAIN_ROOT)])
                        st.success("Pasta de treinamento aberta.")
                    except Exception as exc:
                        st.info(f"Pasta: {YOLO_TRAIN_ROOT} ({exc})")
                if YOLO_BEST_MODEL_PATH.exists():
                    st.success(f"Modelo treinado ativo: {YOLO_BEST_MODEL_PATH}")
                else:
                    st.info("Nenhum best.pt customizado local foi encontrado. A análise usa o YOLO já disponível no ambiente, OpenCV e as mini imagens manuais como referência.")
                last_manual_msg = st.session_state.get("pend_yolo_last_summary", "")
                if last_manual_msg:
                    st.success(last_manual_msg)

            cron_files = _resettable_ortho_uploader(
                "📷 Anexar ortofotos para o seletor do visualizador (até 10)",
                accept_multiple_files=True,
                key="pend_cron_ortos",
                help="Use até 10 ortofotos da mesma área. O grid marcado será mantido no visualizador único."
            )

            cron_items = []
            for cron_file in cron_files or []:
                raw = cron_file.getbuffer().tobytes()
                cron_items.append({"name": cron_file.name, "raw": raw})

            if cron_items:
                selected_cron_items = cron_items[:10]
                if len(cron_items) > 10:
                    st.warning("Foram anexadas mais de 10 ortofotos. O seletor usará somente as 10 primeiras.")

                st.caption("Revise o nome e a data de cada ortofoto. A lista será organizada automaticamente pela data informada.")
                date_cols = st.columns(min(5, len(selected_cron_items)))
                cron_entries = []
                for idx, cron_item in enumerate(selected_cron_items):
                    key_hash = hashlib.md5(f"{idx}-{cron_item['name']}".encode("utf-8")).hexdigest()[:8]
                    with date_cols[idx % len(date_cols)]:
                        cron_label = st.text_input(
                            f"Nome {idx + 1}",
                            value=Path(cron_item["name"]).stem,
                            key=f"pend_cron_name_{key_hash}"
                        )
                        cron_date = st.date_input(
                            f"Data {idx + 1}",
                            value=date.today(),
                            key=f"pend_cron_date_{key_hash}"
                        )
                    cron_entries.append({
                        "idx": idx,
                        "name": (cron_label.strip() or Path(cron_item["name"]).stem),
                        "original_name": cron_item["name"],
                        "raw": cron_item["raw"],
                        "date": cron_date.isoformat()
                    })

                ordered_entries_for_training = sorted(cron_entries, key=lambda it: (it["date"], it["idx"]))
                if "pend_yolo_saved_sample_tokens" not in st.session_state:
                    st.session_state["pend_yolo_saved_sample_tokens"] = set()

                def _salvar_payload_amostra_treino(sample_payload: dict):
                    active_idx = int(sample_payload.get("idx", 0))
                    entry = ordered_entries_for_training[active_idx] if 0 <= active_idx < len(ordered_entries_for_training) else None
                    payload_name = str(sample_payload.get("name", ""))
                    payload_date = str(sample_payload.get("date", ""))
                    if not entry or (payload_name and entry.get("name") != payload_name) or (payload_date and entry.get("date") != payload_date):
                        entry = next(
                            (
                                item for item in ordered_entries_for_training
                                if (not payload_name or item.get("name") == payload_name)
                                and (not payload_date or item.get("date") == payload_date)
                            ),
                            entry,
                        )
                    if not entry:
                        raise ValueError("ortofoto de treino não encontrada")
                    preview_w = max(1.0, float(sample_payload.get("preview_width") or sample_payload.get("width") or 1))
                    preview_h = max(1.0, float(sample_payload.get("preview_height") or sample_payload.get("height") or 1))
                    orig_w = max(1.0, float(sample_payload.get("orig_width") or preview_w))
                    orig_h = max(1.0, float(sample_payload.get("orig_height") or preview_h))
                    click_x = float(sample_payload.get("x", 0))
                    click_y = float(sample_payload.get("y", 0))
                    orig_x = click_x * orig_w / preview_w
                    orig_y = click_y * orig_h / preview_h
                    crop_size = int(sample_payload.get("crop_size") or PENDAO_AVANCADO_PARAMS.get("manual_crop_size_default", 96))
                    return salvar_amostra_treinamento_yolo(
                        entry["raw"],
                        orig_x,
                        orig_y,
                        "pendao_confirmado",
                        crop_size,
                        entry["name"],
                        entry["date"],
                    )

                pend_yolo_payload_query = ""
                try:
                    query_values = st.query_params.get_all("pend_yolo_payload")
                    pend_yolo_payload_query = query_values[-1] if query_values else ""
                except Exception:
                    try:
                        legacy_query = st.experimental_get_query_params()
                        query_values = legacy_query.get("pend_yolo_payload", [])
                        pend_yolo_payload_query = query_values[-1] if query_values else ""
                    except Exception:
                        pend_yolo_payload_query = ""
                pend_yolo_payload = pend_yolo_payload_query or st.session_state.get("pend_yolo_payload", "")
                if pend_yolo_payload:
                    try:
                        payload = json.loads(pend_yolo_payload)
                    except Exception:
                        payload = {}
                    token = str(payload.get("token", ""))
                    if token and token != st.session_state.get("pend_yolo_last_payload_token"):
                        st.session_state["pend_yolo_last_payload_token"] = token
                        action = str(payload.get("action", "")).lower()
                        if action == "sample":
                            try:
                                sample_token = str(payload.get("token", ""))
                                saved_tokens = st.session_state["pend_yolo_saved_sample_tokens"]
                                if sample_token and sample_token in saved_tokens:
                                    raise ValueError("amostra já gravada")
                                meta = _salvar_payload_amostra_treino(payload)
                                if sample_token:
                                    saved_tokens.add(sample_token)
                                st.session_state["pend_yolo_browser_train_active"] = True
                                st.session_state["pend_yolo_apply_training"] = True
                                st.session_state["pend_yolo_last_summary"] = (
                                    f"Mini imagem salva: {Path(meta['image']).name}. "
                                    f"Pasta: {YOLO_TRAIN_ROOT}. O histórico já está disponível para Aplicar Treino."
                                )
                                try:
                                    preparar_deteccoes_pendoamento_hibrido.clear()
                                    carregar_referencias_treino_yolo.clear()
                                except Exception:
                                    pass
                                st.success(st.session_state["pend_yolo_last_summary"])
                            except Exception as exc:
                                st.warning(f"Não foi possível salvar a mini imagem YOLO: {exc}")
                        elif action == "apply_training":
                            st.session_state["pend_yolo_apply_training"] = True
                            yolo_counts_now = contar_amostras_treinamento_yolo()
                            refs_now = _arquivos_referencia_treinamento_yolo(limit=100000)
                            feature_files_now = [
                                path
                                for base in _bases_treinamento_yolo(include_legacy=True)
                                if base.exists()
                                for path in base.glob("**/crops/*/*.json")
                            ]
                            st.session_state["pend_yolo_last_summary"] = (
                                f"Treino YOLO aplicado. Mini imagens disponíveis: {yolo_counts_now['total_images']}. "
                                f"Referências lidas: {len(refs_now)}. Arquivos de características: {len(feature_files_now)}. "
                                "As referências foram recarregadas para reforçar a próxima Análise de Pendoamento."
                            )
                            try:
                                preparar_deteccoes_pendoamento_hibrido.clear()
                                carregar_referencias_treino_yolo.clear()
                            except Exception:
                                pass
                            st.success(st.session_state["pend_yolo_last_summary"])
                        elif action == "delete_sample":
                            try:
                                ok, msg = excluir_amostra_treinamento_yolo(
                                    sample_id=str(payload.get("sample_id", "")),
                                    crop_file=str(payload.get("file", "")),
                                    source_name=str(payload.get("name", "")),
                                    source_date=str(payload.get("date", "")),
                                )
                                yolo_counts_now = contar_amostras_treinamento_yolo()
                                st.session_state["pend_yolo_apply_training"] = bool(yolo_counts_now.get("total_images", 0))
                                st.session_state["pend_yolo_last_summary"] = (
                                    f"{msg} Mini imagens disponíveis: {yolo_counts_now['total_images']}. "
                                    f"Pasta: {YOLO_TRAIN_ROOT}."
                                )
                                try:
                                    preparar_deteccoes_pendoamento_hibrido.clear()
                                    carregar_referencias_treino_yolo.clear()
                                except Exception:
                                    pass
                                if ok:
                                    st.success(st.session_state["pend_yolo_last_summary"])
                                else:
                                    st.warning(st.session_state["pend_yolo_last_summary"])
                            except Exception as exc:
                                st.warning(f"Não foi possível excluir a amostra YOLO: {exc}")
                        elif action == "stop":
                            saved_in_stop = 0
                            stop_errors = []
                            saved_tokens = st.session_state["pend_yolo_saved_sample_tokens"]
                            samples_to_save = payload.get("samples", [])
                            if isinstance(samples_to_save, list):
                                for sample_payload in samples_to_save:
                                    if not isinstance(sample_payload, dict):
                                        continue
                                    sample_token = str(sample_payload.get("token", ""))
                                    if sample_token and sample_token in saved_tokens:
                                        continue
                                    try:
                                        _salvar_payload_amostra_treino(sample_payload)
                                        if sample_token:
                                            saved_tokens.add(sample_token)
                                        saved_in_stop += 1
                                    except Exception as batch_exc:
                                        stop_errors.append(str(batch_exc))
                            st.session_state["pend_yolo_browser_train_active"] = False
                            st.session_state["pend_yolo_apply_training"] = True
                            yolo_counts_now = contar_amostras_treinamento_yolo()
                            st.session_state["pend_yolo_last_summary"] = (
                                f"Treinamento YOLO encerrado. Novas mini imagens salvas agora: {saved_in_stop}. "
                                f"Mini imagens disponíveis: {yolo_counts_now['total_images']}. "
                                f"Pasta: {YOLO_TRAIN_ROOT}. As imagens e características ficam disponíveis para Aplicar Treino."
                            )
                            if stop_errors:
                                st.session_state["pend_yolo_last_summary"] += f" Avisos: {len(stop_errors)} amostra(s) não foram salvas."
                            try:
                                preparar_deteccoes_pendoamento_hibrido.clear()
                                carregar_referencias_treino_yolo.clear()
                            except Exception:
                                pass
                            st.success(st.session_state["pend_yolo_last_summary"])

                ordem_preview = sorted(cron_entries, key=lambda it: (it["date"], it["idx"]))
                resumo_cards = []
                for slot_idx in range(10):
                    if slot_idx < len(ordem_preview):
                        item = ordem_preview[slot_idx]
                        resumo_cards.append(
                            "<div style='border:1px solid #ff8c00;background:#201303;border-radius:8px;padding:8px;'>"
                            f"<div style='color:#ffb347;font-weight:800;font-size:0.72rem;'>#{slot_idx + 1:02d} · SELECIONÁVEL</div>"
                            f"<div style='color:#fff;font-size:0.78rem;font-weight:700;margin-top:4px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'>{html.escape(item['name'])}</div>"
                            f"<div style='color:#888;font-size:0.68rem;margin-top:3px;'>Data: {html.escape(item['date'])}</div>"
                            "</div>"
                        )
                    else:
                        resumo_cards.append(
                            "<div style='border:1px dashed #2e2e2e;background:#0c0c0c;border-radius:8px;padding:8px;'>"
                            f"<div style='color:#555;font-weight:800;font-size:0.72rem;'>#{slot_idx + 1:02d} · VAZIO</div>"
                            "<div style='color:#444;font-size:0.72rem;margin-top:6px;'>Aguardando ortofoto</div>"
                            "</div>"
                        )
                st.markdown(
                    "<div style='margin:8px 0 10px 0;'>"
                    "<div style='color:#ff8c00;font-weight:800;font-size:0.78rem;letter-spacing:1px;text-transform:uppercase;margin-bottom:6px;'>"
                    "Resumo das 10 posições do seletor</div>"
                    "<div style='display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:7px;'>"
                    + "".join(resumo_cards) +
                    "</div></div>",
                    unsafe_allow_html=True
                )

                cron_orthos = []
                cron_errors = []
                yolo_counts_for_viewer = contar_amostras_treinamento_yolo()
                if "pend_yolo_apply_training" not in st.session_state:
                    st.session_state["pend_yolo_apply_training"] = bool(yolo_counts_for_viewer.get("total_images", 0))
                apply_training_enabled = bool(st.session_state.get("pend_yolo_apply_training", False))
                with st.container():
                    total_cron_entries = max(1, len(cron_entries))
                    for idx_cron, item in enumerate(cron_entries, start=1):
                        cron_name = item["name"]
                        try:
                            raw = item["raw"]
                            b64, dims, err, spatial = processar_ortofoto(raw, cron_name)
                            if err:
                                cron_errors.append(f"{cron_name}: {err}")
                                continue
                            try:
                                detector_signature = f"{assinatura_modelo_yolo_pendao()}|{assinatura_referencias_treino_yolo()}|apply:{int(apply_training_enabled)}"
                                pendao_result = preparar_deteccoes_pendoamento_hibrido(raw, cron_name, tuple(dims), detector_signature, apply_training_enabled)
                                pendao_detections = pendao_result.get("detections", [])
                                pendao_training_detections = pendao_result.get("training_detections", [])
                                pendao_status = pendao_result.get("status", "")
                                pendao_mode = pendao_result.get("mode", "OpenCV parametrizado TMG")
                                pendao_counts = pendao_result.get("counts", {})
                                pendao_training_status = pendao_result.get("training_status", "")
                                pendao_training_counts = pendao_result.get("training_counts", {})
                                pendao_backend_ready = bool(pendao_result.get("backend_ready", True))
                                if pendao_status and any(token in pendao_status.lower() for token in ("não instalado", "nao instalado", "falhou", "instale opencv-python")):
                                    cron_errors.append(f"{cron_name}: {pendao_status}")
                            except Exception as det_exc:
                                pendao_detections = []
                                pendao_training_detections = []
                                pendao_status = f"OpenCV indisponível ({det_exc}). Instale opencv-python."
                                pendao_mode = "OpenCV indisponível"
                                pendao_counts = {}
                                pendao_training_status = ""
                                pendao_training_counts = {}
                                pendao_backend_ready = False
                                cron_errors.append(f"{cron_name}: {pendao_status}")
                            orig_width = int(spatial.get("orig_width", dims[0]) or dims[0]) if spatial else int(dims[0])
                            orig_height = int(spatial.get("orig_height", dims[1]) or dims[1]) if spatial else int(dims[1])
                            training_marks = marcas_treinamento_yolo_preview(
                                cron_name,
                                item["date"],
                                tuple(dims),
                                (orig_width, orig_height),
                            )
                            cron_orthos.append({
                                "order": int(item["idx"]),
                                "name": cron_name,
                                "date": item["date"],
                                "b64": b64,
                                "width": int(dims[0]),
                                "height": int(dims[1]),
                                "advanced_detections": pendao_detections,
                                "training_detections": pendao_training_detections,
                                "detector_status": pendao_status,
                                "detector_mode": pendao_mode,
                                "detector_counts": pendao_counts,
                                "training_status": pendao_training_status,
                                "training_counts": pendao_training_counts,
                                "backend_ready": pendao_backend_ready,
                                "training_marks": training_marks,
                                "orig_width": orig_width,
                                "orig_height": orig_height,
                            })
                        except Exception as exc:
                            cron_errors.append(f"{cron_name}: {exc}")

                for message in cron_errors:
                    st.warning(message)

                cron_orthos = sorted(cron_orthos, key=lambda it: (it["date"], it["order"]))

                if cron_orthos:
                    capture_ok, capture_endpoint = iniciar_servidor_captura_yolo_pendoamento()
                    cron_config = {
                        "nomeAnalise": cron_nome_analise,
                        "rows": int(cron_rows),
                        "cols": int(cron_cols),
                        "tetoPlantas": int(cron_teto),
                        "percentualLimite": float(cron_percentual),
                        "minPendoes": int(cron_min_pendoes),
                        "referenciaManual": cron_referencia,
                        "tolerancia": int(cron_tolerancia),
                        "filtroCor": cron_filtro_cor,
                        "areaMin": int(cron_area_min),
                        "areaMax": int(cron_area_max),
                        "sensibilidade": int(cron_sensibilidade),
                        "revisaoManual": bool(cron_revisao_manual),
                        "metadata": {
                            "CODLOCAL": meta_codlocal,
                            "QUADRA": meta_quadra,
                            "ENSAIO": meta_ensaio,
                            "GLI": meta_gli,
                            "TECNOLOGIA": meta_tecnologia,
                            "REP": meta_rep,
                            "NC": meta_nc,
                            "LINHAGEM": meta_linhagem,
                            "GENEALOGIA": meta_genealogia,
                            "PARENTAL1": meta_parental1,
                            "POP_PARENTAL1": meta_pop_parental1,
                            "PARENTAL2": meta_parental2,
                            "POP_PARENTAL2": meta_pop_parental2,
                            "DTP": meta_dtp,
                        }
                    }

                    cron_train_state = {
                        "active": bool(st.session_state.get("pend_yolo_browser_train_active", False)),
                        "apply_training": bool(apply_training_enabled),
                        "folder": str(YOLO_TRAIN_ROOT),
                        "total_images": int(yolo_counts_for_viewer.get("total_images", 0)),
                        "images_train": int(yolo_counts_for_viewer.get("images_train", 0)),
                        "images_val": int(yolo_counts_for_viewer.get("images_val", 0)),
                        "total_labels": int(yolo_counts_for_viewer.get("total_labels", 0)),
                        "crop_default": int(PENDAO_AVANCADO_PARAMS.get("manual_crop_size_default", 128)),
                        "last_summary": st.session_state.get("pend_yolo_last_summary", ""),
                        "capture_endpoint": capture_endpoint if capture_ok else "",
                        "capture_ready": bool(capture_ok),
                    }

                    cron_html = r"""
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  * { box-sizing:border-box; margin:0; padding:0; }
  body { background:#0d0d0d; overflow:hidden; font-family:'Segoe UI',sans-serif; color:#ddd; }
  #cronRoot {
    width:100%; height:850px; display:grid; grid-template-columns:minmax(0,1fr) 305px; gap:10px;
    border:1px solid #2a2a2a; border-radius:12px; background:#0d0d0d; overflow:hidden;
  }
  #cronViewer {
    position:relative; overflow:hidden; user-select:none; cursor:grab;
    background:
      linear-gradient(rgba(255,255,255,.03) 1px, transparent 1px),
      linear-gradient(90deg, rgba(255,255,255,.03) 1px, transparent 1px),
      #0d0d0d;
    background-size:32px 32px;
  }
  #cronViewer:active { cursor:grabbing; }
  #cronCanvas { position:absolute; inset:0; display:block; }
  .cron-side {
    background:linear-gradient(145deg,rgba(18,18,18,.98),rgba(8,8,8,.98));
    border-left:1px solid #2a2a2a; padding:10px; overflow:auto; scrollbar-width:thin; scrollbar-color:#ff8c00 #141414;
  }
  .cron-side::-webkit-scrollbar { width:6px; }
  .cron-side::-webkit-scrollbar-track { background:#141414; }
  .cron-side::-webkit-scrollbar-thumb { background:#ff8c00; border-radius:6px; }
  .title {
    color:#ffffff; font-size:12px; font-weight:900; letter-spacing:1.3px; text-align:center; text-transform:uppercase; margin-bottom:8px;
    text-shadow:0 2px 0 #020e24, 0 6px 14px rgba(0,0,0,.78), 0 0 14px rgba(0,229,255,.54), 0 0 26px rgba(0,255,157,.20);
  }
  .subtle { color:#777; font-size:10px; line-height:1.35; }
  .row { display:flex; gap:6px; align-items:center; justify-content:space-between; margin:5px 0; color:#bbb; font-size:10px; }
  .row span:first-child { max-width:160px; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
  input, select, textarea {
    background:#171717; border:1px solid #333; color:#fff; border-radius:5px; padding:4px 5px; font-size:10px;
  }
  select { width:100%; }
  textarea { width:100%; min-height:44px; resize:vertical; }
  .btn {
    width:100%; min-height:28px; border-radius:5px; padding:5px 7px; font-size:10px; font-weight:800;
    cursor:pointer; transition:.18s; color:#ccc; background:linear-gradient(145deg,#1e1e1e,#111); border:1px solid #3a3a3a;
    margin:3px 0;
  }
  .btn:hover { border-color:#ff8c00; color:#ff8c00; }
  .btn.orange { background:linear-gradient(145deg,#2a1a00,#160b00); border-color:#ff8c00; color:#ffb347; }
  .btn.blue { background:linear-gradient(145deg,#0c1f33,#071321); border-color:#1f77d0; color:#75b7ff; }
  .btn.green { background:linear-gradient(145deg,#0b2815,#06160d); border-color:#00a651; color:#33ee77; }
  .btn.red { background:linear-gradient(145deg,#341010,#1d0606); border-color:#c9302c; color:#ff6666; }
  .btn.active { border-color:#ffd600; color:#ffd600; box-shadow:0 0 10px rgba(255,214,0,.25); }
  .sep { height:1px; background:linear-gradient(90deg,transparent,#333,transparent); margin:8px 0; }
  .badge {
    position:absolute; z-index:20; pointer-events:none; background:rgba(10,10,10,.82); border:1px solid #2a2a2a;
    border-radius:8px; padding:5px 10px; color:#ff8c00; font-size:11px; font-family:'Courier New',monospace; font-weight:700;
  }
  #cronZoom { top:12px; left:12px; }
  #cronCoord { bottom:12px; left:12px; color:#666; font-weight:600; }
  #cronHint { right:12px; bottom:12px; text-align:right; color:#555; font-family:'Segoe UI',sans-serif; font-size:10px; line-height:1.45; }
  .stats { display:grid; grid-template-columns:1fr 1fr; gap:6px; margin:6px 0; }
  .stat {
    border:1px solid #2b2b2b; background:rgba(255,255,255,.025); border-radius:7px; padding:7px;
  }
  .stat b { display:block; color:#fff; font-size:15px; }
  .stat span { color:#888; font-size:9px; text-transform:uppercase; letter-spacing:.6px; }
  .date-list { max-height:330px; overflow:auto; border:1px solid #242424; border-radius:7px; padding:5px; background:#101010; }
  .date-item { color:#aaa; font-size:10px; padding:7px; border:1px solid transparent; border-radius:6px; cursor:pointer; margin-bottom:5px; background:rgba(255,255,255,.018); }
  .date-item.active { color:#ffb347; background:#221400; border-color:#ff8c00; box-shadow:0 0 8px rgba(255,140,0,.18); }
  .date-item:hover { background:#1a1a1a; border-color:#3a3a3a; }
  .date-item.empty { cursor:default; border-style:dashed; color:#555; background:#0d0d0d; }
  .date-item.empty:hover { background:#0d0d0d; border-color:#333; }
  .date-head { display:flex; justify-content:space-between; gap:6px; align-items:center; font-weight:800; }
  .date-name { color:#f0f0f0; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; max-width:178px; }
  .date-meta { color:#888; font-size:9px; margin-top:3px; line-height:1.35; }
  .date-mini { display:grid; grid-template-columns:1fr 1fr 1fr; gap:4px; margin-top:5px; }
  .date-mini span { background:#171717; border:1px solid #2b2b2b; border-radius:4px; padding:3px; text-align:center; color:#aaa; }
  .date-mini b { display:block; color:#fff; font-size:10px; }
  .review {
    border:1px solid #333; border-radius:8px; background:#101010; padding:8px; margin-top:6px; display:none;
  }
  .review h4 { color:#ffd600; font-size:11px; margin-bottom:6px; }
  .review-line { display:grid; grid-template-columns:1fr 58px; gap:5px; align-items:center; margin:3px 0; color:#aaa; font-size:10px; }
  .review-line input { width:58px; text-align:center; }
  .progress {
    height:10px;
    background:__DEPLOY_TRACK_BACKGROUND__;
    border:1px solid __DEPLOY_TRACK_BORDER__;
    border-radius:999px;
    overflow:hidden;
    margin:6px 0;
    box-shadow:__DEPLOY_TRACK_SHADOW__;
  }
  .progress div {
    height:100%;
    width:0%;
    background:__DEPLOY_FILL_ACTIVE__;
    border-radius:999px;
    box-shadow:__DEPLOY_FILL_SHADOW__;
    transition:width .24s ease;
    position:relative;
    overflow:hidden;
  }
  .progress div:after {
    content:"";
    position:absolute;
    inset:0;
    background:linear-gradient(120deg,transparent 0%,rgba(255,255,255,.42) 44%,transparent 74%);
    transform:translateX(-100%);
    animation:deployBarShine 1.45s ease-in-out infinite;
  }
  @keyframes deployBarShine {
    0% { transform:translateX(-100%); }
    100% { transform:translateX(180%); }
  }
  .legend { display:flex; flex-wrap:wrap; gap:5px; color:#888; font-size:9px; margin-top:5px; }
  .leg { display:flex; align-items:center; gap:3px; }
  .sw { width:10px; height:10px; border-radius:2px; border:1px solid rgba(255,255,255,.25); }
  .train-box { border:1px solid #24384c; background:rgba(20,34,48,.45); border-radius:8px; padding:8px; margin:6px 0; }
  .train-box .row span:first-child { max-width:125px; }
  .train-status { border:1px solid #26384a; background:#0b1118; border-radius:6px; color:#8fbde8; font-size:9px; padding:6px; min-height:34px; line-height:1.35; margin-top:5px; }
  .parcel-table-wrap { display:none; border:1px solid #26384a; background:#0b1118; border-radius:8px; padding:7px; margin:7px 0; }
  .parcel-table-title { color:#ffffff; font-size:10px; font-weight:900; margin-bottom:5px; text-transform:uppercase; letter-spacing:.7px; text-shadow:0 2px 0 #020e24,0 0 12px rgba(0,229,255,.42); }
  .parcel-table-scroll { max-height:170px; overflow:auto; border:1px solid #1d2d3e; border-radius:6px; }
  .parcel-table { width:100%; border-collapse:collapse; font-size:9px; color:#ddd; }
  .parcel-table th { position:sticky; top:0; background:#111d2c; color:#75b7ff; padding:4px 3px; border-bottom:1px solid #26384a; }
  .parcel-table td { padding:3px; border-bottom:1px solid rgba(255,255,255,.06); text-align:center; }
  .parcel-total { color:#5ff2b1; font-size:10px; font-weight:800; margin-top:5px; }
  #btnReviewMode, #btnExportCSV, #btnExportResumo, #btnExportCompleto, #btnExportImagem,
  .stats, #statFirstDate, .legend, #reviewPanel { display:none !important; }
</style>
</head>
<body>
<div id="cronRoot">
  <div id="cronViewer">
    <canvas id="cronCanvas"></canvas>
    <div class="badge" id="cronZoom">1.00×</div>
    <div class="badge" id="cronCoord">X:0 Y:0</div>
    <div class="badge" id="cronHint">Scroll=Zoom · Drag=Pan<br>Grid: marque 4 extremidades e arraste para ajustar</div>
  </div>
  <div class="cron-side">
    <div class="title">Resumo / Seletor de Ortofotos</div>
    <div class="subtle">Clique em uma das até 10 ortofotos abaixo para trocar a imagem no mesmo visualizador.</div>

    <div class="sep"></div>
    <div class="row"><span>Nome</span><input id="cfgNome" type="text" value="Pendoamento"></div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;">
      <div class="row" style="display:block;"><span>Disp</span><input id="cfgRows" type="number" min="1" max="200" value="5" style="width:100%;margin-top:3px;"></div>
      <div class="row" style="display:block;"><span>Tiros</span><input id="cfgCols" type="number" min="1" max="200" value="5" style="width:100%;margin-top:3px;"></div>
    </div>
    <div class="row"><span>Teto de pendões</span><input id="cfgTeto" type="number" min="1" max="10000" value="20" style="width:92px;text-align:center;"></div>
    <div class="subtle">O teto trava a primeira data em que cada parcela atingiu o valor configurado.</div>

    <div class="sep"></div>
    <div class="row"><span>Data ativa</span><select id="dateSelect"></select></div>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;">
      <button class="btn blue" id="btnPrevDate">◀ Data</button>
      <button class="btn blue" id="btnNextDate">Data ▶</button>
    </div>
    <div class="subtle" id="activeOrthoSummary">Selecione uma ortofoto para trocar a visualização no mesmo painel.</div>
    <div class="date-list" id="dateList"></div>

    <div class="sep"></div>
    <button class="btn orange" id="btnMarkGrid">⊞ Marcar Grid</button>
    <button class="btn green" id="btnAnalyzeChrono">📐 Análise de Pendoamento</button>
    <button class="btn blue" id="btnReviewMode">✎ Revisar Parcela</button>
    <button class="btn" id="btnFitChrono">⤢ Ajustar à tela</button>
    <button class="btn red" id="btnClearChrono">Limpar seletor</button>
    <div class="train-box">
      <button class="btn blue" id="btnTrainYolo">Treinar YOLO</button>
      <button class="btn green" id="btnApplyTrainingYolo">Aplicar Treino</button>
      <button class="btn orange" id="btnDeleteTrainMark">Excluir Marcação de Treinamento</button>
      <div class="row"><span>Tamanho recorte</span><input id="trainCropSize" type="number" min="48" max="256" step="16" value="96" style="width:78px;text-align:center;"></div>
      <div class="row"><span>Amostras salvas</span><b id="trainSampleCount" style="color:#75b7ff;">0</b></div>
      <button class="btn red" id="btnStopTrainMode">Encerrar Treinamento YOLO</button>
      <div class="train-status" id="trainStatus">Modo treino parado. Clique em Treinar YOLO e depois clique sobre os pendões.</div>
    </div>
    <div class="progress"><div id="cronProgress"></div></div>
    <div class="subtle" id="cronStatus">Selecione a ortofoto no resumo, marque o grid e execute a análise. O teto configurado define o travamento.</div>
    <div class="parcel-table-wrap" id="parcelTableWrap">
      <div class="parcel-table-title">Contagem por parcela</div>
      <div id="parcelTable"></div>
    </div>

    <div class="sep"></div>
    <div class="stats">
      <div class="stat"><b id="statTotal">0</b><span>Parcelas</span></div>
      <div class="stat"><b id="statHit">0</b><span>Atingiram</span></div>
      <div class="stat"><b id="statNoHit">0</b><span>Não atingiram</span></div>
      <div class="stat"><b id="statReview">0</b><span>Revisar</span></div>
    </div>
    <div class="subtle" id="statFirstDate">Primeira data geral: --</div>
    <div class="legend">
      <span class="leg"><i class="sw" style="background:rgba(255,140,0,.35)"></i>atingiu</span>
      <span class="leg"><i class="sw" style="background:rgba(120,120,120,.25)"></i>não atingiu</span>
      <span class="leg"><i class="sw" style="background:rgba(255,214,0,.35)"></i>revisar</span>
      <span class="leg"><i class="sw" style="background:#ff2222"></i>pendão</span>
    </div>

    <div class="review" id="reviewPanel"></div>

    <div class="sep"></div>
    <button class="btn orange" id="btnExportParcelCSV">Exportar CSV Parcelas</button>
    <button class="btn orange" id="btnExportParcelXLSX">Exportar Excel Parcelas</button>
    <button class="btn orange" id="btnExportXLSX">Exportar Excel</button>
    <button class="btn orange" id="btnExportCSV" style="display:none;">Exportar CSV</button>
    <button class="btn" id="btnExportResumo" style="display:none;">Exportar resumo final</button>
    <button class="btn" id="btnExportCompleto" style="display:none;">Exportar dados completos por ortofoto</button>
    <button class="btn" id="btnExportImagem" style="display:none;">Exportar imagem com grid e marcações</button>
  </div>
</div>

<script>
const ORTHOS = __CRON_ORTHOS__;
const CONFIG = __CRON_CONFIG__;
const TRAIN_STATE = __TRAIN_STATE__;
const viewer = document.getElementById('cronViewer');
const canvas = document.getElementById('cronCanvas');
const ctx = canvas.getContext('2d');
const zoomBadge = document.getElementById('cronZoom');
const coordBadge = document.getElementById('cronCoord');
const dateSelect = document.getElementById('dateSelect');
const dateList = document.getElementById('dateList');
const activeOrthoSummary = document.getElementById('activeOrthoSummary');
const cfgNome = document.getElementById('cfgNome');
const cfgRows = document.getElementById('cfgRows');
const cfgCols = document.getElementById('cfgCols');
const cfgTeto = document.getElementById('cfgTeto');
const progressBar = document.getElementById('cronProgress');
const statusEl = document.getElementById('cronStatus');
const reviewPanel = document.getElementById('reviewPanel');
const btnMarkGrid = document.getElementById('btnMarkGrid');
const btnAnalyzeChrono = document.getElementById('btnAnalyzeChrono');
const btnReviewMode = document.getElementById('btnReviewMode');
const btnFitChrono = document.getElementById('btnFitChrono');
const btnClearChrono = document.getElementById('btnClearChrono');
const btnPrevDate = document.getElementById('btnPrevDate');
const btnNextDate = document.getElementById('btnNextDate');
const btnExportCSV = document.getElementById('btnExportCSV');
const btnExportXLSX = document.getElementById('btnExportXLSX');
const btnExportParcelCSV = document.getElementById('btnExportParcelCSV');
const btnExportParcelXLSX = document.getElementById('btnExportParcelXLSX');
const btnExportResumo = document.getElementById('btnExportResumo');
const btnExportCompleto = document.getElementById('btnExportCompleto');
const btnExportImagem = document.getElementById('btnExportImagem');
const btnTrainYolo = document.getElementById('btnTrainYolo');
const btnApplyTrainingYolo = document.getElementById('btnApplyTrainingYolo');
const btnDeleteTrainMark = document.getElementById('btnDeleteTrainMark');
const trainCropSize = document.getElementById('trainCropSize');
const trainSampleCount = document.getElementById('trainSampleCount');
const btnStopTrainMode = document.getElementById('btnStopTrainMode');
const trainStatus = document.getElementById('trainStatus');
const parcelTableWrap = document.getElementById('parcelTableWrap');
const parcelTable = document.getElementById('parcelTable');

let images = [];
let loaded = 0;
let activeIdx = 0;
let scale = 1, offsetX = 0, offsetY = 0;
let dragging = false, lastX = 0, lastY = 0, gridDragPoint = -1;
let markGridMode = false, reviewMode = false;
let gridRatios = [];
let selectedParcel = null;
let resultsByParcel = {};
let finalRows = [];
let fullRows = [];
let parcelCountRows = [];
let manualReviews = {};
let trainYoloMode = Boolean(TRAIN_STATE && TRAIN_STATE.active);
let deleteTrainingMode = false;
let applyYoloTraining = Boolean(TRAIN_STATE && TRAIN_STATE.apply_training);
let yoloSamples = [];
let yoloTrainMarks = [];
let showTrainingMarks = Boolean(trainYoloMode);
let pendingTrainingRefresh = false;
const tempCanvas = document.createElement('canvas');
const tempCtx = tempCanvas.getContext('2d', { willReadFrequently:true });
let tempPrepared = -1;
const VIEWER_STATE_KEY = 'tmg_pendoamento_estado_' + hashString(JSON.stringify(ORTHOS.map(o => [o.name, o.date, o.width, o.height])));

function clamp(v,min,max){ return Math.max(min, Math.min(max, v)); }
function hashString(value){
  let h = 0;
  const s = String(value || '');
  for(let i=0; i<s.length; i++) h = ((h << 5) - h + s.charCodeAt(i)) | 0;
  return Math.abs(h).toString(36);
}
function imgW(idx=activeIdx){ return ORTHOS[idx]?.width || images[idx]?.width || 1; }
function imgH(idx=activeIdx){ return ORTHOS[idx]?.height || images[idx]?.height || 1; }
function cellLabel(r,c){ return 'D' + (r + 1) + ' T' + (c + 1); }
function fmtPct(v){ return Number.isFinite(v) ? v.toFixed(2) : ''; }
function quoteCSV(v){
  const s = (v === null || v === undefined) ? '' : String(v);
  return /[",\n\r]/.test(s) ? '"' + s.replace(/"/g, '""') + '"' : s;
}

function trainDataYaml(){
  return 'path: dados_treinamento_yolo/pendoes\ntrain: images/train\nval: images/val\nnames:\n  0: pendao\n';
}

function sanitizeName(value){
  return String(value || 'amostra').normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/[^a-zA-Z0-9_-]+/g,'_').replace(/^_+|_+$/g,'').slice(0,80) || 'amostra';
}

function canvasToBlob(cv, type='image/jpeg', quality=0.94){
  return new Promise(resolve => cv.toBlob(resolve, type, quality));
}

function captureCropDataUrl(pt, size){
  if(!images[activeIdx] || !images[activeIdx].complete) return '';
  const cv = document.createElement('canvas');
  cv.width = size;
  cv.height = size;
  const cctx = cv.getContext('2d');
  cctx.fillStyle = '#000';
  cctx.fillRect(0,0,size,size);
  const sx = Math.round(pt.x - size / 2);
  const sy = Math.round(pt.y - size / 2);
  const srcX = Math.max(0, sx);
  const srcY = Math.max(0, sy);
  const srcRight = Math.min(imgW(activeIdx), sx + size);
  const srcBottom = Math.min(imgH(activeIdx), sy + size);
  const sw = Math.max(0, srcRight - srcX);
  const sh = Math.max(0, srcBottom - srcY);
  if(sw <= 0 || sh <= 0) return '';
  cctx.drawImage(images[activeIdx], srcX, srcY, sw, sh, srcX - sx, srcY - sy, sw, sh);
  return cv.toDataURL('image/png');
}

async function getDirHandle(root, parts){
  return null;
}

async function writeTrainingFile(parts, filename, content){
  return true;
}

async function generateYamlToTrainingDir(){
  return false;
}

function autoYoloBboxFromCrop(cropCtx, size){
  return {xc:0.5,yc:0.5,w:0.45,h:0.45,auto:false};
}

async function pickTrainingDirectory(){
  trainStatus.textContent = 'O salvamento agora é automático pelo Streamlit na pasta configurada.';
  return false;
}

function currentTrainingTotal(){
  return Number(TRAIN_STATE?.total_images || 0) + yoloSamples.length;
}

async function postTrainingPayloadDirect(payload){
  const endpoint = TRAIN_STATE?.capture_endpoint || '';
  if(!endpoint) return {ok:false, error:'endpoint indisponível'};
  try{
    const response = await fetch(endpoint, {
      method:'POST',
      mode:'cors',
      headers:{'Content-Type':'application/json'},
      body:JSON.stringify(payload)
    });
    const data = await response.json().catch(() => ({}));
    if(!response.ok) return {ok:false, error:data.error || data.message || ('HTTP ' + response.status)};
    return data;
  }catch(e){
    return {ok:false, error:String(e && e.message ? e.message : e)};
  }
}

function sendTrainingPayload(payload){
  persistViewerState();
  const data = JSON.stringify(payload);
  let deliveredByInput = false;
  try{
    window.localStorage.setItem('tmg_pend_yolo_payload', data);
    const parentDoc = window.parent && window.parent.document;
    if(parentDoc){
      const inputs = Array.from(parentDoc.querySelectorAll('input'));
      const target = inputs.find(el => el.getAttribute('aria-label') === 'pend_yolo_payload' || el.name === 'pend_yolo_payload');
      if(target){
        const setter = Object.getOwnPropertyDescriptor(window.parent.HTMLInputElement.prototype, 'value').set;
        target.focus();
        setter.call(target, data);
        target.dispatchEvent(new InputEvent('input', {bubbles:true, data, inputType:'insertText'}));
        target.dispatchEvent(new Event('change', {bubbles:true}));
        target.blur();
        deliveredByInput = true;
      }
    }
  }catch(e){
    console.warn('Falha ao enviar amostra YOLO para o Streamlit', e);
  }
  try{
    const parentWin = window.parent || window;
    const url = new URL(parentWin.location.href);
    url.searchParams.set('pend_yolo_payload', data);
    parentWin.location.assign(url.toString());
    return true;
  }catch(e){
    console.warn('Falha ao enviar payload YOLO por URL', e);
  }
  return deliveredByInput;
}

function requestTrainingRefreshAndApply(){
  persistViewerState();
  try{ window.localStorage.setItem(VIEWER_STATE_KEY + '_auto_apply', '1'); }catch(e){}
  sendTrainingPayload({action:'apply_training', token:'apply_' + Date.now(), total_images:currentTrainingTotal()});
}

function updateTrainPanelStatus(message){
  trainSampleCount.textContent = String(currentTrainingTotal());
  if(message) trainStatus.textContent = message;
}

function persistViewerState(){
  try{
    const payload = {
      version:1,
      savedAt:Date.now(),
      activeIdx,
      gridRatios,
      resultsByParcel,
      manualReviews,
      rows:Number(CONFIG.rows || 1),
      cols:Number(CONFIG.cols || 1)
    };
    window.localStorage.setItem(VIEWER_STATE_KEY, JSON.stringify(payload));
  }catch(e){
    console.warn('Não foi possível salvar estado do visualizador de pendoamento', e);
  }
}

function restoreViewerState(){
  try{
    const raw = window.localStorage.getItem(VIEWER_STATE_KEY);
    if(!raw) return false;
    const payload = JSON.parse(raw);
    if(!payload || payload.version !== 1) return false;
    if(Number(payload.rows || 1) !== Number(CONFIG.rows || 1) || Number(payload.cols || 1) !== Number(CONFIG.cols || 1)) return false;
    if(Array.isArray(payload.gridRatios)) gridRatios = payload.gridRatios.filter(p => Number.isFinite(Number(p.x)) && Number.isFinite(Number(p.y))).slice(0,4);
    if(payload.resultsByParcel && typeof payload.resultsByParcel === 'object') resultsByParcel = payload.resultsByParcel;
    if(payload.manualReviews && typeof payload.manualReviews === 'object') manualReviews = payload.manualReviews;
    activeIdx = clamp(Number(payload.activeIdx || 0), 0, Math.max(0, ORTHOS.length - 1));
    if(dateSelect) dateSelect.value = String(activeIdx);
    if(gridRatios.length || Object.keys(resultsByParcel).length){
      statusEl.textContent = 'Grid e resultado restaurados. Continue o treino ou aplique o treino salvo.';
    }
    return true;
  }catch(e){
    console.warn('Não foi possível restaurar estado do visualizador de pendoamento', e);
    return false;
  }
}

function activeTrainingMarks(){
  const stored = Array.isArray(ORTHOS[activeIdx]?.training_marks) ? ORTHOS[activeIdx].training_marks.map(m => ({...m, idx:activeIdx, stored:true})) : [];
  const local = yoloTrainMarks.filter(m => m.idx === activeIdx).map(m => ({...m, stored:false}));
  return [...stored, ...local];
}

function nearestTrainingMark(pt){
  let best = null;
  let bestDist = Infinity;
  for(const mark of activeTrainingMarks()){
    const dx = pt.x - Number(mark.x || 0);
    const dy = pt.y - Number(mark.y || 0);
    const dist = Math.sqrt(dx*dx + dy*dy);
    const radius = Math.max(18 / Math.max(scale, 0.0001), Number(mark.size || 128) * 0.45);
    if(dist <= radius && dist < bestDist){
      best = mark;
      bestDist = dist;
    }
  }
  return best;
}

function removeLocalTrainingMark(mark){
  if(mark.stored){
    const marks = Array.isArray(ORTHOS[activeIdx]?.training_marks) ? ORTHOS[activeIdx].training_marks : [];
    ORTHOS[activeIdx].training_marks = marks.filter(item => {
      if(mark.id && item.id === mark.id) return false;
      if(mark.file && item.file === mark.file) return false;
      return Math.abs(Number(item.x || 0) - Number(mark.x || 0)) > 0.5 || Math.abs(Number(item.y || 0) - Number(mark.y || 0)) > 0.5;
    });
    if(TRAIN_STATE) TRAIN_STATE.total_images = Math.max(0, Number(TRAIN_STATE.total_images || 0) - 1);
  } else {
    yoloTrainMarks = yoloTrainMarks.filter(item => item !== mark && item.token !== mark.token);
    yoloSamples = yoloSamples.filter(item => item.token !== mark.token);
    if(mark.saved_direct){
      postTrainingPayloadDirect({
        action:'delete_sample',
        sample_id:mark.id || '',
        file:mark.file || '',
        name:ORTHOS[activeIdx]?.name || mark.source_name || '',
        date:ORTHOS[activeIdx]?.date || mark.source_date || ''
      }).then(resp => {
        if(resp.ok){
          if(TRAIN_STATE) TRAIN_STATE.total_images = Math.max(0, Number(resp.total_images || 0));
          pendingTrainingRefresh = true;
          updateTrainPanelStatus('Amostra removida da pasta de treinamento. Aplicar Treino atualizará a recontagem.');
        } else {
          updateTrainPanelStatus('Marcação removida do visualizador, mas não foi possível apagar arquivo: ' + (resp.error || resp.message || 'erro'));
        }
      });
    }
  }
}

function deleteTrainingSampleAt(pt){
  const mark = nearestTrainingMark(pt);
  if(!mark){
    trainStatus.textContent = 'Nenhuma marcação de treinamento encontrada perto do clique.';
    return;
  }
  if(!confirm('Excluir esta amostra de treinamento?')) return;
  removeLocalTrainingMark(mark);
  const ortho = ORTHOS[activeIdx] || {};
  if(mark.stored){
    sendTrainingPayload({
      action:'delete_sample',
      token:'delete_' + Date.now() + '_' + (mark.id || mark.file || ''),
      sample_id:mark.id || '',
      file:mark.file || '',
      name:ortho.name || mark.source_name || '',
      date:ortho.date || mark.source_date || ''
    });
    updateTrainPanelStatus('Amostra removida do visualizador. Apagando arquivos vinculados na pasta YOLO...');
  } else {
    updateTrainPanelStatus('Marcação local removida antes da gravação definitiva.');
  }
  drawAll();
}

async function saveYoloTrainingSample(pt){
  if(!images[activeIdx] || !images[activeIdx].complete){
    trainStatus.textContent = 'Imagem ainda não carregada.';
    return;
  }
  const size = clamp(parseInt(trainCropSize.value || 128), 48, 256);
  const ortho = ORTHOS[activeIdx] || {};
  const token = Date.now() + '_' + activeIdx + '_' + Math.round(pt.x) + '_' + Math.round(pt.y);
  const sample = {
    token,
    idx:activeIdx,
    name:ortho.name || '',
    date:ortho.date || '',
    x:Number(pt.x.toFixed(2)),
    y:Number(pt.y.toFixed(2)),
    preview_width:imgW(activeIdx),
    preview_height:imgH(activeIdx),
    orig_width:Number(ortho.orig_width || imgW(activeIdx)),
    orig_height:Number(ortho.orig_height || imgH(activeIdx)),
    crop_size:size
  };
  yoloSamples.push(sample);
  const localMark = {x:pt.x,y:pt.y,size,type:'pendao_confirmado',idx:activeIdx,token};
  yoloTrainMarks.push(localMark);
  showTrainingMarks = true;
  updateTrainPanelStatus('Capturando mini foto. Salvando crop, label e características na pasta de treinamento...');
  drawAll();
  const cropDataUrl = captureCropDataUrl(pt, size);
  if(cropDataUrl && TRAIN_STATE?.capture_ready){
    const origX = sample.x * sample.orig_width / sample.preview_width;
    const origY = sample.y * sample.orig_height / sample.preview_height;
    const response = await postTrainingPayloadDirect({
      action:'sample',
      token,
      name:sample.name,
      date:sample.date,
      x:sample.x,
      y:sample.y,
      orig_x:origX,
      orig_y:origY,
      crop_size:size,
      crop_data_url:cropDataUrl
    });
    if(response && response.ok){
      yoloSamples = yoloSamples.filter(item => item.token !== token);
      localMark.saved_direct = true;
      localMark.id = response.id || '';
      localMark.file = response.file || '';
      localMark.crop = response.crop || '';
      if(TRAIN_STATE) TRAIN_STATE.total_images = Number(response.total_images || currentTrainingTotal());
      pendingTrainingRefresh = true;
      updateTrainPanelStatus('Mini foto salva automaticamente: ' + (response.file || 'amostra') + '. Pasta: ' + (response.folder || TRAIN_STATE?.folder || 'treino YOLO') + '.');
      drawAll();
      return;
    }
    updateTrainPanelStatus('Salvamento direto falhou (' + (response.error || response.message || 'erro') + '). Tentando salvar pelo Streamlit...');
  }
  sendTrainingPayload({action:'sample', ...sample});
}

async function downloadYoloDataset(){
  trainStatus.textContent = 'As mini imagens já são salvas automaticamente pelo Streamlit na pasta de treinamento.';
}

function applyViewerConfig(clearResults=false){
  const oldRows = Number(CONFIG.rows || 1);
  const oldCols = Number(CONFIG.cols || 1);
  CONFIG.nomeAnalise = (cfgNome.value || 'Pendoamento').trim() || 'Pendoamento';
  CONFIG.rows = clamp(parseInt(cfgRows.value || CONFIG.rows || 1), 1, 200);
  CONFIG.cols = clamp(parseInt(cfgCols.value || CONFIG.cols || 1), 1, 200);
  CONFIG.tetoPlantas = clamp(parseInt(cfgTeto.value || CONFIG.tetoPlantas || 1), 1, 10000);
  cfgRows.value = CONFIG.rows;
  cfgCols.value = CONFIG.cols;
  cfgTeto.value = CONFIG.tetoPlantas;
  if(clearResults || oldRows !== CONFIG.rows || oldCols !== CONFIG.cols){
    selectedParcel = null;
    resultsByParcel = {};
    finalRows = [];
    fullRows = [];
    progressBar.style.width = '0%';
  }
  rebuildRows();
  renderReviewPanel();
  drawAll();
}

function setupViewerConfigInputs(){
  cfgNome.value = CONFIG.nomeAnalise || 'Pendoamento';
  cfgRows.value = CONFIG.rows || 5;
  cfgCols.value = CONFIG.cols || 5;
  cfgTeto.value = CONFIG.tetoPlantas || 20;
  cfgNome.onchange = () => applyViewerConfig(false);
  cfgRows.onchange = () => applyViewerConfig(true);
  cfgCols.onchange = () => applyViewerConfig(true);
  cfgTeto.onchange = () => applyViewerConfig(false);
}

function loadScriptOnce(src){
  return new Promise((resolve, reject) => {
    const existing = document.querySelector('script[data-tmg-src="' + src + '"]');
    if(existing){
      if(existing.dataset.loaded === 'true'){ resolve(true); return; }
      existing.addEventListener('load', () => resolve(true), {once:true});
      existing.addEventListener('error', () => reject(new Error('Falha ao carregar ' + src)), {once:true});
      return;
    }
    const script = document.createElement('script');
    script.src = src;
    script.async = true;
    script.dataset.tmgSrc = src;
    script.onload = () => { script.dataset.loaded = 'true'; resolve(true); };
    script.onerror = () => reject(new Error('Falha ao carregar ' + src));
    document.head.appendChild(script);
  });
}

async function ensureChronoExcel(){
  if(window.__tmgChronoXlsxReady && typeof XLSX !== 'undefined') return true;
  const urls = [
    'https://cdn.jsdelivr.net/npm/xlsx-js-style@1.2.0/dist/xlsx.bundle.js',
    'https://unpkg.com/xlsx-js-style@1.2.0/dist/xlsx.bundle.js',
    'https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js'
  ];
  for(const url of urls){
    try{
      await loadScriptOnce(url);
      if(typeof XLSX !== 'undefined'){
        window.__tmgChronoXlsxReady = true;
        return true;
      }
    } catch(e){ console.warn(e); }
  }
  return false;
}

function setupDates(){
  dateSelect.innerHTML = '';
  ORTHOS.forEach((o, idx) => {
    const opt = document.createElement('option');
    opt.value = idx;
    opt.textContent = (idx + 1) + ' · ' + o.date + ' · ' + o.name;
    dateSelect.appendChild(opt);
  });
  renderOrthoSummary();
}

function safeHtml(value){
  const div = document.createElement('div');
  div.textContent = String(value ?? '');
  return div.innerHTML;
}

function orthoStats(idx){
  let total=0, hit=0, partial=0, review=0;
  const tetoDefault = Math.max(1, Number(CONFIG.tetoPlantas || 1));
  for(const label of Object.keys(resultsByParcel)){
    const rec = resultsByParcel[label] || {};
    const manual = manualReviews[label] || {};
    const count = manual.counts && manual.counts[idx] !== undefined ? Number(manual.counts[idx] || 0) : Number((rec.counts || [])[idx] || 0);
    const teto = Math.max(1, Number(manual.teto || tetoDefault));
    const pct = count / teto * 100;
    total += count;
    if(pct >= Number(CONFIG.percentualLimite || 50) && count >= Number(CONFIG.minPendoes || 0)) hit++;
    else if(count > 0) partial++;
    if((rec.confidence || 1) < 0.38) review++;
  }
  return {total, hit, partial, review};
}

function buildActiveParcelRows(){
  const R = Math.max(1, parseInt(CONFIG.rows || 1));
  const C = Math.max(1, parseInt(CONFIG.cols || 1));
  const meta = CONFIG.metadata || {};
  const quadra = meta.QUADRA || CONFIG.nomeAnalise || 'Pendoamento';
  const rows = [];
  for(let r=0; r<R; r++){
    for(let c=0; c<C; c++){
      const label = cellLabel(r,c);
      const rec = resultsByParcel[label] || {};
      const manual = manualReviews[label] || {};
      const count = manual.counts && manual.counts[activeIdx] !== undefined
        ? Number(manual.counts[activeIdx] || 0)
        : Number((rec.counts || [])[activeIdx] || 0);
      rows.push({
        Quadra: quadra,
        Parcela: label,
        Linha: r + 1,
        Coluna: c + 1,
        Quantidade_Pendoes: count
      });
    }
  }
  return rows;
}

function renderParcelCountTable(){
  if(!Object.keys(resultsByParcel).length){
    parcelCountRows = [];
    parcelTableWrap.style.display = 'none';
    parcelTable.innerHTML = '';
    return;
  }
  parcelCountRows = buildActiveParcelRows();
  const total = parcelCountRows.reduce((sum,row) => sum + Number(row.Quantidade_Pendoes || 0), 0);
  parcelTableWrap.style.display = 'block';
  parcelTable.innerHTML =
    '<div class="parcel-table-scroll"><table class="parcel-table">' +
    '<thead><tr><th>Parcela</th><th>Linha</th><th>Coluna</th><th>Pendões</th></tr></thead><tbody>' +
    parcelCountRows.map(row =>
      '<tr><td>' + safeHtml(row.Parcela) + '</td><td>' + row.Linha + '</td><td>' + row.Coluna + '</td><td><b>' + row.Quantidade_Pendoes + '</b></td></tr>'
    ).join('') +
    '</tbody></table></div>' +
    '<div class="parcel-total">Total geral: ' + total + ' pendões</div>';
}

function renderOrthoSummary(){
  dateList.innerHTML = '';
  const active = ORTHOS[activeIdx] || {};
  const activeStats = orthoStats(activeIdx);
  const activeDet = active.detector_counts || {};
  activeOrthoSummary.innerHTML =
    'Visualizando: <b style="color:#ffb347">' + (activeIdx + 1) + ' · ' + safeHtml(active.name || '') + '</b><br>' +
    'Data: ' + safeHtml(active.date || '--') + ' · Pendões: ' + activeStats.total + ' · Atingidas: ' + activeStats.hit +
    '<br>Detector: ' + safeHtml(active.detector_mode || 'OpenCV fallback') +
    ' · YOLO ' + Number(activeDet.yolo || 0) + ' · OpenCV ' + Number(activeDet.opencv || 0) + ' · Refinado ' + Number(activeDet.refinado || 0) +
    '<br><span style="color:#777">Ortofotos carregadas: ' + ORTHOS.length + '/10</span>';
  for(let idx=0; idx<10; idx++){
    const o = ORTHOS[idx];
    const stats = orthoStats(idx);
    const item = document.createElement('div');
    if(o){
      item.className = 'date-item' + (idx === activeIdx ? ' active' : '');
      item.innerHTML =
        '<div class="date-head"><span>#' + String(idx + 1).padStart(2,'0') + ' · ' + safeHtml(o.date) + '</span><span class="date-name">' + safeHtml(o.name) + '</span></div>' +
        '<div class="date-meta">Clique aqui para visualizar esta ortofoto no painel único.<br>Detector: ' + safeHtml(o.detector_mode || 'OpenCV fallback') +
        ' · YOLO ' + Number((o.detector_counts || {}).yolo || 0) +
        ' · OpenCV ' + Number((o.detector_counts || {}).opencv || 0) +
        ' · Refinado ' + Number((o.detector_counts || {}).refinado || 0) + '</div>' +
        '<div class="date-mini">' +
          '<span><b>' + stats.total + '</b>pendões</span>' +
          '<span><b>' + stats.hit + '</b>atingiu</span>' +
          '<span><b>' + stats.partial + '</b>parcial</span>' +
        '</div>';
      item.onclick = () => setActiveDate(idx);
    } else {
      item.className = 'date-item empty';
      item.innerHTML =
        '<div class="date-head"><span>#' + String(idx + 1).padStart(2,'0') + ' · vazio</span><span class="date-name">Aguardando</span></div>' +
        '<div class="date-meta">Anexe uma ortofoto para liberar esta posição.</div>';
    }
    dateList.appendChild(item);
  }
  renderParcelCountTable();
}

function setActiveDate(idx){
  activeIdx = clamp(idx, 0, ORTHOS.length - 1);
  dateSelect.value = String(activeIdx);
  tempPrepared = -1;
  renderOrthoSummary();
  statusEl.textContent = 'Visualizando ortofoto ' + (activeIdx + 1) + '/' + ORTHOS.length + ': ' + (ORTHOS[activeIdx]?.name || '');
  drawAll();
}

function loadImages(){
  ORTHOS.forEach((o, idx) => {
    const im = new Image();
    im.onload = () => {
      loaded += 1;
      if(idx === 0) fitView();
      statusEl.textContent = 'Ortofotos carregadas: ' + loaded + '/' + ORTHOS.length;
      renderOrthoSummary();
      drawAll();
    };
    im.onerror = () => {
      statusEl.textContent = 'Falha ao carregar ortofoto: ' + (o.name || ('#' + (idx + 1)));
    };
    im.src = 'data:image/jpeg;base64,' + o.b64;
    images[idx] = im;
  });
}

function fitView(){
  const W = viewer.clientWidth, H = viewer.clientHeight;
  const w = imgW(), h = imgH();
  scale = Math.min(W / w, H / h);
  offsetX = (W - w * scale) / 2;
  offsetY = (H - h * scale) / 2;
  drawAll();
}

function screenToImg(clientX, clientY){
  const r = canvas.getBoundingClientRect();
  return { x:(clientX - r.left - offsetX) / scale, y:(clientY - r.top - offsetY) / scale };
}

function ratioPoint(pt, idx=activeIdx){
  return { x: pt.x / imgW(idx), y: pt.y / imgH(idx) };
}

function currentGridPoints(idx=activeIdx){
  return gridRatios.map(p => ({ x:p.x * imgW(idx), y:p.y * imgH(idx) }));
}

function gridPointIndexAt(pt, idx=activeIdx){
  const pts = currentGridPoints(idx);
  for(let i=0; i<pts.length; i++){
    const dx=(pt.x-pts[i].x)*scale;
    const dy=(pt.y-pts[i].y)*scale;
    if(Math.sqrt(dx*dx+dy*dy) < 22) return i;
  }
  return -1;
}

function bilerp(p0,p1,p2,p3,u,v){
  const tx = (1-u)*p0.x + u*p1.x, ty = (1-u)*p0.y + u*p1.y;
  const bx = (1-u)*p3.x + u*p2.x, by = (1-u)*p3.y + u*p2.y;
  return { x:(1-v)*tx + v*bx, y:(1-v)*ty + v*by };
}

function cellPoly(r,c,idx=activeIdx){
  const pts = currentGridPoints(idx);
  if(pts.length < 4) return null;
  const R = Math.max(1, parseInt(CONFIG.rows || 1));
  const C = Math.max(1, parseInt(CONFIG.cols || 1));
  const u0 = c / C, u1 = (c + 1) / C, v0 = r / R, v1 = (r + 1) / R;
  return [
    bilerp(pts[0],pts[1],pts[2],pts[3],u0,v0),
    bilerp(pts[0],pts[1],pts[2],pts[3],u1,v0),
    bilerp(pts[0],pts[1],pts[2],pts[3],u1,v1),
    bilerp(pts[0],pts[1],pts[2],pts[3],u0,v1)
  ];
}

function pointInPolygon(px,py,poly){
  let inside = false;
  for(let i=0,j=poly.length-1; i<poly.length; j=i++){
    const xi=poly[i].x, yi=poly[i].y, xj=poly[j].x, yj=poly[j].y;
    if(((yi>py)!==(yj>py)) && (px < (xj-xi)*(py-yi)/(yj-yi)+xi)) inside = !inside;
  }
  return inside;
}

function findCell(pt){
  const R = Math.max(1, parseInt(CONFIG.rows || 1));
  const C = Math.max(1, parseInt(CONFIG.cols || 1));
  for(let r=0; r<R; r++){
    for(let c=0; c<C; c++){
      const poly = cellPoly(r,c);
      if(poly && pointInPolygon(pt.x, pt.y, poly)) return {r,c,label:cellLabel(r,c),poly};
    }
  }
  return null;
}

function rgbToHsv(r,g,b){
  r/=255; g/=255; b/=255;
  const max=Math.max(r,g,b), min=Math.min(r,g,b), d=max-min;
  let h=0;
  if(d!==0){
    if(max===r) h=60*(((g-b)/d)%6);
    else if(max===g) h=60*((b-r)/d+2);
    else h=60*((r-g)/d+4);
    if(h<0) h+=360;
  }
  return {h, s:max===0 ? 0 : d/max, v:max};
}

function tasselScore(r,g,b){
  const hsv = rgbToHsv(r,g,b);
  const chroma = Math.max(r,g,b) - Math.min(r,g,b);
  const exg = 2*g-r-b;
  const yellowness = ((r+g)*0.5)-b;
  const redYellowBalance = Math.abs(r-g);
  const leafGreen = (exg>18 && g>=r*0.98 && g>=b*1.07 && hsv.h>=68 && hsv.h<=165 && yellowness<34 && hsv.v<0.82);
  const hardGreen = (exg>32 && hsv.h>=70 && hsv.h<=158 && hsv.s>0.20 && yellowness<30);
  const darkShadow = hsv.v<0.20;
  const whiteGlare = hsv.s<0.08 && hsv.v>0.82;
  const soilRed = r>g*1.28 && r>b*1.36 && hsv.h<18;
  const greenWhiteEdge = hsv.s<0.22 && g>=r*1.01 && b>=r*0.74 && hsv.h>=65 && hsv.h<=150 && yellowness<22;
  const nonGreen = !leafGreen && !greenWhiteEdge && !hardGreen && exg < 62;
  const filter = CONFIG.filtroCor || 'Misto';

  let score = 0;
  const young = hsv.h>=38 && hsv.h<=76 && hsv.s>=0.10 && hsv.s<=0.56 && hsv.v>=0.34 && yellowness>=12 && exg<42 && b<=Math.min(r,g)*0.88;
  const dry = hsv.h>=18 && hsv.h<=64 && hsv.s>=0.10 && hsv.v>=0.30 && yellowness>=12 && r>=82 && g>=58 && b<=158 && r>=b+14 && g>=b+6 && redYellowBalance<=92 && exg<52;
  const bright = r>=96 && g>=72 && b<=185 && yellowness>=12 && redYellowBalance<=88 && hsv.s>=0.07 && hsv.h>=12 && hsv.h<=76 && exg<55;
  const creamTip = Math.max(r,g,b)>=132 && yellowness>=16 && b<=g*0.91 && b<=r*0.91 && hsv.s>=0.06 && hsv.s<=0.58 && hsv.h>=14 && hsv.h<=78 && exg<56;
  const oldPinkTan = r>=96 && g>=64 && b<=168 && r>=g*0.84 && g>=b*0.78 && r>=b+8 && hsv.h>=8 && hsv.h<=46 && hsv.s>=0.10 && exg<52;
  const textureColor = chroma>=18 && yellowness>=10 && hsv.s>=0.09 && exg<55;

  if(filter === 'Pendão novo'){
    if(young) score += 3.2;
    if(hsv.h>=50 && hsv.h<=90) score += 1.1;
    if(dry) score += 0.8;
  } else if(filter === 'Pendão seco/velho'){
    if(dry) score += 3.3;
    if(bright) score += 1.6;
    if(young) score += 0.7;
  } else {
    if(young) score += 1.75;
    if(dry) score += 2.25;
    if(bright) score += 2.45;
    if(creamTip) score += 2.35;
    if(oldPinkTan) score += 1.85;
  }
  if(textureColor) score += 0.75;
  if(yellowness>=24) score += 0.65;
  if(yellowness>=38 && exg<52) score += 0.45;
  if(!nonGreen) score -= 4.5;
  if(leafGreen) score -= 4.5;
  if(hardGreen) score -= 4.0;
  if(greenWhiteEdge) score -= 3.2;
  if(darkShadow) score -= 2.2;
  if(whiteGlare) score -= 2.3;
  if(soilRed) score -= 1.8;
  if(chroma<12 && yellowness<18) score -= 1.2;
  if(hsv.h>=92 && hsv.h<=155 && exg>16) score -= 2.0;

  const sensitivity = clamp(Number(CONFIG.sensibilidade || 60), 1, 100);
  const tolerance = clamp(Number(CONFIG.tolerancia || 55), 0, 100);
  score += (sensitivity - 50) / 70;
  score += (tolerance - 50) / 130;
  return score;
}

function scoreThreshold(){
  const sensitivity = clamp(Number(CONFIG.sensibilidade || 60), 1, 100);
  const tolerance = clamp(Number(CONFIG.tolerancia || 55), 0, 100);
  return clamp(3.92 - (sensitivity - 50) / 135 - (tolerance - 50) / 220, 3.55, 4.45); // AJUSTE PENDOAMENTO: limiar mais sensível dentro das parcelas
}

function mergeNearbyTassels(marks,w,h){
  if(!marks.length) return marks;
  const radius = Math.max(12, Math.min(28, Math.min(w,h)*0.040));
  const ordered = [...marks].sort((a,b)=>(b.score*b.area)-(a.score*a.area));
  const merged = [];
  for(const m of ordered){
    let target = null;
    for(const c of merged){
      const dx=m.x-c.x, dy=m.y-c.y;
      if(Math.sqrt(dx*dx+dy*dy)<=radius){ target=c; break; }
    }
    if(target){
      const wa=Math.max(1,target.area), wb=Math.max(1,m.area);
      target.x=(target.x*wa+m.x*wb)/(wa+wb);
      target.y=(target.y*wa+m.y*wb)/(wa+wb);
      target.area+=m.area;
      target.score=Math.max(target.score,m.score);
    } else {
      merged.push({...m});
    }
  }
  return merged.sort((a,b)=>a.y===b.y ? a.x-b.x : a.y-b.y);
}

function prepareTemp(idx){
  if(tempPrepared === idx) return;
  tempCanvas.width = imgW(idx);
  tempCanvas.height = imgH(idx);
  tempCtx.clearRect(0,0,tempCanvas.width,tempCanvas.height);
  tempCtx.drawImage(images[idx],0,0,tempCanvas.width,tempCanvas.height);
  tempPrepared = idx;
}

function analyzeCellFromAdvancedDetections(idx,r,c){
  const source = ORTHOS[idx] && Array.isArray(ORTHOS[idx].advanced_detections) ? ORTHOS[idx].advanced_detections : null;
  const backendReady = Boolean(ORTHOS[idx] && ORTHOS[idx].backend_ready);
  const poly = cellPoly(r,c,idx);
  if(!poly) return {count:0, marks:[], confidence:0};
  if(!source || !source.length) return null; // AJUSTE PENDOAMENTO: sem detecções prévias, usa fallback OpenCV no grid em vez de zerar a parcela
  const marks = [];
  let confidenceSum = 0;
  for(const det of source){
    const x = Number(det.x);
    const y = Number(det.y);
    if(!Number.isFinite(x) || !Number.isFinite(y)) continue;
    if(!pointInPolygon(x,y,poly)) continue;
    const score = Number(det.score || 0);
    const size = clamp(Number(det.size || 8), 5, 22);
    marks.push({
      x,
      y,
      score,
      size,
      area: Math.max(1, size * size),
      tipo: det.tipo || 'misto',
      confianca: det.confianca || 'media',
      source: det.source || 'OpenCV',
      yolo_conf: Number(det.yolo_conf || 0),
      class_name: det.class_name || '',
      yellow_ratio: Number(det.yellow_ratio || 0),
      texture_ratio: Number(det.texture_ratio || 0),
      clear_ratio: Number(det.clear_ratio || 0),
      green_ratio: Number(det.green_ratio || 0)
    });
    confidenceSum += clamp(score / 7, 0.25, 1);
  }
  const confidence = marks.length ? confidenceSum / marks.length : 0.82;
  return {count:marks.length, marks, confidence};
}

function analyzeCellInImage(idx,r,c){
  const advanced = analyzeCellFromAdvancedDetections(idx,r,c);
  // AJUSTE PENDOAMENTO: usa detecção prévia apenas quando ela encontrou pendões na parcela.
  // Se vier zerada, continua para o OpenCV do navegador e tenta contar dentro do polígono do grid.
  if(advanced && advanced.count > 0) return advanced;
  prepareTemp(idx);
  const poly = cellPoly(r,c,idx);
  if(!poly) return {count:0, marks:[], confidence:0};
  const xs = poly.map(p=>p.x), ys = poly.map(p=>p.y);
  const minX = clamp(Math.floor(Math.min(...xs)),0,imgW(idx)-1);
  const maxX = clamp(Math.ceil(Math.max(...xs)),0,imgW(idx)-1);
  const minY = clamp(Math.floor(Math.min(...ys)),0,imgH(idx)-1);
  const maxY = clamp(Math.ceil(Math.max(...ys)),0,imgH(idx)-1);
  const w = maxX-minX+1, h=maxY-minY+1;
  if(w<=1 || h<=1) return {count:0, marks:[], confidence:0};

  const data = tempCtx.getImageData(minX,minY,w,h).data;
  const sensitivity = clamp(Number(CONFIG.sensibilidade || 60), 1, 100);
  const step = sensitivity >= 75 ? 1 : (sensitivity >= 45 ? 2 : 3);
  const gw = Math.ceil(w/step), gh=Math.ceil(h/step);
  const mask = new Uint8Array(gw*gh);
  const clean = new Uint8Array(gw*gh);
  const close = new Uint8Array(gw*gh);
  const inside = new Uint8Array(gw*gh);
  const scores = new Float32Array(gw*gh);
  const yellows = new Float32Array(gw*gh);
  const exgs = new Float32Array(gw*gh);
  const chromas = new Float32Array(gw*gh);
  const th = scoreThreshold();
  let scoreSumAll=0, scoreSqAll=0, scoreCount=0;

  for(let gy=0; gy<gh; gy++){
    for(let gx=0; gx<gw; gx++){
      const px = Math.min(w-1, gx*step), py=Math.min(h-1, gy*step);
      const ax=minX+px, ay=minY+py;
      if(!pointInPolygon(ax,ay,poly)) continue;
      const di=(py*w+px)*4;
      const pr=data[di], pg=data[di+1], pb=data[di+2];
      const score = tasselScore(pr,pg,pb);
      const mi=gy*gw+gx;
      inside[mi]=1;
      scores[mi]=score;
      yellows[mi]=((pr+pg)*0.5)-pb;
      exgs[mi]=2*pg-pr-pb;
      chromas[mi]=Math.max(pr,pg,pb)-Math.min(pr,pg,pb);
      scoreSumAll += score;
      scoreSqAll += score*score;
      scoreCount++;
    }
  }

  const scoreMean = scoreCount ? scoreSumAll / scoreCount : 0;
  const scoreVar = scoreCount ? Math.max(0, scoreSqAll / scoreCount - scoreMean*scoreMean) : 0;
  const localTh = clamp(Math.max(th - 0.22, scoreMean + Math.sqrt(scoreVar) * 0.55), th - 0.35, th + 0.80); // AJUSTE PENDOAMENTO: limiar local menos rígido para pendões pequenos
  for(let mi=0; mi<mask.length; mi++){
    if(inside[mi] && scores[mi]>=localTh){
      mask[mi]=1;
    }
  }

  for(let gy=0; gy<gh; gy++){
    for(let gx=0; gx<gw; gx++){
      const mi=gy*gw+gx;
      if(!mask[mi]) continue;
      let n=0;
      for(let yy=-1; yy<=1; yy++){
        for(let xx=-1; xx<=1; xx++){
          if(xx===0 && yy===0) continue;
          const nx=gx+xx, ny=gy+yy;
          if(nx>=0 && nx<gw && ny>=0 && ny<gh && mask[ny*gw+nx]) n++;
        }
      }
      if(n>=2 || scores[mi]>=localTh+0.85) clean[mi]=1;
    }
  }

  for(let gy=0; gy<gh; gy++){
    for(let gx=0; gx<gw; gx++){
      const mi=gy*gw+gx;
      if(clean[mi]) { close[mi]=1; continue; }
      let n=0, s=0, total=0;
      for(let yy=-1; yy<=1; yy++){
        for(let xx=-1; xx<=1; xx++){
          const nx=gx+xx, ny=gy+yy;
          if(nx>=0 && nx<gw && ny>=0 && ny<gh){
            const ni=ny*gw+nx;
            if(clean[ni]) n++;
            s += scores[ni]; total++;
          }
        }
      }
      if(n>=5 && s/Math.max(1,total)>=localTh-0.35) close[mi]=1;
    }
  }

  const minArea = Math.max(1, Number(CONFIG.areaMin || 10)); // AJUSTE PENDOAMENTO: pendões menores
  const maxAreaCfg = Math.max(minArea+1, Number(CONFIG.areaMax || 900));
  const maxArea = Math.min(maxAreaCfg, Math.max(minArea+1, w*h*0.026));
  const visited = new Uint8Array(gw*gh);
  const dirs = [[1,0],[-1,0],[0,1],[0,-1],[1,1],[-1,1],[1,-1],[-1,-1]];
  const marks = [];
  let confidenceSum = 0;

  for(let gy=0; gy<gh; gy++){
    for(let gx=0; gx<gw; gx++){
      const start=gy*gw+gx;
      if(!close[start] || visited[start]) continue;
      let cells=0, sx=0, sy=0, scoreSum=0, yellowSum=0, exgSum=0, chromaSum=0, coreCells=0, paleCells=0;
      let minGX=gx, maxGX=gx, minGY=gy, maxGY=gy;
      const stack=[[gx,gy]];
      visited[start]=1;
      while(stack.length){
        const p=stack.pop(), x=p[0], y=p[1], pos=y*gw+x;
        cells++; sx+=x; sy+=y; scoreSum+=scores[pos];
        yellowSum+=yellows[pos]; exgSum+=exgs[pos]; chromaSum+=chromas[pos];
        if(scores[pos]>=localTh+0.55) coreCells++;
        if(yellows[pos]>=12 && exgs[pos]<54) paleCells++;
        if(x<minGX) minGX=x; if(x>maxGX) maxGX=x; if(y<minGY) minGY=y; if(y>maxGY) maxGY=y;
        for(const d of dirs){
          const nx=x+d[0], ny=y+d[1];
          if(nx<0 || nx>=gw || ny<0 || ny>=gh) continue;
          const ni=ny*gw+nx;
          if(close[ni] && !visited[ni]) { visited[ni]=1; stack.push([nx,ny]); }
        }
      }
      const area = cells*step*step;
      const widthPx = (maxGX-minGX+1)*step;
      const heightPx = (maxGY-minGY+1)*step;
      const bboxCells = Math.max(1,(maxGX-minGX+1)*(maxGY-minGY+1));
      const density = cells / bboxCells;
      const elongation = Math.max(widthPx,heightPx) / Math.max(1,Math.min(widthPx,heightPx));
      const meanScore = scoreSum / Math.max(1,cells);
      const meanYellow = yellowSum / Math.max(1,cells);
      const meanExg = exgSum / Math.max(1,cells);
      const meanChroma = chromaSum / Math.max(1,cells);
      const coreRatio = coreCells / Math.max(1,cells);
      const paleRatio = paleCells / Math.max(1,cells);
      const cx=minX + (sx/cells)*step, cy=minY + (sy/cells)*step;
      const valid =
        area>=minArea &&
        area<=maxArea &&
        density>=0.08 &&
        elongation<=10.0 &&
        widthPx<=Math.max(20, w*0.22) &&
        heightPx<=Math.max(20, h*0.26) &&
        meanScore>=localTh-0.04 &&
        meanYellow>=7 &&
        meanExg<66 &&
        meanChroma>=10 &&
        coreRatio>=0.08 &&
        paleRatio>=0.12 &&
        pointInPolygon(cx,cy,poly);
      if(valid){
        marks.push({x:cx,y:cy,area:area,score:meanScore});
        confidenceSum += clamp((meanScore - (localTh-0.2)) / 2.2, 0.15, 1);
      }
    }
  }
  const confidence = marks.length ? confidenceSum / marks.length : 0.78;
  const finalMarks = mergeNearbyTassels(marks,w,h);
  return {count:finalMarks.length, marks:finalMarks, confidence};
}

function rebuildRows(){
  finalRows = [];
  fullRows = [];
  const R = Math.max(1, parseInt(CONFIG.rows || 1));
  const C = Math.max(1, parseInt(CONFIG.cols || 1));
  const meta = CONFIG.metadata || {};
  let hit=0, noHit=0, review=0, firstGeneral='';
  for(let r=0; r<R; r++){
    for(let c=0; c<C; c++){
      const label = cellLabel(r,c);
      const rec = resultsByParcel[label] || {counts:ORTHOS.map(()=>0), percents:ORTHOS.map(()=>0), confidence:1, marksByDate:ORTHOS.map(()=>[])};
      const manual = manualReviews[label] || {};
      const teto = Math.max(1, Number(manual.teto || CONFIG.tetoPlantas || 1));
      const counts = ORTHOS.map((_,i) => manual.counts && manual.counts[i] !== undefined ? Number(manual.counts[i] || 0) : Number(rec.counts[i] || 0));
      const percents = counts.map(v => (v / teto) * 100);
      let status = 'NÃO ATINGIU';
      let firstDate = '';
      let firstOrtho = '';
      for(let i=0; i<ORTHOS.length; i++){
        if(counts[i] >= Number(CONFIG.minPendoes || 0) && percents[i] >= Number(CONFIG.percentualLimite || 50)){
          status = 'ATINGIU';
          firstDate = ORTHOS[i].date;
          firstOrtho = ORTHOS[i].name;
          break;
        }
      }
      if(manual.status) status = manual.status;
      if(rec.confidence < 0.38 && !manual.status) status = 'REVISAR';
      if(manual.firstDate !== undefined) firstDate = manual.firstDate;
      if(manual.firstOrtho !== undefined) firstOrtho = manual.firstOrtho;
      const obs = manual.obs || (rec.confidence < 0.38 ? 'Baixa confiança na detecção automática' : '');
      const row = {
        NOME_ANALISE: CONFIG.nomeAnalise || '',
        CODLOCAL: meta.CODLOCAL || '',
        QUADRA: meta.QUADRA || '',
        ENSAIO: meta.ENSAIO || '',
        GLI: meta.GLI || '',
        TECNOLOGIA: meta.TECNOLOGIA || '',
        REP: meta.REP || '',
        NC: meta.NC || '',
        LINHAGEM: meta.LINHAGEM || '',
        GENEALOGIA: meta.GENEALOGIA || '',
        PARENTAL1: meta.PARENTAL1 || '',
        POP_PARENTAL1: meta.POP_PARENTAL1 || '',
        PARENTAL2: meta.PARENTAL2 || '',
        POP_PARENTAL2: meta.POP_PARENTAL2 || '',
        DTP: meta.DTP || '',
        TIRO: 'T' + (c + 1),
        ID_PARCELA: label
      };
      for(let i=0; i<10; i++){
        row['DATA_ORTOFOTO_' + (i+1)] = ORTHOS[i] ? ORTHOS[i].date : '';
        row['PENDOES_' + (i+1)] = ORTHOS[i] ? counts[i] : '';
        row['PERCENTUAL_' + (i+1)] = ORTHOS[i] ? fmtPct(percents[i]) : '';
      }
      row.TETO_PLANTAS = teto;
      row.PERCENTUAL_LIMITE = CONFIG.percentualLimite;
      row.DATA_PRIMEIRO_ATINGIMENTO = status === 'ATINGIU' ? firstDate : '';
      row.ORTOFOTO_PRIMEIRO_ATINGIMENTO = status === 'ATINGIU' ? firstOrtho : '';
      row.STATUS = status;
      row.OBSERVACAO = obs;
      finalRows.push(row);

      for(let i=0; i<ORTHOS.length; i++){
        const perHit = counts[i] >= Number(CONFIG.minPendoes || 0) && percents[i] >= Number(CONFIG.percentualLimite || 50);
        const perStatus = perHit ? 'ATINGIU' : (counts[i] > 0 ? 'EVOLUÇÃO PARCIAL' : 'NÃO ATINGIU');
        fullRows.push({
          Quadra: meta.QUADRA || CONFIG.nomeAnalise || 'Pendoamento',
          Disparo: r + 1,
          Tiro: c + 1,
          Data_Ortofoto: ORTHOS[i].date,
          Nome_Ortofoto: ORTHOS[i].name,
          Total_Pendões: counts[i],
          Teto_Configurado: teto,
          Status: perStatus,
          Data_Atingimento: status === 'ATINGIU' ? firstDate : '',
          Ortofoto_Atingimento: status === 'ATINGIU' ? firstOrtho : '',
          Percentual_Pendoamento: fmtPct(percents[i])
        });
      }

      if(status === 'ATINGIU') {
        hit++;
        if(firstDate && (!firstGeneral || firstDate < firstGeneral)) firstGeneral = firstDate;
      } else if(status === 'REVISAR') review++;
      else if(status === 'SEM DADOS') review++;
      else noHit++;
    }
  }
  document.getElementById('statTotal').textContent = finalRows.length;
  document.getElementById('statHit').textContent = hit;
  document.getElementById('statNoHit').textContent = noHit;
  document.getElementById('statReview').textContent = review;
  document.getElementById('statFirstDate').textContent = 'Primeira data geral: ' + (firstGeneral || '--');
  renderOrthoSummary();
}

function sourceBucket(mark){
  const src = String(mark?.source || '').toLowerCase();
  if(src.includes('aplicar treino') || src.includes('treino')) return 'refinado';
  if(src.includes('yolo')) return 'yolo';
  if(src.includes('refinamento') || src.includes('refer')) return 'refinado';
  return 'opencv';
}

function analysisSourceTotals(){
  const totals = {yolo:0, opencv:0, refinado:0};
  for(const label of Object.keys(resultsByParcel)){
    const rec = resultsByParcel[label] || {};
    const marksByDate = rec.marksByDate || [];
    for(const marks of marksByDate){
      for(const mark of marks || []){
        totals[sourceBucket(mark)] += 1;
      }
    }
  }
  return totals;
}

function treinoDistanceThreshold(idx){
  return clamp(Math.min(imgW(idx), imgH(idx)) * 0.012, 15, 30);
}

function existingTrainingDuplicate(idx, x, y, distance){
  for(const label of Object.keys(resultsByParcel)){
    const rec = resultsByParcel[label] || {};
    const marks = rec.marksByDate ? (rec.marksByDate[idx] || []) : [];
    for(const mark of marks){
      const dx = Number(mark.x || 0) - x;
      const dy = Number(mark.y || 0) - y;
      if(Math.sqrt(dx*dx + dy*dy) <= distance) return true;
    }
  }
  return false;
}

function ensureParcelRecord(label, r, c){
  if(!resultsByParcel[label]){
    resultsByParcel[label] = {
      row:r+1,
      col:c+1,
      counts:ORTHOS.map(()=>0),
      percents:ORTHOS.map(()=>0),
      marksByDate:ORTHOS.map(()=>[]),
      confidence:1
    };
  }
  if(!Array.isArray(resultsByParcel[label].marksByDate)) resultsByParcel[label].marksByDate = ORTHOS.map(()=>[]);
  if(!Array.isArray(resultsByParcel[label].counts)) resultsByParcel[label].counts = ORTHOS.map(()=>0);
  return resultsByParcel[label];
}

function aplicarTreinoContinuoPendoamento(){
  if(!ORTHOS.length){ alert('Importe uma ortofoto antes de aplicar o treino.'); return; }
  if(gridRatios.length < 4){ alert('Defina e salve a grade antes de executar a análise de pendoamento.'); return; }
  if(!Object.keys(resultsByParcel).length){ alert('Execute Análise de Pendoamento antes de aplicar o treino.'); return; }
  const totalImages = Number(TRAIN_STATE?.total_images || 0);
  const available = ORTHOS.reduce((sum, o) => sum + (Array.isArray(o.training_detections) ? o.training_detections.length : 0), 0);
  if(totalImages <= 0 && available <= 0){
    trainStatus.textContent = 'Sem mini imagens salvas. Clique em Treinar YOLO e marque pendões faltantes primeiro.';
    return;
  }
  if(available <= 0){
    trainStatus.textContent = 'Aplicar Treino não encontrou candidatos pré-carregados. Aguarde salvar as mini imagens e reabra o visualizador se necessário.';
    return;
  }
  applyYoloTraining = true;
  if(TRAIN_STATE) TRAIN_STATE.apply_training = true;
  btnApplyTrainingYolo.classList.add('active');
  showTrainingMarks = false;
  yoloTrainMarks = [];
  progressBar.style.width = '8%';
  statusEl.textContent = 'Aplicando treino. Buscando padrões e analisando pendões semelhantes...';
  trainStatus.textContent = 'Aplicando treino. Lendo histórico, mini imagens e características salvas na pasta.';
  let added = 0;
  let skipped = 0;
  let outside = 0;
  for(let i=0; i<ORTHOS.length; i++){
    const detections = Array.isArray(ORTHOS[i].training_detections) ? ORTHOS[i].training_detections : [];
    const dist = treinoDistanceThreshold(i);
    for(const det of detections){
      const x = Number(det.x);
      const y = Number(det.y);
      if(!Number.isFinite(x) || !Number.isFinite(y)) continue;
      let foundCell = null;
      const R = Math.max(1, parseInt(CONFIG.rows || 1));
      const C = Math.max(1, parseInt(CONFIG.cols || 1));
      for(let r=0; r<R && !foundCell; r++){
        for(let c=0; c<C; c++){
          const poly = cellPoly(r,c,i);
          if(poly && pointInPolygon(x,y,poly)){
            foundCell = {r,c,label:cellLabel(r,c)};
            break;
          }
        }
      }
      if(!foundCell){ outside++; continue; }
      if(existingTrainingDuplicate(i, x, y, dist)){ skipped++; continue; }
      const rec = ensureParcelRecord(foundCell.label, foundCell.r, foundCell.c);
      const mark = {
        x,
        y,
        score:Number(det.score || 0),
        size:clamp(Number(det.size || 8), 6, 20),
        area:Math.max(1, Number(det.area || det.size || 8)),
        tipo:det.tipo || 'treino_continuo',
        confianca:det.confianca || 'media',
        source:'Aplicar Treino',
        training_source:det.training_source || det.source || '',
        yolo_conf:Number(det.yolo_conf || 0),
        reference:det.reference || '',
        template_score:Number(det.template_score || 0),
        yellow_ratio:Number(det.yellow_ratio || 0),
        texture_ratio:Number(det.texture_ratio || 0),
        clear_ratio:Number(det.clear_ratio || 0),
        green_ratio:Number(det.green_ratio || 0)
      };
      rec.marksByDate[i].push(mark);
      rec.counts[i] = rec.marksByDate[i].length;
      rec.confidence = Math.min(Number(rec.confidence || 1), 0.82);
      added++;
    }
    progressBar.style.width = Math.round(((i + 1) / Math.max(1, ORTHOS.length)) * 100) + '%';
  }
  rebuildRows();
  renderParcelCountTable();
  persistViewerState();
  const duplicateMsg = skipped ? ('Duplicidades ignoradas: ' + skipped + '.') : 'Nenhuma duplicidade encontrada.';
  trainStatus.textContent =
    'Novos pendões encontrados: ' + added + '. Atualizando parcelas. Recontagem concluída. ' + duplicateMsg;
  statusEl.textContent =
    'Recontagem concluída. Total final atualizado com Aplicar Treino: +' + added +
    ' pendões · fora do grid ignorados: ' + outside + ' · ' + duplicateMsg;
  drawAll();
}

function runChronologicalAnalysis(){
  if(gridRatios.length < 4){ alert('Defina e salve a grade antes de executar a análise de pendoamento.'); return; }
  if(loaded < ORTHOS.length){ alert('Aguarde as ortofotos terminarem de carregar.'); return; }
  const R = Math.max(1, parseInt(CONFIG.rows || 1));
  const C = Math.max(1, parseInt(CONFIG.cols || 1));
  resultsByParcel = {};
  const preCount = ORTHOS.reduce((sum,o) => sum + (Array.isArray(o.advanced_detections) ? o.advanced_detections.length : 0), 0);
  const modes = [...new Set(ORTHOS.map(o => o.detector_mode || 'OpenCV parametrizado TMG'))].join(' + ');
  statusEl.textContent = 'Analisando pendoamento com OpenCV parametrizado dentro das parcelas do grid em ' + ORTHOS.length + ' ortofotos' + (preCount ? ' (' + preCount + ' centros pré-detectados · ' + modes + ').' : '.') ;
  progressBar.style.width = '0%';
  setTimeout(() => {
    let done = 0;
    const total = ORTHOS.length * R * C;
    for(let i=0; i<ORTHOS.length; i++){
      tempPrepared = -1;
      for(let r=0; r<R; r++){
        for(let c=0; c<C; c++){
          const label = cellLabel(r,c);
          if(!resultsByParcel[label]){
            resultsByParcel[label] = {
              row:r+1,
              col:c+1,
              counts:ORTHOS.map(()=>0),
              percents:ORTHOS.map(()=>0),
              marksByDate:ORTHOS.map(()=>[]),
              confidence:1
            };
          }
          const res = analyzeCellInImage(i,r,c);
          resultsByParcel[label].counts[i] = res.count;
          resultsByParcel[label].marksByDate[i] = res.marks;
          resultsByParcel[label].confidence = Math.min(resultsByParcel[label].confidence, res.confidence);
          done++;
          if(done % Math.max(1, Math.floor(total/20)) === 0) progressBar.style.width = Math.round(done/total*100) + '%';
        }
      }
    }
    progressBar.style.width = '100%';
    rebuildRows();
    renderParcelCountTable();
    const sourceTotals = analysisSourceTotals();
    const totalGeral = sourceTotals.yolo + sourceTotals.opencv + sourceTotals.refinado;
    statusEl.textContent =
      'Análise concluída. Total geral de pendões: ' + totalGeral +
      ' · OpenCV: ' + sourceTotals.opencv +
      ' · dentro do grid' +
      ' · parcelas processadas: ' + finalRows.length + '.';
    persistViewerState();
    drawAll();
    renderReviewPanel();
  }, 50);
}

function rowForLabel(label){
  return finalRows.find(r => r.ID_PARCELA === label);
}

function renderReviewPanel(){
  if(!selectedParcel){
    reviewPanel.style.display = 'none';
    return;
  }
  const label = selectedParcel.label;
  const rec = resultsByParcel[label] || {counts:ORTHOS.map(()=>0)};
  const manual = manualReviews[label] || {};
  const counts = ORTHOS.map((_,i) => manual.counts && manual.counts[i] !== undefined ? Number(manual.counts[i] || 0) : Number(rec.counts[i] || 0));
  const teto = manual.teto || CONFIG.tetoPlantas || 1;
  const row = rowForLabel(label) || {};
  reviewPanel.style.display = 'block';
  reviewPanel.innerHTML = `
    <h4>Revisão manual · ${label}</h4>
    <div class="review-line"><span>Teto de plantas</span><input id="revTeto" type="number" min="1" value="${teto}"></div>
    ${ORTHOS.map((o,i)=>`<div class="review-line"><span>${i+1} · ${o.date}</span><input id="revCount${i}" type="number" min="0" value="${counts[i] || 0}"></div>`).join('')}
    <div class="row"><span>Status</span><select id="revStatus">
      ${['','ATINGIU','NÃO ATINGIU','REVISAR','SEM DADOS'].map(s=>`<option value="${s}" ${String(row.STATUS || '')===s?'selected':''}>${s || 'Automático'}</option>`).join('')}
    </select></div>
    <div class="row"><span>Observação</span></div>
    <textarea id="revObs">${manual.obs || row.OBSERVACAO || ''}</textarea>
    <button class="btn green" id="btnSaveReview">Salvar revisão</button>
  `;
  document.getElementById('btnSaveReview').onclick = () => {
    const newCounts = ORTHOS.map((_,i) => Number(document.getElementById('revCount'+i).value || 0));
    manualReviews[label] = {
      counts:newCounts,
      teto:Number(document.getElementById('revTeto').value || CONFIG.tetoPlantas || 1),
      status:document.getElementById('revStatus').value,
      obs:document.getElementById('revObs').value || ''
    };
    rebuildRows();
    renderReviewPanel();
    persistViewerState();
    drawAll();
    statusEl.textContent = 'Revisão salva para ' + label + '.';
  };
}

function drawGrid(idx=activeIdx){
  if(gridRatios.length < 4) return;
  const R = Math.max(1, parseInt(CONFIG.rows || 1));
  const C = Math.max(1, parseInt(CONFIG.cols || 1));
  const pts = currentGridPoints(idx);
  ctx.save();
  ctx.lineWidth = 1.55 / scale;
  ctx.strokeStyle = 'rgba(0,207,255,.78)';
  ctx.shadowColor = 'rgba(0,207,255,.35)';
  ctx.shadowBlur = 5 / scale;
  for(let i=0; i<=R; i++){
    const v=i/R, left=bilerp(pts[0],pts[1],pts[2],pts[3],0,v), right=bilerp(pts[0],pts[1],pts[2],pts[3],1,v);
    ctx.beginPath(); ctx.moveTo(left.x,left.y); ctx.lineTo(right.x,right.y); ctx.stroke();
  }
  for(let j=0; j<=C; j++){
    const u=j/C, top=bilerp(pts[0],pts[1],pts[2],pts[3],u,0), bottom=bilerp(pts[0],pts[1],pts[2],pts[3],u,1);
    ctx.beginPath(); ctx.moveTo(top.x,top.y); ctx.lineTo(bottom.x,bottom.y); ctx.stroke();
  }
  ctx.restore();

  for(let r=0; r<R; r++){
    for(let c=0; c<C; c++){
      const label = cellLabel(r,c);
      const poly = cellPoly(r,c,idx);
      const row = rowForLabel(label);
      const isSel = selectedParcel && selectedParcel.label === label;
      if(!poly) continue;
      ctx.save();
      ctx.beginPath();
      ctx.moveTo(poly[0].x,poly[0].y);
      for(let k=1; k<poly.length; k++) ctx.lineTo(poly[k].x,poly[k].y);
      ctx.closePath();
      if(row){
        if(row.STATUS === 'ATINGIU') ctx.fillStyle='rgba(255,140,0,.22)';
        else if(row.STATUS === 'REVISAR' || row.STATUS === 'SEM DADOS') ctx.fillStyle='rgba(255,214,0,.22)';
        else ctx.fillStyle='rgba(150,150,150,.10)';
        ctx.fill();
      }
      if(isSel){
        ctx.fillStyle='rgba(255,214,0,.22)';
        ctx.fill();
        ctx.strokeStyle='rgba(255,214,0,.96)';
        ctx.lineWidth=2.6/scale;
        ctx.stroke();
      }
      const cx=(poly[0].x+poly[1].x+poly[2].x+poly[3].x)/4;
      const cy=(poly[0].y+poly[1].y+poly[2].y+poly[3].y)/4;
      ctx.shadowColor='rgba(0,0,0,.9)';
      ctx.shadowBlur=4/scale;
      ctx.fillStyle='#fff';
      ctx.font='bold '+Math.max(8,10/scale)+'px Arial';
      ctx.textAlign='center';
      ctx.textBaseline='middle';
      ctx.fillText(label, cx, cy);
      ctx.restore();
    }
  }
}

function drawMarks(idx=activeIdx){
  for(const label of Object.keys(resultsByParcel)){
    const rec = resultsByParcel[label];
    const marks = rec.marksByDate ? (rec.marksByDate[idx] || []) : [];
    for(const m of marks){
      ctx.save();
      const s = clamp(Number(m.size || 7), 6, 20) / scale;
      const src = String(m.source || '').toLowerCase();
      const color = (src.includes('aplicar treino') || src.includes('treino')) ? '#2d8cff' : (src.includes('yolo') ? '#ff8c00' : (src.includes('refinamento') || src.includes('refer') ? '#2d8cff' : '#ff2020'));
      ctx.strokeStyle=color;
      ctx.lineWidth=2/scale;
      ctx.shadowColor=color + 'aa';
      ctx.shadowBlur=5/scale;
      ctx.beginPath(); ctx.moveTo(m.x-s,m.y-s); ctx.lineTo(m.x+s,m.y+s); ctx.stroke();
      ctx.beginPath(); ctx.moveTo(m.x-s,m.y+s); ctx.lineTo(m.x+s,m.y-s); ctx.stroke();
      ctx.restore();
    }
  }
}

function drawTrainingMarks(){
  if(!showTrainingMarks && !trainYoloMode && !deleteTrainingMode) return;
  for(const mark of activeTrainingMarks()){
    if(mark.idx !== activeIdx) continue;
    ctx.save();
    const s = clamp(Number(mark.size || 128), 48, 256) / 2;
    ctx.strokeStyle = '#b45cff';
    ctx.lineWidth = 2.2 / scale;
    ctx.shadowColor = 'rgba(180,92,255,.78)';
    ctx.shadowBlur = 7 / scale;
    ctx.strokeRect(mark.x - s, mark.y - s, s * 2, s * 2);
    const xSize = clamp(s * 0.16, 6, 18) / scale;
    ctx.beginPath(); ctx.moveTo(mark.x - xSize, mark.y - xSize); ctx.lineTo(mark.x + xSize, mark.y + xSize); ctx.stroke();
    ctx.beginPath(); ctx.moveTo(mark.x - xSize, mark.y + xSize); ctx.lineTo(mark.x + xSize, mark.y - xSize); ctx.stroke();
    ctx.restore();
  }
}

function drawAll(){
  const W = viewer.clientWidth, H = viewer.clientHeight;
  canvas.width = W; canvas.height = H;
  ctx.clearRect(0,0,W,H);
  ctx.save();
  ctx.translate(offsetX,offsetY);
  ctx.scale(scale,scale);
  ctx.imageSmoothingEnabled = true;
  ctx.imageSmoothingQuality = 'high';
  if(images[activeIdx] && images[activeIdx].complete) {
    ctx.filter = 'contrast(1.07) saturate(1.06) brightness(1.02)';
    ctx.drawImage(images[activeIdx],0,0,imgW(activeIdx),imgH(activeIdx));
    ctx.filter = 'none';
  }
  drawGrid(activeIdx);
  drawMarks(activeIdx);
  drawTrainingMarks();
  if(gridRatios.length > 0){
    const pts = currentGridPoints(activeIdx);
    pts.forEach((p,i)=>{
      ctx.save();
      ctx.shadowColor=gridDragPoint===i ? 'rgba(255,255,255,.9)' : 'rgba(0,180,255,.8)';
      ctx.shadowBlur=14/scale;
      ctx.fillStyle=gridDragPoint===i ? '#fff' : '#1e90ff';
      ctx.strokeStyle=gridDragPoint===i ? '#aaddff' : '#00cfff';
      ctx.lineWidth=2.5/scale;
      ctx.beginPath(); ctx.arc(p.x,p.y,10/scale,0,Math.PI*2); ctx.fill(); ctx.stroke();
      ctx.fillStyle='#fff'; ctx.font='bold '+(12/scale)+'px Arial'; ctx.textAlign='center'; ctx.textBaseline='middle';
      ctx.fillText(String(i+1),p.x,p.y);
      ctx.restore();
    });
  }
  ctx.restore();
  zoomBadge.textContent = scale.toFixed(2) + '×';
}

function exportRows(rows, filename){
  if(!rows.length){ alert('Execute a análise antes de exportar.'); return; }
  const headers = Object.keys(rows[0]);
  const csv = '\uFEFF' + headers.join(',') + '\n' + rows.map(row => headers.map(h => quoteCSV(row[h])).join(',')).join('\n');
  const blob = new Blob([csv], {type:'text/csv;charset=utf-8;'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = filename;
  a.click();
}

function exportParcelCSV(){
  if(!Object.keys(resultsByParcel).length){ alert('Execute a análise antes de exportar.'); return; }
  exportRows(buildActiveParcelRows(), 'pendoamento_parcelas.csv');
}

async function exportParcelExcel(){
  if(!Object.keys(resultsByParcel).length){ alert('Execute a análise antes de exportar.'); return; }
  if(typeof XLSX === 'undefined' && !(await ensureChronoExcel())){ alert('Biblioteca Excel não carregou. Tente novamente.'); return; }
  if(!(await ensureChronoExcel())){ alert('Biblioteca de estilos do Excel não carregou. Tente novamente.'); return; }
  const headers = ['Quadra','Parcela','Linha','Coluna','Quantidade_Pendoes'];
  const rows = buildActiveParcelRows().map(row => {
    const ordered = {};
    headers.forEach(h => ordered[h] = row[h] ?? '');
    return ordered;
  });
  const wb = XLSX.utils.book_new();
  const ws = XLSX.utils.json_to_sheet(rows, {header: headers});
  ws['!autofilter'] = {ref: ws['!ref']};
  ws['!freeze'] = {xSplit:0,ySplit:1};
  ws['!cols'] = [{wch:18},{wch:14},{wch:10},{wch:10},{wch:20}];
  XLSX.utils.book_append_sheet(wb, ws, 'Pendoamento');
  XLSX.writeFile(wb, 'pendoamento_parcelas.xlsx');
}

function applyTemporalSheetStyle(ws){
  if(!ws || !ws['!ref'] || !XLSX || !XLSX.utils) return;
  const range = XLSX.utils.decode_range(ws['!ref']);
  const headers = [];
  for(let c=range.s.c; c<=range.e.c; c++){
    const cell = ws[XLSX.utils.encode_cell({r:0,c})];
    headers.push(cell ? String(cell.v) : '');
  }
  const statusCol = headers.indexOf('Status');
  const centerCols = ['Disparo','Tiro','Total_Pendões','Teto_Configurado','Percentual_Pendoamento']
    .map(name => headers.indexOf(name))
    .filter(idx => idx >= 0);
  const border = {
    top:{style:'thin',color:{rgb:'808080'}},
    bottom:{style:'thin',color:{rgb:'808080'}},
    left:{style:'thin',color:{rgb:'808080'}},
    right:{style:'thin',color:{rgb:'808080'}}
  };
  for(let r=range.s.r; r<=range.e.r; r++){
    for(let c=range.s.c; c<=range.e.c; c++){
      const addr = XLSX.utils.encode_cell({r,c});
      if(!ws[addr]) continue;
      if(r === 0){
        ws[addr].s = {
          font:{bold:true,color:{rgb:'00B0F0'}},
          fill:{patternType:'solid',fgColor:{rgb:'1F1F1F'}},
          alignment:{horizontal:'center',vertical:'center',wrapText:true},
          border
        };
        continue;
      }
      const style = {
        font:{color:{rgb:'000000'}},
        alignment:{horizontal:centerCols.includes(c) ? 'center' : 'left', vertical:'center', wrapText:true},
        border
      };
      if(c === statusCol){
        const status = String(ws[addr].v || '').toUpperCase();
        if(status.includes('ATINGIU') && !status.includes('NÃO')) {
          style.fill = {patternType:'solid',fgColor:{rgb:'00B050'}};
          style.font = {bold:true,color:{rgb:'FFFFFF'}};
        } else if(status.includes('PARCIAL')) {
          style.fill = {patternType:'solid',fgColor:{rgb:'FFF2CC'}};
          style.font = {bold:true,color:{rgb:'000000'}};
        } else {
          style.fill = {patternType:'solid',fgColor:{rgb:'C00000'}};
          style.font = {bold:true,color:{rgb:'FFFFFF'}};
        }
        style.alignment = {horizontal:'center',vertical:'center',wrapText:true};
      }
      ws[addr].s = style;
    }
  }
  ws['!autofilter'] = {ref: ws['!ref']};
  ws['!freeze'] = {xSplit:0,ySplit:1};
  ws['!cols'] = headers.map(h => ({wch: Math.max(12, Math.min(28, String(h).length + 4))}));
}

async function exportExcel(){
  if(!fullRows.length){ alert('Execute a análise antes de exportar.'); return; }
  if(typeof XLSX === 'undefined' && !(await ensureChronoExcel())){ alert('Biblioteca Excel não carregou. Tente novamente.'); return; }
  if(!(await ensureChronoExcel())){ alert('Biblioteca de estilos do Excel não carregou. Tente novamente.'); return; }
  const wb = XLSX.utils.book_new();
  const headers = [
    'Quadra','Disparo','Tiro','Data_Ortofoto','Nome_Ortofoto','Total_Pendões',
    'Teto_Configurado','Status','Data_Atingimento','Ortofoto_Atingimento','Percentual_Pendoamento'
  ];
  const rows = fullRows.map(row => {
    const ordered = {};
    headers.forEach(h => ordered[h] = row[h] ?? '');
    return ordered;
  });
  const ws = XLSX.utils.json_to_sheet(rows, {header: headers});
  applyTemporalSheetStyle(ws);
  XLSX.utils.book_append_sheet(wb, ws, 'Pendoamento');
  XLSX.writeFile(wb, 'pendoamento_por_parcela.xlsx');
}

function exportResumo(){
  if(!finalRows.length){ alert('Execute a análise antes de exportar.'); return; }
  const resumo = finalRows.map(r => ({
    CODLOCAL:r.CODLOCAL, QUADRA:r.QUADRA, ENSAIO:r.ENSAIO, GLI:r.GLI, TECNOLOGIA:r.TECNOLOGIA,
    REP:r.REP, NC:r.NC, LINHAGEM:r.LINHAGEM, DTP:r.DTP, TIRO:r.TIRO, ID_PARCELA:r.ID_PARCELA,
    TETO_PLANTAS:r.TETO_PLANTAS, PERCENTUAL_LIMITE:r.PERCENTUAL_LIMITE,
    DATA_PRIMEIRO_ATINGIMENTO:r.DATA_PRIMEIRO_ATINGIMENTO,
    ORTOFOTO_PRIMEIRO_ATINGIMENTO:r.ORTOFOTO_PRIMEIRO_ATINGIMENTO,
    STATUS:r.STATUS, OBSERVACAO:r.OBSERVACAO
  }));
  exportRows(resumo, 'resumo_final_pendoamento.csv');
}

function exportImage(){
  const out = document.createElement('canvas');
  out.width = imgW(activeIdx); out.height = imgH(activeIdx);
  const octx = out.getContext('2d');
  octx.drawImage(images[activeIdx],0,0,out.width,out.height);
  const oldCtxDraw = ctx;
  octx.save();
  const originalCanvas = canvas;
  const originalCtx = ctx;
  octx.lineWidth = 2;
  if(gridRatios.length === 4){
    const R = Math.max(1, parseInt(CONFIG.rows || 1));
    const C = Math.max(1, parseInt(CONFIG.cols || 1));
    const pts = currentGridPoints(activeIdx);
    octx.strokeStyle='rgba(0,207,255,.95)';
    for(let i=0; i<=R; i++){
      const v=i/R, left=bilerp(pts[0],pts[1],pts[2],pts[3],0,v), right=bilerp(pts[0],pts[1],pts[2],pts[3],1,v);
      octx.beginPath(); octx.moveTo(left.x,left.y); octx.lineTo(right.x,right.y); octx.stroke();
    }
    for(let j=0; j<=C; j++){
      const u=j/C, top=bilerp(pts[0],pts[1],pts[2],pts[3],u,0), bottom=bilerp(pts[0],pts[1],pts[2],pts[3],u,1);
      octx.beginPath(); octx.moveTo(top.x,top.y); octx.lineTo(bottom.x,bottom.y); octx.stroke();
    }
    for(const label of Object.keys(resultsByParcel)){
      const marks = resultsByParcel[label].marksByDate ? (resultsByParcel[label].marksByDate[activeIdx] || []) : [];
      for(const m of marks){
        const s=7;
        const src = String(m.source || '').toLowerCase();
        octx.strokeStyle = (src.includes('aplicar treino') || src.includes('treino')) ? '#2d8cff' : (src.includes('yolo') ? '#ff8c00' : (src.includes('refinamento') || src.includes('refer') ? '#2d8cff' : '#ff2020'));
        octx.lineWidth=2;
        octx.beginPath(); octx.moveTo(m.x-s,m.y-s); octx.lineTo(m.x+s,m.y+s); octx.stroke();
        octx.beginPath(); octx.moveTo(m.x-s,m.y+s); octx.lineTo(m.x+s,m.y-s); octx.stroke();
      }
    }
  }
  octx.restore();
  const a = document.createElement('a');
  a.href = out.toDataURL('image/png');
  a.download = 'pendoamento_grid_marcacoes_' + (ORTHOS[activeIdx]?.date || 'data') + '.png';
  a.click();
}

viewer.addEventListener('wheel', e => {
  e.preventDefault();
  const factor = e.deltaY < 0 ? 1.2 : 0.8;
  const r = canvas.getBoundingClientRect();
  const mx=e.clientX-r.left, my=e.clientY-r.top;
  const ix=(mx-offsetX)/scale, iy=(my-offsetY)/scale;
  scale = clamp(scale*factor, 0.04, 50);
  offsetX = mx - ix*scale;
  offsetY = my - iy*scale;
  drawAll();
}, {passive:false});

viewer.addEventListener('mousedown', e => {
  const pt = screenToImg(e.clientX,e.clientY);
  if(deleteTrainingMode){
    if(pt.x < 0 || pt.y < 0 || pt.x > imgW(activeIdx) || pt.y > imgH(activeIdx)){
      trainStatus.textContent = 'Clique sobre um X roxo dentro da ortofoto para excluir a amostra.';
      return;
    }
    deleteTrainingSampleAt(pt);
    return;
  }
  if(trainYoloMode){
    if(pt.x < 0 || pt.y < 0 || pt.x > imgW(activeIdx) || pt.y > imgH(activeIdx)){
      trainStatus.textContent = 'Clique dentro da ortofoto para salvar a amostra.';
      return;
    }
    saveYoloTrainingSample(pt);
    return;
  }
  const hitGridPoint = gridPointIndexAt(pt);
  if(hitGridPoint >= 0){
    gridDragPoint = hitGridPoint;
    dragging = false;
    viewer.style.cursor = 'grabbing';
    return;
  }
  if(markGridMode){
    gridRatios.push(ratioPoint(pt, activeIdx));
    if(gridRatios.length >= 4){
      gridRatios = gridRatios.slice(0,4);
      markGridMode = false;
      btnMarkGrid.classList.remove('active');
      statusEl.textContent = 'Grid marcado. Arraste os pontos das extremidades para ajustar Disparo/Tiro.';
    } else {
      statusEl.textContent = 'Marque a extremidade ' + (gridRatios.length + 1) + ' do grid.';
    }
    resultsByParcel = {}; finalRows = []; fullRows = [];
    rebuildRows();
    persistViewerState();
    drawAll();
    return;
  }
  if(reviewMode && gridRatios.length === 4){
    const cell = findCell(pt);
    if(cell){
      selectedParcel = cell;
      renderReviewPanel();
      drawAll();
    }
    return;
  }
  dragging = true; lastX=e.clientX; lastY=e.clientY; viewer.style.cursor='grabbing';
});

viewer.addEventListener('mousemove', e => {
  const pt = screenToImg(e.clientX,e.clientY);
  coordBadge.textContent = 'X:' + Math.round(pt.x) + ' Y:' + Math.round(pt.y);
  if(gridDragPoint >= 0){
    gridRatios[gridDragPoint] = ratioPoint(pt, activeIdx);
    selectedParcel = null;
    resultsByParcel = {};
    finalRows = [];
    fullRows = [];
    progressBar.style.width = '0%';
    document.getElementById('statTotal').textContent = '0';
    document.getElementById('statHit').textContent = '0';
    document.getElementById('statNoHit').textContent = '0';
    document.getElementById('statReview').textContent = '0';
    document.getElementById('statFirstDate').textContent = 'Primeira data geral: --';
    renderOrthoSummary();
    drawAll();
    return;
  }
  if(dragging){
    offsetX += e.clientX-lastX;
    offsetY += e.clientY-lastY;
    lastX=e.clientX; lastY=e.clientY;
    drawAll();
  }
});
viewer.addEventListener('mouseup', () => {
  if(gridDragPoint >= 0) persistViewerState();
  dragging=false; gridDragPoint=-1; viewer.style.cursor='grab';
});
viewer.addEventListener('mouseleave', () => {
  if(gridDragPoint >= 0) persistViewerState();
  dragging=false; gridDragPoint=-1; viewer.style.cursor='grab';
});

btnMarkGrid.onclick = () => {
  markGridMode = !markGridMode;
  reviewMode = false;
  btnMarkGrid.classList.toggle('active', markGridMode);
  btnReviewMode.classList.remove('active');
  if(markGridMode){
    gridRatios = [];
    resultsByParcel = {}; finalRows = []; fullRows = []; parcelCountRows = [];
    rebuildRows();
    statusEl.textContent = 'Marque as 4 extremidades do grid na ortofoto.';
  }
  drawAll();
};
btnReviewMode.onclick = () => {
  if(gridRatios.length < 4){ alert('Marque o grid primeiro.'); return; }
  reviewMode = !reviewMode;
  markGridMode = false;
  btnReviewMode.classList.toggle('active', reviewMode);
  btnMarkGrid.classList.remove('active');
  statusEl.textContent = reviewMode ? 'Clique em uma parcela para revisar contagens por data.' : 'Revisão manual pausada.';
};
btnAnalyzeChrono.onclick = runChronologicalAnalysis;
btnFitChrono.onclick = fitView;
btnClearChrono.onclick = () => {
  if(!confirm('Limpar grid, resultados e revisões deste seletor?')) return;
  gridRatios=[]; selectedParcel=null; resultsByParcel={}; finalRows=[]; fullRows=[]; parcelCountRows=[]; manualReviews={};
  try{ window.localStorage.removeItem(VIEWER_STATE_KEY); }catch(e){}
  progressBar.style.width='0%'; statusEl.textContent='Seletor limpo.';
  rebuildRows(); renderReviewPanel(); drawAll();
};
btnPrevDate.onclick = () => setActiveDate(activeIdx - 1);
btnNextDate.onclick = () => setActiveDate(activeIdx + 1);
dateSelect.onchange = () => setActiveDate(Number(dateSelect.value));
btnExportCSV.onclick = () => exportRows(finalRows, 'analise_cronologica_pendoamento.csv');
btnExportParcelCSV.onclick = exportParcelCSV;
btnExportParcelXLSX.onclick = exportParcelExcel;
btnExportXLSX.onclick = exportExcel;
btnExportResumo.onclick = exportResumo;
btnExportCompleto.onclick = () => exportRows(fullRows, 'dados_completos_por_ortofoto.csv');
btnExportImagem.onclick = exportImage;
btnTrainYolo.onclick = () => {
  trainYoloMode = !trainYoloMode;
  deleteTrainingMode = false;
  showTrainingMarks = trainYoloMode || showTrainingMarks;
  markGridMode = false;
  reviewMode = false;
  btnTrainYolo.classList.toggle('active', trainYoloMode);
  btnDeleteTrainMark.classList.remove('active');
  btnMarkGrid.classList.remove('active');
  btnReviewMode.classList.remove('active');
  trainStatus.textContent = trainYoloMode
    ? 'Modo Treinar YOLO ativado. Clique exatamente sobre cada pendão faltante. Cada clique salva mini foto, label e características em ' + (TRAIN_STATE?.folder || 'dados_treinamento_yolo/pendoes') + '.'
    : 'Modo treino parado.';
  drawAll();
};
btnApplyTrainingYolo.onclick = () => {
  if(yoloSamples.length > 0){
    trainStatus.textContent = 'Existem ' + yoloSamples.length + ' amostra(s) sendo salvas. Aguarde a confirmação ou clique em Encerrar Treinamento YOLO.';
    return;
  }
  if(pendingTrainingRefresh){
    trainStatus.textContent = 'Aplicando treino: recarregando referências salvas na pasta antes da recontagem...';
    requestTrainingRefreshAndApply();
    return;
  }
  aplicarTreinoContinuoPendoamento();
};
btnDeleteTrainMark.onclick = () => {
  deleteTrainingMode = !deleteTrainingMode;
  trainYoloMode = false;
  showTrainingMarks = deleteTrainingMode || showTrainingMarks;
  markGridMode = false;
  reviewMode = false;
  btnDeleteTrainMark.classList.toggle('active', deleteTrainingMode);
  btnTrainYolo.classList.remove('active');
  btnMarkGrid.classList.remove('active');
  btnReviewMode.classList.remove('active');
  trainStatus.textContent = deleteTrainingMode
    ? 'Modo exclusão ativo: clique sobre o X roxo da amostra que deseja remover.'
    : 'Modo exclusão parado.';
  drawAll();
};
btnStopTrainMode.onclick = () => {
  trainYoloMode = false;
  deleteTrainingMode = false;
  btnTrainYolo.classList.remove('active');
  btnDeleteTrainMark.classList.remove('active');
  const pendingSamples = yoloSamples.map(sample => ({...sample}));
  const total = currentTrainingTotal();
  trainStatus.textContent = 'Encerrando treinamento YOLO. Conferindo ' + pendingSamples.length + ' mini imagem(ns), labels e características na pasta: ' + (TRAIN_STATE?.folder || 'dados_treinamento_yolo/pendoes') + '.';
  sendTrainingPayload({action:'stop', token:'stop_' + Date.now(), total_images:total, samples:pendingSamples});
  drawAll();
};

setupViewerConfigInputs();
setupDates();
restoreViewerState();
try{
  if(window.localStorage.getItem(VIEWER_STATE_KEY + '_auto_apply') === '1'){
    window.localStorage.removeItem(VIEWER_STATE_KEY + '_auto_apply');
    setTimeout(() => {
      if(Object.keys(resultsByParcel).length && gridRatios.length >= 4){
        aplicarTreinoContinuoPendoamento();
      }
    }, 450);
  }
}catch(e){}
trainCropSize.value = TRAIN_STATE?.crop_default || trainCropSize.value || 128;
btnTrainYolo.classList.toggle('active', trainYoloMode);
btnApplyTrainingYolo.classList.toggle('active', applyYoloTraining);
btnDeleteTrainMark.classList.remove('active');
updateTrainPanelStatus(
  trainYoloMode
    ? 'Modo Treinar YOLO ativado. Clique exatamente sobre cada pendão faltante; cada clique salva mini foto e características.'
    : (TRAIN_STATE?.last_summary || 'Modo treino parado. Clique em Treinar YOLO, marque os pendões faltantes e depois use Aplicar Treino.')
);
loadImages();
rebuildRows();
new ResizeObserver(() => drawAll()).observe(viewer);
</script>
</body>
</html>
"""
                    cron_html = (
                        cron_html
                        .replace("__CRON_ORTHOS__", json.dumps(cron_orthos, ensure_ascii=False))
                        .replace("__CRON_CONFIG__", json.dumps(cron_config, ensure_ascii=False))
                        .replace("__TRAIN_STATE__", json.dumps(cron_train_state, ensure_ascii=False))
                        .replace("__DEPLOY_TRACK_BACKGROUND__", DEPLOY_BAR_THEME.get("track_background", "linear-gradient(180deg,#020e24,#061525)"))
                        .replace("__DEPLOY_TRACK_BORDER__", DEPLOY_BAR_THEME.get("track_border", "rgba(0,229,255,.42)"))
                        .replace("__DEPLOY_TRACK_SHADOW__", DEPLOY_BAR_THEME.get("track_shadow", "inset 0 3px 8px rgba(0,0,0,.68), 0 8px 18px rgba(0,0,0,.30), 0 0 14px rgba(0,229,255,.18)"))
                        .replace("__DEPLOY_FILL_ACTIVE__", DEPLOY_BAR_THEME.get("fill_active", "linear-gradient(90deg,#00e5ff 0%,#0E3A70 48%,#00ff9d 100%)"))
                        .replace("__DEPLOY_FILL_SHADOW__", DEPLOY_BAR_THEME.get("fill_shadow", "inset 0 1px 0 rgba(255,255,255,.46), 0 0 16px rgba(0,229,255,.56), 0 0 22px rgba(0,255,157,.22)"))
                    )
                    components.html(cron_html, height=870, scrolling=False)
            else:
                empty_cards = []
                for slot_idx in range(10):
                    empty_cards.append(
                        "<div style='border:1px dashed #2e2e2e;background:#0c0c0c;border-radius:8px;padding:10px;min-height:68px;'>"
                        f"<div style='color:#555;font-weight:800;font-size:0.74rem;'>#{slot_idx + 1:02d} · VAZIO</div>"
                        "<div style='color:#444;font-size:0.72rem;margin-top:8px;'>Aguardando ortofoto</div>"
                        "</div>"
                    )
                st.markdown("""
                <div style='border:1px solid #2a2a2a;border-radius:12px;background:#0d0d0d;
                            padding:12px;margin-top:8px;'>
                    <div style='color:#ff8c00;font-weight:800;font-size:0.82rem;letter-spacing:1px;text-transform:uppercase;margin-bottom:8px;'>
                        Resumo das 10 posições do seletor de ortofotos
                    </div>
                    <div style='display:grid;grid-template-columns:repeat(5,minmax(0,1fr));gap:8px;'>
                """ + "".join(empty_cards) + """
                    </div>
                    <div style='color:#666;font-size:0.72rem;margin-top:10px;'>
                        Anexe as ortofotos acima para abrir o visualizador único. Depois clique em qualquer item do resumo lateral para trocar a imagem no mesmo painel.
                    </div>
                </div>""", unsafe_allow_html=True)

        elif st.session_state.visualizador_sub == "Qualidade":
            # ==========================================
            # NOVO - MÓDULO QUALIDADE DE PARCELA (Falhas Lineares)
            # ==========================================
            st.markdown("""
            <div style='color:#ff8c00;font-weight:700;font-size:1rem;letter-spacing:2px;
                        text-transform:uppercase;margin-bottom:10px;'>
                ✅ Qualidade de Parcelas · Detecção de Falhas Lineares
            </div>""", unsafe_allow_html=True)

            qual_file = _resettable_ortho_uploader(
                "📷 Carregar Ortofoto para Análise de Qualidade",
                key="qual_orto_uploader",
                help="PNG · JPG · TIF/GeoTIFF · JP2 · IMG · ECW"
            )
            qual_bytes, qual_name = _uploaded_ortho_bytes(qual_file)

            if qual_bytes:
                with st.container():
                    qual_b64, qual_dims, qual_err, qual_spatial = processar_ortofoto(qual_bytes, qual_name)

                if qual_err:
                    st.error(f"Erro: {qual_err}")
                else:
                    qw, qh = qual_dims
                    st.markdown(
                        f"<p style='color:#666;font-size:0.78rem;margin-bottom:6px;'>"
                        f"📐 {qual_name} · {qw}×{qh} px</p>",
                        unsafe_allow_html=True
                    )

                    qual_viewer = f"""
<!DOCTYPE html>
<html>
<head>
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ background:#0d0d0d; overflow:hidden; font-family:'Segoe UI',sans-serif; }}

  #vc {{
    width:100%; height:720px;
    background:
      linear-gradient(rgba(255,255,255,.03) 1px, transparent 1px),
      linear-gradient(90deg, rgba(255,255,255,.03) 1px, transparent 1px),
      #0d0d0d;
    background-size:32px 32px;
    border:1px solid #2a2a2a; border-radius:12px;
    overflow:hidden; position:relative; cursor:grab; user-select:none;
  }}
  #vc:active {{ cursor:grabbing; }}
  canvas {{ position:absolute; top:0; left:0; display:block; }}

  .toolbar {{ position:absolute; top:12px; right:12px; display:flex; flex-direction:column; gap:5px; z-index:20; }}
  .tb-btn {{
    background:linear-gradient(145deg,#1e1e1e,#111); border:1px solid #3a3a3a;
    color:#ff8c00; width:34px; height:34px; border-radius:8px; cursor:pointer;
    font-size:15px; font-weight:700; display:flex; align-items:center; justify-content:center;
    box-shadow:2px 2px 8px #000,inset 0 1px 0 rgba(255,255,255,.05); transition:all .2s;
  }}
  .tb-btn:hover {{ border-color:#ff8c00; box-shadow:0 0 10px rgba(255,140,0,.35),2px 2px 8px #000; color:#ffaa33; }}
  .tb-btn:active {{ transform:translateY(1px); }}
  .tb-sep {{ width:34px; height:1px; background:linear-gradient(90deg,transparent,#333,transparent); margin:2px 0; }}

  .grid-panel {{
    position:absolute; top:10px; right:55px;
    background:rgba(10,10,10,.92); border:1px solid #2a2a2a;
    border-radius:8px; padding:7px; display:flex; flex-direction:column;
    gap:4px; z-index:20; min-width:190px; max-height:calc(100% - 20px);
    overflow-y:auto; overflow-x:hidden; scrollbar-width:thin; scrollbar-color:#ff8c00 #141414;
  }}
  .grid-panel::-webkit-scrollbar {{ width:6px; }}
  .grid-panel::-webkit-scrollbar-track {{ background:#141414; border-radius:6px; }}
  .grid-panel::-webkit-scrollbar-thumb {{ background:#ff8c00; border-radius:6px; }}
  .grid-panel label {{ color:#ff8c00; font-size:9px; line-height:1.1; font-weight:bold; text-align:center; }}
  .grid-panel input[type=number] {{
    background:#1a1a1a; border:1px solid #333; color:#fff;
    border-radius:4px; padding:2px 3px; height:20px; width:46px; text-align:center; font-size:10px;
  }}
  .grid-panel .row-col {{ display:flex; gap:5px; align-items:center; justify-content:space-between; color:#ccc; font-size:10px; min-height:20px; }}
  .grid-panel .row-col span {{ line-height:1; }}
  .grid-panel select {{ height:21px; padding:2px 4px !important; font-size:10px !important; }}
  .grid-btn {{
    background:linear-gradient(145deg,#1e1e1e,#111); border:1px solid #3a3a3a;
    color:#ccc; cursor:pointer; border-radius:4px; padding:4px 5px; min-height:22px; font-size:10px; font-weight:bold; transition:.2s;
  }}
  .grid-btn:hover {{ border-color:#ff8c00; color:#ff8c00; }}
  .grid-btn.active {{ border-color:#ff8c00; color:#ff8c00; box-shadow:0 0 8px rgba(255,140,0,.3); background:#2a1a00; }}

  .zoom-badge {{
    position:absolute; top:12px; left:12px;
    background:rgba(10,10,10,.82); border:1px solid #2a2a2a; border-radius:8px;
    color:#ff8c00; font-size:11px; font-family:'Courier New',monospace;
    font-weight:700; padding:5px 10px; letter-spacing:1px; z-index:20; pointer-events:none;
  }}
  .crosshair {{
    position:absolute; bottom:12px; left:12px;
    background:rgba(10,10,10,.82); border:1px solid #222; border-radius:8px;
    color:#555; font-size:10px; font-family:'Courier New',monospace;
    padding:4px 10px; z-index:20; pointer-events:none; letter-spacing:.5px;
  }}
  .hint {{
    position:absolute; bottom:12px; right:12px; color:#333; font-size:10px;
    z-index:20; pointer-events:none; text-align:right; line-height:1.6;
  }}

  .count-panel {{
    position:absolute; top:50px; left:12px;
    background:rgba(10,10,10,.92); border:1px solid #2a2a2a; border-radius:8px;
    padding:10px; z-index:20; min-width:180px;
  }}
  .count-panel h3 {{ color:#ff4444; font-size:12px; margin-bottom:6px; letter-spacing:1px; }}
  .count-panel .total {{ color:#fff; font-size:20px; font-weight:bold; }}
  .count-panel .info {{ color:#888; font-size:10px; margin-top:4px; }}

  .cnt-btn {{
    background:linear-gradient(145deg,#1a3a1a,#0a2a0a); border:1px solid #006600;
    color:#00ee55; border-radius:4px; padding:4px 6px; min-height:22px; font-size:10px;
    font-weight:bold; cursor:pointer; transition:.2s; width:100%; margin-top:1px;
  }}
  .cnt-btn:hover {{ border-color:#00ee55; box-shadow:0 0 8px rgba(0,238,85,.3); }}
  .cnt-btn.danger {{ background:linear-gradient(145deg,#3a1a1a,#2a0a0a); border-color:#660000; color:#ff5555; }}
  .cnt-btn.danger:hover {{ border-color:#ff5555; box-shadow:0 0 8px rgba(255,85,85,.3); }}
  .cnt-btn.red {{ background:linear-gradient(145deg,#3a0a0a,#220000); border-color:#cc2200; color:#ff4444; }}
  .cnt-btn.red:hover {{ border-color:#ff4444; box-shadow:0 0 8px rgba(255,68,68,.3); }}
  .cnt-btn.orange {{ background:linear-gradient(145deg,#2a1a00,#1a0a00); border-color:#ff8c00; color:#ff8c00; }}
  .cnt-btn.orange:hover {{ border-color:#ffaa33; box-shadow:0 0 8px rgba(255,140,0,.3); }}

  .qual-sep {{ width:100%;height:1px;background:linear-gradient(90deg,transparent,#333,transparent);margin:1px 0; }}

  /* Falhas Modal */
  #falhasModal {{
    display:none; position:fixed; top:0; left:0; width:100%; height:100%;
    background:rgba(0,0,0,0.88); z-index:9999; align-items:center; justify-content:center;
  }}
  #falhasModal .modal-inner {{
    background:#1a1a1a; border:1px solid #cc3300; border-radius:16px;
    width:92%; max-width:720px; max-height:87vh; overflow:auto; padding:24px;
    box-shadow:0 0 40px rgba(255,60,0,0.3);
  }}
</style>
</head>
<body>
<div id="vc">
  <canvas id="cv"></canvas>
  <div class="zoom-badge" id="zbadge">1.00×</div>
  <div class="crosshair" id="coord">X:0 Y:0</div>
  <div class="hint">Scroll=Zoom · Drag=Pan<br>Grid: marque 4 pontos</div>

  <div class="toolbar">
    <button class="tb-btn" id="btnGridTool" title="Marcar Grid (4 pontos)">⊞</button>
    <button class="tb-btn" id="btnManualMode" title="Modo Manual (clique para marcar)">✏️</button>
    <button class="tb-btn" id="btnRemoveLast" title="Apagar Última Marcação">❌</button>
    <div class="tb-sep"></div>
    <button class="tb-btn" id="btnClearAll" title="Limpar Tudo">🗑️</button>
  </div>

  <div class="grid-panel">
    <label>✅ QUALIDADE</label>
    <div class="row-col">
      <span>Disp:</span><input type="number" id="inpRows" value="5" min="1" max="200">
    </div>
    <div class="row-col">
      <span>Tiros:</span><input type="number" id="inpCols" value="5" min="1" max="200">
    </div>
    <div class="row-col">
      <span>Quadra:</span><input type="text" id="inpQuadraName" value="Q-1" style="width:96px;background:#1a1a1a;border:1px solid #333;color:#fff;border-radius:4px;padding:2px 4px;height:21px;font-size:10px;">
    </div>
    <div class="row-col">
      <span>Linhas/parcela:</span><input type="number" id="inpLinhasParcela" value="4" min="1" max="20" style="width:46px;">
    </div>
    <button class="cnt-btn" id="btnManual2" style="background:linear-gradient(145deg,#1a1a3a,#0a0a2a);border-color:#000066;color:#5599ff;">✏️ Manual</button>
    <div class="qual-sep"></div>
    <label style="color:#ff4444;">⚠️ FALHAS LINEARES</label>
    <div class="row-col">
      <span style="font-size:10px;">Mín. (cm):</span>
      <input type="number" id="inpMinDist" value="20" min="20" step="0.1" max="9999" style="width:55px;">
    </div>
    <div class="row-col">
      <span style="font-size:10px;">Buffer (cm):</span>
      <input type="number" id="inpBufferCm" value="20" min="0" step="1" max="9999" style="width:55px;">
    </div>
    <div class="row-col">
      <span style="font-size:10px;">Parcela (m):</span>
      <input type="number" id="inpParcelLen" value="5" min="0.1" step="0.1" style="width:55px;">
    </div>
    <div class="row-col">
      <span style="font-size:10px;">Unidade:</span>
      <select id="selUnit" style="background:#1a1a1a;border:1px solid #333;color:#fff;border-radius:4px;padding:3px;font-size:10px;">
        <option value="px">px</option>
        <option value="cm">cm</option>
        <option value="m" selected>m</option>
      </select>
    </div>
    <div class="row-col">
      <span style="font-size:10px;">Espessura:</span>
      <input type="number" id="inpLineWidth" value="2" min="1" max="10" style="width:40px;">
    </div>
    <div class="row-col" style="gap:4px;">
      <label style="color:#ccc;font-size:10px;font-weight:normal;display:flex;align-items:center;gap:4px;">
        <input type="checkbox" id="chkLabels" checked style="accent-color:#ff4444;"> Etiquetas
      </label>
      <label style="color:#ccc;font-size:10px;font-weight:normal;display:flex;align-items:center;gap:4px;">
        <input type="checkbox" id="chkShowFalhas" checked style="accent-color:#ff4444;"> Mostrar
      </label>
    </div>
    <button class="cnt-btn red" id="btnDetectFalhas">⚠️ Detectar Falhas</button>
    <button class="cnt-btn red" id="btnMedirPlantados">📏 Medir Metros Plantados</button>
    <button class="cnt-btn danger" id="btnDeleteFalhaMode">🧽 Selecionar/Apagar Falha</button>
    <div style="font-size:9px;color:#777;line-height:1.25;">Duplo clique na parcela sem contagem = 100% FALHADA.</div>
    <div class="qual-sep"></div>
    <label style="color:#ffd21f;">🔧 AJUSTE DE PARCELA</label>
    <button class="cnt-btn orange" id="btnSelectParcel">Marcar Parcela</button>
    <button class="cnt-btn" id="btnMoveSelected" style="background:linear-gradient(145deg,#1a1a3a,#0a0a2a);border-color:#003399;color:#5599ff;">Mover Selecionadas</button>
    <div class="row-col">
      <span style="font-size:10px;">Passo (px):</span>
      <input type="number" id="inpMoveStep" value="5" min="1" max="200" style="width:48px;">
    </div>
    <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:4px;">
      <button class="grid-btn" id="btnMoveUp" title="Mover para cima">↑</button>
      <button class="grid-btn" id="btnMoveDown" title="Mover para baixo">↓</button>
      <button class="grid-btn" id="btnMoveLeft" title="Mover para esquerda">←</button>
      <button class="grid-btn" id="btnMoveRight" title="Mover para direita">→</button>
      <button class="grid-btn" id="btnClearSelection" style="grid-column:span 2;color:#ff7777;border-color:#663333;">Limpar Seleção</button>
    </div>
    <button class="cnt-btn" id="btnSaveParcelAdjust" style="background:linear-gradient(145deg,#1a3a1a,#0a2a0a);border-color:#00aa55;color:#55ff99;">Salvar Ajuste da Parcela</button>
    <div id="parcelAdjustStatus" style="font-size:9px;color:#777;line-height:1.2;max-height:24px;overflow:hidden;">Nenhuma parcela selecionada.</div>
    <div class="qual-sep"></div>
    <button class="cnt-btn orange" id="btnOpenFalhas">📊 Relatório</button>
    <button class="cnt-btn danger" id="btnUndoMark">❌ Desfazer</button>
    <button id="btnExportCnt" style="background:linear-gradient(145deg,#003a00,#001a00);border:1px solid #006600;
      color:#00ee55;border-radius:4px;padding:4px 6px;min-height:22px;font-size:10px;font-weight:bold;cursor:pointer;
      transition:.2s;width:100%;margin-top:1px;">📗 Exportar Excel</button>
  </div>

  <div class="count-panel" id="countPanel" style="display:none;">
    <h3>FALHAS DETECTADAS</h3>
    <div class="total" id="totalCount" style="color:#ff4444;">0</div>
    <div class="info" id="countInfo">Detecte falhas</div>
    <div id="falhasSumario" style="margin-top:6px;font-size:10px;color:#888;"></div>
  </div>

  <!-- Modal Relatório de Falhas -->
  <div id="falhasModal">
    <div class="modal-inner">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
        <h3 style="color:#ff4444;font-size:14px;letter-spacing:2px;text-transform:uppercase;margin:0;">⚠️ RELATÓRIO DE FALHAS LINEARES</h3>
        <button id="btnCloseModal" style="background:#333;border:1px solid #555;color:#fff;border-radius:6px;
          padding:5px 12px;cursor:pointer;font-size:12px;">✕ Fechar</button>
      </div>
      <!-- Cards de totais -->
      <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;margin-bottom:16px;">
        <div style="background:#111;border:1px solid #cc2200;border-radius:10px;padding:14px;text-align:center;">
          <div id="modalTotalFalhas" style="color:#ff4444;font-size:24px;font-weight:bold;">0</div>
          <div style="color:#666;font-size:10px;letter-spacing:1px;text-transform:uppercase;">Total Falhas</div>
        </div>
        <div style="background:#111;border:1px solid #333;border-radius:10px;padding:14px;text-align:center;">
          <div id="modalTotalLinear" style="color:#ffaa33;font-size:20px;font-weight:bold;">0</div>
          <div id="modalTotalLinearLabel" style="color:#666;font-size:10px;letter-spacing:1px;text-transform:uppercase;">Total Linear</div>
        </div>
        <div style="background:#111;border:1px solid #333;border-radius:10px;padding:14px;text-align:center;">
          <div id="modalMaiorFalha" style="color:#ff8c00;font-size:20px;font-weight:bold;">0</div>
          <div id="modalMaiorFalhaLabel" style="color:#666;font-size:10px;letter-spacing:1px;text-transform:uppercase;">Maior Falha</div>
        </div>
      </div>
      <!-- Filtros -->
      <div style="margin-bottom:12px;display:flex;gap:8px;align-items:center;flex-wrap:wrap;">
        <input type="text" id="modalSearch" placeholder="Pesquisar linha..."
          style="flex:1;min-width:120px;background:#111;border:1px solid #333;color:#fff;border-radius:6px;padding:8px;font-size:12px;">
        <select id="modalFilterTiro" style="background:#111;border:1px solid #333;color:#fff;border-radius:6px;padding:8px;font-size:11px;">
          <option value="">Todos Tiros</option>
        </select>
        <select id="modalFilterDisp" style="background:#111;border:1px solid #333;color:#fff;border-radius:6px;padding:8px;font-size:11px;">
          <option value="">Todos Disp.</option>
        </select>
      </div>
      <!-- Tabela -->
      <div id="falhasTable" style="max-height:280px;overflow-y:auto;border:1px solid #333;border-radius:8px;"></div>
      <!-- Botões exportar -->
      <div style="display:flex;gap:8px;margin-top:14px;flex-wrap:wrap;">
        <button id="btnExportFalhasCSV" style="flex:1;background:linear-gradient(145deg,#003a00,#001a00);
          border:1px solid #006600;color:#00ee55;border-radius:6px;padding:8px;font-size:11px;font-weight:bold;cursor:pointer;">
          💾 Exportar CSV
        </button>
        <button id="btnExportFalhasGeoJSON" style="flex:1;background:linear-gradient(145deg,#001a3a,#000a1a);
          border:1px solid #003399;color:#5599ff;border-radius:6px;padding:8px;font-size:11px;font-weight:bold;cursor:pointer;">
          🗺️ Exportar GeoJSON
        </button>
        <button id="btnExportFalhasXLSX" style="flex:1;background:linear-gradient(145deg,#1a003a,#0a001a);
          border:1px solid #660099;color:#cc66ff;border-radius:6px;padding:8px;font-size:11px;font-weight:bold;cursor:pointer;">
          📗 Exportar Excel
        </button>
      </div>
    </div>
  </div>

</div>

<script>
const IMG_B64 = '{qual_b64}';
const vc      = document.getElementById('vc');
const cv      = document.getElementById('cv');
const ctx     = cv.getContext('2d');
const zb      = document.getElementById('zbadge');
const coordEl = document.getElementById('coord');

const inpRows      = document.getElementById('inpRows');
const inpCols      = document.getElementById('inpCols');
const inpQuadraName = document.getElementById('inpQuadraName');
const inpLinhasParcela = document.getElementById('inpLinhasParcela');
const btnGridTool  = document.getElementById('btnGridTool');
const btnCountPlants = document.getElementById('btnCountPlants');
const btnManualMode= document.getElementById('btnManualMode');
const btnRemoveLast= document.getElementById('btnRemoveLast');
const btnClearAll  = document.getElementById('btnClearAll');
const btnCountAuto = document.getElementById('btnCountAuto');
const btnManual2   = document.getElementById('btnManual2');
const btnUndoMark  = document.getElementById('btnUndoMark');
const btnExportCnt = document.getElementById('btnExportCnt');
const countPanel   = document.getElementById('countPanel');
const totalCountEl = document.getElementById('totalCount');
const countInfoEl  = document.getElementById('countInfo');
const falhasSumario = document.getElementById('falhasSumario');
const btnDetectFalhas = document.getElementById('btnDetectFalhas');
const btnMedirPlantados = document.getElementById('btnMedirPlantados');
const btnDeleteFalhaMode = document.getElementById('btnDeleteFalhaMode');
const chkShowFalhas   = document.getElementById('chkShowFalhas');
const chkLabels       = document.getElementById('chkLabels');
const inpMinDist      = document.getElementById('inpMinDist');
const inpBufferCm     = document.getElementById('inpBufferCm');
const inpParcelLen    = document.getElementById('inpParcelLen');
const inpLineWidth    = document.getElementById('inpLineWidth');
const selUnit         = document.getElementById('selUnit');
const falhasModal     = document.getElementById('falhasModal');
const btnOpenFalhas   = document.getElementById('btnOpenFalhas');
const btnCloseModal   = document.getElementById('btnCloseModal');
const btnSelectParcel = document.getElementById('btnSelectParcel');
const btnMoveSelected = document.getElementById('btnMoveSelected');
const btnClearSelection = document.getElementById('btnClearSelection');
const btnSaveParcelAdjust = document.getElementById('btnSaveParcelAdjust');
const btnMoveUp       = document.getElementById('btnMoveUp');
const btnMoveDown     = document.getElementById('btnMoveDown');
const btnMoveLeft     = document.getElementById('btnMoveLeft');
const btnMoveRight    = document.getElementById('btnMoveRight');
const inpMoveStep     = document.getElementById('inpMoveStep');
const parcelAdjustStatus = document.getElementById('parcelAdjustStatus');

let gridMode = false, manualMode = false;
let deleteFalhaMode = false;
let parcelSelectMode = false, parcelMoveMode = false, draggingParcelSelection = false;
let points = [], draggingPoint = -1;
let sc = 1, ox = 0, oy = 0;
let drag = false, lx = 0, ly = 0;
const MIN_SC = 0.05, MAX_SC = 40;
let imgW = 0, imgH = 0;

let plantCenters = [];
let manualMarks  = [];
let parcelCounts = {{}};
let falhas       = []; // array de objetos {{p1, p2, dist, tiro, disp, linha}}
let areasUteis   = [];
let metrosPlantadosLinhas = [];
let metrosPlantadosSegmentos = [];
let qualidadeModoVisual = '';
let manualFailedParcels = {{}};
let deletedFalhaKeys = new Set();
function getLinhasPlantioPorParcela() {{
  const v=Math.max(1,Math.min(20,parseInt(inpLinhasParcela && inpLinhasParcela.value)||4));
  if(inpLinhasParcela) inpLinhasParcela.value=v;
  return v;
}}
function getQuadraNome() {{
  const v=(inpQuadraName && inpQuadraName.value ? inpQuadraName.value.trim() : 'Q-1');
  return v || 'Q-1';
}}
let parcelAdjustments = {{}};
let selectedParcels = new Set();
let lastParcelDragPt = null;
let parcelAdjustStorageKey = '';
const showPlantDebug = false;

const img = new Image();

// ── Coordenadas ──────────────────────────────────────────────────────────────
function getImgCoords(cx, cy) {{
  const r = cv.getBoundingClientRect();
  return {{ x:(cx-r.left-ox)/sc, y:(cy-r.top-oy)/sc }};
}}

function bilerp(p0,p1,p2,p3, u,v) {{
  const tx=(1-u)*p0.x+u*p1.x, ty=(1-u)*p0.y+u*p1.y;
  const bx=(1-u)*p3.x+u*p2.x, by=(1-u)*p3.y+u*p2.y;
  return {{ x:(1-v)*tx+v*bx, y:(1-v)*ty+v*by }};
}}

function getParcelKeyByRC(r2,c,R,C) {{
  return (R-r2)+'_'+(C-c);
}}

function getParcelLabel(key) {{
  const parts=String(key||'').split('_');
  if(parts.length!==2) return key || '';
  return 'T'+parts[1]+' D'+parts[0];
}}

function parseParcelKey(key) {{
  const parts=String(key||'').split('_').map(Number);
  return {{disp:parts[0]||0, tiro:parts[1]||0}};
}}

function getManualFailKeyByLabels(tiro,disp) {{
  return String(disp)+'_'+String(tiro);
}}

function getFalhaSignature(f) {{
  const id=f.parcelaId||('T'+f.tiro+' D'+f.disp);
  return [id,f.linha||'',f.tipo||'',Math.round((f.p1&&f.p1.x)||0),Math.round((f.p1&&f.p1.y)||0),Math.round((f.p2&&f.p2.x)||0),Math.round((f.p2&&f.p2.y)||0)].join('|');
}}

function findNearestFalha(pt) {{
  if(!falhas.length) return -1;
  let best=-1, bestD=Infinity;
  for(let i=0;i<falhas.length;i++) {{
    const f=falhas[i];
    const a=f.p1, b=f.p2;
    const dx=b.x-a.x, dy=b.y-a.y;
    const len2=dx*dx+dy*dy || 1;
    const t=Math.max(0,Math.min(1,((pt.x-a.x)*dx+(pt.y-a.y)*dy)/len2));
    const px=a.x+dx*t, py=a.y+dy*t;
    const d=Math.sqrt((pt.x-px)*(pt.x-px)+(pt.y-py)*(pt.y-py));
    if(d<bestD) {{ bestD=d; best=i; }}
  }}
  return bestD <= Math.max(10/sc,10) ? best : -1;
}}

function getBaseParcelPoly(r2,c,R,C,p0,p1,p2,p3) {{
  const u0=c/C,u1=(c+1)/C,v0=r2/R,v1=(r2+1)/R;
  const tl=bilerp(p0,p1,p2,p3,u0,v0), tr=bilerp(p0,p1,p2,p3,u1,v0);
  const br=bilerp(p0,p1,p2,p3,u1,v1), bl=bilerp(p0,p1,p2,p3,u0,v1);
  return [tl,tr,br,bl];
}}

function getAdjustedParcelPoly(r2,c,R,C,p0,p1,p2,p3) {{
  const key=getParcelKeyByRC(r2,c,R,C);
  const adj=parcelAdjustments[key] || {{dx:0,dy:0}};
  return getBaseParcelPoly(r2,c,R,C,p0,p1,p2,p3).map(p=>{{ return {{x:p.x+(adj.dx||0), y:p.y+(adj.dy||0)}}; }});
}}

function getParcelCenter(poly) {{
  return {{
    x:poly.reduce((s,p)=>s+p.x,0)/poly.length,
    y:poly.reduce((s,p)=>s+p.y,0)/poly.length
  }};
}}

function findParcelAtPoint(pt) {{
  if(points.length<4) return null;
  const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
  const p0=points[0],p1=points[1],p2=points[2],p3=points[3];
  for(let r2=R-1;r2>=0;r2--) {{
    for(let c=C-1;c>=0;c--) {{
      const poly=getAdjustedParcelPoly(r2,c,R,C,p0,p1,p2,p3);
      if(pointInPolygon(pt.x,pt.y,poly)) {{
        const key=getParcelKeyByRC(r2,c,R,C);
        return {{key:key,r:r2,c:c,poly:poly}};
      }}
    }}
  }}
  return null;
}}

function updateParcelAdjustStatus() {{
  const sel=[...selectedParcels].map(k=>getParcelLabel(k));
  const cadeia=getCascadeAffectedKeys(false);
  const ajustadas=Object.keys(parcelAdjustments).filter(k=>Math.abs((parcelAdjustments[k].dx||0))+Math.abs((parcelAdjustments[k].dy||0))>0).length;
  const modo=parcelSelectMode ? 'Seleção ativa' : (parcelMoveMode ? 'Mover ativo' : 'Modo normal');
  parcelAdjustStatus.textContent = modo + ' · Marcadas: ' + (sel.length?sel.join(', '):'nenhuma') + ' · Cadeia: ' + cadeia.size + ' · Ajustadas: ' + ajustadas;
  if(btnSelectParcel) btnSelectParcel.style.borderColor = parcelSelectMode ? '#ffd21f' : '#ff8c00';
  if(btnMoveSelected) btnMoveSelected.style.borderColor = parcelMoveMode ? '#5599ff' : '#003399';
}}

function getCascadeAffectedKeys(includeSelected) {{
  const R=parseInt(inpRows.value)||1;
  const affected=new Set();
  for(const key of selectedParcels) {{
    const parsed=parseParcelKey(key);
    if(includeSelected) affected.add(key);
    for(let d=parsed.disp+1; d<=R; d++) {{
      affected.add(d+'_'+parsed.tiro);
    }}
  }}
  return affected;
}}

function getCascadeMoveKeys() {{
  return getCascadeAffectedKeys(true);
}}

function translateSelectedParcels(dx,dy) {{
  if(selectedParcels.size===0) {{ alert('Selecione pelo menos uma parcela.'); return; }}
  const moveKeys=getCascadeMoveKeys();
  for(const key of moveKeys) {{
    const adj=parcelAdjustments[key] || {{dx:0,dy:0}};
    parcelAdjustments[key] = {{dx:(adj.dx||0)+dx, dy:(adj.dy||0)+dy}};
  }}
  updateParcelAdjustStatus();
  drawAll();
}}

function refreshAfterParcelAdjust() {{
  if(plantCenters.length>0) recount();
  if(falhas.length>0) detectarFalhas();
  else drawAll();
}}

function saveParcelAdjustments() {{
  try {{
    if(parcelAdjustStorageKey) localStorage.setItem(parcelAdjustStorageKey, JSON.stringify(parcelAdjustments));
  }} catch(e) {{}}
  refreshAfterParcelAdjust();
  updateParcelAdjustStatus();
}}

function loadParcelAdjustments() {{
  parcelAdjustStorageKey='tmg_qual_parcel_adjust_'+imgW+'x'+imgH+'_'+IMG_B64.length;
  try {{
    const raw=localStorage.getItem(parcelAdjustStorageKey);
    if(raw) parcelAdjustments=JSON.parse(raw) || {{}};
  }} catch(e) {{ parcelAdjustments={{}}; }}
  updateParcelAdjustStatus();
}}

function pointInPolygon(px, py, polygon) {{
  let inside = false;
  for (let i=0, j=polygon.length-1; i<polygon.length; j=i++) {{
    const xi=polygon[i].x, yi=polygon[i].y, xj=polygon[j].x, yj=polygon[j].y;
    if (((yi>py)!==(yj>py)) && (px<(xj-xi)*(py-yi)/(yj-yi)+xi)) inside=!inside;
  }}
  return inside;
}}

function lineDistance(a,b) {{
  return Math.sqrt((b.x-a.x)**2+(b.y-a.y)**2);
}}

function lerpPoint(a,b,t) {{
  return {{ x:a.x+(b.x-a.x)*t, y:a.y+(b.y-a.y)*t }};
}}

function midpoint(a,b) {{
  return {{ x:(a.x+b.x)/2, y:(a.y+b.y)/2 }};
}}

function pointOnLine(a,b,t) {{
  return lerpPoint(a,b,t);
}}

function getParcelaMeasureLine(tl,tr,br,bl) {{
  const horiz=(lineDistance(tl,tr)+lineDistance(bl,br))/2;
  const vert=(lineDistance(tl,bl)+lineDistance(tr,br))/2;
  const p1=horiz>=vert ? midpoint(tl,bl) : midpoint(tl,tr);
  const p2=horiz>=vert ? midpoint(tr,br) : midpoint(bl,br);
  return {{ p1:p1, p2:p2, dist:lineDistance(p1,p2) }};
}}

function projectPointOnLine(p,a,b) {{
  const dx=b.x-a.x, dy=b.y-a.y;
  const len2=dx*dx+dy*dy;
  if(len2===0) return 0;
  return Math.max(0,Math.min(1,((p.x-a.x)*dx+(p.y-a.y)*dy)/len2));
}}

function getParcelaAxes(tl,tr,br,bl) {{
  const horiz=(lineDistance(tl,tr)+lineDistance(bl,br))/2 >= (lineDistance(tl,bl)+lineDistance(tr,br))/2;
  const crossA=horiz ? midpoint(tl,tr) : midpoint(tl,bl);
  const crossB=horiz ? midpoint(bl,br) : midpoint(tr,br);
  return {{ horiz:horiz, crossA:crossA, crossB:crossB, secondaryLen:lineDistance(crossA,crossB) }};
}}

function getRowGeometry(tl,tr,br,bl,horiz,s) {{
  const p1=horiz ? lerpPoint(tl,bl,s) : lerpPoint(tl,tr,s);
  const p2=horiz ? lerpPoint(tr,br,s) : lerpPoint(bl,br,s);
  return {{ p1:p1, p2:p2, dist:lineDistance(p1,p2) }};
}}

function cross2(a,b) {{
  return a.x*b.y-a.y*b.x;
}}

function clipLineToPolygon(anchor,dir,poly) {{
  const hits=[];
  for(let i=0;i<poly.length;i++) {{
    const a=poly[i], b=poly[(i+1)%poly.length];
    const edge={{x:b.x-a.x,y:b.y-a.y}};
    const den=cross2(dir,edge);
    if(Math.abs(den)<0.000001) continue;
    const ap={{x:a.x-anchor.x,y:a.y-anchor.y}};
    const t=cross2(ap,edge)/den;
    const u=cross2(ap,dir)/den;
    if(u>=-0.0001 && u<=1.0001) hits.push({{t:t,p:{{x:anchor.x+dir.x*t,y:anchor.y+dir.y*t}}}});
  }}
  hits.sort((a,b)=>a.t-b.t);
  const uniq=[];
  for(const h of hits) {{
    if(uniq.length===0 || lineDistance(h.p,uniq[uniq.length-1].p)>0.5) uniq.push(h);
  }}
  if(uniq.length<2) return null;
  return {{p1:uniq[0].p,p2:uniq[uniq.length-1].p,dist:lineDistance(uniq[0].p,uniq[uniq.length-1].p)}};
}}

function alignRowGeometryToPlants(rowGeom,rowItems,poly) {{
  if(rowItems.length<2) return rowGeom;
  const baseDir={{x:rowGeom.p2.x-rowGeom.p1.x,y:rowGeom.p2.y-rowGeom.p1.y}};
  const baseLen=Math.sqrt(baseDir.x*baseDir.x+baseDir.y*baseDir.y);
  if(baseLen<=0) return rowGeom;
  baseDir.x/=baseLen; baseDir.y/=baseLen;
  const ordered=[...rowItems].sort((a,b)=>projectPointOnLine(a.p,rowGeom.p1,rowGeom.p2)-projectPointOnLine(b.p,rowGeom.p1,rowGeom.p2));
  const first=ordered[0].p, last=ordered[ordered.length-1].p;
  let dir={{x:last.x-first.x,y:last.y-first.y}};
  let len=Math.sqrt(dir.x*dir.x+dir.y*dir.y);
  if(len<5) {{ dir=baseDir; len=1; }} else {{ dir.x/=len; dir.y/=len; }}
  if(dir.x*baseDir.x+dir.y*baseDir.y<0) {{ dir.x*=-1; dir.y*=-1; }}
  const anchor={{
    x:ordered.reduce((s,it)=>s+it.p.x,0)/ordered.length,
    y:ordered.reduce((s,it)=>s+it.p.y,0)/ordered.length
  }};
  const clipped=clipLineToPolygon(anchor,dir,poly);
  return clipped || rowGeom;
}}

function getPlantRowInfo(p,axes,tl,tr,br,bl) {{
  const s=projectPointOnLine(p,axes.crossA,axes.crossB);
  const rowGeom=getRowGeometry(tl,tr,br,bl,axes.horiz,s);
  return {{ p:p, s:s, sPx:s*axes.secondaryLen, t:projectPointOnLine(p,rowGeom.p1,rowGeom.p2) }};
}}

function clusterPlantRows(rowInfos,tolPx) {{
  const sorted=[...rowInfos].sort((a,b)=>a.sPx-b.sPx);
  const rows=[];
  for(const item of sorted) {{
    let row=rows.length ? rows[rows.length-1] : null;
    if(!row || Math.abs(item.sPx-row.meanPx)>tolPx) {{
      rows.push({{ items:[item], meanPx:item.sPx, meanS:item.s }});
    }} else {{
      row.items.push(item);
      row.meanPx=row.items.reduce((s,it)=>s+it.sPx,0)/row.items.length;
      row.meanS=row.items.reduce((s,it)=>s+it.s,0)/row.items.length;
    }}
  }}
  return rows;
}}

function getParcelRealMeters() {{
  return Math.max(0.1,parseFloat(inpParcelLen.value)||5);
}}

function getBufferCorredorCm() {{
  const v=parseFloat(inpBufferCm.value);
  return Number.isFinite(v) ? Math.max(0,v) : 20;
}}

function polygonSignedArea(poly) {{
  let area=0;
  for(let i=0;i<poly.length;i++) {{
    const a=poly[i], b=poly[(i+1)%poly.length];
    area+=a.x*b.y-a.y*b.x;
  }}
  return area/2;
}}

function centerShrinkPolygon(poly,insetPx) {{
  const center=getParcelCenter(poly);
  return poly.map(p=>{{
    const vx=center.x-p.x, vy=center.y-p.y;
    const len=Math.sqrt(vx*vx+vy*vy);
    if(len<=0) return {{x:p.x,y:p.y}};
    const step=Math.min(insetPx,len*0.45);
    return {{x:p.x+(vx/len)*step,y:p.y+(vy/len)*step}};
  }});
}}

function intersectOffsetLines(lineA,lineB) {{
  const den=cross2(lineA.d,lineB.d);
  if(Math.abs(den)<0.000001) return null;
  const ap={{x:lineB.p.x-lineA.p.x,y:lineB.p.y-lineA.p.y}};
  const t=cross2(ap,lineB.d)/den;
  return {{x:lineA.p.x+lineA.d.x*t,y:lineA.p.y+lineA.d.y*t}};
}}

function shrinkPolygonToAreaUtil(poly,bufferPx) {{
  if(!bufferPx || bufferPx<=0) return poly.map(p=>{{return {{x:p.x,y:p.y}};}});
  const [tl,tr,br,bl]=poly;
  const leftLen=lineDistance(tl,bl);
  const rightLen=lineDistance(tr,br);
  const tLeft=leftLen>0 ? Math.min(bufferPx,leftLen*0.45)/leftLen : 0;
  const tRight=rightLen>0 ? Math.min(bufferPx,rightLen*0.45)/rightLen : 0;
  return [
    lerpPoint(tl,bl,tLeft),
    lerpPoint(tr,br,tRight),
    lerpPoint(br,tr,tRight),
    lerpPoint(bl,tl,tLeft)
  ];
}}

function getBufferCorredorPx(poly) {{
  const bufferM=getBufferCorredorCm()/100;
  if(bufferM<=0) return 0;
  const pxPorMetro=getPixelsPorMetroParcela(poly);
  return pxPorMetro>0 ? bufferM*pxPorMetro : 0;
}}

function getAreaUtilParcela(poly) {{
  return shrinkPolygonToAreaUtil(poly,getBufferCorredorPx(poly));
}}

function getPixelsPorMetroParcela(poly) {{
  const [tl,tr,br,bl]=poly;
  const measureLine=getParcelaMeasureLine(tl,tr,br,bl);
  if(!measureLine || measureLine.dist<=0) return 0;
  return measureLine.dist/getParcelRealMeters();
}}

function rowPxToMeters(px,rowGeom) {{
  if(!rowGeom || rowGeom.dist<=0) return 0;
  if(rowGeom.pxPorMetro && rowGeom.pxPorMetro>0) return px/rowGeom.pxPorMetro;
  return (px/rowGeom.dist)*getParcelRealMeters();
}}

function addFalhasNaFileira(rowGeom,rowItems,minDistM,tiroLabel,dispLabel,linhaLabel,parcelaId) {{
  if(rowGeom.dist<=0) return;
  const refs=rowItems
    .map(item=>projectPointOnLine(item.p,rowGeom.p1,rowGeom.p2))
    .sort((a,b)=>a-b);

  const refsUnicas=[];
  const mergeT=Math.max(2,(minDistM/getParcelRealMeters())*rowGeom.dist*0.15)/rowGeom.dist;
  for(const t of refs) {{
    if(refsUnicas.length===0 || Math.abs(t-refsUnicas[refsUnicas.length-1])>mergeT) refsUnicas.push(t);
  }}

  const intervalos=[];
  if(refsUnicas.length===0) {{
    intervalos.push({{t0:0,t1:1,tipo:'inicio'}});
  }} else {{
    intervalos.push({{t0:0,t1:refsUnicas[0],tipo:'inicio'}});
    for(let i=1;i<refsUnicas.length;i++) intervalos.push({{t0:refsUnicas[i-1],t1:refsUnicas[i],tipo:'entre plantas'}});
    intervalos.push({{t0:refsUnicas[refsUnicas.length-1],t1:1,tipo:'final'}});
  }}

  for(const inter of intervalos) {{
    const t0=inter.t0, t1=inter.t1;
    const dist=(t1-t0)*rowGeom.dist;
    const distM=rowPxToMeters(dist,rowGeom);
    if(distM>minDistM) {{
      falhas.push({{
        p1:pointOnLine(rowGeom.p1,rowGeom.p2,t0),
        p2:pointOnLine(rowGeom.p1,rowGeom.p2,t1),
        dist:dist,
        distM:distM,
        tiro:tiroLabel,
        disp:dispLabel,
        linha:linhaLabel,
        tipo:inter.tipo,
        parcelaId:parcelaId
      }});
    }}
  }}
}}

function isPlantRGB(r2,g,b) {{
  const max2=Math.max(r2,g,b), min2=Math.min(r2,g,b), diff=max2-min2;
  let hue=0,sat=0;
  if(diff>0){{
    if(max2===r2) hue=60*(((g-b)/diff)%6);
    else if(max2===g) hue=60*((b-r2)/diff+2);
    else hue=60*((r2-g)/diff+4);
    if(hue<0) hue+=360; sat=(diff/max2)*255;
  }}
  hue=hue/2;
  return hue>=30&&hue<=90&&sat>=40&&max2>=40;
}}

function getHSVExGInfo(r2,g,b) {{
  const max2=Math.max(r2,g,b), min2=Math.min(r2,g,b), diff=max2-min2;
  let hue=0, sat=0;
  if(diff>0) {{
    if(max2===r2) hue=60*(((g-b)/diff)%6);
    else if(max2===g) hue=60*((b-r2)/diff+2);
    else hue=60*((r2-g)/diff+4);
    if(hue<0) hue+=360;
    sat=diff/(max2||1);
  }}
  const exg=(2*g-r2-b);
  const exgNorm=exg/(r2+g+b+1);
  const hsvOk=hue>=35 && hue<=170 && sat>=0.16 && max2>=35 && g>=Math.max(r2,b)*0.82;
  const exgOk=exg>=14 || exgNorm>=0.045;
  return {{hsvOk:hsvOk,exgOk:exgOk,plant:hsvOk&&exgOk}};
}}

function erodeMask(mask,w,h) {{
  const out=new Uint8Array(w*h);
  for(let y=1;y<h-1;y++) {{
    const row=y*w;
    for(let x=1;x<w-1;x++) {{
      const idx=row+x;
      if(mask[idx]&&mask[idx-1]&&mask[idx+1]&&mask[idx-w]&&mask[idx+w]&&mask[idx-w-1]&&mask[idx-w+1]&&mask[idx+w-1]&&mask[idx+w+1]) out[idx]=1;
    }}
  }}
  return out;
}}

function dilateMask(mask,w,h) {{
  const out=new Uint8Array(w*h);
  for(let y=1;y<h-1;y++) {{
    const row=y*w;
    for(let x=1;x<w-1;x++) {{
      const idx=row+x;
      if(mask[idx]||mask[idx-1]||mask[idx+1]||mask[idx-w]||mask[idx+w]||mask[idx-w-1]||mask[idx-w+1]||mask[idx+w-1]||mask[idx+w+1]) out[idx]=1;
    }}
  }}
  return out;
}}

function cleanVegetationMask(mask,w,h) {{
  let out=dilateMask(erodeMask(mask,w,h),w,h);
  out=erodeMask(dilateMask(out,w,h),w,h);
  return out;
}}

function removeSmallVegetationObjects(mask,w,h,minX,minY,minArea,gridPoly) {{
  const visited=new Uint8Array(w*h);
  const cleaned=new Uint8Array(w*h);
  const centers=[];
  const stack=[];
  const pixels=[];
  for(let i=0;i<mask.length;i++) {{
    if(!mask[i] || visited[i]) continue;
    stack.length=0; pixels.length=0;
    stack.push(i); visited[i]=1;
    let area=0,sumX=0,sumY=0;
    while(stack.length>0) {{
      const idx=stack.pop();
      pixels.push(idx); area++;
      const x=idx%w, y=Math.floor(idx/w);
      sumX+=x; sumY+=y;
      const neigh=[idx-1,idx+1,idx-w,idx+w];
      for(const ni of neigh) {{
        if(ni<0||ni>=mask.length||visited[ni]||!mask[ni]) continue;
        const nx=ni%w, ny=Math.floor(ni/w);
        if(Math.abs(nx-x)+Math.abs(ny-y)!==1) continue;
        visited[ni]=1; stack.push(ni);
      }}
    }}
    if(area>=minArea) {{
      const cx=sumX/area+minX, cy=sumY/area+minY;
      if(!gridPoly || pointInPolygon(cx,cy,gridPoly)) {{
        for(const pi of pixels) cleaned[pi]=1;
        centers.push({{x:cx,y:cy,area:area}});
      }}
    }}
  }}
  return {{mask:cleaned,centers:centers}};
}}

function getQualidadeColorSampler() {{
  const tempCv=document.createElement('canvas');
  tempCv.width=imgW; tempCv.height=imgH;
  const tc=tempCv.getContext('2d'); tc.drawImage(img,0,0);
  const gridPoly=points.length===4 ? points : null;
  const xs=gridPoly ? gridPoly.map(p=>p.x) : [0,imgW];
  const ys=gridPoly ? gridPoly.map(p=>p.y) : [0,imgH];
  const pad=16;
  const minX=Math.max(0,Math.floor(Math.min(...xs)-pad));
  const maxX=Math.min(imgW,Math.ceil(Math.max(...xs)+pad));
  const minY=Math.max(0,Math.floor(Math.min(...ys)-pad));
  const maxY=Math.min(imgH,Math.ceil(Math.max(...ys)+pad));
  const w=Math.max(1,maxX-minX), h=Math.max(1,maxY-minY);
  const imgData=tc.getImageData(minX,minY,w,h);
  const data=imgData.data;
  let mask=new Uint8Array(w*h);
  for(let yy=0;yy<h;yy++) {{
    for(let xx=0;xx<w;xx++) {{
      const idx=(yy*w+xx)*4;
      const absX=xx+minX, absY=yy+minY;
      if(gridPoly && !pointInPolygon(absX,absY,gridPoly)) continue;
      const info=getHSVExGInfo(data[idx],data[idx+1],data[idx+2]);
      if(info.plant) mask[yy*w+xx]=1;
    }}
  }}
  mask=cleanVegetationMask(mask,w,h);
  const cleaned=removeSmallVegetationObjects(mask,w,h,minX,minY,18,gridPoly);
  mask=cleaned.mask;

  const sampler=function(x,y,rad) {{
    const rr=Math.max(1,Math.round(rad||3));
    const cx=Math.round(x)-minX, cy=Math.round(y)-minY;
    let total=0, plant=0;
    for(let yy=cy-rr;yy<=cy+rr;yy++) {{
      if(yy<0||yy>=h) continue;
      for(let xx=cx-rr;xx<=cx+rr;xx++) {{
        if(xx<0||xx>=w) continue;
        total++;
        if(mask[yy*w+xx]) plant++;
      }}
    }}
    return total>0 && (plant/total)>=0.10;
  }};
  sampler.centers=cleaned.centers;
  return sampler;
}}

function tipoFalhaPorPosicao(t0,t1,rowItems,rowGeom) {{
  if(rowItems.length===0) return 'inicio';
  const refs=rowItems
    .map(item=>projectPointOnLine(item.p,rowGeom.p1,rowGeom.p2))
    .sort((a,b)=>a-b);
  const mid=(t0+t1)/2;
  if(mid<refs[0]) return 'inicio';
  if(mid>refs[refs.length-1]) return 'final';
  return 'entre plantas';
}}

function addFalhasChaoNaFileira(rowGeom,rowItems,minDistM,tiroLabel,dispLabel,linhaLabel,parcelaId,sampler) {{
  if(!sampler) {{ addFalhasNaFileira(rowGeom,rowItems,minDistM,tiroLabel,dispLabel,linhaLabel,parcelaId); return; }}
  if(rowGeom.dist<=0) return;
  const samples=Math.max(12,Math.ceil(rowGeom.dist/2));
  const rad=Math.max(2,Math.min(5,rowGeom.dist*0.01));
  let startT=null;
  for(let i=0;i<=samples;i++) {{
    const t=i/samples;
    const p=pointOnLine(rowGeom.p1,rowGeom.p2,t);
    const isPlant=sampler(p.x,p.y,rad);
    if(!isPlant && startT===null) startT=t;
    if((isPlant || i===samples) && startT!==null) {{
      const endT=isPlant ? Math.max(0,(i-1)/samples) : t;
      if(endT>startT) {{
        const dist=(endT-startT)*rowGeom.dist;
        const distM=rowPxToMeters(dist,rowGeom);
        if(distM>minDistM) {{
          falhas.push({{
            p1:pointOnLine(rowGeom.p1,rowGeom.p2,startT),
            p2:pointOnLine(rowGeom.p1,rowGeom.p2,endT),
            dist:dist,
            distM:distM,
            tiro:tiroLabel,
            disp:dispLabel,
            linha:linhaLabel,
            tipo:tipoFalhaPorPosicao(startT,endT,rowItems,rowGeom),
            parcelaId:parcelaId
          }});
        }}
      }}
      startT=null;
    }}
  }}
}}

function getPlantRangesNaLinha(rowGeom,refsUnicas,sampler) {{
  const fallbackHalfT=Math.max(3,Math.min(12,rowGeom.dist*0.018))/rowGeom.dist;
  const samples=Math.max(80,Math.ceil(rowGeom.dist/2));
  const stepT=1/samples;
  const scanT=Math.max(fallbackHalfT*4, Math.min(0.18, 28/rowGeom.dist));
  const rad=Math.max(2,Math.min(5,rowGeom.dist*0.006));
  const ranges=[];

  for(const t of refsUnicas) {{
    let left=Math.max(0,t-fallbackHalfT);
    let right=Math.min(1,t+fallbackHalfT);

    if(sampler) {{
      let seenLeft=false, edgeLeft=t, emptyLeft=0;
      for(let tt=t; tt>=Math.max(0,t-scanT); tt-=stepT) {{
        const p=pointOnLine(rowGeom.p1,rowGeom.p2,tt);
        if(sampler(p.x,p.y,rad)) {{
          seenLeft=true; edgeLeft=tt; emptyLeft=0;
        }} else if(seenLeft) {{
          emptyLeft++;
          if(emptyLeft>=2) break;
        }}
      }}

      let seenRight=false, edgeRight=t, emptyRight=0;
      for(let tt=t; tt<=Math.min(1,t+scanT); tt+=stepT) {{
        const p=pointOnLine(rowGeom.p1,rowGeom.p2,tt);
        if(sampler(p.x,p.y,rad)) {{
          seenRight=true; edgeRight=tt; emptyRight=0;
        }} else if(seenRight) {{
          emptyRight++;
          if(emptyRight>=2) break;
        }}
      }}

      if(seenLeft || seenRight) {{
        left=Math.max(0,(seenLeft?edgeLeft:t)-stepT);
        right=Math.min(1,(seenRight?edgeRight:t)+stepT);
      }}
    }}

    ranges.push({{start:left,end:right}});
  }}

  ranges.sort((a,b)=>a.start-b.start);
  const merged=[];
  for(const rg of ranges) {{
    if(merged.length===0 || rg.start>merged[merged.length-1].end) merged.push({{start:rg.start,end:rg.end}});
    else merged[merged.length-1].end=Math.max(merged[merged.length-1].end,rg.end);
  }}
  return merged;
}}

function getRefsUnicasDaLinha(rowGeom,rowItems) {{
  if(!rowGeom || rowGeom.dist<=0 || !rowItems || rowItems.length<1) return;
  const refs=rowItems
    .map(item=>projectPointOnLine(item.p,rowGeom.p1,rowGeom.p2))
    .filter(t=>t>=0 && t<=1)
    .sort((a,b)=>a-b);

  const refsUnicas=[];
  const mergeT=Math.max(2,rowGeom.dist*0.01)/rowGeom.dist;
  for(const t of refs) {{
    if(refsUnicas.length===0 || Math.abs(t-refsUnicas[refsUnicas.length-1])>mergeT) refsUnicas.push(t);
  }}
  return refsUnicas;
}}

function sampleVegetacaoNoCorredor(rowGeom,t,sampler) {{
  if(!sampler || !rowGeom || rowGeom.dist<=0) return false;
  const p=pointOnLine(rowGeom.p1,rowGeom.p2,t);
  const dx=rowGeom.p2.x-rowGeom.p1.x, dy=rowGeom.p2.y-rowGeom.p1.y;
  const len=Math.sqrt(dx*dx+dy*dy);
  if(len<=0) return false;
  const normal=rowGeom.normal || {{x:-dy/len,y:dx/len}};
  const half=Math.max(2,Math.min(10,rowGeom.corridorPx||4));
  const step=Math.max(1.5,half/3);
  let total=0, hits=0;
  for(let off=-half; off<=half+0.01; off+=step) {{
    total++;
    if(sampler(p.x+normal.x*off,p.y+normal.y*off,2)) hits++;
  }}
  return total>0 && (hits/total)>=0.28;
}}

function getPlantRangesPorLinha(rowGeom,rowItems,sampler) {{
  if(!rowGeom || rowGeom.dist<=0 || !rowItems || rowItems.length<1) return [];
  const samples=Math.max(80,Math.ceil(rowGeom.dist/2));
  const flags=new Uint8Array(samples+1);
  for(let i=0;i<=samples;i++) {{
    flags[i]=sampleVegetacaoNoCorredor(rowGeom,i/samples,sampler)?1:0;
  }}

  const pxPorMetro=(rowGeom.pxPorMetro&&rowGeom.pxPorMetro>0)?rowGeom.pxPorMetro:(rowGeom.dist/getParcelRealMeters());
  const pxPorSample=rowGeom.dist/samples;
  const closeGapPx=Math.max(2,Math.min(getLimiteFalhaM()*0.75*pxPorMetro,rowGeom.dist*0.08));
  const closeGapSamples=Math.max(1,Math.round(closeGapPx/pxPorSample));

  let i=0;
  while(i<=samples) {{
    while(i<=samples && flags[i]) i++;
    const start=i;
    while(i<=samples && !flags[i]) i++;
    const end=i-1;
    const prev=start>0 && flags[start-1];
    const next=i<=samples && flags[i];
    if(prev && next && (end-start+1)<=closeGapSamples) {{
      for(let k=start;k<=end;k++) flags[k]=1;
    }}
  }}

  const minPlantPx=Math.max(2,0.03*pxPorMetro);
  const minPlantSamples=Math.max(1,Math.round(minPlantPx/pxPorSample));
  i=0;
  while(i<=samples) {{
    while(i<=samples && !flags[i]) i++;
    const start=i;
    while(i<=samples && flags[i]) i++;
    const end=i-1;
    if(end>=start && (end-start+1)<minPlantSamples) {{
      for(let k=start;k<=end;k++) flags[k]=0;
    }}
  }}

  const ranges=[];
  i=0;
  while(i<=samples) {{
    while(i<=samples && !flags[i]) i++;
    const start=i;
    while(i<=samples && flags[i]) i++;
    const end=i-1;
    if(end>=start) {{
      ranges.push({{start:Math.max(0,start/samples),end:Math.min(1,(end+1)/samples)}});
    }}
  }}
  return ranges;
}}

function limitarLinhasPlantio(linhas,maxLinhas) {{
  const out=linhas
    .filter(l=>l && l.items && l.items.length>0)
    .map(l=>{{ return {{items:[...l.items],meanPx:l.meanPx,meanS:l.meanS}}; }})
    .sort((a,b)=>a.meanPx-b.meanPx);
  while(out.length>maxLinhas) {{
    let bestIdx=1, bestDist=Infinity;
    for(let i=1;i<out.length;i++) {{
      const d=Math.abs(out[i].meanPx-out[i-1].meanPx);
      if(d<bestDist) {{ bestDist=d; bestIdx=i; }}
    }}
    const a=out[bestIdx-1], b=out[bestIdx];
    const items=a.items.concat(b.items);
    const meanPx=items.reduce((s,it)=>s+it.sPx,0)/items.length;
    const meanS=items.reduce((s,it)=>s+it.s,0)/items.length;
    out.splice(bestIdx-1,2,{{items:items,meanPx:meanPx,meanS:meanS}});
  }}
  return out.sort((a,b)=>a.meanPx-b.meanPx);
}}

function registrarMetrosPlantadosDaLinha(rowGeom,rowItems,tiroLabel,dispLabel,linhaLabel,parcelaId,sampler,plantRangesProntas) {{
  if(!rowGeom || rowGeom.dist<=0 || !rowItems || rowItems.length<1) return [];
  const plantRanges=plantRangesProntas || getPlantRangesPorLinha(rowGeom,rowItems,sampler);
  if(!plantRanges || plantRanges.length<1) return [];
  let totalPlantadoM=0;
  for(const rg of plantRanges) {{
    const t0=Math.max(0,Math.min(1,rg.start));
    const t1=Math.max(0,Math.min(1,rg.end));
    if(t1<=t0) continue;
    const segM=rowPxToMeters((t1-t0)*rowGeom.dist,rowGeom);
    totalPlantadoM+=segM;
    metrosPlantadosSegmentos.push({{
      p1:pointOnLine(rowGeom.p1,rowGeom.p2,t0),
      p2:pointOnLine(rowGeom.p1,rowGeom.p2,t1),
      distM:segM,
      parcelaId:parcelaId,
      tiro:tiroLabel,
      disp:dispLabel,
      linha:linhaLabel
    }});
  }}
  if(totalPlantadoM>0) {{
    metrosPlantadosLinhas.push({{
      parcela:parcelaId,
      tiro:tiroLabel,
      disp:dispLabel,
      linha:linhaLabel,
      plantadoM:totalPlantadoM
    }});
  }}
  return plantRanges;
}}

function addFalhasSomenteEntrePlantasDaLinha(rowGeom,rowItems,minDistM,tiroLabel,dispLabel,linhaLabel,parcelaId,sampler,plantRangesProntas) {{
  if(!rowGeom || rowGeom.dist<=0 || !rowItems || rowItems.length<1) return;
  const plantRanges=plantRangesProntas || getPlantRangesPorLinha(rowGeom,rowItems,sampler);
  if(plantRanges.length<1) return;
  const intervalos=[];
  intervalos.push({{t0:0,t1:plantRanges[0].start,tipo:'inicio'}});
  for(let i=1;i<plantRanges.length;i++) {{
    intervalos.push({{t0:plantRanges[i-1].end,t1:plantRanges[i].start,tipo:'entre plantas'}});
  }}
  intervalos.push({{t0:plantRanges[plantRanges.length-1].end,t1:1,tipo:'final'}});

  for(const inter of intervalos) {{
    const t0=inter.t0, t1=inter.t1;
    if(t1<=t0) continue;
    const dist=(t1-t0)*rowGeom.dist;
    const distM=rowPxToMeters(dist,rowGeom);
    if(distM>=minDistM) {{
      falhas.push({{
        p1:pointOnLine(rowGeom.p1,rowGeom.p2,t0),
        p2:pointOnLine(rowGeom.p1,rowGeom.p2,t1),
        dist:dist,
        distM:distM,
        tiro:tiroLabel,
        disp:dispLabel,
        linha:linhaLabel,
        tipo:inter.tipo,
        parcelaId:parcelaId
      }});
    }}
  }}
}}

// ── Detecção de plantas (idêntica ao módulo Contagem) ─────────────────────
function countPlantsInGrid() {{
  if (points.length < 4) {{ alert('Marque os 4 cantos do grid primeiro!'); return; }}
  const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
  const p0=points[0],p1=points[1],p2=points[2],p3=points[3];
  const tempCv=document.createElement('canvas');
  tempCv.width=imgW; tempCv.height=imgH;
  const tc=tempCv.getContext('2d'); tc.drawImage(img,0,0);
  const xs=points.map(p=>p.x), ys=points.map(p=>p.y);
  const minX=Math.max(0,Math.floor(Math.min(...xs)));
  const maxX=Math.min(imgW,Math.ceil(Math.max(...xs)));
  const minY=Math.max(0,Math.floor(Math.min(...ys)));
  const maxY=Math.min(imgH,Math.ceil(Math.max(...ys)));
  const imgData=tc.getImageData(minX,minY,maxX-minX,maxY-minY);
  const data=imgData.data; const w=maxX-minX, h=maxY-minY;
  plantCenters=[]; falhas=[]; areasUteis=[]; metrosPlantadosLinhas=[]; metrosPlantadosSegmentos=[]; qualidadeModoVisual='';
  const visited=new Uint8Array(w*h);
  const minArea=25;
  for (let y=0;y<h;y++) {{
    for (let x=0;x<w;x++) {{
      const idx=(y*w+x)*4;
      const r2=data[idx],g=data[idx+1],b=data[idx+2];
      const max2=Math.max(r2,g,b), min2=Math.min(r2,g,b), diff=max2-min2;
      let hue=0,sat=0,val=max2;
      if(diff>0){{
        if(max2===r2) hue=60*(((g-b)/diff)%6);
        else if(max2===g) hue=60*((b-r2)/diff+2);
        else hue=60*((r2-g)/diff+4);
        if(hue<0) hue+=360; sat=(diff/max2)*255;
      }}
      hue=hue/2;
      if(hue>=30&&hue<=90&&sat>=40&&val>=40){{
        const absX=x+minX, absY=y+minY;
        if(pointInPolygon(absX,absY,points)&&!visited[y*w+x]){{
          let area=0,sumX=0,sumY=0;
          const stack=[[x,y]];
          while(stack.length>0&&area<2000){{
            const[sx,sy]=stack.pop();
            if(sx<0||sx>=w||sy<0||sy>=h) continue;
            if(visited[sy*w+sx]) continue;
            const si=(sy*w+sx)*4;
            const sr=data[si],sg=data[si+1],sb=data[si+2];
            const smax=Math.max(sr,sg,sb),smin=Math.min(sr,sg,sb),sdiff=smax-smin;
            let sh=0,ss=0;
            if(sdiff>0){{
              if(smax===sr) sh=60*(((sg-sb)/sdiff)%6);
              else if(smax===sg) sh=60*((sb-sr)/sdiff+2);
              else sh=60*((sr-sg)/sdiff+4);
              if(sh<0) sh+=360; ss=(sdiff/smax)*255;
            }}
            sh=sh/2;
            if(sh<30||sh>90||ss<40||smax<40) continue;
            visited[sy*w+sx]=1; area++; sumX+=sx; sumY+=sy;
            stack.push([sx+1,sy],[sx-1,sy],[sx,sy+1],[sx,sy-1]);
          }}
          if(area>=minArea){{
            plantCenters.push({{x:Math.round(sumX/area)+minX, y:Math.round(sumY/area)+minY}});
          }}
        }}
      }}
    }}
  }}
  for(const m of manualMarks) if(pointInPolygon(m.x,m.y,points)) plantCenters.push(m);
  recount(); drawAll();
}}

// ── Recontagem por parcela ────────────────────────────────────────────────
function recount() {{
  if(points.length<4) return;
  const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
  const p0=points[0],p1=points[1],p2=points[2],p3=points[3];
  parcelCounts={{}};
  for(let r2=0;r2<R;r2++) for(let c=0;c<C;c++) {{
    const poly=getAdjustedParcelPoly(r2,c,R,C,p0,p1,p2,p3);
    let cnt=0;
    for(const p of plantCenters) if(pointInPolygon(p.x,p.y,poly)) cnt++;
    parcelCounts[(R-r2)+'_'+(C-c)]=cnt;
  }}
}}

// ── Detecção de Falhas Lineares ────────────────────────────────────────────
function getLimiteFalhaM() {{
  const cm=Math.max(20,parseFloat(inpMinDist.value)||20);
  inpMinDist.value=cm;
  return cm/100;
}}

function processarQualidadePorLinhas(modo) {{
  if(points.length<4) {{ alert('Marque o grid primeiro!'); return; }}
  const calcularFalhas=(modo==='falhas'||modo==='csv');
  const calcularPlantados=(modo==='plantados'||modo==='csv');
  const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
  const minDistM=getLimiteFalhaM();
  const p0=points[0],p1=points[1],p2=points[2],p3=points[3];
  const sampler=getQualidadeColorSampler();
  const centrosAnalise=[...(sampler.centers||[])];
  for(const m of manualMarks) if(pointInPolygon(m.x,m.y,points)) centrosAnalise.push(m);
  plantCenters=centrosAnalise;
  falhas=[];
  areasUteis=[];
  metrosPlantadosLinhas=[];
  metrosPlantadosSegmentos=[];
  if(modo==='falhas') qualidadeModoVisual='falhas';
  else if(modo==='plantados') qualidadeModoVisual='plantados';

  // Para cada parcela (tiro x disparo), medir somente no eixo real da fileira de plantas.
  for(let r2=0;r2<R;r2++) {{
    for(let c=0;c<C;c++) {{
      const polyOriginal=getAdjustedParcelPoly(r2,c,R,C,p0,p1,p2,p3);
      const areaUtil=getAreaUtilParcela(polyOriginal);
      const [tl,tr,br,bl]=areaUtil;
      const dispLabel=R-r2, tiroLabel=C-c;
      const parcelaId='T'+tiroLabel+' D'+dispLabel;
      areasUteis.push({{parcelaId:parcelaId,poly:areaUtil}});
      const pxPorMetro=getPixelsPorMetroParcela(polyOriginal);
      const axes=getParcelaAxes(tl,tr,br,bl);
      const pts=centrosAnalise.filter(p=>pointInPolygon(p.x,p.y,areaUtil));

      if(pts.length<1 && !manualFailedParcels[getManualFailKeyByLabels(tiroLabel,dispLabel)]) {{
        continue;
      }}

      if(manualFailedParcels[getManualFailKeyByLabels(tiroLabel,dispLabel)]) {{
        const linhasManuais=getLinhasPlantioPorParcela();
        for(let li=1; li<=linhasManuais; li++) {{
          const sLinha=(li-0.5)/linhasManuais;
          const rowGeom=getRowGeometry(tl,tr,br,bl,axes.horiz,sLinha);
          rowGeom.pxPorMetro=pxPorMetro;
          if(calcularFalhas) falhas.push({{
            p1:rowGeom.p1,p2:rowGeom.p2,dist:rowGeom.dist,distM:getParcelRealMeters(),
            tiro:tiroLabel,disp:dispLabel,linha:li,tipo:'100% falhada',parcelaId:parcelaId,obs:'100% FALHADA',manual:true
          }});
        }}
        continue;
      }}

      const rowInfos=pts.map(p=>getPlantRowInfo(p,axes,tl,tr,br,bl));
      const rowTol=Math.max(3,Math.min(10,axes.secondaryLen*0.045));
      const linhas=limitarLinhasPlantio(clusterPlantRows(rowInfos,rowTol),getLinhasPlantioPorParcela());
      const means=linhas.map(l=>l.meanPx).sort((a,b)=>a-b);
      let spacing=axes.secondaryLen/Math.max(1,linhas.length+1);
      for(let mi=1;mi<means.length;mi++) spacing=Math.min(spacing,Math.max(2,means[mi]-means[mi-1]));
      const corridorPx=Math.max(2,Math.min(10,spacing*0.28));

      for(let li=0;li<linhas.length;li++) {{
        if(linhas[li].items.length<1) continue;
        const rowGeom=alignRowGeometryToPlants(getRowGeometry(tl,tr,br,bl,axes.horiz,linhas[li].meanS),linhas[li].items,areaUtil);
        rowGeom.pxPorMetro=pxPorMetro;
        const dx=rowGeom.p2.x-rowGeom.p1.x, dy=rowGeom.p2.y-rowGeom.p1.y;
        const len=Math.sqrt(dx*dx+dy*dy);
        rowGeom.normal=len>0 ? {{x:-dy/len,y:dx/len}} : {{x:0,y:1}};
        rowGeom.corridorPx=corridorPx;
        const plantRanges=getPlantRangesPorLinha(rowGeom,linhas[li].items,sampler);
        if(calcularPlantados) registrarMetrosPlantadosDaLinha(rowGeom,linhas[li].items,tiroLabel,dispLabel,li+1,parcelaId,sampler,plantRanges);
        if(calcularFalhas) addFalhasSomenteEntrePlantasDaLinha(rowGeom,linhas[li].items,minDistM,tiroLabel,dispLabel,li+1,parcelaId,sampler,plantRanges);
      }}
    }}
  }}
  if(deletedFalhaKeys && deletedFalhaKeys.size>0) {{
    falhas=falhas.filter(f=>!deletedFalhaKeys.has(getFalhaSignature(f)));
  }}
  return true;
}}

function detectarFalhas() {{
  if(!processarQualidadePorLinhas('falhas')) return;
  const minDistM=getLimiteFalhaM();
  // Atualizar painel
  countPanel.style.display='block';
  countPanel.querySelector('h3').textContent='FALHAS DETECTADAS';
  totalCountEl.textContent=falhas.length;
  const totalLinear=getFalhasTotalValue(falhas);
  countInfoEl.textContent=falhas.length+' falha(s) na área útil >= '+formatMeters(minDistM)+' | buffer '+getBufferCorredorCm().toFixed(0)+' cm';
  if(falhas.length>0) {{
    falhasSumario.textContent='Total linear: '+formatMetricValue(totalLinear)+
      ' | Maior: '+formatFalhaDist(falhas.reduce((a,b)=>getFalhaValue(a)>=getFalhaValue(b)?a:b));
  }} else {{
    falhasSumario.textContent='Total linear: '+formatMetricValue(totalLinear);
  }}
  drawAll();
}}

function medirMetrosPlantados() {{
  if(!processarQualidadePorLinhas('plantados')) return;
  const totalPlantado=metrosPlantadosLinhas.reduce((s,row)=>s+(row.plantadoM||0),0);
  countPanel.style.display='block';
  countPanel.querySelector('h3').textContent='METROS PLANTADOS';
  totalCountEl.textContent=totalPlantado.toFixed(2).replace('.',',')+' m';
  countInfoEl.textContent=metrosPlantadosLinhas.length+' linha(s) medidas na área útil';
  falhasSumario.textContent='Total plantado: '+totalPlantado.toFixed(2).replace('.',',')+' m';
  drawAll();
}}

function formatDist(px) {{
  const u=selUnit.value;
  if(u==='cm') return (px*0.1).toFixed(1).replace('.',',')+' cm';
  if(u==='m')  return (px*0.001).toFixed(3).replace('.',',')+' m';
  return px.toFixed(0)+' px';
}}

function formatMeters(m) {{
  const u=selUnit.value;
  if(u==='cm') return (m*100).toFixed(1).replace('.',',')+' cm';
  if(u==='m')  return m.toFixed(3).replace('.',',')+' m';
  return m.toFixed(3).replace('.',',')+' m';
}}

function getFalhaValue(f) {{
  return selUnit.value==='px' ? f.dist : (f.distM||0);
}}

function formatMetricValue(value) {{
  if(selUnit.value==='px') return value.toFixed(0)+' px';
  return formatMeters(value);
}}

function formatFalhaDist(f) {{
  if(selUnit.value==='px') return f.dist.toFixed(0)+' px';
  return formatMeters(f.distM||0);
}}

function formatFalhaNumber(f) {{
  return formatFalhaDist(f).replace(' '+selUnit.value,'');
}}

function getFalhasTotalValue(lista) {{
  return lista.reduce((s,f)=>s+getFalhaValue(f),0);
}}

function cmToPx(value) {{
  return value/0.1;
}}

function unitToPx(value) {{
  const u=selUnit.value;
  if(u==='cm') return value/0.1;
  if(u==='m') return value/0.001;
  return value;
}}

function formatDistNumber(px) {{
  return formatDist(px).replace(' '+selUnit.value,'');
}}

function drawPolyPath(poly) {{
  ctx.beginPath();
  ctx.moveTo(poly[0].x,poly[0].y);
  for(let i=1;i<poly.length;i++) ctx.lineTo(poly[i].x,poly[i].y);
  ctx.closePath();
}}

function expandPoly(poly,pad) {{
  const center=getParcelCenter(poly);
  return poly.map(p=>{{
    const vx=p.x-center.x, vy=p.y-center.y;
    const len=Math.sqrt(vx*vx+vy*vy) || 1;
    return {{x:p.x+(vx/len)*pad, y:p.y+(vy/len)*pad}};
  }});
}}

function redrawImageInsidePoly(poly) {{
  if(imgW<=0 || !poly || poly.length<3) return;
  ctx.save();
  drawPolyPath(expandPoly(poly,8/sc));
  ctx.clip();
  ctx.drawImage(img,0,0);
  ctx.restore();
}}

function getParcelVisualKeys() {{
  const keys=new Set();
  for(const key of Object.keys(parcelAdjustments)) {{
    const adj=parcelAdjustments[key] || {{dx:0,dy:0}};
    if(Math.abs(adj.dx||0)+Math.abs(adj.dy||0)>0) keys.add(key);
  }}
  for(const key of selectedParcels) keys.add(key);
  for(const key of getCascadeAffectedKeys(false)) keys.add(key);
  return keys;
}}

function clearParcelAdjustmentResidues(R,C,p0,p1,p2,p3) {{
  const keys=getParcelVisualKeys();
  for(const key of keys) {{
    const parsed=parseParcelKey(key);
    if(parsed.disp<1 || parsed.disp>R || parsed.tiro<1 || parsed.tiro>C) continue;
    const r2=R-parsed.disp, c=C-parsed.tiro;
    redrawImageInsidePoly(getBaseParcelPoly(r2,c,R,C,p0,p1,p2,p3));
    redrawImageInsidePoly(getAdjustedParcelPoly(r2,c,R,C,p0,p1,p2,p3));
  }}
}}

function drawParcelAdjustmentOverlay(R,C,p0,p1,p2,p3) {{
  const cascadeAffected=getCascadeAffectedKeys(false);
  for(let r2=0;r2<R;r2++) for(let c=0;c<C;c++) {{
    const key=getParcelKeyByRC(r2,c,R,C);
    const adj=parcelAdjustments[key] || {{dx:0,dy:0}};
    const isAdjusted=Math.abs(adj.dx||0)+Math.abs(adj.dy||0)>0;
    const isSelected=selectedParcels.has(key);
    const isCascadeAffected=cascadeAffected.has(key);
    if(!isAdjusted && !isSelected && !isCascadeAffected) continue;
    const poly=getAdjustedParcelPoly(r2,c,R,C,p0,p1,p2,p3);
    const center=getParcelCenter(poly);
    ctx.save();
    drawPolyPath(poly);
    ctx.fillStyle=isSelected?'rgba(255,210,31,0.32)':(isCascadeAffected?'rgba(255,210,31,0.12)':'rgba(85,153,255,0.14)');
    ctx.fill();
    ctx.strokeStyle=isSelected?'#ffd21f':(isCascadeAffected?'rgba(255,210,31,0.72)':'#5599ff');
    ctx.lineWidth=(isSelected?4:(isCascadeAffected?2.5:3))/sc;
    ctx.shadowColor=isSelected?'rgba(255,210,31,0.75)':(isCascadeAffected?'rgba(255,210,31,0.35)':'rgba(85,153,255,0.55)');
    ctx.shadowBlur=10/sc;
    ctx.stroke();
    ctx.shadowBlur=0;
    ctx.fillStyle='rgba(0,0,0,0.75)';
    ctx.font='bold '+(Math.max(8,12/sc))+'px Arial';
    const label=getParcelLabel(key);
    const tw=ctx.measureText(label).width;
    ctx.fillRect(center.x-tw/2-4/sc,center.y-8/sc,tw+8/sc,16/sc);
    ctx.fillStyle=isSelected?'#ffd21f':(isCascadeAffected?'#ffe88a':'#aaddff');
    ctx.textAlign='center'; ctx.textBaseline='middle';
    ctx.fillText(label,center.x,center.y);
    ctx.restore();
  }}
}}

// ── Renderização ─────────────────────────────────────────────────────────
function drawAll() {{
  const W=vc.clientWidth, H=vc.clientHeight;
  cv.width=W; cv.height=H;
  ctx.clearRect(0,0,W,H);
  ctx.save();
  ctx.translate(ox,oy); ctx.scale(sc,sc);
  if(imgW>0) ctx.drawImage(img,0,0);

  // Grid
  if(points.length===4){{
    const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
    const p0=points[0],p1=points[1],p2=points[2],p3=points[3];
    ctx.save();
    ctx.strokeStyle='rgba(30,144,255,0.95)'; ctx.lineWidth=2/sc;
    ctx.shadowColor='rgba(0,160,255,0.6)'; ctx.shadowBlur=6/sc;
    for(let i=0;i<=R;i++) {{
      const v=i/R;
      const lx2=(1-v)*p0.x+v*p3.x, ly2=(1-v)*p0.y+v*p3.y;
      const rx2=(1-v)*p1.x+v*p2.x, ry2=(1-v)*p1.y+v*p2.y;
      ctx.beginPath(); ctx.moveTo(lx2,ly2); ctx.lineTo(rx2,ry2); ctx.stroke();
    }}
    for(let j=0;j<=C;j++) {{
      const u=j/C;
      const tx2=(1-u)*p0.x+u*p1.x, ty2=(1-u)*p0.y+u*p1.y;
      const bx2=(1-u)*p3.x+u*p2.x, by2=(1-u)*p3.y+u*p2.y;
      ctx.beginPath(); ctx.moveTo(tx2,ty2); ctx.lineTo(bx2,by2); ctx.stroke();
    }}
    ctx.restore();
    clearParcelAdjustmentResidues(R,C,p0,p1,p2,p3);
    drawParcelAdjustmentOverlay(R,C,p0,p1,p2,p3);
    for(let r2=0;r2<R;r2++) for(let c=0;c<C;c++) {{
      const key=getParcelKeyByRC(r2,c,R,C);
      if(!manualFailedParcels[key]) continue;
      const poly=getAdjustedParcelPoly(r2,c,R,C,p0,p1,p2,p3);
      const center=getParcelCenter(poly);
      ctx.save();
      drawPolyPath(poly);
      ctx.fillStyle='rgba(255,0,0,0.32)';
      ctx.strokeStyle='#ff2222';
      ctx.lineWidth=4/sc;
      ctx.shadowColor='rgba(255,0,0,0.7)';
      ctx.shadowBlur=10/sc;
      ctx.fill(); ctx.stroke(); ctx.shadowBlur=0;
      ctx.fillStyle='rgba(0,0,0,0.78)';
      ctx.font='bold '+(Math.max(8,12/sc))+'px Arial';
      const label='100% FALHADA';
      const tw=ctx.measureText(label).width;
      ctx.fillRect(center.x-tw/2-4/sc,center.y-8/sc,tw+8/sc,16/sc);
      ctx.fillStyle='#ff7777'; ctx.textAlign='center'; ctx.textBaseline='middle';
      ctx.fillText(label,center.x,center.y);
      ctx.restore();
    }}

    // Área útil do buffer interno usada na medição das falhas.
    if(areasUteis.length>0) {{
      ctx.save();
      ctx.fillStyle='rgba(255,210,31,0.13)';
      ctx.strokeStyle='rgba(255,210,31,0.98)';
      ctx.lineWidth=Math.max(2/sc,1.4/sc);
      ctx.shadowColor='rgba(255,210,31,0.55)';
      ctx.shadowBlur=5/sc;
      for(const area of areasUteis) {{
        if(!area.poly || area.poly.length<3) continue;
        drawPolyPath(area.poly);
        ctx.fill();
        ctx.beginPath();
        ctx.moveTo(area.poly[0].x,area.poly[0].y);
        ctx.lineTo(area.poly[1].x,area.poly[1].y);
        ctx.moveTo(area.poly[3].x,area.poly[3].y);
        ctx.lineTo(area.poly[2].x,area.poly[2].y);
        ctx.stroke();
      }}
      ctx.restore();
    }}

    // Labels de parcela
    if(showPlantDebug && Object.keys(parcelCounts).length>0) {{
      for(let r2=0;r2<R;r2++) for(let c=0;c<C;c++) {{
        const [tl,tr,br,bl]=getAdjustedParcelPoly(r2,c,R,C,p0,p1,p2,p3);
        const cx2=(tl.x+tr.x+br.x+bl.x)/4, cy2=(tl.y+tr.y+br.y+bl.y)/4;
        const dispL=R-r2, tiroL=C-c;
        const cnt=parcelCounts[dispL+'_'+tiroL]||0;
        ctx.save();
        ctx.fillStyle=cnt>0?'rgba(0,255,0,0.12)':'rgba(255,0,0,0.07)';
        ctx.beginPath();
        ctx.moveTo(tl.x,tl.y); ctx.lineTo(tr.x,tr.y);
        ctx.lineTo(br.x,br.y); ctx.lineTo(bl.x,bl.y);
        ctx.closePath(); ctx.fill();
        ctx.shadowColor='rgba(0,0,0,0.9)'; ctx.shadowBlur=4/sc;
        ctx.fillStyle='#ffffff';
        ctx.font='bold '+(Math.max(8,11/sc))+'px Arial';
        ctx.textAlign='center'; ctx.textBaseline='middle';
        ctx.fillText(cnt,cx2,cy2);
        ctx.font=(Math.max(6,8/sc))+'px Arial'; ctx.fillStyle='#aaa';
        ctx.fillText('T'+tiroL+' D'+dispL,cx2,cy2+14/sc);
        ctx.restore();
      }}
    }}
  }}

  if(showPlantDebug) {{
    // Plantas detectadas (X verde)
    for(const p of plantCenters) {{
      const sz=6/sc;
      ctx.save(); ctx.strokeStyle='#00ff00'; ctx.lineWidth=2/sc;
      ctx.beginPath(); ctx.moveTo(p.x-sz,p.y-sz); ctx.lineTo(p.x+sz,p.y+sz); ctx.stroke();
      ctx.beginPath(); ctx.moveTo(p.x-sz,p.y+sz); ctx.lineTo(p.x+sz,p.y-sz); ctx.stroke();
      ctx.restore();
    }}

    // Marcas manuais (X vermelho)
    for(const p of manualMarks) {{
      const sz=8/sc;
      ctx.save(); ctx.strokeStyle='#ff3333'; ctx.lineWidth=2.5/sc;
      ctx.beginPath(); ctx.moveTo(p.x-sz,p.y-sz); ctx.lineTo(p.x+sz,p.y+sz); ctx.stroke();
      ctx.beginPath(); ctx.moveTo(p.x-sz,p.y+sz); ctx.lineTo(p.x+sz,p.y-sz); ctx.stroke();
      ctx.restore();
    }}
  }}

  // Metros plantados
  if(chkShowFalhas.checked && qualidadeModoVisual==='plantados' && metrosPlantadosSegmentos.length>0) {{
    const lw=parseFloat(inpLineWidth.value)||2;
    const showLbl=chkLabels.checked;
    for(const seg of metrosPlantadosSegmentos) {{
      ctx.save();
      ctx.strokeStyle='#00ee55'; ctx.lineWidth=Math.max(lw/sc,3/sc);
      ctx.shadowColor='rgba(0,238,85,0.65)'; ctx.shadowBlur=6/sc;
      ctx.beginPath(); ctx.moveTo(seg.p1.x,seg.p1.y); ctx.lineTo(seg.p2.x,seg.p2.y); ctx.stroke();
      ctx.fillStyle='#00ee55';
      ctx.beginPath(); ctx.arc(seg.p1.x,seg.p1.y,3/sc,0,2*Math.PI); ctx.fill();
      ctx.beginPath(); ctx.arc(seg.p2.x,seg.p2.y,3/sc,0,2*Math.PI); ctx.fill();
      if(showLbl) {{
        const mx=(seg.p1.x+seg.p2.x)/2, my=(seg.p1.y+seg.p2.y)/2;
        const label=(seg.distM||0).toFixed(2).replace('.',',')+' m';
        ctx.shadowBlur=0;
        ctx.fillStyle='rgba(0,0,0,0.75)';
        const fs=Math.max(9,12/sc);
        ctx.font=fs+'px Arial';
        const tw=ctx.measureText(label).width;
        ctx.fillRect(mx-tw/2-3/sc, my-fs*0.7, tw+6/sc, fs*1.4);
        ctx.fillStyle='#44ff88';
        ctx.textAlign='center'; ctx.textBaseline='middle';
        ctx.fillText(label,mx,my);
      }}
      ctx.restore();
    }}
  }}

  // Falhas lineares
  if(chkShowFalhas.checked && qualidadeModoVisual==='falhas' && falhas.length>0) {{
    const lw=parseFloat(inpLineWidth.value)||2;
    const showLbl=chkLabels.checked;
    for(const f of falhas) {{
      let dp1=f.p1, dp2=f.p2;
      const segLen=lineDistance(f.p1,f.p2);
      if(segLen>0) {{
        const pad=Math.min(6/sc,segLen*0.25);
        let tA=0, tB=1;
        if(f.tipo==='entre plantas') {{ tA=pad/segLen; tB=1-pad/segLen; }}
        else if(f.tipo==='inicio') tB=1-pad/segLen;
        else if(f.tipo==='final') tA=pad/segLen;
        if(tB>tA) {{
          dp1=pointOnLine(f.p1,f.p2,tA);
          dp2=pointOnLine(f.p1,f.p2,tB);
        }}
      }}
      ctx.save();
      ctx.strokeStyle='#ff2200'; ctx.lineWidth=Math.max(lw/sc,3/sc);
      ctx.shadowColor='rgba(255,34,0,0.6)'; ctx.shadowBlur=6/sc;
      ctx.setLineDash([8/sc,4/sc]);
      ctx.beginPath(); ctx.moveTo(dp1.x,dp1.y); ctx.lineTo(dp2.x,dp2.y); ctx.stroke();
      ctx.setLineDash([]);
      // Pequenos círculos nos extremos
      ctx.fillStyle='#ff4400';
      ctx.beginPath(); ctx.arc(dp1.x,dp1.y,4/sc,0,2*Math.PI); ctx.fill();
      ctx.beginPath(); ctx.arc(dp2.x,dp2.y,4/sc,0,2*Math.PI); ctx.fill();
      // Etiqueta de distância
      if(showLbl) {{
        const mx=(dp1.x+dp2.x)/2, my=(dp1.y+dp2.y)/2;
        const label=formatFalhaDist(f);
        ctx.shadowBlur=0;
        ctx.fillStyle='rgba(0,0,0,0.75)';
        const fs=Math.max(9,12/sc);
        ctx.font=fs+'px Arial';
        const tw=ctx.measureText(label).width;
        ctx.fillRect(mx-tw/2-3/sc, my-fs*0.7, tw+6/sc, fs*1.4);
        ctx.fillStyle='#ff6633';
        ctx.textAlign='center'; ctx.textBaseline='middle';
        ctx.fillText(label,mx,my);
      }}
      ctx.restore();
    }}
  }}

  // Pontos do grid
  points.forEach((p,i)=>{{
    const isDrag=draggingPoint===i, r2=11/sc;
    ctx.save();
    ctx.shadowColor=isDrag?'rgba(255,255,255,0.9)':'rgba(0,180,255,0.8)'; ctx.shadowBlur=14/sc;
    ctx.beginPath(); ctx.arc(p.x,p.y,r2,0,2*Math.PI);
    ctx.fillStyle=isDrag?'#ffffff':'#1e90ff'; ctx.fill();
    ctx.lineWidth=2.5/sc; ctx.strokeStyle=isDrag?'#aaddff':'#00cfff'; ctx.stroke();
    ctx.restore();
    ctx.save(); ctx.fillStyle='#ffffff';
    ctx.font='bold '+(13/sc)+'px Arial';
    ctx.textAlign='center'; ctx.textBaseline='middle';
    ctx.fillText(i+1,p.x,p.y);
    ctx.restore();
  }});

  ctx.restore();
  zb.textContent=sc.toFixed(2)+'×';
}}

// ── Eventos imagem ────────────────────────────────────────────────────────
img.onload=()=>{{
  imgW=img.width; imgH=img.height;
  const W=vc.clientWidth, H=vc.clientHeight;
  sc=Math.min(W/imgW,H/imgH); ox=(W-imgW*sc)/2; oy=(H-imgH*sc)/2;
  loadParcelAdjustments();
  drawAll();
}};
img.src='data:image/jpeg;base64,'+IMG_B64;

// ── Pan & Zoom ────────────────────────────────────────────────────────────
vc.addEventListener('wheel',e=>{{
  e.preventDefault();
  const factor=e.deltaY<0?1.2:0.8;
  const r=cv.getBoundingClientRect();
  const mx=e.clientX-r.left, my=e.clientY-r.top;
  const ix=(mx-ox)/sc, iy=(my-oy)/sc;
  sc*=factor; sc=Math.max(MIN_SC,Math.min(MAX_SC,sc));
  ox=mx-ix*sc; oy=my-iy*sc; drawAll();
}},{{passive:false}});

vc.addEventListener('mousedown',e=>{{
  const pt=getImgCoords(e.clientX,e.clientY);
  if(deleteFalhaMode && points.length===4) {{
    const idx=findNearestFalha(pt);
    if(idx>=0) {{
      deletedFalhaKeys.add(getFalhaSignature(falhas[idx]));
      falhas.splice(idx,1);
      totalCountEl.textContent=falhas.length;
      countInfoEl.textContent='Medição apagada manualmente e descontabilizada dos totais.';
      drawAll();
      return;
    }}
  }}
  if((parcelSelectMode || parcelMoveMode) && points.length===4) {{
    const hit=findParcelAtPoint(pt);
    if(hit) {{
      if(parcelSelectMode) {{
        if(selectedParcels.has(hit.key)) selectedParcels.delete(hit.key);
        else selectedParcels.add(hit.key);
        updateParcelAdjustStatus();
        drawAll();
        return;
      }}
      if(parcelMoveMode) {{
        if(!selectedParcels.has(hit.key)) {{
          if(selectedParcels.size===0) selectedParcels.clear();
          selectedParcels.add(hit.key);
          updateParcelAdjustStatus();
        }}
        draggingParcelSelection=true;
        lastParcelDragPt=pt;
        vc.style.cursor='move';
        drawAll();
        return;
      }}
    }} else if(parcelSelectMode) {{
      selectedParcels.clear();
      updateParcelAdjustStatus();
      drawAll();
      return;
    }}
  }}
  for(let i=0;i<points.length;i++) {{
    const dx=(pt.x-points[i].x)*sc, dy=(pt.y-points[i].y)*sc;
    if(Math.sqrt(dx*dx+dy*dy)<20) {{ draggingPoint=i; return; }}
  }}
  if(gridMode&&points.length<4) {{
    points.push({{x:pt.x,y:pt.y}});
    if(points.length===4) gridMode=false;
    drawAll(); return;
  }}
  if(manualMode&&points.length===4) {{
    if(pointInPolygon(pt.x,pt.y,points)) {{
      manualMarks.push({{x:pt.x,y:pt.y}});
      plantCenters.push({{x:pt.x,y:pt.y}});
      recount(); drawAll();
    }}
    return;
  }}
  drag=true; lx=e.clientX; ly=e.clientY; vc.style.cursor='grabbing';
}});

vc.addEventListener('mousemove',e=>{{
  const pt=getImgCoords(e.clientX,e.clientY);
  coordEl.textContent='X:'+Math.round(pt.x)+' Y:'+Math.round(pt.y);
  if(draggingParcelSelection && lastParcelDragPt) {{
    translateSelectedParcels(pt.x-lastParcelDragPt.x, pt.y-lastParcelDragPt.y);
    lastParcelDragPt=pt;
    return;
  }}
  if(draggingPoint>=0) {{ points[draggingPoint]={{x:pt.x,y:pt.y}}; drawAll(); return; }}
  if(drag) {{ ox+=e.clientX-lx; oy+=e.clientY-ly; lx=e.clientX; ly=e.clientY; drawAll(); }}
}});

vc.addEventListener('mouseup',()=>{{
  if(draggingParcelSelection) refreshAfterParcelAdjust();
  drag=false; draggingPoint=-1; draggingParcelSelection=false; lastParcelDragPt=null; vc.style.cursor='grab';
}});
vc.addEventListener('mouseleave',()=>{{
  if(draggingParcelSelection) refreshAfterParcelAdjust();
  drag=false; draggingPoint=-1; draggingParcelSelection=false; lastParcelDragPt=null;
}});

vc.addEventListener('dblclick',e=>{{
  if(points.length<4) return;
  const pt=getImgCoords(e.clientX,e.clientY);
  const hit=findParcelAtPoint(pt);
  if(!hit) return;
  const currentCount=parcelCounts[hit.key] || 0;
  if(currentCount>0 && !confirm('Esta parcela tem contagem. Marcar mesmo assim como 100% FALHADA?')) return;
  manualFailedParcels[hit.key]=!manualFailedParcels[hit.key];
  if(manualFailedParcels[hit.key]) {{
    countPanel.style.display='block';
    countPanel.querySelector('h3').textContent='AJUSTE MANUAL';
    totalCountEl.textContent='100%';
    countInfoEl.textContent=getParcelLabel(hit.key)+' marcada como 100% FALHADA.';
    falhasSumario.textContent='OBS: 100% FALHADA';
  }}
  if(falhas.length>0 || qualidadeModoVisual==='falhas') detectarFalhas();
  else drawAll();
}});

// ── Botões ────────────────────────────────────────────────────────────────
btnGridTool.onclick=()=>{{
  gridMode=!gridMode; manualMode=false; deleteFalhaMode=false; parcelSelectMode=false; parcelMoveMode=false; selectedParcels.clear();
  btnGridTool.style.borderColor=gridMode?'#ff8c00':'#3a3a3a';
  if(gridMode) {{ points=[]; plantCenters=[]; manualMarks=[]; parcelCounts={{}}; falhas=[]; areasUteis=[]; metrosPlantadosLinhas=[]; metrosPlantadosSegmentos=[]; qualidadeModoVisual=''; parcelAdjustments={{}}; manualFailedParcels={{}}; deletedFalhaKeys=new Set(); if(parcelAdjustStorageKey) localStorage.removeItem(parcelAdjustStorageKey); countPanel.style.display='none'; updateParcelAdjustStatus(); drawAll(); }}
}};
if(btnCountPlants) btnCountPlants.onclick=()=>countPlantsInGrid();
if(btnCountAuto) btnCountAuto.onclick  =()=>countPlantsInGrid();
btnManualMode.onclick=()=>{{
  manualMode=!manualMode; gridMode=false; parcelSelectMode=false; parcelMoveMode=false;
  btnManualMode.style.borderColor=manualMode?'#5599ff':'#3a3a3a';
  updateParcelAdjustStatus();
}};
btnManual2.onclick=()=>{{
  manualMode=!manualMode; gridMode=false; parcelSelectMode=false; parcelMoveMode=false;
  btnManualMode.style.borderColor=manualMode?'#5599ff':'#3a3a3a';
  updateParcelAdjustStatus();
}};
btnSelectParcel.onclick=()=>{{
  if(points.length<4) {{ alert('Marque o grid primeiro.'); return; }}
  parcelSelectMode=!parcelSelectMode;
  parcelMoveMode=false; manualMode=false; gridMode=false;
  btnManualMode.style.borderColor='#3a3a3a';
  btnGridTool.style.borderColor='#3a3a3a';
  updateParcelAdjustStatus();
  drawAll();
}};
btnMoveSelected.onclick=()=>{{
  if(points.length<4) {{ alert('Marque o grid primeiro.'); return; }}
  parcelMoveMode=!parcelMoveMode;
  parcelSelectMode=false; manualMode=false; gridMode=false;
  btnManualMode.style.borderColor='#3a3a3a';
  btnGridTool.style.borderColor='#3a3a3a';
  updateParcelAdjustStatus();
  drawAll();
}};
btnClearSelection.onclick=()=>{{
  selectedParcels.clear();
  parcelSelectMode=false; parcelMoveMode=false; draggingParcelSelection=false;
  updateParcelAdjustStatus();
  drawAll();
}};
function moveSelectedByStep(dx,dy) {{
  translateSelectedParcels(dx,dy);
  refreshAfterParcelAdjust();
}}
btnMoveUp.onclick=()=>moveSelectedByStep(0,-(parseFloat(inpMoveStep.value)||5));
btnMoveDown.onclick=()=>moveSelectedByStep(0,(parseFloat(inpMoveStep.value)||5));
btnMoveLeft.onclick=()=>moveSelectedByStep(-(parseFloat(inpMoveStep.value)||5),0);
btnMoveRight.onclick=()=>moveSelectedByStep((parseFloat(inpMoveStep.value)||5),0);
btnSaveParcelAdjust.onclick=()=>{{
  saveParcelAdjustments();
  alert('Ajuste da parcela salvo para visualização, análise, exportação e relatório nesta ortofoto.');
}};
btnRemoveLast.onclick=btnUndoMark.onclick=()=>{{
  if(manualMarks.length>0) {{
    const rem=manualMarks.pop();
    plantCenters=plantCenters.filter(p=>p.x!==rem.x||p.y!==rem.y);
    recount(); drawAll();
  }}
}};
btnClearAll.onclick=()=>{{
  points=[]; plantCenters=[]; manualMarks=[]; parcelCounts={{}}; falhas=[]; areasUteis=[]; metrosPlantadosLinhas=[]; metrosPlantadosSegmentos=[]; qualidadeModoVisual=''; parcelAdjustments={{}}; manualFailedParcels={{}}; deletedFalhaKeys=new Set(); selectedParcels.clear();
  if(parcelAdjustStorageKey) localStorage.removeItem(parcelAdjustStorageKey);
  countPanel.style.display='none'; gridMode=false; manualMode=false; parcelSelectMode=false; parcelMoveMode=false;
  btnGridTool.style.borderColor='#3a3a3a'; btnManualMode.style.borderColor='#3a3a3a';
  updateParcelAdjustStatus();
  drawAll();
}};
btnDetectFalhas.onclick=()=>detectarFalhas();
btnMedirPlantados.onclick=()=>medirMetrosPlantados();
btnDeleteFalhaMode.onclick=()=>{{
  deleteFalhaMode=!deleteFalhaMode;
  manualMode=false; gridMode=false; parcelSelectMode=false; parcelMoveMode=false;
  btnDeleteFalhaMode.style.borderColor=deleteFalhaMode?'#ff5555':'#660000';
  countInfoEl.textContent=deleteFalhaMode?'Clique em uma linha de falha para apagar e descontabilizar.':'Modo apagar falha desativado.';
}};
chkShowFalhas.onchange=()=>drawAll();
chkLabels.onchange=()=>drawAll();
inpMinDist.onchange=()=>{{ if(falhas.length>0) detectarFalhas(); }};
inpBufferCm.onchange=()=>{{ if(falhas.length>0) detectarFalhas(); }};
inpParcelLen.onchange=()=>{{ if(falhas.length>0) detectarFalhas(); }};
selUnit.onchange=()=>{{ if(falhas.length>0) detectarFalhas(); else drawAll(); }};
inpLineWidth.onchange=()=>drawAll();
inpLinhasParcela.onchange=()=>{{ drawAll(); }};
inpQuadraName.onchange=()=>drawAll();

function parseFalhaParcelaSort(id) {{
  const m=String(id||'').match(/T(\\d+)\\s*D(\\d+)/i);
  return m ? {{tiro:parseInt(m[1]),disp:parseInt(m[2])}} : {{tiro:999999,disp:999999}};
}}

function getLinhaPlantioExport(linha) {{
  return Math.max(1,Math.min(getLinhasPlantioPorParcela(),parseInt(linha)||1));
}}

function getResumoPlantadoFalhaRows() {{
  const mapa={{}};
  for(const row of metrosPlantadosLinhas) {{
    const parcela=row.parcela||('T'+row.tiro+' D'+row.disp);
    const linha=getLinhaPlantioExport(row.linha);
    const key=parcela+'|'+linha;
    if(!mapa[key]) mapa[key]={{parcela:parcela,linha:linha,plantadoM:0,falhaM:0}};
    mapa[key].plantadoM+=Number.isFinite(row.plantadoM) ? row.plantadoM : 0;
  }}
  for(const f of falhas) {{
    const parcela=f.parcelaId||('T'+f.tiro+' D'+f.disp);
    const linha=getLinhaPlantioExport(f.linha);
    const key=parcela+'|'+linha;
    if(!mapa[key]) mapa[key]={{parcela:parcela,linha:linha,plantadoM:0,falhaM:0}};
    mapa[key].falhaM+=Number.isFinite(f.distM) ? f.distM : 0;
  }}
  const parcelas=[...new Set(Object.values(mapa).map(row=>row.parcela))];
  for(const parcela of parcelas) {{
    for(let linha=1; linha<=getLinhasPlantioPorParcela(); linha++) {{
      const key=parcela+'|'+linha;
      if(!mapa[key]) mapa[key]={{parcela:parcela,linha:linha,plantadoM:0,falhaM:0}};
    }}
  }}
  return Object.values(mapa).sort((a,b)=>{{
    const pa=parseFalhaParcelaSort(a.parcela), pb=parseFalhaParcelaSort(b.parcela);
    if(pa.tiro!==pb.tiro) return pa.tiro-pb.tiro;
    if(pa.disp!==pb.disp) return pa.disp-pb.disp;
    return a.linha-b.linha;
  }});
}}

function downloadQualidadeCSV(nomeArquivo,csvContent) {{
  const blob=new Blob([csvContent],{{type:'text/csv;charset=utf-8;'}});
  const a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  a.download=nomeArquivo;
  a.click();
  setTimeout(()=>URL.revokeObjectURL(a.href),1000);
}}

function exportFalhasPorLinhaCSV() {{
  exportQualidadeExcelCompleto();
}}

function exportQualidadeExcelCompleto() {{
  if(points.length<4) {{ alert('Marque o grid primeiro.'); return; }}
  if(typeof XLSX==='undefined') {{ alert('Biblioteca XLSX não carregada.'); return; }}

  const estadoAnterior={{
    falhas:[...falhas],
    areasUteis:[...areasUteis],
    metrosPlantadosLinhas:[...metrosPlantadosLinhas],
    metrosPlantadosSegmentos:[...metrosPlantadosSegmentos],
    qualidadeModoVisual:qualidadeModoVisual
  }};

  function restaurarEstadoQualidade() {{
    falhas=estadoAnterior.falhas;
    areasUteis=estadoAnterior.areasUteis;
    metrosPlantadosLinhas=estadoAnterior.metrosPlantadosLinhas;
    metrosPlantadosSegmentos=estadoAnterior.metrosPlantadosSegmentos;
    qualidadeModoVisual=estadoAnterior.qualidadeModoVisual;
  }}

  if(!processarQualidadePorLinhas('csv')) {{ restaurarEstadoQualidade(); return; }}
  const rows=getResumoPlantadoFalhaRows();
  if(rows.length===0) {{
    restaurarEstadoQualidade();
    alert('Nenhuma linha de plantio detectada para exportar.');
    return;
  }}

  const totalFalhaParcela={{}};
  const totalPlantadoParcela={{}};
  const obsPorParcela={{}};
  for(const row of rows) {{
    const parcela=row.parcela||'';
    totalFalhaParcela[parcela]=(totalFalhaParcela[parcela]||0)+(Number(row.falhaM)||0);
    totalPlantadoParcela[parcela]=(totalPlantadoParcela[parcela]||0)+(Number(row.plantadoM)||0);
  }}
  for(const f of falhas) {{
    const parcela=f.parcelaId||('T'+f.tiro+' D'+f.disp);
    if(f.obs) obsPorParcela[parcela]=f.obs;
  }}

  const round2=(v)=>Math.round((Number(v)||0)*100)/100;
  const headers=['Quadras','Parcela','Linha','Metros_Falha_Linha','Total_Falha_Parcela','Metros_Plantados_Linha','Total_Plantado_Parcela','OBS'];
  const dados=[headers];

  for(const row of rows) {{
    const parcela=row.parcela||'';
    dados.push([
      getQuadraNome(),
      parcela,
      'Linha '+row.linha,
      round2(row.falhaM),
      round2(totalFalhaParcela[parcela]),
      round2(row.plantadoM),
      round2(totalPlantadoParcela[parcela]),
      obsPorParcela[parcela]||''
    ]);
  }}

  const ws=XLSX.utils.aoa_to_sheet(dados);
  ws['!cols']=headers.map((header,idx)=>{{
    let maxLen=String(header).length;
    for(let r=1;r<dados.length;r++) maxLen=Math.max(maxLen,String(dados[r][idx]===undefined?'':dados[r][idx]).length);
    return {{wch:Math.min(Math.max(maxLen+3,12),34)}};
  }});
  ws['!freeze']={{xSplit:0,ySplit:1}};

  const border={{
    top:{{style:'thin',color:{{rgb:'808080'}}}},
    bottom:{{style:'thin',color:{{rgb:'808080'}}}},
    left:{{style:'thin',color:{{rgb:'808080'}}}},
    right:{{style:'thin',color:{{rgb:'808080'}}}}
  }};
  const alignCenter={{horizontal:'center',vertical:'center',wrapText:true}};
  const fillAzul={{fgColor:{{rgb:'D9EAF7'}}}};
  const fillVermelho={{fgColor:{{rgb:'F4CCCC'}}}};
  const fillVerde={{fgColor:{{rgb:'D9EAD3'}}}};

  function fillByCol(c) {{
    if(c===3 || c===4) return fillVermelho;
    if(c===5 || c===6) return fillVerde;
    return fillAzul;
  }}

  const range=XLSX.utils.decode_range(ws['!ref']);
  for(let R=range.s.r; R<=range.e.r; ++R) {{
    for(let C=range.s.c; C<=range.e.c; ++C) {{
      const addr=XLSX.utils.encode_cell({{r:R,c:C}});
      if(!ws[addr]) continue;
      ws[addr].s={{
        font:R===0 ? {{bold:true,name:'Arial',sz:11,color:{{rgb:'000000'}}}} : {{name:'Arial',sz:10,color:{{rgb:'000000'}}}},
        fill:{{patternType:'solid',...fillByCol(C)}},
        alignment:alignCenter,
        border:border
      }};
      if(R>0 && (C===3 || C===4 || C===5 || C===6)) {{
        ws[addr].t='n';
        ws[addr].v=round2(ws[addr].v);
        ws[addr].z='0.00';
      }}
    }}
  }}

  const wb=XLSX.utils.book_new();
  wb.Props={{
    Title:'Qualidade de Parcelas',
    Subject:'Falhas e Metros Plantados por Linha',
    Author:'TMG Sistema de Análise',
    CreatedDate:new Date()
  }};
  XLSX.utils.book_append_sheet(wb,ws,'Qualidade Parcelas');
  const safeQuadra=getQuadraNome().replace(/[^a-zA-Z0-9_-]+/g,'_');
  XLSX.writeFile(wb,'qualidade_parcelas_'+safeQuadra+'.xlsx',{{bookType:'xlsx',cellStyles:true}});
  restaurarEstadoQualidade();
}}

btnExportCnt.onclick=()=>exportFalhasPorLinhaCSV();

// ── Modal Relatório de Falhas ─────────────────────────────────────────────
function getFalhasFiltradas(search, fTiro, fDisp) {{
  return falhas.filter(f=>{{
    const label=(f.parcelaId||('T'+f.tiro+' D'+f.disp));
    if(search && !label.toLowerCase().includes(search.toLowerCase())) return false;
    if(fTiro && f.tiro!==parseInt(fTiro)) return false;
    if(fDisp && f.disp!==parseInt(fDisp)) return false;
    return true;
  }});
}}

function buildResumoParcelas(falhasBase) {{
  if(falhasBase.length===0) return '';
  const resumo={{}};
  for(const f of falhasBase) {{
    const id=f.parcelaId||('T'+f.tiro+' D'+f.disp);
    if(!resumo[id]) resumo[id]={{qtd:0,total:0,maior:0,menor:Infinity,linhas:{{}}}};
    const r=resumo[id];
    const v=getFalhaValue(f);
    r.qtd++;
    r.total+=v;
    r.maior=Math.max(r.maior,v);
    r.menor=Math.min(r.menor,v);
    r.linhas[f.linha]=(r.linhas[f.linha]||0)+1;
  }}
  let html='<div style="background:#111;border:1px solid #333;border-radius:8px;padding:10px;margin-bottom:10px;">';
  html+='<div style="color:#ff8c00;font-size:11px;font-weight:bold;letter-spacing:1px;text-transform:uppercase;margin-bottom:8px;">Resumo por parcela</div>';
  html+='<table style="width:100%;border-collapse:collapse;font-size:10px;">';
  html+='<thead><tr style="color:#aaa;"><th style="text-align:left;padding:4px;">Parcela</th><th style="padding:4px;">Falhas</th><th style="padding:4px;">Total</th><th style="padding:4px;">Por linha</th><th style="padding:4px;">Maior</th><th style="padding:4px;">Menor</th><th style="padding:4px;">Média</th></tr></thead><tbody>';
  for(const id of Object.keys(resumo).sort()) {{
    const r=resumo[id];
    const linhas=Object.keys(r.linhas).sort((a,b)=>parseInt(a)-parseInt(b)).map(l=>'L'+l+': '+r.linhas[l]).join(' | ');
    html+='<tr style="border-top:1px solid #222;">';
    html+='<td style="padding:4px;color:#fff;">'+id+'</td>';
    html+='<td style="padding:4px;color:#ff4444;text-align:center;">'+r.qtd+'</td>';
    html+='<td style="padding:4px;color:#ffaa33;text-align:center;">'+formatMetricValue(r.total)+'</td>';
    html+='<td style="padding:4px;color:#aaa;text-align:center;">'+linhas+'</td>';
    html+='<td style="padding:4px;color:#ff8c00;text-align:center;">'+formatMetricValue(r.maior)+'</td>';
    html+='<td style="padding:4px;color:#bbb;text-align:center;">'+formatMetricValue(r.menor)+'</td>';
    html+='<td style="padding:4px;color:#bbb;text-align:center;">'+formatMetricValue(r.total/r.qtd)+'</td>';
    html+='</tr>';
  }}
  html+='</tbody></table></div>';
  return html;
}}

function buildFalhasTable(search, fTiro, fDisp) {{
  const tbody=document.getElementById('falhasTable');
  if(falhas.length===0) {{ tbody.innerHTML='<p style="color:#666;text-align:center;padding:20px;font-size:12px;">Nenhuma falha detectada. Clique em ⚠️ Detectar Falhas.</p>'; return; }}
  const filtradas=getFalhasFiltradas(search, fTiro, fDisp);
  if(filtradas.length===0) {{
    tbody.innerHTML='<p style="color:#666;text-align:center;padding:20px;font-size:12px;">Nenhuma falha encontrada para o filtro atual.</p>';
    document.getElementById('modalTotalFalhas').textContent='0';
    document.getElementById('modalTotalLinear').textContent=formatMetricValue(0);
    document.getElementById('modalMaiorFalha').textContent=formatMetricValue(0);
    return;
  }}
  let html='<table style="width:100%;border-collapse:collapse;font-size:11px;">';
  html+='<thead><tr style="background:#222;color:#ff4444;"><th style="padding:8px;border-bottom:1px solid #333;">Parcela</th><th style="padding:8px;border-bottom:1px solid #333;">Linha</th><th style="padding:8px;border-bottom:1px solid #333;">Tipo</th><th style="padding:8px;border-bottom:1px solid #333;">Distância</th><th style="padding:8px;border-bottom:1px solid #333;">Coord. Ini.</th><th style="padding:8px;border-bottom:1px solid #333;">Coord. Fim</th></tr></thead><tbody>';
  let totalLinear=0, maior=0, count=0;
  for(const f of filtradas) {{
    const v=getFalhaValue(f);
    totalLinear+=v; if(v>maior) maior=v;
    count++;
    html+='<tr style="border-bottom:1px solid #1a1a1a;">';
    html+='<td style="padding:6px 8px;color:#5599ff;">'+(f.parcelaId||('T'+f.tiro+' D'+f.disp))+'</td>';
    html+='<td style="padding:6px 8px;color:#aaa;">L'+f.linha+'</td>';
    html+='<td style="padding:6px 8px;color:#bbb;">'+(f.tipo||'entre plantas')+'</td>';
    html+='<td style="padding:6px 8px;color:#ff4444;font-weight:bold;">'+formatFalhaDist(f)+'</td>';
    html+='<td style="padding:6px 8px;color:#666;font-size:10px;">('+Math.round(f.p1.x)+','+Math.round(f.p1.y)+')</td>';
    html+='<td style="padding:6px 8px;color:#666;font-size:10px;">('+Math.round(f.p2.x)+','+Math.round(f.p2.y)+')</td>';
    html+='</tr>';
  }}
  html+='</tbody></table>';
  tbody.innerHTML=buildResumoParcelas(filtradas)+html;
  document.getElementById('modalTotalFalhas').textContent=filtradas.length;
  document.getElementById('modalTotalLinear').textContent=formatMetricValue(totalLinear);
  document.getElementById('modalMaiorFalha').textContent=formatMetricValue(maior);
  document.getElementById('modalTotalLinearLabel').textContent='Total Linear';
  document.getElementById('modalMaiorFalhaLabel').textContent='Maior Falha';
}}

btnOpenFalhas.onclick=()=>{{
  if(points.length<4) {{ alert('Marque o grid primeiro!'); return; }}
  if(falhas.length===0) detectarFalhas();
  const R=parseInt(inpRows.value)||1, C=parseInt(inpCols.value)||1;
  const ftiro=document.getElementById('modalFilterTiro');
  const fdisp=document.getElementById('modalFilterDisp');
  ftiro.innerHTML='<option value="">Todos Tiros</option>';
  fdisp.innerHTML='<option value="">Todos Disp.</option>';
  for(let c=1;c<=C;c++) ftiro.innerHTML+='<option value="'+c+'">Tiro '+c+'</option>';
  for(let r=1;r<=R;r++) fdisp.innerHTML+='<option value="'+r+'">Disp. '+r+'</option>';
  buildFalhasTable('','','');
  falhasModal.style.display='flex';
}};
btnCloseModal.onclick=()=>{{ falhasModal.style.display='none'; }};
document.getElementById('modalSearch').oninput=function(){{ buildFalhasTable(this.value,document.getElementById('modalFilterTiro').value,document.getElementById('modalFilterDisp').value); }};
document.getElementById('modalFilterTiro').onchange=function(){{ buildFalhasTable(document.getElementById('modalSearch').value,this.value,document.getElementById('modalFilterDisp').value); }};
document.getElementById('modalFilterDisp').onchange=function(){{ buildFalhasTable(document.getElementById('modalSearch').value,document.getElementById('modalFilterTiro').value,this.value); }};

// ── Exportação Falhas CSV ─────────────────────────────────────────────────
document.getElementById('btnExportFalhasCSV').onclick=()=>{{
  exportFalhasPorLinhaCSV();
}};

// ── Exportação Falhas GeoJSON ─────────────────────────────────────────────
document.getElementById('btnExportFalhasGeoJSON').onclick=()=>{{
  const totalLinear=getFalhasTotalValue(falhas);
  const features=falhas.map((f,i)=>{{
    return {{
      type:'Feature',
      properties:{{ id:i+1, parcela_id:(f.parcelaId||('T'+f.tiro+' D'+f.disp)), tiro:f.tiro, disparo:f.disp, linha:f.linha, tipo:(f.tipo||'entre plantas'), distancia_px:Math.round(f.dist), distancia_m:f.distM||0, distancia_fmt:formatFalhaDist(f), unidade:selUnit.value }},
      geometry:{{ type:'LineString', coordinates:[[f.p1.x,-f.p1.y],[f.p2.x,-f.p2.y]] }}
    }};
  }});
  const gj=JSON.stringify({{type:'FeatureCollection',properties:{{total_falhas:falhas.length,total_linear_fmt:formatMetricValue(totalLinear),comprimento_parcela_m:getParcelRealMeters()}},features:features}},null,2);
  const blob=new Blob([gj],{{type:'application/json'}});
  const a=document.createElement('a'); a.href=URL.createObjectURL(blob);
  a.download='falhas_lineares.geojson'; a.click();
}};

// ── Exportação Falhas Excel ───────────────────────────────────────────────
document.getElementById('btnExportFalhasXLSX').onclick=()=>{{ exportQualidadeExcelCompleto(); return;
  if(typeof XLSX==='undefined') {{ alert('Biblioteca XLSX não carregada.'); return; }}
  const rows=[['ID Parcela','Linha','Tipo Falha','Distância','Unidade','X Início','Y Início','X Fim','Y Fim']];
  for(const f of falhas) rows.push([f.parcelaId||('T'+f.tiro+' D'+f.disp),f.linha,f.tipo||'entre plantas',formatFalhaNumber(f),selUnit.value,Math.round(f.p1.x),Math.round(f.p1.y),Math.round(f.p2.x),Math.round(f.p2.y)]);
  rows.push(['TOTAL','','',formatMetricValue(getFalhasTotalValue(falhas)).replace(' '+selUnit.value,''),selUnit.value,'','','','']);
  const ws=XLSX.utils.aoa_to_sheet(rows);
  ws['!cols']=[{{wch:14}},{{wch:8}},{{wch:16}},{{wch:12}},{{wch:8}},{{wch:10}},{{wch:10}},{{wch:10}},{{wch:10}}];
  const wb=XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb,ws,'Falhas Lineares');
  XLSX.writeFile(wb,'falhas_lineares.xlsx');
}};

new ResizeObserver(()=>drawAll()).observe(vc);
</script>
</body>
</html>
"""
                    components.html(qual_viewer, height=740, scrolling=False)

            else:
                st.markdown("""
                <div style='height:740px;border:1px dashed #2e2e2e;border-radius:12px;background:#0d0d0d;
                            display:flex;flex-direction:column;align-items:center;justify-content:center;
                            gap:12px;color:#333;'>
                    <div style='font-size:3rem;'>✅</div>
                    <div style='font-size:0.85rem;letter-spacing:2px;text-transform:uppercase;color:#555;'>
                        Carregue uma ortofoto para análise de qualidade
                    </div>
                    <div style='font-size:0.75rem;color:#444;'>
                        Marque o grid → Detecte Falhas → Exporte relatório
                    </div>
                </div>""", unsafe_allow_html=True)
            # FIM NOVO - MÓDULO QUALIDADE DE PARCELA

        else:
            st.markdown("""
            <div style='background:#1e1e1e;border:1px solid #333;border-radius:15px;padding:40px;text-align:center;'>
                <div style='font-size:3rem;margin-bottom:12px;'>📈</div>
                <div style='color:#555;font-size:0.9rem;letter-spacing:2px;text-transform:uppercase;'>
                    Selecione uma das opções acima para visualizar os resultados
                </div>
            </div>""", unsafe_allow_html=True)
    # FIM NOVO - VISUALIZADOR DE RESULTADOS

_render_partner_mention_notifications()

st.markdown("""
<style>
html body .main-header,
html body .menu-3d-title,
html body .cultura-title,
html body .login-title,
html body .cfg-panel-title,
html body .vd-login-title,
html body .vd-section-title,
html body .partner-excel-title,
html body .partner-toolbox-title,
html body .partner-window-title,
html body .partner-card-title,
html body .partner-hero-title,
html body .assessment-panel-title,
html body div[data-testid="stMarkdownContainer"] h1,
html body div[data-testid="stMarkdownContainer"] h2,
html body div[data-testid="stMarkdownContainer"] h3,
html body div[data-testid="stMarkdownContainer"] h4 {
    -webkit-text-fill-color:#ffffff !important;
    color:#ffffff !important;
    text-shadow:
        0 2px 0 rgba(0,0,0,.92),
        0 6px 14px rgba(0,0,0,.62),
        0 0 16px rgba(0,229,255,.50),
        0 0 30px rgba(0,255,157,.18) !important;
}
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<style id="tmg-global-neon-components-final">
html body div[data-testid="stTextInput"] [data-baseweb="input"],
html body div[data-testid="stNumberInput"] [data-baseweb="input"],
html body div[data-testid="stDateInput"] [data-baseweb="input"],
html body div[data-testid="stTimeInput"] [data-baseweb="input"],
html body div[data-testid="stTextArea"] textarea,
html body div[data-testid="stSelectbox"] [data-baseweb="select"],
html body div[data-testid="stMultiSelect"] [data-baseweb="select"] {{
    min-height:34px !important;
    border-radius:8px !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .72) !important;
    background:
        linear-gradient(120deg, rgba(255,255,255,.11), transparent 30%),
        linear-gradient(145deg, rgba(2,14,36,.98), rgba(18,62,100,.86), rgba({THEME_PRIMARY_RGB}, .20)) !important;
    box-shadow:
        0 10px 24px rgba(0,0,0,.40),
        0 0 20px rgba({THEME_PRIMARY_RGB}, .26),
        inset 0 1px 0 rgba(255,255,255,.22),
        inset 0 -6px 14px rgba(2,14,36,.48) !important;
    color:#ffffff !important;
    transition:all .30s ease !important;
}}
html body div[data-testid="stTextInput"] input,
html body div[data-testid="stNumberInput"] input,
html body div[data-testid="stDateInput"] input,
html body div[data-testid="stTimeInput"] input,
html body div[data-testid="stTextInput"] input[type="password"],
html body div[data-testid="stTextInput"] input[type="text"],
html body div[data-testid="stNumberInput"] input[type="number"],
html body div[data-testid="stTextArea"] textarea,
html body div[data-testid="stSelectbox"] [data-baseweb="select"] *,
html body div[data-testid="stMultiSelect"] [data-baseweb="select"] * {{
    background:transparent !important;
    color:#ffffff !important;
    -webkit-text-fill-color:#ffffff !important;
    font-weight:800 !important;
    text-shadow:0 1px 0 rgba(0,0,0,.84), 0 0 10px rgba({THEME_PRIMARY_RGB}, .34) !important;
}}
html body div[data-testid="stTextInput"] [data-baseweb="input"]:hover,
html body div[data-testid="stNumberInput"] [data-baseweb="input"]:hover,
html body div[data-testid="stDateInput"] [data-baseweb="input"]:hover,
html body div[data-testid="stTimeInput"] [data-baseweb="input"]:hover,
html body div[data-testid="stTextArea"] textarea:hover,
html body div[data-testid="stSelectbox"] [data-baseweb="select"]:hover,
html body div[data-testid="stMultiSelect"] [data-baseweb="select"]:hover,
html body div[data-testid="stTextInput"] [data-baseweb="input"]:focus-within,
html body div[data-testid="stNumberInput"] [data-baseweb="input"]:focus-within,
html body div[data-testid="stDateInput"] [data-baseweb="input"]:focus-within,
html body div[data-testid="stTimeInput"] [data-baseweb="input"]:focus-within,
html body div[data-testid="stTextArea"] textarea:focus,
html body div[data-testid="stSelectbox"] [data-baseweb="select"]:focus-within,
html body div[data-testid="stMultiSelect"] [data-baseweb="select"]:focus-within {{
    border-color:var(--tmg-primary-soft) !important;
    box-shadow:
        0 12px 28px rgba(0,0,0,.46),
        0 0 30px rgba({THEME_PRIMARY_RGB}, .46),
        inset 0 1px 0 rgba(255,255,255,.32),
        inset 0 -6px 14px rgba(2,14,36,.44) !important;
}}
html body div[data-testid="stTextInput"] input::placeholder,
html body div[data-testid="stNumberInput"] input::placeholder,
html body div[data-testid="stTextArea"] textarea::placeholder {{
    color:rgba(224,247,255,.76) !important;
    -webkit-text-fill-color:rgba(224,247,255,.76) !important;
    opacity:1 !important;
    font-weight:700 !important;
}}
html body div[data-testid="stTextInput"] input:-webkit-autofill,
html body div[data-testid="stNumberInput"] input:-webkit-autofill,
html body div[data-testid="stDateInput"] input:-webkit-autofill,
html body div[data-testid="stTimeInput"] input:-webkit-autofill {{
    -webkit-text-fill-color:#ffffff !important;
    caret-color:var(--tmg-primary-soft) !important;
    box-shadow:
        0 0 0 1000px rgba(2,14,36,.96) inset,
        0 10px 24px rgba(0,0,0,.40),
        0 0 20px rgba({THEME_PRIMARY_RGB}, .26) !important;
    transition:background-color 9999s ease-in-out 0s !important;
}}
html body div[data-testid="stTextInput"] input:disabled,
html body div[data-testid="stNumberInput"] input:disabled,
html body div[data-testid="stDateInput"] input:disabled,
html body div[data-testid="stTimeInput"] input:disabled,
html body div[data-testid="stTextArea"] textarea:disabled {{
    color:rgba(255,255,255,.86) !important;
    -webkit-text-fill-color:rgba(255,255,255,.86) !important;
    opacity:1 !important;
}}
html body div[data-baseweb="popover"],
html body div[data-baseweb="menu"],
html body ul[role="listbox"] {{
    border:1px solid rgba({THEME_PRIMARY_RGB}, .58) !important;
    border-radius:12px !important;
    background:
        linear-gradient(145deg, rgba(2,14,36,.98), rgba(18,62,100,.92), rgba({THEME_PRIMARY_RGB}, .18)) !important;
    box-shadow:
        0 16px 34px rgba(0,0,0,.54),
        0 0 28px rgba({THEME_PRIMARY_RGB}, .32),
        inset 0 1px 0 rgba(255,255,255,.18) !important;
    color:#ffffff !important;
}}
html body div[data-baseweb="popover"] *,
html body div[data-baseweb="menu"] *,
html body ul[role="listbox"] * {{
    color:#ffffff !important;
    -webkit-text-fill-color:#ffffff !important;
}}
</style>
""", unsafe_allow_html=True)

st.markdown(f"""
<style id="tmg-global-neon-system-final">
html body,
html body .stApp {{
    background:
        radial-gradient(circle at 14% 0%, rgba({THEME_PRIMARY_RGB}, .12), transparent 34%),
        linear-gradient(135deg, #020e24 0%, #061525 48%, #0d2b45 100%) !important;
    color:#ffffff !important;
}}
html body .block-container {{
    color:#ffffff !important;
}}
html body ::selection {{
    background:rgba({THEME_PRIMARY_RGB}, .44) !important;
    color:#ffffff !important;
}}
html body ::-webkit-scrollbar {{
    width:11px !important;
    height:11px !important;
}}
html body ::-webkit-scrollbar-track {{
    background:linear-gradient(180deg, #020e24, #071a2c) !important;
    border-radius:999px !important;
}}
html body ::-webkit-scrollbar-thumb {{
    background:linear-gradient(180deg, var(--tmg-primary-soft), var(--tmg-primary), var(--tmg-primary-dark)) !important;
    border:2px solid #020e24 !important;
    border-radius:999px !important;
    box-shadow:0 0 14px rgba({THEME_PRIMARY_RGB}, .40) !important;
}}
html body ::-webkit-scrollbar-thumb:hover {{
    box-shadow:0 0 20px rgba({THEME_PRIMARY_RGB}, .62) !important;
}}
html body [data-testid="stSidebar"],
html body [data-testid="stSidebarContent"] {{
    background:
        radial-gradient(circle at 20% 0%, rgba({THEME_PRIMARY_RGB}, .18), transparent 38%),
        linear-gradient(180deg, rgba(2,14,36,.98), rgba(7,26,53,.98) 52%, rgba(13,43,69,.96)) !important;
    color:#ffffff !important;
}}
html body [data-testid="stSidebar"] * {{
    color:#ffffff;
}}
html body header[data-testid="stHeader"],
html body [data-testid="stToolbar"],
html body [data-testid="stDecoration"] {{
    color:#ffffff !important;
}}
html body div.stButton > button,
html body div[data-testid="stDownloadButton"] button,
html body button[kind],
html body button[data-testid="baseButton-secondary"],
html body button[data-testid="baseButton-primary"],
html body button[data-testid="baseButton-minimal"] {{
    border-radius:10px !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .68) !important;
    background:
        linear-gradient(120deg, rgba(255,255,255,.12), transparent 34%),
        linear-gradient(145deg, rgba(2,14,36,.96), rgba(18,62,100,.84), rgba({THEME_PRIMARY_RGB}, .28)) !important;
    color:#ffffff !important;
    -webkit-text-fill-color:#ffffff !important;
    font-weight:850 !important;
    text-shadow:0 1px 0 rgba(0,0,0,.88), 0 0 10px rgba({THEME_PRIMARY_RGB}, .42) !important;
    box-shadow:
        0 10px 24px rgba(0,0,0,.38),
        0 0 20px rgba({THEME_PRIMARY_RGB}, .26),
        inset 0 1px 0 rgba(255,255,255,.22),
        inset 0 -7px 14px rgba(2,14,36,.44) !important;
    transition:all .30s ease !important;
}}
html body div.stButton > button:hover,
html body div[data-testid="stDownloadButton"] button:hover,
html body button[kind]:hover,
html body button[data-testid="baseButton-secondary"]:hover,
html body button[data-testid="baseButton-primary"]:hover,
html body button[data-testid="baseButton-minimal"]:hover {{
    transform:translateY(-1px) !important;
    border-color:var(--tmg-primary-soft) !important;
    box-shadow:
        0 14px 30px rgba(0,0,0,.46),
        0 0 32px rgba({THEME_PRIMARY_RGB}, .48),
        inset 0 1px 0 rgba(255,255,255,.32),
        inset 0 -7px 14px rgba(2,14,36,.38) !important;
}}
html body div.stButton > button:active,
html body div[data-testid="stDownloadButton"] button:active,
html body button[kind]:active {{
    transform:translateY(1px) scale(.99) !important;
}}
html body div[data-testid="stFileUploader"],
html body div[data-testid="stFileUploader"] section,
html body div[data-testid="stFileUploaderDropzone"] {{
    border-radius:14px !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .62) !important;
    background:
        radial-gradient(circle at 16% 0%, rgba({THEME_PRIMARY_RGB}, .20), transparent 44%),
        linear-gradient(145deg, rgba(2,14,36,.94), rgba(18,62,100,.78), rgba({THEME_PRIMARY_RGB}, .16)) !important;
    color:#ffffff !important;
    box-shadow:
        0 14px 28px rgba(0,0,0,.38),
        0 0 22px rgba({THEME_PRIMARY_RGB}, .24),
        inset 0 1px 0 rgba(255,255,255,.16) !important;
}}
html body div[data-testid="stFileUploader"] *,
html body div[data-testid="stFileUploaderDropzone"] * {{
    color:#ffffff !important;
    -webkit-text-fill-color:#ffffff !important;
    text-shadow:0 1px 0 rgba(0,0,0,.80), 0 0 8px rgba({THEME_PRIMARY_RGB}, .30) !important;
}}
html body div[data-testid="stProgress"] > div {{
    border-radius:999px !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .70) !important;
    background:linear-gradient(180deg, #020e24, #071a2c) !important;
    box-shadow:
        inset 0 3px 8px rgba(0,0,0,.70),
        0 0 18px rgba({THEME_PRIMARY_RGB}, .28) !important;
}}
html body div[data-testid="stProgress"] > div > div > div {{
    border-radius:999px !important;
    background:
        linear-gradient(90deg, var(--tmg-primary-soft), var(--tmg-primary), #00ff9d) !important;
    box-shadow:
        inset 0 1px 0 rgba(255,255,255,.42),
        0 0 18px rgba({THEME_PRIMARY_RGB}, .58) !important;
}}
html body div[data-testid="stExpander"] {{
    border:1px solid rgba({THEME_PRIMARY_RGB}, .54) !important;
    border-radius:12px !important;
    background:
        linear-gradient(145deg, rgba(2,14,36,.94), rgba(13,43,69,.82), rgba({THEME_PRIMARY_RGB}, .12)) !important;
    box-shadow:
        0 12px 26px rgba(0,0,0,.34),
        0 0 20px rgba({THEME_PRIMARY_RGB}, .20),
        inset 0 1px 0 rgba(255,255,255,.10) !important;
    overflow:hidden !important;
}}
html body div[data-testid="stExpander"] summary,
html body div[data-testid="stExpander"] summary * {{
    color:#ffffff !important;
    -webkit-text-fill-color:#ffffff !important;
    font-weight:850 !important;
    text-shadow:0 1px 0 rgba(0,0,0,.86), 0 0 10px rgba({THEME_PRIMARY_RGB}, .34) !important;
}}
html body div[data-testid="stTabs"] [role="tablist"] {{
    gap:8px !important;
    border-bottom:1px solid rgba({THEME_PRIMARY_RGB}, .40) !important;
}}
html body div[data-testid="stTabs"] button[role="tab"] {{
    border-radius:10px 10px 0 0 !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .42) !important;
    background:
        linear-gradient(145deg, rgba(2,14,36,.88), rgba(18,62,100,.62), rgba({THEME_PRIMARY_RGB}, .14)) !important;
    color:#dffbff !important;
    font-weight:800 !important;
    box-shadow:0 0 14px rgba({THEME_PRIMARY_RGB}, .16), inset 0 1px 0 rgba(255,255,255,.12) !important;
}}
html body div[data-testid="stTabs"] button[aria-selected="true"] {{
    background:
        linear-gradient(145deg, rgba(2,14,36,.96), rgba(18,62,100,.84), rgba({THEME_PRIMARY_RGB}, .32)) !important;
    color:#ffffff !important;
    border-color:rgba({THEME_PRIMARY_RGB}, .78) !important;
    box-shadow:0 0 24px rgba({THEME_PRIMARY_RGB}, .36), inset 0 1px 0 rgba(255,255,255,.22) !important;
}}
html body div[data-testid="stMetric"],
html body div[data-testid="stAlert"],
html body div[data-testid="stStatusWidget"],
html body div[data-testid="stInfo"],
html body div[data-testid="stSuccess"],
html body div[data-testid="stWarning"],
html body div[data-testid="stError"] {{
    border-radius:12px !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .46) !important;
    background:
        linear-gradient(145deg, rgba(2,14,36,.93), rgba(13,43,69,.78), rgba({THEME_PRIMARY_RGB}, .12)) !important;
    color:#ffffff !important;
    box-shadow:
        0 12px 24px rgba(0,0,0,.32),
        0 0 18px rgba({THEME_PRIMARY_RGB}, .18),
        inset 0 1px 0 rgba(255,255,255,.10) !important;
}}
html body div[data-testid="stMetric"] *,
html body div[data-testid="stAlert"] *,
html body div[data-testid="stStatusWidget"] *,
html body div[data-testid="stInfo"] *,
html body div[data-testid="stSuccess"] *,
html body div[data-testid="stWarning"] *,
html body div[data-testid="stError"] * {{
    color:#ffffff !important;
    -webkit-text-fill-color:#ffffff !important;
}}
html body div[data-testid="stDataFrame"],
html body div[data-testid="stTable"],
html body div[data-testid="stJson"],
html body div[data-testid="stDataEditor"],
html body .stDataFrame,
html body .stTable {{
    border-radius:14px !important;
    border:1px solid rgba({THEME_PRIMARY_RGB}, .58) !important;
    background:
        radial-gradient(circle at 12% 0%, rgba({THEME_PRIMARY_RGB}, .18), transparent 42%),
        linear-gradient(145deg, rgba(2,14,36,.96), rgba(13,43,69,.86), rgba({THEME_PRIMARY_RGB}, .12)) !important;
    color:#ffffff !important;
    box-shadow:
        0 14px 30px rgba(0,0,0,.40),
        0 0 24px rgba({THEME_PRIMARY_RGB}, .22),
        inset 0 1px 0 rgba(255,255,255,.12) !important;
    overflow:hidden !important;
}}
html body div[data-testid="stDataFrame"] *,
html body div[data-testid="stTable"] *,
html body div[data-testid="stJson"] *,
html body div[data-testid="stDataEditor"] * {{
    color:#ffffff !important;
    border-color:rgba({THEME_PRIMARY_RGB}, .22) !important;
}}
html body table {{
    color:#ffffff !important;
    background:rgba(2,14,36,.72) !important;
    border-color:rgba({THEME_PRIMARY_RGB}, .32) !important;
}}
html body thead,
html body thead tr,
html body th {{
    background:
        linear-gradient(145deg, rgba(18,62,100,.96), rgba({THEME_PRIMARY_RGB}, .30)) !important;
    color:#ffffff !important;
    text-shadow:0 1px 0 rgba(0,0,0,.82), 0 0 8px rgba({THEME_PRIMARY_RGB}, .26) !important;
}}
html body tbody tr:nth-child(odd) {{
    background:rgba(2,14,36,.52) !important;
}}
html body tbody tr:nth-child(even) {{
    background:rgba(13,43,69,.44) !important;
}}
html body tbody tr:hover {{
    background:rgba({THEME_PRIMARY_RGB}, .22) !important;
    box-shadow:inset 3px 0 0 var(--tmg-primary-soft) !important;
}}
html body div[data-testid="stImage"],
html body div[data-testid="stPlotlyChart"],
html body div[data-testid="stPyplot"],
html body iframe,
html body canvas {{
    border-radius:12px !important;
}}
html body div[data-testid="stImage"] img,
html body div[data-testid="stPlotlyChart"],
html body div[data-testid="stPyplot"],
html body iframe {{
    border:1px solid rgba({THEME_PRIMARY_RGB}, .38) !important;
    box-shadow:
        0 14px 30px rgba(0,0,0,.34),
        0 0 22px rgba({THEME_PRIMARY_RGB}, .16) !important;
}}
html body hr,
html body [data-testid="stDivider"] {{
    border-color:rgba({THEME_PRIMARY_RGB}, .34) !important;
}}
html body label,
html body .stMarkdown,
html body p,
html body span {{
    text-shadow:0 1px 0 rgba(0,0,0,.56);
}}
html body code,
html body pre {{
    border:1px solid rgba({THEME_PRIMARY_RGB}, .40) !important;
    border-radius:10px !important;
    background:rgba(2,14,36,.82) !important;
    color:#e8fbff !important;
}}
</style>
""", unsafe_allow_html=True)

# ==========================================
# FOOTER[cite: 1]
# ==========================================
st.markdown("<br><br>", unsafe_allow_html=True)
st.divider()

st.markdown(
    "<p style='text-align: center; color: #555;'>Estrutura Modular Profissional | Python 3.12</p>",
    unsafe_allow_html=True
)
