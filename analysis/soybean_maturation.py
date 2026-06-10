from __future__ import annotations

import base64
import json
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Callable

import cv2
import numpy as np
import pandas as pd
from PIL import Image, ImageDraw


DEFAULT_MATURATION_CONFIG: dict[str, Any] = {
    "verde_h_min": 36,
    "verde_h_max": 92,
    "verde_s_min": 34,
    "verde_v_min": 32,
    "amarelo_h_min": 12,
    "amarelo_h_max": 46,
    "amarelo_s_min": 24,
    "amarelo_v_min": 42,
    "marrom_h_min": 3,
    "marrom_h_max": 26,
    "marrom_s_min": 18,
    "marrom_s_max": 205,
    "marrom_v_min": 24,
    "marrom_v_max": 218,
    "solo_s_max": 35,
    "solo_v_min": 60,
    "sombra_v_max": 30,
    "r6_verde_min": 65,
    "r7_verde_min": 35,
    "r75_verde_min": 20,
    "amarelo_forcar_troca_estagio_pct": 68,
    "maturacao_avancada_util_pct": 82,
    "ignorar_solo_no_indice": False,
    "suavizar_imagem": False,
    "blur_kernel": 3,
    "aplicar_clahe": False,
    "clahe_clip_limit": 2.0,
    "clahe_tile_grid": 8,
    "classificar_outros_como_solo": True,
    "min_vegetacao_util_pct": 12,
    "max_solo_contabilizar_pct": 88,
}

DETAIL_COLUMNS = [
    "Quadra",
    "Parcela",
    "Disparo",
    "Tiro",
    "Data_Ortofoto",
    "Pixels_Validos",
    "Perc_Verde",
    "Perc_Amarelo",
    "Perc_Marrom_Seco",
    "Perc_Solo_Palhada",
    "Perc_Verde_Util",
    "Perc_Vegetacao_Util",
    "Contabilizar",
    "Perc_Sombra",
    "ExG_Medio",
    "ExR_Medio",
    "GLI_Medio",
    "VARI_Medio",
    "Indice_Verde",
    "Indice_Maturacao",
    "Estagio_Calculado",
    "Estagio_Manual",
    "Observacao",
]

SUMMARY_COLUMNS = [
    "Quadra",
    "Parcela",
    "Disparo",
    "Tiro",
    "Contabilizada",
    "Primeira_Data_R7",
    "Primeira_Data_R7_5",
    "Primeira_Data_R8",
    "Dias_Ate_R8",
    "Perc_Verde_Inicial",
    "Perc_Verde_Final",
    "Reducao_Verde",
    "Indice_Maturacao_Inicial",
    "Indice_Maturacao_Final",
    "Aumento_Maturacao",
    "Estagio_Final",
    "Observacao",
]


def _cfg(config: dict[str, Any] | None = None) -> dict[str, Any]:
    merged = dict(DEFAULT_MATURATION_CONFIG)
    if isinstance(config, dict):
        merged.update({k: v for k, v in config.items() if v is not None})
    return merged


def _empty_patch_result(error: str) -> dict[str, Any]:
    return {
        "pixels_validos": 0,
        "perc_verde": 0.0,
        "perc_amarelo": 0.0,
        "perc_marrom_seco": 0.0,
        "perc_solo_palhada": 0.0,
        "perc_verde_util": 0.0,
        "perc_vegetacao_util": 0.0,
        "contabilizar": "Não",
        "perc_sombra": 0.0,
        "exg_medio": 0.0,
        "exr_medio": 0.0,
        "gli_medio": 0.0,
        "vari_medio": 0.0,
        "indice_verde": 0.0,
        "indice_maturacao": 0.0,
        "estagio": "Indefinido",
        "erro": error,
    }


def _as_uint8_bgr(image: Any) -> np.ndarray:
    if isinstance(image, np.ndarray):
        arr = image
        if arr.ndim == 2:
            return cv2.cvtColor(arr.astype(np.uint8), cv2.COLOR_GRAY2BGR)
        if arr.ndim == 3 and arr.shape[2] == 4:
            return cv2.cvtColor(arr.astype(np.uint8), cv2.COLOR_BGRA2BGR)
        if arr.ndim == 3 and arr.shape[2] >= 3:
            return arr[:, :, :3].astype(np.uint8)
        raise ValueError("Imagem NumPy inválida para análise.")
    if isinstance(image, Image.Image):
        rgb = image.convert("RGB")
        return cv2.cvtColor(np.array(rgb), cv2.COLOR_RGB2BGR)
    if isinstance(image, (str, Path)):
        path = Path(image)
        if not path.exists():
            raise FileNotFoundError(str(path))
        with Image.open(path) as img:
            return _as_uint8_bgr(img)
    if isinstance(image, dict):
        if image.get("bgr") is not None:
            return _as_uint8_bgr(image["bgr"])
        if image.get("image") is not None:
            return _as_uint8_bgr(image["image"])
        if image.get("path"):
            return _as_uint8_bgr(image["path"])
        encoded = image.get("b64") or image.get("base64") or image.get("data_url")
        if encoded:
            text = str(encoded)
            if "," in text and text.lower().startswith("data:"):
                text = text.split(",", 1)[1]
            raw = base64.b64decode(text)
            with Image.open(BytesIO(raw)) as img:
                return _as_uint8_bgr(img)
    raise ValueError("Formato de ortofoto não suportado.")


def _preprocess_patch(patch_bgr: np.ndarray, config: dict[str, Any]) -> np.ndarray:
    patch = patch_bgr
    if bool(config.get("suavizar_imagem", False)):
        kernel = int(config.get("blur_kernel", 3) or 3)
        if kernel % 2 == 0:
            kernel += 1
        kernel = max(3, min(kernel, 15))
        patch = cv2.GaussianBlur(patch, (kernel, kernel), 0)
    if bool(config.get("aplicar_clahe", False)):
        lab = cv2.cvtColor(patch, cv2.COLOR_BGR2LAB)
        l_channel, a_channel, b_channel = cv2.split(lab)
        tile = int(config.get("clahe_tile_grid", 8) or 8)
        clahe = cv2.createCLAHE(
            clipLimit=float(config.get("clahe_clip_limit", 2.0) or 2.0),
            tileGridSize=(tile, tile),
        )
        lab = cv2.merge((clahe.apply(l_channel), a_channel, b_channel))
        patch = cv2.cvtColor(lab, cv2.COLOR_LAB2BGR)
    return patch


def calculate_rgb_indices(patch_bgr: np.ndarray, valid_mask: np.ndarray | None = None) -> dict[str, float]:
    if patch_bgr is None or getattr(patch_bgr, "size", 0) == 0:
        return {"exg_medio": 0.0, "exr_medio": 0.0, "gli_medio": 0.0, "vari_medio": 0.0}

    patch = patch_bgr.astype(np.float32)
    b, g, r = cv2.split(patch)
    if valid_mask is None:
        valid_mask = np.ones(patch.shape[:2], dtype=bool)
    else:
        valid_mask = valid_mask.astype(bool)

    exg = 2 * g - r - b
    exr = 1.4 * r - g
    gli = (2 * g - r - b) / (2 * g + r + b + 1e-6)
    vari = (g - r) / (g + r - b + 1e-6)

    if not np.any(valid_mask):
        return {"exg_medio": 0.0, "exr_medio": 0.0, "gli_medio": 0.0, "vari_medio": 0.0}

    return {
        "exg_medio": round(float(np.mean(exg[valid_mask])), 2),
        "exr_medio": round(float(np.mean(exr[valid_mask])), 2),
        "gli_medio": round(float(np.mean(gli[valid_mask])), 4),
        "vari_medio": round(float(np.mean(vari[valid_mask])), 4),
    }


def create_hsv_maturation_masks(patch_bgr: np.ndarray, config: dict[str, Any] | None = None) -> dict[str, np.ndarray]:
    cfg = _cfg(config)
    if patch_bgr is None or getattr(patch_bgr, "size", 0) == 0:
        empty = np.zeros((0, 0), dtype=bool)
        return {"verde": empty, "amarelo": empty, "marrom_seco": empty, "solo_palhada": empty, "sombra": empty, "validos": empty}

    patch = _as_uint8_bgr(patch_bgr)
    if not bool(cfg.get("_preprocessed", False)):
        patch = _preprocess_patch(patch, cfg)
    hsv = cv2.cvtColor(patch, cv2.COLOR_BGR2HSV)
    h = hsv[:, :, 0]
    s = hsv[:, :, 1]
    v = hsv[:, :, 2]
    b, g, r = cv2.split(patch.astype(np.float32))
    exg = 2 * g - r - b

    mask_sombra = v < int(cfg["sombra_v_max"])
    valid = ~mask_sombra

    mask_verde = (
        (h >= int(cfg["verde_h_min"]))
        & (h <= int(cfg["verde_h_max"]))
        & (s >= int(cfg["verde_s_min"]))
        & (v >= int(cfg["verde_v_min"]))
        & (g >= r * 0.78)
        & (g >= b * 0.88)
        & (exg > 6)
        & valid
    )
    mask_amarelo = (
        (h >= int(cfg["amarelo_h_min"]))
        & (h <= int(cfg["amarelo_h_max"]))
        & (s >= int(cfg["amarelo_s_min"]))
        & (v >= int(cfg["amarelo_v_min"]))
        & (g >= b * 0.86)
        & (r >= b * 0.78)
        & valid
        & ~mask_verde
    )
    mask_marrom = (
        (h >= int(cfg["marrom_h_min"]))
        & (h <= int(cfg["marrom_h_max"]))
        & (s >= int(cfg["marrom_s_min"]))
        & (s <= int(cfg["marrom_s_max"]))
        & (v >= int(cfg["marrom_v_min"]))
        & (v <= int(cfg["marrom_v_max"]))
        & (r >= g * 0.70)
        & valid
        & ~mask_verde
        & ~mask_amarelo
    )
    mask_solo = (
        (s < int(cfg["solo_s_max"]))
        & (v >= int(cfg["solo_v_min"]))
        & valid
        & ~mask_verde
        & ~mask_amarelo
        & ~mask_marrom
    )
    if bool(cfg.get("classificar_outros_como_solo", True)):
        mask_solo = mask_solo | (valid & ~mask_verde & ~mask_amarelo & ~mask_marrom)

    return {
        "verde": mask_verde,
        "amarelo": mask_amarelo,
        "marrom_seco": mask_marrom,
        "solo_palhada": mask_solo,
        "sombra": mask_sombra,
        "validos": valid,
    }


def classify_soybean_maturation(
    percent_green: float,
    percent_mature: float,
    config: dict[str, Any] | None = None,
    percent_yellow_brown: float | None = None,
) -> str:
    cfg = _cfg(config)
    green = float(percent_green or 0)
    mature = float(percent_mature or 0)
    yellow_brown = float(percent_yellow_brown or 0)
    switch_pct = float(cfg.get("amarelo_forcar_troca_estagio_pct", 68) or 68)
    advanced_pct = float(cfg.get("maturacao_avancada_util_pct", 82) or 82)
    if yellow_brown >= advanced_pct or mature >= advanced_pct:
        return "R8 / Madura"
    if yellow_brown >= switch_pct or mature >= switch_pct:
        return "R7.5 / Avançada"
    if green >= float(cfg["r6_verde_min"]):
        return "R6 / Imatura"
    if green >= float(cfg["r7_verde_min"]):
        return "R7 / Intermediária"
    if green >= float(cfg["r75_verde_min"]):
        return "R7.5 / Avançada"
    return "R8 / Madura"


def analyze_soybean_maturation_patch(patch_bgr: np.ndarray, config: dict[str, Any] | None = None) -> dict[str, Any]:
    if patch_bgr is None or getattr(patch_bgr, "size", 0) == 0:
        return _empty_patch_result("Recorte vazio")

    cfg = _cfg(config)
    try:
        patch = _preprocess_patch(_as_uint8_bgr(patch_bgr), cfg)
        masks = create_hsv_maturation_masks(patch, {**cfg, "_preprocessed": True})
        pixels_totais = int(patch.shape[0] * patch.shape[1])
        pixels_sombra = int(np.sum(masks["sombra"]))
        valid_mask = masks["validos"]
        pixels_validos_reais = int(np.sum(valid_mask))
        if pixels_totais <= 0 or pixels_validos_reais <= 0:
            return _empty_patch_result("Máscara sem pixels válidos")

        pixels_validos = max(pixels_validos_reais, 1)
        pixels_verde = int(np.sum(masks["verde"]))
        pixels_amarelo = int(np.sum(masks["amarelo"]))
        pixels_marrom = int(np.sum(masks["marrom_seco"]))
        pixels_solo = int(np.sum(masks["solo_palhada"]))
        pixels_vegetacao = pixels_verde + pixels_amarelo + pixels_marrom

        perc_verde = float(pixels_verde / pixels_validos * 100)
        perc_amarelo = float(pixels_amarelo / pixels_validos * 100)
        perc_marrom = float(pixels_marrom / pixels_validos * 100)
        perc_solo = float(pixels_solo / pixels_validos * 100)
        perc_vegetacao_util = float(pixels_vegetacao / pixels_validos * 100)
        if pixels_vegetacao > 0:
            perc_verde_util = float(pixels_verde / pixels_vegetacao * 100)
            indice_maturacao = float((pixels_amarelo + pixels_marrom) / pixels_vegetacao * 100)
        else:
            perc_verde_util = 0.0
            indice_maturacao = 0.0
        perc_sombra = float(pixels_sombra / pixels_totais * 100)

        rgb_indices = calculate_rgb_indices(patch, valid_mask)
        indice_verde = perc_verde
        contabilizar = (
            pixels_vegetacao > 0
            and perc_vegetacao_util >= float(cfg.get("min_vegetacao_util_pct", 12) or 12)
            and perc_solo < float(cfg.get("max_solo_contabilizar_pct", 88) or 88)
        )
        yellow_brown_pct = perc_amarelo + perc_marrom
        estagio = classify_soybean_maturation(perc_verde_util, indice_maturacao, cfg, yellow_brown_pct) if contabilizar else "Vazia/Colhida"

        return {
            "pixels_validos": int(pixels_validos_reais),
            "perc_verde": round(perc_verde, 2),
            "perc_amarelo": round(perc_amarelo, 2),
            "perc_marrom_seco": round(perc_marrom, 2),
            "perc_solo_palhada": round(perc_solo, 2),
            "perc_verde_util": round(perc_verde_util, 2),
            "perc_vegetacao_util": round(perc_vegetacao_util, 2),
            "contabilizar": "Sim" if contabilizar else "Não",
            "perc_sombra": round(perc_sombra, 2),
            "exg_medio": rgb_indices["exg_medio"],
            "exr_medio": rgb_indices["exr_medio"],
            "gli_medio": rgb_indices["gli_medio"],
            "vari_medio": rgb_indices["vari_medio"],
            "indice_verde": round(indice_verde, 2),
            "indice_maturacao": round(indice_maturacao, 2),
            "estagio": estagio,
            "erro": "" if contabilizar else "Parcela vazia/colhida não contabilizada.",
        }
    except MemoryError:
        return _empty_patch_result("Memória insuficiente para analisar o recorte")
    except Exception as exc:
        result = _empty_patch_result(str(exc))
        result["erro"] = f"Erro na análise do recorte: {exc}"
        return result


def crop_grid_parcel(image: np.ndarray, parcel_polygon: list[tuple[float, float]] | np.ndarray) -> np.ndarray:
    img = _as_uint8_bgr(image)
    polygon = np.asarray(parcel_polygon, dtype=np.float32)
    if polygon.ndim != 2 or polygon.shape[0] < 3 or polygon.shape[1] < 2:
        return np.empty((0, 0, 3), dtype=np.uint8)

    h, w = img.shape[:2]
    min_x = max(0, int(np.floor(np.min(polygon[:, 0]))))
    max_x = min(w - 1, int(np.ceil(np.max(polygon[:, 0]))))
    min_y = max(0, int(np.floor(np.min(polygon[:, 1]))))
    max_y = min(h - 1, int(np.ceil(np.max(polygon[:, 1]))))
    if max_x <= min_x or max_y <= min_y:
        return np.empty((0, 0, 3), dtype=np.uint8)

    crop = img[min_y : max_y + 1, min_x : max_x + 1].copy()
    local_polygon = polygon.copy()
    local_polygon[:, 0] -= min_x
    local_polygon[:, 1] -= min_y
    mask = np.zeros(crop.shape[:2], dtype=np.uint8)
    cv2.fillPoly(mask, [np.round(local_polygon).astype(np.int32)], 255)
    crop[mask == 0] = (0, 0, 0)
    return crop


def _bilerp(p0: np.ndarray, p1: np.ndarray, p2: np.ndarray, p3: np.ndarray, u: float, v: float) -> np.ndarray:
    top = (1 - u) * p0 + u * p1
    bottom = (1 - u) * p3 + u * p2
    return (1 - v) * top + v * bottom


def _grid_points_for_image(grid: dict[str, Any], image_shape: tuple[int, int], date_index: int = 0) -> np.ndarray:
    h, w = image_shape[:2]
    date_grids = grid.get("dateGridRatios") or grid.get("date_grid_ratios") or {}
    points = None
    used_date_specific_grid = False
    if date_index and str(date_index) in date_grids:
        points = date_grids[str(date_index)]
        used_date_specific_grid = True
    elif date_index and date_index in date_grids:
        points = date_grids[date_index]
        used_date_specific_grid = True
    if points is None:
        points = grid.get("gridRatios") or grid.get("grid_ratios") or grid.get("points") or grid.get("pontos") or grid.get("grid")
    if not points or len(points) < 4:
        raise ValueError("Grid ausente: marque os 4 cantos antes de analisar.")

    arr = []
    for point in points[:4]:
        if isinstance(point, dict):
            x, y = float(point.get("x", 0)), float(point.get("y", 0))
        else:
            x, y = float(point[0]), float(point[1])
        if abs(x) <= 2 and abs(y) <= 2:
            x *= w
            y *= h
        arr.append((x, y))

    offsets = grid.get("dateOffsets") or grid.get("date_offsets") or {}
    off = offsets.get(str(date_index), offsets.get(date_index, {"x": 0, "y": 0}))
    if off and not used_date_specific_grid:
        ox = float(off.get("x", 0)) * w if isinstance(off, dict) else 0
        oy = float(off.get("y", 0)) * h if isinstance(off, dict) else 0
        arr = [(x + ox, y + oy) for x, y in arr]
    return np.asarray(arr, dtype=np.float32)


def _grid_to_parcels(grid: dict[str, Any], image_shape: tuple[int, int], date_index: int = 0) -> list[dict[str, Any]]:
    if grid.get("parcelas") or grid.get("parcels"):
        source = grid.get("parcelas") or grid.get("parcels")
        parcels = []
        for item in source:
            poly = item.get("polygon") or item.get("poligono") or item.get("points")
            if poly:
                parcelas = {
                    "quadra": item.get("Quadra") or item.get("quadra") or grid.get("quadra") or grid.get("name") or "Maturacao_Soja",
                    "parcela": item.get("Parcela") or item.get("parcela"),
                    "disparo": int(item.get("Disparo") or item.get("disparo") or 0),
                    "tiro": int(item.get("Tiro") or item.get("tiro") or 0),
                    "polygon": np.asarray(poly, dtype=np.float32),
                }
                parcels.append(parcelas)
        if parcels:
            return parcels

    rows = int(grid.get("rows") or grid.get("linhas") or grid.get("disparos") or 1)
    cols = int(grid.get("cols") or grid.get("colunas") or grid.get("tiros") or 1)
    rows = max(1, rows)
    cols = max(1, cols)
    p0, p1, p2, p3 = _grid_points_for_image(grid, image_shape, date_index)
    quadra = grid.get("quadra") or grid.get("name") or grid.get("nome") or "Maturacao_Soja"
    parcels = []
    for r in range(rows):
        for c in range(cols):
            u0, u1 = c / cols, (c + 1) / cols
            v0, v1 = r / rows, (r + 1) / rows
            poly = np.asarray(
                [
                    _bilerp(p0, p1, p2, p3, u0, v0),
                    _bilerp(p0, p1, p2, p3, u1, v0),
                    _bilerp(p0, p1, p2, p3, u1, v1),
                    _bilerp(p0, p1, p2, p3, u0, v1),
                ],
                dtype=np.float32,
            )
            disparo = rows - r
            tiro = cols - c
            parcels.append(
                {
                    "quadra": quadra,
                    "parcela": f"T{tiro} D{disparo}",
                    "disparo": disparo,
                    "tiro": tiro,
                    "polygon": poly,
                }
            )
    return parcels


def _parse_date(value: Any) -> pd.Timestamp:
    if isinstance(value, pd.Timestamp):
        return value
    return pd.to_datetime(value, errors="coerce")


def analyze_grid_temporal_maturation(
    orthophotos: list[Any],
    dates: list[Any] | None,
    grid: dict[str, Any],
    config: dict[str, Any] | None = None,
    progress_callback: Callable[..., Any] | None = None,
) -> pd.DataFrame:
    if not orthophotos:
        raise ValueError("Nenhuma ortofoto carregada para análise temporal.")
    if not isinstance(grid, dict):
        raise ValueError("Grid inválido ou ausente.")

    entries = []
    for idx, item in enumerate(orthophotos):
        item_date = None
        item_name = f"Ortofoto_{idx + 1}"
        if isinstance(item, dict):
            item_date = item.get("date") or item.get("data") or item.get("Data_Ortofoto")
            item_name = item.get("name") or item.get("nome") or item_name
        if dates and idx < len(dates):
            item_date = dates[idx]
        parsed = _parse_date(item_date)
        if pd.isna(parsed):
            raise ValueError(f"Data ausente ou inválida na ortofoto {idx + 1}.")
        entries.append({"idx": idx, "date": parsed, "name": item_name, "item": item})

    if len({entry["date"].date().isoformat() for entry in entries}) != len(entries):
        raise ValueError("Existem datas duplicadas na linha temporal.")

    entries.sort(key=lambda entry: (entry["date"], entry["idx"]))
    rows_count = int(grid.get("rows") or grid.get("linhas") or grid.get("disparos") or 1)
    cols_count = int(grid.get("cols") or grid.get("colunas") or grid.get("tiros") or 1)
    total_steps = len(entries) * max(1, rows_count) * max(1, cols_count)
    rows_out = []
    done = 0
    for order_idx, entry in enumerate(entries):
        bgr = _as_uint8_bgr(entry["item"])
        parcels = _grid_to_parcels(grid, bgr.shape, order_idx)
        for parcel in parcels:
            done += 1
            if progress_callback:
                pct = done / max(1, total_steps) * 100
                try:
                    progress_callback(pct, f"Analisando {entry['name']} · {parcel['parcela']}")
                except TypeError:
                    progress_callback(done, total_steps, f"Analisando {entry['name']} · {parcel['parcela']}")
            patch = crop_grid_parcel(bgr, parcel["polygon"])
            result = analyze_soybean_maturation_patch(patch, config)
            rows_out.append(
                {
                    "Quadra": parcel["quadra"],
                    "Parcela": parcel["parcela"],
                    "Disparo": parcel["disparo"],
                    "Tiro": parcel["tiro"],
                    "Data_Ortofoto": entry["date"].date().isoformat(),
                    "Pixels_Validos": result["pixels_validos"],
                    "Perc_Verde": result["perc_verde"],
                    "Perc_Amarelo": result["perc_amarelo"],
                    "Perc_Marrom_Seco": result["perc_marrom_seco"],
                    "Perc_Solo_Palhada": result["perc_solo_palhada"],
                    "Perc_Verde_Util": result["perc_verde_util"],
                    "Perc_Vegetacao_Util": result["perc_vegetacao_util"],
                    "Contabilizar": result["contabilizar"],
                    "Perc_Sombra": result["perc_sombra"],
                    "ExG_Medio": result["exg_medio"],
                    "ExR_Medio": result["exr_medio"],
                    "GLI_Medio": result["gli_medio"],
                    "VARI_Medio": result["vari_medio"],
                    "Indice_Verde": result["indice_verde"],
                    "Indice_Maturacao": result["indice_maturacao"],
                    "Estagio_Calculado": result["estagio"],
                    "Estagio_Manual": "",
                    "Observacao": result["erro"] or "Classificação automática por HSV/OpenCV.",
                }
            )
        del bgr

    return pd.DataFrame(rows_out, columns=DETAIL_COLUMNS)


def _stage_label(row: pd.Series) -> str:
    manual = str(row.get("Estagio_Manual", "") or "").strip()
    return manual or str(row.get("Estagio_Calculado", "") or "")


def _stage_rank(stage: str) -> int:
    text = str(stage)
    low = text.lower()
    if "dessecado" in low or "vazia" in low or "colhida" in low or "sem leitura" in low or "sem dados" in low:
        return 0
    if "R8" in text:
        return 4
    if "R7.5" in text or "R7_5" in text:
        return 3
    if "R7" in text:
        return 2
    if "R6" in text:
        return 1
    if "maduro" in low or "madura" in low:
        return 4
    if "maturação intermediária" in low or "maturacao intermediaria" in low:
        return 3
    if "início de maturação" in low or "inicio de maturacao" in low:
        return 2
    if "verde" in low or "vegetativo" in low:
        return 1
    return 0


def _is_countable_maturation_row(row: pd.Series) -> bool:
    manual = str(row.get("Estagio_Manual", "") or "").strip()
    stage = manual or str(row.get("Estagio_Calculado", "") or "")
    text = stage.lower()
    if "vazia" in text or "colhida" in text or "sem dados" in text or "indefinido" in text:
        return False
    if manual:
        return _stage_rank(stage) > 0
    if str(row.get("Contabilizar", "") or "").strip().lower() == "não":
        return False
    return _stage_rank(stage) > 0


def build_temporal_summary(results_df: pd.DataFrame) -> pd.DataFrame:
    if results_df is None or results_df.empty:
        return pd.DataFrame(columns=SUMMARY_COLUMNS)
    df = results_df.copy()
    df["Data_Ortofoto"] = pd.to_datetime(df["Data_Ortofoto"], errors="coerce")
    df["Estagio_Final_Linha"] = df.apply(_stage_label, axis=1)
    summary = []
    group_cols = ["Quadra", "Parcela", "Disparo", "Tiro"]
    for keys, group in df.sort_values("Data_Ortofoto").groupby(group_cols, dropna=False):
        group = group.sort_values("Data_Ortofoto")
        first = group.iloc[0]
        last = group.iloc[-1]
        valid_group = group[group.apply(_is_countable_maturation_row, axis=1)]
        ranks = valid_group["Estagio_Final_Linha"].map(_stage_rank) if not valid_group.empty else pd.Series(dtype=int)
        r7_rows = valid_group[ranks >= 2] if not valid_group.empty else pd.DataFrame()
        r75_rows = valid_group[ranks >= 3] if not valid_group.empty else pd.DataFrame()
        r8_rows = valid_group[ranks >= 4] if not valid_group.empty else pd.DataFrame()
        first_countable = valid_group.iloc[0] if not valid_group.empty else first
        last_countable = valid_group.iloc[-1] if not valid_group.empty else last
        first_date = first_countable["Data_Ortofoto"]
        r8_date = r8_rows.iloc[0]["Data_Ortofoto"] if not r8_rows.empty else pd.NaT
        days_to_r8 = ""
        if pd.notna(first_date) and pd.notna(r8_date):
            days_to_r8 = int((r8_date - first_date).days)
        counted = not valid_group.empty
        empty_text = (
            group.get("Estagio_Manual", pd.Series(dtype=str)).astype(str)
            + " "
            + group.get("Estagio_Calculado", pd.Series(dtype=str)).astype(str)
            + " "
            + group.get("Observacao", pd.Series(dtype=str)).astype(str)
        ).str.lower()
        has_empty_manual = empty_text.str.contains("parcela vazia|vazia|colhida", regex=True, na=False).any()
        if has_empty_manual and not counted:
            obs = "Parcela vazia não contabilizada."
        elif group["Estagio_Manual"].astype(str).str.strip().ne("").any():
            obs = "Contém correção manual."
        else:
            obs = "Resumo temporal automático." if counted else "Parcela vazia/colhida não contabilizada."
        summary.append(
            {
                "Quadra": keys[0],
                "Parcela": keys[1],
                "Disparo": keys[2],
                "Tiro": keys[3],
                "Contabilizada": "Sim" if counted else "Não",
                "Primeira_Data_R7": r7_rows.iloc[0]["Data_Ortofoto"].date().isoformat() if not r7_rows.empty else "",
                "Primeira_Data_R7_5": r75_rows.iloc[0]["Data_Ortofoto"].date().isoformat() if not r75_rows.empty else "",
                "Primeira_Data_R8": r8_date.date().isoformat() if pd.notna(r8_date) else "",
                "Dias_Ate_R8": days_to_r8,
                "Perc_Verde_Inicial": round(float(first_countable.get("Perc_Verde", 0) or 0), 2) if counted else "",
                "Perc_Verde_Final": round(float(last_countable.get("Perc_Verde", 0) or 0), 2) if counted else "",
                "Reducao_Verde": round(float(first_countable.get("Perc_Verde", 0) or 0) - float(last_countable.get("Perc_Verde", 0) or 0), 2) if counted else "",
                "Indice_Maturacao_Inicial": round(float(first_countable.get("Indice_Maturacao", 0) or 0), 2) if counted else "",
                "Indice_Maturacao_Final": round(float(last_countable.get("Indice_Maturacao", 0) or 0), 2) if counted else "",
                "Aumento_Maturacao": round(float(last_countable.get("Indice_Maturacao", 0) or 0) - float(first_countable.get("Indice_Maturacao", 0) or 0), 2) if counted else "",
                "Estagio_Final": _stage_label(last_countable),
                "Observacao": obs,
            }
        )
    return pd.DataFrame(summary, columns=SUMMARY_COLUMNS)


def _summary_by_date(results_df: pd.DataFrame) -> pd.DataFrame:
    if results_df is None or results_df.empty:
        return pd.DataFrame()
    df = results_df.copy()
    df["Estagio_Final_Linha"] = df.apply(_stage_label, axis=1)
    df = df[df.apply(_is_countable_maturation_row, axis=1)]
    if df.empty:
        return pd.DataFrame()
    grouped = df.groupby(["Data_Ortofoto", "Estagio_Final_Linha"], dropna=False).size().reset_index(name="Parcelas")
    return grouped


def _summary_by_quadra(summary_df: pd.DataFrame) -> pd.DataFrame:
    if summary_df is None or summary_df.empty:
        return pd.DataFrame()
    if "Contabilizada" in summary_df.columns:
        summary_df = summary_df[summary_df["Contabilizada"].astype(str).str.lower().ne("não")]
    if summary_df.empty:
        return pd.DataFrame()
    return (
        summary_df.groupby(["Quadra", "Estagio_Final"], dropna=False)
        .size()
        .reset_index(name="Parcelas")
        .sort_values(["Quadra", "Estagio_Final"])
    )


def export_maturation_excel(
    results_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    output_path: str | Path | BytesIO,
    config: dict[str, Any] | None = None,
) -> str | BytesIO:
    from openpyxl.styles import Font, PatternFill

    config_df = pd.DataFrame([_cfg(config)])
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        results_df.to_excel(writer, index=False, sheet_name="Detalhado")
        summary_df.to_excel(writer, index=False, sheet_name="Resumo_Temporal")
        config_df.to_excel(writer, index=False, sheet_name="Configuracoes")
        _summary_by_date(results_df).to_excel(writer, index=False, sheet_name="Resumo_por_Data")
        _summary_by_quadra(summary_df).to_excel(writer, index=False, sheet_name="Resumo_por_Quadra")

        wb = writer.book
        header_fill = "0B243D"
        header_font = "FFFFFF"
        fills = {
            "green": "C6EFCE",
            "yellow": "FFF2CC",
            "orange": "FCE4D6",
            "red": "F4CCCC",
        }
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor=header_fill)
                cell.font = Font(bold=True, color=header_font)
            for column_cells in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 11), 35)
            headers = [cell.value for cell in ws[1]]
            for idx, header in enumerate(headers, start=1):
                if header and "Verde" in str(header):
                    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                        row[0].fill = PatternFill("solid", fgColor=fills["green"])
                if header and ("Maturacao" in str(header) or "Maturação" in str(header) or "Amarelo" in str(header)):
                    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                        row[0].fill = PatternFill("solid", fgColor=fills["yellow"])
                if header and "Estagio" in str(header):
                    for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                        value = str(row[0].value or "")
                        if "R8" in value:
                            row[0].fill = PatternFill("solid", fgColor=fills["red"])
                        elif "R7.5" in value:
                            row[0].fill = PatternFill("solid", fgColor=fills["orange"])
        if hasattr(output_path, "seek"):
            output_path.seek(0)
    return output_path


def draw_maturation_grid_overlay(image: Any, grid: dict[str, Any], results_for_date: pd.DataFrame) -> Image.Image:
    bgr = _as_uint8_bgr(image)
    rgb = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
    base = Image.fromarray(rgb).convert("RGBA")
    overlay = Image.new("RGBA", base.size, (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    date_index = int(grid.get("date_index", 0) or 0)
    parcels = _grid_to_parcels(grid, bgr.shape, date_index)
    lookup = {}
    if results_for_date is not None and not results_for_date.empty:
        lookup = {str(row["Parcela"]): _stage_label(row) for _, row in results_for_date.iterrows()}
    colors = {
        "R6": (37, 201, 111, 92),
        "R7": (255, 216, 77, 105),
        "R7.5": (255, 159, 28, 112),
        "R8": (183, 58, 46, 124),
        "Indefinido": (150, 150, 150, 80),
    }
    for parcel in parcels:
        stage = lookup.get(parcel["parcela"], "Indefinido")
        code = "R8" if "R8" in stage else "R7.5" if "R7.5" in stage else "R7" if "R7" in stage else "R6" if "R6" in stage else "Indefinido"
        pts = [tuple(map(float, point)) for point in parcel["polygon"]]
        draw.polygon(pts, fill=colors[code], outline=(0, 229, 255, 220))
        if base.width * base.height <= 6_000_000:
            cx = sum(point[0] for point in pts) / len(pts)
            cy = sum(point[1] for point in pts) / len(pts)
            draw.text((cx, cy), parcel["parcela"], fill=(255, 255, 255, 235), anchor="mm")
    return Image.alpha_composite(base, overlay).convert("RGB")


def render_soybean_maturation_module() -> None:
    try:
        import streamlit as st
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("Streamlit não está disponível para renderizar o módulo.") from exc

    st.subheader("Maturação Temporal da Soja")
    st.info(
        "Motor Python/OpenCV disponível. Para uso integrado completo, abra este módulo pelo app principal, "
        "marque o grid no visualizador e exporte/reabra o projeto JSON para análise backend."
    )
